from flask import Flask, render_template, request, jsonify, send_file
import mysql.connector
from datetime import datetime, date
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
try:
    from zoneinfo import ZoneInfo
except ImportError:
    ZoneInfo = None

app = Flask(__name__)

# Config DB Seenode
def get_db():
    return mysql.connector.connect(
        host="up-de-fra1-mysql-1.db.run-on-seenode.com",
        user="db_rrpq0erbdujn",
        password="5fUNbSRcPP3LN9K2I33Pr0ge",
        database="db_rrpq0erbdujn",
        port=11550,
    )

def ahora_monterrey():
    if ZoneInfo:
        return datetime.now(ZoneInfo("America/Monterrey"))
    return datetime.utcnow()

def asegurar_columnas_extra(cursor):
    """Garantiza que las columnas de tracking existan y la tabla plan_main_log"""
    # Asegurar que existe la tabla plan_main_log
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS plan_main_log (
            id BIGINT AUTO_INCREMENT PRIMARY KEY,
            lot_no VARCHAR(64),
            field_name VARCHAR(64),
            old_value VARCHAR(255),
            new_value VARCHAR(255),
            changed_by VARCHAR(100),
            changed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    cursor.execute("SHOW COLUMNS FROM plan_main")
    cols = set()
    for row in cursor.fetchall():
        if isinstance(row, dict):
            name = row.get('Field') or row.get('COLUMN_NAME') or list(row.values())[0]
        else:
            name = row[0]
        cols.add(name)
    
    alter_statements = []
    if 'started_at' not in cols:
        alter_statements.append("ADD COLUMN started_at DATETIME NULL AFTER updated_at")
    if 'paused_at' not in cols:
        alter_statements.append("ADD COLUMN paused_at INT DEFAULT 0 AFTER started_at")
    if 'ended_at' not in cols:
        alter_statements.append("ADD COLUMN ended_at DATETIME NULL AFTER paused_at")
    if 'pause_reason' not in cols:
        alter_statements.append("ADD COLUMN pause_reason VARCHAR(255) NULL AFTER ended_at")
    if 'end_reason' not in cols:
        alter_statements.append("ADD COLUMN end_reason VARCHAR(255) NULL AFTER pause_reason")
    # Nuevas columnas para grupos y secuencias
    if 'plan_start_date' not in cols:
        alter_statements.append("ADD COLUMN plan_start_date DATE NULL AFTER end_reason")
    if 'sequence' not in cols:
        alter_statements.append("ADD COLUMN sequence INT NULL AFTER status")
    
    # Verificar si paused_at ya existe y cambiar de DATETIME a INT si es necesario
    if 'paused_at' in cols:
        cursor.execute("SHOW COLUMNS FROM plan_main WHERE Field = 'paused_at'")
        paused_at_col = cursor.fetchone()
        if paused_at_col:
            if isinstance(paused_at_col, dict):
                column_type = paused_at_col.get('Type', '')
            else:
                column_type = paused_at_col[1] if len(paused_at_col) > 1 else ''
            
            if 'datetime' in column_type.lower():
                # Forzar migración de DATETIME a INT
                # Migrando columna paused_at de DATETIME a INT...
                try:
                    # Paso 1: Crear columna temporal
                    cursor.execute("ALTER TABLE plan_main ADD COLUMN paused_at_temp INT DEFAULT 0")
                    
                    # Paso 2: Actualizar valores (convertir NULL a 0)
                    cursor.execute("UPDATE plan_main SET paused_at_temp = 0")
                    
                    # Paso 3: Eliminar columna antigua
                    cursor.execute("ALTER TABLE plan_main DROP COLUMN paused_at")
                    
                    # Paso 4: Renombrar columna temporal
                    cursor.execute("ALTER TABLE plan_main CHANGE paused_at_temp paused_at INT DEFAULT 0")
                    
                    # Migración de paused_at completada exitosamente
                except Exception as migrate_error:
                    # Error en migración de paused_at
                    # Si falla la migración, agregar a alter_statements para intentar modificación directa
                    alter_statements.append("MODIFY COLUMN paused_at INT DEFAULT 0")
    
    if alter_statements:
        for stmt in alter_statements:
            try:
                sql = f"ALTER TABLE plan_main {stmt}"
                cursor.execute(sql)
                # Ejecutado:
            except Exception as alter_error:
                # Error en ALTER:
                pass  # Continuar con las demás alteraciones

# Generar LOT NO secuencial
def generar_lot_no(cursor, fecha):
    fecha_str = fecha.strftime("%y%m%d")
    prefix = f"ASSYLINE-{fecha_str}"
    cursor.execute("SELECT COUNT(*) as count FROM plan_main WHERE lot_no LIKE %s", (f"{prefix}%",))
    result = cursor.fetchone()
    
    # Manejar tanto cursors dictionary como normales
    if isinstance(result, dict):
        count = result["count"] + 1
    else:
        count = result[0] + 1
    
    return f"{prefix}-{count:03d}"

def log_change(cursor, lot_no, field_name, old_value, new_value, changed_by="system"):
    """Registra un cambio en la tabla plan_main_log"""
    try:
        cursor.execute("""
            INSERT INTO plan_main_log (lot_no, field_name, old_value, new_value, changed_by, changed_at)
            VALUES (%s, %s, %s, %s, %s, NOW())
        """, (lot_no, field_name, str(old_value) if old_value is not None else None, 
              str(new_value) if new_value is not None else None, changed_by))
    except Exception as e:
        # No fallar por problemas de logging
        pass

def handle_pause_resume(cursor, lot_no, new_status, current_plan_row):
    """Maneja la lógica de pausas y resumir, acumulando segundos pausados"""
    now_local = ahora_monterrey()
    if getattr(now_local, 'tzinfo', None) is not None:
        now_local = now_local.replace(tzinfo=None)
    
    current_status = current_plan_row.get('status')
    # Asegurar que paused_at sea un entero
    current_paused_value = current_plan_row.get('paused_at', 0)
    if current_paused_value is None:
        current_paused_seconds = 0
    elif isinstance(current_paused_value, (int, float)):
        current_paused_seconds = int(current_paused_value)
    else:
        # Si es datetime u otro tipo, usar 0
        current_paused_seconds = 0
    
    if new_status == 'PAUSADO' and current_status != 'PAUSADO':
        # Iniciar pausa - registrar timestamp de inicio de pausa
        log_change(cursor, lot_no, "pause_start", None, now_local.strftime("%Y-%m-%d %H:%M:%S"))
        return {}  # No actualizamos paused_at aún, solo cuando se reanude
        
    elif new_status == 'EN PROGRESO' and current_status == 'PAUSADO':
        # Reanudar desde pausa - calcular tiempo pausado y acumular
        # Buscar el último log de pause_start para este lot_no
        cursor.execute("""
            SELECT new_value FROM plan_main_log 
            WHERE lot_no = %s AND field_name = 'pause_start' 
            ORDER BY changed_at DESC LIMIT 1
        """, (lot_no,))
        
        last_pause_start = cursor.fetchone()
        if last_pause_start:
            try:
                if isinstance(last_pause_start, dict):
                    pause_start_str = last_pause_start['new_value']
                else:
                    pause_start_str = last_pause_start[0]
                
                pause_start = datetime.strptime(pause_start_str, "%Y-%m-%d %H:%M:%S")
                pause_duration_seconds = int((now_local - pause_start).total_seconds())
                
                if pause_duration_seconds > 0:
                    new_total_paused = current_paused_seconds + pause_duration_seconds
                    
                    # Log del fin de pausa
                    log_change(cursor, lot_no, "pause_end", pause_start_str, now_local.strftime("%Y-%m-%d %H:%M:%S"))
                    log_change(cursor, lot_no, "pause_duration", current_paused_seconds, new_total_paused)
                    
                    return {"paused_at": new_total_paused}
            except Exception as e:
                # En caso de error, no fallar pero log el problema
                log_change(cursor, lot_no, "pause_error", str(e), "Failed to calculate pause duration")
    
    return {}

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/control-main")
def control_main():
    # Renderiza la vista de control de operación (archivo con espacios)
    return render_template("Control de operacion de linea Main.html")

@app.route("/production-planning")
def production_planning():
    # Renderiza la vista de planeación de producción
    return render_template("production-planning.html")

# Buscar datos en la tabla raw
def buscar_en_raw_y_obtener_datos(cursor, data):
    cursor.execute("SELECT part_no, model, project FROM raw WHERE part_no=%s", (data["part_no"],))
    raw_data = cursor.fetchone()
    
    if raw_data:
        # Usar datos de raw si se encuentran
        part_no, model_code, project = raw_data
        return {"model_code": model_code, "project": project}
    else:
        # Usar datos proporcionados si no se encuentra en raw
        return {"model_code": data.get("model_code", data["part_no"]), "project": data.get("project", "")}

# Obtener CT y UPH de la tabla raw basándose en part_no
def obtener_ct_uph_desde_raw(cursor, part_no):
    """
    Busca CT y UPH en la tabla raw basándose en el part_no.
    Si no encuentra coincidencia exacta, busca por los últimos 4 dígitos del part_no.
    Retorna tupla (ct, uph) con valores por defecto si no encuentra nada.
    """
    def _norm_ct(val):
        try:
            if val is None: return 0.0
            return float(str(val).replace(',', '.'))
        except: return 0.0
    def _norm_uph(val):
        try:
            if val is None: return 0
            s = str(val).strip()
            if s == '': return 0
            # quitar posibles sufijos o texto
            for ch in [' ', '\\t', '\\n']:
                s = s.replace(ch, '')
            # eliminar caracteres no numéricos excepto punto
            filtrado = ''.join(c for c in s if (c.isdigit()))
            if filtrado == '':
                return 0
            return int(filtrado)
        except: return 0
    try:
        # 1. Coincidencia exacta
        cursor.execute("SELECT part_no, c_t, uph, model FROM raw WHERE part_no = %s ORDER BY id DESC LIMIT 1", (part_no,))
        exact = cursor.fetchone()
        if exact:
            ct = _norm_ct(exact[1] if not isinstance(exact, dict) else exact.get('c_t'))
            uph = _norm_uph(exact[2] if not isinstance(exact, dict) else exact.get('uph'))
            return (ct, uph)

        # 2. Últimos 4 dígitos
        if len(part_no) >= 4:
            last4 = part_no[-4:]
            cursor.execute("SELECT part_no, c_t, uph, model FROM raw WHERE part_no LIKE %s ORDER BY id DESC LIMIT 1", (f"%{last4}%",))
            last = cursor.fetchone()
            if last:
                ct = _norm_ct(last[1] if not isinstance(last, dict) else last.get('c_t'))
                uph = _norm_uph(last[2] if not isinstance(last, dict) else last.get('uph'))
                return (ct, uph)

        # 3. Intentar por modelo si part_no contiene guiones u otras variantes (tomar segmento inicial)
        base_candidate = part_no.split('-')[0]
        if base_candidate and base_candidate != part_no:
            cursor.execute("SELECT part_no, c_t, uph, model FROM raw WHERE model=%s OR part_no=%s ORDER BY id DESC LIMIT 1", (base_candidate, base_candidate))
            base_row = cursor.fetchone()
            if base_row:
                ct = _norm_ct(base_row[1] if not isinstance(base_row, dict) else base_row.get('c_t'))
                uph = _norm_uph(base_row[2] if not isinstance(base_row, dict) else base_row.get('uph'))
                return (ct, uph)

        return (0.0, 0)
    except Exception as e:
        pass
        return (0.0, 0)

@app.route("/api/plan", methods=["POST"])
def registrar_plan():
    data = request.json
    fecha = datetime.strptime(data["working_date"], "%Y-%m-%d")

    db = get_db()
    cursor = db.cursor()  # Cursor normal para generar_lot_no
    try:
        asegurar_columnas_extra(cursor)
    except Exception:
        pass
    
    lot_no = generar_lot_no(cursor, fecha)

    # Buscar en raw o usar datos proporcionados
    raw_info = buscar_en_raw_y_obtener_datos(cursor, data)
    model_code = raw_info["model_code"]
    project = raw_info["project"]
    part_no = data["part_no"]

    # Obtener CT y UPH desde la tabla raw
    ct_value, uph_value = obtener_ct_uph_desde_raw(cursor, part_no)

    # Mapear turno a routing (1=DIA, 2=TIEMPO EXTRA, 3=NOCHE)
    turno = (data.get("turno") or "DIA").strip().upper()
    routing_val = {"DIA": 1, "TIEMPO EXTRA": 2, "NOCHE": 3}.get(turno, 1)

    # Registrar siempre process='MAIN' con CT y UPH obtenidos de raw
    # Determinar siguiente secuencia por fecha/linea opcionalmente (simple: max+1 global)
    try:
        cursor.execute("SELECT COALESCE(MAX(sequence),0)+1 FROM plan_main")
        next_seq = cursor.fetchone()
        next_sequence = next_seq[0] if next_seq else 1
    except Exception:
        next_sequence = None

    cursor.execute(
        """
        INSERT INTO plan_main
        (lot_no, wo_code, po_code, working_date, line, model_code, part_no, project, process, plan_count, ct, uph, routing, status, sequence)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """,
        (
            lot_no,
            data.get("wo_code"),
            data.get("po_code"),
            fecha,
            data["line"],
            model_code,
            part_no,
            project,
            "MAIN",
            data.get("plan_count", 0),
            ct_value,
            uph_value,
            routing_val,
            "PLAN",
            next_sequence,
        ),
    )

    db.commit()
    cursor.close()
    db.close()
    return jsonify({"success": True, "lot_no": lot_no})

@app.route("/api/plan/confirm", methods=["POST"])
def confirmar_plan():
    data = request.json
    db = get_db()
    cursor = db.cursor()
    cursor.execute("UPDATE plan_main SET status='Confirmado' WHERE lot_no=%s", (data["lot_no"],))
    db.commit()
    cursor.close()
    db.close()
    return jsonify({"success": True, "lot_no": data["lot_no"]})

@app.route("/api/plan", methods=["GET"])
def listar_planes():
    db = get_db()
    cursor = db.cursor(dictionary=True)
    try:
        asegurar_columnas_extra(cursor)
    except Exception:
        pass
    # Filtros de fecha opcionales
    base_query = "SELECT * FROM plan_main"
    params = []
    start = request.args.get("start")
    end = request.args.get("end")
    where = []
    if start:
        where.append("working_date >= %s")
        params.append(start)
    if end:
        where.append("working_date <= %s")
        params.append(end)
    if where:
        base_query += " WHERE " + " AND ".join(where)
    
    # Ordenar por grupo y secuencia dentro del grupo
    base_query += """ ORDER BY 
        IFNULL(group_no, 999) ASC, 
        IFNULL(sequence, 999) ASC, 
        working_date DESC, 
        created_at DESC
    """

    cursor.execute(base_query, tuple(params))
    rows = cursor.fetchall()
    cursor.close()
    db.close()

    # Formatear la fecha como dd/MM/yyyy y agregar campos calculados
    for r in rows:
        wd = r.get("working_date")
        try:
            if isinstance(wd, datetime):
                r["working_date"] = wd.strftime("%d/%m/%Y")
            else:
                r["working_date"] = datetime.strptime(str(wd)[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            pass
        
        # Formatear fecha de inicio del plan
        psd = r.get("plan_start_date")
        if psd:
            try:
                if isinstance(psd, datetime):
                    r["plan_start_date"] = psd.strftime("%d/%m/%Y")
                else:
                    r["plan_start_date"] = datetime.strptime(str(psd)[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                r["plan_start_date"] = ""
        else:
            r["plan_start_date"] = ""
        # Producido (placeholder hasta conectar tabla real)
        if "produced" not in r:
            r["produced"] = 0

    return jsonify(rows)

@app.route("/api/plan/order", methods=["POST"])
def actualizar_orden_planes():
    """Actualizar la columna sequence para una lista de lot_nos en el orden dado."""
    data = request.json or {}
    lot_nos = data.get("lot_nos", [])
    if not lot_nos or not isinstance(lot_nos, list):
        return jsonify({"error": "lot_nos requerido (lista)"}), 400

    try:
        db = get_db()
        cursor = db.cursor()
        try:
            asegurar_columnas_extra(cursor)
        except Exception:
            pass
        # Construir UPDATE con CASE para establecer sequence=1..n
        case_parts = []
        params = []
        for idx, lot in enumerate(lot_nos, start=1):
            case_parts.append("WHEN %s THEN %s")
            params.extend([lot, idx])
        case_expr = "CASE lot_no " + " ".join(case_parts) + " END"
        placeholders = ",".join(["%s"] * len(lot_nos))
        sql = f"UPDATE plan_main SET sequence = {case_expr}, updated_at = NOW() WHERE lot_no IN ({placeholders})"
        params.extend(lot_nos)
        cursor.execute(sql, tuple(params))
        db.commit()
        return jsonify({"success": True, "rows_affected": cursor.rowcount})
    except Exception as e:
        try:
            db.rollback()
        except:
            pass
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            cursor.close(); db.close()
        except:
            pass

@app.route("/api/plan/update", methods=["POST"])
def actualizar_plan():
    data = request.json
    plan_id = data.get("id")
    lot_no = data.get("lot_no")
    
    # Validar que se proporcione al menos un identificador
    if not plan_id and not lot_no:
        return jsonify({"error": "ID de plan o LOT NO requerido"}), 400
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        # Determinar WHERE y obtener part_no actual para relookup
        if plan_id:
            where_condition = "id = %s"
            id_value = plan_id
        else:
            where_condition = "lot_no = %s"
            id_value = lot_no

        # Leer fila actual para obtener part_no si no viene en el payload
        current_part_no = data.get('part_no')
        if not current_part_no:
            try:
                cursor.execute(f"SELECT part_no FROM plan_main WHERE {where_condition} LIMIT 1", (id_value,))
                row = cursor.fetchone()
                if row:
                    current_part_no = row[0] if not isinstance(row, dict) else row.get('part_no')
            except Exception:
                pass

        # Construir query dinámicamente basado en los campos proporcionados
        set_fields = []
        values = []

        # Mapear 'turno' a routing si viene desde el formulario
        turno_in = data.get('turno')
        if turno_in:
            routing_val = {"DIA": 1, "TIEMPO EXTRA": 2, "NOCHE": 3}.get(str(turno_in).strip().upper(), 1)
            set_fields.append("routing = %s")
            values.append(routing_val)

        # Campos que se pueden actualizar directamente desde payload
        updatable_fields = {
            "plan_count": "plan_count",
            "status": "status",
            "line": "line",
            "model_code": "model_code",
            "part_no": "part_no",
            "project": "project",
            "process": "process"
        }

        for field_key, db_column in updatable_fields.items():
            if field_key in data:
                set_fields.append(f"{db_column} = %s")
                values.append(data[field_key])

        # Rebuscar en RAW: model_code, project, CT, UPH usando part_no
        try:
            if current_part_no:
                # model_code y project (preferir match exacto)
                cursor2 = db.cursor()
                cursor2.execute("SELECT model, project FROM raw WHERE part_no=%s ORDER BY id DESC LIMIT 1", (current_part_no,))
                raw_row = cursor2.fetchone()
                model_code_new = None
                project_new = None
                if raw_row:
                    if isinstance(raw_row, dict):
                        model_code_new = raw_row.get('model')
                        project_new = raw_row.get('project')
                    else:
                        model_code_new = raw_row[0]
                        project_new = raw_row[1]
                # Fallback: buscar por últimos 4 dígitos si no se encontró
                if not model_code_new or not project_new:
                    try:
                        if current_part_no and len(str(current_part_no)) >= 4:
                            last4 = str(current_part_no)[-4:]
                            cursor2.execute("SELECT model, project FROM raw WHERE part_no LIKE %s ORDER BY id DESC LIMIT 1", (f"%{last4}%",))
                            r2 = cursor2.fetchone()
                            if r2:
                                if isinstance(r2, dict):
                                    model_code_new = model_code_new or r2.get('model')
                                    project_new = project_new or r2.get('project')
                                else:
                                    model_code_new = model_code_new or r2[0]
                                    project_new = project_new or r2[1]
                    except Exception:
                        pass
                # CT/UPH usando función robusta (acepta últimos 4)
                ct_new, uph_new = obtener_ct_uph_desde_raw(cursor, current_part_no)

                # Aplicar solo si tenemos valores válidos
                if model_code_new:
                    set_fields.append("model_code = %s")
                    values.append(model_code_new)
                if project_new:
                    set_fields.append("project = %s")
                    values.append(project_new)
                if ct_new and float(ct_new) > 0:
                    set_fields.append("ct = %s")
                    values.append(ct_new)
                if uph_new and int(uph_new) > 0:
                    set_fields.append("uph = %s")
                    values.append(int(uph_new))
        except Exception:
            pass

        if not set_fields:
            return jsonify({"error": "No hay campos para actualizar"}), 400

        # Agregar updated_at
        set_fields.append("updated_at = NOW()")

        query = f"UPDATE plan_main SET {', '.join(set_fields)} WHERE {where_condition}"
        values.append(id_value)

        cursor.execute(query, values)

        if cursor.rowcount == 0:
            return jsonify({"error": "Plan no encontrado"}), 404

        db.commit()

        return jsonify({"success": True, "rows_affected": cursor.rowcount})

    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        db.close()

@app.route("/api/plan/save-sequences", methods=["POST"])
def guardar_secuencias():
    """Guardar secuencias por grupo y fechas de inicio"""

    
    try:
        data = request.get_json()

        
        if not data or 'sequences' not in data:

            return jsonify({"error": "Datos de secuencias requeridos"}), 400
    
        sequences = data['sequences']

        
        db = get_db()
        db.autocommit = False  # Asegurarse de que autocommit esté desactivado para control manual
        cursor = db.cursor()
        
        # Verificar que las columnas necesarias existan
        cursor.execute("SHOW COLUMNS FROM plan_main WHERE Field IN ('plan_start_date', 'planned_start', 'planned_end', 'group_no', 'sequence', 'effective_minutes', 'breaks_minutes')")
        existing_columns = [row[0] if isinstance(row[0], str) else row[0].decode('utf-8') for row in cursor.fetchall()]
        

        
        # Actualizar cada plan con su grupo y secuencia
        updated_count = 0
        
        for i, plan_data in enumerate(sequences):
            lot_no = plan_data.get('lot_no')
            plan_start_date = plan_data.get('plan_start_date')
            planned_start = plan_data.get('planned_start')
            planned_end = plan_data.get('planned_end')
            group_no = plan_data.get('group_no')
            sequence = plan_data.get('sequence')
            effective_minutes = plan_data.get('effective_minutes')
            breaks_minutes = plan_data.get('breaks_minutes')
            

            
            if lot_no and group_no is not None and sequence is not None:
                # Construir query dinámicamente basado en columnas existentes
                set_clauses = []
                params = []
                
                # Solo usar group_no y sequence (las columnas principales)
                if 'group_no' in existing_columns:
                    set_clauses.append("group_no = %s")
                    params.append(group_no)
                
                if 'sequence' in existing_columns:
                    set_clauses.append("sequence = %s")
                    params.append(sequence)
                
                if 'plan_start_date' in existing_columns and plan_start_date and plan_start_date != '--':
                    set_clauses.append("plan_start_date = %s")
                    params.append(plan_start_date)
                
                if 'planned_start' in existing_columns and planned_start and planned_start != '--':
                    set_clauses.append("planned_start = %s")
                    params.append(planned_start)
                
                if 'planned_end' in existing_columns and planned_end and planned_end != '--':
                    set_clauses.append("planned_end = %s")
                    params.append(planned_end)
                
                if 'effective_minutes' in existing_columns and effective_minutes is not None:
                    set_clauses.append("effective_minutes = %s")
                    params.append(effective_minutes)
                
                if 'breaks_minutes' in existing_columns and breaks_minutes is not None:
                    set_clauses.append("breaks_minutes = %s")
                    params.append(breaks_minutes)
                
                # Siempre actualizar updated_at
                set_clauses.append("updated_at = NOW()")
                
                if set_clauses:
                    # Agregar lot_no para el WHERE al final
                    params.append(lot_no)
                    update_query = f"UPDATE plan_main SET {', '.join(set_clauses)} WHERE lot_no = %s"
                    
                    result = cursor.execute(update_query, params)
                    rows_affected = cursor.rowcount
                    
                    if rows_affected > 0:
                        updated_count += 1
                    else:
                        # Verificar si el lot_no existe
                        cursor.execute("SELECT COUNT(*) FROM plan_main WHERE lot_no = %s", (lot_no,))
                        exists = cursor.fetchone()[0]
                        
        db.commit()

        
        # Verificar que los datos se guardaron realmente

        lot_nos = [plan_data['lot_no'] for plan_data in sequences]
        placeholders = ','.join(['%s'] * len(lot_nos))
        verify_query = f"SELECT lot_no, group_no as grupo, sequence as secuencia, plan_start_date, planned_start, planned_end FROM plan_main WHERE lot_no IN ({placeholders})"
        cursor.execute(verify_query, lot_nos)
        saved_data = cursor.fetchall()

        for row in saved_data[:3]:  # Mostrar solo los primeros 3 para no saturar el log
            pass  # Row data
        

        
        return jsonify({
            "success": True, 
            "message": f"✅ {updated_count} secuencias guardadas correctamente",
            "updated_count": updated_count
        })
        
    except Exception as e:
        # Error en guardar_secuencias
        # Tipo de error:
        import traceback

        if 'db' in locals():
            db.rollback()
        return jsonify({"error": f"Error al guardar secuencias: {str(e)}"}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'db' in locals():
            db.close()

@app.route("/api/plan-main/list", methods=["GET"])
def listar_plan_main():
    """Endpoint específico para el control de operación Main con mapeo de columnas correcto"""
    db = get_db()
    cursor = db.cursor(dictionary=True)
    
    # Construir query con filtros - incluir paused_at
    base_query = """
        SELECT 
            id,
            lot_no as lote,
            part_no as nparte, 
            model_code as modelo,
            line as linea,
            'TURNO1' as turno,
            working_date as fecha_inicio,
            working_date as fecha_fin,
            plan_count as qty,
            COALESCE(produced_count,0) as producido,
            GREATEST(plan_count - COALESCE(produced_count,0),0) as falta,
            status as estatus,
            COALESCE(paused_at, 0) as paused_at,
            wo_code,
            project,
            process,
            routing,
            ct,
            uph,
            created_at,
            updated_at
        FROM plan_main
    """
    
    params_base = []  # para filtros no-fecha (q, linea)
    date_filters = []
    
    # Filtros de la query
    q = request.args.get("q")
    linea = request.args.get("linea")
    solo_pendientes = request.args.get("solo_pendientes")
    desde = request.args.get("desde")
    hasta = request.args.get("hasta")
    
    if q and q.strip():
        # Filtro base (se aplicará tanto al lado de fecha como al de EN PROGRESO)
        params_base.extend([f"%{q.strip()}%", f"%{q.strip()}%", f"%{q.strip()}%"])
        q_clause = "(lot_no LIKE %s OR part_no LIKE %s OR model_code LIKE %s)"
    else:
        q_clause = None
    
    if linea and linea != "ALL" and linea != "Todos":
        line_clause = "line = %s"
        params_base.append(linea)
    else:
        line_clause = None
    
    if solo_pendientes == "true":
        status_only_plan = True
    else:
        status_only_plan = False
        if desde:
            date_filters.append("working_date >= %s")
        if hasta:
            date_filters.append("working_date <= %s")
    
    # Construir cláusulas base (no-fecha)
    base_clauses = []
    if q_clause: base_clauses.append(q_clause)
    if line_clause: base_clauses.append(line_clause)
    base_expr = " AND ".join(base_clauses) if base_clauses else "1=1"

    if status_only_plan:
        # Mostrar solo PLAN explícitamente (no incluir EN PROGRESO) y respetar base filters
        base_query += f" WHERE {base_expr} AND status = 'PLAN' AND status <> 'CANCELADO'"
        params = params_base
    else:
        if date_filters:
            # ((base AND fechas) OR (base AND EN PROGRESO/PAUSADO)) y excluir cancelados
            date_expr = " AND ".join(date_filters)
            base_query += f" WHERE (({base_expr} AND {date_expr}) OR ({base_expr} AND status IN ('EN PROGRESO','PAUSADO'))) AND status <> 'CANCELADO'"
            # params: base+fecha para primer lado, base para segundo lado
            params = []
            params.extend(params_base)  # base para lado 1
            if desde: params.append(desde)
            if hasta: params.append(hasta)
            params.extend(params_base)  # base para lado 2
        else:
            # Default: hoy o EN PROGRESO/PAUSADO, más filtros base
            base_query += f" WHERE (({base_expr} AND DATE(working_date) = CURDATE()) OR ({base_expr} AND status IN ('EN PROGRESO','PAUSADO'))) AND status <> 'CANCELADO'"
            params = params_base + params_base
    
    base_query += " ORDER BY working_date DESC, created_at DESC"
    
    try:
        cursor.execute(base_query, tuple(params))
        rows = cursor.fetchall()
        
        # Formatear fechas para que no aparezca "invalid date"
        for row in rows:
            if row.get("fecha_inicio"):
                try:
                    # Manejar diferentes formatos de fecha
                    if isinstance(row["fecha_inicio"], str):
                        # Si viene como string YYYY-MM-DD
                        fecha_str = row["fecha_inicio"][:10]  
                        row["fecha_inicio"] = fecha_str  # Mantener formato ISO para el frontend
                        row["fecha_fin"] = fecha_str
                    else:
                        # Si viene como objeto date/datetime
                        fecha = row["fecha_inicio"]
                        fecha_str = fecha.strftime("%Y-%m-%d")
                        row["fecha_inicio"] = fecha_str
                        row["fecha_fin"] = fecha_str
                except Exception as date_error:
                    pass
        
        return jsonify(rows)
        
    except Exception as e:
        return jsonify({"error": f"Error en consulta: {str(e)}"}), 500
    finally:
        cursor.close()
        db.close()

@app.route("/api/work-orders", methods=["GET"])
def listar_work_orders():
    """Consultar work orders para importación"""
    db = get_db()
    cursor = db.cursor(dictionary=True)
    
    # Filtros opcionales
    fecha = request.args.get("fecha")
    estado = request.args.get("estado")
    
    base_query = "SELECT * FROM work_orders"
    params = []
    where_clauses = []
    
    if fecha:
        where_clauses.append("fecha_operacion = %s")
        params.append(fecha)
    
    if estado:
        where_clauses.append("estado = %s")
        params.append(estado)
    
    if where_clauses:
        base_query += " WHERE " + " AND ".join(where_clauses)
    
    base_query += " ORDER BY fecha_operacion DESC, id DESC"
    
    cursor.execute(base_query, tuple(params))
    work_orders = cursor.fetchall()
    
    # Verificar cuáles WOs ya fueron importadas
    if work_orders:
        wo_codes = [wo["codigo_wo"] for wo in work_orders if wo.get("codigo_wo")]
        if wo_codes:
            placeholders = ",".join(["%s"] * len(wo_codes))
            cursor.execute(f"""
                SELECT wo_code, lot_no, status 
                FROM plan_main 
                WHERE wo_code IN ({placeholders})
            """, wo_codes)
            imported_wos = {row["wo_code"]: {"lot_no": row["lot_no"], "status": row["status"]} 
                           for row in cursor.fetchall()}
        else:
            imported_wos = {}
    else:
        imported_wos = {}
    
    cursor.close()
    db.close()
    
    # Formatear fechas y agregar información de importación
    for wo in work_orders:
        if wo.get("fecha_operacion"):
            try:
                if isinstance(wo["fecha_operacion"], datetime):
                    wo["fecha_operacion"] = wo["fecha_operacion"].strftime("%Y-%m-%d")
                else:
                    wo["fecha_operacion"] = str(wo["fecha_operacion"])[:10]
            except Exception:
                pass
        
        if wo.get("fecha_modificacion"):
            try:
                if isinstance(wo["fecha_modificacion"], datetime):
                    wo["fecha_modificacion"] = wo["fecha_modificacion"].strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                pass
        
        # Agregar información de si ya fue importado
        wo_code = wo.get("codigo_wo")
        if wo_code and wo_code in imported_wos:
            wo["ya_importado"] = True
            wo["lot_no_existente"] = imported_wos[wo_code]["lot_no"]
            wo["status_plan"] = imported_wos[wo_code]["status"]
        else:
            wo["ya_importado"] = False
    
    return jsonify(work_orders)

@app.route("/api/work-orders/import", methods=["POST"])
def importar_work_orders():
    """Importar work orders seleccionados como planes"""
    data = request.json
    wo_ids = data.get("wo_ids", [])
    
    if not wo_ids:
        return jsonify({"error": "No se seleccionaron work orders"}), 400
    
    db = get_db()
    cursor = db.cursor(dictionary=True)
    
    imported_plans = []
    errors = []
    
    try:
        for wo_id in wo_ids:
            try:
                # Obtener datos del work order
                cursor.execute("SELECT * FROM work_orders WHERE id = %s", (wo_id,))
                wo = cursor.fetchone()
                
                if not wo:
                    errors.append(f"Work Order ID {wo_id} no encontrado")
                    continue
                
                # Verificar si ya existe un plan para este WO
                cursor.execute("SELECT lot_no, status FROM plan_main WHERE wo_code = %s", (wo["codigo_wo"],))
                existing_plan = cursor.fetchone()
                
                if existing_plan:
                    if isinstance(existing_plan, dict):
                        lot_no_existing = existing_plan["lot_no"]
                        status_existing = existing_plan["status"]
                    else:
                        lot_no_existing = existing_plan[0]
                        status_existing = existing_plan[1]
                    
                    errors.append(f"WO {wo['codigo_wo']} ya fue importada como LOT NO: {lot_no_existing} (Status: {status_existing})")
                    continue
                
                # Generar LOT NO
                fecha_op = wo["fecha_operacion"]
                if isinstance(fecha_op, str):
                    fecha_op = datetime.strptime(fecha_op[:10], "%Y-%m-%d")
                
                lot_no = generar_lot_no(cursor, fecha_op)
                
                # Intentar traer datos + CT/UPH en una sola búsqueda priorizando part_no real si lo tenemos
                part_no = None
                model_code = None
                project = None
                ct_value = 0.0
                uph_value = 0

                # Consulta combinada: intentar por modelo, código_modelo y variantes parciales
                cursor.execute("""
                    SELECT part_no, model, project, c_t, uph
                    FROM raw
                    WHERE model = %s OR model = %s OR part_no = %s OR part_no LIKE %s
                    ORDER BY id DESC
                    LIMIT 1
                """, (wo["modelo"], wo["codigo_modelo"], wo["codigo_modelo"], f"%{wo['modelo']}%"))
                raw_row = cursor.fetchone()
                if raw_row:
                    part_no = raw_row.get('part_no') if isinstance(raw_row, dict) else raw_row[0]
                    model_code = raw_row.get('model') if isinstance(raw_row, dict) else raw_row[1]
                    project = raw_row.get('project') if isinstance(raw_row, dict) else raw_row[2]
                    ct_value = raw_row.get('c_t') if isinstance(raw_row, dict) else raw_row[3]
                    uph_value = raw_row.get('uph') if isinstance(raw_row, dict) else raw_row[4]
                # Normalizar valores CT/UPH del primer intento
                try:
                    ct_value = float(ct_value) if ct_value not in (None, '') else 0.0
                except:
                    ct_value = 0.0
                try:
                    uph_value = int(str(uph_value).strip()) if uph_value not in (None, '') and str(uph_value).strip().isdigit() else 0
                except:
                    uph_value = 0

                # Fallback si no se encontró nada o CT/UPH siguen en cero: usar función robusta
                if not part_no:
                    part_no = wo["codigo_modelo"] or wo["modelo"]
                if not model_code:
                    model_code = wo["modelo"]
                if not project:
                    project = wo["nombre_modelo"] or wo["modelo"]
                if (ct_value == 0.0 and uph_value == 0):
                    ct_value, uph_value = obtener_ct_uph_desde_raw(cursor, part_no)
                
                # Validar y preparar datos para INSERT
                insert_data = {
                    'lot_no': lot_no,
                    'wo_code': wo["codigo_wo"] or '',
                    'po_code': wo["codigo_po"] or '',
                    'working_date': wo["fecha_operacion"],
                    'line': wo["linea"] or 'MAIN_LINE',  # Usar la línea del WO
                    'model_code': model_code or '',
                    'part_no': part_no or '',
                    'project': project or '',
                    'process': 'MAIN',
                    'plan_count': wo["cantidad_planeada"] or 0,
                    'ct': ct_value,
                    'uph': uph_value,
                    'routing': 1,  # Valor entero para routing (1 = Standard)
                    'status': 'PLAN'  # Cambiar de 'PLANNED' a 'PLAN'
                }
                
                # Insertar plan
                try:
                    cursor.execute("""
                        INSERT INTO plan_main (
                            lot_no, wo_code, po_code, working_date, line, model_code, 
                            part_no, project, process, plan_count, ct, uph, 
                            routing, status, created_at, updated_at
                        ) VALUES (
                            %s, %s, %s, %s, %s, %s, 
                            %s, %s, %s, %s, %s, %s, 
                            %s, %s, NOW(), NOW()
                        )
                    """, (
                        insert_data['lot_no'], insert_data['wo_code'], insert_data['po_code'], 
                        insert_data['working_date'], insert_data['line'], insert_data['model_code'],
                        insert_data['part_no'], insert_data['project'], insert_data['process'], 
                        insert_data['plan_count'], insert_data['ct'], insert_data['uph'],
                        insert_data['routing'], insert_data['status']
                    ))
                    
                    if cursor.rowcount == 0:
                        raise Exception("INSERT no afectó ninguna fila")
                    
                    imported_plans.append({
                        "lot_no": lot_no,
                        "wo_code": wo["codigo_wo"],
                        "cantidad": wo["cantidad_planeada"]
                    })
                    
                except Exception as insert_error:
                    errors.append(f"WO ID {wo_id}: Error en INSERT - {str(insert_error)}")
                    continue
                
            except Exception as e:
                errors.append(f"Error importando WO ID {wo_id}: {str(e)}")
                continue
        
        db.commit()
        
        return jsonify({
            "success": True,
            "imported": len(imported_plans),
            "plans": imported_plans,
            "errors": errors
        })
        
    except Exception as e:
        db.rollback()
        return jsonify({"error": f"Error en importación: {str(e)}"}), 500
    
    finally:
        cursor.close()
        db.close()

@app.route("/api/raw/ct-uph/<part_no>", methods=["GET"])
def obtener_ct_uph_raw(part_no):
    """Obtener CT y UPH de la tabla raw basado en part_no"""
    db = get_db()
    cursor = db.cursor(dictionary=True)
    
    try:
        # Extraer los últimos 4 dígitos del part_no
        # Ejemplo: EBR80757422 -> 7422
        if len(part_no) >= 4:
            part_search = part_no[-4:]  # Últimos 4 dígitos
        else:
            part_search = part_no
        
        # Buscar en la tabla raw
        query = """
            SELECT part_no, c_t, uph, model, project, main_display 
            FROM raw 
            WHERE part_no LIKE %s 
            ORDER BY id DESC 
            LIMIT 1
        """
        
        cursor.execute(query, (f"%{part_search}%",))
        result = cursor.fetchone()
        
        if result:
            return jsonify({
                "success": True,
                "found": True,
                "data": {
                    "ct": float(result["c_t"]) if result["c_t"] else 0.0,
                    "uph": result["uph"] or "-",
                    "part_no_found": result["part_no"],
                    "model": result["model"],
                    "project": result["project"],
                    "main_display": result["main_display"]
                }
            })
        else:
            return jsonify({
                "success": True,
                "found": False,
                "data": {
                    "ct": 0.0,
                    "uph": "-",
                    "part_no_search": part_search
                }
            })
            
    except Exception as e:
        return jsonify({"error": f"Error obteniendo datos: {str(e)}"}), 500
    
    finally:
        cursor.close()
        db.close()

@app.route("/api/plan/status", methods=["POST"])
def actualizar_status_plan():
    """Actualizar status de un plan por lot_no. Estados permitidos: EN PROGRESO, TERMINADO, PAUSADO"""
    data = request.json or {}
    lot_no = data.get("lot_no")
    nuevo_status = data.get("status")
    pause_reason = data.get("pause_reason")
    end_reason = data.get("end_reason")

    if not lot_no or not nuevo_status:
        return jsonify({"error": "lot_no y status requeridos"}), 400

    # Normalizar status de entrada
    status_map = {
        "EN PROGRESO": "EN PROGRESO",
        "INICIADO": "EN PROGRESO",
        "RUNNING": "EN PROGRESO",
        "TERMINADO": "TERMINADO",
        "FINALIZADO": "TERMINADO",
        "ENDED": "TERMINADO",
        "PAUSA": "PAUSADO",
        "PAUSADO": "PAUSADO"
    }
    normalized = status_map.get(nuevo_status.upper())
    if not normalized:
        return jsonify({"error": "Status no permitido"}), 400

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        try:
            asegurar_columnas_extra(cursor)
        except Exception as col_err:
            pass

        try:
            cursor.execute("SELECT line, status, plan_count, produced_count, started_at, paused_at, ended_at FROM plan_main WHERE lot_no=%s LIMIT 1", (lot_no,))
            plan_row = cursor.fetchone()
        except Exception as sel_err:
            # Columnas no existen todavía: hacer SELECT mínimo
            cursor.execute("SELECT line, status, plan_count, produced_count FROM plan_main WHERE lot_no=%s LIMIT 1", (lot_no,))
            base_row = cursor.fetchone()
            if not base_row:
                return jsonify({"error": "Plan no encontrado", "error_code": "NOT_FOUND"}), 404
            # Normalizar a dict con claves esperadas vacías
            if isinstance(base_row, dict):
                plan_row = {
                    'line': base_row.get('line'),
                    'status': base_row.get('status'),
                    'plan_count': base_row.get('plan_count'),
                    'produced_count': base_row.get('produced_count'),
                    'started_at': None,
                    'paused_at': 0,
                    'ended_at': None
                }
            else:
                line_, status_, pc_, prod_ = base_row
                plan_row = {
                    'line': line_, 'status': status_, 'plan_count': pc_, 'produced_count': prod_,
                    'started_at': None, 'paused_at': 0, 'ended_at': None
                }
        if not plan_row:
            return jsonify({"error": "Plan no encontrado", "error_code": "NOT_FOUND"}), 404

        linea_objetivo = plan_row.get("line")
        plan_count = plan_row.get("plan_count") or 0
        produced_count = plan_row.get("produced_count") or 0
        current_status = plan_row.get("status")

        if normalized == "EN PROGRESO":
            cursor.execute(
                """
                SELECT lot_no FROM plan_main 
                WHERE line=%s AND status='EN PROGRESO' AND lot_no<>%s
                LIMIT 1
                """,
                (linea_objetivo, lot_no)
            )
            conflict = cursor.fetchone()
            if conflict:
                conflict_lot = conflict["lot_no"] if isinstance(conflict, dict) else conflict[0]
                return jsonify({
                    "error": "Ya existe un plan EN PROGRESO en esta línea",
                    "error_code": "LINE_CONFLICT",
                    "line": linea_objetivo,
                    "lot_no_en_progreso": conflict_lot
                }), 409

        # Log del cambio de status
        if current_status != normalized:
            log_change(cursor, lot_no, "status", current_status, normalized)

        sets = ["status=%s", "updated_at=NOW()"]
        values = [normalized]
        now_local = ahora_monterrey()
        # MySQL connector espera datetime naive; remover tzinfo si existe
        if getattr(now_local, 'tzinfo', None) is not None:
            now_local = now_local.replace(tzinfo=None)

        # Manejar pausas y reanudaciones
        pause_updates = handle_pause_resume(cursor, lot_no, normalized, plan_row)
        for field, value in pause_updates.items():
            sets.append(f"{field}=%s")
            values.append(value)

        if normalized == 'EN PROGRESO' and not plan_row.get('started_at'):
            sets.append("started_at=%s")
            values.append(now_local)
            log_change(cursor, lot_no, "started_at", None, now_local.strftime("%Y-%m-%d %H:%M:%S"))
        elif normalized == 'PAUSADO':
            if pause_reason:
                sets.append("pause_reason=%s")
                values.append(pause_reason[:255])
                log_change(cursor, lot_no, "pause_reason", None, pause_reason[:255])
        elif normalized == 'TERMINADO':
            if not plan_row.get('ended_at'):
                sets.append("ended_at=%s")
                values.append(now_local)
                log_change(cursor, lot_no, "ended_at", None, now_local.strftime("%Y-%m-%d %H:%M:%S"))
            if plan_count and produced_count < plan_count and end_reason:
                sets.append("end_reason=%s")
                values.append(end_reason[:255])
                log_change(cursor, lot_no, "end_reason", None, end_reason[:255])

        sql = f"UPDATE plan_main SET {', '.join(sets)} WHERE lot_no=%s"
        values.append(lot_no)
        try:
            cursor.execute(sql, tuple(values))
            db.commit()
        except Exception as exec_err:
            return jsonify({"error": f"Fallo update status: {exec_err}", "error_code": "UPDATE_EXCEPTION"}), 500
        if cursor.rowcount == 0:
            return jsonify({"error": "No se actualizó ninguna fila", "error_code": "NO_ROWS_UPDATED"}), 400
        
        # Re-consultar para devolver timestamps actualizados
        cursor.execute("SELECT status, started_at, paused_at, ended_at FROM plan_main WHERE lot_no=%s", (lot_no,))
        updated = cursor.fetchone() or {}
        
        return jsonify({
            "success": True,
            "lot_no": lot_no,
            "status": updated.get('status', normalized),
            "started_at": str(updated.get('started_at') or ''),
            "paused_at": updated.get('paused_at', 0),
            "ended_at": str(updated.get('ended_at') or '')
        })
    except Exception as e:
        return jsonify({"error": str(e), "error_code": "UNHANDLED_EXCEPTION"}), 500
    finally:
        try:
            cursor.close()
            db.close()
        except:
            pass

@app.route("/api/plan/produce", methods=["POST"])
def producir_plan():
    """Actualizar el produced_count de un plan.
    JSON esperado:
    {
      "lot_no": "...",            # requerido
      "increment": 50,             # opcional, suma a produced_count
      "set": 120                   # opcional, fija produced_count (prioridad sobre increment)
    }
    Lógica:
    - Si se pasa 'set', produced_count = set (no negativo)
    - Sino si 'increment', produced_count = produced_count + increment (no negativo)
    - Ajusta status automáticamente:
        * produced_count >= plan_count  => TERMINADO
        * 0 < produced_count < plan_count y status en (PLAN, PAUSADO) => EN PROGRESO
    - Devuelve produced_count final y status
    """
    data = request.json or {}
    lot_no = data.get("lot_no")
    inc = data.get("increment")
    set_val = data.get("set")

    if not lot_no:
        return jsonify({"error": "lot_no requerido"}), 400

    if inc is None and set_val is None:
        return jsonify({"error": "Debe enviar 'increment' o 'set'"}), 400

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        cursor.execute("SELECT id, plan_count, produced_count, status FROM plan_main WHERE lot_no=%s LIMIT 1", (lot_no,))
        row = cursor.fetchone()
        if not row:
            return jsonify({"error": "Plan no encontrado"}), 404

        plan_count = row.get("plan_count") or 0
        produced = row.get("produced_count") or 0
        status_actual = row.get("status") or "PLAN"

        # Calcular nuevo produced_count
        if set_val is not None:
            try:
                produced_nuevo = max(0, int(set_val))
            except:
                return jsonify({"error": "Valor 'set' inválido"}), 400
        else:
            try:
                inc_val = int(inc)
            except:
                return jsonify({"error": "Valor 'increment' inválido"}), 400
            produced_nuevo = max(0, produced + inc_val)

        # Limitar opcionalmente a plan_count (puede comentarse si se quiere exceder)
        if plan_count and produced_nuevo > plan_count:
            produced_nuevo = plan_count

        # Determinar nuevo status automático
        nuevo_status = status_actual
        if produced_nuevo >= plan_count and plan_count > 0:
            nuevo_status = 'TERMINADO'
        elif produced_nuevo > 0 and status_actual in ('PLAN', 'PAUSADO'):
            nuevo_status = 'EN PROGRESO'

        cursor.execute(
            "UPDATE plan_main SET produced_count=%s, status=%s, updated_at=NOW() WHERE lot_no=%s",
            (produced_nuevo, nuevo_status, lot_no)
        )
        db.commit()

        return jsonify({
            "success": True,
            "lot_no": lot_no,
            "produced_count": produced_nuevo,
            "plan_count": plan_count,
            "status": nuevo_status
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        try:
            cursor.close(); db.close()
        except:
            pass

@app.route("/api/production-planning/generate", methods=["POST"])
def generar_planeacion_produccion():
    """
    Generar planeación de producción automática con cálculo de tiempos, breaks y turnos.
    
    JSON esperado:
    {
      "date": "2024-01-15",           # Fecha de la planeación
      "line": "ALL" | "M1" | "M2",   # Línea específica o todas
      "config": {                     # Configuración opcional
        "breaks": [
          {"start": "09:30", "end": "09:45", "name": "Break 1"},
          {"start": "12:00", "end": "12:30", "name": "Almuerzo"},
          {"start": "15:00", "end": "15:15", "name": "Break 2"}
        ],
        "shiftStart": "07:30",
        "productiveHours": 9,  # 9 horas reales de trabajo (descontando breaks)
        "lineFlows": {"M1": "D1", "M2": "D2", "M3": "D3"}
      }
    }
    
    Retorna la planeación calculada con tiempos, breaks y detección de tiempo extra.
    """
    data = request.json or {}
    fecha = data.get("date")
    linea = data.get("line", "ALL")
    config = data.get("config", {})
    
    if not fecha:
        return jsonify({"error": "Fecha requerida"}), 400
    
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        
        # Obtener planes de la fecha especificada
        base_query = """
            SELECT 
                lot_no, wo_code, po_code, working_date, line, model_code, 
                part_no, project, process, plan_count, ct, uph, routing, 
                status, sequence, created_at
            FROM plan_main 
            WHERE DATE(working_date) = %s 
            AND status NOT IN ('CANCELADO')
        """
        params = [fecha]
        
        if linea != "ALL":
            base_query += " AND line = %s"
            params.append(linea)
            
        base_query += " ORDER BY line, IFNULL(sequence, 999999), created_at"
        
        cursor.execute(base_query, params)
        planes = cursor.fetchall()
        
        if not planes:
            return jsonify({
                "success": True,
                "date": fecha,
                "line": linea,
                "planning": [],
                "message": "No se encontraron planes para la fecha y línea especificadas"
            })
        
        # Configuración por defecto
        default_config = {
            "breaks": [
                {"start": "09:30", "end": "09:45", "name": "Break 1"},
                {"start": "12:00", "end": "12:30", "name": "Almuerzo"},
                {"start": "15:00", "end": "15:15", "name": "Break 2"}
            ],
            "shiftStart": "07:30",
            "shiftEnd": "17:30",
            "productiveHours": 9,  # 9 horas reales de trabajo
            "lineFlows": {"M1": "D1", "M2": "D2", "M3": "D3"}
        }
        
        # Combinar configuración por defecto con la proporcionada
        final_config = {**default_config, **config}
        
        # Generar planeación
        planeacion = calcular_planeacion_automatica(planes, final_config)
        
        return jsonify({
            "success": True,
            "date": fecha,
            "line": linea,
            "config": final_config,
            "planning": planeacion,
            "totalPlans": len(planes)
        })
        
    except Exception as e:
        return jsonify({"error": f"Error generando planeación: {str(e)}"}), 500
    finally:
        try:
            cursor.close()
            db.close()
        except:
            pass

def calcular_planeacion_automatica(planes, config):
    """
    Calcular planeación automática con lógica de tiempos, breaks y tiempo extra.
    """
    from datetime import datetime, timedelta
    
    def time_to_minutes(time_str):
        """Convertir tiempo HH:MM a minutos desde medianoche"""
        if not time_str:
            return 0
        try:
            hours, minutes = map(int, time_str.split(':'))
            return hours * 60 + minutes
        except:
            return 0
    
    def minutes_to_time(minutes):
        """Convertir minutos a formato HH:MM"""
        hours = minutes // 60
        mins = minutes % 60
        return f"{hours:02d}:{mins:02d}"
    
    def calculate_production_time(plan_count, uph):
        """Calcular tiempo de producción en minutos"""
        if not uph or uph == 0:
            return 0
        return round((plan_count / uph) * 60)
    
    # Agrupar planes por líneas conectadas según flujos configurados
    grupos = {}
    
    for plan in planes:
        line = plan["line"]
        group_name = line
        
        # Buscar si esta línea forma parte de un flujo
        if line in config["lineFlows"]:
            dest_line = config["lineFlows"][line]
            group_name = f"{line}-{dest_line}"
        else:
            # Verificar si es línea de destino
            for source, dest in config["lineFlows"].items():
                if dest == line:
                    group_name = f"{source}-{line}"
                    break
        
        if group_name not in grupos:
            grupos[group_name] = []
        grupos[group_name].append(plan)
    
    # Generar planeación para cada grupo
    resultado_planeacion = []
    
    for group_name, group_plans in grupos.items():
        # Header del grupo
        resultado_planeacion.append({
            "type": "group-header",
            "line": f"GRUPO {group_name}",
            "partNo": "",
            "quantity": "",
            "uph": "",
            "time": "",
            "break": "",
            "startTime": "",
            "endTime": "",
            "totalTime": "",
            "status": ""
        })
        
        # Procesar cada línea del grupo
        lines_in_group = list(set(plan["line"] for plan in group_plans))
        lines_in_group.sort()
        
        for line in lines_in_group:
            line_plans = [p for p in group_plans if p["line"] == line]
            
            # Ordenar por secuencia
            line_plans.sort(key=lambda x: (x.get("sequence") or 999999, x.get("created_at", "")))
            
            # Calcular horario de la línea
            current_time = time_to_minutes(config["shiftStart"])
            total_line_time = 0
            line_result = []
            
            for plan in line_plans:
                production_time = calculate_production_time(
                    plan.get("plan_count", 0), 
                    plan.get("uph", 0)
                )
                
                start_time = current_time
                planned_end_time = current_time + production_time
                
                # Verificar si algún break cae durante la producción de este plan
                breaks_during_plan = []
                for break_info in config["breaks"]:
                    break_start = time_to_minutes(break_info["start"])
                    break_end = time_to_minutes(break_info["end"])
                    
                    # Si el break cae durante este plan, lo insertamos
                    if break_start >= start_time and break_start < planned_end_time:
                        breaks_during_plan.append({
                            'start_minutes': break_start,
                            'end_minutes': break_end,
                            'duration': break_end - break_start,
                            'info': break_info
                        })
                
                # Ordenar breaks por hora de inicio
                breaks_during_plan.sort(key=lambda x: x['start_minutes'])
                
                # Calcular tiempo real considerando breaks
                actual_start_time = current_time
                actual_end_time = planned_end_time
                
                # Agregar duración de todos los breaks que caen durante este plan
                for break_data in breaks_during_plan:
                    actual_end_time += break_data['duration']
                
                # Agregar fila del plan
                line_result.append({
                    "type": "plan",
                    "line": line,
                    "partNo": plan.get("part_no", ""),
                    "quantity": plan.get("plan_count", 0),
                    "uph": plan.get("uph", 0),
                    "time": minutes_to_time(production_time),
                    "break": "",
                    "startTime": minutes_to_time(actual_start_time),
                    "endTime": minutes_to_time(actual_end_time),
                    "totalTime": "",
                    "status": "",
                    "lotNo": plan.get("lot_no", ""),
                    "woCode": plan.get("wo_code", "")
                })
                
                # Insertar filas de breaks que ocurrieron durante este plan
                for break_data in breaks_during_plan:
                    line_result.append({
                        "type": "break",
                        "line": line,
                        "partNo": "",
                        "quantity": "",
                        "uph": "",
                        "time": minutes_to_time(break_data['duration']),
                        "break": f"{break_data['info']['name']} ({break_data['info']['start']}-{break_data['info']['end']})",
                        "startTime": break_data['info']["start"],
                        "endTime": break_data['info']["end"],
                        "totalTime": "",
                        "status": ""
                    })
                
                current_time = actual_end_time
                total_line_time += production_time  # Solo tiempo productivo
            
            # Determinar si hay tiempo extra basado en tiempo productivo acumulado
            productive_minutes = config.get("productiveHours", 9) * 60  # 9 horas = 540 min
            has_overtime = total_line_time > productive_minutes
            status = "TIEMPO EXTRA" if has_overtime else "DIA"
            
            # Marcar filas que están en tiempo extra (cuando se acumulan más de 9h productivas)
            if has_overtime:
                accumulated_time = 0
                for row in line_result:
                    if row["type"] == "plan":
                        production_time = time_to_minutes(row["time"]) if row["time"] else 0
                        accumulated_time += production_time
                        if accumulated_time > productive_minutes:
                            row["isOvertime"] = True
            
            # Agregar filas de la línea al resultado
            resultado_planeacion.extend(line_result)
            
            # Agregar fila de totales
            total_hours = total_line_time / 60
            resultado_planeacion.append({
                "type": "total",
                "line": line,
                "partNo": f"TOTAL {line}",
                "quantity": "",
                "uph": "",
                "time": "",
                "break": "",
                "startTime": "",
                "endTime": "",
                "totalTime": f"{total_hours:.2f}h",
                "status": status
            })
    
    return resultado_planeacion

@app.route("/api/production-planning/config", methods=["GET", "POST"])
def configuracion_planeacion():
    """Obtener o actualizar configuración de planeación de producción"""
    
    if request.method == "GET":
        # Retornar configuración por defecto
        return jsonify({
            "success": True,
            "config": {
                "breaks": [
                    {"start": "09:30", "end": "09:45", "name": "Break 1"},
                    {"start": "12:00", "end": "12:30", "name": "Almuerzo"},
                    {"start": "15:00", "end": "15:15", "name": "Break 2"}
                ],
                "shiftStart": "07:30",
                "shiftEnd": "17:30",
                "productiveHours": 9,  # 9 horas reales de trabajo
                "lineFlows": {"M1": "D1", "M2": "D2", "M3": "D3"}
            }
        })
    
    elif request.method == "POST":
        # En una implementación completa, aquí se guardaría la configuración en BD
        # Por ahora solo validamos y retornamos la configuración recibida
        config = request.json or {}
        
        return jsonify({
            "success": True,
            "message": "Configuración actualizada",
            "config": config
        })

@app.route("/api/plan/export-excel", methods=["POST"])
def exportar_planes_excel():
    """Exporta los planes a un archivo Excel respetando los grupos del frontend con formato similar a la imagen"""
    try:
        # Obtener datos desde el frontend (ya ordenados y agrupados)
        data = request.get_json()
        if not data or 'plans' not in data:
            return jsonify({"error": "No se enviaron datos para exportar"}), 400
        
        plans = data.get('plans', [])
        
        if not plans:
            return jsonify({"error": "No hay datos para exportar"}), 400
            
        # Crear archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Plan Producción"
        
        # Estilos base
        header_font = Font(name='Arial', size=10, bold=True)
        data_font = Font(name='Arial', size=9)
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        border_thin = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Colores por línea (basado en la imagen de referencia)
        line_colors = {
            'M1': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),  # Amarillo
            'M2': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),  # Verde claro
            'M3': PatternFill(start_color='DA70D6', end_color='DA70D6', fill_type='solid'),  # Magenta
            'M4': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),  # Amarillo
            'D1': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),  # Verde
            'D2': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),  # Verde claro
            'D3': PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid'),  # Cian
            'H1': PatternFill(start_color='FF69B4', end_color='FF69B4', fill_type='solid'),  # Rosa
        }
        
        # Color por defecto y encabezado
        default_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Gris claro
        yellow_header = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        group_header = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # Gris para separadores
        
        # Crear encabezado principal con semana del año y fecha de México
        import pytz
        mexico_tz = pytz.timezone('America/Monterrey')
        fecha_mexico = datetime.now(mexico_tz)
        semana_del_ano = fecha_mexico.isocalendar()[1]  # Obtener semana del año
        fecha_formateada = fecha_mexico.strftime('%d %b')
        
        ws.merge_cells('A1:P1')
        header_text = f'★ W{semana_del_ano} Assy {fecha_formateada}'
        header_cell = ws.cell(row=1, column=1, value=header_text)
        header_cell.font = Font(name='Arial', size=12, bold=True)
        header_cell.fill = yellow_header
        header_cell.alignment = center_alignment
        header_cell.border = border_thin
        
        # Encabezados de columnas (fila 2) - Columnas combinadas del frontend y referencia
        headers = [
            'Línea', 'Lot No', 'Part No', 'Cantidad', 'CT', 'UPH', 'Tiempo', 
            'Inicio', 'Fin', 'Status', 'Project', 'Model', 'Process', 
            'WO Code', 'PO Code', 'Grupo'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = yellow_header
            cell.alignment = center_alignment
            cell.border = border_thin
        
        # Filtrar y organizar planes
        plans_data = []
        group_headers = []
        
        for item in plans:
            if item.get('isGroupHeader', False):
                # Es un marcador de grupo
                group_headers.append({
                    'title': item.get('groupTitle', ''),
                    'index': item.get('groupIndex', 0)
                })
            else:
                # Es un plan normal - solo incluir si tiene datos reales
                if (item.get('lot_no', '') or item.get('part_no', '') or 
                    item.get('plan_count', 0) > 0 or item.get('line', '')):
                    plans_data.append(item)
        
        # Procesar datos manteniendo el orden del frontend y respetando grupos
        current_row = 3
        processed_groups = set()
        group_totals = {}
        
        # Procesar cada plan válido
        for plan in plans_data:
            plan_group = plan.get('grupo', '').strip()
            group_index = plan.get('groupIndex', 0)
            line_name = plan.get('line', '')
            
            # Si no hay grupo definido, usar el índice de grupo
            if not plan_group:
                matching_header = next((h for h in group_headers if h['index'] == group_index), None)
                if matching_header:
                    plan_group = matching_header['title']
                else:
                    plan_group = f"GRUPO {group_index + 1}"
            
            # Si es un nuevo grupo, agregar encabezado de grupo
            if plan_group not in processed_groups:
                # Solo agregar separador si no es el primer grupo
                if processed_groups:
                    # Agregar fila separadora entre grupos
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=current_row, column=col, value='')
                        cell.fill = group_header
                        cell.border = border_thin
                    current_row += 1
                
                # Agregar fila con nombre del grupo
                ws.merge_cells(f'A{current_row}:P{current_row}')
                group_cell = ws.cell(row=current_row, column=1, value=plan_group)
                group_cell.font = Font(name='Arial', size=10, bold=True)
                group_cell.fill = group_header
                group_cell.alignment = center_alignment
                group_cell.border = border_thin
                current_row += 1
                
                processed_groups.add(plan_group)
            
            # Calcular tiempo de producción
            cantidad = int(plan.get('plan_count', 0))
            uph = int(plan.get('uph', 0)) if plan.get('uph', 0) else 1
            tiempo_minutos = (cantidad / uph * 60) if uph > 0 else 0
            tiempo_str = f"{int(tiempo_minutos//60)}.{int(tiempo_minutos%60):02d}"
            
            # Obtener color según línea
            line_fill = line_colors.get(line_name, default_fill)
            
            # Datos de la fila
            row_data = [
                line_name,  # Línea
                plan.get('lot_no', ''),  # Lot No
                plan.get('part_no', ''),  # Part No
                cantidad,  # Cantidad
                plan.get('ct', ''),  # CT
                uph,  # UPH
                tiempo_str,  # Tiempo
                plan.get('inicio', ''),  # Inicio
                plan.get('fin', ''),  # Fin
                plan.get('status', 'PLAN'),  # Status
                plan.get('project', ''),  # Project
                plan.get('model_code', ''),  # Model
                plan.get('process', ''),  # Process
                plan.get('wo_code', ''),  # WO Code
                plan.get('po_code', ''),  # PO Code
                plan.get('grupo', '')  # Grupo
            ]
            
            # Escribir fila
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = data_font
                cell.fill = line_fill
                cell.alignment = center_alignment if col in [1, 4, 5, 6, 7, 10] else left_alignment
                cell.border = border_thin
                
                # Aplicar formato especial a números
                if col in [4, 6]:  # Cantidad y UPH
                    cell.number_format = '#,##0'
            
            # Acumular totales por grupo
            if plan_group not in group_totals:
                group_totals[plan_group] = {'cantidad': 0, 'tiempo': 0}
            group_totals[plan_group]['cantidad'] += cantidad
            group_totals[plan_group]['tiempo'] += tiempo_minutos
            
            current_row += 1
        
        # Ajustar anchos de columna
        column_widths = [6, 12, 18, 8, 6, 6, 8, 10, 10, 10, 12, 15, 10, 12, 12, 8]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Congelar primeras dos filas
        ws.freeze_panes = 'A3'
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo
        fecha_archivo = fecha_mexico.strftime('%Y%m%d_%H%M')
        filename = f'W{semana_del_ano}_Plan_Produccion_{fecha_archivo}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        # Error al generar Excel
        return jsonify({"error": f"Error al generar el archivo Excel: {str(e)}"}), 500
        ws.freeze_panes = 'A2'
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo con fecha
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'Plan_Produccion_{fecha_actual}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        # Error al generar Excel
        return jsonify({"error": f"Error al generar el archivo Excel: {str(e)}"}), 500

@app.route("/api/plan/pending", methods=["GET"])
def obtener_planes_pendientes():
    """Obtener planes con input pendiente en un rango de fechas"""
    try:
        start_date = request.args.get('start')
        end_date = request.args.get('end')
        
        if not start_date or not end_date:
            return jsonify({"error": "Se requieren fechas de inicio y fin"}), 400
        
        db = get_db()
        cursor = db.cursor(dictionary=True)
        
        # Primero verificar qué columnas existen
        cursor.execute("SHOW COLUMNS FROM plan_main")
        columns = [row['Field'] for row in cursor.fetchall()]
        
        # Determinar qué columnas usar para input
        input_column = None
        if 'input' in columns:
            input_column = 'input'
        elif 'produced' in columns:
            input_column = 'produced'
        elif 'count_input' in columns:
            input_column = 'count_input'
        else:
            # Si no hay columna de input, buscar planes recientes sin producción
            input_column = '0'  # Usar valor fijo para mostrar todos como pendientes
        
        # Consultar planes donde el input sea menor al plan_count (hay pendiente)
        # y que no estén cancelados
        if input_column == '0':
            query = f"""
                SELECT lot_no, working_date, part_no, line, plan_count, 
                       0 as input, status, wo_code, po_code
                FROM plan_main 
                WHERE working_date BETWEEN %s AND %s 
                AND status != 'CANCELADO'
                ORDER BY working_date DESC, lot_no
                LIMIT 50
            """
        else:
            query = f"""
                SELECT lot_no, working_date, part_no, line, plan_count, 
                       COALESCE({input_column}, 0) as input, status, wo_code, po_code
                FROM plan_main 
                WHERE working_date BETWEEN %s AND %s 
                AND status != 'CANCELADO'
                AND (COALESCE({input_column}, 0) < plan_count OR {input_column} IS NULL)
                ORDER BY working_date DESC, lot_no
                LIMIT 50
            """
        
        cursor.execute(query, (start_date, end_date))
        planes = cursor.fetchall()
        
        # Formatear fechas
        for plan in planes:
            if plan.get("working_date"):
                if isinstance(plan["working_date"], date):
                    plan["working_date"] = plan["working_date"].strftime("%Y-%m-%d")
                else:
                    plan["working_date"] = str(plan["working_date"])[:10]
            
            # Asegurar que input tenga un valor
            if plan.get("input") is None:
                plan["input"] = 0
        
        cursor.close()
        

        return jsonify(planes)
        
    except Exception as e:
        # Error al obtener planes pendientes

        return jsonify({"error": f"Error al obtener planes pendientes: {str(e)}"}), 500

@app.route("/api/plan/reschedule", methods=["POST"])
def reprogramar_planes():
    """Reprogramar planes cambiando su fecha de trabajo"""
    try:
        data = request.get_json()
        lot_nos = data.get('lot_nos', [])
        new_working_date = data.get('new_working_date')
        
        if not lot_nos or not new_working_date:
            return jsonify({"error": "Se requieren los LOT NOs y la nueva fecha"}), 400
        
        if not isinstance(lot_nos, list) or len(lot_nos) == 0:
            return jsonify({"error": "Debe proporcionar al menos un LOT NO"}), 400
        
        # Validar formato de fecha
        try:
            datetime.strptime(new_working_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({"error": "Formato de fecha inválido. Use YYYY-MM-DD"}), 400
        
        db = get_db()
        cursor = db.cursor()
        
        # Verificar que los planes existen y no están cancelados
        format_strings = ','.join(['%s'] * len(lot_nos))
        check_query = f"""
            SELECT lot_no, status FROM plan_main 
            WHERE lot_no IN ({format_strings})
        """
        cursor.execute(check_query, lot_nos)
        existing_plans = cursor.fetchall()
        
        if len(existing_plans) != len(lot_nos):
            return jsonify({"error": "Algunos LOT NOs no existen"}), 400
        
        # Verificar que ninguno esté cancelado
        for plan in existing_plans:
            if plan[1] == 'CANCELADO':
                return jsonify({"error": f"El plan {plan[0]} está cancelado y no se puede reprogramar"}), 400
        
        # Actualizar la fecha de trabajo
        update_query = f"""
            UPDATE plan_main 
            SET working_date = %s,
                updated_at = NOW()
            WHERE lot_no IN ({format_strings})
        """
        
        params = [new_working_date] + lot_nos
        cursor.execute(update_query, params)
        
        affected_rows = cursor.rowcount
        db.commit()
        cursor.close()
        
        return jsonify({
            "message": f"Se reprogramaron {affected_rows} planes exitosamente",
            "reprogrammed_count": affected_rows,
            "new_date": new_working_date
        })
        
    except Exception as e:
        # Error al reprogramar planes
        return jsonify({"error": f"Error al reprogramar planes: {str(e)}"}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=4000, debug=True)
