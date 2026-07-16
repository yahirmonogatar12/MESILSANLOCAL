import io
from datetime import date, datetime, timedelta

import openpyxl
import pytest

from app.api.control_produccion.part_planning import (
    _parse_lg_workbook,
    _parse_part10_workbook,
    _pp_filas_segun_modo,
    _pp_ref_lunes,
)


def _workbook_bytes(sheet_name="LG", fechas=None, filas=None):
    """Construye un xlsx en memoria con el layout del plan LG.

    fechas: lista de valores para la fila 3 desde la columna C.
    filas: lista de (part_no, [cantidades]) desde la fila 4.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for i, valor in enumerate(fechas or []):
        ws.cell(row=3, column=3 + i, value=valor)
    for r, (parte, cantidades) in enumerate(filas or [], start=4):
        ws.cell(row=r, column=2, value=parte)
        for i, qty in enumerate(cantidades):
            ws.cell(row=r, column=3 + i, value=qty)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_fechas_datetime_completas():
    data = _workbook_bytes(
        fechas=[datetime(2026, 7, 9), datetime(2026, 7, 10)],
        filas=[("MDS64531211", [536, 700]), ("EAV64632103", [0, 1356])],
    )
    parsed, err = _parse_lg_workbook(data, "plan.xlsx")
    assert err is None
    assert parsed["plan_year"] == 2026
    assert parsed["date_from"] == date(2026, 7, 9)
    assert parsed["date_to"] == date(2026, 7, 10)
    assert parsed["parts_count"] == 2
    assert parsed["records_count"] == 3  # 536, 700, 1356
    assert parsed["zero_records_count"] == 1
    assert parsed["records"][("MDS64531211", date(2026, 7, 9))] == 536


def test_parse_fechas_string_toma_anio_del_filename():
    data = _workbook_bytes(fechas=["09-Jul"], filas=[("P1", [10])])
    parsed, err = _parse_lg_workbook(data, "plan de produccion 20260710 W28.xlsx")
    assert err is None
    assert parsed["records"][("P1", date(2026, 7, 9))] == 10


def test_parse_override_year_gana_sobre_filename():
    data = _workbook_bytes(fechas=["09-Jul"], filas=[("P1", [10])])
    parsed, err = _parse_lg_workbook(data, "plan 20260710.xlsx", override_year=2027)
    assert err is None
    assert ("P1", date(2027, 7, 9)) in parsed["records"]


def test_partes_duplicadas_se_consolidan_sumando():
    data = _workbook_bytes(
        fechas=[datetime(2026, 7, 9)],
        filas=[("MDS64531211", [536]), ("MDS64531211", [120])],
    )
    parsed, err = _parse_lg_workbook(data, "plan.xlsx")
    assert err is None
    assert parsed["records"][("MDS64531211", date(2026, 7, 9))] == 656
    assert parsed["duplicated_parts"] == ["MDS64531211"]
    assert parsed["parts_count"] == 1


def test_cantidad_negativa_se_ajusta_a_cero_con_warning():
    data = _workbook_bytes(
        fechas=[datetime(2026, 7, 9)], filas=[("P1", [-5]), ("P2", [3])]
    )
    parsed, err = _parse_lg_workbook(data, "plan.xlsx")
    assert err is None
    assert parsed["records"][("P1", date(2026, 7, 9))] == 0
    assert any("negativa" in w for w in parsed["warnings"])


def test_errores_bloqueantes():
    # Extension invalida
    _, err = _parse_lg_workbook(b"x", "plan.txt")
    assert err[1] == 400 and "Formato" in err[0]["errors"][0]
    # Vacio
    _, err = _parse_lg_workbook(b"", "plan.xlsx")
    assert "vacio" in err[0]["errors"][0]
    # No es Excel
    _, err = _parse_lg_workbook(b"no soy zip", "plan.xlsx")
    assert "corrupto" in err[0]["errors"][0]
    # Sin hoja LG
    data = _workbook_bytes(sheet_name="Otra", fechas=[datetime(2026, 7, 9)], filas=[("P1", [1])])
    _, err = _parse_lg_workbook(data, "plan.xlsx")
    assert "hoja 'LG'" in err[0]["errors"][0]
    # Sin fechas validas
    data = _workbook_bytes(fechas=["no-fecha"], filas=[("P1", [1])])
    _, err = _parse_lg_workbook(data, "plan.xlsx")
    assert "fechas validas" in err[0]["errors"][0]
    # Sin partes
    data = _workbook_bytes(fechas=[datetime(2026, 7, 9)], filas=[])
    _, err = _parse_lg_workbook(data, "plan.xlsx")
    assert "numeros de parte" in err[0]["errors"][0]


def _part10_bytes():
    """Hoja 'Part 10' minima: encabezado con PART NUMBER + bloque P/S/I."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Part 10"
    # encabezado (fila 30 en el real; la posicion no importa, se busca)
    ws.cell(row=5, column=3, value="PART NUMBER")
    ws.cell(row=5, column=17, value=datetime(2026, 6, 29))
    ws.cell(row=5, column=18, value=datetime(2026, 6, 30))
    # bloque P/S/I
    ws.cell(row=6, column=3, value="EBR80757421")
    ws.cell(row=6, column=4, value="EAX66946005")
    ws.cell(row=6, column=5, value=399)   # LGEMM
    ws.cell(row=6, column=6, value=408)   # ISEMM
    ws.cell(row=6, column=7, value=-63)   # SVC
    ws.cell(row=6, column=11, value=15)   # smt
    ws.cell(row=6, column=12, value=8)    # imd
    ws.cell(row=6, column=13, value="R1")
    ws.cell(row=6, column=14, value="P")
    ws.cell(row=7, column=14, value="S")
    ws.cell(row=7, column=18, value=540)  # schedule 06-30
    ws.cell(row=8, column=14, value="I")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_part10():
    parsed, err = _parse_part10_workbook(_part10_bytes(), "plan.xlsx")
    assert err is None
    assert parsed["ref_date"] == date(2026, 6, 29)
    inv = parsed["inventory"]["EBR80757421"]
    assert inv["lgemm"] == 399 and inv["isemm"] == 408 and inv["svc"] == -63
    assert inv["smt"] == 15 and inv["imd"] == 8
    assert inv["line"] == "R1" and inv["board"] == "EAX66946005"
    assert parsed["schedules"][("EBR80757421", date(2026, 6, 30))] == 540


def test_parse_part10_sin_hoja():
    data = _workbook_bytes(fechas=[datetime(2026, 7, 9)], filas=[("P1", [1])])
    _, err = _parse_part10_workbook(data, "plan.xlsx")
    assert "Part 10" in err[0]["errors"][0]


def test_filas_segun_modo():
    data = _workbook_bytes(
        fechas=[datetime(2026, 7, 9), datetime(2026, 7, 10)],
        filas=[("P1", [5, 0])],
    )
    parsed, _ = _parse_lg_workbook(data, "plan.xlsx")

    todas = _pp_filas_segun_modo(parsed, "upsert", include_zero=True)
    assert len(todas) == 2
    sin_ceros = _pp_filas_segun_modo(parsed, "upsert", include_zero=False)
    assert len(sin_ceros) == 1 and sin_ceros[0][2] == 5
    solo_positivos = _pp_filas_segun_modo(parsed, "only_positive", include_zero=True)
    assert len(solo_positivos) == 1


# ---------- Plan Proyectado (generador de lotes tipo hoja LOTE N) ----------

def test_ppy_qty_pack_redondeo_caja_cerrada():
    from app.api.control_produccion.part_planning import _ppy_qty_pack

    # faltante 510 +10% = 561 -> multiplo de 20 arriba = 580
    assert _ppy_qty_pack(510, 20) == 580
    # sin pack (None/0) -> solo el +10% redondeado arriba
    assert _ppy_qty_pack(510, None) == 561
    assert _ppy_qty_pack(510, 0) == 561
    # exacto: 100 +10% = 110, pack 10 -> 110
    assert _ppy_qty_pack(100, 10) == 110
    # siempre cierra caja aunque el 10% ya pase el multiplo
    assert _ppy_qty_pack(95, 20) == 120  # 104.5 -> 120


def test_ppy_familia_y_uph():
    from app.api.control_produccion.part_planning import _ppy_familia, _ppy_parse_uph

    assert _ppy_familia("EBR80757421") == "EBR807574"
    assert _ppy_familia("EBR80757422") == _ppy_familia("EBR80757412")
    assert _ppy_parse_uph("240") == 240
    assert _ppy_parse_uph(240.0) == 240
    assert _ppy_parse_uph("") is None
    assert _ppy_parse_uph(None) is None
    assert _ppy_parse_uph("0") is None
    assert _ppy_parse_uph("N/A") is None


def test_ppy_armar_lotes_capacidad_y_adelantos():
    from app.api.control_produccion.part_planning import _ppy_armar_lotes

    hoy = date(2026, 7, 14)
    manana = date(2026, 7, 15)

    def cand(part, falt_total, falt_hoy, primera, uph=100, pack=20, line=None):
        return {"part_no": part, "falt_total": falt_total, "falt_hoy": falt_hoy,
                "primera_falta": primera, "line": line, "uph": uph, "pack": pack,
                "ct": 36, "model": None, "main_sub": None}

    # 1) Prioridad: faltante de HOY entra antes que el adelanto cuando no hay horas
    # las horas van por GRUPO: M1 pertenece a G1 (ver PPY_GRUPOS_CAPACIDAD)
    horas = {"G1": 1.0}  # 1 h a 100 uph = 100 pzs
    lotes, fuera = _ppy_armar_lotes(
        [cand("FUTURO0001", 500, 0, manana), cand("URGENTE001", 500, 500, hoy)],
        ["M1"], horas)
    assert lotes[0]["part_no"] == "URGENTE001"
    assert lotes[0]["qty"] == 100  # parcial a caja cerrada por horas
    assert "parcial" in lotes[0]["comentario"]
    assert fuera and fuera[0]["part_no"] == "FUTURO0001"

    # 2) Adelanto: cabe completo (+10% a pack) y trae comentario de adelanto
    horas = {"G1": 9.0}
    lotes, fuera = _ppy_armar_lotes([cand("FUTURO0001", 500, 0, manana)], ["M1"], horas)
    assert not fuera
    assert lotes[0]["qty"] == 560  # 550 -> pack 20 -> 560
    assert lotes[0]["comentario"].startswith("Adelanto")
    assert abs(horas["G1"] - (9.0 - 5.6)) < 1e-6

    # 3) Familia junta en la misma linea aunque otra este mas libre
    horas = {"G1": 9.0, "G2": 9.0}
    lotes, _ = _ppy_armar_lotes(
        [cand("EBR8075741A", 400, 400, hoy, line="M1"),
         cand("EBR8075742B", 400, 400, hoy)],
        ["M1", "M2"], horas)
    assert {l["part_no"][:9] for l in lotes} == {"EBR807574"}
    assert len({l["linea"] for l in lotes}) == 1  # misma linea la familia

    # 4) Sin UPH o sin pack: la parte NO se genera (visible en no_incluidos)
    horas = {"G1": 9.0}
    lotes, fuera = _ppy_armar_lotes(
        [cand("SINUPH0001", 100, 100, hoy, uph=None),
         cand("SINPACK001", 100, 100, hoy, pack=None),
         cand("SINNADA001", 100, 0, manana, uph=None, pack=None),
         cand("COMPLETA01", 100, 100, hoy)],
        ["M1"], horas)
    assert [l["part_no"] for l in lotes] == ["COMPLETA01"]
    motivos = {f["part_no"]: f["motivo"] for f in fuera}
    assert "UPH" in motivos["SINUPH0001"]
    assert "pack" in motivos["SINPACK001"]
    assert "UPH" in motivos["SINNADA001"] and "pack" in motivos["SINNADA001"]


def test_ppy_armar_lotes_lineas_permitidas():
    from app.api.control_produccion.part_planning import (
        _ppy_armar_lotes, _ppy_horas_iniciales,
    )

    hoy = date(2026, 7, 14)

    def cand(part, permitidas=None, line=None):
        return {"part_no": part, "falt_total": 100, "falt_hoy": 100,
                "primera_falta": hoy, "line": line, "uph": 100, "pack": 20,
                "ct": 36, "model": None, "main_sub": None, "permitidas": permitidas}

    # Solo puede ir a M3; la otra, sin regla, cae en la primera activa.
    # Las horas salen de los bloques (flexibles), la linea la fija la regla.
    horas = _ppy_horas_iniciales(["M2", "M3"])
    lotes, fuera = _ppy_armar_lotes(
        [cand("SOLOM30001", permitidas=["M3"]),
         cand("LIBRE00001")],
        ["M2", "M3"], horas)
    por_parte = {l["part_no"]: l["linea"] for l in lotes}
    assert por_parte["SOLOM30001"] == "M3"
    assert por_parte["LIBRE00001"] in ("M2", "M3")

    # Sus lineas no estan activas -> fuera con motivo claro
    horas = _ppy_horas_iniciales(["M2"])
    lotes, fuera = _ppy_armar_lotes([cand("SOLOH10001", permitidas=["H1"])], ["M2"], horas)
    assert not lotes
    assert "H1" in fuera[0]["motivo"] and "no estan activas" in fuera[0]["motivo"]

    # Sin horas en ningun bloque -> fuera, y lo dice
    horas = {"B1": 0.0}
    lotes, fuera = _ppy_armar_lotes([cand("SOLOM20002", permitidas=["M2"])], ["M2"], horas)
    assert not lotes
    assert "sin horas" in fuera[0]["motivo"]


def test_ppy_armar_lotes_modo_laxo_propuesta_schedule():
    from app.api.control_produccion.part_planning import _ppy_armar_lotes

    hoy = date(2026, 7, 14)

    def cand(part, uph=100, pack=20):
        return {"part_no": part, "falt_total": 100, "falt_hoy": 100,
                "primera_falta": hoy, "line": None, "uph": uph, "pack": pack,
                "model": None, "main_sub": None}

    # Sin pack -> asume 20; sin UPH -> entra sin consumir horas
    horas = {"G1": 9.0}
    lotes, fuera = _ppy_armar_lotes(
        [cand("SINPACK001", pack=None),
         cand("SINUPH0001", uph=None, pack=None),
         cand("NORMAL0001")],
        ["M1"], horas, estricto=False)
    assert not fuera
    por_parte = {l["part_no"]: l for l in lotes}
    assert por_parte["SINPACK001"]["qty"] == 120   # 110 -> pack 20 -> 120
    assert por_parte["SINUPH0001"]["qty"] == 120   # mismo calculo, sin horas
    # solo las partes con UPH consumen horas del grupo (120/100 + 120/100 = 2.4)
    assert abs(horas["G1"] - (9.0 - 2.4)) < 1e-6


def test_ppy_acomodo_heuristico_asigna_linea_por_id():
    from app.api.control_produccion.part_planning import _ppy_acomodo_heuristico

    pend = [
        {"id": 1, "part_no": "EBR8075741A", "qty_plan": 200, "uph": 100,
         "estandar_pack": 20, "model": None, "main_sub": None, "permitidas": ["M1"]},
        {"id": 2, "part_no": "EBR8075742B", "qty_plan": 200, "uph": 100,
         "estandar_pack": 20, "model": None, "main_sub": None, "permitidas": None},
        {"id": 3, "part_no": "ABC00000001", "qty_plan": 100, "uph": 100,
         "estandar_pack": 20, "model": None, "main_sub": None, "permitidas": ["H1"]},
    ]
    horas = {"G1": 9.0, "G2": 9.0}
    asig = _ppy_acomodo_heuristico(pend, ["M1", "M2"], horas)
    # id 1 permitido solo M1; familia (1 y 2 = EBR807574) juntas en M1
    assert asig[1] == "M1"
    assert asig[2] == "M1"
    # id 3 solo permite H1 que no esta activa -> no se asigna
    assert 3 not in asig
    # no muta el dict original de horas del caller
    assert horas["G1"] == 9.0


def _mock_ppy_schedule_dependencies(monkeypatch, proyeccion, raw, lineas=("M1",)):
    """Aisla el simulador de MySQL para probar solamente el motor MRP."""
    import app.api.control_produccion.part_planning as pp

    monkeypatch.setattr(pp, "_ppy_config_lineas", lambda *_a, **_k: list(lineas))
    monkeypatch.setattr(pp, "_ppy_proyeccion_rango", lambda *_a, **_k: proyeccion)
    monkeypatch.setattr(pp, "_ppy_datos_raw", lambda *_a, **_k: raw)
    monkeypatch.setattr(pp, "_ppy_lineas_permitidas_map", lambda *_a, **_k: ({}, {}))
    # El simulador consulta demanda sin inventario y lotes confirmados. En
    # estos escenarios ambos conjuntos estan vacios.
    monkeypatch.setattr(pp, "execute_query", lambda *_args, **_kwargs: [])
    return pp


def test_ppy_simular_schedule_contrato_detallado(monkeypatch):
    fecha = date(2026, 7, 15)
    parte = "EBR80757421"
    pp = _mock_ppy_schedule_dependencies(
        monkeypatch,
        {
            parte: {
                "line": "M1",
                "proj": {fecha: -100},
            }
        },
        {
            parte: {
                "model": "QT-T",
                "sub_assy": "MAIN",
                "c_t": 36,
                "uph": 100,
                "estandar_pack": 20,
            }
        },
    )

    propuestas, omitidas, excepciones = pp._ppy_simular_schedule(
        fecha, fecha, detailed=True
    )

    assert omitidas == []
    assert excepciones == []
    assert len(propuestas) == 1
    item = propuestas[0]
    assert item == {
        "fecha": "2026-07-15",
        "linea": "M1",
        # el lote consume las horas de un BLOQUE (4 turnos flexibles de 9 h);
        # la linea la fija assy_line, el bloque es solo el cupo
        "grupo": "B1",
        "turno": "DIA",
        "numero_parte": parte,
        "cantidad": 180,
        "ct": 36.0,
        "uph": 100,
        "horas_requeridas": 1.8,   # 180 pzs / 100 uph
        "inventario_antes": -100,
        "inventario_despues": 80,  # -100 + 180: queda por encima del remain de 60
        "fecha_shortage": "2026-07-15",
        "prioridad": 1,
        "motivo": "Evitar shortage del plan LG",
        "requiere_aprobacion": False,
        "excepciones": [],
        "model": "QT-T",
        "main_sub": "MAIN",
        "pack_size": 20,
        "faltante": 160,
        # Contrato retrocompatible con la cuadricula de Proyeccion.
        "part_no": parte,
        "sched_date": "2026-07-15",
        "qty": 180,
    }


def test_ppy_simular_schedule_reporta_raw_incompleto(monkeypatch):
    fecha = date(2026, 7, 15)
    sin_uph = "SINUPH0001"
    sin_pack = "SINPACK001"
    pp = _mock_ppy_schedule_dependencies(
        monkeypatch,
        {
            sin_uph: {"line": "M1", "proj": {fecha: -100}},
            sin_pack: {"line": "M1", "proj": {fecha: -80}},
        },
        {
            sin_uph: {
                "model": None,
                "sub_assy": None,
                "c_t": 36,
                "uph": None,
                "estandar_pack": 20,
            },
            sin_pack: {
                "model": None,
                "sub_assy": None,
                "c_t": 36,
                "uph": 100,
                "estandar_pack": None,
            },
        },
    )

    propuestas, omitidas, excepciones = pp._ppy_simular_schedule(
        fecha, fecha, detailed=True
    )

    assert propuestas == []
    assert any(sin_uph in texto and "UPH" in texto for texto in omitidas)
    assert any(sin_pack in texto and "pack" in texto for texto in omitidas)
    por_parte = {(item["part_no"], item["code"]) for item in excepciones}
    assert (sin_uph, "UPH_MISSING") in por_parte
    assert (sin_pack, "PACK_MISSING") in por_parte
    assert all(item["date"] == fecha.isoformat() for item in excepciones)


def test_ppy_simular_schedule_limita_capacidad_y_marca_aprobacion(monkeypatch):
    fecha = date(2026, 7, 15)
    parte = "CAPACIDAD01"
    pp = _mock_ppy_schedule_dependencies(
        monkeypatch,
        {parte: {"line": "M1", "proj": {fecha: -1000}}},
        {
            parte: {
                "model": None,
                "sub_assy": None,
                "c_t": 36,
                "uph": 100,
                "estandar_pack": 20,
            }
        },
    )

    primera = pp._ppy_simular_schedule(fecha, fecha, detailed=True)
    segunda = pp._ppy_simular_schedule(fecha, fecha, detailed=True)

    # La propuesta es deterministica y nunca rebasa las 9 h del turno DIA.
    assert primera == segunda
    item = primera[0][0]
    assert item["cantidad"] == 900
    assert item["horas_requeridas"] == 9.0
    assert item["inventario_despues"] == -100
    assert item["requiere_aprobacion"] is True
    assert item["excepciones"] == ["PARTIAL_CAPACITY", "SHORTAGE_REMAINS"]


def test_ppy_simular_schedule_adelanta_shortage_dentro_del_horizonte(monkeypatch):
    """El lote se programa 2 dias de PRODUCCION antes del faltante.

    Es el caso que planteo planning con el 5505: el faltante cae el sabado 18
    (LG consume, ISEMM no produce), asi que el lote NO se programa el miercoles
    15 (seria 3 dias antes) sino el jueves 16, que es 2 dias de produccion
    antes contando vie 17 y sab 18.
    """
    fecha = date(2026, 7, 15)              # miercoles
    fecha_shortage = date(2026, 7, 18)     # sabado
    esperado = date(2026, 7, 16)           # jueves: 2 dias de produccion antes
    parte = "ADELANTO001"
    # 500 no es faltante; el faltante empieza el sabado
    proyeccion = {
        fecha + timedelta(days=offset): (500 if offset < 3 else -100)
        for offset in range(0, 8)
    }
    pp = _mock_ppy_schedule_dependencies(
        monkeypatch,
        {parte: {"line": "M1", "proj": proyeccion}},
        {
            parte: {
                "model": "MODELO-A",
                "sub_assy": "MAIN",
                "c_t": 36,
                "uph": 100,
                "estandar_pack": 20,
            }
        },
    )

    propuestas, omitidas, excepciones = pp._ppy_simular_schedule(
        fecha, fecha + timedelta(days=7), detailed=True
    )

    assert omitidas == []
    assert excepciones == []
    assert len(propuestas) == 1
    item = propuestas[0]
    assert item["sched_date"] == esperado.isoformat(), (
        "el faltante del sabado se produce el jueves, no antes")
    assert item["fecha_shortage"] == fecha_shortage.isoformat()
    assert item["qty"] == 180   # faltante 160 (=60-(-100)) -> +10% -> caja 20
    assert item["motivo"] == "Adelantar produccion para evitar shortage del plan LG"


def test_ppy_simular_schedule_deriva_ct_desde_uph(monkeypatch):
    fecha = date(2026, 7, 15)
    parte = "SINCT000001"
    pp = _mock_ppy_schedule_dependencies(
        monkeypatch,
        {parte: {"line": "M1", "proj": {fecha: -100}}},
        {
            parte: {
                "model": None,
                "sub_assy": None,
                "c_t": None,
                "uph": 100,
                "estandar_pack": 20,
            }
        },
    )

    propuestas, omitidas, excepciones = pp._ppy_simular_schedule(
        fecha, fecha, detailed=True
    )

    assert omitidas == []
    assert excepciones == []
    assert len(propuestas) == 1
    assert propuestas[0]["numero_parte"] == parte
    assert propuestas[0]["ct"] == 36.0  # 3600 / 100 UPH


def test_ppy_configuracion_de_lineas_propaga_errores_sql():
    from app.api.control_produccion.part_planning import (
        _ppy_config_lineas,
        _ppy_lineas_permitidas_map,
    )

    def sql_indisponible(*_args, **_kwargs):
        raise RuntimeError("mysql indisponible")

    with pytest.raises(RuntimeError, match="mysql indisponible"):
        _ppy_config_lineas(query=sql_indisponible)
    with pytest.raises(RuntimeError, match="mysql indisponible"):
        _ppy_lineas_permitidas_map(query=sql_indisponible)


def test_ppy_hash_y_resumen_lineas_son_estables():
    from app.api.control_produccion.part_planning import (
        _ppy_proposal_hash,
        _ppy_resumen_lineas,
    )

    fecha = date(2026, 7, 15)
    base = {
        "fecha": fecha.isoformat(),
        "linea": "M1",
        "turno": "DIA",
        "part_no": "P1",
        "qty": 100,
        "horas_requeridas": 1.25,
    }
    segunda = {**base, "part_no": "P2", "qty": 200, "horas_requeridas": 2.5}

    hash_a = _ppy_proposal_hash(fecha, fecha, [base, segunda], [])
    # El orden de claves de los dicts no altera la huella JSON canonica.
    hash_b = _ppy_proposal_hash(
        fecha, fecha, [dict(reversed(list(base.items()))), segunda], []
    )
    hash_c = _ppy_proposal_hash(fecha, fecha, [{**base, "qty": 101}, segunda], [])
    assert hash_a == hash_b
    assert hash_a != hash_c
    assert len(hash_a) == 64

    assert _ppy_resumen_lineas([base, segunda]) == [
        {
            "fecha": "2026-07-15",
            "linea": "M1",
            "turno": "DIA",
            "horas": 3.75,
            "horas_max": 9.0,
            "lotes": 2,
            "cantidad": 300,
            "excede": False,
        }
    ]


def test_partes_excluidas_de_propuesta_se_normalizan_y_validan():
    from app.api.control_produccion.part_planning import (
        _ppy_normalizar_partes_excluidas,
        _ppy_partes_excluidas_propuesta,
    )

    assert _ppy_normalizar_partes_excluidas(
        [" ebr30299365 ", "EBR30299369", "ebr30299365"]
    ) == ["EBR30299365", "EBR30299369"]
    assert _ppy_partes_excluidas_propuesta(
        '["EBR30299365", "EBR30299369"]'
    ) == ["EBR30299365", "EBR30299369"]
    assert _ppy_partes_excluidas_propuesta(None) == []
    with pytest.raises(ValueError, match="Numero de parte invalido"):
        _ppy_normalizar_partes_excluidas(["EBR 30299365"])


def test_ref_lunes_sale_del_nombre_no_de_hoy():
    # 20260715 (mie, W29) -> lunes 2026-07-13. Validado contra la fila I de
    # "Part 15": con este lunes cuadran 404/404 partes, con el de hoy no.
    assert _pp_ref_lunes(
        "★plan de produccion ILSAN 20260715 W29 ver03.xlsm"
    ) == date(2026, 7, 13)
    # un lunes se queda en si mismo
    assert _pp_ref_lunes("plan ILSAN 20260713 W29 ver01.xlsm") == date(2026, 7, 13)
    # sin fecha usable en el nombre -> lunes actual
    hoy = date.today()
    assert _pp_ref_lunes("plan.xlsm") == hoy - timedelta(days=hoy.weekday())
    assert _pp_ref_lunes("plan ILSAN 20261332 W29.xlsm") == hoy - timedelta(days=hoy.weekday())


def test_ppy_assy_line_manda_no_cae_en_otra_linea():
    """La linea del lote la fija assy_line, no el hueco libre.

    Caso real (plan 15/07/2026): M2 se lleno con EBR23966211 y el fallback
    "la linea mas libre" mandaba EBR37437070 (assy_line=M2) a H1, la linea de
    harness. Una tarjeta main no se produce ahi.
    """
    from app.api.control_produccion.part_planning import (
        _ppy_armar_lotes, _ppy_horas_iniciales,
    )

    hoy = date(2026, 7, 15)

    def cand(part, line, uph=200, pack=10):
        return {"part_no": part, "falt_total": 1800, "falt_hoy": 1800,
                "primera_falta": hoy, "line": line, "uph": uph, "pack": pack,
                "ct": 18, "model": None, "main_sub": None}

    # con H1 activa y libre, la parte de M2 NO debe acabar en H1
    horas = _ppy_horas_iniciales(["M2", "H1"])
    lotes, fuera = _ppy_armar_lotes([cand("EBR37437070", "M2")], ["M2", "H1"], horas)
    assert lotes[0]["linea"] == "M2", "la linea la manda assy_line, no el hueco libre"
    assert not fuera

    # lineas_permitidas (si esta capturada) gana sobre assy_line
    c = cand("EBR30299370", "H1")
    c["permitidas"] = ["M2"]
    horas = _ppy_horas_iniciales(["M2", "H1"])
    lotes, _f = _ppy_armar_lotes([c], ["M2", "H1"], horas)
    assert lotes[0]["linea"] == "M2"

    # una parte sin linea alguna cae en alguna activa
    horas = _ppy_horas_iniciales(["M1", "M2"])
    lotes, _f = _ppy_armar_lotes([cand("SINLINEA01", None)], ["M1", "M2"], horas)
    assert lotes[0]["linea"] in ("M1", "M2")


def test_ppy_sumar_dias_consumo_cuenta_sabado_no_domingo():
    """El horizonte cuenta los dias en que LG consume, no en los que se produce.

    LG consume lunes a sabado (el sabado ~9k pzs); el domingo nunca (0 demanda
    en todo el plan). ISEMM produce L-V. El faltante lo genera el consumo, asi
    que el sabado SI cuenta para el horizonte aunque no se produzca ese dia.
    """
    from app.api.control_produccion.part_planning import _ppy_sumar_dias_consumo

    jueves = date(2026, 7, 16)
    # vie 17 (1), sab 18 (2) -> LG consume, [dom 19 no], lun 20 (3), mar 21 (4)
    assert _ppy_sumar_dias_consumo(jueves, 1) == date(2026, 7, 17)
    assert _ppy_sumar_dias_consumo(jueves, 2) == date(2026, 7, 18)  # sabado cuenta
    assert _ppy_sumar_dias_consumo(jueves, 3) == date(2026, 7, 20)  # salta domingo
    assert _ppy_sumar_dias_consumo(jueves, 4) == date(2026, 7, 21)
    # desde viernes: el sabado es el dia 1, el domingo se salta
    viernes = date(2026, 7, 17)
    assert _ppy_sumar_dias_consumo(viernes, 1) == date(2026, 7, 18)  # sab
    assert _ppy_sumar_dias_consumo(viernes, 2) == date(2026, 7, 20)  # lun
    # desde lunes, corrido hasta el sabado
    lunes = date(2026, 7, 13)
    assert _ppy_sumar_dias_consumo(lunes, 5) == date(2026, 7, 18)  # sab
    assert _ppy_sumar_dias_consumo(lunes, 6) == date(2026, 7, 20)  # lun
    # desde sabado, el siguiente dia de consumo es el lunes
    assert _ppy_sumar_dias_consumo(date(2026, 7, 18), 1) == date(2026, 7, 20)
    assert _ppy_sumar_dias_consumo(jueves, 0) == jueves


def test_ppy_anticipacion_por_linea():
    """Un lote se adelanta 2 dias de PRODUCCION a su faltante; D1 mas."""
    from app.api.control_produccion.part_planning import (
        _ppy_anticipacion, _ppy_sumar_dias_produccion,
        PPY_ANTICIPACION_DIAS, PPY_ANTICIPACION_MAX,
    )

    assert PPY_ANTICIPACION_DIAS == 2
    assert _ppy_anticipacion("D1") == 5
    assert _ppy_anticipacion("d1") == 5           # case-insensitive
    for l in ("M1", "M2", "M3", "M4", "H1", "D2", "D3"):
        assert _ppy_anticipacion(l) == 2, l
    assert _ppy_anticipacion(None) == 2
    assert PPY_ANTICIPACION_MAX == 5

    # el caso que dio la regla: EBR33105505 falta el LUNES 20 y planning lo
    # puso el JUEVES 16, porque 2 dias de produccion desde el jueves son
    # viernes y lunes (el fin de semana no cuenta).
    jueves = date(2026, 7, 16)
    assert _ppy_sumar_dias_produccion(jueves, 1) == date(2026, 7, 17)  # vie
    assert _ppy_sumar_dias_produccion(jueves, 2) == date(2026, 7, 20)  # lun
    # contarlo en dias naturales daria el sabado, que ni siquiera se produce
    assert _ppy_sumar_dias_produccion(jueves, 2) != jueves + timedelta(days=2)
    # desde el lunes no hay finde de por medio
    assert _ppy_sumar_dias_produccion(date(2026, 7, 13), 2) == date(2026, 7, 15)
    # desde el viernes, el primer dia de produccion es el lunes
    assert _ppy_sumar_dias_produccion(date(2026, 7, 17), 1) == date(2026, 7, 20)
    assert _ppy_sumar_dias_produccion(jueves, 0) == jueves


def test_ppy_anticipacion_d1_ve_mas_lejos_que_main(monkeypatch):
    """Un faltante a 4 dias de produccion: main no lo adelanta (2), D1 si (5).

    Regla de planning: "lo unico que adelanta inclusive 7 dias antes es D1".
    """
    fecha = date(2026, 7, 15)          # miercoles
    shortage = date(2026, 7, 21)       # martes: 4 dias de produccion despues
    parte = "LEJOS00001"
    raw = {parte: {"model": "M", "sub_assy": "MAIN", "c_t": 36,
                   "uph": 100, "estandar_pack": 20}}
    proyeccion = {
        fecha + timedelta(days=o): (500 if fecha + timedelta(days=o) < shortage
                                    else -100)
        for o in range(0, 12)
    }
    hasta = fecha + timedelta(days=11)

    # main mira 2 dias de produccion: hasta el vie 17, no alcanza el martes
    pp = _mock_ppy_schedule_dependencies(
        monkeypatch, {parte: {"line": "M1", "proj": proyeccion}}, raw)
    props, _om, _ex = pp._ppy_simular_schedule(fecha, hasta, detailed=True)
    assert [x for x in props if x["sched_date"] == fecha.isoformat()] == []

    # D1 mira 5 dias de produccion: si lo ve
    pp = _mock_ppy_schedule_dependencies(
        monkeypatch, {parte: {"line": "D1", "proj": proyeccion}}, raw,
        lineas=("D1",))
    props, _om, _ex = pp._ppy_simular_schedule(fecha, hasta, detailed=True)
    hoy = [x for x in props if x["sched_date"] == fecha.isoformat()]
    assert len(hoy) == 1
    assert hoy[0]["fecha_shortage"] == shortage.isoformat()


def test_ppy_no_propone_produccion_en_fin_de_semana(monkeypatch):
    """ISEMM produce L-V: el generador no coloca lotes en sabado ni domingo.

    LG si consume el sabado, asi que ese faltante debe cubrirlo el viernes (su
    horizonte cuenta el consumo del sabado). El sabado solo se trabaja en caso
    extraordinario y eso lo decide el planificador a mano.
    """
    from app.api.control_produccion.part_planning import _ppy_es_dia_produccion

    assert _ppy_es_dia_produccion(date(2026, 7, 17)) is True   # viernes
    assert _ppy_es_dia_produccion(date(2026, 7, 18)) is False  # sabado
    assert _ppy_es_dia_produccion(date(2026, 7, 19)) is False  # domingo
    assert _ppy_es_dia_produccion(date(2026, 7, 20)) is True   # lunes

    # el simulador no debe devolver ni una propuesta en sab/dom
    viernes = date(2026, 7, 17)
    parte = "FINDE00001"
    proyeccion = {viernes + timedelta(days=o): -500 for o in range(0, 5)}
    pp = _mock_ppy_schedule_dependencies(
        monkeypatch,
        {parte: {"line": "M1", "proj": proyeccion}},
        {parte: {"model": "M", "sub_assy": "MAIN", "c_t": 36,
                 "uph": 100, "estandar_pack": 20}},
    )
    props, _om, _ex = pp._ppy_simular_schedule(
        viernes, viernes + timedelta(days=4), detailed=True)
    fechas = {p["sched_date"] for p in props}
    assert "2026-07-18" not in fechas, "no debe planear el sabado"
    assert "2026-07-19" not in fechas, "no debe planear el domingo"
    assert fechas <= {"2026-07-17", "2026-07-20", "2026-07-21"}


def test_ppy_bloques_capacidad_flexible():
    """La capacidad permite 5 turnos, pero abre solo los necesarios.

    Planning arma los bloques cada dia con las lineas que convenga (15/07:
    M1+D1 | M2 | M3+M4 | D3+D1+D2; 16/07: M2+M4 | M3 | D1 | D2+D3), y
    el 17/07 requirio M1 | M2 | M3+M4 | D1 | D2+D3 = 5. El repartidor debe
    empacar las lineas y no dispersarlas por los cinco bloques cuando caben en
    cuatro.
    """
    from app.api.control_produccion.part_planning import (
        _ppy_horas_iniciales, _ppy_armar_lotes, PPY_HORAS_TURNO, PPY_BLOQUES,
    )

    horas = _ppy_horas_iniciales()
    assert len(horas) == PPY_BLOQUES == 5
    assert sum(horas.values()) == 5 * PPY_HORAS_TURNO == 45.0

    # lo ya ocupado descuenta del pool, venga de la linea que venga
    horas = _ppy_horas_iniciales(["M3", "M4"], {"M3": 5.0, "M4": 4.0})
    assert sum(horas.values()) == 45.0 - 9.0

    hoy = date(2026, 7, 15)

    def cand(part, line):
        return {"part_no": part, "falt_total": 800, "falt_hoy": 800,
                "primera_falta": hoy, "line": line, "uph": 100, "pack": 20,
                "ct": 36, "model": None, "main_sub": None}

    # cada parte mantiene SU linea (assy_line manda) y consume un bloque
    horas = _ppy_horas_iniciales(["M3", "M4"])
    lotes, _fuera = _ppy_armar_lotes(
        [cand("AAA30299001", "M3"), cand("BBB80757001", "M4")], ["M3", "M4"], horas)
    assert {l["linea"] for l in lotes} == {"M3", "M4"}
    assert all(l["grupo"] in horas for l in lotes)
    assert sum(horas.values()) < 45.0

    from app.api.control_produccion.part_planning import _ppy_repartir_bloques

    # Horas del plan real del 16/07: se acomodan en cuatro bloques y B5 queda
    # cerrado, aunque el maximo disponible sea cinco.
    horas = _ppy_horas_iniciales()
    asignacion = _ppy_repartir_bloques(
        {"M2": 4.0, "M4": 4.5, "M3": 8.0, "D1": 8.5, "D2": 1.0, "D3": 7.0},
        horas,
    )
    assert len({bloque for bloque, _h in asignacion.values()}) == 4

    # Horas del 17/07: ninguna combinacion compatible cabe en cuatro; se abre
    # el quinto y se conserva M3+M4 / D2+D3.
    horas = _ppy_horas_iniciales()
    asignacion = _ppy_repartir_bloques(
        {"M1": 5.0, "M2": 4.5, "M3": 3.3, "M4": 5.0,
         "D1": 8.7, "D2": 1.0, "D3": 7.4},
        horas,
    )
    assert len({bloque for bloque, _h in asignacion.values()}) == 5
    assert asignacion["M3"][0] == asignacion["M4"][0]
    assert asignacion["D2"][0] == asignacion["D3"][0]


def test_ppy_remain_es_objetivo_no_disparador(monkeypatch):
    """El remain de 60 es el OBJETIVO del lote, no un umbral que dispare plan.

    Planning: "si no hay consumo no tiene que planearlo, solo cuando se va a
    producir algo tiene que dar remain de 60 [...] 50 min para scraps". Un
    inventario de 44 son los 60 del lote anterior menos scrap, no un faltante.
    """
    from app.api.control_produccion.part_planning import (
        PPY_REMAIN_MIN, PPY_REMAIN_IDEAL, _ppy_qty_pack,
    )

    assert PPY_REMAIN_IDEAL == 60 and PPY_REMAIN_MIN == 50
    # el lote se calcula para dejar 60: I=-100 -> faltante 160 -> +10% -> caja
    assert _ppy_qty_pack(PPY_REMAIN_IDEAL - (-100), 20) == 180

    fecha = date(2026, 7, 15)
    raw = {"BAJO000001": {"model": "M", "sub_assy": "MAIN", "c_t": 36,
                          "uph": 100, "estandar_pack": 20}}

    # remanente bajo (44) pero SIN consumo: no se planea
    pp = _mock_ppy_schedule_dependencies(
        monkeypatch,
        {"BAJO000001": {"line": "M1",
                        "proj": {fecha + timedelta(days=o): 44 for o in range(0, 5)}}},
        raw,
    )
    props, _om, _ex = pp._ppy_simular_schedule(
        fecha, fecha + timedelta(days=4), detailed=True)
    assert props == [], "sin consumo no se planea aunque el remanente este bajo"

    # el mismo remanente pero con consumo que lo hunde: si se planea, a 60
    pp = _mock_ppy_schedule_dependencies(
        monkeypatch,
        {"BAJO000001": {"line": "M1",
                        "proj": {fecha: 44, fecha + timedelta(days=1): -56}}},
        raw,
    )
    props, _om, _ex = pp._ppy_simular_schedule(
        fecha, fecha + timedelta(days=1), detailed=True)
    assert len(props) == 1
    # faltante = 60 - (-56) = 116 -> x1.10 = 127.6 -> caja de 20 -> 140
    assert props[0]["qty"] == 140
    assert props[0]["inventario_despues"] >= PPY_REMAIN_IDEAL


def test_ppy_d3_no_comparte_bloque_con_main():
    """D3 solo corre con D1/D2, nunca con las main. Regla de planning,
    verificada en los planes reales (15/07: M1+D1 y D3+D1+D2; 16/07: M2+M4 y
    D2+D3 — D3 nunca junto a una M)."""
    from app.api.control_produccion.part_planning import (
        _ppy_pueden_compartir, _ppy_repartir_bloques, PPY_HORAS_TURNO,
    )

    assert _ppy_pueden_compartir("D3", []) is True
    assert _ppy_pueden_compartir("D3", ["D1", "D2"]) is True
    assert _ppy_pueden_compartir("D3", ["M2"]) is False
    assert _ppy_pueden_compartir("M2", ["D3"]) is False      # simetrico
    assert _ppy_pueden_compartir("d3", ["m1"]) is False       # case-insensitive
    assert _ppy_pueden_compartir("D1", ["M1"]) is True        # D1/D2 si van con main
    assert _ppy_pueden_compartir("M4", ["M2"]) is True

    # al repartir, D3 nunca acaba en un bloque con main
    horas = {"B1": PPY_HORAS_TURNO, "B2": PPY_HORAS_TURNO}
    asign = _ppy_repartir_bloques({"M2": 4.0, "D3": 5.0, "D2": 2.0}, horas)
    bloque_de = {l: b for l, (b, _h) in asign.items()}
    assert bloque_de["D3"] != bloque_de["M2"]
    # D2 puede caer en cualquiera de los dos: va con quien tenga mas hueco
    assert bloque_de["D2"] in (bloque_de["D3"], bloque_de["M2"])

    # con un solo bloque disponible, D3 y M2 no pueden compartirlo: entra la
    # que mas horas pide (D3) y la otra queda fuera
    horas = {"B1": 0.0, "B2": PPY_HORAS_TURNO}
    asign = _ppy_repartir_bloques({"M2": 4.0, "D3": 5.0}, horas)
    assert asign["D3"][0] == "B2"
    assert "M2" not in asign


def _cal_bytes(hojas=("260629", "260715"), fechas=None, filas=None):
    """Cal diario: una hoja por fecha, fechas en la fila 2 y parte en la col A.

    Layout distinto al del xlsm semanal (hoja 'LG', fila 3, col B).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for h in hojas:
        ws = wb.create_sheet(h)
        ws.cell(row=1, column=1, value="W29")   # la fila 1 lleva la semana
        for i, v in enumerate(fechas or []):
            ws.cell(row=2, column=2 + i, value=v)
        for r, (parte, qtys) in enumerate(filas or [], start=3):
            ws.cell(row=r, column=1, value=parte)
            for i, q in enumerate(qtys):
                ws.cell(row=r, column=2 + i, value=q)
    wb.create_sheet("BOM")      # hojas sueltas que no son plan
    wb.create_sheet("ISSUES")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_sync_part_reutiliza_reemplazo_del_boton(monkeypatch):
    import app.api.control_produccion.part_planning as pp

    parsed = {
        "sheet_name": "Part 16",
        "ref_date": date(2026, 7, 13),
        "date_to": date(2026, 7, 20),
        "inventory": {
            "P1": {"line": "R1"},
            "P2": {"line": "D2"},
        },
        "schedules": {
            ("P1", date(2026, 7, 16)): 100,
            ("P1", date(2026, 7, 17)): 200,
        },
    }
    monkeypatch.setattr(
        pp, "_parse_part10_workbook", lambda *_args: (parsed, None)
    )
    monkeypatch.setattr(pp, "_ppy_config_lineas", lambda: ["M1", "D2"])
    monkeypatch.setattr(
        pp,
        "_ppy_datos_raw",
        lambda parts: {"P1": {"assy_line": "M1"}} if parts else {},
    )

    preview = pp._pp_sincronizar_schedule_excel(
        b"excel", "plan.xlsm", "ana", aplicar=False
    )

    assert preview == {
        "success": True,
        "sheet_name": "Part 16",
        "parts": 1,
        "schedules": 2,
        "date_from": "2026-07-13",
        "date_to": "2026-07-20",
        "scope": "todos",
        "source_parts": 1,
        "excluded_by_scope": 0,
        "skipped_without_active_line": 0,
        "skipped_parts_without_active_line": [],
        "warnings": [],
        "applied": False,
    }

    class Cursor:
        def __init__(self):
            self.executions = []
            self.many = []
            self.rowcount = 0

        def execute(self, sql, params):
            self.executions.append((" ".join(sql.split()), params))
            self.rowcount = 3

        def executemany(self, sql, rows):
            self.many.append((" ".join(sql.split()), list(rows)))

        def close(self):
            pass

    class Connection:
        def __init__(self):
            self.cursor_instance = Cursor()
            self.committed = False

        def autocommit(self, _value):
            pass

        def commit(self):
            self.committed = True

        def rollback(self):
            pass

        def close(self):
            pass

    connection = Connection()
    monkeypatch.setattr(pp, "get_pooled_connection", lambda: connection)
    monkeypatch.setattr(pp, "get_dict_cursor", lambda conn: conn.cursor_instance)

    result = pp._pp_sincronizar_schedule_excel(
        b"excel", "plan.xlsm", "ana", aplicar=True
    )

    assert result["applied"] is True
    assert result["replaced"] == 3
    assert result["schedules"] == 2
    assert connection.committed is True
    delete_sql, delete_params = connection.cursor_instance.executions[0]
    assert "DELETE FROM lg_schedule_daily" in delete_sql
    assert delete_params == ("P1", date(2026, 7, 13), date(2026, 7, 20))
    inserted = connection.cursor_instance.many[0][1]
    assert inserted == [
        ("P1", date(2026, 7, 16), 100, "M1", "DIA", "ana"),
        ("P1", date(2026, 7, 17), 200, "M1", "DIA", "ana"),
    ]


def test_sync_part_main_filtra_lineas_fuera_de_m1_m4(monkeypatch):
    import app.api.control_produccion.part_planning as pp

    parsed = {
        "sheet_name": "Part 16",
        "ref_date": date(2026, 7, 13),
        "date_to": date(2026, 7, 20),
        "inventory": {"P-MAIN": {"line": "R1"}, "P-DISPLAY": {"line": "D2"}},
        "schedules": {
            ("P-MAIN", date(2026, 7, 16)): 100,
            ("P-DISPLAY", date(2026, 7, 16)): 50,
        },
    }
    monkeypatch.setattr(pp, "_parse_part10_workbook", lambda *_args: (parsed, None))
    monkeypatch.setattr(pp, "_ppy_config_lineas", lambda: ["M1", "D2"])
    monkeypatch.setattr(
        pp,
        "_ppy_datos_raw",
        lambda _parts: {"P-MAIN": {"assy_line": "M1"}, "P-DISPLAY": {"assy_line": "D2"}},
    )

    preview = pp._pp_sincronizar_schedule_excel(
        b"excel", "plan.xlsm", "ana", aplicar=False, alcance="main"
    )

    assert preview["scope"] == "main"
    assert preview["parts"] == 1
    assert preview["schedules"] == 1
    assert preview["excluded_by_scope"] == 1


def test_sync_part_omite_sin_linea_y_sincroniza_las_demas(monkeypatch):
    import app.api.control_produccion.part_planning as pp

    parsed = {
        "sheet_name": "Part 16",
        "ref_date": date(2026, 7, 13),
        "date_to": date(2026, 7, 20),
        "inventory": {"P-VALIDA": {"line": "M1"}, "P-SIN-LINEA": {"line": None}},
        "schedules": {
            ("P-VALIDA", date(2026, 7, 16)): 100,
            ("P-SIN-LINEA", date(2026, 7, 16)): 50,
        },
    }
    monkeypatch.setattr(pp, "_parse_part10_workbook", lambda *_args: (parsed, None))
    monkeypatch.setattr(pp, "_ppy_config_lineas", lambda: ["M1", "M2"])
    monkeypatch.setattr(pp, "_ppy_datos_raw", lambda _parts: {})

    preview = pp._pp_sincronizar_schedule_excel(
        b"excel", "plan.xlsm", "ana", aplicar=False, alcance="todos"
    )

    assert preview["source_parts"] == 2
    assert preview["parts"] == 1
    assert preview["schedules"] == 1
    assert preview["skipped_without_active_line"] == 1
    assert preview["skipped_parts_without_active_line"] == ["P-SIN-LINEA"]
    assert preview["warnings"][0]["code"] == "SCHEDULE_PARTS_SKIPPED_NO_ACTIVE_LINE"


def test_parse_cal_diario_layout_distinto():
    """El Cal se lee igual que el xlsm aunque cambie hoja, fila y columna."""
    data = _cal_bytes(
        fechas=[datetime(2026, 7, 15), datetime(2026, 7, 16)],
        filas=[("MDS64531211", [404, 650]), ("ABQ74229132", [0, 180])],
    )
    parsed, err = _parse_lg_workbook(data, "Cal_260715.xlsx")
    assert err is None
    assert parsed["parts_count"] == 2
    assert parsed["records"][("MDS64531211", date(2026, 7, 16))] == 650
    assert parsed["records"][("ABQ74229132", date(2026, 7, 16))] == 180


def test_pp_hoja_plan_elige_la_del_nombre():
    """Con varias hojas fechadas se toma la del archivo; si no, la mas reciente."""
    from app.api.control_produccion.part_planning import _pp_hoja_plan

    class _WB:
        def __init__(self, nombres):
            self.sheetnames = nombres

    # el xlsm semanal manda: si hay hoja LG, esa
    assert _pp_hoja_plan(_WB(["LG", "260715", "Part 15"]), "plan.xlsm") == "LG"
    # el Cal: la hoja que coincide con el nombre del archivo
    hojas = ["260629", "260706", "260706 (2)", "260715", "BOM", "ISSUES"]
    assert _pp_hoja_plan(_WB(hojas), "Cal_260715.xlsx") == "260715"
    assert _pp_hoja_plan(_WB(hojas), "Cal_260706.xlsx") == "260706"
    # sin fecha en el nombre: la mas reciente (no las variantes "(2)")
    assert _pp_hoja_plan(_WB(hojas), "Cal.xlsx") == "260715"
    # sin hoja de plan
    assert _pp_hoja_plan(_WB(["BOM", "ISSUES"]), "otro.xlsx") is None


def test_pp_fecha_archivo_dos_convenciones():
    """El xlsm semanal escribe la fecha completa; el Cal diario, en corto."""
    from app.api.control_produccion.part_planning import _pp_fecha_archivo

    assert _pp_fecha_archivo(
        "★plan de produccion ILSAN 20260715 W29 ver03.xlsm") == date(2026, 7, 15)
    assert _pp_fecha_archivo("Cal_260715.xlsx") == date(2026, 7, 15)
    assert _pp_fecha_archivo("Cal_260629.xlsx") == date(2026, 6, 29)
    # sin fecha usable
    assert _pp_fecha_archivo("plan.xlsm") is None
    assert _pp_fecha_archivo("") is None
    assert _pp_fecha_archivo("plan ILSAN 20261332 W29.xlsm") is None  # mes 13
    # la ruta no debe confundirse con la fecha
    assert _pp_fecha_archivo("C:/260101/Cal_260715.xlsx") == date(2026, 7, 15)


def test_pp_import_mas_reciente_avisa_retroceso(monkeypatch):
    """Importar un archivo viejo sobre uno nuevo revierte: hay que avisar.

    El upsert es por parte+fecha y gana el ultimo, asi que re-importar el xlsm
    del lunes despues del Cal del jueves pisa lo que el Cal actualizo.
    """
    import app.api.control_produccion.part_planning as pp

    previos = [{"original_filename": "Cal_260715.xlsx", "imported_at": None}]
    monkeypatch.setattr(pp, "execute_query", lambda *a, **k: previos)

    # el xlsm del 14 es mas viejo que el Cal del 15 ya importado -> avisa
    aviso = pp._pp_import_mas_reciente(
        date(2026, 7, 14), date(2026, 6, 29), date(2026, 9, 16))
    assert aviso and aviso["fecha_archivo"] == "2026-07-15"

    # el Cal del 16 es mas nuevo -> no avisa
    assert pp._pp_import_mas_reciente(
        date(2026, 7, 16), date(2026, 7, 16), date(2026, 9, 14)) is None
    # mismo dia -> no avisa (re-importar el mismo archivo es normal)
    assert pp._pp_import_mas_reciente(
        date(2026, 7, 15), date(2026, 7, 15), date(2026, 9, 14)) is None
    # sin fecha en el nombre no se puede comparar -> no avisa
    assert pp._pp_import_mas_reciente(
        None, date(2026, 7, 15), date(2026, 9, 14)) is None


def test_ppy_replanear_incluye_lo_ya_puesto(monkeypatch):
    """El plan del dia es COMPLETO: incluye lo que ya se habia puesto antes.

    Planning: "tiene que dar incluso lo que ya se habia puesto antes, se planea
    diario porque cambia plan de LG". El schedule de mañana capturado ayer esta
    viejo, asi que no cuenta como cubierto: se rehace.
    """
    import app.api.control_produccion.part_planning as pp

    fecha = date(2026, 7, 15)
    parte = "REPLAN0001"
    raw = {parte: {"model": "M", "sub_assy": "MAIN", "c_t": 36,
                   "uph": 100, "estandar_pack": 20}}
    # I sano hoy y faltante manana; el schedule de manana ya trae 500 pzs
    proy_con_sched = {parte: {"line": "M1",
                              "proj": {fecha: 500, fecha + timedelta(days=1): 440}}}
    proy_replan = {parte: {"line": "M1",
                           "proj": {fecha: 500, fecha + timedelta(days=1): -60}}}

    # replanear=False: el schedule cubre, no hay nada que agregar
    pp2 = _mock_ppy_schedule_dependencies(monkeypatch, proy_con_sched, raw)
    props, _om, _ex = pp2._ppy_simular_schedule(
        fecha, fecha + timedelta(days=1), detailed=True, replanear=False)
    assert props == []

    # replanear=True: se ignora ese schedule y el plan sale completo
    pp2 = _mock_ppy_schedule_dependencies(monkeypatch, proy_replan, raw)
    props, _om, _ex = pp2._ppy_simular_schedule(
        fecha, fecha + timedelta(days=1), detailed=True, replanear=True)
    assert len(props) == 1
    assert props[0]["qty"] == 140   # 60-(-60)=120 -> +10% -> caja 20 -> 140


def test_ppy_proyeccion_replanear_desde_ignora_sched_futuro(monkeypatch):
    """replanear_desde ignora el schedule de esa fecha en adelante, no el previo."""
    import app.api.control_produccion.part_planning as pp

    inv = [{"part_no": "P1", "board": None, "line": "M1", "ref_date": date(2026, 7, 13),
            "lgemm": 100, "isemm": 0, "svc": 0, "dif": 0, "pendiente": 0,
            "rework": 0, "smt": 0, "imd": 0}]

    def fake_query(sql, params=(), fetch=None):
        if "lg_part_inventory" in sql:
            return inv
        if "FROM raw" in sql:
            return []
        if "lg_plan_daily" in sql:
            return [{"part_no": "P1", "f": date(2026, 7, 15), "q": 100}]
        if "lg_schedule_daily" in sql:
            return [{"part_no": "P1", "f": date(2026, 7, 14), "q": 30},   # ya produjo
                    {"part_no": "P1", "f": date(2026, 7, 16), "q": 500}]  # se rehace
        return []

    # sin replanear: el S del 16 cuenta -> I sube
    proy = pp._ppy_proyeccion_rango(date(2026, 7, 13), date(2026, 7, 16), query=fake_query)
    assert proy["P1"]["proj"][date(2026, 7, 16)] == 100 + 30 - 100 + 500

    # replaneando desde el 16: ese S se ignora, el del 14 no
    proy = pp._ppy_proyeccion_rango(date(2026, 7, 13), date(2026, 7, 16), query=fake_query,
                                    replanear_desde=date(2026, 7, 16))
    assert proy["P1"]["proj"][date(2026, 7, 16)] == 100 + 30 - 100


def test_ppy_partes_excluidas_no_entran_al_plan(monkeypatch):
    """Lo que no producimos aqui no se planea aunque LG lo pida.

    ACQ30500849 la vende ILSAN Corea aparte: su demanda LG no es trabajo
    nuestro. Nada en los datos la distingue (misma linea/CT/UPH que sus
    vecinas), por eso va en lg_pp_config.partes_excluidas.
    """
    import app.api.control_produccion.part_planning as pp

    # el default trae la parte que reporto planning
    assert "ACQ30500849" in pp.PPY_EXCLUIDAS_DEFAULT

    monkeypatch.setattr(pp, "execute_query",
                        lambda *a, **k: {"valor": "ACQ30500849, acq91482499"})
    ex = pp._ppy_partes_excluidas()
    assert ex == {"ACQ30500849", "ACQ91482499"}   # normaliza a mayusculas

    # un error de SQL no vacia la lista: cae al default, no planea de mas
    def boom(*a, **k):
        raise RuntimeError("sin conexion")
    monkeypatch.setattr(pp, "execute_query", boom)
    assert "ACQ30500849" in pp._ppy_partes_excluidas()

    # config vacia = nada excluido (se puede desactivar)
    monkeypatch.setattr(pp, "execute_query", lambda *a, **k: {"valor": ""})
    assert pp._ppy_partes_excluidas() == set()

    # y el simulador la ignora
    fecha = date(2026, 7, 17)
    proy = {"ACQ30500849": {"line": "D2", "proj": {fecha: -1000}},
            "EBR76683912": {"line": "D1", "proj": {fecha: -1000}}}
    raw = {p: {"model": "M", "sub_assy": "S", "c_t": 14, "uph": 257,
               "estandar_pack": 20} for p in proy}
    pp2 = _mock_ppy_schedule_dependencies(monkeypatch, proy, raw, lineas=("D1", "D2"))
    monkeypatch.setattr(pp2, "_ppy_partes_excluidas", lambda *a, **k: {"ACQ30500849"})
    props, _om, _ex = pp2._ppy_simular_schedule(fecha, fecha, detailed=True)
    assert [p["numero_parte"] for p in props] == ["EBR76683912"]
