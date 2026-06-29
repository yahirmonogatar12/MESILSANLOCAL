from datetime import datetime
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

from openpyxl import Workbook

from app.api.control_material.invoice_core import service as invoice_service
from app.api.control_material.invoice_core.normalizers import normalizar_pallet_no
from app.api.control_material.invoice_core.matcher import validate_system_parts
from app.api.control_material.invoice_core.parser import parse_invoice_workbook


def _workbook_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "INVOICE(total)"
    ws.append(["Texto previo", "", "", ""])
    ws.append(["MAKER", "ORIGIN", "PART NUM.", "DESCRIPTION", "QTY", "UOM", "UNIT COST", "TT COST"])
    ws.append(["ILS", "KR", "abc-001", "Capacitor", "10", "EA", "1.25", "12.50"])
    packing = wb.create_sheet("PACKING LIST 원본")
    packing.append(["PACKING LIST"])
    packing.append([])
    packing.append(["", "", "NO", "품번", "품명", "수량", "", "", "", "", "", "", ""])
    packing.append(["", "", "01", "abc-001", "Capacitor", "10", "EA", "1", "PALLET", "2.5", "KG", "0.1", "CBM"])
    packing.append(["", "", "", "abc-001", "Capacitor", "5", "EA", "", "", "", "", "", ""])
    packing.append(["", "", "02", "COMMERCIAL", "", "10", "EA", "1", "PALLET", "", "", "", ""])
    packing.append(["", "", "", "zero-001", "Free sample", "0", "EA", "", "", "", "", "", ""])
    aux = wb.create_sheet("INVOICE(copy)")
    aux.append(["MAKER", "ORIGIN", "PART NUM.", "DESCRIPTION", "QTY", "UOM", "UNIT COST", "TT COST"])
    aux.append(["ILS", "KR", "ignored-001", "Aux", "99", "EA", "9.99", "989.01"])
    equivalences = wb.create_sheet("Hoja1")
    equivalences.append(["TARIMA", "PART NO", "Part Sys", "ITEM", "SPEC", "QTY"])
    equivalences.append([1, "abc-001", "SYS-001", "CAPACITOR", "x", 10])
    equivalences.append([None, "zero-001", "ZERO-SYS", "CAPACITOR", "x", 0])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def test_normalizar_pallet_no_variantes():
    assert normalizar_pallet_no(1) == "1"
    assert normalizar_pallet_no("01") == "1"
    assert normalizar_pallet_no("PALLET 1") == "1"
    assert normalizar_pallet_no("Pallet-01") == "1"


def test_parse_invoice_workbook_detecta_invoice_y_packing_con_raw_values():
    parsed = parse_invoice_workbook(_workbook_bytes(), "INV-001.xlsx")

    assert len(parsed["invoice_lines"]) == 1
    assert len(parsed["packing_lines"]) == 2

    line = parsed["invoice_lines"][0]
    assert line["raw_part_num"] == "abc-001"
    assert line["numero_parte_invoice"] == "ABC-001"
    assert line["numero_parte_sistema"] == "ABC-001"
    assert line["raw_qty"] == "10"
    assert line["raw_unit_cost"] == "1.25"
    assert str(line["costo_total"]) == "12.5000"

    packing = parsed["packing_lines"][0]
    assert packing["pallet_no_original"] == "01"
    assert packing["pallet_no"] == "1"
    assert packing["numero_parte_packing"] == "ABC-001"
    assert packing["numero_parte_sistema"] == "ABC-001"
    assert str(packing["kg"]) == "2.5000"
    assert str(packing["cbm"]) == "0.1000"
    assert parsed["packing_lines"][1]["pallet_no_original"] == "01"
    assert parsed["packing_lines"][1]["pallet_no"] == "1"


def test_parse_invoice_workbook_requiere_hoja_invoice_total():
    wb = Workbook()
    ws = wb.active
    ws.title = "INVOICE(copy)"
    ws.append(["MAKER", "ORIGIN", "PART NUM.", "DESCRIPTION", "QTY", "UOM", "UNIT COST", "TT COST"])
    ws.append(["ILS", "KR", "abc-001", "Capacitor", "10", "EA", "1.25", "12.50"])
    out = BytesIO()
    wb.save(out)

    parsed = parse_invoice_workbook(out.getvalue(), "INV-001.xlsx")

    assert parsed["invoice_lines"] == []
    assert "INVOICE(total)" in parsed["warnings"][0]


def _converted_workbook_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "INVOICE(CONVERTED)"
    ws.append(["IM20260515-V32"])
    ws.append(["TARIMA", "PART NO", "Part Sys", "ITEM", "SPEC", "QTY", "UOM", "COSTOS", "TOTAL"])
    ws.append(["1", "EAX66946005-1.0", "EAX66946005", "PCB MAIN", "FR4 230*139", 2000, "EA", 2.2682, 4536.4])
    # Tarima vacia: hereda la anterior (1).
    ws.append([None, "0CE108XH618", "0CE108XH618", "CAPACITOR", "1000uF 25V", 7500, "EA", 0.0574, None])
    ws.append(["2", "MGW65415201", "MGW65415201", "REFLECTOR", "MOLD", 2000, "EA", 0.5, 1000])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def test_parse_invoice_workbook_lee_hoja_converted():
    parsed = parse_invoice_workbook(_converted_workbook_bytes(), "V-32.xlsx")

    assert parsed["fuente_hoja"] == "INVOICE(CONVERTED)"
    # Numero de invoice = primera celda (A1).
    assert parsed["numero_invoice_sugerido"] == "IM20260515-V32"
    assert parsed["warnings"] == []
    assert len(parsed["invoice_lines"]) == 3
    assert len(parsed["packing_lines"]) == 3

    primera = parsed["invoice_lines"][0]
    assert primera["raw_part_num"] == "EAX66946005-1.0"
    assert primera["numero_parte_sistema"] == "EAX66946005"
    assert str(primera["cantidad"]) == "2000.0000"
    assert str(primera["costo_unitario"]) == "2.2682"
    assert str(primera["costo_total"]) == "4536.4000"

    # Total se calcula cuando viene vacio (unit * qty).
    segunda = parsed["invoice_lines"][1]
    assert str(segunda["costo_total"]) == "430.5000"

    # TARIMA actua como pallet y se hereda cuando viene vacia.
    assert parsed["packing_lines"][0]["pallet_no"] == "1"
    assert parsed["packing_lines"][1]["pallet_no"] == "1"
    assert parsed["packing_lines"][2]["pallet_no"] == "2"


def test_material_invoice_blueprints_registran_rutas(app):
    rules = {str(rule) for rule in app.url_map.iter_rules()}

    assert "/material/invoices" in rules
    assert "/api/material_admin/invoices/upload" in rules
    assert "/api/material_admin/invoices/preview" in rules
    assert "/api/material_admin/invoices/<int:invoice_id>/apply" in rules
    assert "/api/material_admin/invoices/<int:invoice_id>/file" in rules
    assert "/api/material_admin/invoices/<int:invoice_id>/lines/<int:line_id>" in rules
    # Las rutas de equivalentes/aliases ya no se exponen.
    assert "/api/material_admin/invoices/aliases" not in rules

    # El borrado de invoice se expone como DELETE en la URL del detalle.
    metodos_detalle = set()
    for rule in app.url_map.iter_rules():
        if str(rule) == "/api/material_admin/invoices/<int:invoice_id>":
            metodos_detalle |= set(rule.methods or [])
    assert "DELETE" in metodos_detalle
    metodos_linea = set()
    for rule in app.url_map.iter_rules():
        if str(rule) == "/api/material_admin/invoices/<int:invoice_id>/lines/<int:line_id>":
            metodos_linea |= set(rule.methods or [])
    assert "PATCH" in metodos_linea
    assert "/api/material_admin/inventory/valuation" in rules
    assert "/api/material_admin/inventory/valuation/backfill" in rules
    # Las rutas de numeros de parte originales (aliases) se retiraron.
    assert "/api/informacion_basica/control_material/numeros-originales/import" not in rules
    assert "/api/informacion_basica/control_material/numeros-originales/preview" not in rules


class _MaterialesCursor:
    """Simula materiales: EAX66946005 existe con unidad_medida 'MTR'."""

    def execute(self, query, params):
        pass

    def fetchall(self):
        return [{"numero_parte": "EAX66946005", "unidad_medida": "MTR"}]


def test_validate_system_parts_marca_existencia_y_toma_uom():
    # invoice_lines usan estado 'DIRECTO' (su ENUM no acepta 'MATCH').
    records = [
        {"numero_parte_sistema": "EAX66946005", "uom": "XX"},
        {"numero_parte_sistema": "NO-EXISTE-123", "uom": "XX"},
        {"numero_parte_sistema": "", "uom": ""},
    ]

    validate_system_parts(_MaterialesCursor(), records, "numero_parte_sistema", "DIRECTO")

    # Parte existente: estado DIRECTO y UOM tomado de materiales (no del Excel).
    assert records[0]["estado_match"] == "DIRECTO"
    assert records[0]["uom"] == "MTR"
    # Parte inexistente: diferencia, conserva el UOM del Excel.
    assert records[1]["estado_match"] == "SIN_ALIAS"
    assert records[2]["estado_match"] == "SIN_ALIAS"


def test_validate_system_parts_packing_usa_match():
    # packing_lines usan estado 'MATCH'.
    records = [{"numero_parte_sistema": "EAX66946005"}]

    validate_system_parts(_MaterialesCursor(), records, "numero_parte_sistema", "MATCH")

    assert records[0]["estado_match"] == "MATCH"


def test_validate_system_parts_encuentra_por_parte_base_sin_version():
    # El codigo trae version y/o lote (EAX66946005-1.0, EAX66946005-1.0-2026...)
    # pero materiales tiene la base. Debe resolver a la base y tomar el UOM,
    # probando los prefijos por guion del mas largo al mas corto.
    records = [
        {"numero_parte_sistema": "EAX66946005-1.0", "uom": "EA"},
        {"numero_parte_sistema": "EAX66946005-1.4", "uom": "EA"},
        {"numero_parte_sistema": "EAX66946005-1.0-202601090001", "uom": "EA"},
    ]

    validate_system_parts(_MaterialesCursor(), records, "numero_parte_sistema", "DIRECTO")

    for record in records:
        assert record["numero_parte_sistema"] == "EAX66946005"
        assert record["estado_match"] == "DIRECTO"
        assert record["uom"] == "MTR"
    assert "base" in records[0]["mensaje_match"].lower()


class _UploadConnection:
    def __init__(self):
        self.commits = 0
        self.rollbacks = 0
        self.closed = False

    def commit(self):
        self.commits += 1

    def rollback(self):
        self.rollbacks += 1

    def close(self):
        self.closed = True


class _UploadCursor:
    def __init__(self, duplicate_on_insert=False):
        self.duplicate_on_insert = duplicate_on_insert
        self.events = []
        self.lastrowid = 73
        self._row = None
        self._select_count = 0
        self.closed = False

    def execute(self, query, params=None):
        normalized = " ".join(query.split())
        if normalized.startswith("SELECT id FROM material_invoices"):
            self._select_count += 1
            # Las dos consultas preventivas no encuentran nada. Tras el 1062,
            # la consulta de recuperacion ya ve la invoice de la otra solicitud.
            self._row = {"id": 91} if self._select_count >= 3 else None
            return
        if normalized == "START TRANSACTION":
            self.events.append("transaction")
            return
        if normalized.startswith("INSERT INTO material_invoices"):
            self.events.append("insert")
            if self.duplicate_on_insert:
                raise Exception(
                    1062,
                    "Duplicate entry 'IM20260601-V38' for key 'material_invoices.uk_invoice_numero'",
                )
            return
        if normalized.startswith("UPDATE material_invoices"):
            self.events.append("update")

    def fetchone(self):
        return self._row

    def close(self):
        self.closed = True


def _patch_upload_dependencies(monkeypatch, cursor, conn, save_file):
    parsed = {
        "invoice_lines": [{"costo_total": Decimal("10.0000")}],
        "packing_lines": [],
        "warnings": [],
        "numero_invoice_sugerido": "IM20260601-V38",
    }
    monkeypatch.setattr(invoice_service, "parse_invoice_workbook", lambda *_: parsed)
    monkeypatch.setattr(invoice_service, "_db", lambda: (conn, cursor, None))
    monkeypatch.setattr(invoice_service, "_usuario_actual", lambda: "TEST")
    monkeypatch.setattr(invoice_service, "obtener_fecha_hora_mexico", lambda: datetime(2026, 6, 29, 13, 46))
    monkeypatch.setattr(invoice_service, "build_relative_path", lambda *_: "2026/06/invoice.xlsx")
    monkeypatch.setattr(invoice_service, "save_file", save_file)
    monkeypatch.setattr(invoice_service, "delete_file", lambda path: cursor.events.append("delete"))
    monkeypatch.setattr(invoice_service, "validate_system_parts", lambda *_: None)
    monkeypatch.setattr(invoice_service, "_insert_invoice_lines", lambda *_: cursor.events.append("lines"))
    monkeypatch.setattr(invoice_service, "_insert_packing_lines", lambda *_: cursor.events.append("packing"))
    monkeypatch.setattr(invoice_service, "assign_packing_lines", lambda *_: cursor.events.append("assign"))
    monkeypatch.setattr(invoice_service, "invoice_has_differences", lambda *_: False)


def _upload_file():
    return {"file": SimpleNamespace(filename="invoice.xlsx", read=lambda: b"excel-bytes")}


def test_upload_invoice_colision_concurrente_responde_409_sin_borrar_excel(monkeypatch):
    cursor = _UploadCursor(duplicate_on_insert=True)
    conn = _UploadConnection()

    def unexpected_save(*_):
        raise AssertionError("El archivo no debe escribirse antes de ganar el INSERT")

    _patch_upload_dependencies(monkeypatch, cursor, conn, unexpected_save)

    payload, status = invoice_service.upload_invoice(_upload_file(), {})

    assert status == 409
    assert payload == {
        "success": False,
        "duplicado": True,
        "motivo": "NUMERO_INVOICE",
        "message": "Esta invoice ya fue cargada.",
        "invoice_id": 91,
    }
    assert conn.rollbacks == 1
    assert "delete" not in cursor.events


def test_upload_invoice_guarda_excel_despues_del_insert(monkeypatch):
    cursor = _UploadCursor()
    conn = _UploadConnection()

    def record_save(*_):
        cursor.events.append("save")
        return "C:/storage/invoice.xlsx", len(b"excel-bytes")

    _patch_upload_dependencies(monkeypatch, cursor, conn, record_save)

    payload, status = invoice_service.upload_invoice(_upload_file(), {})

    assert status == 201
    assert payload["invoice_id"] == 73
    assert cursor.events.index("insert") < cursor.events.index("save")
    assert conn.commits == 1
    assert conn.rollbacks == 0


def test_list_invoices_abiertos_ignora_fechas_y_todos_si_las_aplica(monkeypatch):
    class ListCursor:
        def __init__(self):
            self.query = ""
            self.params = []

        def execute(self, query, params=None):
            self.query = " ".join(query.split())
            self.params = list(params or [])

        def fetchall(self):
            return []

        def close(self):
            pass

    cursor = ListCursor()
    conn = _UploadConnection()
    monkeypatch.setattr(invoice_service, "_db", lambda: (conn, cursor, None))

    payload, status = invoice_service.list_invoices(
        {
            "estado": "PENDIENTES",
            "fecha_inicio": "2026-06-15",
            "fecha_fin": "2026-06-29",
        }
    )

    assert status == 200
    assert payload["records"] == []
    assert "COALESCE(mi.cerrado_manual, 0) = 0" in cursor.query
    assert "mi.estado NOT IN ('APLICADA', 'CANCELADA')" in cursor.query
    assert "mi.fecha_carga >= %s" not in cursor.query
    assert "mi.fecha_carga <= %s" not in cursor.query
    assert cursor.params == []

    cursor = ListCursor()
    monkeypatch.setattr(invoice_service, "_db", lambda: (conn, cursor, None))

    payload, status = invoice_service.list_invoices(
        {"fecha_inicio": "2026-06-15", "fecha_fin": "2026-06-29"}
    )

    assert status == 200
    assert "COALESCE(mi.cerrado_manual, 0) = 0" not in cursor.query
    assert "mi.fecha_carga >= %s" in cursor.query
    assert "mi.fecha_carga <= %s" in cursor.query
    assert cursor.params == ["2026-06-15 00:00:00", "2026-06-29 23:59:59"]
