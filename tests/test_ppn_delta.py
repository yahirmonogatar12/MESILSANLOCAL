"""Consumo real de LG a partir de dos snapshots del PPN (app/.../ppn.py)."""
import datetime as dt
import io

import openpyxl
import pytest

from app.api.control_produccion import part_planning as pp
from app.api.control_produccion import ppn
from app.api.portal import ai_plan_tools

DIA = dt.date(2026, 7, 29)


def _wos_ayer():
    return {
        "A": {"modelo": "M1", "partes": {"P1", "P2"}, "sin_bom": False,
              "plan": 100, "result": 0, "dias": {DIA: 100}},
        "B": {"modelo": "M2", "partes": {"P1"}, "sin_bom": False,
              "plan": 50, "result": 10, "dias": {DIA: 40}},
        "C": {"modelo": "M3", "partes": {"P3"}, "sin_bom": False,
              "plan": 30, "result": 0, "dias": {DIA + dt.timedelta(days=1): 30}},
    }


def test_sobrante_atraso_y_cierre():
    base = ppn.baseline(DIA, _wos_ayer())
    # A construyo 60 de 100 -> 40 sin consumir. B desaparecio = cerrado (40, sin
    # atraso). C no estaba programado ese dia.
    hoy = {"A": {"result": 60}}
    renglones, partes = ppn.sobrante(base, hoy)
    assert {r[0] for r in renglones} == {"A", "B"}
    assert partes == {"P1": 40, "P2": 40}


def test_sobrante_topa_en_lo_programado():
    """Un W/O no puede consumir mas de lo que tenia programado ese dia."""
    wos = _wos_ayer()
    wos["A"]["dias"][DIA] = 20  # programado 20, construyo 60 (de varios dias)
    base = ppn.baseline(DIA, wos)
    _, partes = ppn.sobrante(base, {"A": {"result": 60}})
    assert partes == {}


def test_sin_atraso_no_hay_sobrante():
    base = ppn.baseline(DIA, _wos_ayer())
    hoy = {"A": {"result": 100}, "B": {"result": 50}}
    assert ppn.sobrante(base, hoy)[1] == {}


def test_demanda_explota_por_parte():
    assert ppn.demanda({"A": _wos_ayer()["A"]}) == {("P1", DIA): 100, ("P2", DIA): 100}


def _ppn_workbook(fecha0, renglones):
    """PPN minimo en memoria. renglones: (wo, modelo, partes, plan, result, dias)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PPN 30 Jul"
    fechas = [fecha0 + dt.timedelta(days=i) for i in range(ppn.C_DIAN - ppn.C_DIA0)]
    for i, f in enumerate(fechas):
        ws.cell(1, ppn.C_DIA0 + 1 + i, f.day)
    ws.cell(2, 1, "Sum")
    for fila, (wo, modelo, partes, plan, result, dias) in enumerate(renglones, start=3):
        ws.cell(fila, 1, fila - 2)
        ws.cell(fila, ppn.C_WO + 1, wo)
        ws.cell(fila, ppn.C_MODELO + 1, modelo)
        for j, parte in enumerate(partes):
            ws.cell(fila, 9 + j, parte)
        ws.cell(fila, ppn.C_PSTART + 1, dt.datetime.combine(fecha0, dt.time(7, 30)))
        ws.cell(fila, ppn.C_PLAN + 1, plan)
        ws.cell(fila, ppn.C_RESULT + 1, result)
        for f, qty in dias.items():
            ws.cell(fila, ppn.C_DIA0 + 1 + fechas.index(f), qty)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_ppn_rebaja_el_plan_del_dia_que_paso(monkeypatch):
    """El plan de ayer se corrige con lo que LG realmente construyo."""
    hoy = DIA + dt.timedelta(days=1)
    # Ayer: el W/O A tenia 100 programadas y solo construyo 60.
    monkeypatch.setattr(pp, "_pp_ppn_baseline", lambda *_a: (DIA, {
        "A": {"partes": {"P1", "P2"}, "prog": 100, "plan": 100, "result": 0},
    }))
    # P1 ya tenia plan de ayer (120: el Cal traia demanda extra); P2 no tenia.
    monkeypatch.setattr(pp, "execute_query", lambda *_a, **_k: [
        {"part_no": "P1", "plan_qty": 120},
    ])
    data = _ppn_workbook(hoy, [("A", "M1", ["P1", "P2"], 100, 60, {hoy: 40})])

    parsed, err = pp._parse_lg_workbook(data, "PPN 30.07.26.xlsx")

    assert err is None
    # 40 sin consumir: el plan de ayer baja de 120 a 80 y no inventa faltante.
    assert parsed["records"][("P1", DIA)] == 80
    # Sin plan previo no se corrige nada (no se inventan renglones en el pasado).
    assert ("P2", DIA) not in parsed["records"]
    # La demanda de hoy sale de los buckets, ya neta de lo construido.
    assert parsed["records"][("P1", hoy)] == 40
    assert len(parsed["ppn_wo"]) == 1 and parsed["ppn_date"] == hoy


def test_parse_ppn_no_corrige_al_reimportar_el_mismo_dia(monkeypatch):
    """Reimportar el mismo PPN no vuelve a restar el sobrante."""
    monkeypatch.setattr(pp, "_pp_ppn_baseline", lambda *_a: (DIA, {
        "A": {"partes": {"P1"}, "prog": 100, "plan": 100, "result": 0},
    }))
    monkeypatch.setattr(pp, "execute_query", lambda *_a, **_k: [
        {"part_no": "P1", "plan_qty": 120},
    ])
    data = _ppn_workbook(DIA, [("A", "M1", ["P1"], 100, 60, {DIA: 40})])

    parsed, err = pp._parse_lg_workbook(data, "PPN 29.07.26.xlsx")

    assert err is None
    assert parsed["records"][("P1", DIA)] == 40  # el bucket, no el plan corregido


def test_parse_ppn_no_rebaja_dos_veces_con_baseline_explicito(monkeypatch):
    """Reaplicar la misma comparacion no vuelve a descontar el sobrante."""
    hoy = DIA + dt.timedelta(days=1)
    base = {"A": {"partes": {"P1"}, "prog": 100, "plan": 100, "result": 0}}
    # El snapshot guardado ya es del PPN de hoy: la correccion ya se aplico.
    monkeypatch.setattr(pp, "_pp_ppn_baseline", lambda *_a: (hoy, base))
    monkeypatch.setattr(pp, "execute_query", lambda *_a, **_k: [
        {"part_no": "P1", "plan_qty": 80},  # ya rebajado de 120 a 80
    ])
    data = _ppn_workbook(hoy, [("A", "M1", ["P1"], 100, 60, {hoy: 40})])

    parsed, err = pp._parse_lg_workbook(data, "PPN 30.xlsx", ppn_base=(DIA, base))

    assert err is None
    assert ("P1", DIA) not in parsed["records"]
    assert any("ya refleja el PPN" in w for w in parsed["warnings"])


def _oven_workbook(fecha0, renglones, dias=4, hoja="FP_Demand_Management"):
    """Prod Plan de OVEN minimo. renglones: (demand_id, modelo, partes, dias{})."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hoja
    fechas = [fecha0 + dt.timedelta(days=i) for i in range(dias)]
    cols = ["Line", "Demand ID", "Model.Suffix", "MAIN", "POWER", "SOPPORTER",
            "Tool", "Lot Qty "]
    for i, c in enumerate(cols, start=1):
        ws.cell(1, i, c)
    for i, f in enumerate(fechas):
        ws.cell(1, len(cols) + 1 + i, dt.datetime.combine(f, dt.time()))
    for fila, (wo, modelo, partes, por_dia) in enumerate(renglones, start=2):
        ws.cell(fila, 1, "MG1")
        ws.cell(fila, 2, wo)
        ws.cell(fila, 3, modelo)
        for j, parte in enumerate(partes):
            ws.cell(fila, 4 + j, parte)
        ws.cell(fila, 8, sum(por_dia.values()))
        for f, qty in por_dia.items():
            ws.cell(fila, len(cols) + 1 + fechas.index(f), qty)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_sub_ensamble_oven_ignora_el_rotulo_de_la_columna():
    """POWER/SOPPORTER vienen INTERCAMBIADOS entre archivos: manda el AJJ."""
    # El AJJ es el ensamble y su pareja el PCB, este en la columna que este.
    assert ppn._sub_oven(["AJJ30036901", "EBR43713702"]) == {"AJJ30036901": "EBR43713702"}
    assert ppn._sub_oven(["EBR43713702", "AJJ30036901"]) == {"AJJ30036901": "EBR43713702"}
    # Sin ensamble reconocible no se inventa una relacion.
    assert ppn._sub_oven(["EBR43713604", "EBR43713702"]) == {}
    assert ppn._sub_oven(["AJJ30036901", "NO USA"]) == {}
    assert ppn._sub_oven(["AJJ30036901", "AJJ30036902"]) == {}


def test_sub_ensambles_del_oven_en_los_dos_ordenes():
    hoy = DIA + dt.timedelta(days=1)
    # Mismo par, columnas al reves: la relacion tiene que salir igual.
    a = _oven_workbook(hoy, [("W1", "M", ["EBR1", "AJJ30036901", "EBR43713702"],
                              {hoy: 10})])
    b = _oven_workbook(hoy, [("W1", "M", ["EBR1", "EBR43713702", "AJJ30036901"],
                              {hoy: 10})])
    esperado = {"AJJ30036901": "EBR43713702"}
    assert ppn.sub_ensambles(ppn.leer_bytes(a)[1]) == esperado
    assert ppn.sub_ensambles(ppn.leer_bytes(b)[1]) == esperado


def test_leer_oven_reconoce_la_hoja_y_las_partes():
    data = _oven_workbook(DIA, [("6H1X1", "LSGL5831F", ["EBR1", "AJJ1", "#N/A"],
                                 {DIA: 100, DIA + dt.timedelta(days=1): 20})])
    fecha0, wos, fuente = ppn.leer_bytes(data)
    assert (fecha0, fuente) == (DIA, "OVEN")
    w = wos["6H1X1"]
    assert w["partes"] == {"EBR1", "AJJ1"} and w["sin_bom"] is True
    # Sin acumulado construido: sobrante() compara el pendiente, no un result.
    assert w["result"] is None and w["plan"] == 120


def test_sobrante_oven_mide_la_baja_del_pendiente():
    """OVEN no reporta construido: lo que bajo el pendiente es lo que LG metio."""
    hoy = DIA + dt.timedelta(days=1)
    _, ayer, _ = ppn.leer_bytes(_oven_workbook(
        DIA, [("W1", "M", ["P1", "P2"], {DIA: 100, hoy: 50}),
              ("W2", "M", ["P3"], {DIA: 30})]))
    # W1 quedo en 90 pendientes (construyo 60 de las 100 de ayer); W2 desaparecio.
    _, act, _ = ppn.leer_bytes(_oven_workbook(hoy, [("W1", "M", ["P1", "P2"], {hoy: 90})]))
    renglones, partes = ppn.sobrante(ppn.baseline(DIA, ayer), act)
    assert dict((w, (p, r)) for w, p, r in renglones) == {"W1": (100, 60), "W2": (30, 30)}
    assert partes == {"P1": 40, "P2": 40}  # W2 cerro completo: no deja sobrante


def test_parse_oven_rebaja_el_plan_del_dia_que_paso(monkeypatch):
    hoy = DIA + dt.timedelta(days=1)
    base = (DIA, {"W1": {"partes": {"P1"}, "prog": 100, "pend": 150,
                         "plan": 150, "result": None}})
    monkeypatch.setattr(pp, "_pp_ppn_baseline", lambda *_a: (None, {}))
    monkeypatch.setattr(pp, "execute_query", lambda *_a, **_k: [
        {"part_no": "P1", "plan_qty": 120},
    ])
    data = _oven_workbook(hoy, [("W1", "M", ["P1"], {hoy: 90})])

    parsed, err = pp._parse_lg_workbook(data, "Prod Plan JULIO 30 2026.xlsx", ppn_base=base)

    assert err is None
    assert parsed["ppn_fuente"] == "OVEN"
    # Construyo 60 de 100 -> 40 pzas de P1 siguen en piso: 120 - 40.
    assert parsed["records"][("P1", DIA)] == 80
    assert parsed["records"][("P1", hoy)] == 90
    assert parsed["ppn_wo"][0][0] == "OVEN"


def _lookup_de(*archivos):
    """file_lookup falso: `archivos` va del mas reciente al mas viejo."""
    def lookup(_file_ref=None):
        return archivos[0] if archivos else (None, None)
    lookup.recientes = lambda limite=4: list(archivos[:limite])
    return lookup


def test_tool_ppn_comparar_ordena_por_la_fecha_del_archivo(monkeypatch):
    """Da igual en que orden se suban: manda la fecha de adentro del PPN."""
    monkeypatch.setattr(ai_plan_tools, "_has_plan", lambda _u: True)
    hoy = DIA + dt.timedelta(days=1)
    viejo = _ppn_workbook(DIA, [("A", "M1", ["P1", "P2"], 100, 0, {DIA: 100})])
    nuevo = _ppn_workbook(hoy, [("A", "M1", ["P1", "P2"], 100, 60, {hoy: 40})])

    for archivos in (
        ((nuevo, "PPN 30.xlsx"), (viejo, "PPN 29.xlsx")),   # subidos en orden
        ((viejo, "PPN 29.xlsx"), (nuevo, "PPN 30.xlsx")),   # subidos al reves
    ):
        out = ai_plan_tools.execute(
            "plan_ppn_comparar", {}, username="ana", file_lookup=_lookup_de(*archivos),
        )
        r = out["resumen"]
        assert r["dia_evaluado"] == DIA.isoformat()
        assert (r["unidades_planeadas"], r["unidades_construidas"]) == (100, 60)
        assert r["pzas_sin_consumir"] == 80  # 40 unidades x 2 partes
        assert out["confirm_token"]


def test_propuesta_stale_pide_generar_una_nueva():
    """STALE sale como PPYProposalStaleError -> 409 + code, no como 400 sin salida."""
    err = pp._ppy_estado_no_abierto("STALE", "aplicar")
    assert isinstance(err, pp.PPYProposalStaleError)
    assert "Genera una propuesta nueva" in str(err)
    # Los demas estados siguen siendo un ValueError normal.
    otro = pp._ppy_estado_no_abierto("REJECTED", "aplicar")
    assert isinstance(otro, ValueError) and not isinstance(otro, pp.PPYProposalStaleError)


def _stub_dia(monkeypatch, schedule_changes, reportes=None, visto=None):
    """plan_dia_preparar sin BD: solo se ejercita el armado de la salida."""
    monkeypatch.setattr(ai_plan_tools, "_has_plan", lambda _u: True)
    monkeypatch.setattr(ai_plan_tools, "_has_projection", lambda _u: True)
    monkeypatch.setattr(ai_plan_tools, "_reportes", lambda *_a, **_k: (
        {"PPN": [(DIA, {}, b"ayer", "PPN 29.xlsx"), (DIA, {}, b"hoy", "PPN 30.xlsx")]}
        if reportes is None else reportes))
    monkeypatch.setattr(ai_plan_tools, "_base_de", lambda *_a: (DIA, {"W": {}}))
    monkeypatch.setattr(ai_plan_tools, "_ppn_comparar", lambda *_a: {
        "dia_evaluado": DIA.isoformat(), "unidades_planeadas": 100,
        "unidades_construidas": 60, "unidades_atrasadas": 40, "pzas_sin_consumir": 80})
    monkeypatch.setattr(ai_plan_tools, "_importar_ppn", lambda *_a: {"plan_registros": 10})
    monkeypatch.setattr(ai_plan_tools, "_cal_adjunto", lambda *_a, **_k: (None, None))
    monkeypatch.setattr(ai_plan_tools.pp, "_pp_sub_ensambles", lambda _p: {})
    # Sin bloqueos guardados salvo que el test diga lo contrario.
    monkeypatch.setattr(ai_plan_tools.pp, "_pp_bloqueos_activos", lambda **_k: {
        "activos": [], "partes": [], "por_revisar": []})
    real = ai_plan_tools.execute

    def fake(name, arguments, **kw):
        if name == "plan_propuesta_preparar":
            if visto is not None:
                visto.update(arguments)
            return {"schedule_changes": schedule_changes, "confirm_token": "tok",
                    "proposal_id": "p1", "schedule_change_summary": {}}
        return real(name, arguments, **kw)

    monkeypatch.setattr(ai_plan_tools, "execute", fake)
    return real


def test_horas_parte_sale_del_planned_start_del_ppn():
    """El W/O que arranca hoy ocupa a su hora; el arrastrado, desde que abre."""
    hoy = DIA + dt.timedelta(days=1)
    wos = {
        # Arranca el 29 a las 13:00 y sigue el 30: el 30 ya viene corriendo.
        "A": {"partes": {"P1"}, "dias": {DIA: 50, hoy: 40},
              "inicio": dt.datetime(2026, 7, 29, 13, 0)},
        # Otro W/O de la misma parte arranca el 29 mas temprano: gana el minimo.
        "B": {"partes": {"P1"}, "dias": {DIA: 10},
              "inicio": dt.datetime(2026, 7, 29, 10, 15)},
        "C": {"partes": {"P2"}, "dias": {hoy: 30},
              "inicio": dt.datetime(2026, 7, 30, 9, 45)},
        "D": {"partes": {"P3"}, "dias": {hoy: 5}, "inicio": None},  # sin hora
    }
    horas = ppn.horas_parte(wos)
    assert horas[("P1", DIA)] == dt.time(10, 15)
    assert horas[("P1", hoy)] == ppn.HORA_TURNO
    assert horas[("P2", hoy)] == dt.time(9, 45)
    assert ("P3", hoy) not in horas


def _candidato(part_no, hora, linea="M1", falt=100, uph=100):
    return {"part_no": part_no, "falt_total": falt, "falt_hoy": falt,
            "primera_falta": DIA, "line": linea, "uph": uph, "pack": 10,
            "hora_necesaria": hora, "ct": 36.0, "permitidas": [linea],
            "model": None, "main_sub": None}


def test_apretado_secuencia_por_hora_y_no_por_familia():
    """Sin horas para todo, primero lo que LG ocupa mas temprano."""
    # Las dos EBR999 son una familia y piden mas piezas: agrupando van primero.
    cands = [
        _candidato("EBR9990001", dt.time(14, 0), falt=200),
        _candidato("EBR9990002", dt.time(15, 0)),
        _candidato("AJJ3003690", dt.time(7, 30)),   # se ocupa al abrir
    ]
    # 1 h de bloque para 3 lotes de 1 h: solo cabe uno.
    lotes, _fuera = pp._ppy_armar_lotes(
        cands, {"M1"}, {"B1": 1.0}, estricto=True, apretado=True)
    assert [l["part_no"] for l in lotes] == ["AJJ3003690"]

    # Sin apretado manda la familia: entra la primera EBR, no la urgente.
    lotes, _fuera = pp._ppy_armar_lotes(
        cands, {"M1"}, {"B1": 1.0}, estricto=True, apretado=False)
    assert [l["part_no"] for l in lotes] == ["EBR9990001"]


def test_dia_limite_adelanta_lo_que_entra_al_abrir():
    """Si LG lo mete a las 07:30 de mañana, el lote tiene que salir hoy."""
    viernes, jueves = dt.date(2026, 7, 31), dt.date(2026, 7, 30)
    # A las 07:30 no hay ni un minuto de turno antes: se va al dia anterior.
    assert pp._ppy_dia_limite(viernes, dt.time(7, 30), 1.0) == jueves
    # A las 14:27 caben 6 h de turno antes (menos breaks): se alcanza ese dia.
    assert pp._ppy_dia_limite(viernes, dt.time(14, 27), 1.0) == viernes
    # ...pero no si el lote pide mas horas de las que quedan antes de esa hora.
    assert pp._ppy_dia_limite(viernes, dt.time(14, 27), 9.0) == jueves
    # Sin hora conocida no se toca el dia del faltante.
    assert pp._ppy_dia_limite(viernes, None, 9.0) == viernes
    # El lunes se adelanta al VIERNES, no al domingo.
    lunes = dt.date(2026, 8, 3)
    assert pp._ppy_dia_limite(lunes, dt.time(7, 30), 1.0) == dt.date(2026, 7, 31)


def test_lo_que_entra_manana_temprano_gana_las_horas_de_hoy():
    """Compite con los faltantes de hoy, no con los de mañana."""
    manana = DIA + dt.timedelta(days=1)
    hoy_tarde = _candidato("EBR1110001", dt.time(16, 0))
    manana_0730 = {**_candidato("AJJ3003690", dt.time(7, 30)),
                   "primera_falta": manana,
                   "falta_efectiva": pp._ppy_dia_limite(manana, dt.time(7, 30), 1.0)}
    # Sin la fecha limite, el de mañana iria despues por fecha y se quedaria fuera.
    assert manana_0730["falta_efectiva"] == DIA
    lotes, _f = pp._ppy_armar_lotes(
        [hoy_tarde, manana_0730], {"M1"}, {"B1": 1.0}, estricto=True, apretado=True)
    assert [l["part_no"] for l in lotes] == ["AJJ3003690"]
    assert "tiene que salir antes" in (lotes[0]["comentario"] or "")


def test_apretado_sin_hora_conocida_no_se_adelanta():
    """Una parte sin hora (OVEN, Cal) va despues de las que si la traen."""
    cands = [_candidato("SINHORA001", None), _candidato("CONHORA001", dt.time(11, 0))]
    lotes, _f = pp._ppy_armar_lotes(
        cands, {"M1"}, {"B1": 1.0}, estricto=True, apretado=True)
    assert [l["part_no"] for l in lotes] == ["CONHORA001"]


def test_d3_cuesta_horas_de_otra_linea_por_el_tiempo_que_corre():
    """D3 se lleva el personal de otro equipo MIENTRAS corre, no el dia entero."""
    lineas = {"M1", "M2", "D3"}
    bloques = lambda: {"B1": 9.0, "B2": 9.0, "B3": 9.0}   # tres equipos
    # uph 100: 900 pzas = 9 h de linea, 200 pzas = 2 h.
    cands = [_candidato("EBRM100001", dt.time(8, 0), linea="M1", falt=900),
             _candidato("EBRM200001", dt.time(9, 0), linea="M2", falt=900)]
    qty = lambda lotes: {l["linea"]: l["qty"] for l in lotes}

    # Sin D3 las dos main producen completo.
    lotes, fuera = pp._ppy_armar_lotes(cands, lineas, bloques(), estricto=True)
    assert qty(lotes) == {"M1": 900, "M2": 900} and not fuera

    # D3 corriendo 2 h le quita 2 h (220 pzas) a una main, no el dia completo.
    corto = _candidato("ACQD300001", dt.time(7, 30), linea="D3", falt=200)
    lotes, fuera = pp._ppy_armar_lotes(cands + [corto], lineas, bloques(),
                                       estricto=True)
    producido = qty(lotes)
    assert producido["D3"] == 220
    assert sorted(producido.values()) == [220, 680, 900]  # una main pierde 220
    assert not fuera  # nadie deja de trabajar por 2 h de D3

    # D3 el turno entero si para una main completa, y lo dice con su motivo.
    largo = _candidato("ACQD300001", dt.time(7, 30), linea="D3", falt=900)
    lotes, fuera = pp._ppy_armar_lotes(cands + [largo], lineas, bloques(),
                                       estricto=True)
    assert len(qty(lotes)) == 2 and qty(lotes)["D3"] == 900
    d3_motivos = [f["motivo"] for f in fuera
                  if "ocupa el personal de un equipo" in f["motivo"]]
    assert d3_motivos, [f["motivo"] for f in fuera]
    assert pp._ppy_omission_codes(d3_motivos[0]) == ["D3_STOPS_LINE"]


def test_plan_params_congela_el_techo_de_reloj():
    """El apply re-simula con el MISMO reloj: si no, avanzar 10 min = STALE."""
    assert pp._ppy_parse_plan_params({"horas_restantes_hoy": 2.75})[
        "horas_restantes_hoy"] == 2.75
    assert pp._ppy_parse_plan_params(
        '{"horas_restantes_hoy": 1.5}')["horas_restantes_hoy"] == 1.5
    # Propuestas viejas sin la clave: turno completo, como siempre.
    assert pp._ppy_parse_plan_params(None)["horas_restantes_hoy"] is None
    assert pp._ppy_parse_plan_params('{"max_bloques": 3}')["horas_restantes_hoy"] is None


def test_horas_restantes_del_turno_siguen_el_reloj():
    """El UPH da las horas del lote; esto da contra cuantas se comparan."""
    jueves = dt.date(2026, 7, 30)
    en = lambda h, m: dt.datetime(2026, 7, 30, h, m)
    # Antes de abrir el turno estan las 9 h completas; ya cerrado, ninguna.
    assert pp._ppy_horas_restantes(jueves, en(6, 0)) == 9.0
    assert pp._ppy_horas_restantes(jueves, en(18, 0)) == 0.0
    # A media tarde solo queda el reloj, ya sin los breaks que pasaron.
    assert pp._ppy_horas_restantes(jueves, en(16, 0)) == 1.5
    assert pp._ppy_horas_restantes(jueves, en(11, 0)) == 5.75  # 6.5 h - 45 min
    # Otro dia no depende de la hora; el sabado son 8 h y el pasado ya no da.
    assert pp._ppy_horas_restantes(jueves + dt.timedelta(days=1), en(16, 0)) == 9.0
    assert pp._ppy_horas_restantes(dt.date(2026, 8, 1), en(16, 0)) == 8.0
    assert pp._ppy_horas_restantes(jueves - dt.timedelta(days=1), en(9, 0)) == 0.0


def test_el_lunes_el_turno_cierra_a_las_cuatro():
    """Tras el fin de semana falta gente: todas las lineas paran a las 16:00."""
    lunes, martes = dt.date(2026, 8, 3), dt.date(2026, 8, 4)
    en = lambda d, h: dt.datetime(2026, 8, d, h, 0)
    # El tope por BLOQUE del motor baja de 9 h a 7.5 (07:30-16:00 - 1 h de breaks).
    assert pp._ppy_horas_turno_dia(lunes) == pp.PPY_HORAS_LUNES == 7.5
    assert pp._ppy_horas_turno_dia(martes) is None      # dia normal, 9 h
    assert pp._ppy_horas_turno_dia(dt.date(2026, 8, 8)) == pp.PPY_HORAS_SABADO
    # Y el reloj del dia sigue el mismo corte, sin que nadie lo pida.
    assert pp._ppy_horas_restantes(lunes, en(3, 6)) == 7.5
    assert pp._ppy_horas_restantes(martes, en(3, 6)) == 9.0
    assert pp._ppy_horas_restantes(lunes, en(3, 15)) == 0.75   # 15:00-16:00 - break
    assert pp._ppy_horas_restantes(lunes, en(3, 16)) == 0.0
    # Un lote que LG ocupa el lunes despues de las 16:00 ya no alcanza a correr
    # ese dia: se adelanta al viernes (el sabado y domingo no se produce).
    assert pp._ppy_dia_limite(lunes, dt.time(17, 0), 8.0) == dt.date(2026, 7, 31)
    assert pp._ppy_dia_limite(martes, dt.time(17, 0), 8.0) == martes


def test_las_dos_palancas_del_turno_se_autorizan_a_mano():
    """El lunes completo y el tiempo extra se piden; ninguno se activa solo."""
    lunes, martes = dt.date(2026, 8, 3), dt.date(2026, 8, 4)
    hhmm = lambda m: f"{m // 60:02d}:{m % 60:02d}"
    fin = pp._ppy_fin_turno
    # Default: el lunes cierra a las 16:00 y los demas a las 17:30.
    assert hhmm(fin(lunes)) == "16:00" and hhmm(fin(martes)) == "17:30"
    # "Ese lunes si llego la gente": corre el turno completo.
    assert hhmm(fin(lunes, turno_completo=True)) == "17:30"
    assert pp._ppy_horas_turno_dia(lunes, turno_completo=True) is None  # 9 h
    # Tiempo extra: +3 h sobre lo que cierre ESE dia, no siempre 20:30.
    assert hhmm(fin(lunes, tiempo_extra=True)) == "19:00"
    assert hhmm(fin(martes, tiempo_extra=True)) == "20:30"
    # Las dos juntas componen.
    assert hhmm(fin(lunes, tiempo_extra=True, turno_completo=True)) == "20:30"
    # El inventario gana sobre el dia de la semana, y tambien admite TE.
    viernes = dt.date(2026, 8, 28)
    assert hhmm(fin(viernes, inventario=True)) == "14:00"
    assert hhmm(fin(viernes, inventario=True, tiempo_extra=True)) == "17:00"

    # Y el reloj sigue esos cierres: con TE el dia entero da 12 h, no 9.
    ayer = dt.datetime(2026, 8, 2, 9, 0)
    horas = lambda **kw: pp._ppy_horas_restantes(lunes, ayer, fin_min=fin(lunes, **kw))
    assert (horas(), horas(turno_completo=True)) == (7.5, 9.0)
    assert (horas(tiempo_extra=True), horas(tiempo_extra=True, turno_completo=True)) \
        == (10.5, 12.0)


def test_viernes_de_cierre_es_el_mas_cercano_al_fin_de_mes():
    """El mes no cierra entre semana: el corte se va al viernes mas cercano."""
    # Agosto 2026 termina lunes 31, asi que cierra ANTES, el viernes 28.
    assert pp._ppy_viernes_cierre(dt.date(2026, 8, 28))
    assert not pp._ppy_viernes_cierre(dt.date(2026, 9, 4))
    # Septiembre termina miercoles 30: su viernes 25 no cierra, cierra el 2/10.
    assert not pp._ppy_viernes_cierre(dt.date(2026, 9, 25))
    assert pp._ppy_viernes_cierre(dt.date(2026, 10, 2))
    # Octubre termina viernes 30: cierra el mismo dia.
    assert pp._ppy_viernes_cierre(dt.date(2026, 10, 30))
    # Nunca dos viernes seguidos, y jamas otro dia de la semana.
    assert not pp._ppy_viernes_cierre(dt.date(2026, 8, 21))
    assert not pp._ppy_viernes_cierre(dt.date(2026, 8, 31))
    # Cada mes cierra en exactamente un viernes de su ventana, ni dos ni cero.
    for mes in range(1, 13):
        fin = dt.date(2026 + mes // 12, mes % 12 + 1, 1) - dt.timedelta(days=1)
        cerca = [fin + dt.timedelta(days=n) for n in range(-3, 4)]
        assert sum(pp._ppy_viernes_cierre(d) for d in cerca) == 1, fin


def test_dia_de_inventario_corta_el_turno_a_las_dos():
    """Con inventario todas las lineas paran a las 14:00, hoy o a futuro."""
    viernes = dt.date(2026, 8, 28)
    en = lambda h, m: dt.datetime(2026, 8, 28, h, m)
    fin = pp.PPY_FIN_INVENTARIO_MIN
    # 07:30 a 14:00 son 6.5 h de reloj menos los breaks de 09:30 y 12:00.
    assert pp._ppy_horas_restantes(viernes, en(6, 0), fin_min=fin) == 5.75
    # A media manana solo lo que falta, sin el break que ya paso.
    assert pp._ppy_horas_restantes(viernes, en(10, 0), fin_min=fin) == 3.5
    # Pasadas las 14:00 ya no cabe nada, aunque el turno normal siguiera.
    assert pp._ppy_horas_restantes(viernes, en(14, 30), fin_min=fin) == 0.0
    assert pp._ppy_horas_restantes(viernes, en(14, 30)) == 2.75
    # Un viernes futuro tampoco trae el turno completo si lleva inventario.
    assert pp._ppy_horas_restantes(viernes, en(6, 0)) == 9.0
    assert pp._ppy_horas_restantes(
        viernes, dt.datetime(2026, 8, 24, 9, 0), fin_min=fin) == 5.75


def test_plan_dia_marca_lo_que_ya_no_da_tiempo(monkeypatch):
    """Un cambio que pide mas horas de las que quedan de turno se marca."""
    real = _stub_dia(monkeypatch, [
        {"accion": "AGREGAR", "part_no": "P1", "horas": 0.5, "despues_linea": "M1"},
        {"accion": "MODIFICAR", "part_no": "P4", "horas": 0.8, "despues_linea": "M1"},
        {"accion": "AGREGAR", "part_no": "P2", "horas": 4.0, "despues_linea": "D1"},
        {"accion": "ELIMINAR", "part_no": "P3", "horas": 3.0, "antes_linea": "D1"},
    ])
    # 16:00: quedan 1.5 h de turno.
    monkeypatch.setattr(ai_plan_tools, "obtener_fecha_hora_mexico",
                        lambda: dt.datetime(2026, 7, 30, 16, 0))
    out = real("plan_dia_preparar", {"fecha": "2026-07-30"}, username="ana",
               file_lookup=None)

    reloj = out["reloj"]
    assert (reloj["hora_consultada"], reloj["horas_restantes_turno"]) == ("16:00", 1.5)
    # Las lineas corren en paralelo: la carga se acumula por linea, no en un
    # total. Eliminar no consume horas.
    assert reloj["horas_que_piden_por_linea"] == {"M1": 1.3, "D1": 4.0}
    assert reloj["lineas_sin_tiempo"] == ["D1"]
    assert reloj["alcanza_el_turno"] is False
    marcas = {c["part_no"]: c.get("alcanza_hoy") for c in out["cambios"]}
    assert marcas == {"P1": True, "P4": True, "P2": False, "P3": None}


def test_plan_dia_lleva_las_palancas_hasta_el_motor(monkeypatch):
    """Lo que autoriza Planning cambia el plan, no solo el texto del reporte."""
    visto = {}
    real = _stub_dia(monkeypatch, [
        {"accion": "AGREGAR", "part_no": "P1", "horas": 8.0, "despues_linea": "M1"},
    ], visto=visto)
    # Lunes 07:00, antes de abrir: el dia completo por delante.
    monkeypatch.setattr(ai_plan_tools, "obtener_fecha_hora_mexico",
                        lambda: dt.datetime(2026, 8, 3, 7, 0))
    lunes = {"fecha": "2026-08-03", "inventario": None,
             "turno_completo": None, "tiempo_extra": None}

    # Por default el lunes son 7.5 h: un lote de 8 h ya no cabe.
    out = real("plan_dia_preparar", lunes, username="ana", file_lookup=None)
    assert out["reloj"]["horas_restantes_turno"] == 7.5
    assert out["reloj"]["cierra_turno"] == "16:00 (lunes: falta gente)"
    assert out["reloj"]["alcanza_el_turno"] is False
    assert "turno_completo=true" in out["instruccion"]
    assert visto["turno_completo"] is False and visto["horas_restantes_hoy"] == 7.5

    # "Ese lunes si llego la gente": 9 h y el lote cabe.
    out = real("plan_dia_preparar", {**lunes, "turno_completo": True},
               username="ana", file_lookup=None)
    assert out["reloj"]["horas_restantes_turno"] == 9.0
    assert out["reloj"]["cierra_turno"] == "17:30"
    assert out["reloj"]["alcanza_el_turno"] is True
    assert visto["turno_completo"] is True  # y el motor tambien lo sabe

    # Tiempo extra sobre el lunes corto: +3 h desde las 16:00.
    out = real("plan_dia_preparar", {**lunes, "tiempo_extra": True},
               username="ana", file_lookup=None)
    assert out["reloj"]["horas_restantes_turno"] == 10.5
    assert out["reloj"]["cierra_turno"] == "19:00 (lunes: falta gente) +tiempo extra"
    assert visto["tiempo_extra"] is True


def test_plan_dia_pregunta_el_inventario_de_cierre_antes_de_cargar(monkeypatch):
    """El viernes de cierre no se arma a ciegas: primero se pregunta."""
    real = _stub_dia(monkeypatch, [
        {"accion": "AGREGAR", "part_no": "P1", "horas": 4.0, "despues_linea": "M1"},
    ])
    monkeypatch.setattr(ai_plan_tools, "obtener_fecha_hora_mexico",
                        lambda: dt.datetime(2026, 8, 28, 8, 0))
    # Importar consume el delta contra la base: si se corta despues, la segunda
    # llamada ya no tendria con que corregir. Debe cortarse ANTES.
    cargas = []
    monkeypatch.setattr(ai_plan_tools, "_importar_ppn",
                        lambda *_a: cargas.append(1) or {"plan_registros": 10})

    viernes = {"fecha": "2026-08-28", "inventario": None}
    out = real("plan_dia_preparar", viernes, username="ana", file_lookup=None)
    assert out["viernes_cierre_mes"] and "inventario" in out["instruccion"]
    assert "cambios" not in out and not cargas

    # Respondiendo que si, el dia se arma contra el turno cortado a las 14:00.
    out = real("plan_dia_preparar", {**viernes, "inventario": True},
               username="ana", file_lookup=None)
    assert out["reloj"]["horas_restantes_turno"] == 5.25  # 08:00-14:00 - breaks
    assert out["reloj"]["cierra_turno"] == "14:00 (inventario)"
    assert out["cambios"] and cargas

    # Respondiendo que no, es un viernes cualquiera que corre hasta las 17:30.
    out = real("plan_dia_preparar", {**viernes, "inventario": False},
               username="ana", file_lookup=None)
    assert out["reloj"]["horas_restantes_turno"] == 8.5  # 08:00-17:30 - breaks
    assert out["reloj"]["cierra_turno"] == "17:30"


def test_plan_dia_pregunta_el_inventario_de_sub_ensamble(monkeypatch):
    """El MES no tiene ese inventario: el asistente lo pregunta, no lo supone."""
    real = _stub_dia(monkeypatch, [
        {"accion": "AGREGAR", "part_no": "ABQ30226610", "despues_qty": 1040},
        {"accion": "AGREGAR", "part_no": "EBR76683912", "despues_qty": 500},
        {"accion": "ELIMINAR", "part_no": "AJJ30036901", "despues_qty": 0},
    ])
    monkeypatch.setattr(ai_plan_tools.pp, "_pp_sub_ensambles", lambda _p: {
        "ABQ30226610": ["EBR88124627"], "AJJ30036901": ["EBR43713702"]})

    out = real("plan_dia_preparar", {"fecha": None}, username="ana", file_lookup=None)

    # Solo lo que se va a producir; eliminar no consume PCB.
    assert out["sub_ensambles"] == [
        {"part_no": "ABQ30226610", "pzas": 1040, "sub_partes": ["EBR88124627"]}]
    assert "PREGUNTA cuanto inventario hay de los PCB" in out["instruccion"]
    assert "no lo supongas" in out["instruccion"]


def test_plan_dia_sin_sub_ensamble_no_pregunta_de_mas(monkeypatch):
    real = _stub_dia(monkeypatch, [{"accion": "AGREGAR", "part_no": "P1"}])
    out = real("plan_dia_preparar", {"fecha": None}, username="ana", file_lookup=None)
    assert out["sub_ensambles"] == []
    assert "PREGUNTA cuanto inventario" not in out["instruccion"]


def test_plan_dia_solo_devuelve_lo_que_cambia(monkeypatch):
    cambios = [
        {"accion": "CONSERVAR", "part_no": "P1"},
        {"accion": "AGREGAR", "part_no": "P2"},
        {"accion": "ELIMINAR", "part_no": "P3"},
    ]
    real = _stub_dia(monkeypatch, cambios)
    out = real("plan_dia_preparar", {"fecha": None}, username="ana", file_lookup=None)
    assert [c["part_no"] for c in out["cambios"]] == ["P2", "P3"]
    assert out["sin_cambios"] is False
    assert out["confirm_token"] == "tok"


def test_plan_dia_junta_ppn_y_oven_en_una_propuesta(monkeypatch):
    """Los dos flujos entran en la misma llamada y salen en una sola propuesta."""
    real = _stub_dia(monkeypatch, [{"accion": "AGREGAR", "part_no": "P2"}], reportes={
        "PPN": [(DIA, {}, b"p29", "PPN 29.xlsx"), (DIA, {}, b"p30", "PPN 30.xlsx")],
        # De OVEN solo el de hoy: la base sale del snapshot en MES.
        "OVEN": [(DIA, {}, b"o30", "Prod Plan JULIO 30 2026.xlsx")],
    })
    out = real("plan_dia_preparar", {"fecha": None}, username="ana", file_lookup=None)

    assert sorted(out["lg_ayer"]) == ["OVEN", "PPN"]
    assert [r.split(":")[0] for r in out["datos_cargados"]["reportes"]] == ["OVEN", "PPN"]
    # Una sola propuesta con un solo token para los dos flujos.
    assert out["confirm_token"] == "tok" and out["proposal_id"] == "p1"


def test_plan_dia_funciona_con_solo_el_cal(monkeypatch):
    """Sin reportes de LG sigue armando el dia, pero avisa que falta la correccion."""
    real = _stub_dia(monkeypatch, [{"accion": "AGREGAR", "part_no": "P2"}], reportes={})
    monkeypatch.setattr(ai_plan_tools, "_cal_adjunto",
                        lambda *_a, **_k: (b"cal", "Cal_260730.xlsx"))
    monkeypatch.setattr(ai_plan_tools, "_importar_plan_e_inventario",
                        lambda *_a: {"plan_registros": 99})

    out = real("plan_dia_preparar", {"fecha": None}, username="ana", file_lookup=None)

    assert out["lg_ayer"] is None
    assert out["datos_cargados"]["cal"].startswith("99 registros")
    assert "ningun PPN ni Prod Plan" in out["datos_cargados"]["reportes"]
    assert "correccion del consumo real de LG" in out["instruccion"]
    assert out["cambios"] and out["confirm_token"] == "tok"


def test_plan_dia_sin_archivos_utiles_lo_dice(monkeypatch):
    real = _stub_dia(monkeypatch, [], reportes={})
    monkeypatch.setattr(ai_plan_tools, "_cal_adjunto", lambda *_a, **_k: (None, None))
    with pytest.raises(ValueError, match="ni el Cal del dia ni un reporte de LG"):
        real("plan_dia_preparar", {"fecha": None}, username="ana", file_lookup=None)


def test_plan_dia_sin_cambios_no_pide_confirmacion(monkeypatch):
    real = _stub_dia(monkeypatch, [{"accion": "CONSERVAR", "part_no": "P1"}])
    out = real("plan_dia_preparar", {"fecha": None}, username="ana", file_lookup=None)
    assert out["sin_cambios"] is True and out["cambios"] == []
    # Sin token no hay confirmacion pendiente que el servidor pueda ejecutar.
    assert "confirm_token" not in out


def test_ppn_par_no_mezcla_ppn_con_oven(monkeypatch):
    """Un PPN y un Prod Plan no se comparan entre si: son dos flujos distintos."""
    monkeypatch.setattr(ai_plan_tools, "_has_plan", lambda _u: True)
    hoy = DIA + dt.timedelta(days=1)
    lookup = _lookup_de(
        (_oven_workbook(hoy, [("W1", "M", ["P1"], {hoy: 90})]), "Prod Plan 30.xlsx"),
        (_ppn_workbook(DIA, [("A", "M1", ["P1"], 100, 0, {DIA: 100})]), "PPN 29.xlsx"),
    )
    with pytest.raises(ValueError, match="dos reportes del mismo tipo"):
        ai_plan_tools._ppn_par(lookup)

    # Con los dos Prod Plan si empareja, y por la fecha de adentro del archivo.
    lookup = _lookup_de(
        (_oven_workbook(hoy, [("W1", "M", ["P1"], {hoy: 90})]), "Prod Plan 30.xlsx"),
        (_oven_workbook(DIA, [("W1", "M", ["P1"], {DIA: 100, hoy: 50})]), "Prod Plan 29.xlsx"),
    )
    anterior, actual = ai_plan_tools._ppn_par(lookup)
    assert (anterior[0], actual[0]) == (DIA, hoy)
    base = (anterior[0], ppn.baseline(anterior[0], anterior[1]))
    assert ai_plan_tools._ppn_comparar(base, actual, anterior)["fuente"] == "OVEN"


def _snapshot_oven(fecha=DIA):
    return (fecha, {"W1": {"modelo": "M", "partes": {"P1"}, "prog": 100,
                           "pend": 150, "plan": 150, "result": None}})


def test_base_de_usa_el_snapshot_del_mes_si_no_suben_el_de_ayer(monkeypatch):
    """Basta subir el reporte del dia: la base de ayer sale de lg_ppn_wo."""
    snap = _snapshot_oven()
    monkeypatch.setattr(ai_plan_tools.pp, "_pp_ppn_baseline",
                        lambda fuente: snap if fuente == "OVEN" else (None, {}))
    assert ai_plan_tools._base_de("OVEN", []) == snap
    assert ai_plan_tools._base_de("PPN", []) is None  # sin snapshot no hay base

    hoy = DIA + dt.timedelta(days=1)
    _, act, _ = ppn.leer_bytes(_oven_workbook(hoy, [("W1", "M", ["P1"], {hoy: 90})]))
    r = ai_plan_tools._ppn_comparar(snap, (hoy, act, b"", "Prod Plan 30.xlsx"))
    assert (r["unidades_planeadas"], r["unidades_construidas"]) == (100, 60)
    assert r["pzas_sin_consumir"] == 40
    assert "snapshot en MES" in r["ppn_anterior"]
    # Sin el archivo de ayer no hay movimiento de demanda que reportar.
    assert r["cambios_demanda"] == []


def test_ppn_comparar_no_corrige_si_el_snapshot_ya_es_de_hoy():
    """Reimportar el reporte del dia no vuelve a descontar el sobrante."""
    hoy = DIA + dt.timedelta(days=1)
    _, act, _ = ppn.leer_bytes(_oven_workbook(hoy, [("W1", "M", ["P1"], {hoy: 90})]))
    actual = (hoy, act, b"", "Prod Plan 30.xlsx")
    assert ai_plan_tools._ppn_comparar(_snapshot_oven(hoy), actual) is None
    assert ai_plan_tools._ppn_comparar(None, actual) is None


def test_tool_ppn_comparar_pide_el_segundo_archivo(monkeypatch):
    monkeypatch.setattr(ai_plan_tools, "_has_plan", lambda _u: True)
    solo = _ppn_workbook(DIA, [("A", "M1", ["P1"], 100, 0, {DIA: 100})])
    with pytest.raises(ValueError, match="dos PPN"):
        ai_plan_tools.execute(
            "plan_ppn_comparar", {}, username="ana",
            file_lookup=_lookup_de((solo, "PPN 29.xlsx")),
        )


def _stub_bloqueos(monkeypatch, filas, hoy=dt.date(2026, 8, 4)):
    """lg_part_bloqueo sin BD: solo la logica de que se recuerda y que se pregunta."""
    monkeypatch.setattr(pp, "obtener_fecha_hora_mexico",
                        lambda: dt.datetime.combine(hoy, dt.time(8, 0)))
    monkeypatch.setattr(pp, "execute_query", lambda *_a, **_k: filas)


def test_bloqueo_se_recuerda_y_solo_se_pregunta_al_vencer(monkeypatch):
    """Lo que Planning dijo ayer sigue valiendo hoy; el dia prometido se
    PREGUNTA, no se da por resuelto."""
    _stub_bloqueos(monkeypatch, [
        {"part_no": "EBR80757421", "motivo": "falta material",
         "revisar_el": dt.date(2026, 8, 6), "created_by": "yahir",
         "created_at": None},
        {"part_no": "ABQ30226610", "motivo": "molde en mantenimiento",
         "revisar_el": None, "created_by": "yahir", "created_at": None},
    ])
    # Dos dias antes: se recuerda y NO se planea, pero todavia no se pregunta.
    est = pp._pp_bloqueos_activos(hoy=dt.date(2026, 8, 4))
    assert est["partes"] == ["EBR80757421", "ABQ30226610"]
    assert est["por_revisar"] == []

    # El dia prometido: toca preguntar, pero sigue bloqueada hasta que confirmen.
    est = pp._pp_bloqueos_activos(hoy=dt.date(2026, 8, 6))
    assert [b["part_no"] for b in est["por_revisar"]] == ["EBR80757421"]
    assert "EBR80757421" in est["partes"], "no se libera sola al llegar la fecha"
    # Un bloqueo sin fecha nunca se pregunta solo: espera a que lo liberen.
    assert all(b["part_no"] != "ABQ30226610" for b in est["por_revisar"])

    # Pasada la fecha sigue tocando, no se olvida.
    est = pp._pp_bloqueos_activos(hoy=dt.date(2026, 8, 20))
    assert [b["part_no"] for b in est["por_revisar"]] == ["EBR80757421"]


def test_plan_dia_saca_del_plan_lo_bloqueado_y_lo_dice(monkeypatch):
    """El bloqueo llega al motor como exclusion, no solo al texto."""
    visto = {}
    real = _stub_dia(monkeypatch, [
        {"accion": "AGREGAR", "part_no": "P1", "horas": 1.0, "despues_linea": "M1"},
    ], visto=visto)
    monkeypatch.setattr(ai_plan_tools, "obtener_fecha_hora_mexico",
                        lambda: dt.datetime(2026, 8, 6, 8, 0))
    monkeypatch.setattr(ai_plan_tools.pp, "_pp_bloqueos_activos", lambda **_k: {
        "activos": [{"part_no": "EBR80757421", "motivo": "falta material",
                     "revisar_el": "2026-08-06", "toca_revisar": True,
                     "lo_puso": "yahir"}],
        "partes": ["EBR80757421"],
        "por_revisar": [{"part_no": "EBR80757421", "motivo": "falta material",
                         "revisar_el": "2026-08-06", "toca_revisar": True,
                         "lo_puso": "yahir"}],
    })
    out = real("plan_dia_preparar",
               {"fecha": "2026-08-06", "inventario": None,
                "turno_completo": None, "tiempo_extra": None},
               username="ana", file_lookup=None)

    # Al motor le llega como exclusion: su capacidad la ocupa otra parte.
    assert visto["partes_excluidas"] == ["EBR80757421"]
    assert out["bloqueos"]["partes"] == ["EBR80757421"]
    # Y el reporte dice ambas cosas: que quedo fuera y que hoy toca preguntar.
    assert "FUERA del plan" in out["instruccion"]
    assert "falta material" in out["instruccion"]
    assert "HOY toca revisar EBR80757421" in out["instruccion"]
    assert "PREGUNTA si ya se puede" in out["instruccion"]
