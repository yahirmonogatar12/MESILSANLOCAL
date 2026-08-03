"""PPN de LG: demanda pendiente por parte y consumo REAL del dia que paso.

Sin dependencias de Flask ni de BD: solo openpyxl. Lo usan part_planning.py
(import) y tools/ppn_delta.py (CLI).

Cada renglon del PPN es un W/O con su modelo, las partes que ILSAN le surte
(columnas I..Q, mismo layout que la hoja BOM del Cal_*.xlsx) y la cantidad
PENDIENTE repartida por dia (columnas BC..CH). Explotar esos buckets por esas
partes reproduce la hoja diaria del Cal (46/49 partes exactas el 30/07): el Cal
es el PPN explotado, asi que el PPN sirve igual para alimentar lg_plan_daily.

POR QUE IMPORTA EL CONSUMO REAL
Los buckets del PPN ya vienen NETOS de lo construido (W/O Result), asi que la
demanda futura se autocorrige sola. Lo que no se corrige es el pasado: el
renglon de AYER en lg_plan_daily sigue con lo PLANEADO, no con lo que LG metio.
Y la proyeccion arrastra I(t) = I(t-1) - P(t) + S(t) desde ref_date, o sea que
descuenta ese plan viejo completo. Si LG se atraso, ese material sigue en su
piso y ademas el PPN de hoy lo reprograma hacia adelante: se descuenta dos veces
y aparece un faltante fantasma que la IA vuelve a planear.

    sobrante(parte) = programado_ayer - construido_ayer

construido_ayer sale de comparar el W/O Result de dos snapshots del PPN.

OVEN
El "Prod Plan <MES> <DIA>.xlsx" es el mismo problema en el otro producto: la
hoja FP_Demand_Management ya explotada trae Demand ID, sus partes (MAIN, POWER,
SOPPORTER) y el pendiente repartido por dia. Explotarla reproduce su hoja CAL
exacta (320/320 celdas el 29 y el 30 de julio). Lo unico distinto es que no hay
acumulado construido: ahi lo construido se mide por cuanto bajo el pendiente
entre los dos reportes.
"""
import datetime as dt
import io
import re
from collections import defaultdict

import openpyxl

PPN_SHEET_RE = re.compile(r"^\s*ppn\b", re.IGNORECASE)

C_WO, C_MODELO = 4, 7
C_PARTES = range(8, 17)   # I..Q: MAIN, HARNESS x3, DISPLAY, PCB DISPLAY, SUB ENS, PCB SUB, LED
C_PSTART = 31             # Planned Start Time
C_PLAN, C_RESULT = 49, 51  # W/O Plan Qty, W/O Result
C_DIA0, C_DIAN = 54, 86   # BC..CH: pendiente repartido por dia
NO_PARTE = ("NO USA", "N/A", "-", "")

OVEN_PARTES = ("MAIN", "POWER", "SOPPORTER")
# En OVEN el ensamble y su PCB van en el par POWER/SOPPORTER, pero esos dos
# rotulos vienen INTERCAMBIADOS entre archivos (el 29/07 trae POWER=EBR43713702
# y SOPPORTER=AJJ30036901; el 30/07 al reves), asi que no se puede confiar en el
# nombre de la columna. Lo estable es la parte: el AJJ es el ensamble y el otro
# del par es el PCB que hay que producir antes.
# ponytail: prefijo, no catalogo. Si aparece otro ensamble que no sea AJJ, va aqui.
OVEN_ENSAMBLE_PREFIJOS = ("AJJ",)

# Columnas pareadas (ensamble, su PCB): para surtir el ABQ hay que producir
# antes su EBR. Verificado en el PPN del 30/07: 36 ensambles, un solo PCB cada
# uno, y ninguno de esos PCB se surte solo a LG (es nivel interno).
SUB_PARES = ((12, 13), (14, 15))  # DISPLAY/PCB DISPLAY, SUB ENSAMBLE/PCB SUB


def hoja_ppn(wb):
    """Nombre de la hoja PPN del workbook, o None si no es un PPN."""
    for nombre in wb.sheetnames:
        if PPN_SHEET_RE.match(nombre):
            return nombre
    return None


def leer_ppn(ws):
    """-> (fecha0, {wo: {modelo, partes, sin_bom, plan, result, dias{fecha: qty}}}).

    Lanza ValueError si el layout no es el esperado (columnas movidas).
    """
    filas = list(ws.iter_rows(values_only=True))
    if len(filas) < 3:
        raise ValueError("La hoja PPN no tiene renglones de W/O.")
    etiquetas = filas[0][C_DIA0:C_DIAN]
    inicios = [r[C_PSTART] for r in filas[2:]
               if len(r) > C_PSTART and isinstance(r[C_PSTART], dt.datetime)]
    if not inicios:
        raise ValueError("La hoja PPN no trae 'Planned Start Time' (columna AF).")
    fecha0 = min(inicios).date()
    fechas = [fecha0 + dt.timedelta(days=i) for i in range(len(etiquetas))]
    # Las etiquetas de esas columnas son el dia del mes: si no cuadran con
    # fecha0, el layout cambio y explotar los buckets pondria la demanda en
    # fechas equivocadas.
    try:
        dias_hoja = [int(e) for e in etiquetas]
    except (TypeError, ValueError):
        raise ValueError("Las columnas de dias del PPN (BC..CH) no traen dia del mes.")
    if dias_hoja != [f.day for f in fechas]:
        raise ValueError(
            f"Los dias del PPN no empiezan en {fecha0}: encabezados {dias_hoja[:5]}."
        )

    wos = {}
    for r in filas[2:]:
        if not r or r[0] is None or len(r) <= C_RESULT:
            continue
        crudas = [str(r[i]).strip() for i in C_PARTES if isinstance(r[i], str)]
        wos[str(r[C_WO])] = {
            "modelo": str(r[C_MODELO] or "").strip(),
            # "#N/A" = modelo sin BOM en el PPN; se reporta, no se explota.
            "partes": {p for p in crudas if p.upper() not in NO_PARTE and not p.startswith("#")},
            "sin_bom": any(p.startswith("#") for p in crudas),
            "plan": int(r[C_PLAN] or 0),
            "result": int(r[C_RESULT] or 0),
            "inicio": r[C_PSTART] if isinstance(r[C_PSTART], dt.datetime) else None,
            "sub": _sub_del_renglon(r),
            "dias": {f: int(r[C_DIA0 + i])
                     for i, f in enumerate(fechas)
                     if C_DIA0 + i < len(r) and r[C_DIA0 + i]},
        }
    if not wos:
        raise ValueError("La hoja PPN no tiene renglones de W/O.")
    return fecha0, wos


def hoja_oven(wb):
    """Hoja del Prod Plan de OVEN ya explotada a partes, o None."""
    for nombre in wb.sheetnames:
        fila = next(wb[nombre].iter_rows(min_row=1, max_row=1, values_only=True), None)
        cols = [str(v).strip() for v in (fila or ()) if v is not None]
        if "Demand ID" in cols and "MAIN" in cols:
            return nombre
    return None


def leer_oven(ws):
    """-> (fecha0, wos), mismo contrato que leer_ppn() para el Prod Plan de OVEN.

    OVEN no reporta acumulado construido, asi que result queda en None y
    sobrante() mide lo construido por la baja del pendiente entre dos reportes.
    """
    filas = ws.iter_rows(values_only=True)
    encabezado = next(filas, None)
    if not encabezado:
        raise ValueError("La hoja de OVEN esta vacia.")
    nombres = [str(v).strip() if v is not None else "" for v in encabezado]
    if "Demand ID" not in nombres:
        raise ValueError("La hoja de OVEN no trae la columna 'Demand ID'.")
    c_demand = nombres.index("Demand ID")
    c_modelo = nombres.index("Model.Suffix") if "Model.Suffix" in nombres else c_demand
    # El orden de MAIN/POWER/SOPPORTER cambia entre archivos: se buscan por nombre.
    c_partes = [nombres.index(p) for p in OVEN_PARTES if p in nombres]
    # El par ensamble/PCB son las dos columnas de parte que no son MAIN.
    c_par = [nombres.index(p) for p in ("POWER", "SOPPORTER") if p in nombres]
    dias = [(i, v.date()) for i, v in enumerate(encabezado) if isinstance(v, dt.datetime)]
    if not dias or not c_partes:
        raise ValueError("La hoja de OVEN no trae columnas de dia o de parte.")
    fecha0 = dias[0][1]

    wos = {}
    for r in filas:
        if not r or len(r) <= c_demand or r[c_demand] is None:
            continue  # el renglon de totales no trae Demand ID
        crudas = [str(r[i]).strip() for i in c_partes
                  if i < len(r) and isinstance(r[i], str)]
        por_dia = {f: int(r[i]) for i, f in dias if i < len(r) and r[i]}
        wos[str(r[c_demand])] = {
            "modelo": str(r[c_modelo] or "").strip(),
            "partes": {p for p in crudas
                       if p.upper() not in NO_PARTE and not p.startswith("#")},
            "sin_bom": any(p.startswith("#") for p in crudas),
            "plan": sum(por_dia.values()),
            "result": None,
            "dias": por_dia,
            "sub": _sub_oven([r[i] for i in c_par if i < len(r)]),
        }
    if not wos:
        raise ValueError("La hoja de OVEN no tiene renglones de Demand ID.")
    return fecha0, wos


LECTORES = {"PPN": leer_ppn, "OVEN": leer_oven}


def hoja_reporte(wb):
    """(nombre_hoja, fuente) del reporte de LG en el workbook, o (None, None)."""
    nombre = hoja_ppn(wb)
    if nombre:
        return nombre, "PPN"
    nombre = hoja_oven(wb)
    return (nombre, "OVEN") if nombre else (None, None)


def leer_bytes(file_bytes):
    """(fecha0, wos, fuente) del Excel en memoria, o None si no es reporte de LG."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        hoja, fuente = hoja_reporte(wb)
        if not hoja:
            return None
        fecha0, wos = LECTORES[fuente](wb[hoja])
        return fecha0, wos, fuente
    finally:
        wb.close()


def _sub_oven(par):
    """{AJJ: su PCB} de un renglon de OVEN, sin fiarse del rotulo de la columna."""
    limpias = [v.strip() for v in par
               if isinstance(v, str) and v.strip().upper() not in NO_PARTE
               and not v.startswith("#")]
    if len(limpias) != 2:
        return {}
    ensambles = [p for p in limpias if p.startswith(OVEN_ENSAMBLE_PREFIJOS)]
    if len(ensambles) != 1:
        return {}  # sin ensamble, o los dos lo son: no hay relacion clara
    ensamble = ensambles[0]
    return {ensamble: next(p for p in limpias if p != ensamble)}


def _sub_del_renglon(r):
    """{ensamble: PCB} de un renglon del PPN, por las columnas pareadas."""
    sub = {}
    for a, b in SUB_PARES:
        ensamble, pcb = r[a], r[b]
        if not (isinstance(ensamble, str) and isinstance(pcb, str)):
            continue
        ensamble, pcb = ensamble.strip(), pcb.strip()
        if (ensamble.upper() in NO_PARTE or pcb.upper() in NO_PARTE
                or ensamble.startswith("#") or pcb.startswith("#")):
            continue
        sub[ensamble] = pcb
    return sub


def sub_ensambles(wos):
    """{parte ensamblada: PCB que hay que producir antes}.

    Solo el PPN trae la relacion (columnas M/N y O/P). El Prod Plan de OVEN no:
    ahi se sabe que los AJJ llevan PCB pero no que numero de parte es.
    """
    bom = {}
    for w in wos.values():
        bom.update(w.get("sub") or {})
    return bom


HORA_TURNO = dt.time(7, 30)  # cuando abre el piso de LG


def horas_parte(wos):
    """{(parte, fecha): hora mas temprana en que LG ocupa esa parte ese dia}.

    El PPN trae el 'Planned Start Time' de cada W/O (609 horas distintas en el
    del 30/07, de 07:30 a 15:15). Un W/O que arranca ese dia ocupa su material a
    esa hora; el que viene arrastrado de dias anteriores ya esta corriendo
    cuando abre el turno. Sirve para secuenciar: si una parte se ocupa a las
    10:00 hay que arrancarla desde que abre, no acomodarla por familia.
    """
    horas = {}
    for w in wos.values():
        inicio = w.get("inicio")
        if not inicio:
            continue
        for fecha in w["dias"]:
            hora = inicio.time() if fecha <= inicio.date() else HORA_TURNO
            for p in w["partes"]:
                clave = (p, fecha)
                if clave not in horas or hora < horas[clave]:
                    horas[clave] = hora
    return horas


def demanda(wos):
    """Demanda pendiente por (parte, fecha), ya neta de lo que LG construyo."""
    d = defaultdict(int)
    for w in wos.values():
        for fecha, qty in w["dias"].items():
            for p in w["partes"]:
                d[(p, fecha)] += qty
    return d


def baseline(fecha, wos):
    """Snapshot minimo que hay que guardar para poder comparar mañana."""
    return {
        wo: {
            "modelo": w["modelo"],
            "partes": w["partes"],
            "prog": w["dias"].get(fecha, 0),  # lo programado para SU propio dia
            "pend": sum(w["dias"].values()),
            "plan": w["plan"],
            "result": w["result"],
        }
        for wo, w in wos.items()
    }


def sobrante(base, wos_hoy):
    """Material entregado y no consumido el dia del snapshot `base`.

    -> (renglones [(wo, prog, real)], {parte: pzas sin consumir})
    """
    renglones, partes = [], defaultdict(int)
    for wo, b in base.items():
        prog = b["prog"]
        if not prog:
            continue
        hoy = wos_hoy.get(wo)
        if b["result"] is None:
            # OVEN: sin acumulado construido, lo que bajo el pendiente entre los
            # dos reportes es lo que LG metio (todo, si el W/O ya desaparecio).
            hecho = b["pend"] - (sum(hoy["dias"].values()) if hoy else 0)
        elif hoy:
            hecho = hoy["result"] - b["result"]
        else:
            hecho = b["plan"] - b["result"]  # desaparecio del reporte = cerrado
        # Tope: de ese W/O no pudo consumir mas de lo que tenia programado ese
        # dia; lo demas es produccion de otros dias del mismo W/O.
        real = max(0, min(hecho, prog))
        renglones.append((wo, prog, real))
        for p in b["partes"]:
            partes[p] += prog - real
    return renglones, {p: q for p, q in partes.items() if q > 0}
