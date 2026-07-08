"""Endpoints HTTP del sistema PO (Purchase Order) -> WO (Work Order).

Consumido por los modulos de Control de produccion:
  - crear-plan-produccion.js
  - plan-smd-module.js
  - control-embarque.js

Rutas (con url_prefix `/api`):
  POST   /api/work_orders               -> crear WO
  GET    /api/generar_codigo_wo         -> generar codigo WO automatico
  GET    /api/work_orders               -> listar WO (default shape {ok, work_orders})
  GET    /api/work_orders?include_import_status=1 -> array plano con flags
                                           `ya_importado`/`lot_no_existente`
                                           (consumido por ASSY/SMD)
  POST   /api/work_orders/import        -> importacion masiva a plan_main
  GET    /api/wo/listar                 -> listar WO (ruta alternativa, JOIN raw)
  GET    /api/wo/exportar               -> exportar WOs a Excel
  PUT    /api/wo/<codigo>/estado        -> actualizar estado de WO
  POST   /api/wo/actualizar-po          -> actualizar PO de una WO
  POST   /api/wo/actualizar             -> actualizar WO completa
  DELETE /api/wo/eliminar               -> eliminar WO
  GET    /api/po/listar                 -> listar POs (desde tabla embarques)
  POST   /api/po/crear                  -> crear nueva PO

Alias 2026-05-27 (redirects para clientes con cache viejo):
  GET    /api/work-orders               -> 301 a /api/work_orders?include_import_status=1
  POST   /api/work-orders/import        -> 308 a /api/work_orders/import

Migrado desde `app/api_po_wo.py` (2026-05-22). En 2026-05-27 se absorbieron
3 rutas legacy que aun vivian en routes.py (work-orders GET, work-orders/import
POST, wo/exportar GET) consolidando todo WO/PO en este blueprint.
"""

import functools
import re
from datetime import datetime

from flask import Blueprint, jsonify, request

from app.api.shared import execute_query

import logging
logger = logging.getLogger(__name__)


bp = Blueprint("api_po_wo", __name__, url_prefix="/api")


def validar_fecha(fecha_str):
    """Validar formato de fecha YYYY-MM-DD"""
    try:
        datetime.strptime(fecha_str, "%Y-%m-%d")
        return True
    except (ValueError, TypeError):
        return False


def validar_codigo_po(codigo):
    """Validar formato PO-YYMMDD-####"""
    return bool(re.match(r"^PO-\d{6}-\d{4}$", codigo))


def validar_codigo_wo(codigo):
    """Validar formato WO-YYMMDD-####"""
    return bool(re.match(r"^WO-\d{6}-\d{4}$", codigo))


def generar_codigo_wo():
    """Generar codigo WO automatico"""
    fecha_actual = datetime.now()
    fecha_str = fecha_actual.strftime("%y%m%d")

    query = "SELECT codigo_wo FROM work_orders WHERE codigo_wo LIKE %s ORDER BY codigo_wo DESC LIMIT 1"
    resultado = execute_query(query, (f"WO-{fecha_str}-%",), fetch="one")

    if resultado:
        ultimo_codigo = resultado["codigo_wo"]
        ultimo_numero = int(ultimo_codigo.split("-")[-1])
        nuevo_numero = ultimo_numero + 1
    else:
        nuevo_numero = 1

    return f"WO-{fecha_str}-{nuevo_numero:04d}"


def _ensure_wo_model_columns():
    """Asegurar que work_orders tenga columnas codigo_modelo y nombre_modelo.

    No usa IF NOT EXISTS; verifica con SHOW COLUMNS y ALTER TABLE segun sea necesario.
    """
    try:
        cols = execute_query("SHOW COLUMNS FROM work_orders", fetch="all") or []
        names = {c.get("Field") for c in cols}
        if "codigo_modelo" not in names:
            try:
                execute_query("ALTER TABLE work_orders ADD COLUMN codigo_modelo VARCHAR(64)")
            except Exception:
                pass
        if "nombre_modelo" not in names:
            try:
                execute_query("ALTER TABLE work_orders ADD COLUMN nombre_modelo VARCHAR(128)")
            except Exception:
                pass
    except Exception:
        pass


def _nombre_modelo_from_raw(part_no: str) -> str:
    """Obtener nombre (project) desde raw por part_no. Retorna '' si no hay match."""
    try:
        if not part_no:
            return ""
        row = execute_query(
            "SELECT project FROM raw WHERE TRIM(part_no)=TRIM(%s) ORDER BY id DESC LIMIT 1",
            (part_no,), fetch="one")
        return (row.get("project") if row else "") or ""
    except Exception:
        return ""


def manejo_errores(func):
    """Decorator para manejo centralizado de errores"""
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            logger.error(f" Error en {func.__name__}: {e}")
            return jsonify({
                "ok": False,
                "code": "INTERNAL_ERROR",
                "message": "Error interno del servidor"
            }), 500
    return wrapper


@bp.route("/work_orders", methods=["POST"])
@manejo_errores
def crear_wo():
    """POST /api/work_orders - Crear nueva Work Order (WO)"""
    try:
        data = request.get_json(force=True)

        codigo_wo = data.get("codigo_wo", "").strip()
        modelo = data.get("modelo", "").strip()
        codigo_po = data.get("codigo_po", "").strip()
        fecha_operacion_str = data.get("fecha_operacion", "")
        cantidad_planeada = data.get("cantidad_planeada", 0)
        modificador = data.get("usuario_creador", "Usuario no identificado").strip()

        if not codigo_wo:
            codigo_wo = generar_codigo_wo()
        elif not validar_codigo_wo(codigo_wo):
            return jsonify({
                "ok": False,
                "code": "VALIDATION_ERROR",
                "field": "codigo_wo",
                "message": "Formato de codigo WO invalido (debe ser WO-YYMMDD-####)"
            }), 400

        if not modelo:
            return jsonify({
                "ok": False,
                "code": "VALIDATION_ERROR",
                "field": "modelo",
                "message": "Modelo es requerido"
            }), 400

        fecha_operacion = None
        if fecha_operacion_str:
            if not validar_fecha(fecha_operacion_str):
                return jsonify({
                    "ok": False,
                    "code": "VALIDATION_ERROR",
                    "field": "fecha_operacion",
                    "message": "Formato de fecha invalido (debe ser YYYY-MM-DD)"
                }), 400
            fecha_operacion = fecha_operacion_str

        try:
            cantidad_planeada = int(cantidad_planeada)
            if cantidad_planeada <= 0:
                raise ValueError()
        except (ValueError, TypeError):
            return jsonify({
                "ok": False,
                "code": "VALIDATION_ERROR",
                "field": "cantidad_planeada",
                "message": "Cantidad planeada debe ser un numero entero positivo"
            }), 400

        query_check = "SELECT id FROM work_orders WHERE codigo_wo = %s"
        existe = execute_query(query_check, (codigo_wo,), fetch="one")
        if existe:
            return jsonify({
                "ok": False,
                "code": "DUPLICATE_WO",
                "message": f"Ya existe una WO con codigo {codigo_wo}"
            }), 409

        _ensure_wo_model_columns()
        codigo_modelo = modelo
        nombre_modelo = _nombre_modelo_from_raw(modelo)

        query_insert = """
            INSERT INTO work_orders (
                codigo_wo, codigo_po, modelo, codigo_modelo, nombre_modelo,
                cantidad_planeada, fecha_operacion, estado, fecha_modificacion, modificador
            ) VALUES (
                %s, %s, %s, %s, %s,
                %s, %s, 'CREADA', NOW(), %s
            )
        """
        execute_query(query_insert, (
            codigo_wo, codigo_po, modelo, codigo_modelo, nombre_modelo,
            cantidad_planeada, fecha_operacion, modificador
        ))

        logger.info(f" WO creada: {codigo_wo}")
        return jsonify({
            "ok": True,
            "codigo_wo": codigo_wo,
            "message": "Work Order creada exitosamente"
        }), 201

    except Exception as e:
        logger.error(f" Error creando WO: {e}")
        return jsonify({
            "ok": False,
            "code": "INTERNAL_ERROR",
            "message": "Error interno del servidor"
        }), 500


@bp.route("/generar_codigo_wo", methods=["GET"])
@manejo_errores
def generar_codigo_wo_endpoint():
    """GET /api/generar_codigo_wo - Generar codigo WO automatico"""
    codigo = generar_codigo_wo()
    return jsonify({"ok": True, "codigo_wo": codigo})


@bp.route("/work_orders", methods=["GET"])
@manejo_errores
def listar_wos():
    """GET /api/work_orders - Listar Work Orders con filtros opcionales.

    Por defecto solo muestra WO con estado 'CREADA' (excluye 'PLANIFICADA').

    Modos:
      ?include_import_status=1  -> shape compatible con ASSY/SMD legacy:
        devuelve array plano [{...}] con campos extra `ya_importado` y
        `lot_no_existente` (LEFT JOIN plan_main por wo_id).
        Default: estados ['CREADA','PLANIFICADA'] para no perder WOs ya
        importadas que el frontend necesita listar como deshabilitadas.
        Filtros adicionales:
          q       -> busca en codigo_wo/codigo_po/modelo/codigo_modelo
          estado  -> CSV de estados, sobreescribe defaults
          desde   -> fecha_operacion >= desde
          hasta   -> fecha_operacion <= hasta

      Sin flag -> shape estandar `{ok, work_orders: [{...}]}` con JOIN raw
        para nombre_modelo. Default estado 'CREADA' (excluye PLANIFICADA)
        si no se pasa `incluir_planificadas=true`.
    """
    include_import_status = request.args.get("include_import_status", "").lower() in ("1", "true", "yes")

    if include_import_status:
        return _listar_wos_con_import_status()

    estado = request.args.get("estado")
    codigo_wo = request.args.get("codigo_wo")
    modelo = request.args.get("modelo")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    incluir_planificadas = request.args.get("incluir_planificadas", "false").lower() == "true"

    query = (
        "SELECT w.*, COALESCE(w.nombre_modelo, ("
        " SELECT r.project FROM raw r"
        " WHERE TRIM(r.part_no) = COALESCE(NULLIF(TRIM(w.codigo_modelo), ''), TRIM(w.modelo))"
        " ORDER BY r.id DESC LIMIT 1)) AS nombre_modelo"
        " FROM work_orders w WHERE 1=1"
    )
    params = []

    if not incluir_planificadas and not estado:
        query += " AND estado = 'CREADA'"
    elif estado:
        query += " AND estado = %s"
        params.append(estado)

    if codigo_wo:
        query += " AND codigo_wo = %s"
        params.append(codigo_wo)

    if modelo:
        query += " AND modelo LIKE %s"
        params.append(f"%{modelo}%")

    if fecha_desde:
        query += " AND fecha_operacion >= %s"
        params.append(fecha_desde)

    if fecha_hasta:
        query += " AND fecha_operacion <= %s"
        params.append(fecha_hasta)

    query += " ORDER BY fecha_modificacion DESC"

    work_orders = execute_query(query, params, fetch="all") or []

    logger.info(f"Listando {len(work_orders)} WOs (estado filtrado: {estado or 'CREADA por defecto'})")
    return jsonify({"ok": True, "work_orders": work_orders})


def _listar_wos_con_import_status():
    """Variante de listar_wos que devuelve array plano + flags de import status.

    Reemplaza /api/work-orders (con guion) legacy. Consumido por
    plan-assy-workorders.js y plan-smd-module.js: necesitan saber si cada
    WO ya tiene un plan_main para deshabilitar el boton de importar.
    """
    q = request.args.get("q", "").strip()
    estados_param = request.args.get("estado", "")
    desde = request.args.get("desde", "")
    hasta = request.args.get("hasta", "")

    if estados_param:
        estados = [e.strip() for e in estados_param.split(",")]
    else:
        estados = ["CREADA", "PLANIFICADA"]

    query = (
        "SELECT id, codigo_wo, codigo_po, modelo, nombre_modelo, codigo_modelo,"
        " cantidad_planeada, fecha_operacion, estado, usuario_creacion,"
        " orden_proceso, modificador, fecha_modificacion"
        " FROM work_orders WHERE 1=1"
    )
    params = []

    if estados:
        placeholders = ",".join(["%s"] * len(estados))
        query += f" AND estado IN ({placeholders})"
        params.extend(estados)

    if q:
        query += " AND (codigo_wo LIKE %s OR codigo_po LIKE %s OR modelo LIKE %s OR codigo_modelo LIKE %s)"
        q_param = f"%{q}%"
        params.extend([q_param, q_param, q_param, q_param])

    if desde:
        query += " AND fecha_operacion >= %s"
        params.append(desde)

    if hasta:
        query += " AND fecha_operacion <= %s"
        params.append(hasta)

    query += " ORDER BY fecha_operacion ASC, codigo_modelo ASC"

    work_orders = execute_query(query, params, fetch="all") or []

    # Saber cuales WOs ya estan en plan_main
    wo_ids = [wo["id"] for wo in work_orders]
    ya_importados = {}
    if wo_ids:
        placeholders = ",".join(["%s"] * len(wo_ids))
        check_query = (
            f"SELECT DISTINCT wo_id, lot_no FROM plan_main"
            f" WHERE wo_id IN ({placeholders}) AND wo_id IS NOT NULL"
        )
        importados = execute_query(check_query, wo_ids, fetch="all") or []
        for imp in importados:
            ya_importados[imp["wo_id"]] = imp["lot_no"]

    resultado = []
    for wo in work_orders:
        wo_id = wo["id"]
        resultado.append({
            "id": wo_id,
            "codigo_wo": wo["codigo_wo"],
            "codigo_po": wo["codigo_po"] or "",
            "modelo": wo["modelo"] or "",
            "nombre_modelo": wo["nombre_modelo"] or "",
            "codigo_modelo": wo["codigo_modelo"] or "",
            "cantidad_planeada": wo["cantidad_planeada"] or 0,
            "fecha_operacion": wo["fecha_operacion"].strftime("%Y-%m-%d") if wo["fecha_operacion"] else "",
            "estado": wo["estado"] or "",
            "usuario_creacion": wo["usuario_creacion"] or "",
            "orden_proceso": wo["orden_proceso"] or "",
            "modificador": wo["modificador"] or "",
            "fecha_modificacion": wo["fecha_modificacion"].strftime("%Y-%m-%d %H:%M:%S") if wo["fecha_modificacion"] else "",
            "ya_importado": wo_id in ya_importados,
            "lot_no_existente": ya_importados.get(wo_id, None),
        })

    return jsonify(resultado)


@bp.route("/wo/listar", methods=["GET"])
@manejo_errores
def listar_wos_alternativo():
    """GET /api/wo/listar - Ruta alternativa de listar WOs para compatibilidad con el frontend.

    Por defecto solo muestra WO con estado 'CREADA' (excluye 'PLANIFICADA').
    """
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    estado = request.args.get("estado")
    codigo_wo = request.args.get("codigo_wo")
    modelo = request.args.get("modelo")
    incluir_planificadas = request.args.get("incluir_planificadas", "false").lower() == "true"

    query = (
        "SELECT w.*, "
        "(SELECT r.project FROM raw r "
        " WHERE TRIM(r.part_no) = COALESCE(NULLIF(TRIM(w.codigo_modelo), ''), TRIM(w.modelo)) "
        " ORDER BY r.id DESC LIMIT 1) AS nombre_modelo_raw "
        "FROM work_orders w WHERE 1=1"
    )
    params = []

    if fecha_desde:
        query += " AND fecha_operacion >= %s"
        params.append(fecha_desde)

    if fecha_hasta:
        query += " AND fecha_operacion <= %s"
        params.append(fecha_hasta)

    if not incluir_planificadas and not estado:
        query += " AND estado = 'CREADA'"
    elif estado:
        query += " AND estado = %s"
        params.append(estado)

    if codigo_wo:
        query += " AND codigo_wo = %s"
        params.append(codigo_wo)

    if modelo:
        query += " AND modelo LIKE %s"
        params.append(f"%{modelo}%")

    query += " ORDER BY fecha_operacion DESC, fecha_modificacion DESC"

    try:
        work_orders = execute_query(query, params, fetch="all") or []

        for wo in work_orders:
            if wo.get("fecha_operacion") and hasattr(wo["fecha_operacion"], "isoformat"):
                wo["fecha_operacion"] = wo["fecha_operacion"].isoformat()
            if wo.get("fecha_modificacion") and hasattr(wo["fecha_modificacion"], "isoformat"):
                wo["fecha_modificacion"] = wo["fecha_modificacion"].isoformat()

            nombre_raw = wo.get("nombre_modelo_raw")
            if nombre_raw and nombre_raw.strip():
                wo["nombre_modelo"] = nombre_raw
            elif not wo.get("nombre_modelo"):
                wo["nombre_modelo"] = ""

        logger.info(f"Listando {len(work_orders)} WOs (ruta alternativa - estado filtrado: {estado or 'CREADA por defecto'})")
        return jsonify({"success": True, "data": work_orders})

    except Exception as e:
        logger.error(f" Error listando WOs: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/wo/<codigo>/estado", methods=["PUT"])
@manejo_errores
def actualizar_estado_wo(codigo):
    """PUT /api/wo/{codigo}/estado - Actualizar estado de una Work Order"""
    try:
        data = request.get_json(force=True)
        nuevo_estado = data.get("estado", "").strip().upper()
        modificador = data.get("modificador", "Sistema").strip()

        estados_validos = ["CREADA", "PLANIFICADA", "EN_PRODUCCION", "CERRADA"]
        if nuevo_estado not in estados_validos:
            return jsonify({
                "ok": False,
                "code": "VALIDATION_ERROR",
                "message": f"Estado invalido. Estados validos: {', '.join(estados_validos)}"
            }), 400

        query_check = "SELECT id, estado FROM work_orders WHERE codigo_wo = %s"
        wo_actual = execute_query(query_check, (codigo,), fetch="one")

        if not wo_actual:
            return jsonify({
                "ok": False,
                "code": "NOT_FOUND",
                "message": f"No se encontro la WO con codigo {codigo}"
            }), 404

        query_update = """
            UPDATE work_orders
            SET estado = %s, modificador = %s, fecha_modificacion = NOW()
            WHERE codigo_wo = %s
        """
        execute_query(query_update, (nuevo_estado, modificador, codigo))

        logger.info(f" Estado de WO {codigo} actualizado de {wo_actual['estado']} a {nuevo_estado}")
        return jsonify({
            "ok": True,
            "message": f"Estado de WO {codigo} actualizado a {nuevo_estado}",
            "estado_anterior": wo_actual["estado"],
            "estado_nuevo": nuevo_estado
        })

    except Exception as e:
        logger.error(f" Error actualizando estado de WO {codigo}: {e}")
        return jsonify({
            "ok": False,
            "code": "INTERNAL_ERROR",
            "message": "Error interno del servidor"
        }), 500


@bp.route("/wo/actualizar-po", methods=["POST"])
@manejo_errores
def actualizar_po_wo():
    """POST /api/wo/actualizar-po - Actualizar codigo PO de una Work Order"""
    try:
        data = request.get_json(force=True)
        codigo_wo = data.get("codigo_wo", "").strip()
        nuevo_codigo_po = data.get("codigo_po", "").strip() or "SIN-PO"

        if not codigo_wo:
            return jsonify({"success": False, "error": "Codigo WO es requerido"}), 400

        query_check = "SELECT id, codigo_po FROM work_orders WHERE codigo_wo = %s"
        wo_actual = execute_query(query_check, (codigo_wo,), fetch="one")

        if not wo_actual:
            return jsonify({
                "success": False,
                "error": f"No se encontro la WO con codigo {codigo_wo}"
            }), 404

        query_update = """
            UPDATE work_orders
            SET codigo_po = %s, fecha_modificacion = NOW()
            WHERE codigo_wo = %s
        """
        execute_query(query_update, (nuevo_codigo_po, codigo_wo))

        logger.info(f" PO actualizado: WO {codigo_wo} -> PO {nuevo_codigo_po}")
        return jsonify({
            "success": True,
            "message": "Codigo PO actualizado exitosamente",
            "codigo_wo": codigo_wo,
            "codigo_po_anterior": wo_actual["codigo_po"],
            "codigo_po_nuevo": nuevo_codigo_po
        })

    except Exception as e:
        logger.error(f" Error actualizando PO de WO: {e}")
        return jsonify({"success": False, "error": "Error interno del servidor"}), 500


@bp.route("/wo/actualizar", methods=["POST"])
@manejo_errores
def actualizar_wo_completa():
    """POST /api/wo/actualizar - Actualizar Work Order completa (modelo, cantidad, PO)"""
    try:
        data = request.get_json(force=True)
        codigo_wo = data.get("codigo_wo", "").strip()
        modelo = data.get("modelo", "").strip()
        cantidad_planeada = data.get("cantidad_planeada", 0)
        codigo_po = data.get("codigo_po", "").strip() or "SIN-PO"

        if not codigo_wo:
            return jsonify({"success": False, "error": "Codigo WO es requerido"}), 400

        if not modelo:
            return jsonify({"success": False, "error": "Modelo es requerido"}), 400

        if not cantidad_planeada or cantidad_planeada < 1:
            return jsonify({"success": False, "error": "Cantidad planeada debe ser mayor a 0"}), 400

        query_check = "SELECT id, modelo, cantidad_planeada, codigo_po FROM work_orders WHERE codigo_wo = %s"
        wo_actual = execute_query(query_check, (codigo_wo,), fetch="one")

        if not wo_actual:
            return jsonify({
                "success": False,
                "error": f"No se encontro la WO con codigo {codigo_wo}"
            }), 404

        _ensure_wo_model_columns()
        codigo_modelo = modelo
        nombre_modelo = _nombre_modelo_from_raw(modelo)

        query_update = """
            UPDATE work_orders
            SET modelo = %s,
                codigo_modelo = %s,
                nombre_modelo = %s,
                cantidad_planeada = %s,
                codigo_po = %s,
                fecha_modificacion = NOW()
            WHERE codigo_wo = %s
        """
        execute_query(query_update, (modelo, codigo_modelo, nombre_modelo, cantidad_planeada, codigo_po, codigo_wo))

        logger.info(f" WO actualizada: {codigo_wo} -> Modelo: {modelo}, Cantidad: {cantidad_planeada}, PO: {codigo_po}")
        return jsonify({
            "success": True,
            "message": f"WO {codigo_wo} actualizada exitosamente"
        })

    except Exception as e:
        logger.error(f" Error actualizando WO: {e}")
        return jsonify({"success": False, "error": "Error interno del servidor"}), 500


@bp.route("/wo/eliminar", methods=["DELETE"])
@manejo_errores
def eliminar_wo():
    """DELETE /api/wo/eliminar - Eliminar Work Order"""
    try:
        data = request.get_json(force=True)
        codigo_wo = data.get("codigo_wo", "").strip()

        if not codigo_wo:
            return jsonify({"success": False, "error": "Codigo WO es requerido"}), 400

        query_check = "SELECT id FROM work_orders WHERE codigo_wo = %s"
        wo_actual = execute_query(query_check, (codigo_wo,), fetch="one")

        if not wo_actual:
            return jsonify({
                "success": False,
                "error": f"No se encontro la WO con codigo {codigo_wo}"
            }), 404

        query_delete = "DELETE FROM work_orders WHERE codigo_wo = %s"
        execute_query(query_delete, (codigo_wo,))

        logger.info(f" WO eliminada: {codigo_wo}")
        return jsonify({
            "success": True,
            "message": f"WO {codigo_wo} eliminada exitosamente"
        })

    except Exception as e:
        logger.error(f" Error eliminando WO: {e}")
        return jsonify({"success": False, "error": "Error interno del servidor"}), 500


@bp.route("/po/listar", methods=["GET"])
@manejo_errores
def listar_pos():
    """GET /api/po/listar - Listar Purchase Orders desde tabla embarques"""
    try:
        estado = request.args.get("estado")
        fecha_desde = request.args.get("fecha_desde")
        fecha_hasta = request.args.get("fecha_hasta")

        query = """
            SELECT
                e.codigo_po,
                e.nombre_po,
                e.fecha_registro,
                e.modelo,
                e.cliente,
                e.proveedor,
                e.total_cantidad_entregada,
                e.cantidad_entregada,
                e.estado,
                e.codigo_entrega,
                e.fecha_entrega,
                e.usuario_creacion,
                e.modificado,
                (SELECT r.project FROM raw r
                   WHERE TRIM(r.part_no) = TRIM(e.modelo)
                   ORDER BY r.id DESC LIMIT 1) AS nombre_modelo
            FROM embarques e
            WHERE 1=1
        """

        params = []

        if estado:
            query += " AND estado = %s"
            params.append(estado)

        if fecha_desde:
            query += " AND fecha_registro >= %s"
            params.append(fecha_desde)

        if fecha_hasta:
            query += " AND fecha_registro <= %s"
            params.append(fecha_hasta)

        query += " ORDER BY modificado DESC, fecha_registro DESC"

        result = execute_query(query, params, fetch="all")

        if result is not None:
            pos = []
            for row in result:
                po = {
                    "codigo_po": row["codigo_po"],
                    "nombre_po": row["nombre_po"],
                    "fecha_registro": row["fecha_registro"].strftime("%Y-%m-%d") if row["fecha_registro"] else None,
                    "modelo": row["modelo"],
                    "nombre_modelo": row.get("nombre_modelo"),
                    "cliente": row["cliente"],
                    "proveedor": row["proveedor"],
                    "total_cantidad_entregada": row["total_cantidad_entregada"],
                    "cantidad_entregada": row["cantidad_entregada"],
                    "estado": row["estado"],
                    "codigo_entrega": row["codigo_entrega"],
                    "fecha_entrega": row["fecha_entrega"].strftime("%Y-%m-%d") if row["fecha_entrega"] else None,
                    "usuario_creacion": row["usuario_creacion"],
                    "modificado": row["modificado"].strftime("%Y-%m-%d %H:%M:%S") if row["modificado"] else None
                }
                pos.append(po)

            logger.info(f" {len(pos)} POs listadas exitosamente")
            return jsonify({
                "success": True,
                "data": pos,
                "total": len(pos)
            })
        else:
            logger.error(" Error en consulta de POs: No se obtuvieron resultados")
            return jsonify({"success": False, "error": "No se pudieron obtener las POs"}), 500

    except Exception as e:
        logger.error(f" Error listando POs: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/po/crear", methods=["POST"])
def crear_po():
    """POST /api/po/crear - Crear nueva Purchase Order"""
    try:
        data = request.get_json()

        if not data.get("nombre_po"):
            return jsonify({"success": False, "error": "El nombre de PO es obligatorio"}), 400

        if not data.get("fecha_registro"):
            return jsonify({"success": False, "error": "La fecha de registro es obligatoria"}), 400

        if not data.get("modelo"):
            return jsonify({"success": False, "error": "El modelo es obligatorio"}), 400

        if not data.get("cliente"):
            return jsonify({"success": False, "error": "El cliente es obligatorio"}), 400

        fecha_actual = datetime.now()
        fecha_str = fecha_actual.strftime("%y%m%d")

        query_ultimo = "SELECT codigo_po FROM embarques WHERE codigo_po LIKE %s ORDER BY codigo_po DESC LIMIT 1"
        resultado = execute_query(query_ultimo, (f"PO-{fecha_str}-%",), fetch="one")

        if resultado and resultado.get("codigo_po"):
            ultimo_codigo = resultado["codigo_po"]
            ultimo_numero = int(ultimo_codigo.split("-")[-1])
            nuevo_numero = ultimo_numero + 1
        else:
            nuevo_numero = 1

        codigo_po = f"PO-{fecha_str}-{nuevo_numero:04d}"

        insert_data = {
            "codigo_po": codigo_po,
            "nombre_po": data["nombre_po"],
            "fecha_registro": data["fecha_registro"],
            "modelo": data["modelo"],
            "cliente": data["cliente"],
            "proveedor": data.get("proveedor", ""),
            "total_cantidad_entregada": data.get("total_cantidad_entregada", 0),
            "fecha_entrega": data.get("fecha_entrega") if data.get("fecha_entrega") else None,
            "cantidad_entregada": data.get("cantidad_entregada", 0),
            "codigo_entrega": data.get("codigo_entrega", ""),
            "estado": data.get("estado", "PLAN"),
            "usuario_creacion": "Sistema"
        }

        query_insert = """
        INSERT INTO embarques (
            codigo_po, nombre_po, fecha_registro, modelo, cliente, proveedor,
            total_cantidad_entregada, fecha_entrega, cantidad_entregada,
            codigo_entrega, estado, usuario_creacion
        ) VALUES (
            %(codigo_po)s, %(nombre_po)s, %(fecha_registro)s, %(modelo)s,
            %(cliente)s, %(proveedor)s, %(total_cantidad_entregada)s,
            %(fecha_entrega)s, %(cantidad_entregada)s, %(codigo_entrega)s,
            %(estado)s, %(usuario_creacion)s
        )
        """

        affected_rows = execute_query(query_insert, insert_data)

        if affected_rows and affected_rows > 0:
            logger.info(f" PO creada exitosamente: {codigo_po}")
            return jsonify({
                "success": True,
                "message": f"PO {codigo_po} creada exitosamente",
                "data": {"codigo_po": codigo_po, **insert_data}
            })
        else:
            logger.error(" Error creando PO: No se insertaron filas")
            return jsonify({"success": False, "error": "No se pudo insertar la PO"}), 500

    except Exception as e:
        logger.error(f" Error creando PO: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


# ---------------------------------------------------------------------------
# Importacion masiva de WOs a plan_main
# Migrado desde routes.py (api_work_orders_import) 2026-05-27.
# ---------------------------------------------------------------------------


@bp.route("/work_orders/import", methods=["POST"])
def importar_wos_a_plan_main():
    """POST /api/work_orders/import - Crear plan_main a partir de WO seleccionadas.

    Consumido por plan-assy-workorders.js. Recibe:
      {wo_ids: [int,...], import_date: "YYYY-MM-DD"}

    Devuelve: {success, imported, plans:[{lot_no, wo_code}], errors:[...]}.
    """
    from app.api.shared.plan_lot_no import _fp_generate_lot_no, _fp_safe_date

    try:
        data = request.get_json() or {}
        wo_ids = data.get("wo_ids", [])
        import_date = data.get("import_date", None)

        if not wo_ids:
            return jsonify({"error": "No se seleccionaron work orders"}), 400

        fecha_importacion = None
        if import_date:
            try:
                fecha_importacion = datetime.strptime(import_date, "%Y-%m-%d").date()
            except Exception as e:
                return jsonify({"error": f"Fecha de importacion invalida: {str(e)}"}), 400

        imported = 0
        plans = []
        errors = []

        # Recolectar WOs validas primero para poder ordenar por linea
        wo_list = []
        for wo_id in wo_ids:
            row = execute_query(
                "SELECT * FROM work_orders WHERE id = %s", (wo_id,), fetch="one"
            )
            if not row:
                errors.append(f"WO id {wo_id} no encontrado")
                continue
            wo = row
            existing = execute_query(
                "SELECT lot_no, status FROM plan_main WHERE wo_id = %s OR wo_code = %s",
                (wo_id, wo.get("codigo_wo")),
                fetch="one",
            )
            if existing:
                lot_existente = existing.get("lot_no") if isinstance(existing, dict) else existing[0]
                errors.append(f"WO {wo.get('codigo_wo')} ya fue importada como LOT: {lot_existente}")
                continue
            wo_list.append((wo_id, wo))

        # Ordenar por linea para que los lotes sean consecutivos por linea
        wo_list.sort(key=lambda x: x[1].get("linea") or "ZZZ")

        for wo_id, wo in wo_list:
            # En work_orders: 'modelo' es el part_no (ej: EBR42005002)
            # 'linea' viene directo de work_orders, no se busca en raw
            part_no = wo.get("modelo") or wo.get("codigo_modelo") or ""
            line = wo.get("linea") or "MAIN_LINE"

            # Buscar CT, UPH, MODEL, PROJECT en raw (no la linea)
            # Prioriza match exacto de part_no: el LIKE tambien trae variantes
            # con sufijo (EBR43713702_I / _S) y sin esto ganaba la mas nueva.
            raw_data_query = """
                SELECT part_no, model, project, c_t as ct, uph
                FROM raw
                WHERE model = %s OR model = %s OR part_no = %s OR part_no LIKE %s
                ORDER BY (part_no IN (%s, %s)) DESC, id DESC
                LIMIT 1
            """
            raw_params = (
                wo.get("modelo"),
                wo.get("codigo_modelo"),
                wo.get("codigo_modelo"),
                f"%{wo.get('modelo')}%",
                wo.get("modelo"),
                wo.get("codigo_modelo"),
            )
            raw_data = execute_query(raw_data_query, raw_params, fetch="one")

            if raw_data:
                # El WO ya trae el part_no correcto; raw solo lo completa si falta
                part_no = part_no or raw_data.get("part_no") or ""
                model_code = raw_data.get("model") or wo.get("modelo") or ""
                project = raw_data.get("project") or wo.get("nombre_modelo") or ""
                try:
                    ct = float(raw_data.get("ct") or 0)
                except Exception:
                    ct = 0.0
                try:
                    uph_raw = raw_data.get("uph")
                    if uph_raw and str(uph_raw).strip().isdigit():
                        uph = int(str(uph_raw).strip())
                    else:
                        uph = 0
                except Exception:
                    uph = 0
            else:
                part_no = wo.get("codigo_modelo") or wo.get("modelo") or ""
                model_code = wo.get("modelo") or ""
                project = wo.get("nombre_modelo") or ""
                ct = 0.0
                uph = 0

            # Fecha: import_date del request > fecha_operacion del WO > hoy
            if fecha_importacion:
                fecha_dt = fecha_importacion
            else:
                fecha_op = wo.get("fecha_operacion")
                try:
                    if isinstance(fecha_op, str):
                        fecha_dt = _fp_safe_date(fecha_op) or datetime.utcnow().date()
                    else:
                        fecha_dt = fecha_op.date() if hasattr(fecha_op, "date") else datetime.utcnow().date()
                except Exception:
                    fecha_dt = datetime.utcnow().date()

            lot_no = _fp_generate_lot_no(datetime.combine(fecha_dt, datetime.min.time()))

            insert_sql = (
                "INSERT INTO plan_main (lot_no, wo_id, wo_code, po_code, working_date,"
                " line, model_code, part_no, project, process, plan_count, ct, uph,"
                " routing, status, created_at)"
                " VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'PLAN',NOW())"
            )
            params = (
                lot_no,
                wo_id,
                wo.get("codigo_wo"),
                wo.get("codigo_po"),
                fecha_dt,
                line,
                model_code,
                part_no,
                project,
                "MAIN",
                int(wo.get("cantidad_planeada") or 0),
                ct,
                uph,
                1,  # routing por defecto: DIA
            )
            execute_query(insert_sql, params)
            imported += 1
            plans.append({"lot_no": lot_no, "wo_code": wo.get("codigo_wo")})

        return jsonify(
            {"success": True, "imported": imported, "plans": plans, "errors": errors}
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# Exportar WOs a Excel
# Migrado desde routes.py (exportar_wos_excel) 2026-05-27.
# ---------------------------------------------------------------------------


@bp.route("/wo/exportar", methods=["GET"])
def exportar_wos_excel():
    """GET /api/wo/exportar - Exportar WOs a archivo Excel."""
    try:
        import io

        from flask import send_file
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        # po_wo_models.listar_wos vive en app/api/control_produccion/po_wo_models.py
        # (modelo de datos). No se confunde con
        # app.api.control_produccion.po_wo.listar_wos (endpoint Flask).
        from app.api.control_produccion.po_wo_models import listar_wos as listar_wos_modelo

        fecha_desde = request.args.get("fecha_desde")
        fecha_hasta = request.args.get("fecha_hasta")

        wos = listar_wos_modelo(fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)

        wb = Workbook()
        ws = wb.active
        ws.title = "Work Orders"

        headers = [
            "Codigo WO", "Estado", "Fecha Operacion", "Linea",
            "Codigo Modelo", "Nombre Modelo", "Cantidad Planeada", "Codigo PO",
            "Registrado", "Modificador", "Fecha Modificacion", "Fecha Creacion",
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="172A46", end_color="172A46", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        for row_num, wo in enumerate(wos, 2):
            data = [
                wo.get("codigo_wo", ""),
                wo.get("estado", "CREADA"),
                wo.get("fecha_operacion", ""),
                wo.get("linea", "SMT-1"),
                wo.get("codigo_modelo", "") or wo.get("modelo", ""),
                wo.get("nombre_modelo", ""),
                wo.get("cantidad_planeada", 0),
                wo.get("codigo_po", "SIN-PO"),
                "Si" if wo.get("registrado") else "No",
                wo.get("modificador", ""),
                wo.get("fecha_modificacion", ""),
                wo.get("fecha_creacion", ""),
            ]
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        column_widths = {
            "A": 15, "B": 12, "C": 15, "D": 8, "E": 15, "F": 20,
            "G": 12, "H": 12, "I": 10, "J": 15, "K": 18, "L": 18,
        }
        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"work_orders_{timestamp}.xlsx"

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        logger.error(f"Error exportando WOs: {e}")
        return jsonify(
            {"success": False, "error": f"Error exportando WOs: {str(e)}"}
        ), 500


# ---------------------------------------------------------------------------
# Alias 301 legacy: URLs con guion -> URLs canonicas con underscore
# Mantener mientras haya browsers con frontend cacheado apuntando a las
# URLs viejas. Borrar tras 1-2 sprints cuando los logs no muestren hits.
# ---------------------------------------------------------------------------


@bp.route("/work-orders", methods=["GET"])
def alias_legacy_work_orders_get():
    """Alias 301 -> /api/work_orders?include_import_status=1 (preserva query string)."""
    from flask import redirect, request as _req
    qs = _req.query_string.decode("utf-8")
    target = "/api/work_orders?include_import_status=1"
    if qs:
        target += "&" + qs
    return redirect(target, code=301)


@bp.route("/work-orders/import", methods=["POST"])
def alias_legacy_work_orders_import():
    """Alias 308 -> /api/work_orders/import (308 preserva metodo+body en POST)."""
    from flask import redirect
    # 308 Permanent Redirect: REQUIERE que el cliente reenvie el POST sin
    # cambiarlo a GET. 301 en POST tiene comportamiento variable entre clientes.
    return redirect("/api/work_orders/import", code=308)
