"""Endpoints HTTP del modulo Part Planning LG (import del plan diario de LG).

Etapa 1 (2026-07-12): recibir, validar y guardar el plan diario de LG desde
el Excel oficial (hoja "LG"). No calcula LGEMM/ISEMM/pendiente (Etapa 2).

Rutas render:
  GET  /part-planning-ajax                        -> render fragment AJAX

Rutas API:
  POST /api/part-planning/import/preview          -> analizar Excel sin guardar
  POST /api/part-planning/import/confirm          -> guardar plan (transaccional)
  POST /api/part-planning/inventory/import        -> inventario semanal (hoja Part 10)
  POST /api/part-planning/schedule                -> capturar/editar renglon S
  GET  /api/part-planning/plan                    -> plan pivotado P/S/I (parte x fechas)
  GET  /api/part-planning/imports                 -> historial de importaciones

Etapa 2 (2026-07-13): renglones P/S/I. P = lg_plan_daily; S = lg_schedule_daily
(importado de Part 10 y/o capturado en MES); I = proyeccion calculada con la
formula del Excel: I(t) = I(t-1) - P(t) + S(t), I0 = LGEMM+ISEMM+SVC+DIF+
PENDIENTE+REWORK (escalares semanales en lg_part_inventory, hoja "Part 10").

Formato del Excel (verificado contra el archivo real):
  Hoja "LG"; fila 3 = fechas (datetime completos; fallback strings "07-Jul");
  columna B desde fila 4 = numeros de parte; celdas = cantidades o None.

Flujo preview -> confirm: el navegador reenvia el mismo archivo en el confirm
junto con el SHA-256 que devolvio el preview; si el hash no coincide se
rechaza (patron Lista de compras: sin token ni temporales en servidor).
"""

import hashlib
import io
import json
import logging
import math
import os
import re
import uuid
from datetime import date, datetime, timedelta

import openpyxl
from flask import Blueprint, jsonify, render_template, request, session

from app.db_mysql import execute_query
from app.config_mysql import get_pooled_connection
from app.api.pda.shipping_material import get_dict_cursor
from app.api.shared import login_requerido, requiere_permiso_dropdown

logger = logging.getLogger(__name__)

bp = Blueprint("control_produccion_part_planning", __name__)

PP_PERMISO_PAGINA = "LISTA_CONTROLDEPRODUCCION"
PP_PERMISO_SECCION = "Control de plan de produccion"
PP_PERMISO_BOTON = "Part Planning LG"
# Modulo hermano "Proyeccion" (renglones P/S/I, inventario semanal y schedule)
PROY_PERMISO_BOTON = "Proyeccion"
# Modulo hermano "Plan Proyectado" (generador de lotes tipo hoja LOTE N)
PPY_PERMISO_BOTON = "Plan Proyectado"
PPY_HORAS_TURNO = 9.0  # horas productivas sin tiempo extra (plan-assy-helpers.js)
# Capacidad maxima: 5 turnos/equipos de 9 h.
# NO estan atados a una linea fija: planning arma los bloques cada dia con las
# lineas que convenga. En la hoja PLAN se ve que cada bloque corre varias
# lineas en secuencia, con horarios encadenados (09:45 -> 11:05 -> 14:25), y
# que la composicion cambia de un dia a otro:
#     15/07: M1+D1 | M2 | M3+M4 | D3+D1+D2
#     16/07: M2+M4 | M3 | D1    | D2+D3
# El plan del 17/07 confirmo que Planning puede abrir un quinto bloque:
#     17/07: M1 | M2 | M3+M4 | D1 | D2+D3
# No siempre se usan los cinco. El repartidor empaca primero en bloques ya
# abiertos y abre otro solamente cuando la linea completa ya no cabe. Por eso
# los planes del 15 y 16 siguen usando cuatro, mientras el del 17 puede usar
# cinco. La linea del lote la fija assy_line; el bloque solo aporta capacidad.
# Tratar cada linea como si tuviera 9 h propias daba 63 h de capacidad y hacia
# que el generador propusiera piezas que no cabian. El limite real observado es
# 45 h, pero el acomodo minimiza los bloques abiertos.
# El harness va aparte (planning excluye H1 de las lineas activas).
PPY_BLOQUES = 5
# Que lineas pueden compartir un bloque. Los bloques son equipos: 3 atienden
# las main (M1-M4) y 1 el display (D1-D3), pero D1 y D2 si se pueden mezclar
# con main. D3 no: solo va con D1 o D2.
# Se verifica en los planes reales: 15/07 -> M1+D1 y D3+D1+D2 ; 16/07 -> M2+M4
# y D2+D3. En ninguno aparece D3 junto a una M.
PPY_LINEAS_INCOMPATIBLES = {"D3": ("M1", "M2", "M3", "M4")}
# Preferencias observadas cuando dos acomodos dejan exactamente el mismo
# hueco. No fuerzan un bloque ni desplazan una opcion con mejor ajuste.
PPY_LINEAS_AFINES = {
    "M3": ("M4",), "M4": ("M3",),
    "D2": ("D3",), "D3": ("D2",),
}
PPY_FAMILIA_LEN = 9  # ponytail: familia = prefijo del part_no (EBR807574xx); catalogo si hay excepciones
PPY_MARGEN = 1.10  # siempre se planea 10% mas del faltante
# Remanente ("Remain" en la hoja PLAN). NO es un disparador, es el OBJETIVO de
# lo que se planea: cuando una parte se va a producir, el lote se calcula para
# que el inventario termine en 60 y no en cero. Ese colchon es el que absorbe
# el scrap y deja el remanente por encima de 50.
#     qty = (60 - I_proyectado) x 1.10 -> caja cerrada
PPY_REMAIN_IDEAL = 60  # objetivo del lote planeado
PPY_REMAIN_MIN = 50    # piso que protege el colchon (scrap, etc.)
# OJO: una parte NO se planea por tener el remanente bajo. Si no hay consumo no
# se produce: un I de 44 son los 60 del plan anterior menos 16 de scrap, no un
# faltante. Solo dispara el faltante real (I < 0). Medir contra 60 metia al
# plan decenas de partes sin demanda solo para "llegar a 60".
PPY_TURNOS = ("DIA", "TIEMPO EXTRA", "NOCHE")
# Dias de CONSUMO hacia adelante que mira el plan del dia (ver
# _ppy_sumar_dias_consumo: LG consume L-S, el domingo no).
#
# Solo se produce lo que FALTA: si el inventario alcanza, no se planea nada.
# Los grupos deben llenar sus 9 h (planning: "las lineas tienen que producir
# sus 9 hrs"), y para llenarlas hay que adelantar faltantes de los proximos
# dias, no solo los de manana.
#
# Dias de PRODUCCION (L-V) que se adelanta un lote a su faltante. Regla de
# planning: "si el 5505 es para el lunes no lo puedes poner para el viernes, de
# preferencia que sea para el jueves, 2 dias anticipado".
#
# El calendario importa: desde el jueves, 2 dias de PRODUCCION son viernes y
# lunes (el fin de semana no cuenta), asi que el faltante del lunes se produce
# el jueves. Contado en dias de consumo daria viernes, que ya es tarde. Por eso
# se usa _ppy_sumar_dias_produccion y no _ppy_sumar_dias_consumo.
#
# Verificado contra el plan real del 16/07: EBR33105505 falta el lunes 20 y
# planning lo puso el jueves 16.
PPY_ANTICIPACION_DIAS = 2
# D1 es la excepcion: adelanta mucho mas que el resto. Confirmado con planning
# ("lo unico que adelanta inclusive 7 dias antes es la linea de D1"); medido en
# el historico, su anticipacion tiene mediana 4 y p90 7 dias de consumo, que
# son ~5 de produccion (vs los 2 de main).
# ponytail: dict de excepciones; si crece, va a lg_pp_config.
PPY_ANTICIPACION_POR_LINEA = {"D1": 5}
# Hasta donde hay que proyectar para que ninguna linea se quede sin datos.
PPY_ANTICIPACION_MAX = max(
    [PPY_ANTICIPACION_DIAS] + list(PPY_ANTICIPACION_POR_LINEA.values())
)
PPY_LINEAS_DEFAULT = "M1,M2,M3"
# Partes que LG pide pero que NO se producen aqui. Se configuran en
# lg_pp_config.partes_excluidas (CSV) porque no hay nada en los datos que las
# distinga: ACQ30500849 se ve igual que sus vecinas (misma linea, CT y UPH).
PPY_EXCLUIDAS_DEFAULT = "ACQ30500849"  # la vende ILSAN Corea aparte
PPY_PACK_DEFAULT = 20  # std pack asumido en la propuesta de schedule para partes sin registro
PPY_PROPOSAL_ENGINE_VERSION = "mrp-capacity-v6"

PP_SHEET_NAME = "LG"
# Campos de inventario por parte (editables en Proyeccion; suman al I inicial)
PP_INV_FIELDS = ("lgemm", "isemm", "svc", "dif", "pendiente", "rework", "smt", "imd")
PP_MAX_FILE_BYTES = 20 * 1024 * 1024
PP_BATCH_SIZE = 1000
PP_EXTENSIONES = (".xlsx", ".xlsm")
PP_MODOS = ("upsert", "only_positive", "only_new")
PP_MAX_RANGO_DIAS = 120


# =============================
# DDL (llamado desde app/startup_init.py)
# =============================

def init_part_planning_tables():
    """Crea las tablas del modulo Part Planning LG (idempotente)."""
    execute_query(
        """
        CREATE TABLE IF NOT EXISTS lg_plan_daily (
            id BIGINT NOT NULL AUTO_INCREMENT,
            part_no VARCHAR(100) NOT NULL,
            plan_date DATE NOT NULL,
            plan_qty INT NOT NULL DEFAULT 0,
            import_id BIGINT NULL,
            created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                ON UPDATE CURRENT_TIMESTAMP,
            PRIMARY KEY (id),
            UNIQUE KEY uk_lgpd_part_date (part_no, plan_date),
            KEY idx_lgpd_plan_date (plan_date),
            KEY idx_lgpd_part_no (part_no)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )
    execute_query(
        """
        CREATE TABLE IF NOT EXISTS lg_plan_imports (
            id BIGINT NOT NULL AUTO_INCREMENT,
            original_filename VARCHAR(255) NULL,
            sheet_name VARCHAR(50) NOT NULL DEFAULT 'LG',
            plan_year INT NULL,
            date_from DATE NULL,
            date_to DATE NULL,
            parts_count INT NOT NULL DEFAULT 0,
            dates_count INT NOT NULL DEFAULT 0,
            records_count INT NOT NULL DEFAULT 0,
            zero_records_count INT NOT NULL DEFAULT 0,
            warning_count INT NOT NULL DEFAULT 0,
            import_mode VARCHAR(20) NOT NULL DEFAULT 'upsert',
            file_sha256 VARCHAR(64) NULL,
            imported_by VARCHAR(255) NOT NULL DEFAULT 'SISTEMA',
            imported_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            status VARCHAR(20) NOT NULL DEFAULT 'COMPLETADO',
            PRIMARY KEY (id),
            KEY idx_lgpi_imported_at (imported_at)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )
    execute_query(
        """
        CREATE TABLE IF NOT EXISTS lg_plan_import_changes (
            id BIGINT NOT NULL AUTO_INCREMENT,
            import_id BIGINT NOT NULL,
            part_no VARCHAR(100) NOT NULL,
            plan_date DATE NOT NULL,
            old_qty INT NULL,
            new_qty INT NOT NULL,
            change_type VARCHAR(10) NOT NULL,
            created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (id),
            KEY idx_lgpc_import (import_id),
            KEY idx_lgpc_part_date (part_no, plan_date)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )
    # Etapa 2: inventario semanal por parte (hoja "Part 10": LGEMM/ISEMM/...)
    execute_query(
        """
        CREATE TABLE IF NOT EXISTS lg_part_inventory (
            id BIGINT NOT NULL AUTO_INCREMENT,
            part_no VARCHAR(100) NOT NULL,
            board VARCHAR(100) NULL,
            line VARCHAR(20) NULL,
            lgemm INT NOT NULL DEFAULT 0,
            isemm INT NOT NULL DEFAULT 0,
            svc INT NOT NULL DEFAULT 0,
            dif INT NOT NULL DEFAULT 0,
            pendiente INT NOT NULL DEFAULT 0,
            rework INT NOT NULL DEFAULT 0,
            smt INT NOT NULL DEFAULT 0,
            imd INT NOT NULL DEFAULT 0,
            ref_date DATE NULL,
            import_id BIGINT NULL,
            updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                ON UPDATE CURRENT_TIMESTAMP,
            PRIMARY KEY (id),
            UNIQUE KEY uk_lgpi_part (part_no)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )
    # Columnas smt/imd agregadas despues del primer despliegue (ALTER condicional)
    for col_prev, col in (("rework", "smt"), ("smt", "imd")):
        existe = execute_query(
            "SELECT COUNT(*) AS c FROM information_schema.COLUMNS "
            "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'lg_part_inventory' "
            "AND COLUMN_NAME = %s",
            (col,),
            fetch="one",
        )
        if not existe["c"]:
            execute_query(
                f"ALTER TABLE lg_part_inventory "
                f"ADD COLUMN {col} INT NOT NULL DEFAULT 0 AFTER {col_prev}"
            )

    # Etapa 2: schedule diario (renglon S) capturado en MES o importado de Part 10
    execute_query(
        """
        CREATE TABLE IF NOT EXISTS lg_schedule_daily (
            id BIGINT NOT NULL AUTO_INCREMENT,
            part_no VARCHAR(100) NOT NULL,
            sched_date DATE NOT NULL,
            sched_qty INT NOT NULL DEFAULT 0,
            linea VARCHAR(10) NULL,
            turno VARCHAR(15) NOT NULL DEFAULT 'DIA',
            proposal_public_id CHAR(36) NULL,
            updated_by VARCHAR(255) NULL,
            created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                ON UPDATE CURRENT_TIMESTAMP,
            PRIMARY KEY (id),
            UNIQUE KEY uk_lgsd_part_date (part_no, sched_date),
            KEY idx_lgsd_date (sched_date)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )
    # Campos explicitos que sustituyen el color de celda como dato de linea y
    # turno. ALTER idempotente para instalaciones que ya tienen el schedule.
    schedule_columns = (
        ("linea", "VARCHAR(10) NULL AFTER sched_qty"),
        ("turno", "VARCHAR(15) NOT NULL DEFAULT 'DIA' AFTER linea"),
        ("proposal_public_id", "CHAR(36) NULL AFTER turno"),
    )
    for column_name, column_ddl in schedule_columns:
        existe = execute_query(
            "SELECT COUNT(*) AS c FROM information_schema.COLUMNS "
            "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'lg_schedule_daily' "
            "AND COLUMN_NAME = %s",
            (column_name,),
            fetch="one",
        )
        if not existe["c"]:
            execute_query(
                f"ALTER TABLE lg_schedule_daily ADD COLUMN {column_name} {column_ddl}"
            )
    # Plan Proyectado: lotes propuestos/confirmados tipo hoja "LOTE N"
    # (lot_no NULL hasta confirmar; time/falta/%/familia se calculan al vuelo)
    execute_query(
        """
        CREATE TABLE IF NOT EXISTS lg_lote_plan (
            id BIGINT NOT NULL AUTO_INCREMENT,
            plan_date DATE NOT NULL,
            lot_no VARCHAR(20) NULL,
            part_no VARCHAR(100) NOT NULL,
            model VARCHAR(100) NULL,
            main_sub VARCHAR(20) NULL,
            linea VARCHAR(10) NULL,
            turno VARCHAR(15) NOT NULL DEFAULT 'DIA',
            faltante INT NOT NULL DEFAULT 0,
            qty_plan INT NOT NULL DEFAULT 0,
            uph INT NULL,
            estandar_pack INT NULL,
            fisico INT NULL,
            status VARCHAR(12) NOT NULL DEFAULT 'PENDIENTE',
            comentario VARCHAR(255) NULL,
            created_by VARCHAR(80) NULL,
            created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_by VARCHAR(80) NULL,
            updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                ON UPDATE CURRENT_TIMESTAMP,
            confirmed_by VARCHAR(80) NULL,
            confirmed_at DATETIME NULL,
            PRIMARY KEY (id),
            UNIQUE KEY uk_llp_lot (lot_no),
            KEY idx_llp_plan_date (plan_date),
            KEY idx_llp_part (part_no)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )
    # Config del Plan Proyectado (lineas activas configurables)
    execute_query(
        """
        CREATE TABLE IF NOT EXISTS lg_pp_config (
            clave VARCHAR(50) NOT NULL,
            valor VARCHAR(255) NULL,
            PRIMARY KEY (clave)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )
    execute_query(
        "INSERT IGNORE INTO lg_pp_config (clave, valor) VALUES ('lineas_activas', %s)",
        (PPY_LINEAS_DEFAULT,),
    )
    execute_query(
        "INSERT IGNORE INTO lg_pp_config (clave, valor) VALUES ('partes_excluidas', %s)",
        (PPY_EXCLUIDAS_DEFAULT,),
    )
    # Los schedules legacy no tenian linea explicita. Solo se completa cuando
    # la linea del inventario es una linea activa inequívoca; lo ambiguo queda
    # NULL para que el motor falle cerrado y solicite correccion humana.
    execute_query(
        "UPDATE lg_schedule_daily s "
        "JOIN lg_part_inventory i ON i.part_no = s.part_no "
        "JOIN lg_pp_config c ON c.clave = 'lineas_activas' "
        "SET s.linea = UPPER(TRIM(i.line)), s.turno = 'DIA' "
        "WHERE s.linea IS NULL AND i.line IS NOT NULL AND TRIM(i.line) <> '' "
        "AND FIND_IN_SET(UPPER(TRIM(i.line)), REPLACE(UPPER(c.valor), ' ', '')) > 0"
    )
    # Propuestas del motor MRP/capacidad. Se conservan separadas del chat para
    # que borrar una conversacion no elimine el historial de decisiones del
    # planeador. La propuesta no modifica el schedule hasta ser aprobada.
    execute_query(
        """
        CREATE TABLE IF NOT EXISTS lg_plan_proposals (
            id BIGINT NOT NULL AUTO_INCREMENT,
            public_id CHAR(36) NOT NULL,
            version INT NOT NULL DEFAULT 1,
            date_from DATE NOT NULL,
            date_to DATE NOT NULL,
            objective VARCHAR(500) NULL,
            excluded_parts_json LONGTEXT NULL,
            source VARCHAR(20) NOT NULL DEFAULT 'UI',
            engine_version VARCHAR(50) NOT NULL,
            input_hash CHAR(64) NOT NULL,
            status VARCHAR(30) NOT NULL DEFAULT 'DRAFT',
            total_items INT NOT NULL DEFAULT 0,
            total_qty BIGINT NOT NULL DEFAULT 0,
            omitted_count INT NOT NULL DEFAULT 0,
            created_by VARCHAR(255) NOT NULL,
            created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            reviewed_by VARCHAR(255) NULL,
            reviewed_at DATETIME NULL,
            applied_at DATETIME NULL,
            updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                ON UPDATE CURRENT_TIMESTAMP,
            PRIMARY KEY (id),
            UNIQUE KEY uk_lgplanprop_public (public_id),
            KEY idx_lgplanprop_user_status_date (created_by, status, created_at),
            KEY idx_lgplanprop_range (date_from, date_to)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )
    execute_query(
        """
        CREATE TABLE IF NOT EXISTS lg_plan_proposal_items (
            id BIGINT NOT NULL AUTO_INCREMENT,
            public_id CHAR(36) NOT NULL,
            proposal_id BIGINT NOT NULL,
            sequence_no INT NOT NULL,
            part_no VARCHAR(100) NOT NULL,
            sched_date DATE NOT NULL,
            linea VARCHAR(10) NOT NULL,
            turno VARCHAR(15) NOT NULL DEFAULT 'DIA',
            qty_proposed INT NOT NULL,
            qty_final INT NULL,
            ct DECIMAL(12,4) NOT NULL,
            uph INT NOT NULL,
            hours_required DECIMAL(12,4) NOT NULL,
            pack_size INT NOT NULL,
            base_sched_qty INT NOT NULL DEFAULT 0,
            base_sched_linea VARCHAR(10) NULL,
            base_sched_turno VARCHAR(15) NULL,
            inventory_before INT NOT NULL,
            inventory_after INT NOT NULL,
            shortage_date DATE NOT NULL,
            priority_no INT NOT NULL,
            reason VARCHAR(500) NULL,
            requires_approval TINYINT(1) NOT NULL DEFAULT 0,
            exceptions_json LONGTEXT NULL,
            included_final TINYINT(1) NULL,
            final_sched_date DATE NULL,
            final_linea VARCHAR(10) NULL,
            final_turno VARCHAR(15) NULL,
            created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP
                ON UPDATE CURRENT_TIMESTAMP,
            PRIMARY KEY (id),
            UNIQUE KEY uk_lgplanpropitem_public (public_id),
            UNIQUE KEY uk_lgplanpropitem_seq (proposal_id, sequence_no),
            KEY idx_lgplanpropitem_proposal (proposal_id),
            KEY idx_lgplanpropitem_part_date (part_no, sched_date)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )
    proposal_columns = (
        ("excluded_parts_json", "LONGTEXT NULL AFTER objective"),
    )
    for column_name, column_ddl in proposal_columns:
        existe = execute_query(
            "SELECT COUNT(*) AS c FROM information_schema.COLUMNS "
            "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'lg_plan_proposals' "
            "AND COLUMN_NAME = %s",
            (column_name,),
            fetch="one",
        )
        if not existe["c"]:
            execute_query(
                f"ALTER TABLE lg_plan_proposals ADD COLUMN {column_name} {column_ddl}"
            )
    # Migracion idempotente para propuestas creadas con la primera version.
    proposal_item_columns = (
        ("ct", "DECIMAL(12,4) NULL AFTER qty_final"),
        ("base_sched_linea", "VARCHAR(10) NULL AFTER base_sched_qty"),
        ("base_sched_turno", "VARCHAR(15) NULL AFTER base_sched_linea"),
    )
    for column_name, column_ddl in proposal_item_columns:
        existe = execute_query(
            "SELECT COUNT(*) AS c FROM information_schema.COLUMNS "
            "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'lg_plan_proposal_items' "
            "AND COLUMN_NAME = %s",
            (column_name,),
            fetch="one",
        )
        if not existe["c"]:
            execute_query(
                f"ALTER TABLE lg_plan_proposal_items ADD COLUMN {column_name} {column_ddl}"
            )
    execute_query(
        """
        CREATE TABLE IF NOT EXISTS lg_plan_proposal_feedback (
            id BIGINT NOT NULL AUTO_INCREMENT,
            proposal_id BIGINT NOT NULL,
            item_id BIGINT NULL,
            username VARCHAR(255) NOT NULL,
            action VARCHAR(40) NOT NULL,
            reason_code VARCHAR(80) NULL,
            reason_text VARCHAR(1000) NULL,
            before_json LONGTEXT NULL,
            after_json LONGTEXT NULL,
            created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (id),
            KEY idx_lgplanfeedback_proposal (proposal_id, created_at),
            KEY idx_lgplanfeedback_user (username, created_at)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )
    # Lineas permitidas por parte/familia: columna en raw (se configura una
    # vez por familia; aplica a todos los sufijos del mismo prefijo).
    existe = execute_query(
        "SELECT COUNT(*) AS c FROM information_schema.COLUMNS "
        "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'raw' "
        "AND COLUMN_NAME = 'lineas_permitidas'",
        fetch="one",
    )
    if not existe["c"]:
        execute_query(
            "ALTER TABLE raw ADD COLUMN lineas_permitidas VARCHAR(100) NULL "
            "AFTER estandar_pack"
        )


# =============================
# Parser del Excel (compartido por preview y confirm)
# =============================

def _pp_normalizar_parte(valor):
    """Normaliza el numero de parte de la celda (col B)."""
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def _pp_detectar_anio(original_filename, override_year):
    """Anio para fechas string ("07-Jul"): override > filename > anio actual."""
    if override_year:
        return int(override_year)
    m = re.search(r"(20\d{2})", original_filename or "")
    if m:
        return int(m.group(1))
    return date.today().year


def _pp_fecha_archivo(original_filename):
    """Fecha del plan segun el nombre del archivo. None si no la trae.

    Dos convenciones en uso: el xlsm semanal la escribe completa
    ("... ILSAN 20260715 W29 ...") y el Cal diario en corto ("Cal_260715.xlsx").
    """
    base = os.path.basename(original_filename or "")
    for patron, formato in ((r"(20\d{6})", "%Y%m%d"), (r"(\d{6})", "%y%m%d")):
        m = re.search(patron, base)
        if m:
            try:
                return datetime.strptime(m.group(1), formato).date()
            except ValueError:
                continue
    return None


def _pp_ref_lunes(original_filename):
    """Lunes de la semana del archivo ("... ILSAN 20260715 W29 ...") -> 2026-07-13.

    El inventario (LGEMM/ISEMM/...) es la foto del lunes de esa semana, no de hoy:
    validado contra la fila I de la hoja Part N (404/404 partes cuadran con este
    lunes; con el lunes de hoy fallan las partes con S en semanas previas).
    Sin fecha en el nombre cae al lunes actual.
    """
    f = _pp_fecha_archivo(original_filename) or date.today()
    return f - timedelta(days=f.weekday())


def _pp_import_mas_reciente(fecha_archivo, date_from, date_to):
    """Import previo hecho con un archivo MAS NUEVO que este, sobre las mismas
    fechas. None si no hay.

    El upsert es por parte+fecha y gana el ultimo, asi que importar el xlsm del
    lunes despues del Cal del jueves revierte en silencio lo que el Cal
    actualizo. Es un escenario probable: el inventario solo viene en el xlsm, y
    al re-importarlo para refrescarlo se pisa el plan mas fresco.
    """
    if fecha_archivo is None:
        return None
    rows = execute_query(
        "SELECT original_filename, imported_at FROM lg_plan_imports "
        "WHERE status = 'COMPLETADO' AND date_from <= %s AND date_to >= %s "
        "ORDER BY id DESC LIMIT 100",
        (date_to, date_from),
        fetch="all",
    ) or []
    for r in rows:
        otra = _pp_fecha_archivo(r["original_filename"])
        if otra and otra > fecha_archivo:
            return {"archivo": r["original_filename"],
                    "fecha_archivo": otra.isoformat()}
    return None


def _pp_parse_fecha_celda(valor, anio_fallback):
    """Convierte una celda de la fila 3 a date, o None si no es fecha."""
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        texto = valor.strip()
        for fmt in ("%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(texto, fmt).date()
            except ValueError:
                pass
        for fmt in ("%d-%b", "%d/%m"):
            try:
                parcial = datetime.strptime(texto, fmt)
                return date(anio_fallback, parcial.month, parcial.day)
            except ValueError:
                pass
    return None


def _pp_hoja_plan(wb, original_filename=""):
    """Hoja con el plan de LG. None si el archivo no trae ninguna.

    Dos formatos en uso:
      - el xlsm semanal ("plan de produccion ILSAN ..."): hoja 'LG'
      - el Cal diario ("Cal_260715.xlsx"): una hoja por fecha (260629, 260715,
        ...) mas hojas sueltas (BOM, ISSUES). Se toma la del nombre del archivo
        y, si no esta, la mas reciente.
    """
    if PP_SHEET_NAME in wb.sheetnames:
        return PP_SHEET_NAME
    fechadas = [s for s in wb.sheetnames if re.fullmatch(r"\d{6}", s.strip())]
    if not fechadas:
        return None
    m = re.search(r"(\d{6})", os.path.basename(original_filename or ""))
    if m and m.group(1) in (s.strip() for s in fechadas):
        return m.group(1)
    return max(fechadas, key=lambda s: s.strip())


def _pp_layout_plan(ws, anio_fallback):
    """(fila_fechas, col_parte, {col: fecha}) del plan. El layout no es fijo.

    La hoja LG trae las fechas en la fila 3 y la parte en la col B; el Cal, en
    la fila 2 y la col A. En vez de fijar cada formato se busca la fila con mas
    fechas (las de arriba llevan semana o consecutivos, no fechas) y se asume
    que la parte va en la columna justo antes de la primera fecha.
    """
    mejor = (0, None, {})
    for idx, fila in enumerate(
        ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1
    ):
        cols = {}
        for ci in range(0, len(fila)):
            fecha = _pp_parse_fecha_celda(fila[ci], anio_fallback)
            if fecha is not None:
                cols[ci] = fecha
        if len(cols) > mejor[0]:
            mejor = (len(cols), idx, cols)
    _n, fila_fechas, col_fechas = mejor
    if not col_fechas:
        return None, None, {}
    return fila_fechas, max(0, min(col_fechas) - 1), col_fechas


def _parse_lg_workbook(file_bytes, original_filename="", override_year=None):
    """Parsea el Excel del plan LG (hoja 'LG' del xlsm o el Cal diario).

    Retorna (parsed_dict, None) si el archivo es utilizable, o
    (None, (payload_error, http_status)) si hay error bloqueante.
    """
    def error(msg, status=400):
        return None, ({"success": False, "errors": [msg], "warnings": []}, status)

    ext = os.path.splitext(original_filename or "")[1].lower()
    if ext not in PP_EXTENSIONES:
        return error("Formato no permitido. Solo se aceptan archivos .xlsx o .xlsm.")
    if not file_bytes:
        return error("El archivo esta vacio.")
    if len(file_bytes) > PP_MAX_FILE_BYTES:
        return error("El archivo supera el tamano maximo permitido (20 MB).")

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
    except Exception as e:
        logger.warning("part_planning: archivo Excel ilegible: %s", e)
        return error("El archivo esta corrupto o no es un Excel valido.")

    try:
        hoja = _pp_hoja_plan(wb, original_filename)
        if hoja is None:
            return error(
                f"No se encontro la hoja '{PP_SHEET_NAME}' ni una hoja con "
                f"fecha (ej. 260715) en el archivo. "
                f"Hojas disponibles: {', '.join(wb.sheetnames[:10])}"
            )
        ws = wb[hoja]

        anio_fallback = _pp_detectar_anio(original_filename, override_year)
        warnings = []

        fila_fechas, col_parte, col_fechas = _pp_layout_plan(ws, anio_fallback)
        if not col_fechas:
            return error(f"No se encontraron fechas validas en la hoja '{hoja}'.")

        filas = ws.iter_rows(min_row=fila_fechas + 1, values_only=True)

        # Filas de datos: parte en col_parte, cantidades en col_fechas.
        records = {}
        partes_vistas = set()
        partes_duplicadas = set()
        celdas_no_numericas = 0
        for fila in filas:  # continua despues de la fila de fechas
            if len(fila) <= col_parte:
                continue
            part_no = _pp_normalizar_parte(fila[col_parte])
            if not part_no:
                continue
            if part_no in partes_vistas:
                partes_duplicadas.add(part_no)
            partes_vistas.add(part_no)
            for ci, fecha in col_fechas.items():
                valor = fila[ci] if ci < len(fila) else None
                if valor is None or valor == "":
                    qty = 0
                elif isinstance(valor, (int, float)):
                    qty = int(round(valor))
                else:
                    celdas_no_numericas += 1
                    continue
                if qty < 0:
                    warnings.append(
                        f"Cantidad negativa en {part_no} / {fecha.isoformat()}: "
                        f"{qty} se ajusto a 0."
                    )
                    qty = 0
                clave = (part_no, fecha)
                # Duplicados de parte se consolidan sumando.
                records[clave] = records.get(clave, 0) + qty

        if not partes_vistas:
            return error(
                f"No se encontraron numeros de parte en la hoja '{hoja}' "
                f"(columna {chr(65 + col_parte)}, bajo la fila de fechas)."
            )
        if not records:
            return error("El archivo no contiene registros de plan.")

        if celdas_no_numericas:
            warnings.append(
                f"{celdas_no_numericas} celdas con texto no numerico fueron ignoradas."
            )

        fechas = sorted(set(col_fechas.values()))
        records_count = sum(1 for q in records.values() if q > 0)
        parsed = {
            "sheet_name": hoja,
            "plan_year": fechas[0].year,
            "date_from": fechas[0],
            "date_to": fechas[-1],
            "dates": fechas,
            "records": records,
            "parts": sorted(partes_vistas),
            "parts_count": len(partes_vistas),
            "dates_count": len(fechas),
            "records_count": records_count,
            "zero_records_count": len(records) - records_count,
            "duplicated_parts": sorted(partes_duplicadas),
            "warnings": warnings,
            "errors": [],
        }
        return parsed, None
    finally:
        wb.close()


def _pp_compute_warnings(parsed):
    """Advertencias que requieren BD (no bloquean). Muta parsed["warnings"]."""
    warnings = parsed["warnings"]

    dups = parsed["duplicated_parts"]
    if dups:
        listado = ", ".join(dups[:10]) + ("..." if len(dups) > 10 else "")
        warnings.append(
            f"{len(dups)} numeros de parte repetidos se consolidaron sumando: {listado}"
        )

    try:
        rows = execute_query(
            "SELECT DISTINCT TRIM(part_no) AS part_no FROM raw", fetch="all"
        ) or []
        raw_parts = {r["part_no"] for r in rows if r.get("part_no")}
        faltantes = [p for p in parsed["parts"] if p not in raw_parts]
        if faltantes:
            listado = ", ".join(faltantes[:10]) + ("..." if len(faltantes) > 10 else "")
            warnings.append(
                f"{len(faltantes)} numeros de parte no existen en RAW: {listado}"
            )
    except Exception as e:
        logger.warning("part_planning: no se pudo consultar RAW: %s", e)

    hoy = date.today()
    pasadas = [f for f in parsed["dates"] if f < hoy]
    if pasadas:
        warnings.append(
            f"{len(pasadas)} fechas del plan ya pasaron "
            f"({pasadas[0].isoformat()} a {pasadas[-1].isoformat()})."
        )

    try:
        row = execute_query(
            """
            SELECT COUNT(*) AS c, MIN(plan_date) AS f1, MAX(plan_date) AS f2
            FROM lg_plan_daily
            WHERE plan_date BETWEEN %s AND %s
            """,
            (parsed["date_from"], parsed["date_to"]),
            fetch="one",
        )
        if row and int(row.get("c") or 0) > 0:
            warnings.append(
                f"Ya existen {row['c']} registros de plan entre {row['f1']} y "
                f"{row['f2']}; seran afectados segun el modo de importacion."
            )
    except Exception as e:
        logger.warning("part_planning: no se pudo consultar plan existente: %s", e)


def _pp_summary(parsed):
    return {
        "plan_year": parsed["plan_year"],
        "date_from": parsed["date_from"].isoformat(),
        "date_to": parsed["date_to"].isoformat(),
        "parts_count": parsed["parts_count"],
        "dates_count": parsed["dates_count"],
        "records_count": parsed["records_count"],
        "zero_records_count": parsed["zero_records_count"],
        "duplicates_count": len(parsed["duplicated_parts"]),
    }


def _pp_filas_segun_modo(parsed, modo, include_zero):
    """Lista de (part_no, plan_date, qty) a escribir segun el modo."""
    solo_positivos = modo == "only_positive" or not include_zero
    return [
        (part, fecha, qty)
        for (part, fecha), qty in parsed["records"].items()
        if qty > 0 or not solo_positivos
    ]


# =============================
# Parser hoja "Part 10" (inventario semanal LGEMM/ISEMM + schedule)
# =============================

# La hoja cambia de nombre por dia ("Part 10", "Part 14", ...): se busca por patron.
PP_INV_SHEET_RE = re.compile(r"^part\s*\d+\s*$", re.IGNORECASE)


def _pp_buscar_hoja_inventario(sheetnames):
    for nombre in sheetnames:
        if PP_INV_SHEET_RE.match(nombre.strip()):
            return nombre
    return None


def _parse_part10_workbook(file_bytes, original_filename=""):
    """Parsea la hoja 'Part 10': escalares de inventario por parte + renglon S.

    Estructura (verificada contra el archivo real): fila de encabezado con
    'PART NUMBER' en col C; LGEMM..REWORK en cols E-J; linea en col M; tipo de
    renglon (P/S/I) en col N; fechas datetime en el mismo encabezado desde
    col Q. Cada parte ocupa un bloque de 3 renglones P/S/I.

    Retorna (parsed, None) o (None, (payload_error, http_status)).
    """
    def error(msg, status=400):
        return None, ({"success": False, "errors": [msg], "warnings": []}, status)

    ext = os.path.splitext(original_filename or "")[1].lower()
    if ext not in PP_EXTENSIONES:
        return error("Formato no permitido. Solo se aceptan archivos .xlsx o .xlsm.")
    if not file_bytes:
        return error("El archivo esta vacio.")
    if len(file_bytes) > PP_MAX_FILE_BYTES:
        return error("El archivo supera el tamano maximo permitido (20 MB).")

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
    except Exception as e:
        logger.warning("part_planning: archivo Excel ilegible: %s", e)
        return error("El archivo esta corrupto o no es un Excel valido.")

    try:
        hoja_inv = _pp_buscar_hoja_inventario(wb.sheetnames)
        if not hoja_inv:
            return error("No se encontro una hoja 'Part N' (ej. Part 10, Part 14) en el archivo.")
        ws = wb[hoja_inv]

        warnings = []
        filas = ws.iter_rows(values_only=True)

        # Encabezado: fila con 'PART NUMBER' en col C (indice 2).
        col_fechas = {}
        encontrado = False
        for fila in filas:
            valor = fila[2] if len(fila) > 2 else None
            if isinstance(valor, str) and valor.strip().upper() == "PART NUMBER":
                for ci in range(3, len(fila)):
                    f = _pp_parse_fecha_celda(fila[ci], date.today().year)
                    if f is not None:
                        col_fechas[ci] = f
                encontrado = True
                break
        if not encontrado:
            return error(
                f"No se encontro el encabezado 'PART NUMBER' en la hoja '{hoja_inv}'."
            )
        if not col_fechas:
            return error(f"No se encontraron fechas en el encabezado de '{hoja_inv}'.")

        def escalar(v):
            return int(round(v)) if isinstance(v, (int, float)) else 0

        inventory = {}
        schedules = {}
        parte_actual = None
        for fila in filas:  # continua despues del encabezado
            tipo = str(fila[13]).strip().upper() if len(fila) > 13 and fila[13] else ""
            if tipo == "P":
                part = _pp_normalizar_parte(fila[2] if len(fila) > 2 else None)
                if not part:
                    parte_actual = None
                    continue
                if part in inventory:
                    warnings.append(f"Parte repetida en {hoja_inv}: {part} (se toma la primera).")
                    parte_actual = None
                    continue
                parte_actual = part
                inventory[part] = {
                    "board": (str(fila[3]).strip()[:100] if len(fila) > 3 and fila[3] else None),
                    "line": (str(fila[12]).strip()[:20] if len(fila) > 12 and fila[12] else None),
                    "lgemm": escalar(fila[4] if len(fila) > 4 else None),
                    "isemm": escalar(fila[5] if len(fila) > 5 else None),
                    "svc": escalar(fila[6] if len(fila) > 6 else None),
                    "dif": escalar(fila[7] if len(fila) > 7 else None),
                    "pendiente": escalar(fila[8] if len(fila) > 8 else None),
                    "rework": escalar(fila[9] if len(fila) > 9 else None),
                    "smt": escalar(fila[10] if len(fila) > 10 else None),
                    "imd": escalar(fila[11] if len(fila) > 11 else None),
                }
            elif tipo == "S" and parte_actual:
                for ci, fecha in col_fechas.items():
                    v = fila[ci] if ci < len(fila) else None
                    if isinstance(v, (int, float)) and v > 0:
                        clave = (parte_actual, fecha)
                        schedules[clave] = schedules.get(clave, 0) + int(round(v))

        if not inventory:
            return error(f"No se encontraron partes en la hoja '{hoja_inv}'.")

        fechas = sorted(col_fechas.values())
        return {
            "sheet_name": hoja_inv,
            "ref_date": fechas[0],
            "date_to": fechas[-1],
            "dates_count": len(fechas),
            "inventory": inventory,
            "schedules": schedules,
            "warnings": warnings,
        }, None
    finally:
        wb.close()


# =============================
# RUTAS RENDER
# =============================

@bp.route("/part-planning-ajax")
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PP_PERMISO_BOTON)
def part_planning_ajax():
    """Ruta AJAX para cargar dinamicamente el contenido de Part Planning LG."""
    try:
        return render_template("Control de produccion/part_planning_ajax.html")
    except Exception as e:
        logger.error("Error al cargar Part Planning LG: %s", e)
        return f"Error al cargar el contenido: {str(e)}", 500


@bp.route("/proyeccion-ajax")
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PROY_PERMISO_BOTON)
def proyeccion_ajax():
    """Ruta AJAX para el modulo Proyeccion (renglones P/S/I)."""
    try:
        return render_template("Control de produccion/proyeccion_ajax.html")
    except Exception as e:
        logger.error("Error al cargar Proyeccion: %s", e)
        return f"Error al cargar el contenido: {str(e)}", 500


@bp.route("/plan-proyectado-ajax")
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PPY_PERMISO_BOTON)
def plan_proyectado_ajax():
    """Ruta AJAX para el modulo Plan Proyectado (generador de lotes)."""
    try:
        return render_template("Control de produccion/plan_proyectado_ajax.html")
    except Exception as e:
        logger.error("Error al cargar Plan Proyectado: %s", e)
        return f"Error al cargar el contenido: {str(e)}", 500


# =============================
# RUTAS API
# =============================

@bp.route("/api/part-planning/import/preview", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PP_PERMISO_BOTON)
def api_pp_import_preview():
    """Analiza el Excel del plan LG sin guardar nada."""
    try:
        archivo = request.files.get("file")
        if archivo is None or not archivo.filename:
            return jsonify({"success": False, "errors": ["No se recibio archivo."]}), 400
        file_bytes = archivo.read()
        override_year = request.form.get("plan_year", type=int)

        parsed, err = _parse_lg_workbook(file_bytes, archivo.filename, override_year)
        if err is not None:
            payload, status = err
            return jsonify(payload), status

        _pp_compute_warnings(parsed)
        # Mismo aviso que en la IA: importar un archivo mas viejo que el ultimo
        # revierte en silencio lo que el mas nuevo actualizo.
        viejo = _pp_import_mas_reciente(
            _pp_fecha_archivo(archivo.filename), parsed["date_from"], parsed["date_to"])
        if viejo:
            parsed["warnings"].insert(0, (
                f"Ya se importo un archivo mas nuevo ({viejo['archivo']}, del "
                f"{viejo['fecha_archivo']}). Confirmar revierte el plan de las "
                f"fechas que se solapan a esta version mas vieja."
            ))

        sample = sorted(
            (
                (part, fecha, qty)
                for (part, fecha), qty in parsed["records"].items()
                if qty > 0
            ),
            key=lambda x: (x[0], x[1]),
        )[:100]

        return jsonify(
            {
                "success": True,
                "file_sha256": hashlib.sha256(file_bytes).hexdigest(),
                "file": {
                    "name": archivo.filename,
                    "sheet": parsed.get("sheet_name") or PP_SHEET_NAME,
                    "size_bytes": len(file_bytes),
                },
                "summary": _pp_summary(parsed),
                "errors": [],
                "warnings": parsed["warnings"],
                "sample_rows": [
                    {"part_no": p, "plan_date": f.isoformat(), "plan_qty": q}
                    for p, f, q in sample
                ],
            }
        )
    except Exception as e:
        logger.error("Error en api_pp_import_preview: %s", e, exc_info=True)
        return jsonify({"success": False, "errors": [f"Error interno: {e}"]}), 500


@bp.route("/api/part-planning/import/confirm", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PP_PERMISO_BOTON)
def api_pp_import_confirm():
    """Guarda el plan LG (transaccional) con auditoria de cambios."""
    conn = None
    cursor = None
    try:
        archivo = request.files.get("file")
        if archivo is None or not archivo.filename:
            return jsonify({"success": False, "errors": ["No se recibio archivo."]}), 400
        file_bytes = archivo.read()

        hash_esperado = (request.form.get("file_sha256") or "").strip().lower()
        if not hash_esperado:
            return (
                jsonify(
                    {
                        "success": False,
                        "errors": ["file_sha256 requerido. Analiza el archivo primero."],
                    }
                ),
                400,
            )
        hash_real = hashlib.sha256(file_bytes).hexdigest()
        if hash_real != hash_esperado:
            return (
                jsonify(
                    {
                        "success": False,
                        "errors": [
                            "El archivo cambio desde la vista previa. "
                            "Vuelve a analizarlo antes de confirmar."
                        ],
                    }
                ),
                400,
            )

        modo = (request.form.get("import_mode") or "upsert").strip()
        if modo not in PP_MODOS:
            return jsonify({"success": False, "errors": [f"Modo invalido: {modo}"]}), 400
        include_zero = (request.form.get("include_zero") or "1").strip().lower() in (
            "1", "true", "on", "si",
        )
        override_year = request.form.get("plan_year", type=int)

        parsed, err = _parse_lg_workbook(file_bytes, archivo.filename, override_year)
        if err is not None:
            payload, status = err
            return jsonify(payload), status
        _pp_compute_warnings(parsed)

        filas = _pp_filas_segun_modo(parsed, modo, include_zero)
        usuario = session.get("usuario") or "SISTEMA"

        conn = get_pooled_connection()
        if conn is None:
            raise RuntimeError("No fue posible obtener conexion MySQL.")
        cursor = get_dict_cursor(conn)
        conn.autocommit(False)

        # Snapshot de lo existente en el rango para el diff de auditoria.
        cursor.execute(
            "SELECT part_no, plan_date, plan_qty FROM lg_plan_daily "
            "WHERE plan_date BETWEEN %s AND %s",
            (parsed["date_from"], parsed["date_to"]),
        )
        existentes = {
            (r["part_no"], r["plan_date"]): int(r["plan_qty"])
            for r in cursor.fetchall()
        }

        cambios = []  # (part_no, plan_date, old_qty, new_qty, change_type)
        a_escribir = []
        inserted = updated = unchanged = 0
        for part, fecha, qty in filas:
            old = existentes.get((part, fecha))
            if old is None:
                inserted += 1
                cambios.append((part, fecha, None, qty, "INSERT"))
                a_escribir.append((part, fecha, qty))
            elif modo == "only_new":
                unchanged += 1
            elif old != qty:
                updated += 1
                cambios.append((part, fecha, old, qty, "UPDATE"))
                a_escribir.append((part, fecha, qty))
            else:
                unchanged += 1

        cursor.execute(
            """
            INSERT INTO lg_plan_imports
                (original_filename, sheet_name, plan_year, date_from, date_to,
                 parts_count, dates_count, records_count, zero_records_count,
                 warning_count, import_mode, file_sha256, imported_by, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'COMPLETADO')
            """,
            (
                archivo.filename[:255],
                PP_SHEET_NAME,
                parsed["plan_year"],
                parsed["date_from"],
                parsed["date_to"],
                parsed["parts_count"],
                parsed["dates_count"],
                parsed["records_count"],
                parsed["zero_records_count"],
                len(parsed["warnings"]),
                modo,
                hash_real,
                usuario,
            ),
        )
        import_id = cursor.lastrowid

        sql_upsert = (
            "INSERT INTO lg_plan_daily (part_no, plan_date, plan_qty, import_id) "
            "VALUES (%s, %s, %s, %s) "
            "ON DUPLICATE KEY UPDATE plan_qty = VALUES(plan_qty), "
            "import_id = VALUES(import_id)"
        )
        datos = [(p, f, q, import_id) for p, f, q in a_escribir]
        for i in range(0, len(datos), PP_BATCH_SIZE):
            cursor.executemany(sql_upsert, datos[i : i + PP_BATCH_SIZE])

        sql_cambio = (
            "INSERT INTO lg_plan_import_changes "
            "(import_id, part_no, plan_date, old_qty, new_qty, change_type) "
            "VALUES (%s, %s, %s, %s, %s, %s)"
        )
        datos_cambios = [(import_id, p, f, o, n, t) for p, f, o, n, t in cambios]
        for i in range(0, len(datos_cambios), PP_BATCH_SIZE):
            cursor.executemany(sql_cambio, datos_cambios[i : i + PP_BATCH_SIZE])

        conn.commit()

        return jsonify(
            {
                "success": True,
                "import_id": import_id,
                "inserted": inserted,
                "updated": updated,
                "unchanged": unchanged,
                "date_from": parsed["date_from"].isoformat(),
                "date_to": parsed["date_to"].isoformat(),
                "warnings": parsed["warnings"],
            }
        )
    except Exception as e:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        logger.error("Error en api_pp_import_confirm: %s", e, exc_info=True)
        return jsonify({"success": False, "errors": [f"Error interno: {e}"]}), 500
    finally:
        if cursor is not None:
            try:
                cursor.close()
            except Exception:
                pass
        if conn is not None:
            try:
                conn.autocommit(True)
                conn.close()
            except Exception:
                pass


@bp.route("/api/part-planning/inventory/import", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PROY_PERMISO_BOTON)
def api_pp_inventory_import():
    """Importa inventario semanal (LGEMM/ISEMM/...) + schedules de la hoja Part 10."""
    conn = None
    cursor = None
    try:
        archivo = request.files.get("file")
        if archivo is None or not archivo.filename:
            return jsonify({"success": False, "errors": ["No se recibio archivo."]}), 400
        file_bytes = archivo.read()

        parsed, err = _parse_part10_workbook(file_bytes, archivo.filename)
        if err is not None:
            payload, status = err
            return jsonify(payload), status

        usuario = session.get("usuario") or "SISTEMA"
        file_hash = hashlib.sha256(file_bytes).hexdigest()
        # ref_date = lunes de la semana DEL ARCHIVO: el inventario importado ya
        # incluye lo producido antes; arrancar la proyeccion en la primera fecha
        # de la hoja restaria plan y sumaria schedules pasados dos veces.
        ref_lunes = _pp_ref_lunes(archivo.filename)

        conn = get_pooled_connection()
        if conn is None:
            raise RuntimeError("No fue posible obtener conexion MySQL.")
        cursor = get_dict_cursor(conn)
        conn.autocommit(False)

        cursor.execute(
            """
            INSERT INTO lg_plan_imports
                (original_filename, sheet_name, plan_year, date_from, date_to,
                 parts_count, dates_count, records_count, zero_records_count,
                 warning_count, import_mode, file_sha256, imported_by, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 0, %s, 'inventory', %s, %s, 'COMPLETADO')
            """,
            (
                archivo.filename[:255],
                parsed["sheet_name"],
                parsed["ref_date"].year,
                parsed["ref_date"],
                parsed["date_to"],
                len(parsed["inventory"]),
                parsed["dates_count"],
                len(parsed["schedules"]),
                len(parsed["warnings"]),
                file_hash,
                usuario,
            ),
        )
        import_id = cursor.lastrowid

        sql_inv = (
            "INSERT INTO lg_part_inventory "
            "(part_no, board, line, lgemm, isemm, svc, dif, pendiente, rework, "
            " smt, imd, ref_date, import_id) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) "
            "ON DUPLICATE KEY UPDATE board=VALUES(board), line=VALUES(line), "
            "lgemm=VALUES(lgemm), isemm=VALUES(isemm), svc=VALUES(svc), "
            "dif=VALUES(dif), pendiente=VALUES(pendiente), rework=VALUES(rework), "
            "smt=VALUES(smt), imd=VALUES(imd), "
            "ref_date=VALUES(ref_date), import_id=VALUES(import_id)"
        )
        datos_inv = [
            (p, d["board"], d["line"], d["lgemm"], d["isemm"], d["svc"],
             d["dif"], d["pendiente"], d["rework"], d["smt"], d["imd"],
             ref_lunes, import_id)
            for p, d in parsed["inventory"].items()
        ]
        for i in range(0, len(datos_inv), PP_BATCH_SIZE):
            cursor.executemany(sql_inv, datos_inv[i : i + PP_BATCH_SIZE])

        sql_sched = (
            "INSERT INTO lg_schedule_daily (part_no, sched_date, sched_qty, updated_by) "
            "VALUES (%s, %s, %s, %s) "
            "ON DUPLICATE KEY UPDATE sched_qty=VALUES(sched_qty), "
            "updated_by=VALUES(updated_by)"
        )
        datos_sched = [
            (p, f, q, usuario) for (p, f), q in parsed["schedules"].items()
        ]
        for i in range(0, len(datos_sched), PP_BATCH_SIZE):
            cursor.executemany(sql_sched, datos_sched[i : i + PP_BATCH_SIZE])

        conn.commit()

        return jsonify(
            {
                "success": True,
                "import_id": import_id,
                "parts": len(parsed["inventory"]),
                "schedules": len(parsed["schedules"]),
                "ref_date": ref_lunes.isoformat(),
                "date_to": parsed["date_to"].isoformat(),
                "warnings": parsed["warnings"],
            }
        )
    except Exception as e:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        logger.error("Error en api_pp_inventory_import: %s", e, exc_info=True)
        return jsonify({"success": False, "errors": [f"Error interno: {e}"]}), 500
    finally:
        if cursor is not None:
            try:
                cursor.close()
            except Exception:
                pass
        if conn is not None:
            try:
                conn.autocommit(True)
                conn.close()
            except Exception:
                pass


def _pp_sincronizar_schedule_excel(
    file_bytes, filename, usuario, *, aplicar=False, alcance="todos"
):
    """Previsualiza o aplica la misma sincronizacion del boton ``Part N``.

    Solo reemplaza el renglon S para las partes que traen schedule positivo en
    el archivo y dentro de su rango de fechas. No modifica plan LG ni inventario.
    """
    parsed, err = _parse_part10_workbook(file_bytes, filename)
    if err is not None:
        payload, _status = err
        raise ValueError(payload["errors"][0])

    alcance = str(alcance or "todos").strip().lower()
    if alcance not in {"main", "todos"}:
        raise ValueError("alcance debe ser 'main' o 'todos'")
    por_parte = {}
    for (part_no, fecha), qty in parsed["schedules"].items():
        por_parte.setdefault(part_no, []).append((fecha, int(qty)))
    rango_desde = parsed["ref_date"]
    rango_hasta = parsed["date_to"]
    lineas_activas = set(_ppy_config_lineas())
    linea_por_parte = {
        part: str((info or {}).get("line") or "").strip().upper()[:10]
        for part, info in (parsed.get("inventory") or {}).items()
    }
    # Igual que el boton: RAW manda sobre la columna LINE de Part N.
    raw_lineas = _ppy_datos_raw(list(por_parte)).copy()
    for part, raw_info in raw_lineas.items():
        assy_line = str(raw_info.get("assy_line") or "").strip().upper()[:10]
        if assy_line:
            linea_por_parte[part] = assy_line
    partes_archivo = set(por_parte)
    sin_linea = sorted(
        part
        for part in partes_archivo
        if linea_por_parte.get(part) not in lineas_activas
    )
    partes_validas = partes_archivo - set(sin_linea)
    if alcance == "main":
        partes_alcance = {
            part
            for part in partes_validas
            if linea_por_parte.get(part) in {"M1", "M2", "M3", "M4"}
        }
    else:
        partes_alcance = partes_validas
    por_parte = {
        part: celdas for part, celdas in por_parte.items() if part in partes_alcance
    }

    warnings = []
    if sin_linea:
        warnings.append(
            {
                "code": "SCHEDULE_PARTS_SKIPPED_NO_ACTIVE_LINE",
                "count": len(sin_linea),
                "parts": sin_linea[:50],
                "message": (
                    "Se omitieron schedules de partes sin Assy line activa; "
                    "las demas partes sí se pueden sincronizar."
                ),
            }
        )

    summary = {
        "success": True,
        "sheet_name": parsed["sheet_name"],
        "parts": len(por_parte),
        "schedules": sum(len(celdas) for celdas in por_parte.values()),
        "date_from": rango_desde.isoformat(),
        "date_to": rango_hasta.isoformat(),
        "scope": alcance,
        "source_parts": len(partes_archivo),
        "excluded_by_scope": len(partes_validas - partes_alcance),
        "skipped_without_active_line": len(sin_linea),
        "skipped_parts_without_active_line": sin_linea[:50],
        "warnings": warnings,
        "applied": False,
    }
    if not aplicar:
        return summary
    if not por_parte:
        return {**summary, "replaced": 0, "applied": True}

    conn = None
    cursor = None
    try:
        conn = get_pooled_connection()
        if conn is None:
            raise RuntimeError("No fue posible obtener conexion MySQL.")
        cursor = get_dict_cursor(conn)
        conn.autocommit(False)

        partes = list(por_parte)
        borradas = 0
        for i in range(0, len(partes), PP_BATCH_SIZE):
            lote = partes[i : i + PP_BATCH_SIZE]
            placeholders = ", ".join(["%s"] * len(lote))
            cursor.execute(
                "DELETE FROM lg_schedule_daily "
                f"WHERE part_no IN ({placeholders}) AND sched_date BETWEEN %s AND %s",
                tuple(lote) + (rango_desde, rango_hasta),
            )
            borradas += cursor.rowcount or 0

        datos = [
            (p, f, q, linea_por_parte[p], "DIA", usuario)
            for p, celdas in por_parte.items()
            for (f, q) in celdas
            if q > 0
        ]
        for i in range(0, len(datos), PP_BATCH_SIZE):
            cursor.executemany(
                "INSERT INTO lg_schedule_daily "
                "(part_no, sched_date, sched_qty, linea, turno, updated_by) "
                "VALUES (%s, %s, %s, %s, %s, %s)",
                datos[i : i + PP_BATCH_SIZE],
            )
        conn.commit()
        return {
            **summary,
            "schedules": len(datos),
            "replaced": borradas,
            "applied": True,
        }
    except Exception:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        raise
    finally:
        if cursor is not None:
            try:
                cursor.close()
            except Exception:
                pass
        if conn is not None:
            try:
                conn.autocommit(True)
                conn.close()
            except Exception:
                pass


@bp.route("/api/part-planning/schedule/sync-excel", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PROY_PERMISO_BOTON)
def api_pp_schedule_sync_excel():
    """Sincroniza SOLO el schedule (renglon S) desde la hoja Part N."""
    try:
        archivo = request.files.get("file")
        if archivo is None or not archivo.filename:
            return jsonify({"success": False, "errors": ["No se recibio archivo."]}), 400
        result = _pp_sincronizar_schedule_excel(
            archivo.read(),
            archivo.filename,
            session.get("usuario") or "SISTEMA",
            aplicar=True,
        )
        return jsonify(result)
    except ValueError as exc:
        return jsonify({"success": False, "errors": [str(exc)]}), 400
    except Exception as exc:
        logger.error("Error en api_pp_schedule_sync_excel: %s", exc, exc_info=True)
        return jsonify({"success": False, "errors": [f"Error interno: {exc}"]}), 500


@bp.route("/api/part-planning/schedule", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PROY_PERMISO_BOTON)
def api_pp_schedule_save():
    """Captura/edita el schedule (renglon S) de una parte en una fecha."""
    try:
        data = request.get_json(silent=True) or {}
        part_no = (data.get("part_no") or "").strip()
        fecha_txt = (data.get("sched_date") or "").strip()
        if not part_no or not fecha_txt:
            return jsonify({"success": False, "error": "part_no y sched_date requeridos"}), 400
        try:
            fecha = datetime.strptime(fecha_txt, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({"success": False, "error": "sched_date invalida"}), 400

        qty_raw = data.get("sched_qty")
        usuario = session.get("usuario") or "SISTEMA"

        if qty_raw is None or str(qty_raw).strip() == "" or int(qty_raw) <= 0:
            execute_query(
                "DELETE FROM lg_schedule_daily WHERE part_no = %s AND sched_date = %s",
                (part_no, fecha),
            )
            return jsonify({"success": True, "part_no": part_no,
                            "sched_date": fecha.isoformat(), "sched_qty": None})

        qty = int(qty_raw)
        linea = str(data.get("linea") or "").strip().upper()[:10]
        turno = str(data.get("turno") or "DIA").strip().upper()[:15]
        lineas_activas = _ppy_config_lineas()
        if linea not in lineas_activas:
            return jsonify(
                {
                    "success": False,
                    "error": "Selecciona una linea activa antes de capturar el schedule",
                }
            ), 400
        if turno not in PPY_TURNOS:
            return jsonify({"success": False, "error": "turno invalido"}), 400
        permitidas_parte, permitidas_familia = _ppy_lineas_permitidas_map()
        permitidas = permitidas_parte.get(part_no) or permitidas_familia.get(
            _ppy_familia(part_no)
        )
        if permitidas and linea not in permitidas:
            return jsonify(
                {
                    "success": False,
                    "error": f"La parte {part_no} no esta permitida en {linea}",
                }
            ), 400
        execute_query(
            "INSERT INTO lg_schedule_daily "
            "(part_no, sched_date, sched_qty, linea, turno, proposal_public_id, updated_by) "
            "VALUES (%s, %s, %s, %s, %s, NULL, %s) "
            "ON DUPLICATE KEY UPDATE sched_qty=VALUES(sched_qty), "
            "linea=VALUES(linea), turno=VALUES(turno), proposal_public_id=NULL, "
            "updated_by=VALUES(updated_by)",
            (part_no, fecha, qty, linea, turno, usuario),
        )
        return jsonify({"success": True, "part_no": part_no,
                        "sched_date": fecha.isoformat(), "sched_qty": qty})
    except Exception as e:
        logger.error("Error en api_pp_schedule_save: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


def _ppy_omission_codes(motivo):
    """Convierte mensajes internos en codigos estables para UI y aprendizaje."""
    texto = str(motivo or "")
    codigos = []
    if "CT" in texto:
        codigos.append("CT_MISSING")
    if "UPH" in texto:
        codigos.append("UPH_MISSING")
    if "pack" in texto.lower():
        codigos.append("PACK_MISSING")
    if "no estan activas" in texto:
        codigos.append("NO_ACTIVE_ALLOWED_LINE")
    if "sin horas" in texto:
        codigos.append("NO_CAPACITY")
    return codigos or ["NOT_SCHEDULABLE"]


def _ppy_schedule_snapshot(fecha_ini, fecha_fin, query=None):
    """Foto del Schedule MAIN que Planning entrego como punto de partida."""
    run_query = query or execute_query
    rows = run_query(
        "SELECT part_no, sched_date, sched_qty, linea, turno FROM lg_schedule_daily "
        "WHERE sched_date BETWEEN %s AND %s ORDER BY sched_date, part_no",
        (fecha_ini, fecha_fin),
        fetch="all",
    ) or []
    parts = sorted({str(row.get("part_no") or "").strip() for row in rows})
    raw = _ppy_datos_raw(parts, run_query) if parts else {}
    active_lines = set(_ppy_config_lineas(run_query))
    snapshot = []
    for row in rows:
        part_no = str(row.get("part_no") or "").strip()
        explicit_line = str(row.get("linea") or "").strip().upper()
        raw_line = str((raw.get(part_no) or {}).get("assy_line") or "").strip().upper()
        if explicit_line not in active_lines and raw_line not in active_lines:
            continue
        fecha = row.get("sched_date")
        if isinstance(fecha, datetime):
            fecha = fecha.date()
        snapshot.append(
            {
                "part_no": part_no,
                "sched_date": fecha.isoformat() if isinstance(fecha, date) else str(fecha or ""),
                "qty": int(row.get("sched_qty") or 0),
                "linea": explicit_line or raw_line or None,
                "turno": str(row.get("turno") or "DIA").strip().upper(),
            }
        )
    return snapshot


def _ppy_schedule_changes(propuestas, schedule_snapshot):
    """Compara el Schedule recibido contra el plan final optimizado."""
    base = {
        (item["part_no"], item["sched_date"]): item
        for item in schedule_snapshot
    }
    final = {
        (item["part_no"], str(item["sched_date"])): {
            "part_no": item["part_no"],
            "sched_date": str(item["sched_date"]),
            "qty": int(item["qty"]),
            "linea": str(item.get("linea") or "").strip().upper() or None,
            "turno": str(item.get("turno") or "DIA").strip().upper(),
        }
        for item in propuestas
    }
    changes = []
    for key in sorted(set(base) | set(final), key=lambda value: (value[1], value[0])):
        before = base.get(key)
        after = final.get(key)
        if before and after:
            action = "CONSERVAR" if all(
                before.get(field) == after.get(field)
                for field in ("qty", "linea", "turno")
            ) else "MODIFICAR"
        elif after:
            action = "AGREGAR"
        else:
            action = "ELIMINAR"
        changes.append(
            {
                "accion": action,
                "part_no": key[0],
                "sched_date": key[1],
                "antes_qty": int((before or {}).get("qty") or 0),
                "despues_qty": int((after or {}).get("qty") or 0),
                "antes_linea": (before or {}).get("linea"),
                "despues_linea": (after or {}).get("linea"),
                "turno": (after or before or {}).get("turno"),
            }
        )
    return changes


def _ppy_proposal_hash(
    fecha_ini, fecha_fin, propuestas, excepciones, schedule_snapshot=None
):
    """Huella estable de la salida deterministica para detectar propuestas viejas."""
    payload = {
        "engine": PPY_PROPOSAL_ENGINE_VERSION,
        "date_from": fecha_ini.isoformat(),
        "date_to": fecha_fin.isoformat(),
        "proposals": propuestas,
        "exceptions": excepciones,
        "schedule_snapshot": schedule_snapshot or [],
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _ppy_normalizar_partes_excluidas(values):
    """Normaliza restricciones explícitas de Planning para una propuesta."""
    if values in (None, ""):
        return []
    if isinstance(values, str):
        values = re.split(r"[\s,;]+", values)
    if not isinstance(values, (list, tuple, set)):
        raise ValueError("partes_excluidas debe ser una lista")
    salida = []
    vistas = set()
    for raw in values:
        part_no = str(raw or "").strip().upper()
        if not part_no:
            continue
        if not re.fullmatch(r"[A-Z0-9._-]{3,100}", part_no):
            raise ValueError(f"Numero de parte invalido para excluir: {part_no}")
        if part_no not in vistas:
            vistas.add(part_no)
            salida.append(part_no)
    if len(salida) > 50:
        raise ValueError("Solo se pueden excluir hasta 50 partes por propuesta")
    return salida


def _ppy_partes_excluidas_propuesta(value):
    """Lee la lista persistida; propuestas legacy equivalen a lista vacía."""
    if not value:
        return []
    if isinstance(value, (list, tuple, set)):
        return _ppy_normalizar_partes_excluidas(value)
    try:
        parsed = json.loads(str(value))
    except (TypeError, ValueError):
        parsed = value
    return _ppy_normalizar_partes_excluidas(parsed)


def _ppy_simular_schedule(fecha_ini, fecha_fin, detailed=False, query=None,
                          replanear=True, excluded_parts=None):
    """Propone schedule con MRP y capacidad, sin escribir datos operativos.

    replanear=True (default): toma el Schedule capturado como linea base para
    comparar, pero reconstruye el resultado final del rango. El motor puede
    conservar, modificar, agregar o eliminar renglones para mejorar el plan.
    El Schedule anterior a fecha_ini si cuenta como material ya producido.

    replanear=False: calcula solamente incrementos sobre el Schedule vigente;
    se conserva para consultas de diagnostico, no para la propuesta completa.

    Formula base: I(t) = I(t-1) - P(t) + S(t). Cada faltante se cubre con
    +10 %, caja cerrada, linea permitida y un maximo de 9 h del turno DIA.
    Lo propuesto un dia se propaga a los dias siguientes. Las partes sin
    inventario, UPH, empaque, linea valida o capacidad se reportan como
    excepciones; nunca se convierten silenciosamente en trabajo factible.

    Mantiene los aliases historicos part_no/sched_date/qty para la pantalla.
    Si ``detailed`` es True retorna tambien excepciones estructuradas.
    """
    run_query = query or execute_query
    # La proyeccion va MAS ALLA del rango que se planea: cada dia mira hacia
    # adelante su anticipacion (2 dias de produccion, D1 5) para encontrar el
    # faltante que debe cubrir. Cortarla en fecha_fin dejaba la ventana en cero
    # al pedir un solo dia ("plan para mañana" -> rango 17..17 -> 0 lotes).
    proy_fin = _ppy_sumar_dias_produccion(fecha_fin, PPY_ANTICIPACION_MAX)
    replanear_desde = fecha_ini if replanear else None
    if query is None:
        lineas_activas = _ppy_config_lineas()
        proy = _ppy_proyeccion_rango(fecha_ini, proy_fin,
                                     replanear_desde=replanear_desde)
    else:
        lineas_activas = _ppy_config_lineas(query)
        proy = _ppy_proyeccion_rango(fecha_ini, proy_fin, query,
                                     replanear_desde=replanear_desde)
    # Las partes que no producimos no entran al plan aunque LG las pida.
    excluidas_solicitadas = _ppy_normalizar_partes_excluidas(excluded_parts)
    excluidas = _ppy_partes_excluidas(query) | set(excluidas_solicitadas)
    partes = sorted(p for p in proy if p.upper() not in excluidas)
    if query is None:
        datos_raw = _ppy_datos_raw(partes)
        perm_parte, perm_familia = _ppy_lineas_permitidas_map()
    else:
        datos_raw = _ppy_datos_raw(partes, query)
        perm_parte, perm_familia = _ppy_lineas_permitidas_map(query)

    excepciones = []
    excepciones_vistas = set()

    def agregar_excepcion(part_no, code, message, fecha=None):
        clave = (str(part_no or ""), code, str(fecha or ""), str(message or ""))
        if clave in excepciones_vistas:
            return
        excepciones_vistas.add(clave)
        excepciones.append(
            {
                "part_no": str(part_no or ""),
                "code": code,
                "message": str(message or ""),
                "date": fecha.isoformat() if isinstance(fecha, date) else (str(fecha) if fecha else None),
            }
        )

    omitidas = {
        f"{part_no} (excluida expresamente por Planning)"
        for part_no in excluidas_solicitadas
    }
    for part_no in excluidas_solicitadas:
        agregar_excepcion(
            part_no,
            "PLANNING_EXCLUDED",
            "La parte fue excluida expresamente por Planning para esta propuesta",
        )

    # Una parte con demanda pero sin inventario no aparece en la proyeccion;
    # hacerla visible evita asumir inventario cero y crear un plan inseguro.
    plan_rows = run_query(
        "SELECT DISTINCT TRIM(part_no) AS part_no FROM lg_plan_daily "
        "WHERE plan_date BETWEEN %s AND %s AND plan_qty > 0 ORDER BY TRIM(part_no)",
        (fecha_ini, fecha_fin),
        fetch="all",
    ) or []
    for row in plan_rows:
        part_no = str(row.get("part_no") or "").strip()
        if part_no and part_no.upper() not in excluidas and part_no not in proy:
            agregar_excepcion(
                part_no,
                "INVENTORY_MISSING",
                "La parte tiene demanda LG pero no tiene inventario base en lg_part_inventory",
            )

    info = {}
    for p in partes:
        r = datos_raw.get(p) or {}
        pack = r.get("estandar_pack")
        info[p] = {
            "pack": int(pack) if pack else None,
            "ct": _ppy_parse_ct(r.get("c_t")),
            "uph": _ppy_parse_uph(r.get("uph")),
            "model": r.get("model"),
            "main_sub": r.get("sub_assy"),
            "line": (
                str(r.get("assy_line") or "").strip().upper()[:10]
                or proy[p]["line"]
            ),
            "permitidas": perm_parte.get(p) or perm_familia.get(_ppy_familia(p)),
        }

    # La V2 propone el turno DIA. Un lote confirmado de NOCHE o TIEMPO EXTRA
    # no consume la capacidad diurna. Un confirmado DIA sin linea/UPH queda
    # como excepcion en lugar de contar como cero horas.
    horas_conf = {}
    confirmadas_qty = {}
    rows = run_query(
        "SELECT part_no, plan_date, linea, turno, qty_plan, uph FROM lg_lote_plan "
        "WHERE plan_date BETWEEN %s AND %s AND status = 'CONFIRMADO' "
        "ORDER BY plan_date, linea, part_no",
        (fecha_ini, fecha_fin),
        fetch="all",
    ) or []
    for r in rows:
        turno = str(r.get("turno") or "DIA").strip().upper()
        f = r["plan_date"].date() if isinstance(r["plan_date"], datetime) else r["plan_date"]
        clave_confirmada = (str(r.get("part_no") or "").strip(), f)
        confirmadas_qty[clave_confirmada] = (
            confirmadas_qty.get(clave_confirmada, 0) + int(r.get("qty_plan") or 0)
        )
        if turno != "DIA":
            continue
        linea = str(r.get("linea") or "").strip().upper()
        uph = _ppy_parse_uph(r.get("uph"))
        if not linea or linea not in lineas_activas or not uph:
            agregar_excepcion(
                r.get("part_no"),
                "CONFIRMED_CAPACITY_UNKNOWN",
                "Lote confirmado DIA sin linea activa o UPH; no fue posible descontar su capacidad",
                f,
            )
            continue
        horas_conf[(f, linea)] = horas_conf.get((f, linea), 0.0) + int(r.get("qty_plan") or 0) / uph

    # El schedule ya capturado tambien reserva capacidad. Se descuenta solo
    # la cantidad que no esta representada por lotes confirmados para evitar
    # contar dos veces. Los registros legacy sin linea/UPH bloquean nuevas
    # propuestas de ese dia: la capacidad real no es verificable.
    schedule_rows = run_query(
        "SELECT part_no, sched_date, sched_qty, linea, turno FROM lg_schedule_daily "
        "WHERE sched_date BETWEEN %s AND %s AND sched_qty > 0 "
        "ORDER BY sched_date, linea, part_no",
        (fecha_ini, fecha_fin),
        fetch="all",
    ) or []
    schedule_parts = sorted(
        {str(row.get("part_no") or "").strip() for row in schedule_rows}
    )
    schedule_raw = (
        _ppy_datos_raw(schedule_parts)
        if query is None
        else _ppy_datos_raw(schedule_parts, query)
    )
    capacidad_desconocida = set()
    asignacion_schedule = {}
    # Al replanear, el schedule del rango se rehace: no reserva horas ni
    # condiciona la linea. Solo los lotes CONFIRMADOS siguen mandando.
    if replanear:
        schedule_rows = []
    for row in schedule_rows:
        part_no = str(row.get("part_no") or "").strip()
        fecha = row["sched_date"].date() if isinstance(row["sched_date"], datetime) else row["sched_date"]
        linea = str(row.get("linea") or "").strip().upper()
        if not linea:
            # El schedule importado del Excel no trae linea (solo la propuesta
            # aprobada la escribe), pero la parte si la tiene en raw.assy_line.
            # Sin esto el dia entero entraba en capacidad_desconocida y el motor
            # devolvia 0 lotes aunque hubiera 41 partes con faltante real.
            linea = str(
                (schedule_raw.get(part_no) or {}).get("assy_line") or ""
            ).strip().upper()
        turno = str(row.get("turno") or "DIA").strip().upper()
        asignacion_schedule[(part_no, fecha)] = {
            "qty": int(row.get("sched_qty") or 0),
            "linea": linea or None,
            "turno": turno,
        }
        residual = max(
            int(row.get("sched_qty") or 0) - confirmadas_qty.get((part_no, fecha), 0),
            0,
        )
        if residual <= 0 or turno != "DIA":
            continue
        uph = _ppy_parse_uph((schedule_raw.get(part_no) or {}).get("uph"))
        if not linea or linea not in lineas_activas or not uph:
            capacidad_desconocida.add(fecha)
            agregar_excepcion(
                part_no,
                "SCHEDULE_CAPACITY_UNKNOWN",
                "Schedule existente DIA sin linea activa o UPH; no se puede verificar capacidad",
                fecha,
            )
            continue
        horas_conf[(fecha, linea)] = horas_conf.get((fecha, linea), 0.0) + residual / uph

    propuestas = []
    acum = {p: 0 for p in partes}
    d = fecha_ini
    while d <= fecha_fin:
        if not _ppy_es_dia_produccion(d):
            # ISEMM no produce sabado ni domingo. El faltante de esos dias lo
            # cubre el viernes: su horizonte si cuenta el consumo del sabado.
            d += timedelta(days=1)
            continue
        candidatos = []
        for p in partes:
            # Cuanto mira hacia adelante este dia para encontrar el faltante
            # que debe cubrir: 2 dias de produccion, 5 si es D1. NO se acota
            # con fecha_fin (son cosas distintas: fecha_fin dice hasta que dia
            # se planea, no hasta donde se mira); por eso la proyeccion se pidio
            # hasta proy_fin.
            horizonte_fin = _ppy_sumar_dias_produccion(
                d, _ppy_anticipacion(info[p].get("line"))
            )
            ventana = []
            f = d
            while f <= horizonte_fin:
                base = proy[p]["proj"].get(f)
                if base is not None:
                    ventana.append((f, int(base + acum[p])))
                f += timedelta(days=1)
            # Solo dispara el faltante real: el inventario cae en negativo
            # dentro de la ventana. Una parte que no se consume nunca llega
            # ahi, asi que no se planea aunque su remanente este bajo (un I de
            # 44 son los 60 del plan anterior menos scrap, no un faltante).
            negativos = [(f, valor) for f, valor in ventana if valor < 0]
            if not negativos:
                continue
            primera_falta, inventario_primera_falta = negativos[0]
            # Al reponer, el lote deja el inventario en PPY_REMAIN_IDEAL, no en
            # cero: ese colchon es el que absorbe el scrap.
            falt_total = PPY_REMAIN_IDEAL - min(valor for _f, valor in ventana)
            inventario_hoy = next((valor for f, valor in ventana if f == d), 0)
            candidato = {
                "part_no": p,
                "falt_total": falt_total,
                "falt_hoy": max(0, -inventario_hoy),
                "primera_falta": primera_falta,
                "inventario_antes": inventario_primera_falta,
                **info[p],
            }
            existente = asignacion_schedule.get((p, d))
            if existente:
                # El schedule agregado admite una sola asignacion por parte/dia.
                # Mientras se normaliza a allocations, una ampliacion debe
                # conservar exactamente la linea y turno existentes.
                if existente["turno"] != "DIA" or not existente["linea"]:
                    agregar_excepcion(
                        p,
                        "SCHEDULE_ASSIGNMENT_CONFLICT",
                        "El schedule existente no tiene una asignacion DIA compatible",
                        d,
                    )
                    continue
                candidato["line"] = existente["linea"]
            candidatos.append(candidato)
        if candidatos:
            # Las horas son por GRUPO (M3+M4 comparten 9 h, no 9 h cada una).
            horas_rest = _ppy_horas_iniciales(
                lineas_activas,
                {l: h for (f, l), h in horas_conf.items() if f == d},
            )
            if d in capacidad_desconocida:
                horas_rest = {g: 0.0 for g in horas_rest}
            lotes, fuera = _ppy_armar_lotes(
                candidatos, lineas_activas, horas_rest, estricto=True
            )
            for lote in lotes:
                qty = int(lote["qty"])
                uph = int(lote["uph"])
                inventario_antes = int(lote["inventario_antes"])
                inventario_despues = inventario_antes + qty
                qty_objetivo = _ppy_qty_pack(lote["falt_total"], lote["pack"])
                item_excepciones = []
                if qty < qty_objetivo:
                    item_excepciones.append("PARTIAL_CAPACITY")
                if inventario_despues < 0:
                    item_excepciones.append("SHORTAGE_REMAINS")
                motivo = (
                    "Evitar shortage del plan LG"
                    if d == lote["primera_falta"]
                    else "Adelantar produccion para evitar shortage del plan LG"
                )
                propuesta = {
                    # Contrato explicito recomendado para el planificador.
                    "fecha": d.isoformat(),
                    "linea": str(lote.get("linea") or "").upper(),
                    # grupo de 9 h que consume el lote: D1 puede caer en G1 o
                    # G4, asi que no se puede deducir solo de la linea.
                    "grupo": lote.get("grupo"),
                    "turno": "DIA",
                    "numero_parte": lote["part_no"],
                    "cantidad": qty,
                    "ct": float(lote["ct"]),
                    "uph": uph,
                    "horas_requeridas": round(qty / uph, 4),
                    "inventario_antes": inventario_antes,
                    "inventario_despues": inventario_despues,
                    "fecha_shortage": lote["primera_falta"].isoformat(),
                    "prioridad": 0,
                    "motivo": motivo,
                    "requiere_aprobacion": bool(item_excepciones),
                    "excepciones": item_excepciones,
                    "model": lote.get("model"),
                    "main_sub": lote.get("main_sub"),
                    "pack_size": int(lote["pack"]),
                    "faltante": int(lote["falt_total"]),
                    # Aliases que consume la pantalla Proyeccion actual.
                    "part_no": lote["part_no"],
                    "sched_date": d.isoformat(),
                    "qty": qty,
                }
                propuestas.append(propuesta)
                acum[lote["part_no"]] += qty
            for item in fuera:
                texto = item["part_no"] + " (" + item["motivo"] + ")"
                omitidas.add(texto)
                for code in _ppy_omission_codes(item["motivo"]):
                    agregar_excepcion(item["part_no"], code, item["motivo"], d)
        d += timedelta(days=1)

    for prioridad, propuesta in enumerate(propuestas, start=1):
        propuesta["prioridad"] = prioridad
    excepciones.sort(key=lambda item: (item.get("date") or "", item["part_no"], item["code"]))
    if detailed:
        return propuestas, sorted(omitidas), excepciones
    return propuestas, sorted(omitidas)


class PPYProposalStaleError(ValueError):
    """La propuesta ya no corresponde a los datos actuales del MES."""


def _ppy_attach_schedule_actual(propuestas, fecha_ini, fecha_fin, query=None):
    """Agrega el schedule base que se usara para aplicar de forma idempotente."""
    if not propuestas:
        return propuestas
    partes = sorted({p["part_no"] for p in propuestas})
    placeholders = ", ".join(["%s"] * len(partes))
    run_query = query or execute_query
    rows = run_query(
        "SELECT part_no, sched_date, sched_qty, linea, turno FROM lg_schedule_daily "
        f"WHERE sched_date BETWEEN %s AND %s AND part_no IN ({placeholders}) "
        "ORDER BY part_no, sched_date",
        tuple([fecha_ini, fecha_fin] + partes),
        fetch="all",
    ) or []
    actual = {}
    for row in rows:
        fecha = row["sched_date"].date() if isinstance(row["sched_date"], datetime) else row["sched_date"]
        actual[(row["part_no"], fecha.isoformat())] = {
            "qty": int(row.get("sched_qty") or 0),
            "linea": str(row.get("linea") or "").strip().upper() or None,
            "turno": str(row.get("turno") or "DIA").strip().upper(),
        }
    for propuesta in propuestas:
        base = actual.get((propuesta["part_no"], propuesta["sched_date"])) or {}
        propuesta["sched_actual"] = int(base.get("qty") or 0)
        propuesta["sched_actual_linea"] = base.get("linea")
        propuesta["sched_actual_turno"] = base.get("turno")
    return propuestas


def _ppy_resumen_lineas(propuestas):
    resumen = {}
    for item in propuestas:
        clave = (item["fecha"], item["linea"], item["turno"])
        fila = resumen.setdefault(
            clave,
            {
                "fecha": item["fecha"],
                "linea": item["linea"],
                "turno": item["turno"],
                "horas": 0.0,
                "horas_max": PPY_HORAS_TURNO,
                "lotes": 0,
                "cantidad": 0,
            },
        )
        fila["horas"] += float(item.get("horas_requeridas") or 0)
        fila["lotes"] += 1
        fila["cantidad"] += int(item.get("qty") or 0)
    salida = []
    for clave in sorted(resumen):
        fila = resumen[clave]
        fila["horas"] = round(fila["horas"], 2)
        fila["excede"] = fila["horas"] > PPY_HORAS_TURNO + 1e-6
        salida.append(fila)
    return salida


def _ppy_crear_propuesta(
    fecha_ini,
    fecha_fin,
    usuario,
    *,
    source="UI",
    objective=None,
    excluded_parts=None,
):
    """Calcula y persiste un borrador; no modifica schedule ni lotes."""
    excluded_parts = _ppy_normalizar_partes_excluidas(excluded_parts)
    propuestas, omitidas, excepciones = _ppy_simular_schedule(
        fecha_ini, fecha_fin, detailed=True, excluded_parts=excluded_parts
    )
    schedule_snapshot = _ppy_schedule_snapshot(fecha_ini, fecha_fin)
    _ppy_attach_schedule_actual(propuestas, fecha_ini, fecha_fin)
    schedule_changes = _ppy_schedule_changes(propuestas, schedule_snapshot)
    input_hash = _ppy_proposal_hash(
        fecha_ini,
        fecha_fin,
        propuestas,
        excepciones,
        schedule_snapshot,
    )
    public_id = str(uuid.uuid4())
    source = str(source or "UI").strip().upper()[:20] or "UI"
    objective = str(objective or "").strip()[:500] or None

    conn = get_pooled_connection()
    if conn is None:
        raise RuntimeError("No fue posible obtener conexion MySQL.")
    cursor = get_dict_cursor(conn)
    try:
        conn.autocommit(False)
        cursor.execute(
            "UPDATE lg_plan_proposals SET status='SUPERSEDED', reviewed_by=%s, "
            "reviewed_at=NOW() WHERE created_by=%s AND source=%s AND status='DRAFT'",
            (usuario, usuario, source),
        )
        cursor.execute(
            "INSERT INTO lg_plan_proposals "
            "(public_id, version, date_from, date_to, objective, excluded_parts_json, source, "
            "engine_version, input_hash, status, total_items, total_qty, "
            "omitted_count, created_by) "
            "VALUES (%s, 1, %s, %s, %s, %s, %s, %s, %s, 'DRAFT', %s, %s, %s, %s)",
            (
                public_id,
                fecha_ini,
                fecha_fin,
                objective,
                json.dumps(excluded_parts, ensure_ascii=False),
                source,
                PPY_PROPOSAL_ENGINE_VERSION,
                input_hash,
                len(propuestas),
                sum(int(item["qty"]) for item in propuestas),
                len(omitidas),
                usuario,
            ),
        )
        proposal_id = cursor.lastrowid
        filas = []
        for item in propuestas:
            item_public_id = str(uuid.uuid4())
            item["item_id"] = item_public_id
            filas.append(
                (
                    item_public_id,
                    proposal_id,
                    int(item["prioridad"]),
                    item["part_no"],
                    item["sched_date"],
                    item["linea"],
                    item["turno"],
                    int(item["qty"]),
                    float(item["ct"]),
                    int(item["uph"]),
                    float(item["horas_requeridas"]),
                    int(item["pack_size"]),
                    int(item.get("sched_actual") or 0),
                    item.get("sched_actual_linea"),
                    item.get("sched_actual_turno"),
                    int(item["inventario_antes"]),
                    int(item["inventario_despues"]),
                    item["fecha_shortage"],
                    int(item["prioridad"]),
                    item.get("motivo"),
                    1 if item.get("requiere_aprobacion") else 0,
                    json.dumps(item.get("excepciones") or [], ensure_ascii=False),
                )
            )
        if filas:
            cursor.executemany(
                "INSERT INTO lg_plan_proposal_items "
                "(public_id, proposal_id, sequence_no, part_no, sched_date, linea, turno, "
                "qty_proposed, ct, uph, hours_required, pack_size, base_sched_qty, "
                "base_sched_linea, base_sched_turno, "
                "inventory_before, inventory_after, shortage_date, priority_no, reason, "
                "requires_approval, exceptions_json) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, "
                "%s, %s, %s, %s, %s, %s, %s, %s)",
                filas,
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.autocommit(True)
        conn.close()

    return {
        "proposal_id": public_id,
        "version": 1,
        "engine_version": PPY_PROPOSAL_ENGINE_VERSION,
        "date_from": fecha_ini.isoformat(),
        "date_to": fecha_fin.isoformat(),
        "proposals": propuestas,
        "total_qty": sum(int(item["qty"]) for item in propuestas),
        "partes": len({item["part_no"] for item in propuestas}),
        "omitidas": omitidas[:20],
        "omitidas_count": len(omitidas),
        "excluded_parts": excluded_parts,
        "exceptions": excepciones,
        "requires_approval_count": sum(
            1 for item in propuestas if item.get("requiere_aprobacion")
        ),
        "line_summary": _ppy_resumen_lineas(propuestas),
        "schedule_changes": schedule_changes,
        "schedule_change_summary": {
            action: sum(1 for item in schedule_changes if item["accion"] == action)
            for action in ("CONSERVAR", "MODIFICAR", "AGREGAR", "ELIMINAR")
        },
        "input_hash": input_hash,
    }


def _ppy_mark_proposal_pending(public_id, usuario):
    execute_query(
        "UPDATE lg_plan_proposals SET status = 'PENDING_CONFIRMATION' "
        "WHERE public_id = %s AND created_by = %s AND status = 'DRAFT'",
        (public_id, usuario),
    )


def _ppy_reject_proposal(
    public_id,
    usuario,
    reason_code=None,
    reason_text=None,
    *,
    version=1,
):
    """Registra el rechazo sin borrar la propuesta: sera dato de aprendizaje."""
    conn = get_pooled_connection()
    if conn is None:
        raise RuntimeError("No fue posible obtener conexion MySQL.")
    cursor = get_dict_cursor(conn)
    try:
        conn.autocommit(False)
        cursor.execute(
            "SELECT * FROM lg_plan_proposals WHERE public_id = %s "
            "AND created_by = %s FOR UPDATE",
            (public_id, usuario),
        )
        row = cursor.fetchone()
        if not row:
            raise ValueError("Propuesta no encontrada")
        if row["status"] == "REJECTED":
            conn.rollback()
            return {"proposal_id": public_id, "status": "REJECTED", "already_rejected": True}
        if row["status"] == "APPLIED":
            raise ValueError("La propuesta ya fue aplicada y no puede rechazarse")
        if row["status"] not in ("DRAFT", "PENDING_CONFIRMATION"):
            raise ValueError(f"La propuesta no puede rechazarse en estado {row['status']}")
        if int(row.get("version") or 1) != int(version or 1):
            raise PPYProposalStaleError(
                "La version de la propuesta cambio; genera una propuesta nueva"
            )
        cursor.execute(
            "UPDATE lg_plan_proposals SET status='REJECTED', reviewed_by=%s, "
            "reviewed_at=NOW() WHERE id=%s",
            (usuario, row["id"]),
        )
        cursor.execute(
            "INSERT INTO lg_plan_proposal_feedback "
            "(proposal_id, username, action, reason_code, reason_text) "
            "VALUES (%s, %s, 'REJECTED', %s, %s)",
            (
                row["id"],
                usuario,
                str(reason_code or "USER_REJECTED")[:80],
                str(reason_text or "").strip()[:1000] or None,
            ),
        )
        conn.commit()
        return {"proposal_id": public_id, "status": "REJECTED"}
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.autocommit(True)
        conn.close()


def _ppy_applied_result(public_id, proposal_id, query=None):
    """Reconstruye el resultado final para que los reintentos sean exactos."""
    run_query = query or execute_query
    row = run_query(
        "SELECT COUNT(CASE WHEN included_final = 1 THEN 1 END) AS aplicadas, "
        "COUNT(CASE WHEN included_final = 0 THEN 1 END) AS excluidas, "
        "COUNT(CASE WHEN included_final = 1 AND "
        "(qty_final <> qty_proposed OR final_sched_date <> sched_date OR "
        "final_linea <> linea OR final_turno <> turno) THEN 1 END) AS modificadas, "
        "COALESCE(SUM(CASE WHEN included_final = 1 THEN qty_final ELSE 0 END), 0) "
        "AS total_qty FROM lg_plan_proposal_items WHERE proposal_id = %s",
        (proposal_id,),
        fetch="one",
    ) or {}
    return {
        "proposal_id": public_id,
        "status": "APPLIED",
        "aplicadas": int(row.get("aplicadas") or 0),
        "modificadas": int(row.get("modificadas") or 0),
        "excluidas": int(row.get("excluidas") or 0),
        "total_qty": int(row.get("total_qty") or 0),
        "already_applied": True,
    }


def _ppy_cursor_query(cursor, sql, params=(), fetch=None):
    """Adapta un cursor transaccional al contrato pequeño de execute_query."""
    cursor.execute(sql, params or ())
    if fetch == "one":
        return cursor.fetchone()
    if fetch == "all":
        return cursor.fetchall()
    return cursor.rowcount


def _ppy_aplicar_propuesta(public_id, usuario, *, version=1, items=None):
    """Aprueba un borrador, valida frescura/capacidad y actualiza el schedule.

    La propuesta se aplica una sola vez. Un reintento sobre una propuesta ya
    aplicada devuelve el resultado anterior sin volver a sumar cantidades.
    """
    header = execute_query(
        "SELECT * FROM lg_plan_proposals WHERE public_id=%s AND created_by=%s",
        (public_id, usuario),
        fetch="one",
    )
    if not header:
        raise ValueError("Propuesta no encontrada")
    if header.get("status") == "APPLIED":
        return _ppy_applied_result(public_id, header["id"])
    if header.get("status") not in ("DRAFT", "PENDING_CONFIRMATION"):
        raise ValueError(f"La propuesta no puede aplicarse en estado {header.get('status')}")
    if int(header.get("version") or 1) != int(version or 1):
        raise PPYProposalStaleError("La version de la propuesta cambio; vuelve a revisarla")

    conn = get_pooled_connection()
    if conn is None:
        raise RuntimeError("No fue posible obtener conexion MySQL.")
    cursor = get_dict_cursor(conn)
    try:
        conn.autocommit(False)
        # Todas las lecturas que forman la huella y la validacion ocurren en
        # el mismo snapshot serializable que el write. Esto elimina la ventana
        # entre "recalcular" y "aplicar".
        cursor.execute("SET TRANSACTION ISOLATION LEVEL SERIALIZABLE")
        cursor.execute(
            "SELECT * FROM lg_plan_proposals WHERE id=%s FOR UPDATE",
            (header["id"],),
        )
        locked = cursor.fetchone()
        if locked["status"] == "APPLIED":
            result = _ppy_applied_result(
                public_id,
                locked["id"],
                query=lambda sql, params=(), fetch=None: _ppy_cursor_query(
                    cursor, sql, params, fetch
                ),
            )
            conn.rollback()
            return result
        if locked["status"] not in ("DRAFT", "PENDING_CONFIRMATION"):
            raise ValueError(f"La propuesta no puede aplicarse en estado {locked['status']}")
        if int(locked.get("version") or 1) != int(version or 1):
            raise PPYProposalStaleError(
                "La version de la propuesta cambio; vuelve a revisarla"
            )
        if locked.get("engine_version") != PPY_PROPOSAL_ENGINE_VERSION:
            raise PPYProposalStaleError(
                "El motor de planeacion cambio; genera una propuesta nueva"
            )

        cursor.execute(
            "SELECT * FROM lg_plan_proposal_items WHERE proposal_id=%s "
            "ORDER BY sequence_no FOR UPDATE",
            (locked["id"],),
        )
        stored_items = cursor.fetchall() or []

        def tx_query(sql, params=(), fetch=None):
            return _ppy_cursor_query(cursor, sql, params, fetch)

        fecha_ini = (
            locked["date_from"].date()
            if isinstance(locked["date_from"], datetime)
            else locked["date_from"]
        )
        fecha_fin = (
            locked["date_to"].date()
            if isinstance(locked["date_to"], datetime)
            else locked["date_to"]
        )
        actuales, _omitidas, excepciones = _ppy_simular_schedule(
            fecha_ini,
            fecha_fin,
            detailed=True,
            query=tx_query,
            excluded_parts=_ppy_partes_excluidas_propuesta(
                locked.get("excluded_parts_json")
            ),
        )
        _ppy_attach_schedule_actual(
            actuales, fecha_ini, fecha_fin, query=tx_query
        )
        schedule_snapshot = _ppy_schedule_snapshot(
            fecha_ini, fecha_fin, query=tx_query
        )
        hash_actual = _ppy_proposal_hash(
            fecha_ini,
            fecha_fin,
            actuales,
            excepciones,
            schedule_snapshot,
        )
        if hash_actual != locked.get("input_hash"):
            cursor.execute(
                "UPDATE lg_plan_proposals SET status='STALE', reviewed_by=%s, "
                "reviewed_at=NOW() WHERE id=%s",
                (usuario, locked["id"]),
            )
            conn.commit()
            raise PPYProposalStaleError(
                "El plan, inventario, RAW, capacidad o schedule cambio; "
                "genera una propuesta nueva"
            )
        provided = {}
        for raw in items or []:
            if not isinstance(raw, dict):
                raise ValueError("Cada item de propuesta debe ser un objeto")
            item_public_id = str(raw.get("item_id") or "").strip()
            if not item_public_id or item_public_id in provided:
                raise ValueError("Items de propuesta invalidos o duplicados")
            provided[item_public_id] = raw
        valid_ids = {row["public_id"] for row in stored_items}
        if set(provided) - valid_ids:
            raise ValueError("La propuesta contiene items que no le pertenecen")

        lineas_activas = _ppy_config_lineas(tx_query)
        permitidas_parte, permitidas_familia = _ppy_lineas_permitidas_map(tx_query)
        finales = []
        horas = {}
        for row in stored_items:
            change = provided.get(row["public_id"], {})
            included_raw = change.get("included", True)
            if not isinstance(included_raw, bool):
                raise ValueError(
                    f"included debe ser booleano para {row['part_no']}"
                )
            included = included_raw
            original_date = (
                row["sched_date"].date()
                if isinstance(row["sched_date"], datetime)
                else row["sched_date"]
            )
            if included:
                if "sched_date" in change:
                    sched_date = _ppy_parse_fecha(change.get("sched_date"))
                    if sched_date is None:
                        raise ValueError(
                            f"Fecha invalida para {row['part_no']}"
                        )
                else:
                    sched_date = original_date
                linea = str(
                    change.get("linea") or row["linea"] or ""
                ).strip().upper()[:10]
                turno = str(
                    change.get("turno") or row["turno"] or "DIA"
                ).strip().upper()[:15]
                qty_raw = change.get("qty", row["qty_proposed"])
                if isinstance(qty_raw, bool):
                    raise ValueError(f"Cantidad invalida para {row['part_no']}")
                try:
                    qty = int(qty_raw)
                except (TypeError, ValueError):
                    raise ValueError(f"Cantidad invalida para {row['part_no']}")
            else:
                # Una exclusion solo cambia included. No se aceptan campos
                # arbitrarios que contaminen el dataset de aprendizaje.
                sched_date = original_date
                linea = str(row["linea"] or "").strip().upper()
                turno = str(row["turno"] or "DIA").strip().upper()
                qty = int(row["qty_proposed"])

            if included:
                if not fecha_ini <= sched_date <= fecha_fin:
                    raise ValueError(f"Fecha fuera del rango de la propuesta para {row['part_no']}")
                if qty <= 0:
                    raise ValueError(f"Cantidad invalida para {row['part_no']}")
                pack = int(row.get("pack_size") or 0)
                if pack <= 0 or qty % pack != 0:
                    raise ValueError(
                        f"{row['part_no']}: la cantidad debe ser multiplo del empaque {pack or 'N/D'}"
                    )
                if linea not in lineas_activas:
                    raise ValueError(f"{row['part_no']}: linea {linea or 'N/D'} no activa")
                permitidas = (
                    permitidas_parte.get(row["part_no"])
                    or permitidas_familia.get(_ppy_familia(row["part_no"]))
                )
                if permitidas and linea not in permitidas:
                    raise ValueError(
                        f"{row['part_no']}: linea {linea} no permitida ({','.join(permitidas)})"
                    )
                if turno not in PPY_TURNOS:
                    raise ValueError(f"Turno invalido para {row['part_no']}")
                uph = int(row.get("uph") or 0)
                ct = _ppy_parse_ct(row.get("ct"))
                if not ct:
                    raise ValueError(f"{row['part_no']}: CT requerido")
                if uph <= 0:
                    raise ValueError(f"{row['part_no']}: UPH requerido")
                clave_horas = (sched_date, linea, turno)
                horas[clave_horas] = horas.get(clave_horas, 0.0) + qty / uph

            before = {
                "included": True,
                "sched_date": row["sched_date"].isoformat(),
                "linea": row["linea"],
                "turno": row["turno"],
                "qty": int(row["qty_proposed"]),
            }
            after = {
                "included": included,
                "sched_date": sched_date.isoformat(),
                "linea": linea,
                "turno": turno,
                "qty": qty,
            }
            action = "ACCEPTED" if before == after else ("EXCLUDED" if not included else "MODIFIED")
            finales.append((row, after, action, before))

        # Capacidad ya confirmada por fecha/linea/turno tambien cuenta.
        requested_date_turns = {(key[0], key[2]) for key in horas}
        cursor.execute(
            "SELECT part_no, plan_date, linea, turno, qty_plan, uph FROM lg_lote_plan "
            "WHERE plan_date BETWEEN %s AND %s AND status='CONFIRMADO' FOR UPDATE",
            (fecha_ini, fecha_fin),
        )
        confirmadas_qty = {}
        for row in cursor.fetchall() or []:
            linea = str(row.get("linea") or "").strip().upper()
            turno = str(row.get("turno") or "DIA").strip().upper()
            uph = _ppy_parse_uph(row.get("uph"))
            fecha = row["plan_date"].date() if isinstance(row["plan_date"], datetime) else row["plan_date"]
            clave_confirmada = (str(row.get("part_no") or "").strip(), fecha)
            confirmadas_qty[clave_confirmada] = (
                confirmadas_qty.get(clave_confirmada, 0) + int(row.get("qty_plan") or 0)
            )
            if not linea or not uph:
                if (fecha, turno) in requested_date_turns:
                    raise ValueError(
                        f"Lote confirmado {row.get('part_no') or ''} sin linea o UPH; "
                        "capacidad no verificable"
                    )
                continue
            clave = (fecha, linea, turno)
            horas[clave] = horas.get(clave, 0.0) + int(row.get("qty_plan") or 0) / uph

        # El Schedule vigente es la linea base que esta propuesta REEMPLAZA;
        # sumarlo aqui duplicaria sus horas. Los lotes CONFIRMADOS de arriba
        # si son inamovibles y consumen capacidad adicional al plan nuevo.
        excedidas = [
            (clave, total)
            for clave, total in horas.items()
            if total > PPY_HORAS_TURNO + 1e-6
        ]
        if excedidas:
            detalle = ", ".join(
                f"{f.isoformat()} {linea}/{turno}: {total:.2f} h"
                for (f, linea, turno), total in excedidas
            )
            raise ValueError(f"La propuesta excede {PPY_HORAS_TURNO:g} h: {detalle}")

        cursor.execute(
            "SELECT part_no, sched_date, sched_qty, linea, turno "
            "FROM lg_schedule_daily "
            "WHERE sched_date BETWEEN %s AND %s FOR UPDATE",
            (fecha_ini, fecha_fin),
        )
        schedule_actual = {}
        managed_schedule_keys = {
            (item["part_no"], _ppy_parse_fecha(item["sched_date"]))
            for item in schedule_snapshot
        }
        proposal_parts = {row["part_no"] for row in stored_items}
        for row in cursor.fetchall() or []:
            fecha = row["sched_date"].date() if isinstance(row["sched_date"], datetime) else row["sched_date"]
            key = (row["part_no"], fecha)
            if key not in managed_schedule_keys and row["part_no"] not in proposal_parts:
                continue
            schedule_actual[key] = {
                "qty": int(row.get("sched_qty") or 0),
                "linea": str(row.get("linea") or "").strip().upper() or None,
                "turno": str(row.get("turno") or "DIA").strip().upper(),
            }

        # Si cambio una celda original entre preparar y aplicar, no sumar a
        # ciegas: la propuesta se invalida.
        for row in stored_items:
            fecha = row["sched_date"].date() if isinstance(row["sched_date"], datetime) else row["sched_date"]
            current = schedule_actual.get((row["part_no"], fecha)) or {
                "qty": 0,
                "linea": None,
                "turno": None,
            }
            base = {
                "qty": int(row.get("base_sched_qty") or 0),
                "linea": str(row.get("base_sched_linea") or "").strip().upper() or None,
                "turno": (
                    str(row.get("base_sched_turno") or "").strip().upper() or None
                ),
            }
            if current != base:
                raise PPYProposalStaleError(
                    f"Cambio el schedule de {row['part_no']} para {fecha.isoformat()}; regenera la propuesta"
                )

        incrementos = {}
        for row, after, _action, _before in finales:
            if after["included"]:
                clave = (row["part_no"], _ppy_parse_fecha(after["sched_date"]))
                existente = incrementos.get(clave)
                if existente and (
                    existente["linea"] != after["linea"]
                    or existente["turno"] != after["turno"]
                ):
                    raise ValueError(
                        f"{row['part_no']}: una misma fecha no puede tener dos lineas o turnos "
                        "en el schedule actual"
                    )
                if not existente:
                    existente = {
                        "qty": 0,
                        "linea": after["linea"],
                        "turno": after["turno"],
                    }
                    incrementos[clave] = existente
                existente["qty"] += int(after["qty"])
        target_schedule = {}
        for key, sched in incrementos.items():
            target_schedule[key] = {
                "qty": int(confirmadas_qty.get(key, 0)) + sched["qty"],
                "linea": sched["linea"],
                "turno": sched["turno"],
            }
        for key, current in schedule_actual.items():
            if key not in target_schedule and int(confirmadas_qty.get(key, 0)) > 0:
                target_schedule[key] = {
                    "qty": int(confirmadas_qty[key]),
                    "linea": current.get("linea"),
                    "turno": current.get("turno"),
                }
        change_counts = {
            "CONSERVAR": 0, "MODIFICAR": 0, "AGREGAR": 0, "ELIMINAR": 0
        }
        for key in set(schedule_actual) | set(target_schedule):
            before = schedule_actual.get(key)
            after = target_schedule.get(key)
            action = (
                "ELIMINAR" if after is None else
                "AGREGAR" if before is None else
                "CONSERVAR" if before == after else
                "MODIFICAR"
            )
            change_counts[action] += 1

        # El resultado es un Schedule final completo. Lo que ya no aparece se
        # elimina; lo modificado se reemplaza y lo nuevo se agrega. Un lote ya
        # CONFIRMADO nunca se borra: su cantidad queda como base inamovible.
        for (part_no, sched_date), current in schedule_actual.items():
            if (part_no, sched_date) in incrementos:
                continue
            qty_confirmada = int(confirmadas_qty.get((part_no, sched_date), 0))
            if qty_confirmada > 0:
                cursor.execute(
                    "UPDATE lg_schedule_daily SET sched_qty=%s, proposal_public_id=%s, "
                    "updated_by=%s WHERE part_no=%s AND sched_date=%s",
                    (qty_confirmada, public_id, usuario, part_no, sched_date),
                )
            else:
                cursor.execute(
                    "DELETE FROM lg_schedule_daily WHERE part_no=%s AND sched_date=%s",
                    (part_no, sched_date),
                )

        cantidades_finales = {}
        for (part_no, sched_date), sched in incrementos.items():
            nueva_qty = int(confirmadas_qty.get((part_no, sched_date), 0)) + sched["qty"]
            cantidades_finales[(part_no, sched_date)] = nueva_qty
            cursor.execute(
                "INSERT INTO lg_schedule_daily "
                "(part_no, sched_date, sched_qty, linea, turno, proposal_public_id, updated_by) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s) ON DUPLICATE KEY UPDATE "
                "sched_qty=VALUES(sched_qty), linea=VALUES(linea), turno=VALUES(turno), "
                "proposal_public_id=VALUES(proposal_public_id), updated_by=VALUES(updated_by)",
                (
                    part_no,
                    sched_date,
                    nueva_qty,
                    sched["linea"],
                    sched["turno"],
                    public_id,
                    usuario,
                ),
            )

        for row, after, action, before in finales:
            cursor.execute(
                "UPDATE lg_plan_proposal_items SET qty_final=%s, included_final=%s, "
                "final_sched_date=%s, final_linea=%s, final_turno=%s WHERE id=%s",
                (
                    int(after["qty"]),
                    1 if after["included"] else 0,
                    after["sched_date"],
                    after["linea"],
                    after["turno"],
                    row["id"],
                ),
            )
            cursor.execute(
                "INSERT INTO lg_plan_proposal_feedback "
                "(proposal_id, item_id, username, action, before_json, after_json) "
                "VALUES (%s, %s, %s, %s, %s, %s)",
                (
                    locked["id"],
                    row["id"],
                    usuario,
                    action,
                    json.dumps(before, ensure_ascii=False),
                    json.dumps(after, ensure_ascii=False),
                ),
            )
        cursor.execute(
            "UPDATE lg_plan_proposals SET status='APPLIED', reviewed_by=%s, "
            "reviewed_at=NOW(), applied_at=NOW() WHERE id=%s",
            (usuario, locked["id"]),
        )
        cursor.execute(
            "INSERT INTO lg_plan_proposal_feedback "
            "(proposal_id, username, action, reason_code) "
            "VALUES (%s, %s, 'APPLIED', 'USER_CONFIRMED')",
            (locked["id"], usuario),
        )
        conn.commit()
        aplicadas = sum(1 for _row, after, _action, _before in finales if after["included"])
        modificadas = sum(1 for _row, _after, action, _before in finales if action == "MODIFIED")
        excluidas = sum(1 for _row, _after, action, _before in finales if action == "EXCLUDED")
        return {
            "proposal_id": public_id,
            "status": "APPLIED",
            "aplicadas": aplicadas,
            "modificadas": modificadas,
            "excluidas": excluidas,
            "total_qty": sum(cantidades_finales.values()),
            "schedule_change_summary": change_counts,
        }
    except PPYProposalStaleError:
        conn.rollback()
        try:
            cursor.execute(
                "UPDATE lg_plan_proposals SET status='STALE', reviewed_by=%s, "
                "reviewed_at=NOW() WHERE id=%s AND status IN ('DRAFT','PENDING_CONFIRMATION')",
                (usuario, header["id"]),
            )
            conn.commit()
        except Exception:
            conn.rollback()
        raise
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.autocommit(True)
        conn.close()


@bp.route("/api/part-planning/schedule/proponer", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PROY_PERMISO_BOTON)
def api_pp_schedule_proponer():
    """Calcula un borrador revisable; no modifica schedule ni lotes."""
    try:
        data = request.get_json(silent=True) or {}
        hoy = date.today()
        date_from = _ppy_parse_fecha(data.get("date_from")) or hoy
        date_to = _ppy_parse_fecha(data.get("date_to")) or (date_from + timedelta(days=6))
        if date_to < date_from:
            date_from, date_to = date_to, date_from
        date_from = max(date_from, hoy)  # no se propone produccion en el pasado
        if date_to < date_from:
            return jsonify({"success": False, "error": "El rango ya paso; no hay dias por proponer."}), 400
        if (date_to - date_from).days > PP_MAX_RANGO_DIAS:
            return jsonify({"success": False, "error": f"Rango maximo {PP_MAX_RANGO_DIAS} dias"}), 400

        usuario = session.get("usuario") or "SISTEMA"
        propuesta = _ppy_crear_propuesta(
            date_from,
            date_to,
            usuario,
            source="UI",
            objective=data.get("objective"),
            excluded_parts=data.get("excluded_parts"),
        )
        return jsonify({"success": True, **propuesta})
    except Exception as e:
        logger.error("Error en api_pp_schedule_proponer: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/part-planning/schedule/proponer/aplicar", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PROY_PERMISO_BOTON)
def api_pp_schedule_aplicar():
    """Aprueba un borrador persistido y registra aceptaciones/correcciones."""
    try:
        data = request.get_json(silent=True) or {}
        public_id = str(data.get("proposal_id") or "").strip()
        if not public_id:
            return jsonify({"success": False, "error": "proposal_id requerido"}), 400
        items = data.get("items")
        if items is not None and not isinstance(items, list):
            return jsonify({"success": False, "error": "items debe ser una lista"}), 400
        if items is not None and len(items) > 5000:
            return jsonify({"success": False, "error": "Demasiados items"}), 400
        usuario = session.get("usuario") or "SISTEMA"
        result = _ppy_aplicar_propuesta(
            public_id,
            usuario,
            version=data.get("version") or 1,
            items=items,
        )
        return jsonify({"success": True, **result})
    except PPYProposalStaleError as e:
        return jsonify({"success": False, "error": str(e), "code": "PROPOSAL_STALE"}), 409
    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        logger.error("Error en api_pp_schedule_aplicar: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/part-planning/schedule/proponer/rechazar", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PROY_PERMISO_BOTON)
def api_pp_schedule_rechazar():
    """Conserva el rechazo y su motivo como feedback para el planificador."""
    try:
        data = request.get_json(silent=True) or {}
        public_id = str(data.get("proposal_id") or "").strip()
        if not public_id:
            return jsonify({"success": False, "error": "proposal_id requerido"}), 400
        usuario = session.get("usuario") or "SISTEMA"
        result = _ppy_reject_proposal(
            public_id,
            usuario,
            reason_code=data.get("reason_code"),
            reason_text=data.get("reason_text"),
            version=data.get("version") or 1,
        )
        return jsonify({"success": True, **result})
    except PPYProposalStaleError as e:
        return jsonify({"success": False, "error": str(e), "code": "PROPOSAL_STALE"}), 409
    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        logger.error("Error en api_pp_schedule_rechazar: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/part-planning/inventory/field", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PROY_PERMISO_BOTON)
def api_pp_inventory_field():
    """Captura/edita un campo de inventario (LGEMM, ISEMM, SVC, ...) de una parte.

    Al editar manualmente, la fecha de referencia del calculo de I pasa al
    lunes de la semana actual (el inventario capturado es de esta semana).
    """
    try:
        data = request.get_json(silent=True) or {}
        part_no = (data.get("part_no") or "").strip()
        campo = (data.get("field") or "").strip().lower()
        if not part_no:
            return jsonify({"success": False, "error": "part_no requerido"}), 400
        if campo not in PP_INV_FIELDS:
            return jsonify({"success": False, "error": f"Campo invalido: {campo}"}), 400

        valor_raw = data.get("value")
        try:
            valor = int(valor_raw) if valor_raw not in (None, "") else 0
        except (TypeError, ValueError):
            return jsonify({"success": False, "error": "Valor invalido"}), 400

        hoy = date.today()
        lunes = hoy - timedelta(days=hoy.weekday())
        # campo esta validado contra PP_INV_FIELDS (whitelist)
        execute_query(
            f"INSERT INTO lg_part_inventory (part_no, {campo}, ref_date) "
            "VALUES (%s, %s, %s) "
            f"ON DUPLICATE KEY UPDATE {campo}=VALUES({campo}), ref_date=VALUES(ref_date)",
            (part_no, valor, lunes),
        )
        return jsonify(
            {
                "success": True,
                "part_no": part_no,
                "field": campo,
                "value": valor,
                "ref_date": lunes.isoformat(),
            }
        )
    except Exception as e:
        logger.error("Error en api_pp_inventory_field: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/part-planning/plan", methods=["GET"])
@login_requerido
def api_pp_plan():
    """Plan importado pivotado: filas = partes, columnas = fechas del rango."""
    try:
        hoy = date.today()
        lunes = hoy - timedelta(days=hoy.weekday())
        try:
            date_from = datetime.strptime(
                request.args.get("date_from") or lunes.isoformat(), "%Y-%m-%d"
            ).date()
            date_to = datetime.strptime(
                request.args.get("date_to")
                or (lunes + timedelta(days=6)).isoformat(),
                "%Y-%m-%d",
            ).date()
        except ValueError:
            return jsonify({"success": False, "error": "Fechas invalidas"}), 400
        if date_to < date_from:
            date_from, date_to = date_to, date_from
        if (date_to - date_from).days > PP_MAX_RANGO_DIAS:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": f"Rango maximo {PP_MAX_RANGO_DIAS} dias",
                    }
                ),
                400,
            )

        part = (request.args.get("part") or "").strip()
        page = max(request.args.get("page", type=int) or 1, 1)
        page_size = min(max(request.args.get("page_size", type=int) or 50, 1), 200)
        # Solo faltantes: partes cuya proyeccion I cae en negativo en el rango.
        # Se calcula sobre todas las partes del rango y se pagina en Python.
        only_shortage = (request.args.get("only_shortage") or "").strip().lower() in (
            "1", "true", "on",
        )

        where = "WHERE plan_date BETWEEN %s AND %s"
        params = [date_from, date_to]
        if part:
            where += " AND part_no LIKE %s"
            params.append(f"%{part}%")

        total_row = execute_query(
            f"SELECT COUNT(DISTINCT part_no) AS c FROM lg_plan_daily {where}",
            tuple(params),
            fetch="one",
        )
        total_parts = int((total_row or {}).get("c") or 0)

        if only_shortage:
            partes = execute_query(
                f"SELECT DISTINCT part_no FROM lg_plan_daily {where} ORDER BY part_no",
                tuple(params),
                fetch="all",
            ) or []
        else:
            partes = execute_query(
                f"SELECT DISTINCT part_no FROM lg_plan_daily {where} "
                "ORDER BY part_no LIMIT %s OFFSET %s",
                tuple(params + [page_size, (page - 1) * page_size]),
                fetch="all",
            ) or []
        parte_list = [r["part_no"] for r in partes]

        filas = []
        if parte_list:
            placeholders = ", ".join(["%s"] * len(parte_list))
            datos = execute_query(
                "SELECT part_no, plan_date, plan_qty FROM lg_plan_daily "
                f"WHERE plan_date BETWEEN %s AND %s AND part_no IN ({placeholders}) "
                "ORDER BY part_no, plan_date",
                tuple([date_from, date_to] + parte_list),
                fetch="all",
            ) or []
            por_parte = {p: {} for p in parte_list}
            for r in datos:
                fecha = r["plan_date"]
                if isinstance(fecha, datetime):
                    fecha = fecha.date()
                por_parte[r["part_no"]][fecha.isoformat()] = int(r["plan_qty"])

            def _a_date(v):
                return v.date() if isinstance(v, datetime) else v

            # Demanda total de LG por parte desde hoy (todo el plan importado,
            # independiente del rango visible) — "cuanto ocupa LG".
            req_rows = execute_query(
                "SELECT part_no, SUM(plan_qty) AS req FROM lg_plan_daily "
                f"WHERE plan_date >= %s AND part_no IN ({placeholders}) "
                "GROUP BY part_no",
                tuple([date.today()] + parte_list),
                fetch="all",
            ) or []
            req_por_parte = {r["part_no"]: int(r["req"] or 0) for r in req_rows}

            # Inventario semanal (escalares Part 10) por parte
            inv_rows = execute_query(
                f"SELECT * FROM lg_part_inventory WHERE part_no IN ({placeholders})",
                tuple(parte_list),
                fetch="all",
            ) or []
            inv_por_parte = {}
            for r in inv_rows:
                ref = _a_date(r.get("ref_date"))
                escalares = {k: int(r.get(k) or 0) for k in PP_INV_FIELDS}
                inv_por_parte[r["part_no"]] = {
                    "board": r.get("board"),
                    "line": r.get("line"),
                    **escalares,
                    "i0": sum(escalares.values()),
                    "ref_date": ref.isoformat() if ref else None,
                    "_ref": ref,
                }

            # Schedule (renglon S) en el rango consultado
            sched_rows = execute_query(
                "SELECT part_no, sched_date, sched_qty FROM lg_schedule_daily "
                f"WHERE sched_date BETWEEN %s AND %s AND part_no IN ({placeholders})",
                tuple([date_from, date_to] + parte_list),
                fetch="all",
            ) or []
            sched_por_parte = {p: {} for p in parte_list}
            for r in sched_rows:
                f = _a_date(r["sched_date"])
                sched_por_parte[r["part_no"]][f.isoformat()] = int(r["sched_qty"])

            # Arrastre de I: acumulados de P y S entre ref_date y date_from-1
            prev_plan = {}
            prev_sched = {}
            refs = [d["_ref"] for d in inv_por_parte.values() if d["_ref"]]
            if refs and min(refs) < date_from:
                min_ref = min(refs)
                for tabla, col_f, col_q, destino in (
                    ("lg_plan_daily", "plan_date", "plan_qty", prev_plan),
                    ("lg_schedule_daily", "sched_date", "sched_qty", prev_sched),
                ):
                    rows_prev = execute_query(
                        f"SELECT part_no, {col_f} AS f, {col_q} AS q FROM {tabla} "
                        f"WHERE {col_f} >= %s AND {col_f} < %s "
                        f"AND part_no IN ({placeholders})",
                        tuple([min_ref, date_from] + parte_list),
                        fetch="all",
                    ) or []
                    for r in rows_prev:
                        inv = inv_por_parte.get(r["part_no"])
                        if inv and inv["_ref"] and _a_date(r["f"]) >= inv["_ref"]:
                            destino[r["part_no"]] = destino.get(r["part_no"], 0) + int(r["q"])

            fechas_rango = [
                date_from + timedelta(days=i)
                for i in range((date_to - date_from).days + 1)
            ]
            for p in parte_list:
                qty = por_parte[p]
                sched = sched_por_parte.get(p, {})
                inv = inv_por_parte.get(p)
                # Proyeccion I (formula del Excel): I(t) = I(t-1) - P(t) + S(t),
                # I0 = LGEMM+ISEMM+SVC+DIF+PENDIENTE+REWORK desde ref_date.
                proj = {}
                if inv and inv["_ref"]:
                    carry = inv["i0"] - prev_plan.get(p, 0) + prev_sched.get(p, 0)
                    for f in fechas_rango:
                        iso = f.isoformat()
                        if f < inv["_ref"]:
                            proj[iso] = None
                        else:
                            carry = carry - qty.get(iso, 0) + sched.get(iso, 0)
                            proj[iso] = carry
                filas.append(
                    {
                        "part_no": p,
                        "total": sum(qty.values()),
                        "qty": qty,
                        "req_lg": req_por_parte.get(p, 0),
                        "sched": sched,
                        "sched_total": sum(sched.values()),
                        "inv": ({k: v for k, v in inv.items() if k != "_ref"} if inv else None),
                        "proj": proj,
                    }
                )

            if only_shortage:
                # Solo faltantes accionables: dias de HOY en adelante (un
                # negativo de un dia pasado ya no se puede producir).
                hoy_iso = date.today().isoformat()
                filas = [
                    f for f in filas
                    if any(
                        v is not None and v < 0 and iso >= hoy_iso
                        for iso, v in f["proj"].items()
                    )
                ]
                total_parts = len(filas)
                filas = filas[(page - 1) * page_size : page * page_size]

        num_dias = (date_to - date_from).days + 1
        fechas = [
            (date_from + timedelta(days=i)).isoformat() for i in range(num_dias)
        ]

        return jsonify(
            {
                "success": True,
                "date_from": date_from.isoformat(),
                "date_to": date_to.isoformat(),
                "dates": fechas,
                "rows": filas,
                "total_parts": total_parts,
                "page": page,
                "page_size": page_size,
            }
        )
    except Exception as e:
        logger.error("Error en api_pp_plan: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/part-planning/imports", methods=["GET"])
@login_requerido
def api_pp_imports():
    """Historial de importaciones del plan LG."""
    try:
        limit = min(max(request.args.get("limit", type=int) or 20, 1), 100)
        rows = execute_query(
            "SELECT * FROM lg_plan_imports ORDER BY imported_at DESC LIMIT %s",
            (limit,),
            fetch="all",
        ) or []
        for r in rows:
            for k, v in r.items():
                if isinstance(v, (datetime, date)):
                    r[k] = v.isoformat(sep=" ") if isinstance(v, datetime) else v.isoformat()
        return jsonify({"success": True, "imports": rows})
    except Exception as e:
        logger.error("Error en api_pp_imports: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


# =============================
# PLAN PROYECTADO (generador de lotes tipo hoja "LOTE N" del Excel)
#
# Reglas (confirmadas con el usuario):
#   - faltante = |I proyectado negativo| a la fecha (formula de Proyeccion)
#   - se planea faltante + 10%, redondeado ARRIBA a caja cerrada (estandar_pack)
#   - familias juntas (prefijo del part_no) y por linea
#   - por linea la suma de horas (qty/UPH) debe caber en 9 h sin tiempo extra
#   - los lotes quedan PENDIENTE sin numero; al confirmar se numeran
#     I<YYYYMMDD>-#### (consecutivo global del dia) y su qty se SUMA al
#     renglon S (lg_schedule_daily) de Proyeccion
# =============================


def _ppy_anticipacion(linea):
    """Dias de produccion que se adelanta una linea (D1 mira mas lejos)."""
    return PPY_ANTICIPACION_POR_LINEA.get(
        str(linea or "").strip().upper(), PPY_ANTICIPACION_DIAS
    )


def _ppy_bloque_con_horas(horas_rest):
    """Bloque de 9 h mas libre. Los bloques son flexibles: cualquiera atiende
    cualquier linea, asi que solo importa donde queda cupo."""
    return max(horas_rest, key=lambda b: horas_rest[b]) if horas_rest else None


def _ppy_pueden_compartir(linea, lineas_del_bloque):
    """Si `linea` puede correr en el mismo bloque que las que ya estan ahi."""
    l = str(linea or "").strip().upper()
    for otra in lineas_del_bloque:
        o = str(otra or "").strip().upper()
        if o in PPY_LINEAS_INCOMPATIBLES.get(l, ()) or l in PPY_LINEAS_INCOMPATIBLES.get(o, ()):
            return False
    return True


def _ppy_repartir_bloques(horas_pedidas, horas_rest):
    """Reparte las lineas del dia entre los bloques de 9 h.

    horas_pedidas: {linea: horas que necesita para cubrir su faltante}
    horas_rest:    {bloque: horas libres} — se muta al asignar.

    Dos reglas:
      - Cada linea entra COMPLETA en un solo bloque. Los bloques son equipos
        que trabajan en paralelo, asi que partir una linea entre dos seria
        correrla en dos lados a la vez. En el plan real cada linea aparece en
        un unico bloque (16/07: M2+M4 | M3 | D1 | D2+D3).
      - No cualquier par comparte bloque: ver PPY_LINEAS_INCOMPATIBLES (D3
        nunca con las main).

    Las lineas mas pesadas primero. Se usa best-fit: primero intenta completar
    un bloque ya abierto y compatible; solo abre uno nuevo cuando la linea no
    cabe completa en los abiertos. Asi cuatro bloques bastan cuando las horas
    se pueden empacar y el quinto aparece solo cuando es necesario.

    Retorna {linea: (bloque, horas_asignadas)}; una linea sin cupo no aparece.
    """
    asignacion = {}
    # Un bloque con horas previamente descontadas ya esta abierto, aunque el
    # Excel legacy no permita saber que linea lo ocupa. El marcador no agrega
    # incompatibilidades; solo evita abrir otro bloque antes de aprovecharlo.
    ocupantes = {
        b: (["__RESERVADO__"] if h < PPY_HORAS_TURNO else [])
        for b, h in horas_rest.items()
    }
    for linea, pedidas in sorted(horas_pedidas.items(), key=lambda kv: (-kv[1], kv[0])):
        opciones = [b for b in horas_rest
                    if horas_rest[b] > 0 and _ppy_pueden_compartir(linea, ocupantes[b])]
        if not opciones:
            continue
        completas = [b for b in opciones if horas_rest[b] + 1e-9 >= pedidas]
        usadas = [b for b in completas if ocupantes[b]]
        if usadas:
            # Best-fit deja el menor hueco y conserva libres los bloques nuevos.
            afines = set(PPY_LINEAS_AFINES.get(linea, ()))
            bloque = min(
                usadas,
                key=lambda b: (
                    horas_rest[b] - pedidas,
                    0 if afines.intersection(ocupantes[b]) else 1,
                    b,
                ),
            )
        elif completas:
            bloque = min(completas)
        else:
            # Si la linea por si sola rebasa 9 h (o ya no queda un bloque
            # completo), toma el mayor cupo compatible y se planea parcial.
            bloque = max(opciones, key=lambda b: (horas_rest[b], b))
        dadas = min(pedidas, horas_rest[bloque])
        horas_rest[bloque] -= dadas
        ocupantes[bloque].append(linea)
        asignacion[linea] = (bloque, dadas)
    return asignacion


def _ppy_horas_iniciales(lineas_activas=None, horas_ocupadas=None):
    """Horas disponibles: hasta PPY_BLOQUES turnos de 9 h.

    horas_ocupadas: {linea: horas} ya usadas (lotes confirmados, schedule). Se
    descuentan del bloque mas libre, porque el Excel no dice en que bloque
    quedo cada lote confirmado.
    """
    horas = {"B%d" % i: PPY_HORAS_TURNO for i in range(1, PPY_BLOQUES + 1)}
    for _linea, h in sorted((horas_ocupadas or {}).items(),
                            key=lambda kv: -kv[1]):
        b = _ppy_bloque_con_horas(horas)
        if b:
            horas[b] -= h
    return horas


def _ppy_es_dia_produccion(dia):
    """ISEMM produce de lunes a viernes.

    El sabado solo se trabaja en caso extraordinario (8am-4pm) y se decide a
    mano, asi que el generador no propone lotes en sabado ni domingo: el
    faltante del sabado lo cubre el viernes, porque el horizonte si cuenta el
    consumo de LG de ese dia (ver _ppy_sumar_dias_consumo).
    """
    return dia.weekday() < 5


def _ppy_sumar_dias_produccion(dia, dias):
    """Fecha tras avanzar N dias de PRODUCCION (L-V) desde `dia`.

    Es el calendario del anticipo: un lote se programa 2 dias de produccion
    antes de la fecha en que falta. Desde el jueves, +2 son viernes y LUNES
    (el fin de semana no cuenta), asi que el faltante del lunes se produce el
    jueves. Contarlo en dias de consumo daria viernes, que ya es tarde.
    """
    f = dia
    while dias > 0:
        f += timedelta(days=1)
        if _ppy_es_dia_produccion(f):
            dias -= 1
    return f


def _ppy_sumar_dias_consumo(dia, dias):
    """Fecha tope tras avanzar N dias de CONSUMO desde `dia`.

    El horizonte mide cuando va a FALTAR material, asi que cuenta los dias en
    que LG consume, no los dias en que ISEMM produce. No son los mismos:
      - LG consume lunes a SABADO (el sabado ~9k pzs, un tercio de un dia
        normal). El domingo nunca: 0 demanda en todo el plan.
      - ISEMM produce lunes a viernes; el sabado solo en caso extraordinario
        con tiempo extra.
    Por eso aqui solo se salta el domingo. Contar dias naturales metia el
    domingo en la ventana (un dia que no consume ni produce) y contar dias
    habiles dejaba fuera el sabado, que si genera faltante.

    ponytail: solo salta domingo; los feriados y los sabados sin demanda no
    estan modelados. Si hiciera falta, el calendario va en lg_pp_config.
    """
    f = dia
    while dias > 0:
        f += timedelta(days=1)
        if f.weekday() != 6:  # 6=domingo: LG no consume
            dias -= 1
    return f


def _ppy_familia(part_no):
    return (part_no or "")[:PPY_FAMILIA_LEN]


def _ppy_qty_pack(faltante, pack):
    """Cantidad a planear: faltante + 10%, redondeada arriba a caja cerrada."""
    try:
        pack = int(pack) if pack and int(pack) > 0 else 1
    except (TypeError, ValueError):
        pack = 1
    # round() evita que el float de *1.10 suba de mas (100*1.1 = 110.0000...01)
    objetivo = round(faltante * PPY_MARGEN, 6)
    return int(math.ceil(objetivo / pack)) * pack


def _ppy_parse_uph(valor):
    """UPH de raw es varchar: parseo defensivo, None si no es numero positivo."""
    try:
        n = int(float(str(valor).strip()))
        return n if n > 0 else None
    except (TypeError, ValueError):
        return None


def _ppy_parse_ct(valor):
    """Cycle time positivo desde RAW; se conserva con precision decimal."""
    try:
        n = float(str(valor).strip())
        return round(n, 4) if n > 0 else None
    except (TypeError, ValueError):
        return None


def _ppy_parse_fecha(texto):
    if isinstance(texto, datetime):
        return texto.date()
    if isinstance(texto, date):
        return texto
    if not isinstance(texto, str):
        return None
    try:
        return datetime.strptime(texto.strip(), "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


def _ppy_config_lineas(query=None):
    """Lineas activas configurables; un error SQL nunca habilita defaults."""
    run_query = query or execute_query
    row = run_query(
        "SELECT valor FROM lg_pp_config WHERE clave = 'lineas_activas'",
        fetch="one",
    )
    texto = (row or {}).get("valor") or PPY_LINEAS_DEFAULT
    lineas = []
    for l in texto.split(","):
        l = l.strip().upper()[:10]
        if l and l not in lineas:
            lineas.append(l)
    return lineas or [x.strip() for x in PPY_LINEAS_DEFAULT.split(",")]


def _ppy_partes_excluidas(query=None):
    """Partes que no se producen aqui aunque LG las pida.

    Su demanda no es trabajo nuestro: ACQ30500849, por ejemplo, la vende ILSAN
    Corea aparte. Sin esta lista el generador la planeaba (1920 pzs en D2) y
    planning nunca la mete. Un error de SQL no vacia la lista por si acaso: se
    cae al default.
    """
    run_query = query or execute_query
    try:
        row = run_query(
            "SELECT valor FROM lg_pp_config WHERE clave = 'partes_excluidas'",
            fetch="one",
        )
    except Exception:
        row = None
    texto = (row or {}).get("valor")
    if texto is None:
        texto = PPY_EXCLUIDAS_DEFAULT
    return {p.strip().upper() for p in str(texto).split(",") if p.strip()}


def _ppy_lineas_permitidas_map(query=None):
    """Lineas permitidas desde raw.lineas_permitidas (CSV, ej. 'M1,M2,M4').

    Se configura una sola vez por familia: basta capturarla en cualquier
    parte de la familia y aplica a todos los sufijos (EBR807574xx).
    Retorna ({part_no: [lineas]}, {familia: [lineas]}).
    """
    run_query = query or execute_query
    rows = run_query(
        "SELECT TRIM(part_no) AS part_no, lineas_permitidas FROM raw "
        "WHERE lineas_permitidas IS NOT NULL AND TRIM(lineas_permitidas) <> '' "
        "ORDER BY TRIM(part_no)",
        fetch="all",
    ) or []
    por_parte = {}
    por_familia = {}
    for r in rows:
        lineas = [l.strip().upper() for l in (r["lineas_permitidas"] or "").split(",") if l.strip()]
        if not lineas:
            continue
        por_parte[r["part_no"]] = lineas
        por_familia.setdefault(_ppy_familia(r["part_no"]), lineas)
    return por_parte, por_familia


def _ppy_proyeccion_rango(fecha_ini, fecha_fin, query=None, replanear_desde=None):
    """I proyectado por parte para cada dia de [fecha_ini, fecha_fin].

    Misma formula que Proyeccion: I(t) = I(t-1) - P(t) + S(t) desde ref_date.
    Como I ya incluye los S confirmados/capturados, regenerar no duplica.
    Retorna {part_no: {"line": str|None, "proj": {date: int}}}.

    replanear_desde: fecha a partir de la cual se IGNORA el schedule ya
    capturado. El plan se rehace cada dia porque el de LG cambia, asi que el
    S de mañana que se puso ayer esta viejo: contarlo como cubierto daria un
    plan incompleto (solo el delta). El S anterior a esa fecha si cuenta: ya
    se produjo. None = proyeccion normal (el estado real, para consultar).
    """
    run_query = query or execute_query
    inv_rows = run_query("SELECT * FROM lg_part_inventory", fetch="all") or []
    base = {}
    refs = {}
    raw_lineas = {}
    try:
        raw_rows = run_query(
            "SELECT TRIM(part_no) AS part_no, assy_line FROM raw "
            "WHERE assy_line IS NOT NULL AND TRIM(assy_line) <> ''",
            fetch="all",
        ) or []
        raw_lineas = {
            str(r.get("part_no") or "").strip():
            str(r.get("assy_line") or "").strip().upper()[:10]
            for r in raw_rows
            if str(r.get("part_no") or "").strip()
        }
    except Exception:
        # Permite leer historiales antiguos si la migracion aun no corrio.
        raw_lineas = {}
    for r in inv_rows:
        ref = r.get("ref_date")
        if isinstance(ref, datetime):
            ref = ref.date()
        i0 = sum(int(r.get(k) or 0) for k in PP_INV_FIELDS)
        part_no = str(r.get("part_no") or "").strip()
        base[part_no] = {
            "line": raw_lineas.get(part_no) or r.get("line"),
            "i0": i0,
        }
        if ref and ref <= fecha_fin:
            refs[part_no] = ref

    movs = {}  # (part, date) -> delta (-P + S)
    if refs:
        min_ref = min(refs.values())
        for tabla, col_f, col_q, signo in (
            ("lg_plan_daily", "plan_date", "plan_qty", -1),
            ("lg_schedule_daily", "sched_date", "sched_qty", 1),
        ):
            rows = run_query(
                f"SELECT part_no, {col_f} AS f, {col_q} AS q FROM {tabla} "
                f"WHERE {col_f} BETWEEN %s AND %s",
                (min_ref, fecha_fin),
                fetch="all",
            ) or []
            es_sched = tabla == "lg_schedule_daily"
            for r in rows:
                ref = refs.get(r["part_no"])
                if not ref:
                    continue
                f = r["f"].date() if isinstance(r["f"], datetime) else r["f"]
                if es_sched and replanear_desde and f >= replanear_desde:
                    continue  # ese schedule se va a rehacer
                if f >= ref:
                    clave = (r["part_no"], f)
                    movs[clave] = movs.get(clave, 0) + signo * int(r["q"] or 0)

        if replanear_desde:
            # El Schedule futuro se puede reemplazar, pero los lotes ya
            # confirmados no. Se reincorporan a la proyeccion para que el plan
            # nuevo calcule solo lo adicional a esa produccion inamovible.
            confirmed_rows = run_query(
                "SELECT part_no, plan_date AS f, SUM(qty_plan) AS q "
                "FROM lg_lote_plan WHERE status='CONFIRMADO' "
                "AND plan_date BETWEEN %s AND %s GROUP BY part_no, plan_date",
                (replanear_desde, fecha_fin),
                fetch="all",
            ) or []
            for row in confirmed_rows:
                ref = refs.get(row["part_no"])
                if not ref:
                    continue
                f = row["f"].date() if isinstance(row["f"], datetime) else row["f"]
                if f >= ref:
                    key = (row["part_no"], f)
                    movs[key] = movs.get(key, 0) + int(row.get("q") or 0)

    resultado = {}
    for p, d in base.items():
        ref = refs.get(p)
        proj = {}
        carry = d["i0"]
        if ref:
            f = ref
            while f <= fecha_fin:
                carry += movs.get((p, f), 0)
                if f >= fecha_ini:
                    proj[f] = carry
                f += timedelta(days=1)
        else:
            f = fecha_ini
            while f <= fecha_fin:
                proj[f] = carry
                f += timedelta(days=1)
        resultado[p] = {"line": d["line"], "proj": proj}
    return resultado


def _ppy_armar_lotes(candidatos, lineas_activas, horas_rest, estricto=True):
    """Ajusta los candidatos a las horas disponibles por linea (9 h sin TE).

    candidatos: [{part_no, falt_total, falt_hoy, primera_falta(date), line,
                  uph, pack, model, main_sub}]
    horas_rest: {bloque: horas disponibles} (se muta al asignar). La capacidad
        son hasta 5 turnos de 9 h flexibles, no 9 h por linea: ver PPY_BLOQUES y
        _ppy_horas_iniciales().

    Orden: familias con la falta mas proxima primero (los faltantes de HOY van
    antes que los adelantos); familia junta y de preferencia en la misma linea.
    Si un lote no cabe completo se planea parcial a caja cerrada ("lo que se
    pueda meter"). Retorna (lotes, fuera).

    Linea: se respeta la de la parte (raw.lineas_permitidas si esta capturada,
    si no raw.assy_line). Una parte con linea conocida NO se reasigna a otra
    linea al saturarse la suya: se reporta en `fuera` como "sin horas
    disponibles en el turno de X". Solo las partes sin linea alguna caen al
    reparto por linea mas libre.

    Capacidad: en dos pases. Primero se decide la linea de cada lote y cuantas
    horas pediria; luego las lineas se reparten entre los bloques de 9 h (una
    linea completa por bloque, ver _ppy_repartir_bloques) y las cantidades se
    recortan al cupo que le toco a cada una.

    estricto=True (Plan Proyectado): sin UPH o sin pack la parte no se genera.
    estricto=False (propuesta de schedule): pack faltante se asume
    PPY_PACK_DEFAULT y sin UPH la parte entra sin presupuesto de horas.
    """
    familias = {}
    for c in candidatos:
        familias.setdefault(_ppy_familia(c["part_no"]), []).append(c)
    orden = sorted(
        familias.items(),
        key=lambda item: (
            min(c["primera_falta"] for c in item[1]),
            -max(c["falt_total"] for c in item[1]),
            item[0],
        ),
    )

    lotes = []
    fuera = []
    plan = []       # pase 1: [(candidato, linea, qty_obj, pack, adelanto)]
    pedidas = {}    # linea -> horas que necesitaria para todo su faltante
    for _familia, grupo in orden:
        grupo.sort(key=lambda c: (c["primera_falta"], -c["falt_total"], c["part_no"]))
        linea_familia = None
        for c in grupo:
            pack = c["pack"] if c["pack"] and c["pack"] > 0 else None
            if estricto:
                # Sin UPH no hay presupuesto de horas y sin pack no hay caja
                # cerrada: la parte no se genera (visible en no_incluidos).
                faltan = []
                if not c["uph"]:
                    faltan.append("UPH")
                if not pack:
                    faltan.append("pack")
                if faltan:
                    fuera.append({"part_no": c["part_no"],
                                  "qty": _ppy_qty_pack(c["falt_total"], pack),
                                  "motivo": "sin " + " ni ".join(faltan) + " en raw"})
                    continue
            elif not pack:
                pack = PPY_PACK_DEFAULT

            # La capacidad se calcula con UPH. Si RAW trae UPH y empaque pero
            # no CT (caso real ACQ91482499 del 17/07), no se debe perder el
            # lote: CT es derivable y se conserva para el contrato de salida.
            if not c.get("ct") and c.get("uph"):
                c = {**c, "ct": 3600.0 / float(c["uph"])}

            qty_obj = _ppy_qty_pack(c["falt_total"], pack)
            adelanto = c["falt_hoy"] <= 0

            if not c["uph"]:
                # Solo en modo laxo: entra completo sin consumir horas
                lotes.append({**c, "qty": qty_obj,
                              "linea": linea_familia or c.get("line"),
                              "comentario": None})
                continue

            # Linea: la de la familia > la preferida de la parte > la mas
            # libre, siempre dentro de las permitidas (raw.lineas_permitidas)
            permitidas = c.get("permitidas")
            linea_parte = (c.get("line") or "").upper() or None
            if permitidas is None and linea_parte:
                # assy_line manda: es la linea real de ensamble de la parte, no
                # una sugerencia. Sin esto, al saturarse su linea el fallback
                # "la mas libre" llegaba a meter tarjetas main en H1 (harness).
                # lineas_permitidas, si esta capturada, sigue teniendo prioridad.
                permitidas = [linea_parte]

            def _permitida(l):
                return permitidas is None or l in permitidas

            # La linea la fija assy_line; el cupo de horas sale del bloque, y
            # eso se resuelve en el pase 2 (una linea entera por bloque).
            activas = {str(l).strip().upper() for l in lineas_activas}
            opciones = []
            for cand in (linea_familia, (c.get("line") or "").upper() or None):
                if cand and cand in activas and _permitida(cand) and cand not in opciones:
                    opciones.append(cand)
            opciones += sorted(l for l in activas
                               if l not in opciones and _permitida(l))
            if not opciones:
                fuera.append({"part_no": c["part_no"], "qty": qty_obj,
                              "motivo": "sus lineas (" + ",".join(permitidas or []) +
                                        ") no estan activas"})
                continue
            elegido = linea_familia if linea_familia in opciones else opciones[0]
            if linea_familia is None:
                linea_familia = elegido
            plan.append((c, elegido, qty_obj, pack, adelanto))
            pedidas[elegido] = pedidas.get(elegido, 0.0) + qty_obj / c["uph"]

    # Pase 2: repartir las lineas entre los bloques de 9 h del dia. Cada linea
    # queda completa en un bloque; el bloque le acota las horas.
    asignacion = _ppy_repartir_bloques(pedidas, horas_rest)
    libre = {linea: horas for linea, (_b, horas) in asignacion.items()}

    # Pase 3: cantidades ajustadas al cupo que le toco a cada linea.
    for c, linea, qty_obj, pack, adelanto in plan:
        bloque = (asignacion.get(linea) or (None, 0.0))[0]
        cap = int(max(libre.get(linea, 0.0), 0.0) * c["uph"]) // pack * pack
        qty = min(qty_obj, cap) if cap >= pack else 0
        if not qty:
            # No quedan horas en el bloque de su linea: aqui es donde entraria
            # el tiempo extra si planning lo autoriza.
            fuera.append({"part_no": c["part_no"], "qty": qty_obj,
                          "motivo": "sin horas disponibles en el turno de " + linea})
            continue
        libre[linea] -= qty / c["uph"]
        notas = []
        if adelanto:
            notas.append(f"Adelanto: falta el {c['primera_falta'].strftime('%d/%m')}")
        if qty < qty_obj:
            notas.append("parcial por horas")
        lotes.append({**c, "qty": qty, "linea": linea, "grupo": bloque,
                      "comentario": " | ".join(notas) or None})
    return lotes, fuera


def _ppy_datos_raw(parte_list, query=None):
    """Datos de estandar RAW, indexados por numero de parte normalizado."""
    if not parte_list:
        return {}
    placeholders = ", ".join(["%s"] * len(parte_list))
    run_query = query or execute_query
    rows = run_query(
        "SELECT TRIM(part_no) AS part_no, model, sub_assy, c_t, uph, estandar_pack, assy_line "
        f"FROM raw WHERE TRIM(part_no) IN ({placeholders})",
        tuple(parte_list),
        fetch="all",
    ) or []
    return {r["part_no"]: r for r in rows}


_PPY_ORDEN_SQL = (
    "ORDER BY COALESCE(NULLIF(linea, ''), 'ZZZ'), "
    f"LEFT(part_no, {PPY_FAMILIA_LEN}), part_no, id"
)


def _ppy_listado(fecha):
    """Lotes del dia + resumen de horas por linea (para GET y respuestas)."""
    rows = execute_query(
        "SELECT * FROM lg_lote_plan WHERE plan_date = %s "
        "AND status IN ('PENDIENTE', 'CONFIRMADO') " + _PPY_ORDEN_SQL,
        (fecha,),
        fetch="all",
    ) or []
    lotes = []
    lineas = {}
    for r in rows:
        uph = r.get("uph")
        qty = int(r.get("qty_plan") or 0)
        horas = round(qty / uph, 2) if uph else None
        fisico = r.get("fisico")
        pack = r.get("estandar_pack")
        lotes.append(
            {
                "id": r["id"],
                "lot_no": r.get("lot_no"),
                "part_no": r["part_no"],
                "familia": _ppy_familia(r["part_no"]),
                "model": r.get("model"),
                "main_sub": r.get("main_sub"),
                "linea": r.get("linea"),
                "turno": r.get("turno"),
                "faltante": int(r.get("faltante") or 0),
                "qty_plan": qty,
                "uph": uph,
                "estandar_pack": pack,
                "time_horas": horas,
                "fisico": fisico,
                "falta": (int(fisico) - qty) if fisico is not None else None,
                "pct": (round(int(fisico) * 100.0 / qty, 1) if fisico is not None and qty else None),
                "status": r.get("status"),
                "comentario": r.get("comentario"),
                "warn_pack": bool(pack and int(pack) > 1 and qty % int(pack) != 0),
                "warn_uph": uph is None,
            }
        )
        clave = r.get("linea") or "SIN LINEA"
        acum = lineas.setdefault(clave, {"linea": clave, "horas": 0.0, "lotes": 0})
        acum["lotes"] += 1
        if horas:
            acum["horas"] = round(acum["horas"] + horas, 2)
    resumen = []
    for clave in sorted(lineas, key=lambda x: (x == "SIN LINEA", x)):
        d = lineas[clave]
        d["horas_max"] = PPY_HORAS_TURNO
        # "SIN LINEA" agrupa lo no acomodado: su suma de horas no es un exceso.
        d["excede"] = clave != "SIN LINEA" and d["horas"] > PPY_HORAS_TURNO
        resumen.append(d)
    return {"fecha": fecha.isoformat(), "lotes": lotes, "lineas": resumen}


@bp.route("/api/plan-proyectado", methods=["GET"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PPY_PERMISO_BOTON)
def api_ppy_listado():
    try:
        fecha = _ppy_parse_fecha(request.args.get("fecha")) or date.today()
        return jsonify(
            {"success": True, "lineas_activas": _ppy_config_lineas(), **_ppy_listado(fecha)}
        )
    except Exception as e:
        logger.error("Error en api_ppy_listado: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/plan-proyectado/config", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PPY_PERMISO_BOTON)
def api_ppy_config():
    """Guarda las lineas activas (CSV, ej. 'M1,M2,M3') para el generador."""
    try:
        data = request.get_json(silent=True) or {}
        texto = (data.get("lineas") or "").strip().upper()
        lineas = []
        for l in texto.split(","):
            l = l.strip()
            if not l:
                continue
            if not re.match(r"^[A-Z0-9\-]{1,10}$", l):
                return jsonify({"success": False, "error": f"Linea invalida: {l}"}), 400
            if l not in lineas:
                lineas.append(l)
        if not lineas:
            return jsonify({"success": False, "error": "Captura al menos una linea"}), 400
        execute_query(
            "INSERT INTO lg_pp_config (clave, valor) VALUES ('lineas_activas', %s) "
            "ON DUPLICATE KEY UPDATE valor = VALUES(valor)",
            (",".join(lineas),),
        )
        return jsonify({"success": True, "lineas_activas": lineas})
    except Exception as e:
        logger.error("Error en api_ppy_config: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/plan-proyectado/generate", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PPY_PERMISO_BOTON)
def api_ppy_generate():
    """Genera la propuesta del dia: borra PENDIENTEs de la fecha y los recrea.

    Cubre los faltantes de HOY y adelanta faltantes de los proximos
    PPY_ANTICIPACION_DIAS dias hasta llenar las horas de las lineas activas.
    """
    conn = None
    cursor = None
    try:
        data = request.get_json(silent=True) or {}
        fecha = _ppy_parse_fecha(data.get("fecha")) or date.today()
        usuario = session.get("usuario") or "SISTEMA"

        lineas_activas = _ppy_config_lineas()
        # Los lotes ya confirmados del dia consumen horas de su GRUPO
        confirmados = execute_query(
            "SELECT part_no, linea, qty_plan, uph FROM lg_lote_plan "
            "WHERE plan_date = %s AND status = 'CONFIRMADO'",
            (fecha,),
            fetch="all",
        ) or []
        partes_confirmadas = set()
        ocupadas = {}
        for r in confirmados:
            partes_confirmadas.add(r["part_no"])
            l = (r.get("linea") or "").upper()
            if l and r.get("uph"):
                ocupadas[l] = ocupadas.get(l, 0.0) + int(r["qty_plan"] or 0) / int(r["uph"])
        horas_rest = _ppy_horas_iniciales(lineas_activas, ocupadas)

        # Se proyecta al horizonte MAXIMO y luego cada parte mira solo hasta el
        # de su linea: proyectar solo a PPY_ANTICIPACION_DIAS le cortaria los datos
        # a D1, que adelanta mas.
        proy = _ppy_proyeccion_rango(
            fecha, _ppy_sumar_dias_produccion(fecha, PPY_ANTICIPACION_MAX))
        excluidas = _ppy_partes_excluidas()
        candidatos = []
        for p, d in proy.items():
            if p in partes_confirmadas or not d["proj"]:
                continue
            if p.upper() in excluidas:
                continue  # no se produce aqui (ver _ppy_partes_excluidas)
            tope = _ppy_sumar_dias_produccion(fecha, _ppy_anticipacion(d["line"]))
            proj = {f: v for f, v in d["proj"].items() if f <= tope}
            if not proj:
                continue
            # Solo dispara el faltante real (I < 0); al reponer, el lote
            # deja el inventario en PPY_REMAIN_IDEAL, no en cero.
            min_i = min(proj.values())
            if min_i >= 0:
                continue
            i_hoy = proj.get(fecha, 0)
            candidatos.append(
                {
                    "part_no": p,
                    "falt_total": PPY_REMAIN_IDEAL - min_i,
                    "falt_hoy": max(0, -i_hoy),
                    "primera_falta": min(f for f, v in proj.items() if v < 0),
                    "line": d["line"],
                }
            )

        datos_raw = _ppy_datos_raw([c["part_no"] for c in candidatos])
        perm_parte, perm_familia = _ppy_lineas_permitidas_map()
        for c in candidatos:
            raw_info = datos_raw.get(c["part_no"]) or {}
            pack = raw_info.get("estandar_pack")
            c["pack"] = int(pack) if pack else None
            c["ct"] = _ppy_parse_ct(raw_info.get("c_t"))
            c["uph"] = _ppy_parse_uph(raw_info.get("uph"))
            c["model"] = raw_info.get("model")
            c["main_sub"] = raw_info.get("sub_assy")
            # Control de modelos RAW es la fuente oficial de la linea de
            # ensamble. La linea de inventario/proyeccion solo respalda
            # registros antiguos sin assy_line.
            c["line"] = (
                str(raw_info.get("assy_line") or "").strip().upper()[:10]
                or c.get("line")
            )
            c["permitidas"] = (perm_parte.get(c["part_no"])
                               or perm_familia.get(_ppy_familia(c["part_no"])))

        lotes, fuera = _ppy_armar_lotes(candidatos, lineas_activas, horas_rest)
        nuevos = [
            (
                fecha,
                l["part_no"],
                l.get("model"),
                l.get("main_sub"),
                (l.get("linea") or None),
                "DIA",
                l["falt_total"],
                l["qty"],
                l.get("uph"),
                l.get("pack"),
                l.get("comentario"),
                usuario,
            )
            for l in lotes
        ]

        conn = get_pooled_connection()
        if conn is None:
            raise RuntimeError("No fue posible obtener conexion MySQL.")
        cursor = get_dict_cursor(conn)
        conn.autocommit(False)
        cursor.execute(
            "DELETE FROM lg_lote_plan WHERE plan_date = %s AND status = 'PENDIENTE'",
            (fecha,),
        )
        if nuevos:
            for i in range(0, len(nuevos), PP_BATCH_SIZE):
                cursor.executemany(
                    "INSERT INTO lg_lote_plan (plan_date, part_no, model, main_sub, "
                    "linea, turno, faltante, qty_plan, uph, estandar_pack, comentario, created_by) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    nuevos[i : i + PP_BATCH_SIZE],
                )
        conn.commit()

        return jsonify(
            {
                "success": True,
                "generados": len(nuevos),
                "no_incluidos": fuera,
                "lineas_activas": lineas_activas,
                **_ppy_listado(fecha),
            }
        )
    except Exception as e:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        logger.error("Error en api_ppy_generate: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if cursor is not None:
            try:
                cursor.close()
            except Exception:
                pass
        if conn is not None:
            try:
                conn.autocommit(True)
                conn.close()
            except Exception:
                pass


@bp.route("/api/plan-proyectado/generate-faltantes", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PPY_PERMISO_BOTON)
def api_ppy_generate_faltantes():
    """Crea lotes PENDIENTES desde los faltantes del dia SIN asignar linea.

    La linea se asigna despues (manual o con "Acomodar con IA"). Cantidad =
    faltante +10% a caja cerrada (pack asumido PPY_PACK_DEFAULT si no hay).
    Reemplaza los PENDIENTES del dia; no toca los CONFIRMADOS.
    """
    conn = None
    cursor = None
    try:
        data = request.get_json(silent=True) or {}
        fecha = _ppy_parse_fecha(data.get("fecha")) or date.today()
        usuario = session.get("usuario") or "SISTEMA"

        confirmadas = execute_query(
            "SELECT DISTINCT part_no FROM lg_lote_plan "
            "WHERE plan_date = %s AND status = 'CONFIRMADO'",
            (fecha,),
            fetch="all",
        ) or []
        partes_confirmadas = {r["part_no"] for r in confirmadas}

        # Proyecta al horizonte maximo; cada parte mira hasta el de su linea.
        proy = _ppy_proyeccion_rango(
            fecha, _ppy_sumar_dias_produccion(fecha, PPY_ANTICIPACION_MAX))
        excluidas = _ppy_partes_excluidas()
        faltantes = {}  # part_no -> faltante total (peor I del rango)
        for p, d in proy.items():
            if p in partes_confirmadas or not d["proj"]:
                continue
            if p.upper() in excluidas:
                continue  # no se produce aqui
            tope = _ppy_sumar_dias_produccion(fecha, _ppy_anticipacion(d["line"]))
            valores = [v for f, v in d["proj"].items() if f <= tope]
            if not valores:
                continue
            # Solo el faltante real; el lote deja el remanente ideal.
            min_i = min(valores)
            if min_i < 0:
                faltantes[p] = PPY_REMAIN_IDEAL - min_i

        datos_raw = _ppy_datos_raw(list(faltantes))
        nuevos = []
        for p, falt in faltantes.items():
            raw_info = datos_raw.get(p) or {}
            pack_raw = raw_info.get("estandar_pack")
            pack = int(pack_raw) if pack_raw else PPY_PACK_DEFAULT
            nuevos.append(
                (
                    fecha, p, raw_info.get("model"), raw_info.get("sub_assy"),
                    None,  # sin linea: se acomoda despues
                    "DIA", falt, _ppy_qty_pack(falt, pack),
                    _ppy_parse_uph(raw_info.get("uph")),
                    (int(pack_raw) if pack_raw else None),
                    None, usuario,
                )
            )

        conn = get_pooled_connection()
        if conn is None:
            raise RuntimeError("No fue posible obtener conexion MySQL.")
        cursor = get_dict_cursor(conn)
        conn.autocommit(False)
        cursor.execute(
            "DELETE FROM lg_lote_plan WHERE plan_date = %s AND status = 'PENDIENTE'",
            (fecha,),
        )
        if nuevos:
            for i in range(0, len(nuevos), PP_BATCH_SIZE):
                cursor.executemany(
                    "INSERT INTO lg_lote_plan (plan_date, part_no, model, main_sub, "
                    "linea, turno, faltante, qty_plan, uph, estandar_pack, comentario, created_by) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    nuevos[i : i + PP_BATCH_SIZE],
                )
        conn.commit()
        return jsonify(
            {"success": True, "generados": len(nuevos),
             "lineas_activas": _ppy_config_lineas(), **_ppy_listado(fecha)}
        )
    except Exception as e:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        logger.error("Error en api_ppy_generate_faltantes: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if cursor is not None:
            try:
                cursor.close()
            except Exception:
                pass
        if conn is not None:
            try:
                conn.autocommit(True)
                conn.close()
            except Exception:
                pass


@bp.route("/api/plan-proyectado/generate-schedule", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PPY_PERMISO_BOTON)
def api_ppy_generate_schedule():
    """Crea un lote PENDIENTE por cada Schedule (renglon S) del dia.

    El Schedule capturado/importado en Proyeccion ES el plan del dia: cada
    celda S de la fecha se vuelve un lote con su cantidad EXACTA. Si el
    schedule proviene de una propuesta aprobada conserva linea y turno
    explicitos; los registros legacy sin linea pueden acomodarse despues.
    Reemplaza los PENDIENTES; no toca los CONFIRMADOS.
    """
    conn = None
    cursor = None
    try:
        data = request.get_json(silent=True) or {}
        fecha = _ppy_parse_fecha(data.get("fecha")) or date.today()
        usuario = session.get("usuario") or "SISTEMA"

        confirmadas = execute_query(
            "SELECT DISTINCT part_no FROM lg_lote_plan "
            "WHERE plan_date = %s AND status = 'CONFIRMADO'",
            (fecha,), fetch="all",
        ) or []
        partes_confirmadas = {r["part_no"] for r in confirmadas}

        sched = execute_query(
            "SELECT part_no, sched_qty, linea, turno FROM lg_schedule_daily "
            "WHERE sched_date = %s AND sched_qty > 0 ORDER BY part_no",
            (fecha,), fetch="all",
        ) or []
        sched = [r for r in sched if r["part_no"] not in partes_confirmadas]

        datos_raw = _ppy_datos_raw([r["part_no"] for r in sched])
        nuevos = []
        for r in sched:
            p = r["part_no"]
            qty = int(r["sched_qty"] or 0)
            raw_info = datos_raw.get(p) or {}
            pack_raw = raw_info.get("estandar_pack")
            nuevos.append(
                (
                    fecha, p, raw_info.get("model"), raw_info.get("sub_assy"),
                    (str(r.get("linea") or "").strip().upper() or None),
                    (str(r.get("turno") or "DIA").strip().upper()),
                    qty,  # faltante = cantidad del schedule (referencia)
                    qty,  # qty_plan = cantidad EXACTA del schedule
                    _ppy_parse_uph(raw_info.get("uph")),
                    (int(pack_raw) if pack_raw else None),
                    "Desde schedule", usuario,
                )
            )

        conn = get_pooled_connection()
        if conn is None:
            raise RuntimeError("No fue posible obtener conexion MySQL.")
        cursor = get_dict_cursor(conn)
        conn.autocommit(False)
        cursor.execute(
            "DELETE FROM lg_lote_plan WHERE plan_date = %s AND status = 'PENDIENTE'",
            (fecha,),
        )
        if nuevos:
            for i in range(0, len(nuevos), PP_BATCH_SIZE):
                cursor.executemany(
                    "INSERT INTO lg_lote_plan (plan_date, part_no, model, main_sub, "
                    "linea, turno, faltante, qty_plan, uph, estandar_pack, comentario, created_by) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    nuevos[i : i + PP_BATCH_SIZE],
                )
        conn.commit()
        return jsonify(
            {"success": True, "generados": len(nuevos),
             "lineas_activas": _ppy_config_lineas(), **_ppy_listado(fecha)}
        )
    except Exception as e:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        logger.error("Error en api_ppy_generate_schedule: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if cursor is not None:
            try:
                cursor.close()
            except Exception:
                pass
        if conn is not None:
            try:
                conn.autocommit(True)
                conn.close()
            except Exception:
                pass


def _ppy_acomodo_heuristico(pend, lineas_activas, horas_rest):
    """Fallback local: asigna linea a los lotes sin linea por horas + familia.

    Reutiliza _ppy_armar_lotes: manda la linea de la parte (lineas_permitidas
    si esta capturada, si no assy_line) y se agrupa la familia. No cambia
    cantidades, solo linea: por eso el resultado puede pasar de 9 h y hay que
    mirar el flag `excede` del resumen. Para el plan del dia usar la ruta
    /generate, que si recorta. Retorna {id: linea}.
    """
    candidatos = []
    for l in pend:
        candidatos.append({
            "part_no": l["part_no"], "falt_total": int(l["qty_plan"] or 0),
            "falt_hoy": int(l["qty_plan"] or 0), "primera_falta": date.today(),
            "line": l.get("assy_line"), "uph": l.get("uph"),
            "pack": (int(l["estandar_pack"]) if l.get("estandar_pack") else None),
            "model": l.get("model"), "main_sub": l.get("main_sub"),
            "permitidas": l.get("permitidas"), "_id": l["id"],
        })
    lotes, _ = _ppy_armar_lotes(candidatos, lineas_activas, dict(horas_rest), estricto=False)
    # _ppy_armar_lotes conserva las claves extra via {**c}, incluido _id
    return {l["_id"]: l.get("linea") for l in lotes if l.get("linea")}


@bp.route("/api/plan-proyectado/acomodar-ia", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PPY_PERMISO_BOTON)
def api_ppy_acomodar_ia():
    """Asigna linea a los lotes PENDIENTES sin linea usando la IA (OpenAI).

    Respeta lineas permitidas (raw) y 9 h por linea activa (menos lo ya usado
    por lotes con linea). Si la IA falla o no hay API key, cae a la heuristica
    local. No cambia cantidades: solo asigna linea.
    """
    conn = None
    cursor = None
    try:
        data = request.get_json(silent=True) or {}
        fecha = _ppy_parse_fecha(data.get("fecha")) or date.today()
        usuario = session.get("usuario") or "SISTEMA"
        lineas_activas = _ppy_config_lineas()

        rows = execute_query(
            "SELECT id, part_no, model, main_sub, linea, qty_plan, uph, estandar_pack "
            "FROM lg_lote_plan WHERE plan_date = %s AND status = 'PENDIENTE'",
            (fecha,),
            fetch="all",
        ) or []
        perm_parte, perm_familia = _ppy_lineas_permitidas_map()
        # La linea de ensamble de la parte manda si no hay lineas_permitidas
        # capturadas. Sin esto el acomodo repartia por "la mas libre" y mandaba
        # todo a la primera linea activa por orden alfabetico.
        datos_raw = _ppy_datos_raw([r["part_no"] for r in rows])
        for r in rows:
            r["permitidas"] = (perm_parte.get(r["part_no"])
                               or perm_familia.get(_ppy_familia(r["part_no"])))
            r["assy_line"] = (
                str((datos_raw.get(r["part_no"]) or {}).get("assy_line") or "")
                .strip().upper() or None
            )

        sin_linea = [r for r in rows if not (r.get("linea") or "").strip()]
        if not sin_linea:
            return jsonify({"success": True, "asignados": 0, "metodo": "nada",
                            "lineas_activas": lineas_activas, **_ppy_listado(fecha)})

        # Horas ya ocupadas por lotes con linea (y confirmados), por GRUPO
        confirmados = execute_query(
            "SELECT linea, qty_plan, uph FROM lg_lote_plan "
            "WHERE plan_date = %s AND status = 'CONFIRMADO'",
            (fecha,), fetch="all",
        ) or []
        ocupadas = {}
        for r in list(rows) + confirmados:
            l = (r.get("linea") or "").upper()
            if l and r.get("uph"):
                ocupadas[l] = ocupadas.get(l, 0.0) + int(r["qty_plan"] or 0) / int(r["uph"])
        horas_rest = _ppy_horas_iniciales(lineas_activas, ocupadas)

        asignaciones = {}
        ia_ok = False
        try:
            asignaciones = _ppy_acomodo_ia_llm(sin_linea, lineas_activas, horas_rest)
            ia_ok = True
        except Exception as e:
            logger.warning("acomodar-ia: IA no disponible, uso heuristica: %r", e)

        # Validar cada asignacion de la IA: linea activa, permitida, y que
        # NO exceda las horas disponibles de la linea (la IA puede equivocarse).
        by_id = {r["id"]: r for r in sin_linea}
        horas_libre = dict(horas_rest)
        validas = {}
        for lote_id, l in asignaciones.items():
            r = by_id.get(lote_id)
            if not r:
                continue
            l = str(l).strip().upper()
            if l not in lineas_activas:
                continue
            if r["permitidas"] and l not in r["permitidas"]:
                continue
            h = (int(r["qty_plan"] or 0) / int(r["uph"])) if r.get("uph") else 0.0
            if h > horas_libre.get(l, 0.0) + 1e-6:
                continue  # excederia la linea: se descarta, va a heuristica
            horas_libre[l] = horas_libre.get(l, 0.0) - h
            validas[lote_id] = l

        if ia_ok and validas:
            metodo = "ia"
        else:
            metodo = "heuristica"

        # Solo si la IA fallo por completo (0 asignaciones validas), la
        # heuristica local intenta acomodar todo. Si la IA acomodo bien, lo
        # que dejo fuera es porque no cabe: NO se fuerza.
        if not validas:
            heur = _ppy_acomodo_heuristico(sin_linea, lineas_activas, horas_rest)
            validas.update(heur)

        conn = get_pooled_connection()
        if conn is None:
            raise RuntimeError("No fue posible obtener conexion MySQL.")
        cursor = get_dict_cursor(conn)
        conn.autocommit(False)
        for lote_id, linea in validas.items():
            cursor.execute(
                "UPDATE lg_lote_plan SET linea = %s, updated_by = %s "
                "WHERE id = %s AND status = 'PENDIENTE'",
                (linea, usuario, lote_id),
            )
        conn.commit()

        return jsonify(
            {"success": True, "asignados": len(validas),
             "sin_asignar": len(sin_linea) - len(validas), "metodo": metodo,
             "lineas_activas": lineas_activas, **_ppy_listado(fecha)}
        )
    except Exception as e:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        logger.error("Error en api_ppy_acomodar_ia: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if cursor is not None:
            try:
                cursor.close()
            except Exception:
                pass
        if conn is not None:
            try:
                conn.autocommit(True)
                conn.close()
            except Exception:
                pass


def _ppy_acomodo_ia_llm(sin_linea, lineas_activas, horas_rest):
    """Pide a la IA la mejor linea por lote. Retorna {id: linea}.

    Lanza excepcion si la IA no esta disponible (sin API key) o falla; el
    caller cae a la heuristica local.
    """
    from app.api.portal.ai_openai import complete_json

    # Solo se le manda a la IA lo que puede caber: lotes con UPH (horas
    # calculables), ordenados por familia, hasta ~el total de horas
    # disponibles + un margen. El resto no cabe igual y se queda sin linea:
    # asi el prompt no crece con faltantes que no se van a acomodar.
    horas_totales = sum(max(horas_rest.get(l, 0.0), 0.0) for l in lineas_activas)
    candidatos = sorted(
        (r for r in sin_linea if r.get("uph")),
        key=lambda r: (_ppy_familia(r["part_no"]), r["part_no"]),
    )
    lotes_payload = []
    acum_horas = 0.0
    for r in candidatos:
        h = round(int(r["qty_plan"] or 0) / int(r["uph"]), 2)
        lotes_payload.append({
            "id": r["id"],
            "familia": _ppy_familia(r["part_no"]),
            "horas": h,
            "lineas": r["permitidas"] or lineas_activas,
        })
        acum_horas += h
        if acum_horas >= horas_totales * 1.2:  # margen para que la IA elija
            break
    if not lotes_payload:
        return {}

    system = (
        "Eres un planificador de produccion. Asignas lotes a lineas de ensamble "
        "balanceando la carga. Reglas ESTRICTAS:\n"
        "1. Cada lote va SOLO a una de sus 'lineas' permitidas.\n"
        "2. La suma de 'horas' de los lotes de una linea no debe exceder sus "
        "'horas_disponibles'. Si no cabe todo, deja lotes sin asignar; no "
        "inventes lineas ni excedas horas.\n"
        "3. Manten juntos en la MISMA linea los lotes de la misma 'familia'.\n"
        "4. Balancea las horas entre lineas.\n"
        "Responde SOLO JSON {\"asignaciones\":[{\"id\":<int>,\"linea\":\"<LINEA>\"}]}, "
        "omitiendo los lotes que no acomodes."
    )
    user = "Acomoda estos lotes y responde en JSON.\n" + json.dumps({
        "horas_disponibles": {l: round(max(horas_rest.get(l, 0.0), 0.0), 2) for l in lineas_activas},
        "lotes": lotes_payload,
    }, ensure_ascii=False)

    resultado = complete_json(system=system, user=user, username="plan_proyectado")
    salida = {}
    for a in (resultado.get("asignaciones") or []):
        try:
            salida[int(a["id"])] = str(a["linea"]).strip().upper()
        except (KeyError, TypeError, ValueError):
            continue
    return salida


@bp.route("/api/plan-proyectado/add", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PPY_PERMISO_BOTON)
def api_ppy_add():
    """Alta manual de un lote PENDIENTE (enriquece desde raw e inventario)."""
    try:
        data = request.get_json(silent=True) or {}
        fecha = _ppy_parse_fecha(data.get("fecha")) or date.today()
        part_no = (data.get("part_no") or "").strip()
        if not part_no:
            return jsonify({"success": False, "error": "part_no requerido"}), 400
        try:
            qty = int(data.get("qty"))
        except (TypeError, ValueError):
            return jsonify({"success": False, "error": "Cantidad invalida"}), 400
        if qty <= 0:
            return jsonify({"success": False, "error": "Cantidad invalida"}), 400

        usuario = session.get("usuario") or "SISTEMA"
        raw_info = (_ppy_datos_raw([part_no])).get(part_no) or {}
        linea = (data.get("linea") or "").strip() or None
        if not linea:
            linea = (raw_info.get("assy_line") or "").strip() or None
        if not linea:
            inv = execute_query(
                "SELECT line FROM lg_part_inventory WHERE part_no = %s",
                (part_no,),
                fetch="one",
            )
            linea = (inv or {}).get("line") or None
        pack = raw_info.get("estandar_pack")
        execute_query(
            "INSERT INTO lg_lote_plan (plan_date, part_no, model, main_sub, linea, "
            "turno, faltante, qty_plan, uph, estandar_pack, created_by) "
            "VALUES (%s, %s, %s, %s, %s, 'DIA', 0, %s, %s, %s, %s)",
            (
                fecha,
                part_no,
                raw_info.get("model"),
                raw_info.get("sub_assy"),
                linea,
                qty,
                _ppy_parse_uph(raw_info.get("uph")),
                (int(pack) if pack else None),
                usuario,
            ),
        )
        return jsonify({"success": True, **_ppy_listado(fecha)})
    except Exception as e:
        logger.error("Error en api_ppy_add: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/plan-proyectado/update", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PPY_PERMISO_BOTON)
def api_ppy_update():
    """Edita un campo del lote. PENDIENTE: qty/linea/turno/comentario/fisico;
    CONFIRMADO: solo fisico y comentario (la qty confirmada ya sumo al S)."""
    try:
        data = request.get_json(silent=True) or {}
        lote_id = data.get("id")
        campo = (data.get("field") or "").strip().lower()
        valor = data.get("value")
        if not lote_id or campo not in ("qty_plan", "linea", "turno", "comentario", "fisico"):
            return jsonify({"success": False, "error": "Campo invalido"}), 400

        row = execute_query(
            "SELECT * FROM lg_lote_plan WHERE id = %s", (lote_id,), fetch="one"
        )
        if not row:
            return jsonify({"success": False, "error": "Lote no encontrado"}), 404
        if row["status"] == "CANCELADO":
            return jsonify({"success": False, "error": "Lote cancelado"}), 400
        if row["status"] == "CONFIRMADO" and campo not in ("fisico", "comentario"):
            return (
                jsonify({"success": False, "error": "Lote confirmado: solo FISICO y comentario son editables"}),
                400,
            )

        if campo == "qty_plan":
            try:
                valor = int(valor)
            except (TypeError, ValueError):
                return jsonify({"success": False, "error": "Cantidad invalida"}), 400
            if valor <= 0:
                return jsonify({"success": False, "error": "Cantidad invalida"}), 400
        elif campo == "fisico":
            if valor in (None, ""):
                valor = None
            else:
                try:
                    valor = int(valor)
                except (TypeError, ValueError):
                    return jsonify({"success": False, "error": "Valor invalido"}), 400
                if valor < 0:
                    return jsonify({"success": False, "error": "Valor invalido"}), 400
        elif campo == "turno":
            valor = (str(valor) or "").strip().upper()
            if valor not in PPY_TURNOS:
                return jsonify({"success": False, "error": "Turno invalido"}), 400
        else:
            valor = (str(valor or "").strip()[:255]) or None
            if campo == "linea" and valor:
                valor = valor.upper()[:10]

        usuario = session.get("usuario") or "SISTEMA"
        # campo validado contra whitelist arriba
        execute_query(
            f"UPDATE lg_lote_plan SET {campo} = %s, updated_by = %s WHERE id = %s",
            (valor, usuario, lote_id),
        )
        fecha = row["plan_date"]
        if isinstance(fecha, datetime):
            fecha = fecha.date()
        return jsonify({"success": True, **_ppy_listado(fecha)})
    except Exception as e:
        logger.error("Error en api_ppy_update: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/plan-proyectado/delete", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PPY_PERMISO_BOTON)
def api_ppy_delete():
    """PENDIENTE se borra; CONFIRMADO se cancela y su qty se resta del S."""
    try:
        data = request.get_json(silent=True) or {}
        lote_id = data.get("id")
        row = execute_query(
            "SELECT * FROM lg_lote_plan WHERE id = %s", (lote_id,), fetch="one"
        )
        if not row:
            return jsonify({"success": False, "error": "Lote no encontrado"}), 404
        fecha = row["plan_date"]
        if isinstance(fecha, datetime):
            fecha = fecha.date()

        if row["status"] == "PENDIENTE":
            execute_query("DELETE FROM lg_lote_plan WHERE id = %s", (lote_id,))
        elif row["status"] == "CONFIRMADO":
            usuario = session.get("usuario") or "SISTEMA"
            execute_query(
                "UPDATE lg_lote_plan SET status = 'CANCELADO', updated_by = %s "
                "WHERE id = %s",
                (usuario, lote_id),
            )
            execute_query(
                "UPDATE lg_schedule_daily SET sched_qty = GREATEST(sched_qty - %s, 0), "
                "updated_by = %s WHERE part_no = %s AND sched_date = %s",
                (int(row["qty_plan"] or 0), usuario, row["part_no"], fecha),
            )
            execute_query(
                "DELETE FROM lg_schedule_daily WHERE part_no = %s AND sched_date = %s "
                "AND sched_qty = 0",
                (row["part_no"], fecha),
            )
        return jsonify({"success": True, **_ppy_listado(fecha)})
    except Exception as e:
        logger.error("Error en api_ppy_delete: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/plan-proyectado/confirm", methods=["POST"])
@login_requerido
@requiere_permiso_dropdown(PP_PERMISO_PAGINA, PP_PERMISO_SECCION, PPY_PERMISO_BOTON)
def api_ppy_confirm():
    """Numera los PENDIENTEs del dia (I<YYYYMMDD>-####) y suma su qty al S.

    Bloquea si alguna linea excede las 9 h sin tiempo extra ("debe cumplirse
    en sus horas"): el usuario ajusta cantidades y reintenta.
    """
    conn = None
    cursor = None
    try:
        data = request.get_json(silent=True) or {}
        fecha = _ppy_parse_fecha(data.get("fecha")) or date.today()
        usuario = session.get("usuario") or "SISTEMA"

        listado = _ppy_listado(fecha)
        excedidas = [l for l in listado["lineas"] if l["excede"]]
        if excedidas:
            detalle = ", ".join(f"{l['linea']}: {l['horas']} h" for l in excedidas)
            return (
                jsonify({"success": False,
                         "error": f"Lineas exceden {PPY_HORAS_TURNO:g} h sin tiempo extra: {detalle}. Ajusta cantidades antes de confirmar."}),
                400,
            )

        conn = get_pooled_connection()
        if conn is None:
            raise RuntimeError("No fue posible obtener conexion MySQL.")
        cursor = get_dict_cursor(conn)
        conn.autocommit(False)

        cursor.execute(
            "SELECT * FROM lg_lote_plan WHERE plan_date = %s AND status = 'PENDIENTE' "
            + _PPY_ORDEN_SQL + " FOR UPDATE",
            (fecha,),
        )
        pendientes = cursor.fetchall() or []
        if not pendientes:
            conn.rollback()
            return jsonify({"success": False, "error": "No hay lotes pendientes para esa fecha."}), 400

        prefijo = f"I{fecha.strftime('%Y%m%d')}"
        cursor.execute(
            "SELECT lot_no FROM lg_lote_plan WHERE lot_no LIKE %s "
            "ORDER BY lot_no DESC LIMIT 1 FOR UPDATE",
            (f"{prefijo}-%",),
        )
        ultimo = cursor.fetchone()
        consecutivo = 0
        if ultimo and ultimo.get("lot_no"):
            try:
                consecutivo = int(str(ultimo["lot_no"]).split("-")[-1])
            except ValueError:
                consecutivo = 0

        sumas_s = {}
        for r in pendientes:
            consecutivo += 1
            cursor.execute(
                "UPDATE lg_lote_plan SET lot_no = %s, status = 'CONFIRMADO', "
                "confirmed_by = %s, confirmed_at = NOW() WHERE id = %s",
                (f"{prefijo}-{consecutivo:04d}", usuario, r["id"]),
            )
            sumas_s[r["part_no"]] = sumas_s.get(r["part_no"], 0) + int(r["qty_plan"] or 0)

        cursor.executemany(
            "INSERT INTO lg_schedule_daily (part_no, sched_date, sched_qty, updated_by) "
            "VALUES (%s, %s, %s, %s) "
            "ON DUPLICATE KEY UPDATE sched_qty = sched_qty + VALUES(sched_qty), "
            "updated_by = VALUES(updated_by)",
            [(p, fecha, q, usuario) for p, q in sumas_s.items()],
        )
        conn.commit()

        return jsonify({"success": True, "confirmados": len(pendientes), **_ppy_listado(fecha)})
    except Exception as e:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        logger.error("Error en api_ppy_confirm: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if cursor is not None:
            try:
                cursor.close()
            except Exception:
                pass
        if conn is not None:
            try:
                conn.autocommit(True)
                conn.close()
            except Exception:
                pass
