"""Endpoints HTTP de Almacen de Embarques.

Migrado desde `app/routes.py` (2026-05-27). 33 rutas + 100+ helpers privados,
~5400 lineas, sin cambios funcionales (copia 1:1, solo `@app.route` -> `@bp.route`).

Modulos cubiertos (sidebar Control de proceso -> Almacen de Embarques):
  - Entradas almacen embarques        (`/almacen-embarques-entradas-ajax`)
  - Salidas almacen embarques         (`/almacen-embarques-salidas-ajax`)
  - Retorno almacen embarques         (`/almacen-embarques-retorno-ajax`)
  - Modificar movimientos embarques   (`/almacen-embarques-movimientos-ajax`)
  - Inventario general embarques      (`/almacen-embarques-inventario-general-ajax`)
  - Catalogo numeros de parte         (`/almacen-embarques-catalogo-ajax`)

NO migrar a este modulo:
  - `init_shipping_material_tables` (vive en `app/api/pda/shipping_material.py`,
    bootstrap de tablas llamado por `startup_init.py`).
  - Las 3 rutas `/control-salida-lineas*` (modulo distinto: "Control de salida
    de lineas", queda en `routes.py`). Consume `_normalizar_texto_embarques_historial`
    y `_exportar_historial_embarques_excel` desde este blueprint via import.

Helpers de auth genericos `_obtener_usuario_display_actual` y
`_validar_password_usuario_actual` viajan con este bloque porque solo se usan
desde embarques (verificado 2026-05-27 con grep).
"""

import csv
import hashlib
import io
import json
import os
import re
import struct
import traceback
import zlib
from datetime import date, datetime, timedelta
from datetime import time as dt_time
from decimal import Decimal
from functools import wraps
from io import BytesIO
from pathlib import Path

import pandas as pd
from flask import (
    Blueprint,
    Response,
    jsonify,
    make_response,
    render_template,
    request,
    send_file,
    session,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.auth_system import auth_system
from app.api.shared.datetime_helpers import obtener_fecha_hora_mexico
from app.config_mysql import get_pooled_connection
from app.db import get_db_connection
from app.db_mysql import execute_query

# Helpers compartidos del modulo PDA shipping (NO mover, son consumidos
# tambien por las apps moviles PDA).
from app.api.pda.shipping_material import (
    SHIPPING_TABLES,
    adjust_shipping_movement_record,
    assign_exit_departure_value,
    delete_shipping_movement_record,
    ensure_inventory_record,
    generate_movement_folio,
    get_departure_history_records,
    get_dict_cursor,
    init_shipping_material_tables,
    normalize_integer,
    normalize_part_number,
    normalize_search,
    rebuild_part_inventory_state,
    to_sql_datetime,
)

# `login_requerido` vive en `app.routes`. Se importa dentro del wrapper
# (tarde) para evitar circular: shared -> routes -> control_proceso -> shared.
def login_requerido(f):
    """Proxy del decorador real definido en `app.routes`."""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        from app import routes as _r

        return _r.login_requerido(f)(*args, **kwargs)

    return decorated_function


def normalize_excel_value(value):
    """Fallback: importa late de `routes.py` para no duplicar la logica."""
    from app import routes as _r

    return _r.normalize_excel_value(value)


def parse_json_summary(value):
    """Fallback: importa late de `routes.py` para no duplicar la logica."""
    from app import routes as _r

    return _r.parse_json_summary(value)


bp = Blueprint("control_proceso_almacen_embarques", __name__)


def _normalizar_numero_embarques_historial(value):
    """Convertir Decimals a int/float legibles para JSON y Excel."""
    if value is None:
        return 0
    if isinstance(value, Decimal):
        try:
            entero = int(value)
            return entero if value == entero else float(value)
        except Exception:
            return float(value)
    return value


def _normalizar_texto_embarques_historial(value):
    """Normalizar valores escalar a texto simple."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value)


def _aplicar_filtros_historial_embarques(sql, params, search_columns):
    """Aplicar filtros comunes de historial para los módulos de embarques."""
    search = request.args.get("search", "").strip()
    fecha_desde = request.args.get("fecha_desde", "").strip()
    fecha_hasta = request.args.get("fecha_hasta", "").strip()

    # Los historiales operativos sólo deben mostrar el periodo vigente.
    # Los movimientos anteriores quedan trazados en el historial de cierres.
    sql += f"""
        AND COALESCE(movement_at, created_at) >= COALESCE(
            (SELECT MAX(closed_at) FROM `{SHIPPING_TABLES['inventory_closures']}`),
            '1000-01-01'
        )
    """

    if fecha_desde:
        sql += " AND DATE(COALESCE(movement_at, created_at)) >= %s"
        params.append(fecha_desde)

    if fecha_hasta:
        sql += " AND DATE(COALESCE(movement_at, created_at)) <= %s"
        params.append(fecha_hasta)

    if search:
        like_value = f"%{search}%"
        sql += " AND (" + " OR ".join(
            [f"COALESCE({column}, '') LIKE %s" for column in search_columns]
        ) + ")"
        params.extend([like_value] * len(search_columns))

    return sql, params


def _obtener_historial_entradas_almacen_embarques(limit=300):
    sql = """
        SELECT
            DATE(COALESCE(movement_at, created_at)) AS fecha,
            DATE_FORMAT(COALESCE(movement_at, created_at), '%%H:%%i:%%s') AS hora,
            entry_folio AS folio,
            part_number,
            quantity AS cantidad,
            previous_quantity,
            new_quantity,
            available_quantity,
            product_model,
            description,
            customer,
            zone_code,
            location_code,
            reference_code,
            batch_no,
            registered_by
        FROM embarques_entrada_material
        WHERE COALESCE(is_fifo_layer_only, 0) = 0
    """
    params = []
    sql, params = _aplicar_filtros_historial_embarques(
        sql,
        params,
        [
            "entry_folio",
            "part_number",
            "product_model",
            "description",
            "customer",
            "zone_code",
            "location_code",
            "reference_code",
            "registered_by",
        ],
    )
    sql += " ORDER BY COALESCE(movement_at, created_at) DESC, id DESC LIMIT %s"
    params.append(limit)

    rows = execute_query(sql, tuple(params), fetch="all") or []
    return [
        {
            "fecha": _normalizar_texto_embarques_historial(row.get("fecha")),
            "hora": _normalizar_texto_embarques_historial(row.get("hora")),
            "folio": _normalizar_texto_embarques_historial(row.get("folio")),
            "part_number": _normalizar_texto_embarques_historial(
                row.get("part_number")
            ),
            "cantidad": _normalizar_numero_embarques_historial(row.get("cantidad")),
            "previous_quantity": _normalizar_numero_embarques_historial(
                row.get("previous_quantity")
            ),
            "new_quantity": _normalizar_numero_embarques_historial(
                row.get("new_quantity")
            ),
            "available_quantity": _normalizar_numero_embarques_historial(
                row.get("available_quantity")
            ),
            "product_model": _normalizar_texto_embarques_historial(
                row.get("product_model")
            ),
            "description": _normalizar_texto_embarques_historial(
                row.get("description")
            ),
            "customer": _normalizar_texto_embarques_historial(row.get("customer")),
            "zone_code": _normalizar_texto_embarques_historial(row.get("zone_code")),
            "location_code": _normalizar_texto_embarques_historial(
                row.get("location_code")
            ),
            "reference_code": _normalizar_texto_embarques_historial(
                row.get("reference_code")
            ),
            "batch_no": _normalizar_texto_embarques_historial(row.get("batch_no")),
            "registered_by": _normalizar_texto_embarques_historial(
                row.get("registered_by")
            ),
        }
        for row in rows
    ]


def _obtener_historial_salidas_almacen_embarques(limit=300):
    sql = """
        SELECT
            id,
            DATE(COALESCE(movement_at, created_at)) AS fecha,
            DATE_FORMAT(COALESCE(movement_at, created_at), '%%H:%%i:%%s') AS hora,
            exit_folio AS folio,
            part_number,
            quantity AS cantidad,
            previous_quantity,
            new_quantity,
            product_model,
            description,
            customer,
            zone_code,
            location_code,
            destination_area,
            departure_code,
            departure_assigned_at,
            departure_assigned_by,
            reason,
            requested_by,
            registered_by
        FROM embarques_salida_material
        WHERE 1=1
    """
    params = []
    sql, params = _aplicar_filtros_historial_embarques(
        sql,
        params,
        [
            "exit_folio",
            "part_number",
            "product_model",
            "description",
            "customer",
            "zone_code",
            "location_code",
            "destination_area",
            "departure_code",
            "departure_assigned_by",
            "reason",
            "requested_by",
            "registered_by",
        ],
    )
    sql += " ORDER BY COALESCE(movement_at, created_at) DESC LIMIT %s"
    params.append(limit)

    rows = execute_query(sql, tuple(params), fetch="all") or []
    return [
        {
            "id": row.get("id"),
            "fecha": _normalizar_texto_embarques_historial(row.get("fecha")),
            "hora": _normalizar_texto_embarques_historial(row.get("hora")),
            "folio": _normalizar_texto_embarques_historial(row.get("folio")),
            "part_number": _normalizar_texto_embarques_historial(
                row.get("part_number")
            ),
            "cantidad": _normalizar_numero_embarques_historial(row.get("cantidad")),
            "previous_quantity": _normalizar_numero_embarques_historial(
                row.get("previous_quantity")
            ),
            "new_quantity": _normalizar_numero_embarques_historial(
                row.get("new_quantity")
            ),
            "product_model": _normalizar_texto_embarques_historial(
                row.get("product_model")
            ),
            "description": _normalizar_texto_embarques_historial(
                row.get("description")
            ),
            "customer": _normalizar_texto_embarques_historial(row.get("customer")),
            "zone_code": _normalizar_texto_embarques_historial(row.get("zone_code")),
            "location_code": _normalizar_texto_embarques_historial(
                row.get("location_code")
            ),
            "destination_area": _normalizar_texto_embarques_historial(
                row.get("destination_area")
            ),
            "departure_code": _normalizar_texto_embarques_historial(
                row.get("departure_code")
            ),
            "departure_assigned_at": _normalizar_texto_embarques_historial(
                row.get("departure_assigned_at")
            ),
            "departure_assigned_by": _normalizar_texto_embarques_historial(
                row.get("departure_assigned_by")
            ),
            "reason": _normalizar_texto_embarques_historial(row.get("reason")),
            "requested_by": _normalizar_texto_embarques_historial(
                row.get("requested_by")
            ),
            "registered_by": _normalizar_texto_embarques_historial(
                row.get("registered_by")
            ),
        }
        for row in rows
    ]


def _obtener_historial_retorno_almacen_embarques(limit=300):
    sql = """
        SELECT
            id,
            DATE(COALESCE(movement_at, created_at)) AS fecha,
            DATE_FORMAT(COALESCE(movement_at, created_at), '%%H:%%i:%%s') AS hora,
            return_folio AS folio,
            part_number,
            return_quantity,
            loss_quantity,
            previous_quantity,
            new_quantity,
            product_model,
            description,
            customer,
            zone_code,
            location_code,
            reason,
            remarks,
            registered_by
        FROM embarques_retorno_material
        WHERE 1=1
    """
    params = []
    sql, params = _aplicar_filtros_historial_embarques(
        sql,
        params,
        [
            "return_folio",
            "part_number",
            "product_model",
            "description",
            "customer",
            "zone_code",
            "location_code",
            "reason",
            "remarks",
            "registered_by",
        ],
    )
    sql += " ORDER BY COALESCE(movement_at, created_at) DESC LIMIT %s"
    params.append(limit)

    rows = execute_query(sql, tuple(params), fetch="all") or []
    return [
        {
            "id": row.get("id"),
            "fecha": _normalizar_texto_embarques_historial(row.get("fecha")),
            "hora": _normalizar_texto_embarques_historial(row.get("hora")),
            "folio": _normalizar_texto_embarques_historial(row.get("folio")),
            "part_number": _normalizar_texto_embarques_historial(
                row.get("part_number")
            ),
            "return_quantity": _normalizar_numero_embarques_historial(
                row.get("return_quantity")
            ),
            "loss_quantity": _normalizar_numero_embarques_historial(
                row.get("loss_quantity")
            ),
            "previous_quantity": _normalizar_numero_embarques_historial(
                row.get("previous_quantity")
            ),
            "new_quantity": _normalizar_numero_embarques_historial(
                row.get("new_quantity")
            ),
            "product_model": _normalizar_texto_embarques_historial(
                row.get("product_model")
            ),
            "description": _normalizar_texto_embarques_historial(
                row.get("description")
            ),
            "customer": _normalizar_texto_embarques_historial(row.get("customer")),
            "zone_code": _normalizar_texto_embarques_historial(row.get("zone_code")),
            "location_code": _normalizar_texto_embarques_historial(
                row.get("location_code")
            ),
            "reason": _normalizar_texto_embarques_historial(row.get("reason")),
            "remarks": _normalizar_texto_embarques_historial(row.get("remarks")),
            "registered_by": _normalizar_texto_embarques_historial(
                row.get("registered_by")
            ),
        }
        for row in rows
    ]


def _obtener_movimientos_editables_almacen_embarques(limit=500):
    limit = min(max(int(limit or 500), 1), 2000)
    tipo = (request.args.get("tipo", "") or "").strip().lower()
    search = (request.args.get("search", "") or "").strip()
    fecha_desde = (request.args.get("fecha_desde", "") or "").strip()
    fecha_hasta = (request.args.get("fecha_hasta", "") or "").strip()

    sql = f"""
        SELECT
            movimiento.movement_type,
            movimiento.movement_label,
            movimiento.record_id,
            DATE(movimiento.movement_timestamp) AS fecha,
            DATE_FORMAT(movimiento.movement_timestamp, '%%H:%%i:%%s') AS hora,
            movimiento.folio,
            movimiento.part_number,
            movimiento.quantity_primary,
            movimiento.quantity_secondary,
            movimiento.product_model,
            movimiento.customer,
            movimiento.zone_code,
            movimiento.location_value,
            movimiento.detail,
            movimiento.departure_code,
            movimiento.registered_by,
            ajuste.adjusted_by AS last_adjusted_by,
            ajuste.adjusted_at AS last_adjusted_at
        FROM (
            SELECT
                'entry' AS movement_type,
                'Entrada' AS movement_label,
                e.id AS record_id,
                COALESCE(e.movement_at, e.created_at) AS movement_timestamp,
                e.entry_folio AS folio,
                e.part_number,
                e.quantity AS quantity_primary,
                NULL AS quantity_secondary,
                e.product_model,
                e.customer,
                e.zone_code,
                e.location_code AS location_value,
                CONCAT_WS(' / ', NULLIF(e.zone_code, ''), NULLIF(e.location_code, '')) AS detail,
                NULL AS departure_code,
                e.registered_by
            FROM `{SHIPPING_TABLES['entries']}` e
            WHERE COALESCE(e.is_fifo_layer_only, 0) = 0

            UNION ALL

            SELECT
                'exit' AS movement_type,
                'Salida' AS movement_label,
                s.id AS record_id,
                COALESCE(s.movement_at, s.created_at) AS movement_timestamp,
                s.exit_folio AS folio,
                s.part_number,
                s.quantity AS quantity_primary,
                NULL AS quantity_secondary,
                s.product_model,
                s.customer,
                s.zone_code,
                s.destination_area AS location_value,
                CONCAT_WS(' / ', NULLIF(s.destination_area, ''), NULLIF(s.reason, '')) AS detail,
                s.departure_code,
                s.registered_by
            FROM `{SHIPPING_TABLES['exits']}` s

            UNION ALL

            SELECT
                'return' AS movement_type,
                'Retorno' AS movement_label,
                r.id AS record_id,
                COALESCE(r.movement_at, r.created_at) AS movement_timestamp,
                r.return_folio AS folio,
                r.part_number,
                r.return_quantity AS quantity_primary,
                r.loss_quantity AS quantity_secondary,
                r.product_model,
                r.customer,
                r.zone_code,
                r.location_code AS location_value,
                CONCAT_WS(' / ', NULLIF(r.reason, ''), NULLIF(r.remarks, '')) AS detail,
                NULL AS departure_code,
                r.registered_by
            FROM `{SHIPPING_TABLES['returns']}` r
        ) movimiento
        LEFT JOIN (
            SELECT
                a.movement_type,
                a.record_id,
                a.adjusted_by,
                a.adjusted_at
            FROM `{SHIPPING_TABLES['movement_adjustments']}` a
            INNER JOIN (
                SELECT movement_type, record_id, MAX(id) AS latest_id
                FROM `{SHIPPING_TABLES['movement_adjustments']}`
                GROUP BY movement_type, record_id
            ) latest
                ON latest.latest_id = a.id
        ) ajuste
            ON ajuste.movement_type = movimiento.movement_type
           AND ajuste.record_id = movimiento.record_id
        WHERE 1 = 1
    """
    params = []

    if tipo in {"entry", "exit", "return"}:
        sql += " AND movimiento.movement_type = %s"
        params.append(tipo)

    if fecha_desde:
        sql += " AND DATE(movimiento.movement_timestamp) >= %s"
        params.append(fecha_desde)

    if fecha_hasta:
        sql += " AND DATE(movimiento.movement_timestamp) <= %s"
        params.append(fecha_hasta)

    if search:
        like_value = f"%{search}%"
        sql += """
            AND (
                COALESCE(movimiento.folio, '') LIKE %s
                OR COALESCE(movimiento.part_number, '') LIKE %s
                OR COALESCE(movimiento.product_model, '') LIKE %s
                OR COALESCE(movimiento.customer, '') LIKE %s
                OR COALESCE(movimiento.zone_code, '') LIKE %s
                OR COALESCE(movimiento.location_value, '') LIKE %s
                OR COALESCE(movimiento.detail, '') LIKE %s
                OR COALESCE(movimiento.departure_code, '') LIKE %s
                OR COALESCE(movimiento.registered_by, '') LIKE %s
            )
        """
        params.extend([like_value] * 9)

    sql += " ORDER BY movimiento.movement_timestamp DESC, movimiento.record_id DESC LIMIT %s"
    params.append(limit)

    rows = execute_query(sql, tuple(params), fetch="all") or []
    return [
        {
            "movement_type": _normalizar_texto_embarques_historial(row.get("movement_type")),
            "movement_label": _normalizar_texto_embarques_historial(row.get("movement_label")),
            "record_id": row.get("record_id"),
            "fecha": _normalizar_texto_embarques_historial(row.get("fecha")),
            "hora": _normalizar_texto_embarques_historial(row.get("hora")),
            "folio": _normalizar_texto_embarques_historial(row.get("folio")),
            "part_number": _normalizar_texto_embarques_historial(row.get("part_number")),
            "quantity_primary": _normalizar_numero_embarques_historial(row.get("quantity_primary")),
            "quantity_secondary": _normalizar_numero_embarques_historial(row.get("quantity_secondary")),
            "product_model": _normalizar_texto_embarques_historial(row.get("product_model")),
            "customer": _normalizar_texto_embarques_historial(row.get("customer")),
            "zone_code": _normalizar_texto_embarques_historial(row.get("zone_code")),
            "location_value": _normalizar_texto_embarques_historial(row.get("location_value")),
            "detail": _normalizar_texto_embarques_historial(row.get("detail")),
            "departure_code": _normalizar_texto_embarques_historial(row.get("departure_code")),
            "registered_by": _normalizar_texto_embarques_historial(row.get("registered_by")),
            "last_adjusted_by": _normalizar_texto_embarques_historial(row.get("last_adjusted_by")),
            "last_adjusted_at": _normalizar_texto_embarques_historial(row.get("last_adjusted_at")),
        }
        for row in rows
    ]


def _obtener_detalle_movimiento_almacen_embarques(movement_type, record_id):
    normalized_type = movement_type.strip().lower()
    record_id = int(record_id)

    if normalized_type == "entry":
        row = execute_query(
            f"""
            SELECT
                id,
                entry_folio AS folio,
                part_number,
                quantity,
                product_model,
                description,
                customer,
                zone_code,
                location_code,
                reference_code,
                batch_no,
                notes,
                registered_by,
                movement_at
            FROM `{SHIPPING_TABLES['entries']}`
            WHERE id = %s
              AND COALESCE(is_fifo_layer_only, 0) = 0
            LIMIT 1
            """,
            (record_id,),
            fetch="one",
        )
        if not row:
            return None
        return {
            "movement_type": "entry",
            "movement_label": "Entrada",
            "record_id": row.get("id"),
            "folio": _normalizar_texto_embarques_historial(row.get("folio")),
            "part_number": _normalizar_texto_embarques_historial(row.get("part_number")),
            "quantity": _normalizar_numero_embarques_historial(row.get("quantity")),
            "product_model": _normalizar_texto_embarques_historial(row.get("product_model")),
            "description": _normalizar_texto_embarques_historial(row.get("description")),
            "customer": _normalizar_texto_embarques_historial(row.get("customer")),
            "zone_code": _normalizar_texto_embarques_historial(row.get("zone_code")),
            "location_code": _normalizar_texto_embarques_historial(row.get("location_code")),
            "reference_code": _normalizar_texto_embarques_historial(row.get("reference_code")),
            "batch_no": _normalizar_texto_embarques_historial(row.get("batch_no")),
            "notes": _normalizar_texto_embarques_historial(row.get("notes")),
            "registered_by": _normalizar_texto_embarques_historial(row.get("registered_by")),
            "movement_at": _normalizar_texto_embarques_historial(row.get("movement_at")),
        }

    if normalized_type == "exit":
        row = execute_query(
            f"""
            SELECT
                id,
                exit_folio AS folio,
                part_number,
                quantity,
                product_model,
                description,
                customer,
                zone_code,
                location_code,
                destination_area,
                departure_code,
                reason,
                requested_by,
                remarks,
                registered_by,
                movement_at
            FROM `{SHIPPING_TABLES['exits']}`
            WHERE id = %s
            LIMIT 1
            """,
            (record_id,),
            fetch="one",
        )
        if not row:
            return None
        return {
            "movement_type": "exit",
            "movement_label": "Salida",
            "record_id": row.get("id"),
            "folio": _normalizar_texto_embarques_historial(row.get("folio")),
            "part_number": _normalizar_texto_embarques_historial(row.get("part_number")),
            "quantity": _normalizar_numero_embarques_historial(row.get("quantity")),
            "product_model": _normalizar_texto_embarques_historial(row.get("product_model")),
            "description": _normalizar_texto_embarques_historial(row.get("description")),
            "customer": _normalizar_texto_embarques_historial(row.get("customer")),
            "zone_code": _normalizar_texto_embarques_historial(row.get("zone_code")),
            "location_code": _normalizar_texto_embarques_historial(row.get("location_code")),
            "destination_area": _normalizar_texto_embarques_historial(row.get("destination_area")),
            "departure_code": _normalizar_texto_embarques_historial(row.get("departure_code")),
            "reason": _normalizar_texto_embarques_historial(row.get("reason")),
            "requested_by": _normalizar_texto_embarques_historial(row.get("requested_by")),
            "remarks": _normalizar_texto_embarques_historial(row.get("remarks")),
            "registered_by": _normalizar_texto_embarques_historial(row.get("registered_by")),
            "movement_at": _normalizar_texto_embarques_historial(row.get("movement_at")),
        }

    if normalized_type == "return":
        row = execute_query(
            f"""
            SELECT
                id,
                return_folio AS folio,
                part_number,
                return_quantity,
                loss_quantity,
                product_model,
                description,
                customer,
                zone_code,
                location_code,
                reason,
                remarks,
                registered_by,
                movement_at
            FROM `{SHIPPING_TABLES['returns']}`
            WHERE id = %s
            LIMIT 1
            """,
            (record_id,),
            fetch="one",
        )
        if not row:
            return None
        return {
            "movement_type": "return",
            "movement_label": "Retorno",
            "record_id": row.get("id"),
            "folio": _normalizar_texto_embarques_historial(row.get("folio")),
            "part_number": _normalizar_texto_embarques_historial(row.get("part_number")),
            "return_quantity": _normalizar_numero_embarques_historial(row.get("return_quantity")),
            "loss_quantity": _normalizar_numero_embarques_historial(row.get("loss_quantity")),
            "product_model": _normalizar_texto_embarques_historial(row.get("product_model")),
            "description": _normalizar_texto_embarques_historial(row.get("description")),
            "customer": _normalizar_texto_embarques_historial(row.get("customer")),
            "zone_code": _normalizar_texto_embarques_historial(row.get("zone_code")),
            "location_code": _normalizar_texto_embarques_historial(row.get("location_code")),
            "reason": _normalizar_texto_embarques_historial(row.get("reason")),
            "remarks": _normalizar_texto_embarques_historial(row.get("remarks")),
            "registered_by": _normalizar_texto_embarques_historial(row.get("registered_by")),
            "movement_at": _normalizar_texto_embarques_historial(row.get("movement_at")),
        }

    return None


def _obtener_inventario_general_almacen_embarques(limit=5000):
    limit = min(max(int(limit or 5000), 1), 20000)
    search = (request.args.get("search", "") or "").strip()
    part_number_key_expr = "CONVERT(%s USING utf8mb4) COLLATE utf8mb4_unicode_ci"

    closure_subquery = f"""
        SELECT
            cierre.part_number,
            {part_number_key_expr % 'cierre.part_number'} AS part_number_key,
            cierre.initial_quantity,
            cierre.closed_at,
            cierre.closure_label
        FROM `{SHIPPING_TABLES['inventory_closures']}` cierre
        INNER JOIN (
            SELECT
                {part_number_key_expr % 'part_number'} AS part_number_key,
                MAX(closed_at) AS latest_closed_at
            FROM `{SHIPPING_TABLES['inventory_closures']}`
            GROUP BY {part_number_key_expr % 'part_number'}
        ) ultimo
            ON ultimo.part_number_key = {part_number_key_expr % 'cierre.part_number'}
           AND ultimo.latest_closed_at = cierre.closed_at
    """

    params = []

    sql = f"""
        SELECT
            i.part_number,
            i.product_model,
            i.customer,
            i.current_quantity,
            cierre.initial_quantity AS closure_initial_quantity,
            cierre.closed_at AS period_start,
            cierre.closure_label,
            COALESCE(entradas.entries_qty, 0) AS entries_qty,
            COALESCE(salidas.exits_qty, 0) AS exits_qty,
            COALESCE(retornos.return_entries_qty, 0) AS return_entries_qty,
            COALESCE(retornos.return_exits_qty, 0) AS return_exits_qty
        FROM `{SHIPPING_TABLES['inventory']}` i
        LEFT JOIN ({closure_subquery}) cierre
            ON cierre.part_number_key = {part_number_key_expr % 'i.part_number'}
        LEFT JOIN (
            SELECT
                {part_number_key_expr % 'e.part_number'} AS part_number_key,
                COALESCE(SUM(e.quantity), 0) AS entries_qty
            FROM `{SHIPPING_TABLES['entries']}` e
            LEFT JOIN ({closure_subquery}) cierre_e
                ON cierre_e.part_number_key = {part_number_key_expr % 'e.part_number'}
            WHERE COALESCE(e.is_fifo_layer_only, 0) = 0
              AND COALESCE(e.movement_at, e.created_at) >= COALESCE(cierre_e.closed_at, '1000-01-01')
            GROUP BY {part_number_key_expr % 'e.part_number'}
        ) entradas
            ON entradas.part_number_key = {part_number_key_expr % 'i.part_number'}
        LEFT JOIN (
            SELECT
                {part_number_key_expr % 's.part_number'} AS part_number_key,
                COALESCE(SUM(s.quantity), 0) AS exits_qty
            FROM `{SHIPPING_TABLES['exits']}` s
            LEFT JOIN ({closure_subquery}) cierre_s
                ON cierre_s.part_number_key = {part_number_key_expr % 's.part_number'}
            WHERE COALESCE(s.movement_at, s.created_at) >= COALESCE(cierre_s.closed_at, '1000-01-01')
            GROUP BY {part_number_key_expr % 's.part_number'}
        ) salidas
            ON salidas.part_number_key = {part_number_key_expr % 'i.part_number'}
        LEFT JOIN (
            SELECT
                {part_number_key_expr % 'r.part_number'} AS part_number_key,
                COALESCE(SUM(r.return_quantity), 0) AS return_entries_qty,
                COALESCE(SUM(r.loss_quantity), 0) AS return_exits_qty
            FROM `{SHIPPING_TABLES['returns']}` r
            LEFT JOIN ({closure_subquery}) cierre_r
                ON cierre_r.part_number_key = {part_number_key_expr % 'r.part_number'}
            WHERE COALESCE(r.movement_at, r.created_at) >= COALESCE(cierre_r.closed_at, '1000-01-01')
            GROUP BY {part_number_key_expr % 'r.part_number'}
        ) retornos
            ON retornos.part_number_key = {part_number_key_expr % 'i.part_number'}
        WHERE 1 = 1
    """

    if search:
        like_value = f"%{search}%"
        sql += """
            AND (
                COALESCE(i.part_number, '') COLLATE utf8mb4_unicode_ci LIKE %s
            )
        """
        params.append(like_value)

    sql += " ORDER BY i.part_number ASC LIMIT %s"
    params.append(limit)

    rows = execute_query(sql, tuple(params), fetch="all") or []
    result_rows = []
    has_closure = False
    latest_period_start = None

    for row in rows:
        current_quantity = _normalizar_numero_embarques_historial(row.get("current_quantity"))
        entries_qty = _normalizar_numero_embarques_historial(row.get("entries_qty"))
        exits_qty = _normalizar_numero_embarques_historial(row.get("exits_qty"))
        return_entries_qty = _normalizar_numero_embarques_historial(row.get("return_entries_qty"))
        return_exits_qty = _normalizar_numero_embarques_historial(row.get("return_exits_qty"))
        closure_initial_quantity = row.get("closure_initial_quantity")
        period_start = _normalizar_texto_embarques_historial(row.get("period_start"))
        closure_label = _normalizar_texto_embarques_historial(row.get("closure_label"))

        if closure_initial_quantity is None:
            initial_quantity = (
                current_quantity
                - entries_qty
                + exits_qty
                - return_entries_qty
                + return_exits_qty
            )
        else:
            initial_quantity = _normalizar_numero_embarques_historial(closure_initial_quantity)
            has_closure = True

        calculated_current_quantity = (
            initial_quantity
            + entries_qty
            - exits_qty
            + return_entries_qty
            - return_exits_qty
        )

        if period_start and (latest_period_start is None or period_start > latest_period_start):
            latest_period_start = period_start

        result_rows.append(
            {
                "part_number": _normalizar_texto_embarques_historial(row.get("part_number")),
                "product_model": _normalizar_texto_embarques_historial(row.get("product_model")),
                "customer": _normalizar_texto_embarques_historial(row.get("customer")),
                "initial_quantity": initial_quantity,
                "entries_qty": entries_qty,
                "exits_qty": exits_qty,
                "return_entries_qty": return_entries_qty,
                "return_exits_qty": return_exits_qty,
                "current_quantity": calculated_current_quantity,
                "period_start": period_start,
                "closure_label": closure_label,
            }
        )

    return {
        "rows": result_rows,
        "summary": {
            "total_items": len(result_rows),
            "has_closure": has_closure,
            "latest_period_start": latest_period_start or "",
        },
    }


CATALOG_FIELD_KEYS = [
    "part_number",
    "product_model",
    "product_status",
    "description",
    "standard_pack",
    "customer",
    "zone_code",
]


def _normalizar_estado_catalogo_embarques(raw_value):
    value = normalize_search(raw_value).lower()
    if value in {"inactivo", "inactive", "baja", "0", "false"}:
        return "inactivo"
    return "activo"


def _normalizar_standard_pack_catalogo_embarques(raw_value):
    parsed = normalize_integer(raw_value)
    if parsed is None or parsed <= 0:
        return 1
    return parsed


def _serializar_catalogo_embarques(row):
    if not row:
        return {}
    return {
        "catalog_id": row.get("catalog_id") or row.get("id"),
        "inventory_id": row.get("inventory_id"),
        "part_number": _normalizar_texto_embarques_historial(row.get("part_number")),
        "product_model": _normalizar_texto_embarques_historial(row.get("product_model")),
        "product_status": _normalizar_texto_embarques_historial(row.get("product_status")),
        "description": _normalizar_texto_embarques_historial(row.get("description")),
        "standard_pack": _normalizar_numero_embarques_historial(row.get("standard_pack")),
        "customer": _normalizar_texto_embarques_historial(row.get("customer")),
        "zone_code": _normalizar_texto_embarques_historial(row.get("zone_code")),
        "current_quantity": _normalizar_numero_embarques_historial(row.get("current_quantity")),
        "entries_count": _normalizar_numero_embarques_historial(row.get("entries_count")),
        "exits_count": _normalizar_numero_embarques_historial(row.get("exits_count")),
        "returns_count": _normalizar_numero_embarques_historial(row.get("returns_count")),
        "departure_count": _normalizar_numero_embarques_historial(row.get("departure_count")),
        "closure_count": _normalizar_numero_embarques_historial(row.get("closure_count")),
    }


def _registrar_auditoria_catalogo_embarques(
    cursor,
    action,
    previous_values,
    new_values,
    changed_fields,
    notes,
):
    previous_values = previous_values or {}
    new_values = new_values or {}
    catalog_id = (
        new_values.get("catalog_id")
        or new_values.get("id")
        or previous_values.get("catalog_id")
        or previous_values.get("id")
    )
    inventory_id = new_values.get("inventory_id") or previous_values.get("inventory_id")
    part_before = previous_values.get("part_number")
    part_after = new_values.get("part_number")
    cursor.execute(
        f"""
        INSERT INTO `{SHIPPING_TABLES['catalog_adjustments']}` (
          catalog_id,
          inventory_id,
          action,
          part_number_before,
          part_number_after,
          previous_values_json,
          new_values_json,
          changed_fields_json,
          notes,
          adjusted_by,
          adjusted_at
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """,
        (
            catalog_id,
            inventory_id,
            action,
            part_before,
            part_after,
            json.dumps(previous_values, ensure_ascii=False, default=str, separators=(",", ":")),
            json.dumps(new_values, ensure_ascii=False, default=str, separators=(",", ":")),
            json.dumps(changed_fields or [], ensure_ascii=False, separators=(",", ":")),
            normalize_search(notes),
            _obtener_usuario_display_actual(),
            obtener_fecha_hora_mexico(),
        ),
    )


def _obtener_catalogo_almacen_embarques(limit=5000):
    limit = min(max(int(limit or 5000), 1), 20000)
    search = normalize_search(request.args.get("search"))
    status = normalize_search(request.args.get("status")).lower()
    params = []
    part_number_key_expr = "CONVERT(%s USING utf8mb4) COLLATE utf8mb4_unicode_ci"
    latest_closure_subquery = f"""
        SELECT
          {part_number_key_expr % 'part_number'} AS part_number_key,
          MAX(closed_at) AS latest_closed_at
        FROM `{SHIPPING_TABLES['inventory_closures']}`
        GROUP BY {part_number_key_expr % 'part_number'}
    """

    sql = f"""
        SELECT
          c.id AS catalog_id,
          i.id AS inventory_id,
          c.part_number,
          c.product_model,
          c.product_status,
          c.description,
          c.standard_pack,
          c.customer,
          c.zone_code,
          COALESCE(i.current_quantity, 0) AS current_quantity,
          COALESCE(entry_counts.total_rows, 0) AS entries_count,
          COALESCE(exit_counts.total_rows, 0) AS exits_count,
          COALESCE(return_counts.total_rows, 0) AS returns_count,
          COALESCE(departure_counts.total_rows, 0) AS departure_count,
          COALESCE(closure_counts.total_rows, 0) AS closure_count,
          COALESCE(entry_period_counts.total_rows, 0) AS period_entries_count,
          COALESCE(exit_period_counts.total_rows, 0) AS period_exits_count,
          COALESCE(return_period_counts.total_rows, 0) AS period_returns_count,
          COALESCE(departure_period_counts.total_rows, 0) AS period_departure_count,
          c.created_at,
          c.updated_at
        FROM `{SHIPPING_TABLES['catalog']}` c
        LEFT JOIN `{SHIPPING_TABLES['inventory']}` i
          ON i.catalog_id = c.id
        LEFT JOIN (
          SELECT catalog_id, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['entries']}`
          GROUP BY catalog_id
        ) entry_counts
          ON entry_counts.catalog_id = c.id
        LEFT JOIN (
          SELECT catalog_id, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['exits']}`
          GROUP BY catalog_id
        ) exit_counts
          ON exit_counts.catalog_id = c.id
        LEFT JOIN (
          SELECT catalog_id, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['returns']}`
          GROUP BY catalog_id
        ) return_counts
          ON return_counts.catalog_id = c.id
        LEFT JOIN (
          SELECT catalog_id, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['departure_history']}`
          GROUP BY catalog_id
        ) departure_counts
          ON departure_counts.catalog_id = c.id
        LEFT JOIN (
          SELECT e.catalog_id, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['entries']}` e
          LEFT JOIN ({latest_closure_subquery}) latest_closure
            ON latest_closure.part_number_key = {part_number_key_expr % 'e.part_number'}
          WHERE e.movement_at >= COALESCE(latest_closure.latest_closed_at, '1000-01-01')
          GROUP BY e.catalog_id
        ) entry_period_counts
          ON entry_period_counts.catalog_id = c.id
        LEFT JOIN (
          SELECT s.catalog_id, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['exits']}` s
          LEFT JOIN ({latest_closure_subquery}) latest_closure
            ON latest_closure.part_number_key = {part_number_key_expr % 's.part_number'}
          WHERE s.movement_at >= COALESCE(latest_closure.latest_closed_at, '1000-01-01')
          GROUP BY s.catalog_id
        ) exit_period_counts
          ON exit_period_counts.catalog_id = c.id
        LEFT JOIN (
          SELECT r.catalog_id, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['returns']}` r
          LEFT JOIN ({latest_closure_subquery}) latest_closure
            ON latest_closure.part_number_key = {part_number_key_expr % 'r.part_number'}
          WHERE r.movement_at >= COALESCE(latest_closure.latest_closed_at, '1000-01-01')
          GROUP BY r.catalog_id
        ) return_period_counts
          ON return_period_counts.catalog_id = c.id
        LEFT JOIN (
          SELECT d.catalog_id, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['departure_history']}` d
          LEFT JOIN ({latest_closure_subquery}) latest_closure
            ON latest_closure.part_number_key = {part_number_key_expr % 'd.part_number'}
          WHERE d.assigned_at >= COALESCE(latest_closure.latest_closed_at, '1000-01-01')
          GROUP BY d.catalog_id
        ) departure_period_counts
          ON departure_period_counts.catalog_id = c.id
        LEFT JOIN (
          SELECT part_number, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['inventory_closures']}`
          GROUP BY part_number
        ) closure_counts
          ON closure_counts.part_number COLLATE utf8mb4_unicode_ci = c.part_number COLLATE utf8mb4_unicode_ci
        WHERE 1 = 1
    """

    if search:
        like_value = f"%{search}%"
        sql += """
          AND (
            COALESCE(c.part_number, '') COLLATE utf8mb4_unicode_ci LIKE %s
            OR COALESCE(c.product_model, '') COLLATE utf8mb4_unicode_ci LIKE %s
            OR COALESCE(c.description, '') COLLATE utf8mb4_unicode_ci LIKE %s
            OR COALESCE(c.customer, '') COLLATE utf8mb4_unicode_ci LIKE %s
          )
        """
        params.extend([like_value, like_value, like_value, like_value])

    if status in {"activo", "inactivo"}:
        sql += " AND c.product_status = %s"
        params.append(status)

    sql += " ORDER BY c.part_number ASC LIMIT %s"
    params.append(limit)

    rows = execute_query(sql, tuple(params), fetch="all") or []
    result_rows = []
    for row in rows:
        current_quantity = _normalizar_numero_embarques_historial(row.get("current_quantity"))
        movement_count = sum(
            _normalizar_numero_embarques_historial(row.get(key))
            for key in (
                "entries_count",
                "exits_count",
                "returns_count",
                "departure_count",
                "closure_count",
            )
        )
        period_movement_count = sum(
            _normalizar_numero_embarques_historial(row.get(key))
            for key in (
                "period_entries_count",
                "period_exits_count",
                "period_returns_count",
                "period_departure_count",
            )
        )
        result_rows.append(
            {
                "catalog_id": row.get("catalog_id"),
                "inventory_id": row.get("inventory_id"),
                "part_number": _normalizar_texto_embarques_historial(row.get("part_number")),
                "product_model": _normalizar_texto_embarques_historial(row.get("product_model")),
                "product_status": _normalizar_texto_embarques_historial(row.get("product_status")),
                "description": _normalizar_texto_embarques_historial(row.get("description")),
                "standard_pack": _normalizar_numero_embarques_historial(row.get("standard_pack")) or 1,
                "customer": _normalizar_texto_embarques_historial(row.get("customer")),
                "zone_code": _normalizar_texto_embarques_historial(row.get("zone_code")),
                "current_quantity": current_quantity,
                "movement_count": movement_count,
                "period_movement_count": period_movement_count,
                "entries_count": _normalizar_numero_embarques_historial(row.get("entries_count")),
                "exits_count": _normalizar_numero_embarques_historial(row.get("exits_count")),
                "returns_count": _normalizar_numero_embarques_historial(row.get("returns_count")),
                "departure_count": _normalizar_numero_embarques_historial(row.get("departure_count")),
                "closure_count": _normalizar_numero_embarques_historial(row.get("closure_count")),
                "delete_mode": (
                    "blocked_stock"
                    if current_quantity != 0
                    else "blocked_movements"
                    if period_movement_count > 0
                    else "soft_delete"
                    if movement_count > 0
                    else "hard_delete"
                ),
                "created_at": _normalizar_texto_embarques_historial(row.get("created_at")),
                "updated_at": _normalizar_texto_embarques_historial(row.get("updated_at")),
            }
        )

    return {"rows": result_rows, "summary": {"total_items": len(result_rows)}}


def _obtener_catalogo_detalle_almacen_embarques(cursor, catalog_id, for_update=False):
    lock_clause = " FOR UPDATE" if for_update else ""
    part_number_key_expr = "CONVERT(%s USING utf8mb4) COLLATE utf8mb4_unicode_ci"
    latest_closure_subquery = f"""
        SELECT
          {part_number_key_expr % 'part_number'} AS part_number_key,
          MAX(closed_at) AS latest_closed_at
        FROM `{SHIPPING_TABLES['inventory_closures']}`
        GROUP BY {part_number_key_expr % 'part_number'}
    """
    cursor.execute(
        f"""
        SELECT
          c.id AS catalog_id,
          i.id AS inventory_id,
          c.part_number,
          c.product_model,
          c.product_status,
          c.description,
          c.standard_pack,
          c.customer,
          c.zone_code,
          COALESCE(i.current_quantity, 0) AS current_quantity,
          COALESCE(entry_counts.total_rows, 0) AS entries_count,
          COALESCE(exit_counts.total_rows, 0) AS exits_count,
          COALESCE(return_counts.total_rows, 0) AS returns_count,
          COALESCE(departure_counts.total_rows, 0) AS departure_count,
          COALESCE(closure_counts.total_rows, 0) AS closure_count,
          COALESCE(entry_period_counts.total_rows, 0) AS period_entries_count,
          COALESCE(exit_period_counts.total_rows, 0) AS period_exits_count,
          COALESCE(return_period_counts.total_rows, 0) AS period_returns_count,
          COALESCE(departure_period_counts.total_rows, 0) AS period_departure_count
        FROM `{SHIPPING_TABLES['catalog']}` c
        LEFT JOIN `{SHIPPING_TABLES['inventory']}` i
          ON i.catalog_id = c.id
        LEFT JOIN (
          SELECT catalog_id, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['entries']}`
          GROUP BY catalog_id
        ) entry_counts
          ON entry_counts.catalog_id = c.id
        LEFT JOIN (
          SELECT catalog_id, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['exits']}`
          GROUP BY catalog_id
        ) exit_counts
          ON exit_counts.catalog_id = c.id
        LEFT JOIN (
          SELECT catalog_id, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['returns']}`
          GROUP BY catalog_id
        ) return_counts
          ON return_counts.catalog_id = c.id
        LEFT JOIN (
          SELECT catalog_id, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['departure_history']}`
          GROUP BY catalog_id
        ) departure_counts
          ON departure_counts.catalog_id = c.id
        LEFT JOIN (
          SELECT e.catalog_id, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['entries']}` e
          LEFT JOIN ({latest_closure_subquery}) latest_closure
            ON latest_closure.part_number_key = {part_number_key_expr % 'e.part_number'}
          WHERE e.movement_at >= COALESCE(latest_closure.latest_closed_at, '1000-01-01')
          GROUP BY e.catalog_id
        ) entry_period_counts
          ON entry_period_counts.catalog_id = c.id
        LEFT JOIN (
          SELECT s.catalog_id, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['exits']}` s
          LEFT JOIN ({latest_closure_subquery}) latest_closure
            ON latest_closure.part_number_key = {part_number_key_expr % 's.part_number'}
          WHERE s.movement_at >= COALESCE(latest_closure.latest_closed_at, '1000-01-01')
          GROUP BY s.catalog_id
        ) exit_period_counts
          ON exit_period_counts.catalog_id = c.id
        LEFT JOIN (
          SELECT r.catalog_id, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['returns']}` r
          LEFT JOIN ({latest_closure_subquery}) latest_closure
            ON latest_closure.part_number_key = {part_number_key_expr % 'r.part_number'}
          WHERE r.movement_at >= COALESCE(latest_closure.latest_closed_at, '1000-01-01')
          GROUP BY r.catalog_id
        ) return_period_counts
          ON return_period_counts.catalog_id = c.id
        LEFT JOIN (
          SELECT d.catalog_id, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['departure_history']}` d
          LEFT JOIN ({latest_closure_subquery}) latest_closure
            ON latest_closure.part_number_key = {part_number_key_expr % 'd.part_number'}
          WHERE d.assigned_at >= COALESCE(latest_closure.latest_closed_at, '1000-01-01')
          GROUP BY d.catalog_id
        ) departure_period_counts
          ON departure_period_counts.catalog_id = c.id
        LEFT JOIN (
          SELECT part_number, COUNT(*) AS total_rows
          FROM `{SHIPPING_TABLES['inventory_closures']}`
          GROUP BY part_number
        ) closure_counts
          ON closure_counts.part_number COLLATE utf8mb4_unicode_ci = c.part_number COLLATE utf8mb4_unicode_ci
        WHERE c.id = %s
        LIMIT 1{lock_clause}
        """,
        (catalog_id,),
    )
    return cursor.fetchone()


def _validar_part_number_catalogo_unico(cursor, part_number, catalog_id=None, inventory_id=None):
    cursor.execute(
        f"""
        SELECT id
        FROM `{SHIPPING_TABLES['catalog']}`
        WHERE part_number = %s
          AND (%s IS NULL OR id <> %s)
        LIMIT 1
        """,
        (part_number, catalog_id, catalog_id),
    )
    if cursor.fetchone():
        return False, f"El número de parte {part_number} ya existe en el catálogo."

    cursor.execute(
        f"""
        SELECT id
        FROM `{SHIPPING_TABLES['inventory']}`
        WHERE part_number = %s
          AND (%s IS NULL OR id <> %s)
        LIMIT 1
        """,
        (part_number, inventory_id, inventory_id),
    )
    if cursor.fetchone():
        return False, f"El número de parte {part_number} ya existe en inventario actual."

    return True, ""


def _construir_payload_catalogo_embarques(data, current=None):
    part_number = normalize_part_number(data.get("partNumber") or data.get("part_number"))
    if not part_number and current:
        part_number = normalize_part_number(current.get("part_number"))

    return {
        "part_number": part_number,
        "product_model": normalize_search(data.get("productModel") or data.get("product_model")),
        "product_status": _normalizar_estado_catalogo_embarques(
            data.get("productStatus") or data.get("product_status") or "activo"
        ),
        "description": normalize_search(data.get("description")),
        "standard_pack": _normalizar_standard_pack_catalogo_embarques(
            data.get("standardPack") or data.get("standard_pack")
        ),
        "customer": normalize_search(data.get("customer")) or "LG",
        "zone_code": normalize_search(data.get("zoneCode") or data.get("zone_code")) or "pending",
    }


def _crear_catalogo_almacen_embarques(data):
    payload = _construir_payload_catalogo_embarques(data)
    if not payload["part_number"]:
        return {"success": False, "error": "El número de parte es obligatorio."}, 400

    conn = get_pooled_connection()
    if conn is None:
        raise RuntimeError("No fue posible obtener conexión MySQL.")

    cursor = get_dict_cursor(conn)
    try:
        conn.autocommit(False)
        is_unique, error_message = _validar_part_number_catalogo_unico(
            cursor,
            payload["part_number"],
        )
        if not is_unique:
            conn.rollback()
            return {"success": False, "error": error_message}, 409

        cursor.execute(
            f"""
            INSERT INTO `{SHIPPING_TABLES['catalog']}` (
              part_number,
              product_model,
              product_status,
              description,
              standard_pack,
              customer,
              zone_code,
              source_file
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, 'portal_catalogo_embarques')
            """,
            (
                payload["part_number"],
                payload["product_model"],
                payload["product_status"],
                payload["description"],
                payload["standard_pack"],
                payload["customer"],
                payload["zone_code"],
            ),
        )
        catalog_id = cursor.lastrowid
        cursor.execute(
            f"""
            INSERT INTO `{SHIPPING_TABLES['inventory']}` (
              catalog_id,
              part_number,
              product_model,
              product_status,
              description,
              customer,
              zone_code,
              standard_pack,
              current_quantity,
              catalog_loaded_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 0, %s)
            """,
            (
                catalog_id,
                payload["part_number"],
                payload["product_model"],
                payload["product_status"],
                payload["description"],
                payload["customer"],
                payload["zone_code"],
                payload["standard_pack"],
                obtener_fecha_hora_mexico(),
            ),
        )
        inventory_id = cursor.lastrowid
        new_values = {
            "catalog_id": catalog_id,
            "inventory_id": inventory_id,
            **payload,
            "current_quantity": 0,
        }
        _registrar_auditoria_catalogo_embarques(
            cursor,
            "create",
            {},
            new_values,
            list(payload.keys()),
            data.get("notes") or "Alta desde portal",
        )
        conn.commit()
        return {
            "success": True,
            "message": "Número de parte agregado correctamente.",
            "row": new_values,
        }, 201
    except Exception:
        conn.rollback()
        raise
    finally:
        try:
            conn.autocommit(True)
        except Exception:
            pass
        cursor.close()
        conn.close()


def _actualizar_catalogo_almacen_embarques(catalog_id, data):
    notes = normalize_search(data.get("notes"))
    if not notes:
        return {"success": False, "error": "El motivo del cambio es obligatorio."}, 400

    conn = get_pooled_connection()
    if conn is None:
        raise RuntimeError("No fue posible obtener conexión MySQL.")

    cursor = get_dict_cursor(conn)
    try:
        conn.autocommit(False)
        current = _obtener_catalogo_detalle_almacen_embarques(cursor, catalog_id, True)
        if not current:
            conn.rollback()
            return {"success": False, "error": "Número de parte no encontrado."}, 404

        previous_values = _serializar_catalogo_embarques(current)
        payload = _construir_payload_catalogo_embarques(data, current)
        if not payload["part_number"]:
            conn.rollback()
            return {"success": False, "error": "El número de parte es obligatorio."}, 400

        is_unique, error_message = _validar_part_number_catalogo_unico(
            cursor,
            payload["part_number"],
            catalog_id=current.get("catalog_id"),
            inventory_id=current.get("inventory_id"),
        )
        if not is_unique:
            conn.rollback()
            return {"success": False, "error": error_message}, 409

        changed_fields = []
        for key in CATALOG_FIELD_KEYS:
            previous_value = previous_values.get(key)
            next_value = payload.get(key)
            if key == "standard_pack":
                previous_value = _normalizar_numero_embarques_historial(previous_value) or 1
            if normalize_search(previous_value) != normalize_search(next_value):
                changed_fields.append(key)

        if not changed_fields:
            conn.rollback()
            return {"success": False, "error": "No hay cambios para guardar."}, 400

        cursor.execute(
            f"""
            UPDATE `{SHIPPING_TABLES['catalog']}`
            SET part_number = %s,
                product_model = %s,
                product_status = %s,
                description = %s,
                standard_pack = %s,
                customer = %s,
                zone_code = %s
            WHERE id = %s
            """,
            (
                payload["part_number"],
                payload["product_model"],
                payload["product_status"],
                payload["description"],
                payload["standard_pack"],
                payload["customer"],
                payload["zone_code"],
                current.get("catalog_id"),
            ),
        )

        cursor.execute(
            f"""
            UPDATE `{SHIPPING_TABLES['inventory']}`
            SET part_number = %s,
                product_model = %s,
                product_status = %s,
                description = %s,
                customer = %s,
                zone_code = %s,
                standard_pack = %s
            WHERE catalog_id = %s
            """,
            (
                payload["part_number"],
                payload["product_model"],
                payload["product_status"],
                payload["description"],
                payload["customer"],
                payload["zone_code"],
                payload["standard_pack"],
                current.get("catalog_id"),
            ),
        )

        if "part_number" in changed_fields:
            for table_key in ("entries", "exits", "returns", "departure_history"):
                cursor.execute(
                    f"""
                    UPDATE `{SHIPPING_TABLES[table_key]}`
                    SET part_number = %s
                    WHERE catalog_id = %s
                    """,
                    (payload["part_number"], current.get("catalog_id")),
                )
            cursor.execute(
                f"""
                UPDATE `{SHIPPING_TABLES['inventory_closures']}`
                SET part_number = %s
                WHERE part_number = %s
                """,
                (payload["part_number"], previous_values.get("part_number")),
            )
            cursor.execute(
                f"""
                UPDATE `{SHIPPING_TABLES['manual_adjustment_items']}`
                SET part_number = %s
                WHERE part_number = %s
                """,
                (payload["part_number"], previous_values.get("part_number")),
            )

        next_values = {
            **previous_values,
            **payload,
            "catalog_id": current.get("catalog_id"),
            "inventory_id": current.get("inventory_id"),
            "current_quantity": previous_values.get("current_quantity"),
        }
        _registrar_auditoria_catalogo_embarques(
            cursor,
            "update",
            previous_values,
            next_values,
            changed_fields,
            notes,
        )
        conn.commit()
        return {
            "success": True,
            "message": "Número de parte actualizado correctamente.",
            "changedFields": changed_fields,
        }, 200
    except Exception:
        conn.rollback()
        raise
    finally:
        try:
            conn.autocommit(True)
        except Exception:
            pass
        cursor.close()
        conn.close()


def _eliminar_catalogo_almacen_embarques(catalog_id, data):
    valid_password, password_error = _validar_password_usuario_actual(data.get("password"))
    if not valid_password:
        return {"success": False, "error": password_error}, 403

    notes = normalize_search(data.get("notes"))
    if not notes:
        return {"success": False, "error": "El comentario de eliminación es obligatorio."}, 400

    conn = get_pooled_connection()
    if conn is None:
        raise RuntimeError("No fue posible obtener conexión MySQL.")

    cursor = get_dict_cursor(conn)
    try:
        conn.autocommit(False)
        current = _obtener_catalogo_detalle_almacen_embarques(cursor, catalog_id, True)
        if not current:
            conn.rollback()
            return {"success": False, "error": "Número de parte no encontrado."}, 404

        previous_values = _serializar_catalogo_embarques(current)
        current_quantity = _normalizar_numero_embarques_historial(current.get("current_quantity"))
        movement_count = sum(
            _normalizar_numero_embarques_historial(current.get(key))
            for key in (
                "entries_count",
                "exits_count",
                "returns_count",
                "departure_count",
                "closure_count",
            )
        )
        period_movement_count = sum(
            _normalizar_numero_embarques_historial(current.get(key))
            for key in (
                "period_entries_count",
                "period_exits_count",
                "period_returns_count",
                "period_departure_count",
            )
        )

        if current_quantity != 0:
            conn.rollback()
            return {
                "success": False,
                "error": (
                    "Este número de parte tiene stock. No se puede eliminar hasta haber "
                    "consumido el total de inventario."
                ),
            }, 409

        if period_movement_count > 0:
            conn.rollback()
            return {
                "success": False,
                "error": (
                    "Este número de parte tiene movimientos registrados. No se puede "
                    "eliminar hasta haber hecho el cierre mensual."
                ),
            }, 409

        if movement_count > 0:
            cursor.execute(
                f"""
                UPDATE `{SHIPPING_TABLES['catalog']}`
                SET product_status = 'inactivo'
                WHERE id = %s
                """,
                (current.get("catalog_id"),),
            )
            cursor.execute(
                f"""
                UPDATE `{SHIPPING_TABLES['inventory']}`
                SET product_status = 'inactivo'
                WHERE catalog_id = %s
                """,
                (current.get("catalog_id"),),
            )
            new_values = dict(previous_values)
            new_values["product_status"] = "inactivo"
            _registrar_auditoria_catalogo_embarques(
                cursor,
                "soft_delete",
                previous_values,
                new_values,
                ["product_status"],
                notes,
            )
            conn.commit()
            return {
                "success": True,
                "message": "Número de parte eliminado del catálogo activo.",
                "deleteMode": "soft_delete",
            }, 200

        cursor.execute(
            f"DELETE FROM `{SHIPPING_TABLES['inventory']}` WHERE catalog_id = %s",
            (current.get("catalog_id"),),
        )
        cursor.execute(
            f"DELETE FROM `{SHIPPING_TABLES['catalog']}` WHERE id = %s",
            (current.get("catalog_id"),),
        )
        _registrar_auditoria_catalogo_embarques(
            cursor,
            "hard_delete",
            previous_values,
            {},
            ["deleted"],
            notes,
        )
        conn.commit()
        return {
            "success": True,
            "message": "Número de parte eliminado correctamente.",
            "deleteMode": "hard_delete",
        }, 200
    except Exception:
        conn.rollback()
        raise
    finally:
        try:
            conn.autocommit(True)
        except Exception:
            pass
        cursor.close()
        conn.close()


def _obtener_usuario_display_actual():
    return (
        session.get("nombre_completo")
        or session.get("usuario")
        or "Sistema"
    ).strip()


def _validar_password_usuario_actual(raw_password):
    usuario = (session.get("usuario") or "").strip()
    password = str(raw_password or "").strip()

    if not usuario:
        return False, "No se encontro un usuario valido en sesion"
    if not password:
        return False, "Debes confirmar tu contraseña actual"

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if conn is None:
            raise RuntimeError("No se pudo obtener conexion a la base de datos")

        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT password_hash
            FROM usuarios_sistema
            WHERE username = %s
            LIMIT 1
            """,
            (usuario,),
        )
        usuario_row = cursor.fetchone()
        if isinstance(usuario_row, dict):
            password_hash_actual = usuario_row.get("password_hash") or ""
        elif usuario_row:
            password_hash_actual = usuario_row[0] or ""
        else:
            password_hash_actual = ""

        if not password_hash_actual:
            return False, "No fue posible validar tu contraseña actual"
        if password_hash_actual != auth_system.hash_password(password):
            return False, "La contraseña es incorrecta"

        return True, ""
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


MESES_CIERRE_EMBARQUES = [
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]


def _parse_datetime_cierre_embarques(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, dt_time.min)

    raw_value = str(value or "").strip()
    if not raw_value:
        return None

    for parser in (
        lambda raw: datetime.fromisoformat(raw.replace("Z", "+00:00")),
        lambda raw: datetime.strptime(raw, "%Y-%m-%d %H:%M:%S"),
        lambda raw: datetime.strptime(raw, "%Y-%m-%d"),
    ):
        try:
            return parser(raw_value).replace(tzinfo=None)
        except ValueError:
            continue
    return None


def _parse_month_key_cierre_embarques(value):
    match = re.match(r"^(\d{4})-(\d{2})$", str(value or "").strip())
    if not match:
        return None
    year = int(match.group(1))
    month = int(match.group(2))
    if month < 1 or month > 12:
        return None
    return year, month


def _sumar_meses_cierre_embarques(year, month, delta):
    month_index = (year * 12) + (month - 1) + delta
    return month_index // 12, (month_index % 12) + 1


def _month_key_cierre_embarques(year, month):
    return f"{year:04d}-{month:02d}"


def _month_label_cierre_embarques(year, month):
    return f"{MESES_CIERRE_EMBARQUES[month - 1]} {year}"


def _resolver_periodo_cierre_embarques(batch=None, previous_closed_at=None, reference_at=None):
    """
    El campo closed_at marca el arranque del siguiente periodo operativo.
    El nombre del cierre debe corresponder al periodo cerrado, no al mes de closed_at.
    """
    batch = batch or {}
    closed_at = _parse_datetime_cierre_embarques(batch.get("closed_at")) or _parse_datetime_cierre_embarques(reference_at)
    if closed_at is None:
        closed_at = obtener_fecha_hora_mexico()

    previous_dt = _parse_datetime_cierre_embarques(previous_closed_at)
    raw_month = _parse_month_key_cierre_embarques(batch.get("closure_month"))

    if previous_dt:
        period_year, period_month = previous_dt.year, previous_dt.month
        period_start = previous_dt
    elif raw_month:
        period_year, period_month = raw_month
        if (
            closed_at
            and period_year == closed_at.year
            and period_month == closed_at.month
            and closed_at.day <= 7
        ):
            period_year, period_month = _sumar_meses_cierre_embarques(
                closed_at.year,
                closed_at.month,
                -1,
            )
        period_start = datetime(period_year, period_month, 1)
    else:
        period_year, period_month = closed_at.year, closed_at.month
        if closed_at.day <= 7:
            period_year, period_month = _sumar_meses_cierre_embarques(
                closed_at.year,
                closed_at.month,
                -1,
            )
        period_start = datetime(period_year, period_month, 1)

    period_end_inclusive = closed_at - timedelta(seconds=1) if closed_at else None
    month_label = _month_label_cierre_embarques(period_year, period_month)
    return {
        "closure_month": _month_key_cierre_embarques(period_year, period_month),
        "closure_month_label": month_label,
        "closure_label": f"Cierre {month_label}",
        "period_start": period_start,
        "period_end_exclusive": closed_at,
        "period_end_inclusive": period_end_inclusive,
    }


def _aplicar_metadata_periodo_cierre_embarques(payload, period_info):
    payload = payload or {}
    metadata = payload.setdefault("metadata", {})
    metadata["closureMonth"] = period_info["closure_month"]
    metadata["closureMonthLabel"] = period_info["closure_month_label"]
    metadata["closureLabel"] = period_info["closure_label"]
    return payload


def _obtener_contexto_cierre_inventario_embarques():
    fecha_actual = obtener_fecha_hora_mexico()
    latest_closure = None
    try:
        latest_closure = execute_query(
            f"""
            SELECT closed_at
            FROM `{SHIPPING_TABLES['inventory_closure_batches']}`
            WHERE status = 'confirmed'
            ORDER BY closed_at DESC, id DESC
            LIMIT 1
            """,
            fetch="one",
        )
    except Exception:
        latest_closure = None

    period_info = _resolver_periodo_cierre_embarques(
        {"closed_at": fecha_actual},
        previous_closed_at=(latest_closure or {}).get("closed_at"),
    )
    return {
        "closureDate": fecha_actual.strftime("%Y-%m-%d"),
        "closureDateTime": fecha_actual.strftime("%Y-%m-%d %H:%M:%S"),
        "closureMonth": period_info["closure_month"],
        "closureMonthLabel": period_info["closure_month_label"],
        "closureLabel": period_info["closure_label"],
        "closureUser": _obtener_usuario_display_actual(),
    }


def _normalizar_header_csv_cierre_embarques(value):
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _construir_preview_cierre_inventario_embarques(current_rows, csv_quantities=None):
    preview_rows = []
    differing_parts = []
    negative_difference_parts = []
    matching_rows = 0
    total_rows = len(current_rows)

    csv_quantities = csv_quantities or {}

    for row in current_rows:
        part_number = _normalizar_texto_embarques_historial(row.get("part_number"))
        system_quantity = _normalizar_numero_embarques_historial(
            row.get("current_quantity")
        )
        csv_current_qty = csv_quantities.get(part_number)
        difference_qty = (
            csv_current_qty - system_quantity
            if csv_current_qty is not None
            else None
        )
        applied_initial_qty = (
            max(csv_current_qty, 0) if csv_current_qty is not None else None
        )

        if csv_current_qty is not None:
            if difference_qty == 0:
                matching_rows += 1
            else:
                differing_parts.append(part_number)
                if difference_qty < 0:
                    negative_difference_parts.append(part_number)

        preview_rows.append(
            {
                "part_number": part_number,
                "product_model": _normalizar_texto_embarques_historial(
                    row.get("product_model")
                ),
                "customer": _normalizar_texto_embarques_historial(row.get("customer")),
                "system_quantity": system_quantity,
                "csv_current_qty": csv_current_qty,
                "difference_quantity": difference_qty,
                "applied_initial_quantity": applied_initial_qty,
                "status": (
                    "pendiente"
                    if csv_current_qty is None
                    else "igual"
                    if difference_qty == 0
                    else "diferencia"
                ),
            }
        )

    accuracy_pct = (
        round((matching_rows / total_rows) * 100, 2) if total_rows and csv_quantities else 0
    )
    serializable_rows = [
        {
            "part_number": row["part_number"],
            "system_quantity": row["system_quantity"],
            "csv_current_qty": row["csv_current_qty"],
            "difference_quantity": row["difference_quantity"],
            "applied_initial_quantity": row["applied_initial_quantity"],
        }
        for row in preview_rows
    ]
    rows_hash = hashlib.sha256(
        json.dumps(
            serializable_rows,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()

    return {
        "rows": preview_rows,
        "summary": {
            "totalRows": total_rows,
            "matchingRows": matching_rows,
            "differenceRows": len(differing_parts),
            "negativeDifferenceRows": len(negative_difference_parts),
            "accuracyPct": accuracy_pct,
            "differencePartNumbers": differing_parts,
            "negativeDifferencePartNumbers": negative_difference_parts,
            "rowsHash": rows_hash,
        },
    }


def _parsear_csv_cierre_inventario_embarques(file_storage, expected_part_numbers):
    raw_bytes = file_storage.read()
    csv_hash = hashlib.sha256(raw_bytes).hexdigest()

    try:
        decoded = raw_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            decoded = raw_bytes.decode("latin-1")
        except UnicodeDecodeError as exc:
            return {
                "valid": False,
                "errors": [f"No fue posible leer el CSV: {exc}"],
                "csvHash": csv_hash,
                "rows": {},
            }

    reader = csv.reader(io.StringIO(decoded))
    try:
        headers = next(reader)
    except StopIteration:
        return {
            "valid": False,
            "errors": ["El archivo CSV está vacío."],
            "csvHash": csv_hash,
            "rows": {},
        }

    normalized_headers = [_normalizar_header_csv_cierre_embarques(h) for h in headers]
    if set(normalized_headers) != {"part_number", "current_qty"} or len(normalized_headers) != 2:
        return {
            "valid": False,
            "errors": [
                "El CSV debe contener exactamente las columnas part_number y current_qty."
            ],
            "csvHash": csv_hash,
            "rows": {},
        }

    part_idx = normalized_headers.index("part_number")
    qty_idx = normalized_headers.index("current_qty")

    parsed_rows = {}
    errors = []

    for line_number, row in enumerate(reader, start=2):
        if not row or not any(str(cell or "").strip() for cell in row):
            continue

        if len(row) <= max(part_idx, qty_idx):
            errors.append(
                f"Línea {line_number}: faltan columnas requeridas para part_number y current_qty."
            )
            continue

        part_number = _normalizar_texto_embarques_historial(row[part_idx]).upper()
        qty_raw = str(row[qty_idx] or "").strip()

        if not part_number:
            errors.append(f"Línea {line_number}: part_number es obligatorio.")
            continue

        if qty_raw == "":
            errors.append(
                f"Línea {line_number}: current_qty no puede estar vacío. Usa 0 si no hay físico."
            )
            continue

        if not re.fullmatch(r"-?\d+", qty_raw):
            errors.append(
                f"Línea {line_number}: current_qty debe ser un entero para {part_number}."
            )
            continue

        current_qty = int(qty_raw)
        if current_qty < 0:
            errors.append(
                f"Línea {line_number}: current_qty no permite negativos para {part_number}."
            )
            continue

        if part_number in parsed_rows:
            errors.append(f"Línea {line_number}: part_number duplicado {part_number}.")
            continue

        parsed_rows[part_number] = current_qty

    expected_set = set(expected_part_numbers)
    uploaded_set = set(parsed_rows.keys())
    missing_parts = sorted(expected_set - uploaded_set)
    extra_parts = sorted(uploaded_set - expected_set)

    if missing_parts:
        muestra = ", ".join(missing_parts[:8])
        suffix = "..." if len(missing_parts) > 8 else ""
        errors.append(
            f"Faltan {len(missing_parts)} números de parte del catálogo completo. Ejemplo: {muestra}{suffix}"
        )

    if extra_parts:
        muestra = ", ".join(extra_parts[:8])
        suffix = "..." if len(extra_parts) > 8 else ""
        errors.append(
            f"El CSV contiene {len(extra_parts)} números de parte fuera del catálogo actual. Ejemplo: {muestra}{suffix}"
        )

    return {
        "valid": not errors,
        "errors": errors,
        "csvHash": csv_hash,
        "rows": parsed_rows,
    }


MANUAL_SHIPPING_ADJUSTMENT_CONFIG = {
    "entradas": {
        "movement_type": "entry",
        "label": "Entrada",
        "table_key": "entries",
        "folio_column": "entry_folio",
        "folio_prefix": "EMB-ENT-AJ",
        "batch_prefix": "AJ-ENT",
        "quantity_sign": 1,
        "template_filename": "plantilla_ajuste_entradas_embarques",
        "default_notes": "Ajuste manual por lote de entradas",
    },
    "salidas": {
        "movement_type": "exit",
        "label": "Salida",
        "table_key": "exits",
        "folio_column": "exit_folio",
        "folio_prefix": "EMB-SAL-AJ",
        "batch_prefix": "AJ-SAL",
        "quantity_sign": -1,
        "template_filename": "plantilla_ajuste_salidas_embarques",
        "default_notes": "Ajuste manual por lote de salidas",
    },
}


def _obtener_config_ajuste_manual_embarques(module_name):
    config = MANUAL_SHIPPING_ADJUSTMENT_CONFIG.get(
        normalize_search(module_name).lower()
    )
    if not config:
        raise ValueError("Módulo de ajuste no soportado.")
    return config


def _normalizar_fecha_movimiento_ajuste_embarques(raw_value):
    value = normalize_search(raw_value)
    if not value:
        raise ValueError("La fecha del movimiento es obligatoria.")

    normalized = value.replace("T", " ").strip()
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(normalized, fmt)
            if fmt == "%Y-%m-%d":
                return parsed.replace(hour=12, minute=0, second=0)
            return parsed.replace(microsecond=0)
        except ValueError:
            continue

    return to_sql_datetime(value)


def _primer_valor_presente(row, keys):
    for key in keys:
        if key in row and row.get(key) is not None:
            value = row.get(key)
            if not isinstance(value, str) or value.strip() != "":
                return value
    return None


def _parsear_archivo_ajuste_manual_embarques(file_storage):
    raw_bytes = file_storage.read()
    file_hash = hashlib.sha256(raw_bytes).hexdigest()
    filename = normalize_search(file_storage.filename)
    extension = Path(filename).suffix.lower()
    errors = []
    source_rows = []

    try:
        if extension in {".xlsx", ".xls"}:
            dataframe = pd.read_excel(io.BytesIO(raw_bytes), dtype=str)
            source_rows = dataframe.fillna("").to_dict("records")
        else:
            try:
                decoded = raw_bytes.decode("utf-8-sig")
            except UnicodeDecodeError:
                decoded = raw_bytes.decode("latin-1")
            reader = csv.DictReader(io.StringIO(decoded))
            source_rows = list(reader)
    except Exception as exc:
        return {
            "valid": False,
            "errors": [f"No fue posible leer el archivo: {exc}"],
            "rows": [],
            "fileHash": file_hash,
        }

    if not source_rows:
        return {
            "valid": False,
            "errors": ["El archivo no contiene registros para validar."],
            "rows": [],
            "fileHash": file_hash,
        }

    parsed_rows = []
    seen_parts = set()

    for index, raw_row in enumerate(source_rows, start=2):
        normalized_row = {
            _normalizar_header_csv_cierre_embarques(key): value
            for key, value in raw_row.items()
        }
        part_number = normalize_part_number(
            _primer_valor_presente(
                normalized_row,
                ["part_number", "no_parte", "numero_parte", "part", "pn"],
            )
        )
        quantity_raw = _primer_valor_presente(
            normalized_row,
            ["quantity", "cantidad", "qty", "current_qty"],
        )
        quantity = normalize_integer(quantity_raw)

        if not part_number and quantity is None:
            continue

        if not part_number:
            errors.append(f"Línea {index}: part_number es obligatorio.")
            continue

        if quantity is None:
            errors.append(f"Línea {index}: quantity es obligatorio para {part_number}.")
            continue

        if quantity <= 0:
            errors.append(
                f"Línea {index}: quantity debe ser mayor a cero para {part_number}."
            )
            continue

        if part_number in seen_parts:
            errors.append(
                f"Línea {index}: part_number duplicado {part_number}. Deja un solo renglón por número de parte."
            )
            continue

        seen_parts.add(part_number)
        parsed_rows.append(
            {
                "rowNumber": index,
                "part_number": part_number,
                "quantity": quantity,
            }
        )

    if not parsed_rows and not errors:
        errors.append("No se encontraron renglones válidos en el archivo.")

    serializable_rows = [
        {
            "part_number": row["part_number"],
            "quantity": row["quantity"],
        }
        for row in parsed_rows
    ]
    rows_hash = hashlib.sha256(
        json.dumps(
            serializable_rows,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()

    return {
        "valid": not errors,
        "errors": errors,
        "rows": parsed_rows,
        "fileHash": file_hash,
        "rowsHash": rows_hash,
    }


def _obtener_mapa_inventario_ajuste_manual_embarques():
    rows = execute_query(
        f"""
        SELECT
          id,
          catalog_id,
          part_number,
          current_quantity,
          product_model,
          product_status,
          description,
          customer,
          zone_code,
          standard_pack
        FROM `{SHIPPING_TABLES['inventory']}`
        WHERE part_number IS NOT NULL
          AND TRIM(part_number) <> ''
        ORDER BY part_number ASC
        """,
        fetch="all",
    ) or []
    return {normalize_part_number(row.get("part_number")): row for row in rows}


def _obtener_cierres_impactados_ajuste_manual(part_numbers, movement_at):
    normalized_parts = {normalize_part_number(part) for part in part_numbers if part}
    if not normalized_parts:
        return []

    rows = execute_query(
        f"""
        SELECT
          id,
          closure_batch_id,
          part_number,
          initial_quantity,
          system_quantity,
          difference_quantity,
          raw_current_quantity,
          closed_at,
          closure_label
        FROM `{SHIPPING_TABLES['inventory_closures']}`
        WHERE closed_at > %s
        ORDER BY closed_at ASC, id ASC
        """,
        (movement_at,),
        fetch="all",
    ) or []

    return [
        row
        for row in rows
        if normalize_part_number(row.get("part_number")) in normalized_parts
    ]


def _construir_preview_ajuste_manual_embarques(config, parsed_rows, movement_at):
    inventory_map = _obtener_mapa_inventario_ajuste_manual_embarques()
    impacted_closures = _obtener_cierres_impactados_ajuste_manual(
        [row.get("part_number") for row in parsed_rows],
        movement_at,
    )
    impacted_by_part = {}
    for closure in impacted_closures:
        impacted_by_part.setdefault(
            normalize_part_number(closure.get("part_number")),
            [],
        ).append(closure)

    errors = []
    preview_rows = []
    total_quantity = 0
    sign = int(config["quantity_sign"])

    for row in parsed_rows:
        part_number = normalize_part_number(row.get("part_number"))
        quantity = normalize_integer(row.get("quantity")) or 0
        inventory = inventory_map.get(part_number)

        if not inventory:
            errors.append(
                f"Línea {row.get('rowNumber')}: {part_number} no existe en el catálogo de embarques."
            )
            continue

        previous_quantity = normalize_integer(inventory.get("current_quantity")) or 0
        new_quantity = previous_quantity + (quantity * sign)
        closures_for_part = impacted_by_part.get(part_number, [])
        total_quantity += quantity

        preview_rows.append(
            {
                "rowNumber": row.get("rowNumber"),
                "part_number": part_number,
                "quantity": quantity,
                "current_quantity": previous_quantity,
                "new_quantity": new_quantity,
                "product_model": _normalizar_texto_embarques_historial(
                    inventory.get("product_model")
                ),
                "customer": _normalizar_texto_embarques_historial(
                    inventory.get("customer")
                ),
                "closureImpactCount": len(closures_for_part),
                "status": "impacta_cierre" if closures_for_part else "ok",
            }
        )

    closure_impact = {
        "affected": bool(impacted_closures),
        "affectedRows": len(impacted_closures),
        "affectedClosures": [
            {
                "id": closure.get("id"),
                "closureBatchId": closure.get("closure_batch_id"),
                "partNumber": _normalizar_texto_embarques_historial(
                    closure.get("part_number")
                ),
                "closureLabel": _normalizar_texto_embarques_historial(
                    closure.get("closure_label")
                ),
                "closedAt": _normalizar_texto_embarques_historial(
                    closure.get("closed_at")
                ),
            }
            for closure in impacted_closures
        ],
    }

    serializable_rows = [
        {
            "part_number": row["part_number"],
            "quantity": row["quantity"],
            "movement_type": config["movement_type"],
            "movement_at": movement_at.strftime("%Y-%m-%d %H:%M:%S"),
        }
        for row in preview_rows
    ]
    rows_hash = hashlib.sha256(
        json.dumps(
            serializable_rows,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()

    return {
        "valid": not errors and bool(preview_rows),
        "errors": errors,
        "rows": preview_rows,
        "closureImpact": closure_impact,
        "summary": {
            "totalRows": len(preview_rows),
            "totalQuantity": total_quantity,
            "rowsHash": rows_hash,
            "movementType": config["movement_type"],
            "movementAt": movement_at.strftime("%Y-%m-%d %H:%M:%S"),
        },
    }


def _generar_batch_code_ajuste_manual_embarques(config, rows_hash):
    now = obtener_fecha_hora_mexico()
    suffix = hashlib.sha1(
        f"{rows_hash}:{now.isoformat()}".encode("utf-8")
    ).hexdigest()[:6].upper()
    return f"{config['batch_prefix']}-{now.strftime('%Y%m%d-%H%M%S')}-{suffix}"


def _guardar_preview_ajuste_manual_embarques(
    config,
    movement_at,
    reason,
    file_name,
    file_hash,
    preview_payload,
):
    conn = get_pooled_connection()
    if conn is None:
        raise RuntimeError("No fue posible obtener conexión MySQL para guardar el preview.")

    cursor = get_dict_cursor(conn)
    usuario_actual = _obtener_usuario_display_actual()
    batch_code = _generar_batch_code_ajuste_manual_embarques(
        config,
        preview_payload["summary"]["rowsHash"],
    )
    payload = {
        "metadata": {
            "batchCode": batch_code,
            "movementType": config["movement_type"],
            "movementLabel": config["label"],
            "movementAt": movement_at.strftime("%Y-%m-%d %H:%M:%S"),
            "createdBy": usuario_actual,
            "sourceFileName": file_name,
            "sourceFileHash": file_hash,
            "reason": reason,
        },
        "summary": preview_payload["summary"],
        "rows": preview_payload["rows"],
        "closureImpact": preview_payload["closureImpact"],
    }

    try:
        conn.autocommit(False)
        cursor.execute(
            f"""
            UPDATE `{SHIPPING_TABLES['manual_adjustment_batches']}`
            SET status = 'superseded',
                cancelled_at = NOW()
            WHERE movement_type = %s
              AND created_by = %s
              AND status = 'draft'
            """,
            (config["movement_type"], usuario_actual),
        )
        cursor.execute(
            f"""
            INSERT INTO `{SHIPPING_TABLES['manual_adjustment_batches']}` (
              batch_code,
              movement_type,
              movement_at,
              status,
              reason,
              created_by,
              source_file_name,
              source_file_hash,
              rows_hash,
              payload_json,
              closure_impact_json,
              total_rows,
              total_quantity
            ) VALUES (%s, %s, %s, 'draft', %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                batch_code,
                config["movement_type"],
                movement_at,
                reason,
                usuario_actual,
                file_name,
                file_hash,
                preview_payload["summary"]["rowsHash"],
                json.dumps(payload, ensure_ascii=False, default=str, separators=(",", ":")),
                json.dumps(
                    preview_payload["closureImpact"],
                    ensure_ascii=False,
                    default=str,
                    separators=(",", ":"),
                ),
                preview_payload["summary"]["totalRows"],
                preview_payload["summary"]["totalQuantity"],
            ),
        )
        batch_id = cursor.lastrowid
        for row in preview_payload["rows"]:
            cursor.execute(
                f"""
                INSERT INTO `{SHIPPING_TABLES['manual_adjustment_items']}` (
                  batch_id,
                  `row_number`,
                  part_number,
                  quantity,
                  status
                ) VALUES (%s, %s, %s, %s, 'draft')
                """,
                (
                    batch_id,
                    row.get("rowNumber"),
                    row.get("part_number"),
                    row.get("quantity"),
                ),
            )
        conn.commit()
        return batch_id, batch_code
    except Exception:
        conn.rollback()
        raise
    finally:
        try:
            conn.autocommit(True)
        except Exception:
            pass
        cursor.close()
        conn.close()


def _recalcular_cierres_impactados_ajuste_manual(cursor, config, movement_at, item_rows):
    sign = int(config["quantity_sign"])
    updated_closure_ids = []

    for item in item_rows:
        part_number = normalize_part_number(item.get("part_number"))
        quantity = normalize_integer(item.get("quantity")) or 0
        if not part_number or quantity <= 0:
            continue

        cursor.execute(
            f"""
            SELECT
              id,
              system_quantity,
              raw_current_quantity,
              initial_quantity
            FROM `{SHIPPING_TABLES['inventory_closures']}`
            WHERE part_number = %s
              AND closed_at > %s
            ORDER BY closed_at ASC, id ASC
            FOR UPDATE
            """,
            (part_number, movement_at),
        )
        for closure in cursor.fetchall():
            current_system = normalize_integer(closure.get("system_quantity"))
            if current_system is None:
                current_system = normalize_integer(closure.get("initial_quantity")) or 0
            new_system = current_system + (quantity * sign)
            raw_current = normalize_integer(closure.get("raw_current_quantity"))
            if raw_current is None:
                raw_current = normalize_integer(closure.get("initial_quantity")) or 0
            difference = raw_current - new_system
            cursor.execute(
                f"""
                UPDATE `{SHIPPING_TABLES['inventory_closures']}`
                SET system_quantity = %s,
                    difference_quantity = %s
                WHERE id = %s
                """,
                (new_system, difference, closure.get("id")),
            )
            updated_closure_ids.append(closure.get("id"))

    return updated_closure_ids


def _confirmar_ajuste_manual_embarques(config, batch_id):
    conn = get_pooled_connection()
    if conn is None:
        raise RuntimeError("No fue posible obtener conexión MySQL para confirmar el lote.")

    cursor = get_dict_cursor(conn)
    usuario_actual = _obtener_usuario_display_actual()
    inserted_records = []
    rebuild_parts = set()

    try:
        conn.autocommit(False)
        cursor.execute(
            f"""
            SELECT *
            FROM `{SHIPPING_TABLES['manual_adjustment_batches']}`
            WHERE id = %s
              AND movement_type = %s
            LIMIT 1
            FOR UPDATE
            """,
            (batch_id, config["movement_type"]),
        )
        batch = cursor.fetchone()
        if not batch:
            conn.rollback()
            return {"success": False, "error": "No se encontró el lote solicitado."}, 404

        if normalize_search(batch.get("status")).lower() != "draft":
            conn.rollback()
            return {
                "success": False,
                "error": "El lote ya no está disponible para confirmar.",
            }, 409

        cursor.execute(
            f"""
            SELECT *
            FROM `{SHIPPING_TABLES['manual_adjustment_items']}`
            WHERE batch_id = %s
            ORDER BY `row_number` ASC, id ASC
            FOR UPDATE
            """,
            (batch_id,),
        )
        item_rows = cursor.fetchall()
        if not item_rows:
            conn.rollback()
            return {"success": False, "error": "El lote no contiene registros."}, 400

        movement_at = to_sql_datetime(batch.get("movement_at"))
        reason = normalize_search(batch.get("reason"))

        for item in item_rows:
            part_number = normalize_part_number(item.get("part_number"))
            quantity = normalize_integer(item.get("quantity")) or 0
            inventory = ensure_inventory_record(cursor, part_number)
            if not inventory:
                conn.rollback()
                return {
                    "success": False,
                    "error": f"{part_number} ya no existe en el catálogo de embarques.",
                }, 404

            folio = generate_movement_folio(config["folio_prefix"])
            catalog_id = inventory.get("catalog_id") or inventory.get("catalog_ref_id")
            inventory_id = inventory.get("id")
            product_model = inventory.get("product_model") or inventory.get("catalog_model")
            description = inventory.get("description") or inventory.get("catalog_description")
            customer = inventory.get("customer") or inventory.get("catalog_customer")
            zone_code = inventory.get("zone_code") or inventory.get("catalog_zone_code")
            notes = f"{config['default_notes']} {batch.get('batch_code')}: {reason}"

            if config["movement_type"] == "entry":
                cursor.execute(
                    f"""
                    INSERT INTO `{SHIPPING_TABLES['entries']}` (
                      entry_folio,
                      inventory_id,
                      catalog_id,
                      part_number,
                      quantity,
                      available_quantity,
                      is_fifo_layer_only,
                      previous_quantity,
                      new_quantity,
                      product_model,
                      description,
                      customer,
                      zone_code,
                      location_code,
                      reference_code,
                      batch_no,
                      notes,
                      registered_by,
                      movement_at
                    ) VALUES (%s, %s, %s, %s, %s, 0, 0, 0, 0, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        folio,
                        inventory_id,
                        catalog_id,
                        part_number,
                        quantity,
                        product_model,
                        description,
                        customer,
                        zone_code,
                        None,
                        f"AJUSTE:{batch.get('batch_code')}",
                        None,
                        notes,
                        usuario_actual,
                        movement_at,
                    ),
                )
            else:
                cursor.execute(
                    f"""
                    INSERT INTO `{SHIPPING_TABLES['exits']}` (
                      exit_folio,
                      inventory_id,
                      catalog_id,
                      part_number,
                      quantity,
                      previous_quantity,
                      new_quantity,
                      product_model,
                      description,
                      customer,
                      zone_code,
                      location_code,
                      fifo_allocation_json,
                      destination_area,
                      departure_code,
                      departure_assigned_at,
                      departure_assigned_by,
                      reason,
                      requested_by,
                      remarks,
                      registered_by,
                      movement_at
                    ) VALUES (%s, %s, %s, %s, %s, 0, 0, %s, %s, %s, %s, %s, NULL, %s, NULL, NULL, NULL, %s, %s, %s, %s, %s)
                    """,
                    (
                        folio,
                        inventory_id,
                        catalog_id,
                        part_number,
                        quantity,
                        product_model,
                        description,
                        customer,
                        zone_code,
                        None,
                        "Embarques",
                        "Ajuste manual de salida",
                        usuario_actual,
                        notes,
                        usuario_actual,
                        movement_at,
                    ),
                )

            inserted_records.append(
                {
                    "item_id": item.get("id"),
                    "part_number": part_number,
                    "quantity": quantity,
                    "folio": folio,
                    "inventory_id": inventory_id,
                    "catalog_id": catalog_id,
                }
            )
            rebuild_parts.add(part_number)

        for part_number in sorted(rebuild_parts):
            rebuild_part_inventory_state(cursor, part_number)

        table_name = SHIPPING_TABLES[config["table_key"]]
        folio_column = config["folio_column"]
        for record in inserted_records:
            cursor.execute(
                f"""
                SELECT previous_quantity, new_quantity
                FROM `{table_name}`
                WHERE `{folio_column}` = %s
                LIMIT 1
                """,
                (record["folio"],),
            )
            movement_row = cursor.fetchone() or {}
            cursor.execute(
                f"""
                UPDATE `{SHIPPING_TABLES['manual_adjustment_items']}`
                SET inventory_id = %s,
                    catalog_id = %s,
                    folio = %s,
                    previous_quantity = %s,
                    new_quantity = %s,
                    status = 'applied'
                WHERE id = %s
                """,
                (
                    record.get("inventory_id"),
                    record.get("catalog_id"),
                    record.get("folio"),
                    normalize_integer(movement_row.get("previous_quantity")) or 0,
                    normalize_integer(movement_row.get("new_quantity")) or 0,
                    record.get("item_id"),
                ),
            )

        updated_closure_ids = _recalcular_cierres_impactados_ajuste_manual(
            cursor,
            config,
            movement_at,
            item_rows,
        )

        cursor.execute(
            f"""
            UPDATE `{SHIPPING_TABLES['manual_adjustment_batches']}`
            SET status = 'applied',
                confirmed_by = %s,
                confirmed_at = NOW()
            WHERE id = %s
            """,
            (usuario_actual, batch_id),
        )

        conn.commit()
        return {
            "success": True,
            "batchId": batch_id,
            "batchCode": batch.get("batch_code"),
            "movementType": config["movement_type"],
            "insertedRows": len(inserted_records),
            "updatedClosureRows": len(updated_closure_ids),
            "message": "Lote aplicado correctamente.",
        }, 200
    except Exception:
        conn.rollback()
        raise
    finally:
        try:
            conn.autocommit(True)
        except Exception:
            pass
        cursor.close()
        conn.close()


def _registrar_ajuste_manual_individual_embarques(config, data):
    part_number = normalize_part_number(data.get("partNumber") or data.get("part_number"))
    quantity = normalize_integer(data.get("quantity"))
    reason = normalize_search(data.get("reason"))
    movement_at = _normalizar_fecha_movimiento_ajuste_embarques(data.get("movementAt"))

    if not part_number:
        return {"success": False, "error": "El número de parte es obligatorio."}, 400
    if quantity is None or quantity <= 0:
        return {"success": False, "error": "La cantidad debe ser mayor a cero."}, 400
    if not reason:
        return {"success": False, "error": "El motivo del registro es obligatorio."}, 400

    parsed_rows = [
        {
            "rowNumber": 1,
            "part_number": part_number,
            "quantity": quantity,
        }
    ]
    preview = _construir_preview_ajuste_manual_embarques(
        config,
        parsed_rows,
        movement_at,
    )
    if not preview["valid"]:
        return {
            "success": False,
            "error": "; ".join(preview.get("errors") or ["No fue posible validar el registro."]),
            "preview": preview,
        }, 400

    manual_hash = hashlib.sha256(
        json.dumps(
            {
                "movementType": config["movement_type"],
                "movementAt": movement_at.strftime("%Y-%m-%d %H:%M:%S"),
                "partNumber": part_number,
                "quantity": quantity,
                "reason": reason,
                "createdBy": _obtener_usuario_display_actual(),
            },
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()
    batch_id, _batch_code = _guardar_preview_ajuste_manual_embarques(
        config,
        movement_at,
        reason,
        "registro_manual",
        manual_hash,
        preview,
    )
    return _confirmar_ajuste_manual_embarques(config, batch_id)


def _guardar_borrador_cierre_inventario_embarques(payload, csv_file_name, csv_hash):
    conn = get_pooled_connection()
    if conn is None:
        raise RuntimeError("No fue posible obtener conexión MySQL para guardar el borrador.")

    cursor = get_dict_cursor(conn)
    try:
        conn.autocommit(False)
        cursor.execute(
            f"""
            UPDATE `{SHIPPING_TABLES['inventory_closure_batches']}`
            SET status = 'superseded'
            WHERE status = 'draft'
              AND created_by = %s
              AND closure_month = %s
            """,
            (payload["metadata"]["closureUser"], payload["metadata"]["closureMonth"]),
        )
        cursor.execute(
            f"""
            INSERT INTO `{SHIPPING_TABLES['inventory_closure_batches']}` (
              closure_label,
              closure_month,
              closed_at,
              status,
              created_by,
              csv_file_name,
              csv_hash,
              rows_hash,
              payload_json,
              accuracy_pct,
              total_rows,
              differing_rows
            ) VALUES (%s, %s, %s, 'draft', %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                payload["metadata"]["closureLabel"],
                payload["metadata"]["closureMonth"],
                payload["metadata"]["closureDateTime"],
                payload["metadata"]["closureUser"],
                csv_file_name,
                csv_hash,
                payload["summary"]["rowsHash"],
                json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
                payload["summary"]["accuracyPct"],
                payload["summary"]["totalRows"],
                payload["summary"]["differenceRows"],
            ),
        )
        batch_id = cursor.lastrowid
        conn.commit()
        return batch_id
    except Exception:
        conn.rollback()
        raise
    finally:
        try:
            conn.autocommit(True)
        except Exception:
            pass
        cursor.close()
        conn.close()


def _obtener_historial_cierres_inventario_embarques(limit=20, include_last_draft_for=None):
    rows = execute_query(
        f"""
        SELECT
          id,
          closure_label,
          closure_month,
          status,
          created_by,
          confirmed_by,
          closed_at,
          confirmed_at,
          accuracy_pct,
          total_rows,
          differing_rows,
          csv_hash,
          rows_hash
        FROM `{SHIPPING_TABLES['inventory_closure_batches']}`
        WHERE status = 'confirmed'
        ORDER BY COALESCE(confirmed_at, created_at) DESC, id DESC
        LIMIT %s
        """,
        (limit,),
        fetch="all",
    ) or []

    if include_last_draft_for:
        draft_row = execute_query(
            f"""
            SELECT
              id,
              closure_label,
              closure_month,
              status,
              created_by,
              confirmed_by,
              closed_at,
              confirmed_at,
              accuracy_pct,
              total_rows,
              differing_rows,
              csv_hash,
              rows_hash
            FROM `{SHIPPING_TABLES['inventory_closure_batches']}`
            WHERE status = 'draft'
              AND created_by = %s
            ORDER BY created_at DESC, id DESC
            LIMIT 1
            """,
            (include_last_draft_for,),
            fetch="one",
        )
        if draft_row:
            rows.insert(0, draft_row)

    history_rows = []
    for row in rows:
        period_info = _resolver_periodo_cierre_embarques(row)
        history_rows.append(
            {
                "id": row.get("id"),
                "closure_label": period_info["closure_label"],
                "closure_month": period_info["closure_month"],
                "status": _normalizar_texto_embarques_historial(row.get("status")),
                "created_by": _normalizar_texto_embarques_historial(row.get("created_by")),
                "confirmed_by": _normalizar_texto_embarques_historial(row.get("confirmed_by")),
                "closed_at": _normalizar_texto_embarques_historial(row.get("closed_at")),
                "confirmed_at": _normalizar_texto_embarques_historial(row.get("confirmed_at")),
                "accuracy_pct": _normalizar_numero_embarques_historial(row.get("accuracy_pct")),
                "total_rows": _normalizar_numero_embarques_historial(row.get("total_rows")),
                "differing_rows": _normalizar_numero_embarques_historial(row.get("differing_rows")),
                "csv_hash": _normalizar_texto_embarques_historial(row.get("csv_hash")),
                "rows_hash": _normalizar_texto_embarques_historial(row.get("rows_hash")),
            }
        )
    return history_rows


def _agregar_timestamp_nombre_archivo(filename):
    stem, extension = os.path.splitext(filename)
    timestamp = obtener_fecha_hora_mexico().strftime("%y%m%d_%H%M%S")
    return f"{stem}_{timestamp}{extension}"


def _exportar_historial_embarques_excel(sheet_name, filename, headers, rows):
    """Generar archivo Excel para cualquiera de los historiales de embarques."""
    from io import BytesIO

    from flask import Response
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(headers.values(), 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            cell.alignment = Alignment(vertical="top")
            value_length = len(str(cell.value or ""))
            if value_length > max_length:
                max_length = value_length
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 28)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    download_filename = _agregar_timestamp_nombre_archivo(filename)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{download_filename}"'},
    )


def _texto_pdf_embarques(value, fallback="-"):
    """Normalizar texto para documentos PDF de embarques."""
    text = str(value if value is not None else "").strip()
    return text or fallback


def _cantidad_pdf_salidas_retorno(row):
    """Obtener la cantidad imprimible de una salida de retorno."""
    for key in ("movement_quantity", "loss_quantity", "quantity", "cantidad"):
        value = row.get(key)
        if value not in (None, ""):
            try:
                return float(value)
            except (TypeError, ValueError):
                return 0
    return 0


def _formatear_numero_pdf_embarques(value):
    try:
        numeric_value = float(value)
    except (TypeError, ValueError):
        return "0"
    if numeric_value.is_integer():
        return f"{int(numeric_value):,}"
    return f"{numeric_value:,.2f}".rstrip("0").rstrip(".")


def _tipo_retorno_pdf_embarques(reason):
    text = _texto_pdf_embarques(reason, "Retorno")
    return text.split("/")[0].strip() or "Retorno"


def _normalizar_filas_pdf_salidas_retorno(rows):
    normalized_rows = []
    for raw_row in rows[:1000]:
        if not isinstance(raw_row, dict):
            continue
        quantity = _cantidad_pdf_salidas_retorno(raw_row)
        if quantity <= 0:
            continue
        normalized_rows.append(
            {
                "fecha": _texto_pdf_embarques(raw_row.get("fecha")),
                "hora": _texto_pdf_embarques(raw_row.get("hora")),
                "folio": _texto_pdf_embarques(raw_row.get("folio")),
                "part_number": _texto_pdf_embarques(raw_row.get("part_number")),
                "quantity": quantity,
                "product_model": _texto_pdf_embarques(raw_row.get("product_model")),
                "reason": _tipo_retorno_pdf_embarques(raw_row.get("reason")),
                "registered_by": _texto_pdf_embarques(raw_row.get("registered_by")),
            }
        )

    return sorted(
        normalized_rows,
        key=lambda row: (
            row["part_number"].lower(),
            row["folio"].lower(),
            row["fecha"],
            row["hora"],
        ),
    )


def _pdf_escape_text_embarques(value):
    """Escapar texto para una cadena PDF WinAnsi sin dependencias externas."""
    encoded = _texto_pdf_embarques(value).encode("cp1252", "replace")
    escaped = bytearray()
    for byte in encoded:
        if byte in (40, 41, 92):  # (, ), \
            escaped.append(92)
            escaped.append(byte)
        elif byte in (10, 13, 9):
            escaped.append(32)
        else:
            escaped.append(byte)
    return escaped.decode("latin-1")


def _pdf_color_embarques(hex_color):
    color = hex_color.strip().lstrip("#")
    if len(color) != 6:
        return (0, 0, 0)
    return tuple(int(color[index : index + 2], 16) / 255 for index in (0, 2, 4))


def _pdf_cmd_color_embarques(hex_color, stroke=False):
    r, g, b = _pdf_color_embarques(hex_color)
    operator = "RG" if stroke else "rg"
    return f"{r:.3f} {g:.3f} {b:.3f} {operator}"


def _pdf_text_width_embarques(text, font_size, bold=False):
    factor = 0.56 if bold else 0.50
    return len(str(text)) * font_size * factor


def _pdf_fit_text_embarques(text, font_size, max_width, bold=False):
    text = _texto_pdf_embarques(text)
    if _pdf_text_width_embarques(text, font_size, bold) <= max_width:
        return text

    ellipsis = "..."
    available = max_width - _pdf_text_width_embarques(ellipsis, font_size, bold)
    if available <= 0:
        return ""

    fitted = text
    while fitted and _pdf_text_width_embarques(fitted, font_size, bold) > available:
        fitted = fitted[:-1]
    return f"{fitted}{ellipsis}" if fitted else ellipsis


def _pdf_rect_embarques(x, y, width, height, page_height, fill="#ffffff", stroke="#cfd6e4", line_width=0.5):
    pdf_y = page_height - y - height
    commands = [
        f"q {_pdf_cmd_color_embarques(fill)} {x:.2f} {pdf_y:.2f} {width:.2f} {height:.2f} re f Q",
    ]
    if stroke:
        commands.append(
            f"q {_pdf_cmd_color_embarques(stroke, stroke=True)} {line_width:.2f} w "
            f"{x:.2f} {pdf_y:.2f} {width:.2f} {height:.2f} re S Q"
        )
    return "\n".join(commands)


def _pdf_line_embarques(x1, y1, x2, y2, page_height, stroke="#11213c", line_width=1):
    return (
        f"q {_pdf_cmd_color_embarques(stroke, stroke=True)} {line_width:.2f} w "
        f"{x1:.2f} {page_height - y1:.2f} m {x2:.2f} {page_height - y2:.2f} l S Q"
    )


def _pdf_text_embarques(
    x,
    y,
    text,
    page_height,
    font_size=8,
    font="F1",
    fill="#162033",
    max_width=None,
    align="left",
    bold=False,
):
    safe_text = _texto_pdf_embarques(text)
    if max_width is not None:
        safe_text = _pdf_fit_text_embarques(safe_text, font_size, max_width, bold)
    text_width = _pdf_text_width_embarques(safe_text, font_size, bold)
    if align == "right" and max_width is not None:
        text_x = x + max(max_width - text_width, 0)
    elif align == "center" and max_width is not None:
        text_x = x + max((max_width - text_width) / 2, 0)
    else:
        text_x = x
    pdf_y = page_height - y
    escaped_text = _pdf_escape_text_embarques(safe_text)
    return (
        f"q {_pdf_cmd_color_embarques(fill)} BT /{font} {font_size:.2f} Tf "
        f"1 0 0 1 {text_x:.2f} {pdf_y:.2f} Tm ({escaped_text}) Tj ET Q"
    )


def _pdf_cell_text_embarques(x, y, width, height, text, page_height, font_size=7, font="F1", fill="#162033", align="left", bold=False):
    text_x = x + 4
    text_y = y + (height / 2) + (font_size / 2) - 1.5
    return _pdf_text_embarques(
        text_x,
        text_y,
        text,
        page_height,
        font_size=font_size,
        font=font,
        fill=fill,
        max_width=max(width - 8, 5),
        align=align,
        bold=bold,
    )


def _crear_pdf_embarques(page_streams, page_width=792, page_height=612):
    """Crear un PDF mínimo con fuentes Helvetica estándar."""
    objects = [
        None,
        None,
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold /Encoding /WinAnsiEncoding >>",
    ]
    pages_id = 2
    font_regular_id = 3
    font_bold_id = 4
    page_ids = []

    def add_object(payload):
        objects.append(payload)
        return len(objects)

    for stream in page_streams:
        stream_bytes = stream.encode("latin-1", "replace")
        content_id = add_object(
            b"<< /Length "
            + str(len(stream_bytes)).encode("ascii")
            + b" >>\nstream\n"
            + stream_bytes
            + b"\nendstream"
        )
        page_id = add_object(
            (
                f"<< /Type /Page /Parent {pages_id} 0 R /MediaBox [0 0 {page_width} {page_height}] "
                f"/Resources << /Font << /F1 {font_regular_id} 0 R /F2 {font_bold_id} 0 R >> >> "
                f"/Contents {content_id} 0 R >>"
            ).encode("ascii")
        )
        page_ids.append(page_id)

    objects[0] = f"<< /Type /Catalog /Pages {pages_id} 0 R >>".encode("ascii")
    kids = " ".join(f"{page_id} 0 R" for page_id in page_ids)
    objects[1] = f"<< /Type /Pages /Kids [{kids}] /Count {len(page_ids)} >>".encode("ascii")

    output = io.BytesIO()
    output.write(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for index, payload in enumerate(objects, 1):
        offsets.append(output.tell())
        output.write(f"{index} 0 obj\n".encode("ascii"))
        output.write(payload)
        output.write(b"\nendobj\n")

    xref_offset = output.tell()
    output.write(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    output.write(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.write(f"{offset:010d} 00000 n \n".encode("ascii"))
    output.write(
        (
            f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
            f"startxref\n{xref_offset}\n%%EOF\n"
        ).encode("ascii")
    )
    output.seek(0)
    return output


def _generar_pdf_salidas_retorno_embarques(rows):
    """Generar PDF carta horizontal para el formato de salidas de retorno sin dependencias externas."""
    normalized_rows = _normalizar_filas_pdf_salidas_retorno(rows)
    if not normalized_rows:
        raise ValueError("No hay salidas de retorno válidas para generar el PDF.")

    page_width = 792
    page_height = 612
    margin_x = 30
    margin_y = 52
    table_top = margin_y + 122
    row_height = 14
    table_header_height = 17
    footer_reserved = 56
    table_bottom = page_height - margin_y - footer_reserved
    navy = "#11213c"
    gray = "#63708a"
    border = "#cfd6e4"
    row_alt = "#f4f7fb"
    white = "#ffffff"
    black = "#162033"

    columns = [
        ("Fecha", "fecha", 58, "left"),
        ("Hora", "hora", 44, "left"),
        ("Folio", "folio", 162, "left"),
        ("No. parte", "part_number", 95, "left"),
        ("Cantidad", "quantity", 50, "right"),
        ("Modelo", "product_model", 168, "left"),
        ("Tipo", "reason", 60, "left"),
        ("Usuario", "registered_by", 95, "left"),
    ]
    available_width = page_width - (margin_x * 2)
    total_width = sum(column[2] for column in columns)
    if total_width < available_width:
        extra_width = available_width - total_width
        columns[5] = (columns[5][0], columns[5][1], columns[5][2] + extra_width, columns[5][3])

    rows_per_page = max(1, int((table_bottom - table_top - table_header_height) / row_height))
    total_pages = max(1, (len(normalized_rows) + rows_per_page - 1) // rows_per_page)
    total_quantity = sum(row["quantity"] for row in normalized_rows)
    generated_at = obtener_fecha_hora_mexico().strftime("%d/%m/%Y, %I:%M %p").replace("AM", "a.m.").replace("PM", "p.m.")

    page_streams = []
    for page_index in range(total_pages):
        page_rows = normalized_rows[
            page_index * rows_per_page : (page_index + 1) * rows_per_page
        ]
        commands = []

        commands.append(_pdf_text_embarques(margin_x, margin_y, "ALMACÉN DE EMBARQUES", page_height, 8, "F2", gray, bold=True))
        commands.append(_pdf_text_embarques(margin_x, margin_y + 22, "SALIDA DE RETORNO", page_height, 20, "F2", navy, bold=True))
        commands.append(_pdf_text_embarques(page_width - margin_x - 58, margin_y, "GENERADO", page_height, 7, "F2", gray, bold=True))
        commands.append(_pdf_text_embarques(page_width - margin_x - 112, margin_y + 18, generated_at, page_height, 8, "F2", black, bold=True))
        commands.append(
            _pdf_text_embarques(
                page_width - margin_x - 42,
                margin_y + 34,
                f"Pág. {page_index + 1}/{total_pages}",
                page_height,
                7,
                "F2",
                gray,
                bold=True,
            )
        )
        commands.append(_pdf_line_embarques(margin_x, margin_y + 50, page_width - margin_x, margin_y + 50, page_height, navy, 1.5))

        summary_top = margin_y + 64
        summary_width = (available_width - 8) / 2
        for offset, label, value in (
            (0, "REGISTROS", _formatear_numero_pdf_embarques(len(normalized_rows))),
            (summary_width + 8, "CANTIDAD TOTAL", _formatear_numero_pdf_embarques(total_quantity)),
        ):
            x = margin_x + offset
            commands.append(_pdf_rect_embarques(x, summary_top, summary_width, 36, page_height, white, border, 0.6))
            commands.append(_pdf_text_embarques(x + 8, summary_top + 12, label, page_height, 6.5, "F2", gray, bold=True))
            commands.append(_pdf_text_embarques(x + 8, summary_top + 28, value, page_height, 12, "F2", navy, bold=True))

        x = margin_x
        y = table_top
        for label, _key, width, align in columns:
            commands.append(_pdf_rect_embarques(x, y, width, table_header_height, page_height, navy, navy, 0.4))
            commands.append(
                _pdf_cell_text_embarques(
                    x,
                    y,
                    width,
                    table_header_height,
                    label,
                    page_height,
                    6.2,
                    "F2",
                    white,
                    align,
                    True,
                )
            )
            x += width

        y += table_header_height
        for row_index, row in enumerate(page_rows):
            x = margin_x
            row_fill = row_alt if row_index % 2 else white
            for _label, key, width, align in columns:
                commands.append(_pdf_rect_embarques(x, y, width, row_height, page_height, row_fill, border, 0.35))
                value = _formatear_numero_pdf_embarques(row[key]) if key == "quantity" else row[key]
                is_bold = key == "part_number"
                commands.append(
                    _pdf_cell_text_embarques(
                        x,
                        y,
                        width,
                        row_height,
                        value,
                        page_height,
                        5.8,
                        "F2" if is_bold else "F1",
                        black,
                        align,
                        is_bold,
                    )
                )
                x += width
            y += row_height

        footer_y = page_height - margin_y - 30
        signature_width = (available_width - 60) / 3
        for index, label in enumerate(("Entrega", "Recibe", "Validación")):
            x = margin_x + index * (signature_width + 30)
            commands.append(_pdf_line_embarques(x, footer_y, x + signature_width, footer_y, page_height, navy, 0.7))
            commands.append(
                _pdf_text_embarques(
                    x,
                    footer_y + 12,
                    label,
                    page_height,
                    7,
                    "F1",
                    gray,
                    max_width=signature_width,
                    align="center",
                )
            )

        page_streams.append("\n".join(commands))

    return _crear_pdf_embarques(page_streams, page_width=page_width, page_height=page_height)


@bp.route("/almacen-embarques-entradas-ajax")
@login_requerido
def almacen_embarques_entradas_ajax():
    """Ruta AJAX para visualizar historial de entradas de almacén de embarques."""
    try:
        return render_template(
            "Control de proceso/almacen_embarques_entradas_ajax.html"
        )
    except Exception as e:
        print(f"Error al cargar template Almacén Embarques Entradas AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@bp.route("/almacen-embarques-salidas-ajax")
@login_requerido
def almacen_embarques_salidas_ajax():
    """Ruta AJAX para visualizar historial de salidas de almacén de embarques."""
    try:
        return render_template(
            "Control de proceso/almacen_embarques_salidas_ajax.html"
        )
    except Exception as e:
        print(f"Error al cargar template Almacén Embarques Salidas AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@bp.route("/almacen-embarques-retorno-ajax")
@login_requerido
def almacen_embarques_retorno_ajax():
    """Ruta AJAX para visualizar historial de retornos de almacén de embarques."""
    try:
        response = make_response(render_template(
            "Control de proceso/almacen_embarques_retorno_ajax.html"
        ))
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response
    except Exception as e:
        print(f"Error al cargar template Almacén Embarques Retorno AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@bp.route("/almacen-embarques-movimientos-ajax")
@login_requerido
def almacen_embarques_movimientos_ajax():
    """Ruta AJAX para visualizar y ajustar movimientos de almacén de embarques."""
    try:
        return render_template(
            "Control de proceso/almacen_embarques_movimientos_ajax.html"
        )
    except Exception as e:
        print(f"Error al cargar template Almacén Embarques Movimientos AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@bp.route("/almacen-embarques-inventario-general-ajax")
@login_requerido
def almacen_embarques_inventario_general_ajax():
    """Ruta AJAX para visualizar inventario general de almacén de embarques."""
    try:
        return render_template(
            "Control de proceso/almacen_embarques_inventario_general_ajax.html"
        )
    except Exception as e:
        print(
            f"Error al cargar template Almacén Embarques Inventario General AJAX: {e}"
        )
        return f"Error al cargar el contenido: {str(e)}", 500


@bp.route("/almacen-embarques-catalogo-ajax")
@login_requerido
def almacen_embarques_catalogo_ajax():
    """Ruta AJAX para administrar catálogo de números de parte de embarques."""
    try:
        response = make_response(render_template(
            "Control de proceso/almacen_embarques_catalogo_ajax.html"
        ))
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response
    except Exception as e:
        print(f"Error al cargar template Almacén Embarques Catálogo AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@bp.route("/api/almacen-embarques/entradas")
@login_requerido
def api_almacen_embarques_entradas():
    """Obtener historial de entradas de almacén de embarques."""
    try:
        return jsonify(_obtener_historial_entradas_almacen_embarques())
    except Exception as e:
        print(f"Error API entradas almacén embarques: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


@bp.route("/api/almacen-embarques/entradas/export")
@login_requerido
def export_almacen_embarques_entradas():
    """Exportar historial de entradas de almacén de embarques a Excel."""
    try:
        rows = _obtener_historial_entradas_almacen_embarques(limit=5000)
        return _exportar_historial_embarques_excel(
            "Entradas Embarques",
            "entradas_almacen_embarques.xlsx",
            {
                "Fecha": "fecha",
                "Hora": "hora",
                "Folio": "folio",
                "No. Parte": "part_number",
                "Cantidad": "cantidad",
                "Modelo": "product_model",
                "Cliente": "customer",
                "Zona": "zone_code",
                "Ubicación": "location_code",
                "Usuario": "registered_by",
            },
            rows,
        )
    except Exception as e:
        print(
            f"Error exportando entradas almacén embarques: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"error": str(e)}), 500


@bp.route("/api/almacen-embarques/salidas")
@login_requerido
def api_almacen_embarques_salidas():
    """Obtener historial de salidas de almacén de embarques."""
    try:
        return jsonify(_obtener_historial_salidas_almacen_embarques())
    except Exception as e:
        print(f"Error API salidas almacén embarques: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


@bp.route("/api/almacen-embarques/salidas/export")
@login_requerido
def export_almacen_embarques_salidas():
    """Exportar historial de salidas de almacén de embarques a Excel."""
    try:
        rows = _obtener_historial_salidas_almacen_embarques(limit=5000)
        return _exportar_historial_embarques_excel(
            "Salidas Embarques",
            "salidas_almacen_embarques.xlsx",
            {
                "Fecha": "fecha",
                "Hora": "hora",
                "Folio": "folio",
                "No. Parte": "part_number",
                "Cantidad": "cantidad",
                "Departure": "departure_code",
                "Modelo": "product_model",
                "Cliente": "customer",
                "Destino": "destination_area",
                "Motivo": "reason",
                "Usuario": "registered_by",
            },
            rows,
        )
    except Exception as e:
        print(
            f"Error exportando salidas almacén embarques: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"error": str(e)}), 500


@bp.route("/api/almacen-embarques/<module_name>/ajustes/template")
@login_requerido
def api_almacen_embarques_ajustes_template(module_name):
    """Descargar plantilla CSV para ajustes manuales por lote."""
    try:
        config = _obtener_config_ajuste_manual_embarques(module_name)
        inventory_map = _obtener_mapa_inventario_ajuste_manual_embarques()
        output = io.StringIO()
        writer = csv.writer(output, lineterminator="\n")
        writer.writerow(["part_number", "quantity"])
        for part_number in sorted(inventory_map.keys()):
            writer.writerow([part_number, ""])

        filename = (
            f"{config['template_filename']}_{datetime.now().strftime('%Y%m%d')}.csv"
        )
        return send_file(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            mimetype="text/csv",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        print(
            f"Error generando plantilla ajustes {module_name}: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/<module_name>/ajustes/preview", methods=["POST"])
@login_requerido
def api_almacen_embarques_ajustes_preview(module_name):
    """Validar archivo CSV/XLSX y persistir preview de ajuste manual."""
    try:
        config = _obtener_config_ajuste_manual_embarques(module_name)
        uploaded_file = request.files.get("adjustment_file")
        reason = normalize_search(request.form.get("reason"))
        try:
            movement_at = _normalizar_fecha_movimiento_ajuste_embarques(
                request.form.get("movement_at")
            )
        except ValueError as exc:
            return jsonify({"success": False, "error": str(exc)}), 400

        if not uploaded_file or not uploaded_file.filename:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "Debes seleccionar un archivo CSV o Excel.",
                    }
                ),
                400,
            )

        if not reason:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "El motivo del ajuste es obligatorio.",
                    }
                ),
                400,
            )

        init_shipping_material_tables()
        parsed = _parsear_archivo_ajuste_manual_embarques(uploaded_file)
        if not parsed["valid"]:
            return jsonify(
                {
                    "success": True,
                    "valid": False,
                    "errors": parsed.get("errors") or [],
                    "batchId": None,
                }
            )

        preview = _construir_preview_ajuste_manual_embarques(
            config,
            parsed.get("rows") or [],
            movement_at,
        )
        if not preview["valid"]:
            return jsonify(
                {
                    "success": True,
                    "valid": False,
                    "errors": preview.get("errors") or [],
                    "preview": preview,
                    "batchId": None,
                }
            )

        batch_id, batch_code = _guardar_preview_ajuste_manual_embarques(
            config,
            movement_at,
            reason,
            uploaded_file.filename,
            parsed["fileHash"],
            preview,
        )

        return jsonify(
            {
                "success": True,
                "valid": True,
                "batchId": batch_id,
                "batchCode": batch_code,
                "preview": preview,
                "message": "Preview validado correctamente.",
            }
        )
    except Exception as e:
        print(
            f"Error preview ajustes {module_name}: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/<module_name>/ajustes/confirm", methods=["POST"])
@login_requerido
def api_almacen_embarques_ajustes_confirm(module_name):
    """Confirmar un lote validado de ajustes manuales."""
    try:
        config = _obtener_config_ajuste_manual_embarques(module_name)
        data = request.get_json(silent=True) or {}
        batch_id = int(data.get("batchId") or 0)
        if batch_id <= 0:
            return jsonify({"success": False, "error": "batchId inválido."}), 400

        payload, status_code = _confirmar_ajuste_manual_embarques(config, batch_id)
        return jsonify(payload), status_code
    except Exception as e:
        print(
            f"Error confirmando ajustes {module_name}: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/<module_name>/ajustes/manual", methods=["POST"])
@login_requerido
def api_almacen_embarques_ajustes_manual(module_name):
    """Registrar un ajuste manual unitario con la misma trazabilidad de lotes."""
    try:
        config = _obtener_config_ajuste_manual_embarques(module_name)
        data = request.get_json(silent=True) or {}
        init_shipping_material_tables()
        payload, status_code = _registrar_ajuste_manual_individual_embarques(
            config,
            data,
        )
        return jsonify(payload), status_code
    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        print(
            f"Error registrando ajuste manual {module_name}: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/<module_name>/ajustes/cancel", methods=["POST"])
@login_requerido
def api_almacen_embarques_ajustes_cancel(module_name):
    """Cancelar un preview draft de ajustes manuales."""
    try:
        config = _obtener_config_ajuste_manual_embarques(module_name)
        data = request.get_json(silent=True) or {}
        batch_id = int(data.get("batchId") or 0)
        if batch_id <= 0:
            return jsonify({"success": False, "error": "batchId inválido."}), 400

        conn = get_pooled_connection()
        if conn is None:
            raise RuntimeError("No fue posible obtener conexión MySQL.")

        cursor = get_dict_cursor(conn)
        try:
            conn.autocommit(False)
            cursor.execute(
                f"""
                UPDATE `{SHIPPING_TABLES['manual_adjustment_batches']}`
                SET status = 'cancelled',
                    cancelled_at = NOW()
                WHERE id = %s
                  AND movement_type = %s
                  AND status = 'draft'
                """,
                (batch_id, config["movement_type"]),
            )
            if cursor.rowcount <= 0:
                conn.rollback()
                return (
                    jsonify(
                        {
                            "success": False,
                            "error": "No se encontró un preview pendiente para cancelar.",
                        }
                    ),
                    404,
                )
            conn.commit()
            return jsonify({"success": True, "message": "Preview cancelado."})
        except Exception:
            conn.rollback()
            raise
        finally:
            try:
                conn.autocommit(True)
            except Exception:
                pass
            cursor.close()
            conn.close()
    except Exception as e:
        print(
            f"Error cancelando ajustes {module_name}: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/salidas/<int:exit_id>/departure", methods=["POST", "PUT", "PATCH"])
@login_requerido
def assign_almacen_embarques_departure(exit_id):
    """Asignar o reasignar departure a una salida de almacén de embarques."""
    try:
        data = request.get_json(silent=True) or {}
        assigned_by = session.get(
            "nombre_completo", session.get("usuario", "Sistema")
        )
        payload, status_code = assign_exit_departure_value(
            exit_id,
            data.get("departureCode") or data.get("departure"),
            assigned_by,
            departure_quantity=(
                data.get("departureQuantity")
                if data.get("departureQuantity") is not None
                else data.get("quantity")
            ),
            notes=data.get("notes"),
            assigned_at=data.get("assignedAt"),
        )
        return jsonify(payload), status_code
    except Exception as e:
        print(
            f"Error asignando departure a salida {exit_id}: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/departures/history")
@login_requerido
def api_almacen_embarques_departure_history():
    """Consultar historial de asignaciones de departure ligadas a salidas."""
    try:
        payload = get_departure_history_records(
            limit=request.args.get("limit"),
            departure_code=request.args.get("departureCode")
            or request.args.get("departure"),
            exit_id=request.args.get("exitId"),
        )
        return jsonify(payload)
    except Exception as e:
        print(
            f"Error consultando historial de departures: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/retorno")
@login_requerido
def api_almacen_embarques_retorno():
    """Obtener historial de retornos de almacén de embarques."""
    try:
        return jsonify(_obtener_historial_retorno_almacen_embarques())
    except Exception as e:
        print(f"Error API retorno almacén embarques: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


@bp.route("/api/almacen-embarques/retorno/export")
@login_requerido
def export_almacen_embarques_retorno():
    """Exportar historial de retornos de almacén de embarques a Excel."""
    try:
        rows = _obtener_historial_retorno_almacen_embarques(limit=5000)
        movement = (request.args.get("movement", "") or "").strip().lower()

        if movement == "entry":
            rows = [
                row
                for row in rows
                if (row.get("return_quantity") or 0) > 0
                and (row.get("loss_quantity") or 0) <= 0
            ]
            for row in rows:
                row["movement_quantity"] = row.get("return_quantity") or 0
            sheet_name = "Entradas Retorno"
            filename = "entradas_retorno_almacen_embarques.xlsx"
            headers = {
                "Fecha": "fecha",
                "Hora": "hora",
                "Folio": "folio",
                "No. Parte": "part_number",
                "Cantidad entrada": "movement_quantity",
                "Modelo": "product_model",
                "Tipo": "reason",
                "Usuario": "registered_by",
            }
        elif movement == "exit":
            rows = [row for row in rows if (row.get("loss_quantity") or 0) > 0]
            for row in rows:
                row["movement_quantity"] = row.get("loss_quantity") or 0
            sheet_name = "Salidas Retorno"
            filename = "salidas_retorno_almacen_embarques.xlsx"
            headers = {
                "Fecha": "fecha",
                "Hora": "hora",
                "Folio": "folio",
                "No. Parte": "part_number",
                "Cantidad salida": "movement_quantity",
                "Modelo": "product_model",
                "Tipo": "reason",
                "Usuario": "registered_by",
            }
        else:
            sheet_name = "Retorno Embarques"
            filename = "retorno_almacen_embarques.xlsx"
            headers = {
                "Fecha": "fecha",
                "Hora": "hora",
                "Folio": "folio",
                "No. Parte": "part_number",
                "Cantidad retorno": "return_quantity",
                "Cantidad pérdida": "loss_quantity",
                "Modelo": "product_model",
                "Cliente": "customer",
                "Tipo": "reason",
                "Usuario": "registered_by",
            }

        return _exportar_historial_embarques_excel(
            sheet_name,
            filename,
            headers,
            rows,
        )
    except Exception as e:
        print(
            f"Error exportando retorno almacén embarques: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"error": str(e)}), 500


@bp.route("/api/almacen-embarques/retorno/print-pdf", methods=["POST"])
@login_requerido
def export_almacen_embarques_retorno_print_pdf():
    """Generar PDF imprimible de salidas de retorno seleccionadas."""
    try:
        data = request.get_json(silent=True) or {}
        rows = data.get("rows") or []
        if not isinstance(rows, list):
            return jsonify({"success": False, "error": "Formato de registros inválido."}), 400

        pdf_output = _generar_pdf_salidas_retorno_embarques(rows)
        filename = f"formato_salidas_retorno_{obtener_fecha_hora_mexico().strftime('%Y%m%d%H%M')}.pdf"
        return send_file(
            pdf_output,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename,
        )
    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        print(
            f"Error generando PDF salidas retorno almacén embarques: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/movimientos")
@login_requerido
def api_almacen_embarques_movimientos():
    """Obtener historial unificado editable de movimientos de embarques."""
    try:
        return jsonify(
            {
                "success": True,
                "rows": _obtener_movimientos_editables_almacen_embarques(),
            }
        )
    except Exception as e:
        print(
            f"Error API movimientos almacén embarques: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/movimientos/export")
@login_requerido
def export_almacen_embarques_movimientos():
    """Exportar historial editable de movimientos de embarques."""
    try:
        rows = _obtener_movimientos_editables_almacen_embarques(limit=5000)
        return _exportar_historial_embarques_excel(
            "Movimientos Embarques",
            "movimientos_almacen_embarques.xlsx",
            {
                "Fecha": "fecha",
                "Hora": "hora",
                "Tipo": "movement_label",
                "Folio": "folio",
                "No. Parte": "part_number",
                "Cantidad": "quantity_primary",
                "Modelo": "product_model",
                "Cliente": "customer",
                "Zona": "zone_code",
                "Ubicacion / Destino": "location_value",
                "Departure": "departure_code",
                "Usuario": "registered_by",
            },
            rows,
        )
    except Exception as e:
        print(
            f"Error exportando movimientos de embarques: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/movimientos/<movement_type>/<int:record_id>")
@login_requerido
def api_almacen_embarques_movimiento_detalle(movement_type, record_id):
    """Obtener detalle actual de un movimiento editable de embarques."""
    try:
        movement = _obtener_detalle_movimiento_almacen_embarques(
            movement_type, record_id
        )
        if not movement:
            return jsonify({"success": False, "error": "Movimiento no encontrado"}), 404
        return jsonify({"success": True, "movement": movement})
    except Exception as e:
        print(
            f"Error obteniendo detalle de movimiento {movement_type}/{record_id}: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route(
    "/api/almacen-embarques/movimientos/<movement_type>/<int:record_id>",
    methods=["PATCH", "PUT", "POST"],
)
@login_requerido
def api_almacen_embarques_movimiento_update(movement_type, record_id):
    """Actualizar un movimiento de embarques y registrar el historial del ajuste."""
    try:
        data = request.get_json(silent=True) or {}
        notes = (data.get("notes") or "").strip()
        if not notes:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "El motivo del ajuste es obligatorio",
                    }
                ),
                400,
            )
        payload, status_code = adjust_shipping_movement_record(
            movement_type,
            record_id,
            data.get("changes") or data,
            session.get("nombre_completo", session.get("usuario", "Sistema")),
            notes=notes,
        )
        return jsonify(payload), status_code
    except Exception as e:
        print(
            f"Error actualizando movimiento {movement_type}/{record_id}: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route(
    "/api/almacen-embarques/movimientos/<movement_type>/<int:record_id>",
    methods=["DELETE"],
)
@login_requerido
def api_almacen_embarques_movimiento_delete(movement_type, record_id):
    """Eliminar un movimiento de embarques con confirmación de contraseña."""
    try:
        data = request.get_json(silent=True) or {}
        password = data.get("password")
        notes = (data.get("notes") or "").strip()
        if not notes:
            return jsonify({"success": False, "error": "El comentario de eliminación es obligatorio"}), 400

        is_valid_password, password_error = _validar_password_usuario_actual(password)
        if not is_valid_password:
            return jsonify({"success": False, "error": password_error}), 400

        payload, status_code = delete_shipping_movement_record(
            movement_type,
            record_id,
            _obtener_usuario_display_actual(),
            notes=notes,
        )
        return jsonify(payload), status_code
    except Exception as e:
        print(
            f"Error eliminando movimiento {movement_type}/{record_id}: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/inventario-general")
@login_requerido
def api_almacen_embarques_inventario_general():
    """Obtener inventario general del periodo para almacén de embarques."""
    try:
        payload = _obtener_inventario_general_almacen_embarques()
        return jsonify({"success": True, **payload})
    except Exception as e:
        print(
            f"Error API inventario general embarques: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/catalogo")
@login_requerido
def api_almacen_embarques_catalogo():
    """Obtener catálogo de números de parte de almacén de embarques."""
    try:
        init_shipping_material_tables()
        payload = _obtener_catalogo_almacen_embarques()
        return jsonify({"success": True, **payload})
    except Exception as e:
        print(f"Error API catálogo embarques: {e}\n{traceback.format_exc()}")
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/catalogo", methods=["POST"])
@login_requerido
def api_almacen_embarques_catalogo_create():
    """Crear un número de parte en catálogo e inventario de embarques."""
    try:
        init_shipping_material_tables()
        data = request.get_json(silent=True) or {}
        payload, status_code = _crear_catalogo_almacen_embarques(data)
        return jsonify(payload), status_code
    except Exception as e:
        print(f"Error creando catálogo embarques: {e}\n{traceback.format_exc()}")
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/catalogo/<int:catalog_id>", methods=["PATCH"])
@login_requerido
def api_almacen_embarques_catalogo_update(catalog_id):
    """Actualizar campos del catálogo de embarques y su snapshot de inventario."""
    try:
        init_shipping_material_tables()
        data = request.get_json(silent=True) or {}
        payload, status_code = _actualizar_catalogo_almacen_embarques(catalog_id, data)
        return jsonify(payload), status_code
    except Exception as e:
        print(
            f"Error actualizando catálogo embarques {catalog_id}: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/catalogo/<int:catalog_id>", methods=["DELETE"])
@login_requerido
def api_almacen_embarques_catalogo_delete(catalog_id):
    """Eliminar o dar de baja lógica a un número de parte del catálogo."""
    try:
        init_shipping_material_tables()
        data = request.get_json(silent=True) or {}
        payload, status_code = _eliminar_catalogo_almacen_embarques(catalog_id, data)
        return jsonify(payload), status_code
    except Exception as e:
        print(
            f"Error eliminando catálogo embarques {catalog_id}: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/catalogo/export")
@login_requerido
def export_almacen_embarques_catalogo():
    """Exportar catálogo de números de parte de embarques."""
    try:
        payload = _obtener_catalogo_almacen_embarques(limit=20000)
        return _exportar_historial_embarques_excel(
            "Catalogo Embarques",
            "catalogo_almacen_embarques.xlsx",
            {
                "No. Parte": "part_number",
                "Modelo": "product_model",
                "Estatus": "product_status",
                "Descripción": "description",
                "Std Pack": "standard_pack",
                "Cliente": "customer",
                "Zona": "zone_code",
                "Cantidad actual": "current_quantity",
            },
            payload.get("rows") or [],
        )
    except Exception as e:
        print(f"Error exportando catálogo embarques: {e}\n{traceback.format_exc()}")
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/inventario-general/export")
@login_requerido
def export_almacen_embarques_inventario_general():
    """Exportar inventario general del periodo de embarques."""
    try:
        payload = _obtener_inventario_general_almacen_embarques(limit=20000)
        return _exportar_historial_embarques_excel(
            "Inventario General Embarques",
            "inventario_general_almacen_embarques.xlsx",
            {
                "No. Parte": "part_number",
                "Modelo": "product_model",
                "Cliente": "customer",
                "Inventario inicial": "initial_quantity",
                "Entradas": "entries_qty",
                "Salidas": "exits_qty",
                "Entradas retorno": "return_entries_qty",
                "Salidas retorno": "return_exits_qty",
                "Cantidad total": "current_quantity",
                "Inicio de periodo": "period_start",
                "Cierre": "closure_label",
            },
            payload.get("rows") or [],
        )
    except Exception as e:
        print(
            f"Error exportando inventario general embarques: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/inventario-general/cierre/bootstrap")
@login_requerido
def api_almacen_embarques_inventario_cierre_bootstrap():
    """Obtener contexto inicial y baseline del cierre de inventario de embarques."""
    try:
        payload = _obtener_inventario_general_almacen_embarques(limit=20000)
        preview = _construir_preview_cierre_inventario_embarques(payload.get("rows") or [])
        return jsonify(
            {
                "success": True,
                "metadata": _obtener_contexto_cierre_inventario_embarques(),
                "preview": preview,
                "history": _obtener_historial_cierres_inventario_embarques(
                    include_last_draft_for=_obtener_usuario_display_actual()
                ),
            }
        )
    except Exception as e:
        print(
            f"Error bootstrap cierre inventario embarques: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/inventario-general/cierre/template")
@login_requerido
def api_almacen_embarques_inventario_cierre_template():
    """Descargar plantilla CSV del catálogo completo para cierre de inventario."""
    try:
        payload = _obtener_inventario_general_almacen_embarques(limit=20000)
        output = io.StringIO()
        writer = csv.writer(output, lineterminator="\n")
        writer.writerow(["part_number", "current_qty"])
        for row in payload.get("rows") or []:
            writer.writerow([row.get("part_number") or "", ""])

        content = output.getvalue().encode("utf-8-sig")
        filename = f"plantilla_cierre_inventario_embarques_{datetime.now().strftime('%Y%m%d')}.csv"
        return send_file(
            io.BytesIO(content),
            mimetype="text/csv",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        print(
            f"Error generando plantilla CSV cierre inventario embarques: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/inventario-general/cierre/preview", methods=["POST"])
@login_requerido
def api_almacen_embarques_inventario_cierre_preview():
    """Validar CSV de cierre y generar preview persistente en borrador."""
    try:
        csv_file = request.files.get("closure_file")
        if not csv_file or not csv_file.filename:
            return jsonify(
                {
                    "success": False,
                    "error": "Debes seleccionar un archivo CSV para validar.",
                }
            ), 400

        current_payload = _obtener_inventario_general_almacen_embarques(limit=20000)
        current_rows = current_payload.get("rows") or []
        expected_parts = [row.get("part_number") for row in current_rows if row.get("part_number")]

        parsed_csv = _parsear_csv_cierre_inventario_embarques(csv_file, expected_parts)
        preview = _construir_preview_cierre_inventario_embarques(
            current_rows,
            parsed_csv.get("rows") if parsed_csv.get("rows") else None,
        )
        metadata = _obtener_contexto_cierre_inventario_embarques()

        response_payload = {
            "success": True,
            "valid": parsed_csv["valid"],
            "metadata": metadata,
            "preview": preview,
            "errors": parsed_csv.get("errors") or [],
            "history": _obtener_historial_cierres_inventario_embarques(
                include_last_draft_for=_obtener_usuario_display_actual()
            ),
            "batchId": None,
        }

        if parsed_csv["valid"]:
            draft_payload = {
                "metadata": {
                    **metadata,
                    "csvFileName": csv_file.filename,
                    "csvHash": parsed_csv["csvHash"],
                },
                "summary": preview["summary"],
                "rows": preview["rows"],
            }
            batch_id = _guardar_borrador_cierre_inventario_embarques(
                draft_payload,
                csv_file.filename,
                parsed_csv["csvHash"],
            )
            response_payload["batchId"] = batch_id

        return jsonify(response_payload)
    except Exception as e:
        print(
            f"Error preview cierre inventario embarques: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/inventario-general/cierre/confirm", methods=["POST"])
@login_requerido
def api_almacen_embarques_inventario_cierre_confirm():
    """Confirmar lote de cierre validado y registrarlo como base del siguiente periodo."""
    try:
        data = request.get_json(silent=True) or {}
        batch_id = int(data.get("batchId") or 0)
        if batch_id <= 0:
            return jsonify({"success": False, "error": "batchId inválido."}), 400

        conn = get_pooled_connection()
        if conn is None:
            raise RuntimeError("No fue posible obtener conexión MySQL para confirmar el cierre.")

        cursor = get_dict_cursor(conn)
        usuario_actual = _obtener_usuario_display_actual()

        try:
            conn.autocommit(False)
            cursor.execute(
                f"""
                SELECT *
                FROM `{SHIPPING_TABLES['inventory_closure_batches']}`
                WHERE id = %s
                LIMIT 1
                """,
                (batch_id,),
            )
            batch_row = cursor.fetchone()
            if not batch_row:
                return jsonify({"success": False, "error": "No se encontró el borrador del cierre."}), 404

            if (batch_row.get("status") or "").lower() != "draft":
                return jsonify({"success": False, "error": "El lote indicado ya no está disponible para confirmar."}), 409

            cursor.execute(
                f"""
                SELECT id
                FROM `{SHIPPING_TABLES['inventory_closure_batches']}`
                WHERE closure_month = %s
                  AND status = 'confirmed'
                LIMIT 1
                """,
                (batch_row.get("closure_month"),),
            )
            existing_confirmed = cursor.fetchone()
            if existing_confirmed:
                return jsonify({"success": False, "error": "Ya existe un cierre confirmado para ese mes."}), 409

            payload_json = batch_row.get("payload_json") or "{}"
            payload = json.loads(payload_json)
            preview_rows = payload.get("rows") or []
            metadata = payload.get("metadata") or {}

            for row in preview_rows:
                csv_current_qty = row.get("csv_current_qty")
                if csv_current_qty is None:
                    raise ValueError(
                        f"El lote contiene part_numbers sin inventario físico cargado: {row.get('part_number')}"
                    )

                cursor.execute(
                    f"""
                    INSERT INTO `{SHIPPING_TABLES['inventory_closures']}` (
                      closure_batch_id,
                      part_number,
                      initial_quantity,
                      system_quantity,
                      difference_quantity,
                      raw_current_quantity,
                      closed_at,
                      closure_label,
                      closed_by,
                      notes
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        batch_id,
                        row.get("part_number"),
                        max(int(row.get("applied_initial_quantity") or 0), 0),
                        int(row.get("system_quantity") or 0),
                        int(row.get("difference_quantity") or 0),
                        int(csv_current_qty),
                        metadata.get("closureDateTime"),
                        metadata.get("closureLabel"),
                        usuario_actual,
                        json.dumps(
                            {
                                "status": row.get("status"),
                                "csvHash": metadata.get("csvHash"),
                                "rowsHash": payload.get("summary", {}).get("rowsHash"),
                            },
                            ensure_ascii=False,
                            separators=(",", ":"),
                        ),
                    ),
                )

            cursor.execute(
                f"""
                UPDATE `{SHIPPING_TABLES['inventory_closure_batches']}`
                SET status = 'confirmed',
                    confirmed_by = %s,
                    confirmed_at = %s
                WHERE id = %s
                """,
                (usuario_actual, metadata.get("closureDateTime"), batch_id),
            )

            conn.commit()
            return jsonify(
                {
                    "success": True,
                    "message": "Cierre de inventario confirmado correctamente.",
                    "history": _obtener_historial_cierres_inventario_embarques(
                        include_last_draft_for=_obtener_usuario_display_actual()
                    ),
                }
            )
        except Exception:
            conn.rollback()
            raise
        finally:
            try:
                conn.autocommit(True)
            except Exception:
                pass
            cursor.close()
            conn.close()
    except Exception as e:
        print(
            f"Error confirmando cierre inventario embarques: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/inventario-general/cierre/cancel", methods=["POST"])
@login_requerido
def api_almacen_embarques_inventario_cierre_cancel():
    """Cancelar un borrador de preview de cierre para reiniciar el proceso."""
    try:
        data = request.get_json(silent=True) or {}
        batch_id = int(data.get("batchId") or 0)
        if batch_id <= 0:
            return jsonify({"success": False, "error": "batchId inválido."}), 400

        usuario_actual = _obtener_usuario_display_actual()
        conn = get_pooled_connection()
        if conn is None:
            raise RuntimeError("No fue posible obtener conexión MySQL para cancelar el borrador.")

        cursor = get_dict_cursor(conn)
        try:
            conn.autocommit(False)
            cursor.execute(
                f"""
                SELECT id, status, created_by
                FROM `{SHIPPING_TABLES['inventory_closure_batches']}`
                WHERE id = %s
                LIMIT 1
                """,
                (batch_id,),
            )
            batch_row = cursor.fetchone()
            if not batch_row:
                return jsonify({"success": False, "error": "No se encontró el borrador del cierre."}), 404

            if (batch_row.get("status") or "").lower() != "draft":
                return jsonify({"success": False, "error": "Solo se pueden cancelar previews en estado draft."}), 409

            if (
                (batch_row.get("created_by") or "").strip().lower()
                != (usuario_actual or "").strip().lower()
            ):
                return jsonify({"success": False, "error": "Solo el usuario que generó el preview puede cancelarlo."}), 403

            cursor.execute(
                f"""
                UPDATE `{SHIPPING_TABLES['inventory_closure_batches']}`
                SET status = 'cancelled'
                WHERE id = %s
                """,
                (batch_id,),
            )
            conn.commit()
            return jsonify(
                {
                    "success": True,
                    "message": "Preview cancelado correctamente.",
                    "history": _obtener_historial_cierres_inventario_embarques(
                        include_last_draft_for=_obtener_usuario_display_actual()
                    ),
                }
            )
        except Exception:
            conn.rollback()
            raise
        finally:
            try:
                conn.autocommit(True)
            except Exception:
                pass
            cursor.close()
            conn.close()
    except Exception as e:
        print(
            f"Error cancelando preview de cierre inventario embarques: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/inventario-general/cierre/history/<int:batch_id>")
@login_requerido
def api_almacen_embarques_inventario_cierre_history_detail(batch_id):
    """Consultar detalle completo de un cierre de inventario."""
    try:
        row = execute_query(
            f"""
            SELECT *
            FROM `{SHIPPING_TABLES['inventory_closure_batches']}`
            WHERE id = %s
            LIMIT 1
            """,
            (batch_id,),
            fetch="one",
        )
        if not row:
            return jsonify({"success": False, "error": "No se encontró el cierre solicitado."}), 404

        payload = json.loads(row.get("payload_json") or "{}")
        period_info = _resolver_periodo_cierre_embarques(row)
        payload = _aplicar_metadata_periodo_cierre_embarques(payload, period_info)
        return jsonify(
            {
                "success": True,
                "batch": {
                    "id": row.get("id"),
                    "closure_label": period_info["closure_label"],
                    "closure_month": period_info["closure_month"],
                    "status": row.get("status"),
                    "created_by": row.get("created_by"),
                    "confirmed_by": row.get("confirmed_by"),
                    "closed_at": _normalizar_texto_embarques_historial(row.get("closed_at")),
                    "confirmed_at": _normalizar_texto_embarques_historial(row.get("confirmed_at")),
                    "accuracy_pct": row.get("accuracy_pct"),
                    "csv_hash": row.get("csv_hash"),
                    "rows_hash": row.get("rows_hash"),
                },
                "payload": payload,
            }
        )
    except Exception as e:
        print(
            f"Error consultando detalle de cierre inventario embarques: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/almacen-embarques/inventario-general/cierre/history/<int:batch_id>/export")
@login_requerido
def export_almacen_embarques_inventario_cierre_report(batch_id):
    """Exportar reporte completo del cierre de inventario en Excel."""
    try:
        from io import BytesIO

        from flask import Response
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        batch = execute_query(
            f"""
            SELECT *
            FROM `{SHIPPING_TABLES['inventory_closure_batches']}`
            WHERE id = %s
            LIMIT 1
            """,
            (batch_id,),
            fetch="one",
        )
        if not batch:
            return jsonify({"success": False, "error": "No se encontró el cierre solicitado."}), 404
        if (batch.get("status") or "").lower() != "confirmed":
            return jsonify({"success": False, "error": "Sólo se pueden exportar cierres confirmados."}), 409

        raw_period_end = batch.get("closed_at")
        previous_batch = execute_query(
            f"""
            SELECT id, closed_at
            FROM `{SHIPPING_TABLES['inventory_closure_batches']}`
            WHERE status = 'confirmed'
              AND closed_at < %s
            ORDER BY closed_at DESC, id DESC
            LIMIT 1
            """,
            (raw_period_end,),
            fetch="one",
        )
        period_info = _resolver_periodo_cierre_embarques(
            batch,
            previous_closed_at=(previous_batch or {}).get("closed_at"),
        )
        period_start = (previous_batch or {}).get("closed_at") or period_info["period_start"]
        period_end = period_info["period_end_exclusive"] or raw_period_end
        period_end_inclusive = period_info["period_end_inclusive"]
        previous_batch_id = (previous_batch or {}).get("id")

        def period_filter(column_expr):
            clauses = []
            params = []
            if period_start:
                clauses.append(f"{column_expr} >= %s")
                params.append(period_start)
            if period_end:
                clauses.append(f"{column_expr} < %s")
                params.append(period_end)
            return " AND ".join(clauses) if clauses else "1 = 1", params

        def normalize_excel_value(value):
            if isinstance(value, Decimal):
                try:
                    as_int = int(value)
                    return as_int if value == as_int else float(value)
                except Exception:
                    return float(value)
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d %H:%M:%S")
            if isinstance(value, date):
                return value.strftime("%Y-%m-%d")
            return value

        def parse_json_summary(raw_value):
            if not raw_value:
                return ""
            try:
                parsed = json.loads(raw_value)
            except Exception:
                return str(raw_value)
            if isinstance(parsed, dict):
                return " | ".join(
                    f"{key}: {normalize_excel_value(value)}"
                    for key, value in parsed.items()
                )
            if isinstance(parsed, list):
                return ", ".join(str(item) for item in parsed)
            return str(parsed)

        def index_quantity(rows, part_key, quantity_key):
            indexed = {}
            for item in rows:
                part_number = _normalizar_texto_embarques_historial(item.get(part_key))
                indexed[part_number] = _normalizar_numero_embarques_historial(
                    item.get(quantity_key)
                ) or 0
            return indexed

        closures = execute_query(
            f"""
            SELECT
              part_number,
              initial_quantity,
              system_quantity,
              difference_quantity,
              raw_current_quantity
            FROM `{SHIPPING_TABLES['inventory_closures']}`
            WHERE closure_batch_id = %s
            ORDER BY part_number ASC
            """,
            (batch_id,),
            fetch="all",
        ) or []

        previous_closures = []
        if previous_batch_id:
            previous_closures = execute_query(
                f"""
                SELECT part_number, initial_quantity
                FROM `{SHIPPING_TABLES['inventory_closures']}`
                WHERE closure_batch_id = %s
                """,
                (previous_batch_id,),
                fetch="all",
            ) or []

        movement_clause, movement_params = period_filter("COALESCE(movement_at, created_at)")
        entries_agg = execute_query(
            f"""
            SELECT part_number, COALESCE(SUM(quantity), 0) AS quantity
            FROM `{SHIPPING_TABLES['entries']}`
            WHERE {movement_clause}
              AND COALESCE(is_fifo_layer_only, 0) = 0
            GROUP BY part_number
            """,
            tuple(movement_params),
            fetch="all",
        ) or []
        exits_agg = execute_query(
            f"""
            SELECT part_number, COALESCE(SUM(quantity), 0) AS quantity
            FROM `{SHIPPING_TABLES['exits']}`
            WHERE {movement_clause}
            GROUP BY part_number
            """,
            tuple(movement_params),
            fetch="all",
        ) or []
        returns_agg = execute_query(
            f"""
            SELECT
              part_number,
              COALESCE(SUM(return_quantity), 0) AS return_quantity,
              COALESCE(SUM(loss_quantity), 0) AS loss_quantity
            FROM `{SHIPPING_TABLES['returns']}`
            WHERE {movement_clause}
            GROUP BY part_number
            """,
            tuple(movement_params),
            fetch="all",
        ) or []

        entries_by_part = index_quantity(entries_agg, "part_number", "quantity")
        exits_by_part = index_quantity(exits_agg, "part_number", "quantity")
        return_entries_by_part = index_quantity(returns_agg, "part_number", "return_quantity")
        return_losses_by_part = index_quantity(returns_agg, "part_number", "loss_quantity")
        previous_initial_by_part = index_quantity(previous_closures, "part_number", "initial_quantity")

        cierre_rows = []
        for row in closures:
            part_number = _normalizar_texto_embarques_historial(row.get("part_number"))
            entries_quantity = entries_by_part.get(part_number, 0)
            exits_quantity = exits_by_part.get(part_number, 0)
            return_entries_quantity = return_entries_by_part.get(part_number, 0)
            return_losses_quantity = return_losses_by_part.get(part_number, 0)
            if previous_batch_id:
                initial_quantity = previous_initial_by_part.get(part_number, 0)
            else:
                # Primer cierre: no hay cierre anterior que defina base.
                # Se infiere la base inicial usando el snapshot del sistema al cierre.
                system_quantity = _normalizar_numero_embarques_historial(row.get("system_quantity"))
                if system_quantity is None:
                    system_quantity = 0
                initial_quantity = (
                    system_quantity
                    - entries_quantity
                    + exits_quantity
                    - return_entries_quantity
                    + return_losses_quantity
                )
            cierre_rows.append(
                {
                    "part_number": part_number,
                    "initial_quantity": initial_quantity,
                    "entries": entries_quantity,
                    "exits": exits_quantity,
                    "return_entries": return_entries_quantity,
                    "return_losses": return_losses_quantity,
                    "physical_quantity": _normalizar_numero_embarques_historial(row.get("raw_current_quantity")),
                    "difference_quantity": _normalizar_numero_embarques_historial(row.get("difference_quantity")),
                }
            )

        entries_rows = execute_query(
            f"""
            SELECT
              COALESCE(movement_at, created_at) AS movement_at,
              'Entrada' AS movement_type,
              entry_folio AS folio,
              part_number,
              quantity,
              '' AS departure_code,
              product_model,
              customer,
              registered_by
            FROM `{SHIPPING_TABLES['entries']}`
            WHERE {movement_clause}
              AND COALESCE(is_fifo_layer_only, 0) = 0
            """,
            tuple(movement_params),
            fetch="all",
        ) or []
        exits_rows = execute_query(
            f"""
            SELECT
              COALESCE(movement_at, created_at) AS movement_at,
              'Salida' AS movement_type,
              exit_folio AS folio,
              part_number,
              quantity,
              departure_code,
              product_model,
              customer,
              registered_by
            FROM `{SHIPPING_TABLES['exits']}`
            WHERE {movement_clause}
            """,
            tuple(movement_params),
            fetch="all",
        ) or []

        entradas_salidas_rows = []
        for item in entries_rows + exits_rows:
            movement_at = item.get("movement_at")
            entradas_salidas_rows.append(
                {
                    "fecha": movement_at.strftime("%Y-%m-%d") if isinstance(movement_at, datetime) else "",
                    "hora": movement_at.strftime("%H:%M:%S") if isinstance(movement_at, datetime) else "",
                    "tipo": item.get("movement_type"),
                    "folio": item.get("folio"),
                    "part_number": item.get("part_number"),
                    "quantity": _normalizar_numero_embarques_historial(item.get("quantity")),
                    "departure_code": item.get("departure_code") or "",
                    "product_model": item.get("product_model") or "",
                    "customer": item.get("customer") or "",
                    "registered_by": item.get("registered_by") or "",
                }
            )
        entradas_salidas_rows.sort(
            key=lambda item: (item.get("fecha") or "", item.get("hora") or "", item.get("folio") or "")
        )

        returns_rows_raw = execute_query(
            f"""
            SELECT
              COALESCE(movement_at, created_at) AS movement_at,
              return_folio,
              part_number,
              return_quantity,
              loss_quantity,
              product_model,
              customer,
              reason,
              registered_by
            FROM `{SHIPPING_TABLES['returns']}`
            WHERE {movement_clause}
            ORDER BY COALESCE(movement_at, created_at) ASC, id ASC
            """,
            tuple(movement_params),
            fetch="all",
        ) or []
        retorno_rows = []
        for item in returns_rows_raw:
            movement_at = item.get("movement_at")
            base = {
                "fecha": movement_at.strftime("%Y-%m-%d") if isinstance(movement_at, datetime) else "",
                "hora": movement_at.strftime("%H:%M:%S") if isinstance(movement_at, datetime) else "",
                "folio": item.get("return_folio"),
                "part_number": item.get("part_number"),
                "product_model": item.get("product_model") or "",
                "customer": item.get("customer") or "",
                "return_type": item.get("reason") or "",
                "registered_by": item.get("registered_by") or "",
            }
            return_quantity = _normalizar_numero_embarques_historial(item.get("return_quantity")) or 0
            loss_quantity = _normalizar_numero_embarques_historial(item.get("loss_quantity")) or 0
            if return_quantity:
                retorno_rows.append({**base, "movement_type": "Entrada retorno", "quantity": return_quantity})
            if loss_quantity:
                retorno_rows.append({**base, "movement_type": "Salida retorno", "quantity": loss_quantity})

        adjustment_clause, adjustment_params = period_filter("adjusted_at")
        adjustment_rows_raw = execute_query(
            f"""
            SELECT
              adjusted_at,
              'Movimiento' AS adjustment_type,
              movement_type,
              adjustment_action,
              folio,
              part_number,
              changed_fields_json,
              previous_values_json,
              new_values_json,
              notes,
              adjusted_by
            FROM `{SHIPPING_TABLES['movement_adjustments']}`
            WHERE {adjustment_clause}
            ORDER BY adjusted_at ASC, id ASC
            """,
            tuple(adjustment_params),
            fetch="all",
        ) or []
        departure_clause, departure_params = period_filter("assigned_at")
        departure_rows_raw = execute_query(
            f"""
            SELECT
              assigned_at,
              'Departure' AS adjustment_type,
              'exit' AS movement_type,
              assignment_action AS adjustment_action,
              exit_folio AS folio,
              part_number,
              previous_departure_code,
              departure_code,
              notes,
              assigned_by
            FROM `{SHIPPING_TABLES['departure_history']}`
            WHERE {departure_clause}
            ORDER BY assigned_at ASC, id ASC
            """,
            tuple(departure_params),
            fetch="all",
        ) or []

        modificaciones_rows = []
        for item in adjustment_rows_raw:
            adjusted_at = item.get("adjusted_at")
            modificaciones_rows.append(
                {
                    "fecha": adjusted_at.strftime("%Y-%m-%d") if isinstance(adjusted_at, datetime) else "",
                    "hora": adjusted_at.strftime("%H:%M:%S") if isinstance(adjusted_at, datetime) else "",
                    "adjustment_type": item.get("adjustment_type"),
                    "movement_type": item.get("movement_type"),
                    "folio": item.get("folio"),
                    "part_number": item.get("part_number"),
                    "action": item.get("adjustment_action"),
                    "changed_fields": parse_json_summary(item.get("changed_fields_json")),
                    "before": parse_json_summary(item.get("previous_values_json")),
                    "after": parse_json_summary(item.get("new_values_json")),
                    "notes": item.get("notes") or "",
                    "user": item.get("adjusted_by") or "",
                }
            )
        for item in departure_rows_raw:
            adjusted_at = item.get("assigned_at")
            modificaciones_rows.append(
                {
                    "fecha": adjusted_at.strftime("%Y-%m-%d") if isinstance(adjusted_at, datetime) else "",
                    "hora": adjusted_at.strftime("%H:%M:%S") if isinstance(adjusted_at, datetime) else "",
                    "adjustment_type": item.get("adjustment_type"),
                    "movement_type": "Salida",
                    "folio": item.get("folio"),
                    "part_number": item.get("part_number"),
                    "action": item.get("adjustment_action"),
                    "changed_fields": "departure_code",
                    "before": item.get("previous_departure_code") or "",
                    "after": item.get("departure_code") or "",
                    "notes": item.get("notes") or "",
                    "user": item.get("assigned_by") or "",
                }
            )
        modificaciones_rows.sort(
            key=lambda item: (item.get("fecha") or "", item.get("hora") or "", item.get("folio") or "")
        )

        wb = Workbook()
        header_fill = PatternFill(start_color="10243F", end_color="10243F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        title_font = Font(color="10243F", bold=True, size=14)

        def write_sheet(ws, title, headers, rows):
            ws.cell(row=1, column=1, value=title)
            ws.cell(row=1, column=1).font = title_font
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            if period_start and period_end_inclusive:
                period_text = (
                    f"Periodo: {normalize_excel_value(period_start)} "
                    f"a {normalize_excel_value(period_end_inclusive)}"
                )
            else:
                period_text = f"Periodo: inicio de operación a {normalize_excel_value(period_end)}"
            ws.cell(row=2, column=1, value=period_text)
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

            header_row = 4
            for col_idx, (label, _key) in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_idx, value=label)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if rows:
                for row_idx, row in enumerate(rows, header_row + 1):
                    for col_idx, (_label, key) in enumerate(headers, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=normalize_excel_value(row.get(key, "")))
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                ws.cell(row=header_row + 1, column=1, value="Sin registros para el periodo.")
                ws.merge_cells(start_row=header_row + 1, start_column=1, end_row=header_row + 1, end_column=len(headers))

            ws.freeze_panes = "A5"
            ws.auto_filter.ref = f"A4:{ws.cell(row=max(header_row + 1, len(rows) + header_row), column=len(headers)).coordinate}"

            for column_idx in range(1, ws.max_column + 1):
                column_letter = get_column_letter(column_idx)
                max_length = 0
                for cell in ws[column_letter]:
                    value_length = len(str(cell.value or ""))
                    max_length = max(max_length, value_length)
                ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 36)

        ws = wb.active
        ws.title = "Cierre"
        write_sheet(
            ws,
            f"Reporte de cierre - {period_info['closure_label']}",
            [
                ("No. parte", "part_number"),
                ("Inventario inicial", "initial_quantity"),
                ("Entradas", "entries"),
                ("Salidas", "exits"),
                ("Entradas de retorno", "return_entries"),
                ("Salidas de retorno", "return_losses"),
                ("Inventario físico", "physical_quantity"),
                ("Diferencia", "difference_quantity"),
            ],
            cierre_rows,
        )

        ws = wb.create_sheet("Entradas y Salidas")
        write_sheet(
            ws,
            "Historial de entradas y salidas",
            [
                ("Fecha", "fecha"),
                ("Hora", "hora"),
                ("Movimiento", "tipo"),
                ("Folio", "folio"),
                ("No. parte", "part_number"),
                ("Cantidad", "quantity"),
                ("Departure", "departure_code"),
                ("Modelo", "product_model"),
                ("Cliente", "customer"),
                ("Usuario", "registered_by"),
            ],
            entradas_salidas_rows,
        )

        ws = wb.create_sheet("Retornos")
        write_sheet(
            ws,
            "Historial de entradas y salidas de retorno",
            [
                ("Fecha", "fecha"),
                ("Hora", "hora"),
                ("Movimiento", "movement_type"),
                ("Folio", "folio"),
                ("No. parte", "part_number"),
                ("Cantidad", "quantity"),
                ("Modelo", "product_model"),
                ("Cliente", "customer"),
                ("Tipo", "return_type"),
                ("Usuario", "registered_by"),
            ],
            retorno_rows,
        )

        ws = wb.create_sheet("Modificaciones")
        write_sheet(
            ws,
            "Historial de modificaciones",
            [
                ("Fecha", "fecha"),
                ("Hora", "hora"),
                ("Tipo ajuste", "adjustment_type"),
                ("Movimiento", "movement_type"),
                ("Folio", "folio"),
                ("No. parte", "part_number"),
                ("Acción", "action"),
                ("Campos", "changed_fields"),
                ("Antes", "before"),
                ("Después", "after"),
                ("Motivo / nota", "notes"),
                ("Usuario", "user"),
            ],
            modificaciones_rows,
        )

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename_label = re.sub(
            r"[^A-Za-z0-9_-]+",
            "_",
            str(period_info["closure_label"] or f"cierre_{batch_id}"),
        ).strip("_")
        filename = f"reporte_{filename_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        print(
            f"Error exportando reporte de cierre inventario embarques: {e}\n{traceback.format_exc()}"
        )
        return jsonify({"success": False, "error": str(e)}), 500
