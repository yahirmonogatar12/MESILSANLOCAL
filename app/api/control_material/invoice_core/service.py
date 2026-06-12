"""Casos de uso transaccionales para invoices de material."""

import hashlib
import logging
import re
import unicodedata
from decimal import Decimal
from io import BytesIO

from flask import session
from werkzeug.utils import secure_filename

from app.api.control_material.costing_core.resolver import (
    fetch_lot_row,
    recalculate_lot_cost,
    upsert_inventory_cost,
)
from app.api.control_material.invoice_core.constants import (
    ERROR_INTERNO,
    ESTADOS_INVOICE,
    MONEDA_DEFAULT,
)
from app.api.control_material.invoice_core.matcher import (
    assign_packing_lines,
    invoice_has_differences,
    resolve_aliases,
)
from app.api.control_material.invoice_core.normalizers import (
    decimal_or_zero,
    normalizar_numero_parte,
    normalizar_pallet_no,
    raw_text,
    row_to_json,
)
from app.api.control_material.invoice_core.parser import parse_invoice_workbook
from app.api.control_material.invoice_core.repository import (
    candidate_lots_for_packing,
    fetch_invoice,
    packing_available_locked,
    recalculate_invoice_state,
    recalculate_packing_state,
)
from app.api.shared import (
    conexion_o_error,
    dict_cursor,
    obtener_fecha_hora_mexico,
    sanitizar_texto,
)

logger = logging.getLogger(__name__)

ORIGINAL_PART_HEADER_ALIASES = {
    "ALIAS": "original",
    "ALIASPARTNUM": "original",
    "ALIASPARTNO": "original",
    "PARTNUM": "original",
    "PARTNO": "original",
    "PARTNUMBER": "original",
    "NUMEROPARTE": "original",
    "NUMERODEPARTE": "original",
    "PARTEINVOICE": "original",
    "PARTEPROVEEDOR": "original",
    "품번": "original",
    "SISTEMA": "sistema",
    "PARTSYS": "sistema",
    "SYSTEMPART": "sistema",
    "SYSTEMPARTNUM": "sistema",
    "SYSTEMPARTNO": "sistema",
    "NUMEROPARTESISTEMA": "sistema",
    "NUMERODEPARTESISTEMA": "sistema",
    "PARTESISTEMA": "sistema",
    "전산품번": "sistema",
    "TIPO": "tipo",
    "TYPE": "tipo",
    "TIPOMATERIAL": "tipo",
    "TIPODEMATERIAL": "tipo",
    "CLASE": "tipo",
    "PROVEEDOR": "tipo",
    "VENDOR": "tipo",
    "SUPPLIER": "tipo",
}


def _usuario_actual():
    return session.get("usuario") or "SISTEMA"


def _db():
    conn, error_response = conexion_o_error()
    if error_response:
        return None, None, ({"success": False, "error": "Base de datos no disponible"}, 503)
    return conn, dict_cursor(conn), None


def _header_key(value):
    text = raw_text(value).upper().replace("\u00a0", " ")
    text = "".join(
        char for char in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(char)
    )
    return re.sub(r"[\s\.\-_/#:()]+", "", text)


def _alias_header_map(row):
    mapped = {}
    for idx, cell in enumerate(row):
        key = _header_key(cell.value)
        field = ORIGINAL_PART_HEADER_ALIASES.get(key)
        if field and field not in mapped:
            mapped[field] = idx
    return mapped


def _alias_cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    return raw_text(row[idx].value)


def _parse_alias_workbook(file_bytes, include_conflicts=False):
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    records = {}
    conflicts = {}
    conflicted_keys = set()
    skipped = 0
    for sheet in wb.worksheets:
        header = None
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 1, 50)):
            mapped = _alias_header_map(row)
            if "original" in mapped and "sistema" in mapped:
                header = {"row": row[0].row, "mapped": mapped}
                break
        if not header:
            continue
        mapped = header["mapped"]
        for row in sheet.iter_rows(min_row=header["row"] + 1):
            numero_parte_original = normalizar_numero_parte(_alias_cell(row, mapped.get("original")))
            sistema = normalizar_numero_parte(_alias_cell(row, mapped.get("sistema")))
            tipo = sanitizar_texto(_alias_cell(row, mapped.get("tipo")), 255)
            if not numero_parte_original and not sistema:
                continue
            if not numero_parte_original or not sistema:
                skipped += 1
                continue
            key = (numero_parte_original, tipo or "")
            if key in conflicted_keys:
                conflict = conflicts[key]
                if sistema not in conflict["sistemas"]:
                    conflict["sistemas"].append(sistema)
                conflict["filas"].append(row[0].row)
                continue
            existing = records.get(key)
            if existing and existing["numero_parte_sistema"] != sistema:
                conflicted_keys.add(key)
                conflicts[key] = {
                    "numero_parte_original": numero_parte_original,
                    "tipo": tipo or "",
                    "sistemas": [existing["numero_parte_sistema"], sistema],
                    "filas": [existing.get("_row_no"), row[0].row],
                    "detalle": "Mismo numero original + tipo apunta a varios sistemas",
                }
                records.pop(key, None)
                continue
            records[key] = {
                "numero_parte_original": numero_parte_original,
                "numero_parte_sistema": sistema,
                "tipo": tipo or "",
                "_row_no": row[0].row,
            }
    clean_records = [
        {k: v for k, v in record.items() if k != "_row_no"}
        for record in records.values()
    ]
    if include_conflicts:
        return clean_records, skipped, list(conflicts.values())
    return clean_records, skipped


def list_invoices(args):
    conn, cursor, error = _db()
    if error:
        return error
    try:
        q = sanitizar_texto(args.get("q"), 120)
        estado = sanitizar_texto(args.get("estado"), 40).upper()
        params = []
        where = ["1=1"]
        if q:
            where.append("(mi.numero_invoice LIKE %s OR COALESCE(mi.tipo, '') LIKE %s OR mi.archivo_nombre LIKE %s)")
            like = f"%{q}%"
            params.extend([like, like, like])
        if estado in ESTADOS_INVOICE:
            where.append("mi.estado = %s")
            params.append(estado)

        cursor.execute(
            f"""
            SELECT mi.*,
                   COALESCE(active_links.links_activos, 0) AS links_activos
            FROM material_invoices mi
            LEFT JOIN (
                SELECT invoice_id, COUNT(*) AS links_activos
                FROM material_invoice_lot_links
                WHERE estado = 'APLICADO'
                GROUP BY invoice_id
            ) active_links ON active_links.invoice_id = mi.id
            WHERE {' AND '.join(where)}
            ORDER BY mi.fecha_carga DESC, mi.id DESC
            LIMIT 300
            """,
            params,
        )
        records = [row_to_json(r) for r in (cursor.fetchall() or [])]
        return {"success": True, "records": records, "total": len(records)}, 200
    except Exception as exc:
        logger.exception("Error listando invoices: %s", exc)
        return {"success": False, "error": ERROR_INTERNO}, 500
    finally:
        cursor.close()
        conn.close()


def upload_invoice(files, form):
    uploaded = files.get("file") or files.get("archivo")
    if not uploaded:
        return {"success": False, "error": "Archivo requerido."}, 400

    filename = secure_filename(uploaded.filename or "invoice.xlsx")
    file_bytes = uploaded.read()
    if not file_bytes:
        return {"success": False, "error": "El archivo esta vacio."}, 400

    try:
        parsed = parse_invoice_workbook(file_bytes, filename)
    except Exception as exc:
        logger.exception("Error leyendo invoice desde %s: %s", filename, exc)
        return {"success": False, "error": "No se pudo leer el Excel de la invoice."}, 400
    numero_invoice = sanitizar_texto(
        form.get("numero_invoice") or parsed.get("numero_invoice_sugerido"),
        255,
    )
    tipo = sanitizar_texto(form.get("tipo") or form.get("proveedor"), 255)
    if not numero_invoice:
        return {"success": False, "error": "numero_invoice requerido."}, 400
    if not parsed["invoice_lines"]:
        detail = " ".join(parsed.get("warnings") or [])
        message = "No se detectaron lineas de invoice en la hoja INVOICE(total)."
        return {"success": False, "error": f"{message} {detail}".strip()}, 400

    file_hash = hashlib.sha256(file_bytes).hexdigest()
    conn, cursor, error = _db()
    if error:
        return error
    try:
        cursor.execute("SELECT id FROM material_invoices WHERE numero_invoice = %s LIMIT 1", (numero_invoice,))
        existing = cursor.fetchone()
        if existing:
            return (
                {
                    "success": False,
                    "duplicado": True,
                    "motivo": "NUMERO_INVOICE",
                    "invoice_id": existing["id"],
                    "message": "Esta invoice ya fue cargada.",
                },
                409,
            )

        cursor.execute("SELECT id FROM material_invoices WHERE archivo_hash_sha256 = %s LIMIT 1", (file_hash,))
        existing = cursor.fetchone()
        if existing:
            return (
                {
                    "success": False,
                    "duplicado": True,
                    "motivo": "ARCHIVO_HASH",
                    "invoice_id": existing["id"],
                    "message": "Este archivo parece ya cargado.",
                },
                409,
            )

        usuario = _usuario_actual()
        fecha = obtener_fecha_hora_mexico()
        cursor.execute("START TRANSACTION")
        resolve_aliases(cursor, parsed["invoice_lines"], tipo, "numero_parte_sistema")
        resolve_aliases(cursor, parsed["packing_lines"], tipo, "numero_parte_sistema")

        total_monto = sum((r.get("costo_total") or Decimal("0.0000")) for r in parsed["invoice_lines"])
        cursor.execute(
            """
            INSERT INTO material_invoices (
                numero_invoice, tipo, archivo_nombre, archivo_hash_sha256,
                estado, moneda, total_lineas, total_packing, total_monto,
                usuario_carga, fecha_carga
            ) VALUES (%s, %s, %s, %s, 'BORRADOR', %s, %s, %s, %s, %s, %s)
            """,
            (
                numero_invoice,
                tipo or None,
                filename,
                file_hash,
                MONEDA_DEFAULT,
                len(parsed["invoice_lines"]),
                len(parsed["packing_lines"]),
                str(total_monto),
                usuario,
                fecha,
            ),
        )
        invoice_id = cursor.lastrowid

        _insert_invoice_lines(cursor, invoice_id, parsed["invoice_lines"])
        _insert_packing_lines(cursor, invoice_id, parsed["packing_lines"])

        assign_packing_lines(cursor, invoice_id)
        estado = "CON_DIFERENCIAS" if invoice_has_differences(cursor, invoice_id) else "VALIDADA"
        cursor.execute(
            """
            UPDATE material_invoices
            SET estado = %s, usuario_validacion = %s, fecha_validacion = %s
            WHERE id = %s
            """,
            (estado, usuario, fecha, invoice_id),
        )
        conn.commit()
        return (
            {
                "success": True,
                "invoice_id": invoice_id,
                "estado": estado,
                "lineas": len(parsed["invoice_lines"]),
                "packing": len(parsed["packing_lines"]),
            },
            201,
        )
    except Exception as exc:
        conn.rollback()
        logger.exception("Error cargando invoice: %s", exc)
        return {"success": False, "error": ERROR_INTERNO}, 500
    finally:
        cursor.close()
        conn.close()


def _insert_invoice_lines(cursor, invoice_id, invoice_lines):
    for line in invoice_lines:
        cursor.execute(
            """
            INSERT INTO material_invoice_lines (
                invoice_id, line_no, maker, origin, raw_part_num,
                numero_parte_invoice, numero_parte_sistema, descripcion,
                cantidad, uom, costo_unitario, costo_total, moneda,
                raw_qty, raw_unit_cost, raw_total_cost, estado_match, mensaje_match
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                invoice_id,
                line["line_no"],
                line.get("maker"),
                line.get("origin"),
                line.get("raw_part_num"),
                line["numero_parte_invoice"],
                line["numero_parte_sistema"],
                line.get("descripcion"),
                str(line["cantidad"]),
                line.get("uom"),
                str(line["costo_unitario"]),
                str(line["costo_total"]),
                line.get("moneda") or MONEDA_DEFAULT,
                line.get("raw_qty"),
                line.get("raw_unit_cost"),
                line.get("raw_total_cost"),
                line.get("estado_match") or "PENDIENTE",
                line.get("mensaje_match"),
            ),
        )


def _insert_packing_lines(cursor, invoice_id, packing_lines):
    for packing in packing_lines:
        cursor.execute(
            """
            INSERT INTO material_invoice_packing_lines (
                invoice_id, line_no, packing_no, raw_part_num,
                numero_parte_packing, numero_parte_sistema, descripcion,
                cantidad_packing, raw_qty, pallet_no_original, pallet_no,
                kg, cbm, estado_match, mensaje_match
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                invoice_id,
                packing["line_no"],
                packing.get("packing_no"),
                packing.get("raw_part_num"),
                packing["numero_parte_packing"],
                packing["numero_parte_sistema"],
                packing.get("descripcion"),
                str(packing["cantidad_packing"]),
                packing.get("raw_qty"),
                packing.get("pallet_no_original"),
                packing.get("pallet_no"),
                str(packing["kg"]) if packing.get("kg") is not None else None,
                str(packing["cbm"]) if packing.get("cbm") is not None else None,
                packing.get("estado_match") or "PENDIENTE",
                packing.get("mensaje_match"),
            ),
        )


def get_invoice_detail(invoice_id):
    conn, cursor, error = _db()
    if error:
        return error
    try:
        invoice = fetch_invoice(cursor, invoice_id)
        if not invoice:
            return {"success": False, "error": "Invoice no encontrada."}, 404
        cursor.execute("SELECT * FROM material_invoice_lines WHERE invoice_id = %s ORDER BY line_no, id", (invoice_id,))
        lines = [row_to_json(r) for r in (cursor.fetchall() or [])]
        cursor.execute(
            """
            SELECT
              mipl.*,
              COALESCE(entradas.entradas_recibidas, 0) AS entradas_recibidas,
              COALESCE(entradas.cantidad_recibida, 0) AS cantidad_recibida,
              GREATEST(mipl.cantidad_packing - COALESCE(entradas.cantidad_recibida, 0), 0) AS cantidad_pendiente_entrada,
              COALESCE(aplicado.links_activos, 0) AS links_activos,
              COALESCE(aplicado.cantidad_aplicada_activa, 0) AS cantidad_aplicada_activa
            FROM material_invoice_packing_lines mipl
            LEFT JOIN (
                SELECT
                  cma.numero_invoice,
                  cma.pallet_no,
                  COALESCE(il.numero_parte, cma.numero_parte) AS numero_parte_sistema,
                  COUNT(DISTINCT cma.codigo_material_recibido) AS entradas_recibidas,
                  SUM(COALESCE(il.total_entrada, cma.cantidad_actual, 0)) AS cantidad_recibida
                FROM control_material_almacen cma
                LEFT JOIN inventario_lotes il
                  ON il.codigo_material_recibido = cma.codigo_material_recibido
                WHERE cma.numero_invoice = %s
                  AND (cma.cancelado = 0 OR cma.cancelado IS NULL)
                GROUP BY cma.numero_invoice, cma.pallet_no, COALESCE(il.numero_parte, cma.numero_parte)
            ) entradas
              ON entradas.numero_invoice = %s
             AND entradas.pallet_no = mipl.pallet_no
             AND entradas.numero_parte_sistema = mipl.numero_parte_sistema
            LEFT JOIN (
                SELECT
                  packing_line_id,
                  COUNT(*) AS links_activos,
                  SUM(cantidad_aplicada) AS cantidad_aplicada_activa
                FROM material_invoice_lot_links
                WHERE invoice_id = %s
                  AND estado = 'APLICADO'
                GROUP BY packing_line_id
            ) aplicado
              ON aplicado.packing_line_id = mipl.id
            WHERE mipl.invoice_id = %s
            ORDER BY
              CASE
                WHEN mipl.pallet_no REGEXP '^[0-9]+$' THEN CAST(mipl.pallet_no AS UNSIGNED)
                ELSE 999999
              END,
              mipl.pallet_no,
              mipl.line_no,
              mipl.id
            """,
            (invoice["numero_invoice"], invoice["numero_invoice"], invoice_id, invoice_id),
        )
        packing = [row_to_json(r) for r in (cursor.fetchall() or [])]
        cursor.execute(
            """
            SELECT *
            FROM material_invoice_lot_links
            WHERE invoice_id = %s
            ORDER BY fecha_aplicacion DESC, id DESC
            LIMIT 500
            """,
            (invoice_id,),
        )
        links = [row_to_json(r) for r in (cursor.fetchall() or [])]
        return {"success": True, "invoice": row_to_json(invoice), "lines": lines, "packing": packing, "links": links}, 200
    except Exception as exc:
        logger.exception("Error detalle invoice %s: %s", invoice_id, exc)
        return {"success": False, "error": ERROR_INTERNO}, 500
    finally:
        cursor.close()
        conn.close()


def get_invoice_candidates(invoice_id, args):
    packing_line_id = args.get("packing_line_id")
    conn, cursor, error = _db()
    if error:
        return error
    try:
        invoice = fetch_invoice(cursor, invoice_id)
        if not invoice:
            return {"success": False, "error": "Invoice no encontrada."}, 404

        params = [invoice["numero_invoice"]]
        where = ["cma.numero_invoice = %s", "(cma.cancelado = 0 OR cma.cancelado IS NULL)"]
        if packing_line_id:
            cursor.execute(
                "SELECT * FROM material_invoice_packing_lines WHERE id = %s AND invoice_id = %s",
                (packing_line_id, invoice_id),
            )
            packing = cursor.fetchone()
            if not packing:
                return {"success": False, "error": "Packing line no encontrada."}, 404
            where.append("cma.pallet_no = %s")
            where.append("cma.numero_parte = %s")
            params.extend([packing.get("pallet_no"), packing.get("numero_parte_sistema")])

        cursor.execute(
            f"""
            SELECT
              cma.codigo_material_recibido,
              cma.numero_parte,
              cma.numero_lote_material,
              cma.numero_invoice,
              cma.pallet_no_original,
              cma.pallet_no,
              cma.vendedor,
              cma.cantidad_actual,
              il.stock_actual,
              active.id AS active_link_id
            FROM control_material_almacen cma
            LEFT JOIN inventario_lotes il ON il.codigo_material_recibido = cma.codigo_material_recibido
            LEFT JOIN material_invoice_lot_links active
              ON active.codigo_material_recibido = cma.codigo_material_recibido
             AND active.estado = 'APLICADO'
            WHERE {' AND '.join(where)}
            ORDER BY cma.fecha_recibo ASC, cma.id ASC
            LIMIT 500
            """,
            params,
        )
        return {"success": True, "records": [row_to_json(r) for r in (cursor.fetchall() or [])]}, 200
    except Exception as exc:
        logger.exception("Error candidatos invoice %s: %s", invoice_id, exc)
        return {"success": False, "error": ERROR_INTERNO}, 500
    finally:
        cursor.close()
        conn.close()


def apply_invoice(invoice_id, data):
    conn, cursor, error = _db()
    if error:
        return error
    try:
        usuario = _usuario_actual()
        fecha = obtener_fecha_hora_mexico()
        cursor.execute("START TRANSACTION")
        invoice = fetch_invoice(cursor, invoice_id, for_update=True)
        if not invoice:
            conn.rollback()
            return {"success": False, "error": "Invoice no encontrada."}, 404
        if invoice.get("estado") == "CANCELADA":
            conn.rollback()
            return {"success": False, "error": "No se puede aplicar una invoice cancelada."}, 400

        applied, skipped = _apply_items_or_auto(cursor, invoice, data, usuario, fecha)
        estado = recalculate_invoice_state(cursor, invoice_id)
        conn.commit()
        return {
            "success": True,
            "estado": estado,
            "aplicados": len(applied),
            "omitidos": len(skipped),
            "applied": applied,
            "skipped": skipped[:50],
        }, 200
    except Exception as exc:
        conn.rollback()
        logger.exception("Error aplicando invoice %s: %s", invoice_id, exc)
        return {"success": False, "error": ERROR_INTERNO}, 500
    finally:
        cursor.close()
        conn.close()


def _apply_items_or_auto(cursor, invoice, data, usuario, fecha):
    items = data.get("items") if isinstance(data.get("items"), list) else None
    if not items:
        return _auto_apply_invoice(cursor, invoice, usuario, fecha)

    applied = []
    skipped = []
    for item in items:
        result = _apply_one_locked(cursor, invoice, item, usuario, fecha)
        (applied if result.get("success") else skipped).append(result)
    return applied, skipped


def _apply_one_locked(cursor, invoice, item, usuario, fecha):
    codigo = sanitizar_texto(item.get("codigo_material_recibido"), 255)
    packing_line_id = item.get("packing_line_id")
    if not codigo or not packing_line_id:
        return {"success": False, "codigo_material_recibido": codigo, "error": "codigo_material_recibido y packing_line_id son requeridos."}

    packing, _total_packing, available = packing_available_locked(cursor, packing_line_id)
    if not packing or int(packing["invoice_id"]) != int(invoice["id"]):
        return {"success": False, "codigo_material_recibido": codigo, "error": "Packing line no pertenece a la invoice."}
    if not packing.get("invoice_line_id") and not item.get("invoice_line_id"):
        return {"success": False, "codigo_material_recibido": codigo, "error": "Packing line no esta conciliada con invoice line."}

    invoice_line_id = item.get("invoice_line_id") or packing["invoice_line_id"]
    cursor.execute(
        """
        SELECT *
        FROM material_invoice_lines
        WHERE id = %s AND invoice_id = %s
        LIMIT 1
        FOR UPDATE
        """,
        (invoice_line_id, invoice["id"]),
    )
    line = cursor.fetchone()
    if not line:
        return {"success": False, "codigo_material_recibido": codigo, "error": "Invoice line no encontrada."}

    cursor.execute(
        """
        SELECT id, invoice_id
        FROM material_invoice_lot_links
        WHERE codigo_material_recibido = %s AND estado = 'APLICADO'
        FOR UPDATE
        """,
        (codigo,),
    )
    if cursor.fetchone():
        return {
            "success": False,
            "codigo_material_recibido": codigo,
            "error": "El lote ya tiene una invoice aplicada. Desaplica primero o usa reaplicacion explicita.",
        }

    lot = fetch_lot_row(cursor, codigo, for_update=True)
    if not lot:
        return {"success": False, "codigo_material_recibido": codigo, "error": "Lote no encontrado en control_material_almacen."}

    lot_pallet = normalizar_pallet_no(lot.get("pallet_no") or lot.get("pallet_no_original"))
    if lot.get("numero_invoice") != invoice.get("numero_invoice"):
        return {"success": False, "codigo_material_recibido": codigo, "error": "El numero_invoice del lote no coincide."}
    if lot_pallet != (packing.get("pallet_no") or ""):
        return {"success": False, "codigo_material_recibido": codigo, "error": "El pallet del lote no coincide."}
    if normalizar_numero_parte(lot.get("numero_parte_sistema")) != normalizar_numero_parte(packing.get("numero_parte_sistema")):
        return {"success": False, "codigo_material_recibido": codigo, "error": "La parte del lote no coincide."}

    cantidad = decimal_or_zero(item.get("cantidad_aplicada"))
    if cantidad <= 0:
        cantidad = decimal_or_zero(lot.get("cantidad_lote"))
    if cantidad <= 0:
        return {"success": False, "codigo_material_recibido": codigo, "error": "Cantidad aplicada invalida."}
    if cantidad > available:
        return {
            "success": False,
            "codigo_material_recibido": codigo,
            "error": "La cantidad del lote excede la cantidad packing disponible.",
            "cantidad_disponible": str(available),
        }

    cursor.execute(
        """
        INSERT INTO material_invoice_lot_links (
            invoice_id, invoice_line_id, packing_line_id, codigo_material_recibido,
            numero_parte_sistema, cantidad_aplicada, costo_unitario, moneda,
            usuario_aplicacion, fecha_aplicacion, usuario_registro, fecha_registro, estado
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'APLICADO')
        """,
        (
            invoice["id"],
            invoice_line_id,
            packing_line_id,
            codigo,
            line["numero_parte_sistema"],
            str(cantidad),
            str(line["costo_unitario"]),
            line.get("moneda") or MONEDA_DEFAULT,
            usuario,
            fecha,
            usuario,
            fecha,
        ),
    )
    link_id = cursor.lastrowid
    upsert_inventory_cost(
        cursor,
        lot,
        {
            "costo_unitario": Decimal(str(line["costo_unitario"])),
            "moneda": line.get("moneda") or MONEDA_DEFAULT,
            "fuente_costo": "INVOICE",
            "es_estimado": 0,
        },
        usuario,
        {
            "invoice_id": invoice["id"],
            "invoice_line_id": invoice_line_id,
            "packing_line_id": packing_line_id,
            "link_id": link_id,
        },
    )
    recalculate_packing_state(cursor, packing_line_id)
    return {"success": True, "codigo_material_recibido": codigo, "link_id": link_id, "cantidad_aplicada": str(cantidad)}


def _auto_apply_invoice(cursor, invoice, usuario, fecha):
    cursor.execute(
        """
        SELECT *
        FROM material_invoice_packing_lines
        WHERE invoice_id = %s
          AND invoice_line_id IS NOT NULL
          AND estado_match NOT IN ('SIN_ALIAS','SIN_LINEA','DIFERENCIA')
        ORDER BY pallet_no, line_no, id
        """,
        (invoice["id"],),
    )
    applied = []
    skipped = []
    for packing in cursor.fetchall() or []:
        for codigo in candidate_lots_for_packing(cursor, invoice, packing):
            result = _apply_one_locked(
                cursor,
                invoice,
                {
                    "codigo_material_recibido": codigo,
                    "packing_line_id": packing["id"],
                    "invoice_line_id": packing["invoice_line_id"],
                },
                usuario,
                fecha,
            )
            (applied if result.get("success") else skipped).append(result)
    return applied, skipped


def unapply_invoice(invoice_id, data):
    conn, cursor, error = _db()
    if error:
        return error
    try:
        usuario = _usuario_actual()
        fecha = obtener_fecha_hora_mexico()
        cursor.execute("START TRANSACTION")
        invoice = fetch_invoice(cursor, invoice_id, for_update=True)
        if not invoice:
            conn.rollback()
            return {"success": False, "error": "Invoice no encontrada."}, 404
        if invoice.get("estado") == "CANCELADA":
            conn.rollback()
            return {"success": False, "error": "No se puede desaplicar una invoice cancelada."}, 400

        result = _unapply_links_locked(cursor, invoice_id, data, usuario, fecha)
        estado = recalculate_invoice_state(cursor, invoice_id)
        conn.commit()
        result.update({"success": True, "estado": estado})
        return result, 200
    except Exception as exc:
        conn.rollback()
        logger.exception("Error desaplicando invoice %s: %s", invoice_id, exc)
        return {"success": False, "error": ERROR_INTERNO}, 500
    finally:
        cursor.close()
        conn.close()


def _unapply_links_locked(cursor, invoice_id, data, usuario, fecha):
    motivo = sanitizar_texto(data.get("motivo_desaplicado") or data.get("motivo"), 255)
    link_ids = data.get("link_ids") if isinstance(data.get("link_ids"), list) else []
    codigo = sanitizar_texto(data.get("codigo_material_recibido"), 255)

    params = [invoice_id]
    where = ["invoice_id = %s", "estado = 'APLICADO'"]
    if link_ids:
        placeholders = ", ".join(["%s"] * len(link_ids))
        where.append(f"id IN ({placeholders})")
        params.extend(link_ids)
    if codigo:
        where.append("codigo_material_recibido = %s")
        params.append(codigo)

    cursor.execute(
        f"""
        SELECT *
        FROM material_invoice_lot_links
        WHERE {' AND '.join(where)}
        ORDER BY fecha_aplicacion ASC, id ASC
        FOR UPDATE
        """,
        params,
    )
    links = cursor.fetchall() or []
    affected_codes = []
    packing_ids = set()
    for link in links:
        cursor.execute(
            """
            UPDATE material_invoice_lot_links
            SET estado = 'DESAPLICADO',
                fecha_desaplicado = %s,
                usuario_desaplicado = %s,
                motivo_desaplicado = %s
            WHERE id = %s AND estado = 'APLICADO'
            """,
            (fecha, usuario, motivo or None, link["id"]),
        )
        affected_codes.append((link["codigo_material_recibido"], link["fecha_aplicacion"], link["id"]))
        if link.get("packing_line_id"):
            packing_ids.add(link["packing_line_id"])

    recalculated = 0
    skipped_later = 0
    for codigo_lote, old_fecha, old_id in affected_codes:
        cursor.execute(
            """
            SELECT id
            FROM material_invoice_lot_links
            WHERE codigo_material_recibido = %s
              AND estado = 'APLICADO'
              AND (
                fecha_aplicacion > %s
                OR (fecha_aplicacion = %s AND id > %s)
              )
            ORDER BY fecha_aplicacion DESC, id DESC
            LIMIT 1
            """,
            (codigo_lote, old_fecha, old_fecha, old_id),
        )
        if cursor.fetchone():
            skipped_later += 1
            continue
        if recalculate_lot_cost(cursor, codigo_lote, usuario):
            recalculated += 1

    for packing_id in packing_ids:
        recalculate_packing_state(cursor, packing_id)

    return {
        "links_desaplicados": len(links),
        "lotes_recalculados": recalculated,
        "lotes_omitidos_por_reaplicacion_posterior": skipped_later,
    }


def reapply_invoice(invoice_id, data):
    conn, cursor, error = _db()
    if error:
        return error
    try:
        usuario = _usuario_actual()
        fecha = obtener_fecha_hora_mexico()
        cursor.execute("START TRANSACTION")
        invoice = fetch_invoice(cursor, invoice_id, for_update=True)
        if not invoice:
            conn.rollback()
            return {"success": False, "error": "Invoice no encontrada."}, 404
        if invoice.get("estado") == "CANCELADA":
            conn.rollback()
            return {"success": False, "error": "No se puede reaplicar una invoice cancelada."}, 400

        unapply_result = _unapply_links_locked(
            cursor,
            invoice_id,
            {"motivo": data.get("motivo") or "Reaplicacion explicita"},
            usuario,
            fecha,
        )
        applied, skipped = _apply_items_or_auto(cursor, invoice, data, usuario, fecha)
        estado = recalculate_invoice_state(cursor, invoice_id)
        conn.commit()
        return {
            "success": True,
            "estado": estado,
            "unapply": unapply_result,
            "aplicados": len(applied),
            "omitidos": len(skipped),
            "applied": applied,
            "skipped": skipped[:50],
        }, 200
    except Exception as exc:
        conn.rollback()
        logger.exception("Error reaplicando invoice %s: %s", invoice_id, exc)
        return {"success": False, "error": ERROR_INTERNO}, 500
    finally:
        cursor.close()
        conn.close()


def list_aliases(args):
    conn, cursor, error = _db()
    if error:
        return error
    try:
        q = sanitizar_texto(args.get("q"), 120)
        include_inactive = str(args.get("include_inactive") or "").lower() in ("1", "true", "yes", "si", "on")
        params = []
        where = ["1=1"]
        if not include_inactive:
            where.append("a.activo = 1")
        if q:
            where.append("(a.numero_parte_original LIKE %s OR a.numero_parte_sistema LIKE %s OR COALESCE(a.tipo, '') LIKE %s)")
            like = f"%{q}%"
            params.extend([like, like, like])

        cursor.execute(
            f"""
            SELECT
              a.*,
              CASE WHEN m.numero_parte IS NULL THEN 0 ELSE 1 END AS sistema_existe
            FROM material_part_aliases a
            LEFT JOIN materiales m
              ON m.numero_parte = a.numero_parte_sistema
            WHERE {' AND '.join(where)}
            ORDER BY a.activo DESC, a.numero_parte_original ASC, a.tipo ASC, a.id DESC
            LIMIT 500
            """,
            params,
        )
        records = [row_to_json(row) for row in (cursor.fetchall() or [])]
        return {"success": True, "records": records, "total": len(records)}, 200
    except Exception as exc:
        logger.exception("Error listando aliases: %s", exc)
        return {"success": False, "error": ERROR_INTERNO}, 500
    finally:
        cursor.close()
        conn.close()


def _upsert_alias(cursor, numero_parte_original, sistema, tipo, usuario, fecha):
    cursor.execute(
        """
        INSERT INTO material_part_aliases (
            numero_parte_original, numero_parte_sistema, tipo,
            usuario_registro, fecha_registro, activo
        ) VALUES (%s, %s, %s, %s, %s, 1)
        ON DUPLICATE KEY UPDATE
            numero_parte_sistema = VALUES(numero_parte_sistema),
            usuario_registro = VALUES(usuario_registro),
            fecha_registro = VALUES(fecha_registro),
            activo = 1
        """,
        (numero_parte_original, sistema, tipo or "", usuario, fecha),
    )


def _materiales_existentes(cursor, numeros_parte):
    numeros = sorted({normalizar_numero_parte(numero) for numero in numeros_parte if numero})
    existentes = set()
    for idx in range(0, len(numeros), 500):
        bloque = numeros[idx:idx + 500]
        placeholders = ", ".join(["%s"] * len(bloque))
        cursor.execute(
            f"SELECT numero_parte FROM materiales WHERE numero_parte IN ({placeholders})",
            tuple(bloque),
        )
        existentes.update(row["numero_parte"] for row in (cursor.fetchall() or []))
    return existentes


def create_alias(data):
    numero_parte_original = normalizar_numero_parte(
        data.get("numero_parte_original", data.get("alias_part_num"))
    )
    sistema = normalizar_numero_parte(data.get("numero_parte_sistema"))
    tipo = sanitizar_texto(data.get("tipo") or data.get("proveedor"), 255)
    if not numero_parte_original or not sistema:
        return {"success": False, "error": "numero_parte_original y numero_parte_sistema son requeridos."}, 400

    conn, cursor, error = _db()
    if error:
        return error
    try:
        if sistema not in _materiales_existentes(cursor, [sistema]):
            return {
                "success": False,
                "error": f"El numero de parte sistema {sistema} no existe en materiales.",
                "motivo": "PARTE_SISTEMA_NO_EXISTE",
                "numero_parte_sistema": sistema,
            }, 400
        _upsert_alias(cursor, numero_parte_original, sistema, tipo, _usuario_actual(), obtener_fecha_hora_mexico())
        conn.commit()
        return {
            "success": True,
            "numero_parte_original": numero_parte_original,
            "numero_parte_sistema": sistema,
            "tipo": tipo,
        }, 200
    except Exception as exc:
        conn.rollback()
        logger.exception("Error creando alias: %s", exc)
        return {"success": False, "error": ERROR_INTERNO}, 500
    finally:
        cursor.close()
        conn.close()


def import_aliases(files, form):
    uploaded = files.get("file") or files.get("archivo")
    if not uploaded:
        return {"success": False, "error": "Archivo requerido."}, 400
    filename = secure_filename(uploaded.filename or "aliases.xlsx")
    file_bytes = uploaded.read()
    if not file_bytes:
        return {"success": False, "error": "El archivo esta vacio."}, 400

    try:
        records, skipped, conflicts = _parse_alias_workbook(file_bytes, include_conflicts=True)
    except Exception as exc:
        logger.exception("Error leyendo aliases desde %s: %s", filename, exc)
        return {"success": False, "error": "No se pudo leer el Excel de equivalencias."}, 400

    if not records:
        return {"success": False, "error": "No se detectaron columnas de numero de parte original y parte sistema."}, 400

    conn, cursor, error = _db()
    if error:
        return error
    try:
        usuario = _usuario_actual()
        fecha = obtener_fecha_hora_mexico()
        tipo_default = sanitizar_texto(form.get("tipo") or form.get("proveedor"), 255)
        cursor.execute("START TRANSACTION")
        materiales_existentes = _materiales_existentes(
            cursor,
            (record["numero_parte_sistema"] for record in records),
        )
        imported = 0
        omitidos_sistema = []
        for record in records:
            if record["numero_parte_sistema"] not in materiales_existentes:
                omitidos_sistema.append({
                    "numero_parte_original": record["numero_parte_original"],
                    "numero_parte_sistema": record["numero_parte_sistema"],
                    "tipo": record.get("tipo") or tipo_default or "",
                    "motivo": "PARTE_SISTEMA_NO_EXISTE",
                })
                continue
            tipo = record.get("tipo") or tipo_default or ""
            _upsert_alias(
                cursor,
                record["numero_parte_original"],
                record["numero_parte_sistema"],
                tipo,
                usuario,
                fecha,
            )
            imported += 1
        conn.commit()
        total_omitidos = skipped + len(omitidos_sistema) + len(conflicts)
        return {
            "success": True,
            "archivo": filename,
            "importados": imported,
            "omitidos": total_omitidos,
            "omitidos_excel": skipped,
            "omitidos_sistema": len(omitidos_sistema),
            "omitidos_conflicto": len(conflicts),
            "conflictos_preview": conflicts[:50],
            "omitidos_sistema_preview": omitidos_sistema[:50],
            "total_detectados": len(records) + len(conflicts),
        }, 200
    except Exception as exc:
        conn.rollback()
        logger.exception("Error importando aliases: %s", exc)
        return {"success": False, "error": ERROR_INTERNO}, 500
    finally:
        cursor.close()
        conn.close()


def deactivate_alias(alias_id):
    conn, cursor, error = _db()
    if error:
        return error
    try:
        cursor.execute(
            """
            UPDATE material_part_aliases
            SET activo = 0
            WHERE id = %s
            """,
            (alias_id,),
        )
        conn.commit()
        return {"success": True}, 200
    except Exception as exc:
        conn.rollback()
        logger.exception("Error desactivando alias %s: %s", alias_id, exc)
        return {"success": False, "error": ERROR_INTERNO}, 500
    finally:
        cursor.close()
        conn.close()
