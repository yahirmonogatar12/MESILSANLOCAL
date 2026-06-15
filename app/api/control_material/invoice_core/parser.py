"""Parser Excel para invoices y packing lists.

No toca BD: mantiene testeable la lectura de archivos y los raw values.
"""

import re
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from app.api.shared import sanitizar_texto
from app.api.control_material.invoice_core.constants import MONEDA_DEFAULT
from app.api.control_material.invoice_core.normalizers import (
    decimal_or_zero,
    normalizar_numero_parte,
    normalizar_pallet_no,
    raw_text,
    version_parte,
)

INVOICE_CONVERTED_SHEET_NAME = "INVOICE(CONVERTED)"
INVOICE_TOTAL_SHEET_NAME = "INVOICE(total)"
PACKING_ORIGINAL_SHEET_NAME = "PACKING LIST 원본"
PART_EQUIVALENCE_SHEETS = ("Hoja1", "Sheet1")
SUMMARY_PART_LABELS = {"COMMERCIAL", "NONCOMMERCIAL", "TOTAL", "SUBTOTAL"}

# Mapeo fijo de encabezados de la hoja INVOICE(CONVERTED) (formato interno
# ya conciliado: columna A = numero de invoice, fila 2 = encabezados).
CONVERTED_HEADER_ALIASES = {
    "TARIMA": "pallet",
    "PALLET": "pallet",
    "PARTNO": "raw_part_num",
    "PARTNUM": "raw_part_num",
    "PARTNUMBER": "raw_part_num",
    "PARTSYS": "part_sys",
    "PARTSYSTEM": "part_sys",
    "ITEM": "item",
    "SPEC": "spec",
    "DESCRIPTION": "spec",
    "QTY": "qty",
    "QUANTITY": "qty",
    "UOM": "uom",
    "UNIT": "uom",
    "UNIDAD": "uom",
    "COSTOS": "unit_cost",
    "COSTO": "unit_cost",
    "UNITCOST": "unit_cost",
    "TOTAL": "total_cost",
    "TOTALCOST": "total_cost",
}


def _sheet_key(name):
    return re.sub(r"\s+", "", raw_text(name)).upper()


def _is_invoice_converted_sheet(sheet):
    return _sheet_key(sheet.title) == _sheet_key(INVOICE_CONVERTED_SHEET_NAME)


def _is_invoice_total_sheet(sheet):
    return _sheet_key(sheet.title) == _sheet_key(INVOICE_TOTAL_SHEET_NAME)


def _is_packing_original_sheet(sheet):
    return _sheet_key(sheet.title) == _sheet_key(PACKING_ORIGINAL_SHEET_NAME)


def _is_summary_part(value):
    return raw_text(value).strip().upper() in SUMMARY_PART_LABELS


def _header_key(value):
    text = raw_text(value).upper()
    text = text.replace("\u00a0", " ")
    return re.sub(r"[\s\.\-_/#:]+", "", text)


INVOICE_HEADER_ALIASES = {
    "MAKER": "maker",
    "ORIGIN": "origin",
    "PARTNUM": "part_num",
    "PARTNO": "part_num",
    "PARTNUMBER": "part_num",
    "DESCRIPTION": "description",
    "DESC": "description",
    "QTY": "qty",
    "QUANTITY": "qty",
    "UOM": "uom",
    "UNITCOST": "unit_cost",
    "UNITPRICE": "unit_cost",
    "TTCOST": "total_cost",
    "TOTALCOST": "total_cost",
    "AMOUNT": "total_cost",
}

PACKING_HEADER_ALIASES = {
    "NO": "no",
    "품번": "part_num",
    "PARTNUM": "part_num",
    "PARTNO": "part_num",
    "품명": "description",
    "DESCRIPTION": "description",
    "수량": "qty",
    "QTY": "qty",
    "QUANTITY": "qty",
    "PALLET": "pallet",
    "PALLETNO": "pallet",
    "KG": "kg",
    "CBM": "cbm",
}


def _map_header(row, aliases):
    mapped = {}
    for idx, cell in enumerate(row):
        key = _header_key(cell.value)
        if key in aliases and aliases[key] not in mapped:
            mapped[aliases[key]] = idx
    return mapped


def _is_invoice_header(mapped):
    return "part_num" in mapped and "qty" in mapped and (
        "unit_cost" in mapped or "total_cost" in mapped or "uom" in mapped
    )


def _is_packing_header(mapped):
    return "part_num" in mapped and "qty" in mapped and "pallet" in mapped


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx].value


def _looks_like_header(row):
    inv = _map_header(row, INVOICE_HEADER_ALIASES)
    pack = _map_header(row, PACKING_HEADER_ALIASES)
    return _is_invoice_header(inv) or _is_packing_header(pack)


def _header_index_map(row):
    indexes = {}
    for idx, cell in enumerate(row):
        key = _header_key(cell.value)
        if key and key not in indexes:
            indexes[key] = idx
    return indexes


def _first_existing(indexes, *keys):
    for key in keys:
        if key in indexes:
            return indexes[key]
    return None


def _build_part_system_aliases(wb):
    aliases = {}
    target_names = {_sheet_key(name) for name in PART_EQUIVALENCE_SHEETS}
    for sheet in wb.worksheets:
        if _sheet_key(sheet.title) not in target_names:
            continue
        header_map = None
        raw_idx = None
        system_idx = None
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 1, 20)):
            indexes = _header_index_map(row)
            raw_idx = _first_existing(indexes, "PARTNO", "PARTNUM", "품번")
            system_idx = _first_existing(indexes, "PARTSYS", "전산품번")
            if raw_idx is not None and system_idx is not None:
                header_map = row
                break
        if header_map is None:
            continue
        start_row = header_map[0].row + 1
        for row in sheet.iter_rows(min_row=start_row):
            raw_part = normalizar_numero_parte(_cell(row, raw_idx))
            system_part = normalizar_numero_parte(_cell(row, system_idx))
            if raw_part and system_part:
                aliases[raw_part] = system_part
    return aliases


def _parse_rows_after_header(sheet, header_row_number, mapped, kind, start_no, part_aliases=None):
    part_aliases = part_aliases or {}
    records = []
    blank_streak = 0
    line_no = start_no
    for row in sheet.iter_rows(min_row=header_row_number + 1):
        if _looks_like_header(row):
            break

        part_raw = raw_text(_cell(row, mapped.get("part_num")))
        qty_raw = raw_text(_cell(row, mapped.get("qty")))
        if not part_raw and not qty_raw:
            blank_streak += 1
            if blank_streak >= 8:
                break
            continue
        blank_streak = 0
        if not part_raw:
            continue
        if kind == "invoice" and _is_summary_part(part_raw):
            continue

        line_no += 1
        qty = decimal_or_zero(qty_raw)
        if kind == "invoice":
            if qty <= 0:
                continue
            unit_raw = raw_text(_cell(row, mapped.get("unit_cost")))
            total_raw = raw_text(_cell(row, mapped.get("total_cost")))
            unit = decimal_or_zero(unit_raw)
            total = decimal_or_zero(total_raw)
            if unit == 0 and total == 0:
                continue
            if total == 0 and unit != 0 and qty != 0:
                total = (unit * qty).quantize(Decimal("0.0001"))
            invoice_part = normalizar_numero_parte(part_raw)
            system_part = part_aliases.get(invoice_part, invoice_part)
            records.append(
                {
                    "line_no": line_no,
                    "maker": raw_text(_cell(row, mapped.get("maker"))),
                    "origin": raw_text(_cell(row, mapped.get("origin"))),
                    "raw_part_num": part_raw,
                    "numero_parte_invoice": invoice_part,
                    "numero_parte_sistema": system_part,
                    "descripcion": raw_text(_cell(row, mapped.get("description"))),
                    "cantidad": qty,
                    "uom": raw_text(_cell(row, mapped.get("uom"))),
                    "costo_unitario": unit,
                    "costo_total": total,
                    "moneda": MONEDA_DEFAULT,
                    "raw_qty": qty_raw,
                    "raw_unit_cost": unit_raw,
                    "raw_total_cost": total_raw,
                    "estado_match": "PENDIENTE",
                    "mensaje_match": "",
                }
            )
        else:
            if _is_summary_part(part_raw) or qty <= 0:
                continue
            pallet_raw = raw_text(_cell(row, mapped.get("pallet")))
            packing_part = normalizar_numero_parte(part_raw)
            system_part = part_aliases.get(packing_part, packing_part)
            records.append(
                {
                    "line_no": line_no,
                    "packing_no": raw_text(_cell(row, mapped.get("no"))),
                    "raw_part_num": part_raw,
                    "numero_parte_packing": packing_part,
                    "numero_parte_sistema": system_part,
                    "descripcion": raw_text(_cell(row, mapped.get("description"))),
                    "cantidad_packing": qty,
                    "raw_qty": qty_raw,
                    "pallet_no_original": pallet_raw,
                    "pallet_no": normalizar_pallet_no(pallet_raw),
                    "kg": decimal_or_zero(_cell(row, mapped.get("kg"))) if "kg" in mapped else None,
                    "cbm": decimal_or_zero(_cell(row, mapped.get("cbm"))) if "cbm" in mapped else None,
                    "estado_match": "PENDIENTE",
                    "mensaje_match": "",
                }
            )
    return records


def _find_packing_original_header(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 1, 100)):
        indexes = _header_index_map(row)
        no_idx = _first_existing(indexes, "NO")
        part_idx = _first_existing(indexes, "품번", "PARTNO", "PARTNUM")
        description_idx = _first_existing(indexes, "품명", "DESCRIPTION", "DESC")
        qty_idx = _first_existing(indexes, "수량", "QTY", "QUANTITY")
        if no_idx is not None and part_idx is not None and qty_idx is not None:
            return {
                "row_number": row[0].row,
                "no": no_idx,
                "part_num": part_idx,
                "description": description_idx,
                "qty": qty_idx,
            }
    return None


def _infer_value_before_label_index(sheet, start_row, label_key):
    for row in sheet.iter_rows(min_row=start_row, max_row=min(sheet.max_row or start_row, start_row + 30)):
        for idx, cell in enumerate(row):
            if _header_key(cell.value) == label_key and idx > 0:
                return idx - 1
    return None


def _parse_packing_original_sheet(sheet, part_aliases=None):
    """Parsea PACKING LIST 원본: columna NO es el pallet y se hereda por grupo."""
    part_aliases = part_aliases or {}
    header = _find_packing_original_header(sheet)
    if not header:
        return []

    records = []
    line_no = 0
    blank_streak = 0
    current_pallet_raw = ""
    start_row = header["row_number"] + 1
    kg_idx = _infer_value_before_label_index(sheet, start_row, "KG")
    cbm_idx = _infer_value_before_label_index(sheet, start_row, "CBM")

    for row in sheet.iter_rows(min_row=start_row):
        pallet_marker = raw_text(_cell(row, header.get("no")))
        if pallet_marker:
            current_pallet_raw = pallet_marker

        part_raw = raw_text(_cell(row, header.get("part_num")))
        qty_raw = raw_text(_cell(row, header.get("qty")))
        if not part_raw and not qty_raw and not pallet_marker:
            blank_streak += 1
            if blank_streak >= 8:
                break
            continue
        blank_streak = 0

        if not part_raw or _is_summary_part(part_raw):
            continue

        qty = decimal_or_zero(qty_raw)
        if qty <= 0:
            continue

        line_no += 1
        packing_part = normalizar_numero_parte(part_raw)
        system_part = part_aliases.get(packing_part, packing_part)
        records.append(
            {
                "line_no": line_no,
                "packing_no": current_pallet_raw,
                "raw_part_num": part_raw,
                "numero_parte_packing": packing_part,
                "numero_parte_sistema": system_part,
                "descripcion": raw_text(_cell(row, header.get("description"))),
                "cantidad_packing": qty,
                "raw_qty": qty_raw,
                "pallet_no_original": current_pallet_raw,
                "pallet_no": normalizar_pallet_no(current_pallet_raw),
                "kg": decimal_or_zero(_cell(row, kg_idx)) if kg_idx is not None else None,
                "cbm": decimal_or_zero(_cell(row, cbm_idx)) if cbm_idx is not None else None,
                "estado_match": "PENDIENTE",
                "mensaje_match": "",
            }
        )
    return records


def _converted_invoice_number(sheet):
    """Numero de invoice = primera celda de la hoja CONVERTED (A1)."""
    return sanitizar_texto(sheet.cell(row=1, column=1).value, 255)


def _find_converted_header(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 1, 20)):
        mapped = _map_header(row, CONVERTED_HEADER_ALIASES)
        if "raw_part_num" in mapped and "qty" in mapped:
            mapped["row_number"] = row[0].row
            return mapped
    return None


def _parse_converted_sheet(sheet):
    """Parsea INVOICE(CONVERTED): una sola tabla con parte, costo y tarima.

    Devuelve (invoice_lines, packing_lines). La TARIMA actua como pallet y se
    hereda por grupo cuando viene vacia. Cada fila valida produce a la vez una
    invoice line (con costo) y una packing line (pallet+parte+cantidad), de modo
    que la conciliacion packing existente funciona sin la hoja PACKING original.
    """
    header = _find_converted_header(sheet)
    if not header:
        return [], []

    invoice_lines = []
    packing_lines = []
    line_no = 0
    blank_streak = 0
    current_pallet_raw = ""
    start_row = header["row_number"] + 1

    for row in sheet.iter_rows(min_row=start_row):
        pallet_marker = raw_text(_cell(row, header.get("pallet")))
        if pallet_marker:
            current_pallet_raw = pallet_marker

        part_raw = raw_text(_cell(row, header.get("raw_part_num")))
        qty_raw = raw_text(_cell(row, header.get("qty")))
        if not part_raw and not qty_raw and not pallet_marker:
            blank_streak += 1
            if blank_streak >= 8:
                break
            continue
        blank_streak = 0

        if not part_raw or _is_summary_part(part_raw):
            continue

        qty = decimal_or_zero(qty_raw)
        if qty <= 0:
            continue

        line_no += 1
        unit_raw = raw_text(_cell(row, header.get("unit_cost")))
        total_raw = raw_text(_cell(row, header.get("total_cost")))
        unit = decimal_or_zero(unit_raw)
        total = decimal_or_zero(total_raw)
        if total == 0 and unit != 0 and qty != 0:
            total = (unit * qty).quantize(Decimal("0.0001"))

        invoice_part = normalizar_numero_parte(part_raw)
        # Part Sys ya viene resuelto en la hoja; si falta, se usa la parte raw.
        # El usuario apila el codigo como parte-version-lote (ej.
        # EAX66946005-1.0-...), pero materiales/almacen/costos usan la parte base.
        # No se pre-corta aqui: validate_system_parts (con acceso a materiales)
        # resuelve la base real probando los prefijos por guion. La version queda
        # disponible aparte y en raw_part_num.
        sys_raw = raw_text(_cell(row, header.get("part_sys")))
        system_part = normalizar_numero_parte(sys_raw) or invoice_part
        parte_version = version_parte(sys_raw) or version_parte(part_raw)
        item = raw_text(_cell(row, header.get("item")))
        spec = raw_text(_cell(row, header.get("spec")))
        uom = raw_text(_cell(row, header.get("uom")))
        descripcion = " ".join(part for part in (item, spec) if part).strip()
        pallet_raw = current_pallet_raw

        invoice_lines.append(
            {
                "line_no": line_no,
                "maker": "",
                "origin": "",
                "raw_part_num": part_raw,
                "numero_parte_invoice": invoice_part,
                "numero_parte_sistema": system_part,
                "descripcion": descripcion,
                "cantidad": qty,
                "uom": uom,
                "costo_unitario": unit,
                "costo_total": total,
                "moneda": MONEDA_DEFAULT,
                "raw_qty": qty_raw,
                "raw_unit_cost": unit_raw,
                "raw_total_cost": total_raw,
                "estado_match": "PENDIENTE",
                "mensaje_match": "",
            }
        )
        packing_lines.append(
            {
                "line_no": line_no,
                "packing_no": pallet_raw,
                "raw_part_num": part_raw,
                "numero_parte_packing": invoice_part,
                "numero_parte_sistema": system_part,
                "descripcion": descripcion,
                "cantidad_packing": qty,
                "raw_qty": qty_raw,
                "pallet_no_original": pallet_raw,
                "pallet_no": normalizar_pallet_no(pallet_raw),
                "kg": None,
                "cbm": None,
                "estado_match": "PENDIENTE",
                "mensaje_match": "",
            }
        )
    return invoice_lines, packing_lines


def _guess_invoice_number(wb, filename):
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(max_row=30, max_col=12):
            for idx, cell in enumerate(row):
                text = raw_text(cell.value)
                if "INVOICE" not in text.upper():
                    continue
                same = re.search(r"INVOICE\s*(?:NO|#|NUMBER|NUM)?\s*[:\-]?\s*(\S+)", text, re.I)
                if same and same.group(1).upper() not in {"NO", "NUMBER"}:
                    return sanitizar_texto(same.group(1), 255)
                if idx + 1 < len(row):
                    adjacent = sanitizar_texto(row[idx + 1].value, 255)
                    if adjacent:
                        return adjacent
    return sanitizar_texto(Path(filename or "invoice").stem, 255) or "SIN_NUMERO"


def parse_invoice_workbook(file_bytes, filename=""):
    """Parsea invoice/packing usando openpyxl data_only=True.

    Devuelve datos normalizados y conserva raw values para depuracion.
    Esta funcion no toca BD para que sea testeable.
    """
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    invoice_lines = []
    packing_lines = []
    seen_headers = set()
    warnings = []

    # Fuente principal: hoja INVOICE(CONVERTED), formato interno ya conciliado.
    converted_sheet = next(
        (sheet for sheet in wb.worksheets if _is_invoice_converted_sheet(sheet)),
        None,
    )
    if converted_sheet is not None:
        invoice_lines, packing_lines = _parse_converted_sheet(converted_sheet)
        numero_invoice = (
            _converted_invoice_number(converted_sheet)
            or _guess_invoice_number(wb, filename)
        )
        if not invoice_lines:
            warnings.append(
                "La hoja INVOICE(CONVERTED) no contiene lineas validas."
            )
        return {
            "numero_invoice_sugerido": numero_invoice,
            "invoice_lines": invoice_lines,
            "packing_lines": packing_lines,
            "warnings": warnings,
            "fuente_hoja": INVOICE_CONVERTED_SHEET_NAME,
        }

    # Fallback: formato anterior INVOICE(total) + PACKING LIST original.
    invoice_sheet_found = any(_is_invoice_total_sheet(sheet) for sheet in wb.worksheets)
    packing_original_found = any(_is_packing_original_sheet(sheet) for sheet in wb.worksheets)
    part_aliases = {}

    if packing_original_found:
        for sheet in wb.worksheets:
            if _is_packing_original_sheet(sheet):
                packing_lines.extend(_parse_packing_original_sheet(sheet, part_aliases))

    for sheet in wb.worksheets:
        is_invoice_total = _is_invoice_total_sheet(sheet)
        is_packing_original = _is_packing_original_sheet(sheet)
        rows = list(sheet.iter_rows())
        for row in rows:
            row_number = row[0].row if row else 0
            if (sheet.title, row_number) in seen_headers:
                continue
            inv_map = _map_header(row, INVOICE_HEADER_ALIASES)
            pack_map = _map_header(row, PACKING_HEADER_ALIASES)
            if not packing_original_found and not is_packing_original and _is_packing_header(pack_map):
                seen_headers.add((sheet.title, row_number))
                packing_lines.extend(
                    _parse_rows_after_header(
                        sheet, row_number, pack_map, "packing", len(packing_lines), part_aliases
                    )
                )
                continue
            if is_invoice_total and _is_invoice_header(inv_map):
                seen_headers.add((sheet.title, row_number))
                invoice_lines.extend(
                    _parse_rows_after_header(
                        sheet, row_number, inv_map, "invoice", len(invoice_lines), part_aliases
                    )
                )
    if not invoice_sheet_found:
        warnings.append("No se encontro la hoja INVOICE(CONVERTED) ni INVOICE(total).")

    return {
        "numero_invoice_sugerido": _guess_invoice_number(wb, filename),
        "invoice_lines": invoice_lines,
        "packing_lines": packing_lines,
        "warnings": warnings,
        "fuente_hoja": INVOICE_TOTAL_SHEET_NAME,
    }


# Compatibilidad con tests/imports antiguos.
_parse_invoice_workbook = parse_invoice_workbook
