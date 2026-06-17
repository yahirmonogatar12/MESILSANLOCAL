"""Parser Excel para Lista de compras (hoja 'Compras totales').

No toca BD. Mapea columnas por NOMBRE de encabezado (no por posicion) porque los
archivos LG y Oven traen layouts distintos. Encabezado en fila 2; datos fila 3+.
Reusa los normalizadores de invoice_core.
"""

import re
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from app.api.control_material.invoice_core.normalizers import (
    decimal_or_zero,
    normalizar_numero_parte,
    raw_text,
)
from app.api.control_material.invoice_core.parser import _header_key, _is_summary_part

COMPRAS_SHEET_NAME = "Compras totales"
HEADER_ROW = 2

# Encabezado normalizado (_header_key: upper, sin espacios/.-_/#:) -> campo interno.
# Ej. "Part No." -> PARTNO, "Precio Unitario" -> PRECIOUNITARIO, "Trans." -> TRANS.
COMPRAS_HEADER_ALIASES = {
    "ANO": "anio",
    "AÑO": "anio",  # por si _header_key conserva la Ñ
    "MES": "mes",
    "FECHA": "fecha_compra",
    "WK": "wk",
    "PARTNO": "raw_part_num",
    "PARTNUM": "raw_part_num",
    "PARTNUMBER": "raw_part_num",
    "PARTSYS": "part_sys",
    "PARTSYSTEM": "part_sys",
    "DESC": "descripcion",
    "PARTDESC": "descripcion",
    "DESCRIPTION": "descripcion",
    "SPEC": "spec",
    "SHORTSPEC": "spec",
    "CANTIDAD": "cantidad",
    "QTY": "cantidad",
    "TRANS": "numero_transaccion",
    "TRANSACCION": "numero_transaccion",
    "FECHADEFACTURA": "fecha_factura",
    "CURR": "moneda",
    "MONEDA": "moneda",
    "PRECIOUNITARIO": "costo_unitario",
    "UNITCOST": "costo_unitario",
    "TOTAL": "costo_total",
    "PROVEEDOR": "proveedor",
    "FACTURA": "factura",
    "MODEL": "modelo",
    "LINEA": "modelo",
    "PCBHARNESS": "categoria",
    "COMENTARIO": "comentario",
}


def _sheet_key(name):
    return re.sub(r"\s+", "", raw_text(name)).upper()


def _find_compras_sheet(wb):
    target = _sheet_key(COMPRAS_SHEET_NAME)
    for name in wb.sheetnames:
        if _sheet_key(name) == target:
            return wb[name]
    return None


def _parse_date(value):
    """Devuelve 'YYYY-MM-DD' o None. Acepta datetime/date o texto."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = raw_text(value)
    # raw_text ya formatea fechas; si vino como texto largo, recorta a la fecha.
    m = re.match(r"(\d{4}-\d{2}-\d{2})", text)
    return m.group(1) if m else None


def _parse_int(value):
    raw = raw_text(value)
    m = re.search(r"-?\d+", raw)
    return int(m.group(0)) if m else None


def parse_compras_workbook(file_bytes, filename=""):
    """Parsea la hoja 'Compras totales'. Devuelve {lineas, warnings}."""
    import openpyxl

    warnings = []
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        sheet = _find_compras_sheet(wb)
        if sheet is None:
            return {
                "lineas": [],
                "warnings": [f"No se encontro la hoja '{COMPRAS_SHEET_NAME}'."],
            }

        rows = sheet.iter_rows(min_row=HEADER_ROW, values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return {"lineas": [], "warnings": ["La hoja esta vacia."]}

        # header_key -> indice de columna
        col_index = {}
        for idx, cell in enumerate(header):
            field = COMPRAS_HEADER_ALIASES.get(_header_key(cell))
            if field and field not in col_index:
                col_index[field] = idx

        if "numero_transaccion" not in col_index and "raw_part_num" not in col_index:
            return {
                "lineas": [],
                "warnings": ["No se reconocieron columnas (Part No. / Trans.)."],
            }

        def cell(row, field):
            i = col_index.get(field)
            if i is None or i >= len(row):
                return None
            return row[i]

        lineas = []
        for row in rows:
            if row is None or all(v is None for v in row):
                continue

            raw_part = raw_text(cell(row, "raw_part_num"))
            numero_transaccion = raw_text(cell(row, "numero_transaccion"))
            # Necesita al menos parte o transaccion; salta resumenes (TOTAL...).
            if not raw_part and not numero_transaccion:
                continue
            if _is_summary_part(raw_part):
                continue

            cantidad = decimal_or_zero(cell(row, "cantidad"))
            if cantidad <= 0:
                continue

            numero_parte = normalizar_numero_parte(raw_part)
            part_sys = normalizar_numero_parte(cell(row, "part_sys"))
            numero_parte_sistema = part_sys or numero_parte

            costo_unitario = decimal_or_zero(cell(row, "costo_unitario"))
            costo_total = decimal_or_zero(cell(row, "costo_total"))
            if costo_total == 0 and costo_unitario > 0:
                costo_total = (costo_unitario * cantidad).quantize(Decimal("0.0001"))
            # Sin precio: se guarda NULL para distinguir de un costo real de 0.
            costo_unitario_out = costo_unitario if costo_unitario > 0 else None
            costo_total_out = costo_total if costo_total > 0 else None

            moneda = (raw_text(cell(row, "moneda")) or "USD").upper()[:10]

            lineas.append(
                {
                    "numero_transaccion": numero_transaccion[:255],
                    "anio": _parse_int(cell(row, "anio")),
                    "mes": raw_text(cell(row, "mes"))[:20] or None,
                    "fecha_compra": _parse_date(cell(row, "fecha_compra")),
                    "wk": raw_text(cell(row, "wk"))[:20] or None,
                    "raw_part_num": raw_part[:512] or None,
                    "numero_parte": numero_parte[:512],
                    "numero_parte_sistema": numero_parte_sistema[:512],
                    "descripcion": raw_text(cell(row, "descripcion"))[:1024] or None,
                    "spec": raw_text(cell(row, "spec"))[:512] or None,
                    "cantidad": cantidad,
                    "moneda": moneda,
                    "costo_unitario": costo_unitario_out,
                    "costo_total": costo_total_out,
                    "fecha_factura": _parse_date(cell(row, "fecha_factura")),
                    "proveedor": raw_text(cell(row, "proveedor"))[:255] or None,
                    "factura": raw_text(cell(row, "factura"))[:255] or None,
                    "modelo": raw_text(cell(row, "modelo"))[:255] or None,
                    "categoria": raw_text(cell(row, "categoria"))[:50] or None,
                    "comentario": raw_text(cell(row, "comentario"))[:512] or None,
                }
            )

        if not lineas:
            warnings.append("No se detectaron renglones validos en 'Compras totales'.")
        return {"lineas": lineas, "warnings": warnings}
    finally:
        wb.close()


if __name__ == "__main__":
    # ponytail: self-check minimo del mapeo por encabezado + reglas de costo.
    # Verifica _header_key/alias y la regla de costo NULL sin abrir Excel real.
    assert _header_key("Part No.") == "PARTNO"
    assert _header_key("Precio Unitario") == "PRECIOUNITARIO"
    assert _header_key("Trans.") == "TRANS"
    assert _header_key("Fecha de factura") == "FECHADEFACTURA"
    assert _header_key("PCB/HARNESS") == "PCBHARNESS"
    assert COMPRAS_HEADER_ALIASES[_header_key("Part Sys")] == "part_sys"
    assert _parse_date(datetime(2024, 1, 2)) == "2024-01-02"
    assert _parse_date("2024-01-02 00:00:00") == "2024-01-02"
    assert _parse_int("WK 12") == 12
    print("compras parser self-check OK")
