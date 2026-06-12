"""Vistas administrativas de Control de material: inventario actual,
historial de entradas/salidas/retornos y exportes Excel.

Migrado desde `app/Almacen_api.py` (mal nombrado: el archivo decia
"Almacen" pero su contenido es material_admin).

Rutas (sin cambios respecto al legacy):
  GET  /material/inventario_actual              -> render HTML
  GET  /material/historial_admin/<tipo>         -> render HTML
  GET  /api/material_admin/inventory/summary    -> JSON
  GET  /api/material_admin/inventory/lots       -> JSON
  GET  /api/material_admin/inventory/export     -> XLSX
  GET  /api/material_admin/history/<tipo>       -> JSON
  GET  /api/material_admin/history/<tipo>/export-> XLSX
"""

import io
import traceback
from datetime import date, datetime
from datetime import time as dt_time
from decimal import Decimal
from functools import wraps

from flask import Blueprint, jsonify, make_response, render_template, request, session

from app.api.shared import (
    auth_system,
    execute_query,
    login_requerido,
    obtener_fecha_hora_mexico,
)
# Fase 2 (2026-05-28): import directo del blueprint dueno; antes resolvia
# via app.api.shared.__getattr__ -> app.routes (re-export zombie).
from app.api.control_produccion.cuchillas_corte import _cuchillas_rows_to_json

import logging
logger = logging.getLogger(__name__)


bp = Blueprint("material_admin", __name__)

MATERIAL_ADMIN_PAGE = "LISTA_DE_MATERIALES"
MATERIAL_ADMIN_SECTION = "Control de material"
HISTORY_PERMISSION_BY_TYPE = {
    "entradas": "Historial de entradas",
    "salidas": "Historial de salidas",
    "retornos": "Historial de retornos",
}


def _material_admin_wants_json():
    accept = request.headers.get("Accept", "")
    return request.path.startswith("/api/") or request.is_json or "application/json" in accept


def _material_admin_permission_denied(boton):
    permiso = f"{MATERIAL_ADMIN_PAGE} > {MATERIAL_ADMIN_SECTION} > {boton}"
    if _material_admin_wants_json():
        return jsonify({"error": f"No tienes permisos para acceder a: {boton}", "permiso_requerido": permiso}), 403
    return (
        f"""
        <div style="
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            height: 400px;
            background: #2c2c2c;
            color: #e0e0e0;
            border-radius: 10px;
            margin: 20px;
            text-align: center;
        ">
            <h3>Acceso Denegado</h3>
            <p>No tienes permisos para acceder a: <strong>{boton}</strong></p>
            <p style="font-size: 0.9rem; opacity: 0.7;">Permiso requerido: {permiso}</p>
        </div>
        """,
        403,
    )


def requiere_permiso_material_admin(boton_resolver):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            boton = boton_resolver(*args, **kwargs) if callable(boton_resolver) else boton_resolver
            if not boton:
                return fn(*args, **kwargs)

            username = session.get("usuario")
            if not username:
                return jsonify({"error": "Usuario no autenticado", "redirect": "/login"}), 401

            try:
                if auth_system.obtener_rol_principal_usuario(username) == "superadmin":
                    return fn(*args, **kwargs)

                if auth_system.verificar_permiso_boton(username, MATERIAL_ADMIN_PAGE, MATERIAL_ADMIN_SECTION, boton):
                    return fn(*args, **kwargs)
                return _material_admin_permission_denied(boton)
            except Exception as exc:
                logger.error(f"Error verificando permiso material admin: {exc}")
                traceback.print_exc()
                return jsonify({"error": "Error interno verificando permisos"}), 500

        return wrapper

    return decorator


@bp.route("/material/inventario_actual")
@login_requerido
@requiere_permiso_material_admin("Inventario actual")
def material_inventario_actual_admin():
    """Vista administrativa read-only de inventario actual."""
    try:
        return render_template("Control de material/material_admin_inventory_ajax.html")
    except Exception as e:
        logger.error(f"Error al cargar Inventario actual administrativo: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@bp.route("/material/historial_admin/<tipo>")
@login_requerido
@requiere_permiso_material_admin(lambda tipo: HISTORY_PERMISSION_BY_TYPE.get(tipo))
def material_historial_admin(tipo):
    """Vista administrativa read-only de entradas, salidas o retornos."""
    titulos = {
        "entradas": "Historial de entradas",
        "salidas": "Historial de salidas",
        "retornos": "Historial de retornos",
    }
    if tipo not in titulos:
        return "Historial no valido", 404
    try:
        return render_template(
            "Control de material/material_admin_history_ajax.html",
            tipo=tipo,
            titulo=titulos[tipo],
        )
    except Exception as e:
        logger.error(f"Error al cargar {tipo}: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


def _material_admin_bool(value):
    return str(value or "").strip().lower() in ("1", "true", "yes", "si", "on")


def _material_admin_int(value, default, minimum=0, maximum=5000):
    try:
        parsed = int(value)
        return max(minimum, min(maximum, parsed))
    except Exception:
        return default


def _material_admin_add_version(rows):
    parsed_rows = _cuchillas_rows_to_json(rows)
    for row in parsed_rows:
        numero_parte = str(row.get("numero_parte") or "")
        if "-" in numero_parte:
            base, version = numero_parte.split("-", 1)
        else:
            base, version = numero_parte, "N/A"
        row["numero_parte_base"] = base
        row["version"] = version
    return parsed_rows


def _material_admin_inventory_summary_rows(args):
    numero_parte = (args.get("numero_parte") or "").strip()
    ubicacion = (args.get("ubicacion") or "").strip()
    include_zero = _material_admin_bool(args.get("include_zero_stock"))
    fecha_inicio = (args.get("fecha_inicio") or "").strip()
    fecha_fin = (args.get("fecha_fin") or "").strip()
    params = []

    if fecha_inicio and fecha_fin:
        inicio = f"{fecha_inicio} 00:00:00"
        fin = f"{fecha_fin} 23:59:59"
        query = """
            SELECT
              cma.numero_parte,
              MAX(COALESCE(m.especificacion_material, cma.especificacion)) AS especificacion,
              MAX(cma.ubicacion_salida) AS ubicacion,
              SUM(CASE WHEN cma.fecha_recibo BETWEEN %s AND %s THEN cma.cantidad_actual ELSE 0 END) AS total_entrada,
              COALESCE(SUM(cms_periodo.cantidad_salida), 0) AS total_salida,
              (
                SUM(CASE WHEN cma.fecha_recibo <= %s THEN cma.cantidad_actual ELSE 0 END)
                - COALESCE(SUM(cms_historico.cantidad_salida), 0)
              ) AS stock_total,
              COUNT(DISTINCT CASE WHEN cma.fecha_recibo BETWEEN %s AND %s THEN cma.numero_lote_material END) AS lotes_distintos,
              SUM(CASE WHEN (
                CASE WHEN cma.fecha_recibo <= %s THEN cma.cantidad_actual ELSE 0 END
                - COALESCE(cms_historico.sal_individual, 0)
              ) > 0 THEN 1 ELSE 0 END) AS lotes_con_stock,
              MAX(IFNULL(m.unidad_medida, cma.unidad_medida)) AS unidad_medida
            FROM control_material_almacen cma
            LEFT JOIN materiales m ON SUBSTRING_INDEX(cma.numero_parte, '-', 1) = m.numero_parte
            LEFT JOIN (
              SELECT codigo_material_recibido, SUM(cantidad_salida) AS cantidad_salida
              FROM control_material_salida
              WHERE fecha_salida BETWEEN %s AND %s AND (cancelado = 0 OR cancelado IS NULL)
              GROUP BY codigo_material_recibido
            ) cms_periodo ON cma.codigo_material_recibido = cms_periodo.codigo_material_recibido
            LEFT JOIN (
              SELECT codigo_material_recibido, SUM(cantidad_salida) AS cantidad_salida, SUM(cantidad_salida) AS sal_individual
              FROM control_material_salida
              WHERE fecha_salida <= %s AND (cancelado = 0 OR cancelado IS NULL)
              GROUP BY codigo_material_recibido
            ) cms_historico ON cma.codigo_material_recibido = cms_historico.codigo_material_recibido
            WHERE (cma.cancelado = 0 OR cma.cancelado IS NULL)
              AND cma.iqc_status IN ('Released', 'NotRequired')
              AND (cma.fecha_recibo BETWEEN %s AND %s OR cms_periodo.codigo_material_recibido IS NOT NULL)
        """
        params.extend([inicio, fin, fin, inicio, fin, fin, inicio, fin, fin, inicio, fin])
        if numero_parte:
            query += " AND cma.numero_parte LIKE %s"
            params.append(f"%{numero_parte}%")
        if ubicacion:
            query += " AND cma.ubicacion_salida LIKE %s"
            params.append(f"%{ubicacion}%")
        query += " GROUP BY cma.numero_parte"
        if not include_zero:
            query += " HAVING stock_total > 0"
        query += " ORDER BY cma.numero_parte LIMIT 5000"
        return _material_admin_add_version(execute_query(query, tuple(params), fetch="all") or [])

    query = """
        SELECT
          il.numero_parte,
          MAX(COALESCE(m.especificacion_material, cma.especificacion)) AS especificacion,
          SUM(il.total_entrada) AS total_entrada,
          SUM(il.total_salida) AS total_salida,
          SUM(il.stock_actual) AS stock_total,
          COUNT(DISTINCT il.numero_lote) AS lotes_distintos,
          SUM(CASE WHEN il.stock_actual > 0 THEN 1 ELSE 0 END) AS lotes_con_stock,
          MAX(IFNULL(il.unidad_medida, IFNULL(m.unidad_medida, 'EA'))) AS unidad_medida,
          MAX(m.ubicacion_material) AS ubicacion
        FROM inventario_lotes il
        LEFT JOIN control_material_almacen cma ON il.codigo_material_recibido = cma.codigo_material_recibido
        LEFT JOIN materiales m ON SUBSTRING_INDEX(il.numero_parte, '-', 1) = m.numero_parte
        WHERE 1=1
    """
    if numero_parte:
        query += " AND il.numero_parte LIKE %s"
        params.append(f"%{numero_parte}%")
    if ubicacion:
        query += " AND cma.ubicacion_salida LIKE %s"
        params.append(f"%{ubicacion}%")
    query += " GROUP BY il.numero_parte"
    if not include_zero:
        query += " HAVING stock_total > 0"
    query += " ORDER BY il.numero_parte LIMIT 5000"
    return _material_admin_add_version(execute_query(query, tuple(params) if params else None, fetch="all") or [])


def _material_admin_column_filter_clauses(args, field_exprs):
    clauses = []
    params = []
    for field, expression in field_exprs.items():
        value = (args.get(f"cf_{field}") or "").strip()
        if not value:
            continue
        clauses.append(f"CAST(COALESCE({expression}, '') AS CHAR) LIKE %s")
        params.append(f"%{value}%")
    return clauses, params


def _material_admin_apply_column_filters(query, args, field_exprs, params):
    clauses, filter_params = _material_admin_column_filter_clauses(args, field_exprs)
    if clauses:
        query += " AND " + " AND ".join(clauses)
        params.extend(filter_params)
    return query


def _material_admin_wrap_column_filters(query, args, field_exprs, params, alias_name):
    clauses, filter_params = _material_admin_column_filter_clauses(args, field_exprs)
    if clauses:
        query = f"SELECT * FROM ({query}) AS {alias_name} WHERE " + " AND ".join(clauses)
        params.extend(filter_params)
    return query


def _material_admin_inventory_lots_result(args, force_export=False):
    numero_parte = (args.get("numero_parte") or "").strip()
    codigo = (args.get("codigo_material_recibido") or "").strip()
    ubicacion = (args.get("ubicacion") or "").strip()
    include_zero = _material_admin_bool(args.get("include_zero_stock"))
    only_na = _material_admin_bool(args.get("only_na"))
    fecha_inicio = (args.get("fecha_inicio") or "").strip()
    fecha_fin = (args.get("fecha_fin") or "").strip()
    limit = 5000 if force_export else _material_admin_int(args.get("limit"), 500, 1, 5000)
    offset = 0 if force_export else _material_admin_int(args.get("offset"), 0, 0, 10000000)
    params = []

    if fecha_inicio and fecha_fin:
        inicio = f"{fecha_inicio} 00:00:00"
        fin = f"{fecha_fin} 23:59:59"
        base = """
            SELECT
              cma.numero_parte,
              SUBSTRING_INDEX(cma.numero_parte, '-', 1) AS numero_parte_base,
              CASE WHEN cma.numero_parte LIKE '%%-%%' THEN SUBSTRING(cma.numero_parte, LOCATE('-', cma.numero_parte) + 1) ELSE 'N/A' END AS version,
              cma.numero_lote_material AS numero_lote,
              cma.codigo_material_recibido,
              CASE WHEN cma.fecha_recibo BETWEEN %s AND %s THEN cma.cantidad_actual ELSE 0 END AS total_entrada,
              COALESCE(cms_periodo.cantidad_salida, 0) AS total_salida,
              (
                CASE WHEN cma.fecha_recibo <= %s THEN cma.cantidad_actual ELSE 0 END
                - COALESCE(cms_historico.cantidad_salida, 0)
              ) AS stock_actual,
              cma.codigo_material,
              cma.ubicacion_salida AS ubicacion,
              COALESCE(m.especificacion_material, cma.especificacion) AS especificacion,
              COALESCE(cma.en_cuarentena, 0) AS in_quarantine,
              IFNULL(cma.unidad_medida, IFNULL(m.unidad_medida, 'EA')) AS unidad_medida,
              cma.fecha_recibo AS fecha_recibo,
              cms_periodo.ultima_fecha_salida AS fecha_salida,
              cma.usuario_registro AS usuario_entrada,
              cms_periodo.usuario_salida AS usuario_salida
            FROM control_material_almacen cma
            LEFT JOIN materiales m ON SUBSTRING_INDEX(cma.numero_parte, '-', 1) = m.numero_parte
            LEFT JOIN (
              SELECT codigo_material_recibido,
                     SUM(cantidad_salida) AS cantidad_salida,
                     MAX(fecha_salida) AS ultima_fecha_salida,
                     MAX(usuario_registro) AS usuario_salida
              FROM control_material_salida
              WHERE fecha_salida BETWEEN %s AND %s AND (cancelado = 0 OR cancelado IS NULL)
              GROUP BY codigo_material_recibido
            ) cms_periodo ON cma.codigo_material_recibido = cms_periodo.codigo_material_recibido
            LEFT JOIN (
              SELECT codigo_material_recibido, SUM(cantidad_salida) AS cantidad_salida
              FROM control_material_salida
              WHERE fecha_salida <= %s AND (cancelado = 0 OR cancelado IS NULL)
              GROUP BY codigo_material_recibido
            ) cms_historico ON cma.codigo_material_recibido = cms_historico.codigo_material_recibido
            WHERE (cma.cancelado = 0 OR cma.cancelado IS NULL)
              AND cma.iqc_status IN ('Released', 'NotRequired')
              AND (cma.fecha_recibo BETWEEN %s AND %s OR cms_periodo.codigo_material_recibido IS NOT NULL)
        """
        params.extend([inicio, fin, fin, inicio, fin, fin, inicio, fin])
        if not include_zero:
            base += " AND (CASE WHEN cma.fecha_recibo <= %s THEN cma.cantidad_actual ELSE 0 END - COALESCE(cms_historico.cantidad_salida, 0)) > 0"
            params.append(fin)
        if numero_parte:
            base += " AND cma.numero_parte LIKE %s"
            params.append(f"%{numero_parte}%")
        if codigo:
            base += " AND cma.codigo_material_recibido LIKE %s"
            params.append(f"%{codigo}%")
        if only_na:
            base += " AND (cma.numero_lote_material IS NULL OR cma.numero_lote_material = '' OR UPPER(cma.numero_lote_material) = 'N/A')"
        if ubicacion:
            base += " AND cma.ubicacion_salida LIKE %s"
            params.append(f"%{ubicacion}%")
        base = _material_admin_wrap_column_filters(base, args, {
            "numero_parte": "numero_parte",
            "numero_parte_base": "numero_parte_base",
            "version": "version",
            "numero_lote": "numero_lote",
            "codigo_material_recibido": "codigo_material_recibido",
            "especificacion": "especificacion",
            "ubicacion": "ubicacion",
            "in_quarantine": "CASE WHEN in_quarantine IN (1, '1', TRUE) THEN 'Si' ELSE 'No' END",
            "unidad_medida": "unidad_medida",
            "total_entrada": "total_entrada",
            "total_salida": "total_salida",
            "stock_actual": "stock_actual",
            "fecha_recibo": "fecha_recibo",
            "fecha_salida": "fecha_salida",
            "usuario_entrada": "usuario_entrada",
            "usuario_salida": "usuario_salida",
        }, params, "inv_lotes")
        count = execute_query(f"SELECT COUNT(*) AS total FROM ({base}) AS t", tuple(params), fetch="one") or {}
        data = execute_query(base + " ORDER BY numero_parte, numero_lote LIMIT %s OFFSET %s", tuple(params + [limit, offset]), fetch="all") or []
        return {"data": _material_admin_add_version(data), "total": int(count.get("total") or 0), "limit": limit, "offset": offset}

    base = """
        SELECT
          il.numero_parte,
          SUBSTRING_INDEX(il.numero_parte, '-', 1) AS numero_parte_base,
          CASE WHEN il.numero_parte LIKE '%%-%%' THEN SUBSTRING(il.numero_parte, LOCATE('-', il.numero_parte) + 1) ELSE 'N/A' END AS version,
          il.numero_lote,
          il.codigo_material_recibido,
          il.total_entrada,
          il.total_salida,
          il.stock_actual,
          cma.codigo_material,
          cma.ubicacion_salida AS ubicacion,
          COALESCE(m.especificacion_material, cma.especificacion) AS especificacion,
          COALESCE(cma.en_cuarentena, 0) AS in_quarantine,
          IFNULL(il.unidad_medida, IFNULL(m.unidad_medida, 'EA')) AS unidad_medida,
          cma.fecha_recibo AS fecha_recibo,
          cms.ultima_fecha_salida AS fecha_salida,
          cma.usuario_registro AS usuario_entrada,
          cms.usuario_salida
        FROM inventario_lotes il
        LEFT JOIN control_material_almacen cma ON il.codigo_material_recibido = cma.codigo_material_recibido
        LEFT JOIN materiales m ON SUBSTRING_INDEX(il.numero_parte, '-', 1) = m.numero_parte
        LEFT JOIN (
          SELECT codigo_material_recibido,
                 MAX(fecha_salida) AS ultima_fecha_salida,
                 MAX(usuario_registro) AS usuario_salida
          FROM control_material_salida
          WHERE (cancelado = 0 OR cancelado IS NULL)
          GROUP BY codigo_material_recibido
        ) cms ON il.codigo_material_recibido = cms.codigo_material_recibido
        WHERE 1=1
    """
    if not include_zero:
        base += " AND il.stock_actual > 0"
    if numero_parte:
        base += " AND il.numero_parte LIKE %s"
        params.append(f"%{numero_parte}%")
    if codigo:
        base += " AND il.codigo_material_recibido LIKE %s"
        params.append(f"%{codigo}%")
    if only_na:
        base += " AND (il.numero_lote IS NULL OR il.numero_lote = '' OR UPPER(il.numero_lote) = 'N/A')"
    if ubicacion:
        base += " AND cma.ubicacion_salida LIKE %s"
        params.append(f"%{ubicacion}%")
    base = _material_admin_wrap_column_filters(base, args, {
        "numero_parte": "numero_parte",
        "numero_parte_base": "numero_parte_base",
        "version": "version",
        "numero_lote": "numero_lote",
        "codigo_material_recibido": "codigo_material_recibido",
        "especificacion": "especificacion",
        "ubicacion": "ubicacion",
        "in_quarantine": "CASE WHEN in_quarantine IN (1, '1', TRUE) THEN 'Si' ELSE 'No' END",
        "unidad_medida": "unidad_medida",
        "total_entrada": "total_entrada",
        "total_salida": "total_salida",
        "stock_actual": "stock_actual",
        "fecha_recibo": "fecha_recibo",
        "fecha_salida": "fecha_salida",
        "usuario_entrada": "usuario_entrada",
        "usuario_salida": "usuario_salida",
    }, params, "inv_lotes")
    count = execute_query(f"SELECT COUNT(*) AS total FROM ({base}) AS t", tuple(params) if params else None, fetch="one") or {}
    data = execute_query(base + " ORDER BY il.numero_parte, il.numero_lote LIMIT %s OFFSET %s", tuple(params + [limit, offset]), fetch="all") or []
    return {"data": _material_admin_add_version(data), "total": int(count.get("total") or 0), "limit": limit, "offset": offset}


def _material_admin_rows_with_times(rows, time_fields=None):
    parsed_rows = _cuchillas_rows_to_json(rows)
    for row in parsed_rows:
        for target, source in (time_fields or {}).items():
            raw_value = str(row.get(source) or "")
            row[target] = raw_value.split(" ", 1)[1][:8] if " " in raw_value else ""
    return parsed_rows


def _material_admin_history_rows(tipo, args, export=False):
    fecha_inicio = (args.get("fecha_inicio") or "").strip()
    fecha_fin = (args.get("fecha_fin") or "").strip()
    texto = (args.get("texto") or "").strip()
    limit = _material_admin_int(args.get("limit"), 500, 1, 5000)
    offset = _material_admin_int(args.get("offset"), 0, 0, 10000000)
    params = []

    if tipo == "entradas":
        query = """
            SELECT
              cma.numero_parte,
              cma.especificacion,
              cma.codigo_material,
              cma.codigo_material_recibido,
              cma.numero_lote_material,
              cma.numero_invoice,
              cma.propiedad_material,
              cma.iqc_status,
              cma.en_cuarentena AS in_quarantine,
              cma.cantidad_actual,
              cma.cantidad_estandarizada,
              m.ubicacion_material AS location,
              cma.fecha_recibo AS fecha_recibo,
              cma.material_importacion_local,
              cma.estado_desecho,
              cma.vendedor,
              cma.cancelado,
              cma.usuario_registro
            FROM control_material_almacen cma
            LEFT JOIN materiales m ON cma.numero_parte = m.numero_parte
            WHERE 1=1
        """
        date_field = "cma.fecha_recibo"
        text_fields = ["cma.codigo_material_recibido", "cma.codigo_material", "cma.numero_parte", "cma.numero_lote_material", "cma.numero_invoice", "cma.ubicacion_salida", "m.ubicacion_material", "cma.especificacion", "cma.vendedor", "cma.usuario_registro"]
        order = "cma.fecha_recibo DESC, cma.id DESC"
        time_fields = {"fecha_recibo_hora": "fecha_recibo"}
        column_filter_fields = {
            "numero_parte": "cma.numero_parte",
            "especificacion": "cma.especificacion",
            "codigo_material": "cma.codigo_material",
            "codigo_material_recibido": "cma.codigo_material_recibido",
            "numero_lote_material": "cma.numero_lote_material",
            "numero_invoice": "cma.numero_invoice",
            "propiedad_material": "cma.propiedad_material",
            "iqc_status": "cma.iqc_status",
            "in_quarantine": "CASE WHEN cma.en_cuarentena IN (1, '1', TRUE) THEN 'Si' ELSE 'No' END",
            "cantidad_actual": "cma.cantidad_actual",
            "cantidad_estandarizada": "cma.cantidad_estandarizada",
            "location": "m.ubicacion_material",
            "fecha_recibo": "cma.fecha_recibo",
            "fecha_recibo_hora": "DATE_FORMAT(cma.fecha_recibo, '%H:%i:%s')",
            "material_importacion_local": "cma.material_importacion_local",
            "estado_desecho": "cma.estado_desecho",
            "vendedor": "cma.vendedor",
            "cancelado": "CASE WHEN cma.cancelado IN (1, '1', TRUE) THEN 'Si' ELSE 'No' END",
            "usuario_registro": "cma.usuario_registro",
        }
    elif tipo == "salidas":
        query = """
            SELECT
              cms.fecha_salida AS fecha_salida,
              cms.proceso_salida,
              cms.codigo_material_recibido,
              cms.numero_parte AS material_code,
              cms.numero_parte,
              cms.cantidad_salida,
              cms.modelo,
              cms.numero_lote,
              cms.especificacion_material,
              cms.depto_salida,
              cms.vendedor,
              cms.usuario_registro
            FROM control_material_salida cms
            LEFT JOIN control_material_almacen cma ON cms.codigo_material_recibido = cma.codigo_material_recibido
            WHERE 1=1
        """
        date_field = "cms.fecha_salida"
        text_fields = ["cms.codigo_material_recibido", "cms.numero_parte", "cms.numero_lote", "cms.modelo", "cms.depto_salida", "cms.proceso_salida", "cms.especificacion_material", "cms.vendedor", "cms.usuario_registro"]
        order = "cms.fecha_salida DESC, cms.id DESC"
        time_fields = {"fecha_salida_hora": "fecha_salida"}
        column_filter_fields = {
            "fecha_salida": "cms.fecha_salida",
            "fecha_salida_hora": "DATE_FORMAT(cms.fecha_salida, '%H:%i:%s')",
            "proceso_salida": "cms.proceso_salida",
            "codigo_material_recibido": "cms.codigo_material_recibido",
            "material_code": "cms.numero_parte",
            "numero_parte": "cms.numero_parte",
            "cantidad_salida": "cms.cantidad_salida",
            "modelo": "cms.modelo",
            "numero_lote": "cms.numero_lote",
            "especificacion_material": "cms.especificacion_material",
            "depto_salida": "cms.depto_salida",
            "vendedor": "cms.vendedor",
            "usuario_registro": "cms.usuario_registro",
        }
    elif tipo == "retornos":
        query = """
            SELECT
              mr.fecha_creacion AS fecha_creacion,
              mr.warehousing_code,
              cma.codigo_material AS material_code,
              mr.numero_parte,
              mr.numero_lote_material,
              cma.cantidad_estandarizada AS packaging_unit,
              mr.cantidad_devuelta,
              cma.especificacion AS material_spec,
              mr.descripcion_motivo
            FROM material_return mr
            LEFT JOIN control_material_almacen cma ON mr.warehousing_id = cma.id
            WHERE 1=1
        """
        date_field = "mr.fecha_creacion"
        text_fields = ["mr.warehousing_code", "cma.codigo_material", "mr.numero_parte", "mr.numero_lote_material", "cma.especificacion", "mr.descripcion_motivo"]
        order = "mr.fecha_creacion DESC, mr.id DESC"
        time_fields = {}
        column_filter_fields = {
            "fecha_creacion": "mr.fecha_creacion",
            "warehousing_code": "mr.warehousing_code",
            "material_code": "cma.codigo_material",
            "numero_parte": "mr.numero_parte",
            "numero_lote_material": "mr.numero_lote_material",
            "packaging_unit": "cma.cantidad_estandarizada",
            "cantidad_devuelta": "mr.cantidad_devuelta",
            "material_spec": "cma.especificacion",
            "descripcion_motivo": "mr.descripcion_motivo",
        }
    else:
        raise ValueError("Tipo de historial no valido")

    if fecha_inicio:
        query += f" AND DATE({date_field}) >= %s"
        params.append(fecha_inicio)
    if fecha_fin:
        query += f" AND DATE({date_field}) <= %s"
        params.append(fecha_fin)
    if texto:
        query += " AND (" + " OR ".join(f"{field} LIKE %s" for field in text_fields) + ")"
        params.extend([f"%{texto}%"] * len(text_fields))
    query = _material_admin_apply_column_filters(query, args, column_filter_fields, params)

    if export:
        query += f" ORDER BY {order}"
        return _material_admin_rows_with_times(execute_query(query, tuple(params), fetch="all") or [], time_fields)

    count = execute_query(f"SELECT COUNT(*) AS total FROM ({query}) AS t", tuple(params), fetch="one") or {}
    data = execute_query(
        query + f" ORDER BY {order} LIMIT %s OFFSET %s",
        tuple(params + [limit, offset]),
        fetch="all",
    ) or []
    return {
        "data": _material_admin_rows_with_times(data, time_fields),
        "total": int(count.get("total") or 0),
        "limit": limit,
        "offset": offset,
    }


def _material_admin_export_filename(base_name, compact=False):
    now = obtener_fecha_hora_mexico()
    timestamp = now.strftime("%Y%m%d_%H%M%S") if compact else now.strftime("%Y-%m-%d %H-%M-%S")
    return f"{base_name}_{timestamp}.xlsx"


def _material_admin_parse_datetime(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, dt_time.min)

    text = str(value).strip()
    if not text:
        return None

    normalized = text.replace("Z", "")
    try:
        return datetime.fromisoformat(normalized)
    except ValueError:
        pass

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _material_admin_excel_cell_value(row, field):
    date_fields = {"fecha_recibo", "fecha_salida", "fecha_registro"}
    datetime_fields = {"fecha_creacion"}

    if field.endswith("_hora"):
        base_field = field[:-5]
        parsed = _material_admin_parse_datetime(row.get(base_field))
        return parsed.strftime("%H:%M") if parsed else ""

    if field in date_fields:
        parsed = _material_admin_parse_datetime(row.get(field))
        return parsed.strftime("%d/%m/%Y") if parsed else (row.get(field, "") or "")

    if field in datetime_fields:
        parsed = _material_admin_parse_datetime(row.get(field))
        return parsed.strftime("%d/%m/%Y %H:%M") if parsed else (row.get(field, "") or "")

    if field in ("estado_desecho", "cancelado"):
        value = row.get(field)
        return "Yes" if str(value or "").strip().lower() in ("1", "true", "yes", "si", "on") else "No"

    return row.get(field, "")


def _material_admin_excel_response(filename, sheet_name, headers, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    header_fill = PatternFill(start_color="1E5128", end_color="1E5128", fill_type="solid")
    alternate_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, (_, label) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 30 if col_idx == 1 else 18

    for row_idx, row in enumerate(rows, 2):
        for col_idx, (field, _) in enumerate(headers, 1):
            value = _material_admin_excel_cell_value(row, field)
            cell = ws.cell(row=row_idx, column=col_idx)
            if field == "version":
                cell.value = "" if value in (None, "") else str(value)
                cell.number_format = "@"
            elif value not in (None, ""):
                try:
                    numeric_value = Decimal(str(value).replace(",", ""))
                    cell.value = int(numeric_value)
                except Exception:
                    cell.value = value
            else:
                cell.value = ""
            if (row_idx - 2) % 2 == 1:
                cell.fill = alternate_fill

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@bp.route("/api/material_admin/inventory/summary")
@login_requerido
@requiere_permiso_material_admin("Inventario actual")
def api_material_admin_inventory_summary():
    try:
        return jsonify(_material_admin_inventory_summary_rows(request.args))
    except Exception as e:
        logger.error(f"Error en summary inventario admin: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


@bp.route("/api/material_admin/inventory/lots")
@login_requerido
@requiere_permiso_material_admin("Inventario actual")
def api_material_admin_inventory_lots():
    try:
        return jsonify(_material_admin_inventory_lots_result(request.args))
    except Exception as e:
        logger.error(f"Error en lots inventario admin: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


@bp.route("/api/material_admin/inventory/export")
@login_requerido
@requiere_permiso_material_admin("Inventario actual")
def api_material_admin_inventory_export():
    try:
        view = (request.args.get("view") or "summary").strip()
        if view == "detail":
            result = _material_admin_inventory_lots_result(request.args, force_export=True)
            headers = [
                ("numero_parte_base", "No. de Parte"),
                ("version", "Version"),
                ("numero_lote", "Numero de Lote"),
                ("codigo_material_recibido", "Codigo de Almacen"),
                ("especificacion", "Especificacion"),
                ("ubicacion", "Ubicacion"),
                ("in_quarantine", "Cuarentena"),
                ("unidad_medida", "Unidad"),
                ("total_entrada", "Total Entrada"),
                ("total_salida", "Total Salida"),
                ("stock_actual", "Stock Actual"),
                ("fecha_recibo", "Fecha Entrada"),
                ("fecha_salida", "Fecha Salida"),
                ("usuario_entrada", "Usuario entrada"),
                ("usuario_salida", "Usuario salida"),
            ]
            return _material_admin_excel_response(_material_admin_export_filename("Inventory_Detail"), "Sheet1", headers, result["data"])

        rows = _material_admin_inventory_summary_rows(request.args)
        headers = [
            ("numero_parte_base", "No. de Parte"),
            ("version", "Version"),
            ("especificacion", "Especificacion"),
            ("unidad_medida", "Unidad"),
        ]
        if request.args.get("fecha_inicio") and request.args.get("fecha_fin"):
            headers.extend([
                ("total_entrada", "Entradas"),
                ("total_salida", "Salidas"),
            ])
        headers.extend([
            ("stock_total", "Stock total"),
            ("lotes_distintos", "Lotes"),
            ("lotes_con_stock", "Lotes con stock"),
        ])
        return _material_admin_excel_response(_material_admin_export_filename("Inventory_Summary"), "Sheet1", headers, rows)
    except Exception as e:
        logger.error(f"Error exportando inventario admin: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


@bp.route("/api/material_admin/history/<tipo>")
@login_requerido
@requiere_permiso_material_admin(lambda tipo: HISTORY_PERMISSION_BY_TYPE.get(tipo))
def api_material_admin_history(tipo):
    if tipo not in ("entradas", "salidas", "retornos"):
        return jsonify({"error": "Tipo de historial no valido"}), 404
    try:
        return jsonify(_material_admin_history_rows(tipo, request.args))
    except Exception as e:
        logger.error(f"Error en historial admin {tipo}: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


@bp.route("/api/material_admin/history/<tipo>/export")
@login_requerido
@requiere_permiso_material_admin(lambda tipo: HISTORY_PERMISSION_BY_TYPE.get(tipo))
def api_material_admin_history_export(tipo):
    headers_map = {
        "entradas": [
            ("codigo_material_recibido", "Codigo de Almacen"),
            ("codigo_material", "Codigo de Material"),
            ("numero_parte", "No. de Parte"),
            ("numero_lote_material", "Lote Proveedor"),
            ("numero_invoice", "No. de Invoice"),
            ("propiedad_material", "Propiedad de Material"),
            ("cantidad_actual", "Cant. Actual"),
            ("cantidad_estandarizada", "Unidad de Empaque"),
            ("location", "Ubicacion"),
            ("fecha_recibo", "Fecha de Almacen"),
            ("fecha_recibo_hora", "Hora"),
            ("especificacion", "Especificacion"),
            ("material_importacion_local", "Material Consignado"),
            ("estado_desecho", "Disposicion"),
            ("cancelado", "Cancelled"),
        ],
        "salidas": [
            ("fecha_salida", "Fecha de Salida"),
            ("fecha_salida_hora", "Hora"),
            ("proceso_salida", "Proceso de Salida"),
            ("codigo_material_recibido", "Codigo de Almacen"),
            ("material_code", "Codigo de Material"),
            ("numero_parte", "No. de Parte"),
            ("cantidad_salida", "Cant."),
            ("modelo", "Modelo"),
            ("numero_lote", "Lote Proveedor"),
            ("especificacion_material", "Especificacion"),
            ("depto_salida", "Depto. de Salida"),
            ("usuario_registro", "Registrado Por"),
        ],
        "retornos": [
            ("fecha_creacion", "Fecha/Hora Devolucion"),
            ("warehousing_code", "Codigo de Almacen"),
            ("material_code", "Codigo de Material"),
            ("numero_parte", "No. de Parte"),
            ("numero_lote_material", "Lote Proveedor"),
            ("packaging_unit", "Unidad de Empaque"),
            ("cantidad_devuelta", "Cant. Devolucion"),
            ("material_spec", "Especificacion"),
            ("descripcion_motivo", "Motivo"),
        ],
    }
    filename_map = {
        "entradas": _material_admin_export_filename("Material_Warehousing"),
        "salidas": _material_admin_export_filename("Material_Outgoing_History"),
        "retornos": _material_admin_export_filename("Material_Returns", compact=True),
    }
    sheet_map = {
        "entradas": "Sheet1",
        "salidas": "Sheet1",
        "retornos": "Material Returns",
    }
    if tipo not in headers_map:
        return jsonify({"error": "Tipo de historial no valido"}), 404
    try:
        rows = _material_admin_history_rows(tipo, request.args, export=True)
        return _material_admin_excel_response(filename_map[tipo], sheet_map[tipo], headers_map[tipo], rows)
    except Exception as e:
        logger.error(f"Error exportando historial admin {tipo}: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


# Limpieza 2026-06-12: POST /guardar_material eliminado. Su unico consumidor era
# material-edit-drawer.js (drawer "Nuevo registro" de CONTROL_DE_MATERIAL.html),
# borrado junto con ese modulo para reconstruirlo desde cero. La funcion de BD
# `app.db_mysql.guardar_material` sigue existiendo para el modulo nuevo.
