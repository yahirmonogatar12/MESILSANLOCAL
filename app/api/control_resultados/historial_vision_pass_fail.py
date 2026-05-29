"""Endpoints HTTP del modulo "Historial de maquina Vision % Pass/Fail".

Consumido por LISTA_DE_CONTROL_DE_RESULTADOS / Historial de maquinas calidad.
JS cliente: app/static/js/history_vision_pass_fail.js
Template:   app/templates/Control de resultados/history_vision_pass_fail.html

Rutas:
  GET /historial_vision_pass_fail/ajax  -> render template (canonica)
  GET /api/vision/pass-fail-summary     -> resumen agrupado por linea/numero parte
  GET /api/vision/pass-fail-summary/export -> exportar resumen a Excel

Alias 2026-05-27:
  GET /historial-vision-pass-fail       -> 301 a la canonica
  GET /historial-vision-pass-fail-ajax  -> 301 a la canonica

Migrado desde app/routes.py el 2026-05-27. Las 2 APIs ya tenian
@login_requerido en el original.

Helpers compartidos con historial_vision.py viven en
app/api/shared/vision_helpers.py.
Helpers de exportacion Excel viven en app/api/shared/excel_helpers.py.
"""

from datetime import datetime

from flask import Blueprint, jsonify, redirect, render_template

from app.api.shared import execute_query, login_requerido
from app.api.shared.excel_helpers import (
    _create_vision_pass_fail_excel_image,
    _send_excel_download,
)
from app.api.shared.vision_helpers import (
    _build_history_vision_pass_fail_summary_query,
    _vision_format_value,
)

import logging
logger = logging.getLogger(__name__)


bp = Blueprint("historial_vision_pass_fail", __name__)


# ---------------------------------------------------------------------------
# Render template
# ---------------------------------------------------------------------------


@bp.route("/historial_vision_pass_fail/ajax")
@login_requerido
def historial_vision_pass_fail_ajax():
    """Render canonico del template Historial Vision Pass/Fail."""
    try:
        return render_template("Control de resultados/history_vision_pass_fail.html")
    except Exception as e:
        logger.error(f"Error al cargar Historial Vision % Pass/Fail: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@bp.route("/historial-vision-pass-fail")
@bp.route("/historial-vision-pass-fail-ajax")
def alias_legacy_vision_pass_fail():
    """Aliases 301 -> /historial_vision_pass_fail/ajax."""
    return redirect("/historial_vision_pass_fail/ajax", code=301)


# ---------------------------------------------------------------------------
# APIs
# ---------------------------------------------------------------------------


@bp.route("/api/vision/pass-fail-summary")
@login_requerido
def vision_pass_fail_summary_api():
    """Obtener resumen agrupado por linea y numero de parte para history_vision."""
    import traceback
    try:
        sql, params = _build_history_vision_pass_fail_summary_query()
        rows = execute_query(sql, params, fetch="all") or []

        result = []
        for row in rows:
            total = int(row.get("total") or 0)
            ok_count = int(row.get("ok_count") or 0)
            ng_count = int(row.get("ng_count") or 0)
            porcentaje_ok = round((ok_count / total) * 100, 2) if total else 0
            porcentaje_ng = round((ng_count / total) * 100, 2) if total else 0

            result.append({
                "linea": _vision_format_value(row.get("linea", "")) or "",
                "numero_parte": _vision_format_value(row.get("numero_parte", "")) or "",
                "total": total,
                "ok_count": ok_count,
                "ng_count": ng_count,
                "porcentaje_ok": porcentaje_ok,
                "porcentaje_ng": porcentaje_ng,
            })

        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@bp.route("/api/vision/pass-fail-summary/export")
@login_requerido
def export_vision_pass_fail_summary_excel():
    """Exportar resumen Pass/Fail de Vision a un archivo de Excel."""
    import traceback
    try:
        sql, params = _build_history_vision_pass_fail_summary_query()
        rows = execute_query(sql, params, fetch="all") or []

        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Vision Pass Fail"

        header_fill = PatternFill(start_color="3f6b6e", end_color="3f6b6e", fill_type="solid")
        cell_fill = PatternFill(start_color="a1a09c", end_color="a1a09c", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        headers = [
            "Linea", "Numero de parte", "Total", "OK", "NG",
            "% Pass", "% Fail", "PORCENTAJE",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_idx, row in enumerate(rows, start=2):
            total = int(row.get("total") or 0)
            ok_count = int(row.get("ok_count") or 0)
            ng_count = int(row.get("ng_count") or 0)
            porcentaje_ok = round((ok_count / total) * 100, 2) if total else 0
            porcentaje_ng = round((ng_count / total) * 100, 2) if total else 0

            values = [
                _vision_format_value(row.get("linea", "")) or "",
                _vision_format_value(row.get("numero_parte", "")) or "",
                total,
                ok_count,
                ng_count,
                porcentaje_ok,
                porcentaje_ng,
            ]

            for col_num, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_num, value=value)
                cell.fill = cell_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            image_cell = ws.cell(row=row_idx, column=8, value="")
            image_cell.fill = cell_fill
            image_cell.alignment = Alignment(horizontal="center", vertical="center")
            image_cell.border = border

            excel_image = _create_vision_pass_fail_excel_image(porcentaje_ok, porcentaje_ng)
            ws.add_image(excel_image, f"H{row_idx}")
            ws.row_dimensions[row_idx].height = 24

        column_widths = [20, 28, 14, 12, 12, 14, 14, 58]
        for col_num, width in enumerate(column_widths, start=1):
            column_letter = ws.cell(row=1, column=col_num).column_letter
            ws.column_dimensions[column_letter].width = width

        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"historial_vision_pass_fail_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return _send_excel_download(output, filename)
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500
