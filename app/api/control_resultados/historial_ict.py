"""Endpoints HTTP del modulo "Historial de maquina ICT".

Consumido por LISTA_DE_CONTROL_DE_RESULTADOS / Historial de maquinas calidad.
JS cliente: app/static/js/ict.js
Template:   app/templates/Control de resultados/history_ict.html

Rutas:
  GET  /historial_ict/ajax       -> render template (canonica)
  GET  /api/ict/data             -> listar registros con filtros + paginacion
  GET  /api/ict/defects          -> defectos asociados a un barcode
  GET  /api/ict/export           -> exportar listado filtrado a Excel
  GET  /api/ict/export-defects   -> exportar parametros de un barcode a Excel
  POST /api/ict/export-compare   -> comparar parametros entre 2+ ejecuciones

Alias 2026-05-27:
  GET /historial-ict             -> 301 a /historial_ict/ajax
  GET /historial-ict-ajax        -> 301 a /historial_ict/ajax
  GET /ict/front-full-defects2   -> 301 a /historial_ict/ajax (path historico)

Migrado desde app/routes.py el 2026-05-27. Las 5 APIs ya tenian
@login_requerido en el original.

Helpers compartidos con los otros 2 modulos ICT viven en
app/api/shared/ict_helpers.py.
"""

from datetime import datetime

from flask import Blueprint, jsonify, redirect, render_template, request, send_file

from app.api.shared import excel_response_ict, execute_query, login_requerido
from app.api.shared.ict_helpers import (
    _append_indexable_text_filter,
    _ict_attach_operator,
    _ict_format_row,
    _ict_load_local_parameters,
)
from app.services.ict_lgd_parser import (
    IctLgdError,
    IctLgdNotFoundError,
    IctLgdPathError,
)

import logging
logger = logging.getLogger(__name__)


bp = Blueprint("historial_ict", __name__)


_ICT_HISTORY_COLUMN_FILTER_SQL = {
    "fecha": "CAST(fecha AS CHAR) LIKE %s",
    "hora": "CAST(TIME(ts) AS CHAR) LIKE %s",
    "linea": "linea LIKE %s",
    "ict": "CAST(ict AS CHAR) LIKE %s",
    "resultado": "resultado LIKE %s",
    "no_parte": "no_parte LIKE %s",
    "barcode": "barcode LIKE %s",
    "fuente_archivo": "fuente_archivo LIKE %s",
    "defect_code": "defect_code LIKE %s",
    "defect_valor": "defect_valor LIKE %s",
}


def _append_ict_history_column_filters(sql, params):
    """Agrega filtros de encabezado permitidos sin interpolar nombres externos."""
    for key, clause in _ICT_HISTORY_COLUMN_FILTER_SQL.items():
        value = request.args.get(f"cf_{key}", "").strip()
        if value:
            sql += f" AND {clause}"
            params.append(f"%{value}%")
    return sql


# ---------------------------------------------------------------------------
# Render template
# ---------------------------------------------------------------------------


@bp.route("/historial_ict/ajax")
@login_requerido
def historial_ict_ajax():
    """Render canonico del template Historial ICT (defectos)."""
    try:
        return render_template("Control de resultados/history_ict.html")
    except Exception as e:
        logger.error(f"Error al cargar Historial ICT: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@bp.route("/historial-ict")
@bp.route("/historial-ict-ajax")
@bp.route("/ict/front-full-defects2")
def alias_legacy_historial_ict():
    """Aliases 301 -> /historial_ict/ajax."""
    return redirect("/historial_ict/ajax", code=301)


# ---------------------------------------------------------------------------
# APIs
# ---------------------------------------------------------------------------


@bp.route("/api/ict/data")
@login_requerido
def ict_data_api():
    """Obtener registros del historial ICT con filtros y paginacion.

    Soporta `fecha` (igualdad, retro-compatible) o `fecha_desde`/`fecha_hasta`
    (rango). Si se envian ambos, prevalece el rango.

    Paginacion: `page` (1-based) y `per_page` (default 1000, max 1000).
    Cuando se envia `page`, la respuesta es un objeto con metadata; si no se
    envia, se devuelve el array plano (retro-compatible) con LIMIT 500.
    """
    try:
        fecha = request.args.get("fecha", "").strip()
        fecha_desde = request.args.get("fecha_desde", "").strip()
        fecha_hasta = request.args.get("fecha_hasta", "").strip()
        no_parte = request.args.get("no_parte", "").strip()
        linea = request.args.get("linea", "").strip()
        ict_filter = request.args.get("ict", "").strip()
        resultado = request.args.get("resultado", "").strip()
        barcode_like = request.args.get("barcode_like", "").strip()
        page_raw = request.args.get("page", "").strip()
        per_page_raw = request.args.get("per_page", "").strip()
        paginated = bool(page_raw)
        if barcode_like and len(barcode_like) < 6:
            if paginated:
                return jsonify({"rows": [], "total": 0, "page": 1, "per_page": 0, "total_pages": 0})
            return jsonify([])

        where_sql = "WHERE 1=1"
        params = []

        def _add(clause, *vals):
            nonlocal where_sql
            where_sql += " " + clause
            params.extend(vals)

        if fecha_desde or fecha_hasta:
            if fecha_desde:
                _add("AND fecha>=%s", fecha_desde)
            if fecha_hasta:
                _add("AND fecha<=%s", fecha_hasta)
        elif fecha:
            _add("AND fecha=%s", fecha)
        if no_parte:
            _add("AND no_parte LIKE %s", f"{no_parte}%")
        if linea:
            _add("AND linea=%s", linea)
        if ict_filter:
            try:
                _add("AND ict=%s", int(ict_filter))
            except ValueError:
                return jsonify({"error": "ict debe ser numerico"}), 400
        if resultado:
            _add("AND resultado=%s", resultado)
        if barcode_like:
            # Reaplicar la misma logica de _append_indexable_text_filter pero
            # sobre el WHERE acumulado (no sobre el SELECT completo).
            value = barcode_like.strip()
            if len(value) >= 12:
                where_sql += " AND barcode=%s"
                params.append(value)
            else:
                where_sql += " AND barcode LIKE %s"
                params.append(f"{value}%")

        where_sql = _append_ict_history_column_filters(where_sql, params)

        select_cols = (
            "SELECT fecha, TIME(ts) AS hora, linea, ict, resultado, no_parte, barcode, "
            "ts, fuente_archivo, defect_code, defect_valor "
            "FROM history_ict "
        )

        if paginated:
            try:
                page = max(1, int(page_raw))
            except ValueError:
                page = 1
            try:
                per_page = int(per_page_raw) if per_page_raw else 1000
            except ValueError:
                per_page = 1000
            per_page = max(1, min(per_page, 1000))

            count_sql = "SELECT COUNT(*) AS n FROM history_ict " + where_sql
            count_row = execute_query(count_sql, tuple(params), fetch="one") or {}
            total = int(count_row.get("n", 0))

            offset = (page - 1) * per_page
            data_sql = select_cols + where_sql + " ORDER BY ts DESC LIMIT %s OFFSET %s"
            rows = execute_query(data_sql, tuple(params) + (per_page, offset), fetch="all") or []
            _ict_attach_operator(rows)

            total_pages = (total + per_page - 1) // per_page if per_page else 0
            return jsonify({
                "rows": [_ict_format_row(row) for row in rows],
                "total": total,
                "page": page,
                "per_page": per_page,
                "total_pages": total_pages,
            })

        # Modo legacy sin paginacion (retro-compatible)
        data_sql = select_cols + where_sql + " ORDER BY ts DESC LIMIT 500"
        rows = execute_query(data_sql, tuple(params), fetch="all") or []
        _ict_attach_operator(rows)
        return jsonify([_ict_format_row(row) for row in rows])
    except Exception as e:
        logger.exception("Error en endpoint ICT")
        return jsonify({"error": str(e)}), 500


@bp.route("/api/ict/defects")
@login_requerido
def ict_defects_api():
    """Obtener defectos asociados a un barcode especifico."""
    barcode = request.args.get("barcode", "").strip()
    ts = request.args.get("ts", "").strip()
    if not barcode:
        return jsonify([])

    try:
        rows, _ = _ict_load_local_parameters(barcode, ts)
        return jsonify([_ict_format_row(row) for row in rows])
    except IctLgdNotFoundError as e:
        return jsonify({"error": str(e)}), 404
    except IctLgdPathError as e:
        return jsonify({"error": str(e)}), 400
    except IctLgdError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        logger.exception("Error en endpoint ICT")
        return jsonify({"error": str(e)}), 500


@bp.route("/api/ict/export")
@login_requerido
def export_ict_excel():
    """Exportar el historial ICT filtrado a un archivo de Excel."""
    try:
        fecha = request.args.get("fecha", "").strip()
        fecha_desde = request.args.get("fecha_desde", "").strip()
        fecha_hasta = request.args.get("fecha_hasta", "").strip()
        no_parte = request.args.get("no_parte", "").strip()
        linea = request.args.get("linea", "").strip()
        ict_filter = request.args.get("ict", "").strip()
        resultado = request.args.get("resultado", "").strip()
        barcode_like = request.args.get("barcode_like", "").strip()
        if barcode_like and len(barcode_like) < 6:
            return jsonify({"error": "Barcode demasiado corto para exportar"}), 400

        sql = (
            "SELECT fecha, TIME(ts) AS hora, linea, ict, resultado, no_parte, barcode, "
            "ts, fuente_archivo, defect_code, defect_valor "
            "FROM history_ict WHERE 1=1"
        )
        params = []

        if fecha_desde or fecha_hasta:
            if fecha_desde:
                sql += " AND fecha>=%s"
                params.append(fecha_desde)
            if fecha_hasta:
                sql += " AND fecha<=%s"
                params.append(fecha_hasta)
        elif fecha:
            sql += " AND fecha=%s"
            params.append(fecha)
        if no_parte:
            sql += " AND no_parte LIKE %s"
            params.append(f"{no_parte}%")
        if linea:
            sql += " AND linea=%s"
            params.append(linea)
        if ict_filter:
            try:
                sql += " AND ict=%s"
                params.append(int(ict_filter))
            except ValueError:
                return jsonify({"error": "ict debe ser numerico"}), 400
        if resultado:
            sql += " AND resultado=%s"
            params.append(resultado)
        if barcode_like:
            sql = _append_indexable_text_filter(sql, params, "barcode", barcode_like)

        sql = _append_ict_history_column_filters(sql, params)

        sql += " ORDER BY ts DESC LIMIT 500"
        rows = execute_query(sql, tuple(params), fetch="all") or []
        _ict_attach_operator(rows)

        items = [_ict_format_row(row) for row in rows]
        headers = [
            "Fecha", "Hora", "Linea", "ICT", "Resultado", "Operador",
            "No Parte", "Barcode", "Fuente", "Defect Code", "Defect Valor",
        ]
        keys = [
            "fecha", "hora", "linea", "ict", "resultado", "operador",
            "no_parte", "barcode", "fuente_archivo", "defect_code", "defect_valor",
        ]
        filename = f"historial_ict_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        return excel_response_ict(
            items, headers, keys, widths=[16] * len(headers),
            sheet="Historial ICT", filename=filename,
        )
    except IctLgdNotFoundError as e:
        return jsonify({"error": str(e)}), 404
    except IctLgdPathError as e:
        return jsonify({"error": str(e)}), 400
    except IctLgdError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        logger.exception("Error en endpoint ICT")
        return jsonify({"error": str(e)}), 500


@bp.route("/api/ict/export-defects")
@login_requerido
def export_ict_defects_excel():
    """Exportar detalles de defectos ICT a un archivo de Excel."""
    barcode = request.args.get("barcode", "").strip()
    ts = request.args.get("ts", "").strip()
    resultado_filter = request.args.get("resultado", "").strip()

    if not barcode:
        return jsonify({"error": "Barcode requerido"}), 400

    try:
        rows, _ = _ict_load_local_parameters(barcode, ts)
        if resultado_filter:
            rows = [row for row in rows if row.get("resultado_local") == resultado_filter]

        items = []
        for row in rows:
            formatted = _ict_format_row(row)
            hlim = formatted.get("hlim_pct", "")
            llim = formatted.get("llim_pct", "")
            formatted["hlim_fmt"] = f"{hlim}%" if hlim else ""
            formatted["llim_fmt"] = f"{llim}%" if llim else ""
            items.append(formatted)

        headers = [
            "Fecha", "Hora", "Linea", "ICT", "Barcode",
            "Componente", "Pinref", "ACT", "Unit", "STD",
            "Unit", "MEAS", "M", "R", "HLIM",
            "LLIM", "H.P", "L.P", "WS", "DS",
            "RC", "P", "J", "Resultado", "Tipo Defecto",
        ]
        keys = [
            "fecha", "hora", "linea", "ict", "barcode",
            "componente", "pinref", "act_value", "act_unit", "std_value",
            "std_unit", "meas_value", "m_value", "r_value", "hlim_fmt",
            "llim_fmt", "hp_value", "lp_value", "ws_value", "ds_value",
            "rc_value", "p_flag", "j_flag", "resultado_local", "defecto_tipo",
        ]
        filename = f"parametros_{barcode}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        return excel_response_ict(
            items, headers, keys, widths=[12] * len(headers),
            sheet=f"Parametros {barcode[:20]}", filename=filename,
        )
    except IctLgdNotFoundError as e:
        return jsonify({"error": str(e)}), 404
    except IctLgdPathError as e:
        return jsonify({"error": str(e)}), 400
    except IctLgdError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        logger.exception("Error en endpoint ICT")
        return jsonify({"error": str(e)}), 500


@bp.route("/api/ict/export-compare", methods=["POST"])
@login_requerido
def export_ict_compare_excel():
    """Exportar comparacion de parametros ICT entre varias ejecuciones (auditoria).

    Body JSON: { runs: [{barcode, ts, fecha, hora, linea, resultado}, ...], only_diffs: bool }
    """
    try:
        payload = request.get_json(silent=True) or {}
        runs_input = payload.get("runs") or []
        only_diffs = bool(payload.get("only_diffs", True))

        if len(runs_input) < 2:
            return jsonify({"error": "Se requieren al menos 2 ejecuciones"}), 400

        runs = []
        for idx, rec in enumerate(runs_input, start=1):
            barcode = (rec.get("barcode") or "").strip()
            ts = (rec.get("ts") or "").strip()
            if not barcode:
                continue
            rows, _ = _ict_load_local_parameters(barcode, ts)
            runs.append({
                "run_index": idx,
                "barcode": barcode,
                "ts": ts,
                "fecha": rec.get("fecha") or "",
                "hora": rec.get("hora") or "",
                "linea": rec.get("linea") or "",
                "resultado": rec.get("resultado") or "",
                "defects": [_ict_format_row(row) for row in rows],
            })

        if len(runs) < 2:
            return jsonify({"error": "No se pudieron cargar parametros de las ejecuciones"}), 400

        compare_fields = [
            ("act_value", "ACT", ""),
            ("act_unit", "UNIT", ""),
            ("std_value", "STD", ""),
            ("std_unit", "UNIT", ""),
            ("m_value", "M", ""),
            ("r_value", "R", ""),
            ("hlim_pct", "HLIM %", "%"),
            ("llim_pct", "LLIM %", "%"),
            ("hp_value", "HP", ""),
            ("lp_value", "LP", ""),
            ("ws_value", "WS", ""),
            ("ds_value", "DS", ""),
            ("rc_value", "RC", ""),
            ("p_flag", "P", ""),
            ("j_flag", "J", ""),
        ]

        groups = {}
        for run in runs:
            for d in run["defects"]:
                key = (d.get("componente") or "", d.get("pinref") or "")
                if key not in groups:
                    groups[key] = {"componente": key[0], "pinref": key[1], "by_run": {}}
                groups[key]["by_run"][run["run_index"]] = d

        def _norm(v):
            if v is None:
                return ""
            return str(v).strip()

        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Comparacion ICT"

        header_fill = PatternFill(start_color="3f6b6e", end_color="3f6b6e", fill_type="solid")
        cell_fill = PatternFill(start_color="a1a09c", end_color="a1a09c", fill_type="solid")
        diff_fill = PatternFill(start_color="f4b3ad", end_color="f4b3ad", fill_type="solid")
        missing_fill = PatternFill(start_color="d9d9d9", end_color="d9d9d9", fill_type="solid")
        info_fill = PatternFill(start_color="dde6e8", end_color="dde6e8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        title_font = Font(bold=True, color="000000", size=12)
        info_font = Font(bold=True, color="000000", size=9)
        diff_font = Font(bold=True, color="8b0000", size=10)
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.cell(row=1, column=1, value="COMPARACION DE PARAMETROS ICT - AUDITORIA").font = title_font
        ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=2, column=2, value=f"Solo diferencias: {'SI' if only_diffs else 'NO'}")

        info_headers = ["#", "Barcode", "Fecha", "Hora", "Linea", "Resultado"]
        for col_num, h in enumerate(info_headers, start=1):
            c = ws.cell(row=4, column=col_num, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

        for i, run in enumerate(runs, start=5):
            for col_num, value in enumerate([
                f"#{run['run_index']}",
                run["barcode"],
                run["fecha"],
                run["hora"],
                run["linea"],
                run["resultado"],
            ], start=1):
                c = ws.cell(row=i, column=col_num, value=value)
                c.fill = info_fill
                c.font = info_font
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border

        table_start_row = 5 + len(runs) + 2
        table_headers = ["Componente", "Pinref", "Ejecucion"] + [f[1] for f in compare_fields]
        for col_num, h in enumerate(table_headers, start=1):
            c = ws.cell(row=table_start_row, column=col_num, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

        sorted_keys = sorted(groups.keys(), key=lambda k: (k[0] or "", k[1] or ""))

        current_row = table_start_row + 1
        for key in sorted_keys:
            group = groups[key]

            diff_keys = set()
            for f in compare_fields:
                values = set()
                for run in runs:
                    d = group["by_run"].get(run["run_index"])
                    if d:
                        values.add(_norm(d.get(f[0])))
                if len(values) > 1:
                    diff_keys.add(f[0])

            missing_in_some = any(run["run_index"] not in group["by_run"] for run in runs)
            has_any_diff = bool(diff_keys) or missing_in_some

            if only_diffs and not has_any_diff:
                continue

            for run in runs:
                d = group["by_run"].get(run["run_index"])
                run_label = f"#{run['run_index']} - {run['barcode']} {run['fecha']} {run['hora']}"
                row_values = [group["componente"], group["pinref"], run_label]

                for f in compare_fields:
                    if not d:
                        row_values.append("--")
                        continue
                    raw = d.get(f[0])
                    if raw is None or raw == "":
                        row_values.append("")
                    else:
                        suffix = f[2]
                        row_values.append(f"{raw}{suffix}" if suffix else raw)

                for col_num, value in enumerate(row_values, start=1):
                    cell = ws.cell(row=current_row, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border

                    if col_num <= 3:
                        cell.fill = cell_fill
                    else:
                        field = compare_fields[col_num - 4]
                        if not d:
                            cell.fill = missing_fill
                        elif field[0] in diff_keys:
                            cell.fill = diff_fill
                            cell.font = diff_font
                        else:
                            cell.fill = cell_fill

                current_row += 1

        widths = [18, 12, 38] + [12] * len(compare_fields)
        for col_idx, width in enumerate(widths, start=1):
            column_letter = ws.cell(row=table_start_row, column=col_idx).column_letter
            ws.column_dimensions[column_letter].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"comparacion_ict_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except IctLgdNotFoundError as e:
        return jsonify({"error": str(e)}), 404
    except IctLgdPathError as e:
        return jsonify({"error": str(e)}), 400
    except IctLgdError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        logger.exception("Error en endpoint ICT")
        return jsonify({"error": str(e)}), 500
