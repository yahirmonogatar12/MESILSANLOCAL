"""Endpoints HTTP del modulo "Historial de maquina ICT % Pass/Fail".

Consumido por LISTA_DE_CONTROL_DE_RESULTADOS / Historial de maquinas calidad.
JS cliente: app/static/js/ict-Pass-Fail.js
Template:   app/templates/Control de resultados/history_ict_Pass_Fail.html

Rutas:
  GET /historial_ict_pass_fail/ajax  -> render template (canonica)
  GET /api/ict/pass-fail             -> resumen Pass/Fail por jornada/turno
  GET /api/ict/pass-fail/detail      -> detalle de barcodes por fila del resumen
  GET /api/ict/pass-fail/export      -> exportar resumen a Excel

Alias 2026-05-27:
  GET /historial-maquina-ict-pass-fail      -> 301 a /historial_ict_pass_fail/ajax
  GET /historial-maquina-ict-pass-fail-ajax -> 301 a /historial_ict_pass_fail/ajax

Migrado desde app/routes.py el 2026-05-27. Las 3 APIs ya tenian
@login_requerido en el original; se conservan.

Helpers compartidos con los otros 2 modulos ICT viven en
app/api/shared/ict_helpers.py.
"""

from datetime import datetime, timedelta
from datetime import time as dt_time

from flask import Blueprint, jsonify, redirect, render_template, request

from app.api.shared import execute_query, login_requerido
from app.api.shared.ict_helpers import _ict_format_row


bp = Blueprint("historial_ict_pass_fail", __name__)


# ---------------------------------------------------------------------------
# Helpers privados del modulo
# ---------------------------------------------------------------------------


def _ict_pass_fail_fecha_jornada_expr(ts_column="ts"):
    return (
        f"CASE WHEN TIME({ts_column}) >= '07:30:00' "
        f"THEN DATE({ts_column}) ELSE DATE(DATE_SUB({ts_column}, INTERVAL 1 DAY)) END"
    )


def _ict_pass_fail_turno_expr(ts_column="ts"):
    return (
        "CASE "
        f"WHEN TIME({ts_column}) >= '07:30:00' AND TIME({ts_column}) < '17:30:00' THEN 'DIA' "
        f"WHEN TIME({ts_column}) >= '17:30:00' AND TIME({ts_column}) < '22:00:00' THEN 'TIEMPO EXTRA' "
        "ELSE 'NOCHE' "
        "END"
    )


def _ict_pass_fail_real_counts(intentos, ok_count, ng_count, fue_reparacion):
    """Clasificar el resultado ICT real de una PCB/barcode.

    Nota: si la pieza fue reparada pero el ICT nunca la marco NG
    (ng_count == 0), se ignora la reparacion y se trata como OK_REAL.
    El ICT no detecta el 100% de las causas y a veces se repara por
    algo que el ICT no evalua, asi que ese caso no se cuenta como
    falso negativo del ICT.
    """
    if fue_reparacion and ng_count > 0:
        return {
            "total_real": 1,
            "ok_real": 0,
            "defectos_detectados": 1,
            "falsos_negativos": 0,
            "falsos_fail": 0,
            "correcto_real": 1,
            "falla_real": 0,
            "criterio_real": "DEFECTO_DETECTADO",
        }

    falsos_fail = ng_count
    return {
        "total_real": intentos,
        "ok_real": ok_count,
        "defectos_detectados": 0,
        "falsos_negativos": 0,
        "falsos_fail": falsos_fail,
        "correcto_real": ok_count,
        "falla_real": falsos_fail,
        "criterio_real": "FALSO_FAIL" if falsos_fail else "OK_REAL",
    }


def _build_history_ict_pass_fail_summary_query():
    """Construir query para resumen Pass/Fail de ICT por jornada, turno, linea y numero de parte."""
    fecha_desde = (
        request.args.get("fecha_desde", "").strip()
        or request.args.get("fecha", "").strip()
    )
    fecha_hasta = request.args.get("fecha_hasta", "").strip() or fecha_desde
    numero_parte = (
        request.args.get("numero_parte", "").strip()
        or request.args.get("no_parte", "").strip()
    )
    turno = request.args.get("turno", "").strip().upper()
    barcode = (
        request.args.get("barcode", "").strip()
        or request.args.get("barcode_like", "").strip()
    )

    fecha_jornada_expr = _ict_pass_fail_fecha_jornada_expr("h.ts")
    turno_expr = _ict_pass_fail_turno_expr("h.ts")
    linea_expr = "COALESCE(NULLIF(TRIM(h.linea), ''), 'SIN LINEA')"
    ict_expr = "COALESCE(h.ict, 0)"
    numero_parte_expr = "COALESCE(NULLIF(TRIM(h.no_parte), ''), 'SIN NUMERO DE PARTE')"

    sql = (
        "SELECT resumen.fecha, resumen.linea, resumen.ict, resumen.turno, resumen.numero_parte, "
        "SUM(resumen.total_real) AS total, "
        "SUM(resumen.correcto_real) AS ok_count, "
        "SUM(resumen.falla_real) AS ng_count, "
        "SUM(resumen.ok_real) AS ok_real, "
        "SUM(resumen.defectos_detectados) AS defectos_detectados, "
        "SUM(resumen.falsos_negativos) AS falsos_negativos, "
        "SUM(resumen.falsos_fail) AS falsos_fail, "
        "SUM(resumen.piezas_con_defecto) AS piezas_con_defecto, "
        "SUM(resumen.piezas_sin_defecto) AS piezas_sin_defecto, "
        "SUM(resumen.intentos_sin_defecto) AS intentos_sin_defecto, "
        "SUM(resumen.intentos) AS total_intentos, "
        "SUM(resumen.ok_count) AS ok_count_raw, "
        "SUM(resumen.ng_count) AS ng_count_raw, "
        "COUNT(*) AS piezas_unicas, "
        "SUM(CASE WHEN resumen.intentos > 1 THEN 1 ELSE 0 END) AS piezas_repetidas, "
        "SUM(resumen.fue_reparacion) AS piezas_reparacion "
        "FROM ("
        " SELECT por_barcode.*, "
        # Reparacion con ng_count = 0 se ignora: se trata como pieza sin
        # defecto (OK_REAL). Solo cuenta como "con defecto" si el ICT la
        # marco NG (defect_count > 0 AND ng_count > 0).
        " CASE WHEN por_barcode.defect_count > 0 AND por_barcode.ng_count > 0 THEN 1 ELSE por_barcode.intentos END AS total_real, "
        " CASE WHEN por_barcode.defect_count > 0 AND por_barcode.ng_count > 0 THEN 0 ELSE por_barcode.ok_count END AS ok_real, "
        " CASE WHEN por_barcode.defect_count > 0 AND por_barcode.ng_count > 0 THEN 1 ELSE 0 END AS defectos_detectados, "
        " 0 AS falsos_negativos, "
        " CASE WHEN NOT (por_barcode.defect_count > 0 AND por_barcode.ng_count > 0) THEN por_barcode.ng_count ELSE 0 END AS falsos_fail, "
        " CASE WHEN por_barcode.defect_count > 0 AND por_barcode.ng_count > 0 THEN 1 ELSE 0 END AS piezas_con_defecto, "
        " CASE WHEN por_barcode.defect_count > 0 AND por_barcode.ng_count > 0 THEN 0 ELSE 1 END AS piezas_sin_defecto, "
        " CASE WHEN NOT (por_barcode.defect_count > 0 AND por_barcode.ng_count > 0) THEN por_barcode.intentos ELSE 0 END AS intentos_sin_defecto, "
        " CASE "
        "  WHEN por_barcode.defect_count > 0 AND por_barcode.ng_count > 0 THEN 1 "
        "  ELSE por_barcode.ok_count "
        " END AS correcto_real, "
        " CASE "
        "  WHEN por_barcode.defect_count > 0 AND por_barcode.ng_count > 0 THEN 0 "
        "  ELSE por_barcode.ng_count "
        " END AS falla_real "
        " FROM ("
        "  SELECT "
        f"{fecha_jornada_expr} AS fecha, "
        f"{linea_expr} AS linea, "
        f"{ict_expr} AS ict, "
        f"{turno_expr} AS turno, "
        f"{numero_parte_expr} AS numero_parte, "
        "h.barcode AS barcode, "
        "COUNT(*) AS intentos, "
        "SUM(CASE WHEN UPPER(COALESCE(h.resultado, '')) = 'OK' THEN 1 ELSE 0 END) AS ok_count, "
        "SUM(CASE WHEN UPPER(COALESCE(h.resultado, '')) = 'NG' THEN 1 ELSE 0 END) AS ng_count, "
        "COALESCE(MAX(reparacion.defect_count), 0) AS defect_count, "
        "CASE WHEN MAX(reparacion.defect_count) IS NULL THEN 0 ELSE 1 END AS fue_reparacion "
        "  FROM history_ict h "
        "  LEFT JOIN ("
        "   SELECT codigo, COUNT(*) AS defect_count "
        "   FROM defect_data GROUP BY codigo"
        "  ) reparacion ON reparacion.codigo = h.barcode "
        "  WHERE 1=1"
    )
    params = []

    if fecha_desde:
        start_date = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
        params.append(datetime.combine(start_date, dt_time(7, 30)))
        sql += " AND h.ts >= %s"
    if fecha_hasta:
        end_date = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
        params.append(datetime.combine(end_date + timedelta(days=1), dt_time(7, 30)))
        sql += " AND h.ts < %s"
    if numero_parte:
        sql += " AND h.no_parte LIKE %s"
        params.append(f"{numero_parte}%")
    if barcode:
        sql += " AND h.barcode LIKE %s"
        params.append(f"{barcode}%")
    if turno in {"DIA", "TIEMPO EXTRA", "NOCHE"}:
        sql += f" AND {turno_expr}=%s"
        params.append(turno)

    sql += (
        f" GROUP BY {fecha_jornada_expr},"
        f" {linea_expr},"
        f" {ict_expr},"
        f" {turno_expr},"
        f" {numero_parte_expr},"
        " h.barcode"
        " ) por_barcode"
        ") resumen "
        "GROUP BY resumen.fecha, resumen.linea, resumen.ict, resumen.turno, resumen.numero_parte "
        "ORDER BY resumen.fecha ASC, resumen.linea ASC, resumen.ict ASC, "
        "FIELD(resumen.turno, 'DIA', 'TIEMPO EXTRA', 'NOCHE'), "
        "total DESC, resumen.numero_parte ASC"
    )

    return sql, tuple(params) if params else None


# ---------------------------------------------------------------------------
# Render template
# ---------------------------------------------------------------------------


@bp.route("/historial_ict_pass_fail/ajax")
@login_requerido
def historial_ict_pass_fail_ajax():
    """Render canonico del template Historial ICT Pass/Fail."""
    try:
        return render_template("Control de resultados/history_ict_Pass_Fail.html")
    except Exception as e:
        print(f"Error al cargar Historial maquina ICT % Pass/Fail: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@bp.route("/historial-maquina-ict-pass-fail")
@bp.route("/historial-maquina-ict-pass-fail-ajax")
def alias_legacy_pass_fail():
    """Aliases 301 -> /historial_ict_pass_fail/ajax."""
    return redirect("/historial_ict_pass_fail/ajax", code=301)


# ---------------------------------------------------------------------------
# APIs
# ---------------------------------------------------------------------------


@bp.route("/api/ict/pass-fail")
@login_requerido
def ict_pass_fail_api():
    """Obtener resumen Pass/Fail de ICT con el mismo formato que Vision."""
    try:
        sql, params = _build_history_ict_pass_fail_summary_query()
        rows = execute_query(sql, params, fetch="all") or []

        result = []
        for row in rows:
            total = int(row.get("total") or 0)
            ok_count = int(row.get("ok_count") or 0)
            ng_count = int(row.get("ng_count") or 0)
            porcentaje_ok = round((ok_count / total) * 100, 2) if total else 0
            porcentaje_ng = round((ng_count / total) * 100, 2) if total else 0
            piezas_con_defecto = int(row.get("piezas_con_defecto") or 0)
            defectos_detectados = int(row.get("defectos_detectados") or 0)
            falsos_negativos = int(row.get("falsos_negativos") or 0)
            intentos_sin_defecto = int(row.get("intentos_sin_defecto") or 0)
            falsos_fail = int(row.get("falsos_fail") or 0)
            porcentaje_deteccion = (
                round((defectos_detectados / piezas_con_defecto) * 100, 2)
                if piezas_con_defecto else 0
            )
            porcentaje_falso_negativo = (
                round((falsos_negativos / piezas_con_defecto) * 100, 2)
                if piezas_con_defecto else 0
            )
            porcentaje_falso_fail = (
                round((falsos_fail / intentos_sin_defecto) * 100, 2)
                if intentos_sin_defecto else 0
            )

            result.append({
                "fecha": _ict_format_row({"fecha": row.get("fecha")}).get("fecha", ""),
                "linea": row.get("linea", "") or "",
                "ict": row.get("ict", "") or "",
                "turno": row.get("turno", "") or "",
                "numero_parte": row.get("numero_parte", "") or "",
                "total": total,
                "ok_count": ok_count,
                "ng_count": ng_count,
                "total_real": total,
                "ok_real": int(row.get("ok_real") or 0),
                "defectos_detectados": defectos_detectados,
                "falsos_negativos": falsos_negativos,
                "falsos_fail": falsos_fail,
                "correcto_real": ok_count,
                "falla_real": ng_count,
                "ng_real": ng_count,
                "total_intentos": int(row.get("total_intentos") or 0),
                "ok_count_raw": int(row.get("ok_count_raw") or 0),
                "ng_count_raw": int(row.get("ng_count_raw") or 0),
                "piezas_unicas": int(row.get("piezas_unicas") or 0),
                "piezas_repetidas": int(row.get("piezas_repetidas") or 0),
                "piezas_reparacion": int(row.get("piezas_reparacion") or 0),
                "piezas_con_defecto": piezas_con_defecto,
                "piezas_sin_defecto": int(row.get("piezas_sin_defecto") or 0),
                "intentos_sin_defecto": intentos_sin_defecto,
                "porcentaje_ok": porcentaje_ok,
                "porcentaje_ng": porcentaje_ng,
                "porcentaje_deteccion": porcentaje_deteccion,
                "porcentaje_falso_negativo": porcentaje_falso_negativo,
                "porcentaje_falso_fail": porcentaje_falso_fail,
            })

        return jsonify(result)
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@bp.route("/api/ict/pass-fail/detail")
@login_requerido
def ict_pass_fail_detail_api():
    """Obtener detalle de barcodes para una fila del resumen ICT Pass/Fail."""
    try:
        fecha = request.args.get("fecha", "").strip()
        linea = request.args.get("linea", "").strip()
        ict = request.args.get("ict", "").strip()
        turno = request.args.get("turno", "").strip().upper()
        numero_parte = request.args.get("numero_parte", "").strip()
        barcode = request.args.get("barcode", "").strip()

        try:
            min_intentos = max(1, int(request.args.get("min_intentos", "").strip() or 1))
        except ValueError:
            min_intentos = 1

        if not all([fecha, linea, ict, turno, numero_parte]):
            return jsonify({"error": "Faltan filtros para consultar el detalle."}), 400

        if turno not in {"DIA", "TIEMPO EXTRA", "NOCHE"}:
            return jsonify({"error": "Turno invalido para consultar detalle."}), 400

        fecha_jornada = datetime.strptime(fecha, "%Y-%m-%d").date()
        fecha_inicio = datetime.combine(fecha_jornada, dt_time(7, 30))
        fecha_fin = datetime.combine(fecha_jornada + timedelta(days=1), dt_time(7, 30))

        linea_expr = "COALESCE(NULLIF(TRIM(h.linea), ''), 'SIN LINEA')"
        ict_expr = "COALESCE(h.ict, 0)"
        numero_parte_expr = "COALESCE(NULLIF(TRIM(h.no_parte), ''), 'SIN NUMERO DE PARTE')"
        turno_expr = _ict_pass_fail_turno_expr("h.ts")

        sql = (
            "SELECT detalle.barcode, detalle.numero_parte, detalle.linea, "
            "detalle.ict, detalle.turno, detalle.primer_test, detalle.ultimo_test, "
            "detalle.intentos, detalle.ok_count, detalle.ng_count, "
            "detalle.resultado_primer, detalle.resultado_final, "
            "COALESCE(reparacion.defect_count, 0) AS defect_count, "
            "CASE WHEN reparacion.defect_count IS NULL THEN 0 ELSE 1 END AS fue_reparacion, "
            "reparacion.primera_reparacion, "
            "COALESCE(reparacion.defectos, '') AS defectos "
            "FROM ("
            " SELECT h.barcode AS barcode, "
            f" MIN({numero_parte_expr}) AS numero_parte, "
            f" MIN({linea_expr}) AS linea, "
            f" MIN({ict_expr}) AS ict, "
            f" MIN({turno_expr}) AS turno, "
            " MIN(h.ts) AS primer_test, "
            " MAX(h.ts) AS ultimo_test, "
            " COUNT(*) AS intentos, "
            " SUM(CASE WHEN UPPER(COALESCE(h.resultado, '')) = 'OK' THEN 1 ELSE 0 END) AS ok_count, "
            " SUM(CASE WHEN UPPER(COALESCE(h.resultado, '')) = 'NG' THEN 1 ELSE 0 END) AS ng_count, "
            " SUBSTRING_INDEX(GROUP_CONCAT(UPPER(COALESCE(h.resultado, '')) ORDER BY h.ts ASC, h.id ASC SEPARATOR ','), ',', 1) AS resultado_primer, "
            " SUBSTRING_INDEX(GROUP_CONCAT(UPPER(COALESCE(h.resultado, '')) ORDER BY h.ts DESC, h.id DESC SEPARATOR ','), ',', 1) AS resultado_final "
            " FROM history_ict h "
            " WHERE h.ts >= %s AND h.ts < %s "
            f" AND {linea_expr}=%s "
            f" AND {ict_expr}=%s "
            f" AND {turno_expr}=%s "
            f" AND {numero_parte_expr}=%s "
        )
        params = [fecha_inicio, fecha_fin, linea, ict, turno, numero_parte]

        if barcode:
            sql += " AND h.barcode LIKE %s "
            params.append(f"{barcode}%")

        sql += (
            " GROUP BY h.barcode "
            " HAVING COUNT(*) >= %s "
            ") detalle "
            "LEFT JOIN ("
            " SELECT codigo, COUNT(*) AS defect_count, MIN(fecha) AS primera_reparacion, "
            " GROUP_CONCAT("
            "  DISTINCT NULLIF("
            "   CONCAT_WS(', ', NULLIF(TRIM(defecto), ''), NULLIF(TRIM(ubicacion), '')), "
            "   ''"
            "  ) ORDER BY defecto, ubicacion SEPARATOR ', '"
            " ) AS defectos "
            " FROM defect_data GROUP BY codigo"
            ") reparacion ON reparacion.codigo = detalle.barcode "
            "ORDER BY detalle.intentos DESC, detalle.ultimo_test DESC, detalle.barcode ASC"
        )
        params.append(min_intentos)

        rows = execute_query(sql, tuple(params), fetch="all") or []
        formatted_rows = []
        for row in rows:
            formatted = _ict_format_row(row)
            intentos = int(formatted.get("intentos") or 0)
            ok_count = int(formatted.get("ok_count") or 0)
            ng_count = int(formatted.get("ng_count") or 0)
            defect_count = int(formatted.get("defect_count") or 0)
            fue_reparacion = bool(int(formatted.get("fue_reparacion") or 0))
            resultado_primer = (formatted.get("resultado_primer") or "").upper()
            resultado_final = (formatted.get("resultado_final") or "").upper()
            real_counts = _ict_pass_fail_real_counts(
                intentos,
                ok_count,
                ng_count,
                fue_reparacion,
            )

            formatted["intentos"] = intentos
            formatted["ok_count"] = ok_count
            formatted["ng_count"] = ng_count
            formatted["defect_count"] = defect_count
            formatted["fue_reparacion"] = fue_reparacion
            formatted["resultado_primer"] = resultado_primer
            formatted["resultado_final"] = resultado_final
            formatted.update(real_counts)
            formatted["ng_real"] = real_counts["falla_real"]
            formatted["pass_real"] = real_counts["correcto_real"]
            formatted_rows.append(formatted)

        total_intentos = sum(row["intentos"] for row in formatted_rows)
        ok_total = sum(row["ok_count"] for row in formatted_rows)
        ng_total = sum(row["ng_count"] for row in formatted_rows)
        total_real = sum(row["total_real"] for row in formatted_rows)
        ok_real = sum(row["ok_real"] for row in formatted_rows)
        defectos_detectados = sum(row["defectos_detectados"] for row in formatted_rows)
        falsos_negativos = sum(row["falsos_negativos"] for row in formatted_rows)
        falsos_fail = sum(row["falsos_fail"] for row in formatted_rows)
        correcto_real = sum(row["correcto_real"] for row in formatted_rows)
        ng_real = sum(row["falla_real"] for row in formatted_rows)
        piezas_unicas = len(formatted_rows)
        piezas_repetidas = sum(1 for row in formatted_rows if row["intentos"] > 1)
        piezas_reparacion = sum(1 for row in formatted_rows if row["fue_reparacion"])
        piezas_con_defecto = sum(1 for row in formatted_rows if row["fue_reparacion"])
        piezas_sin_defecto = piezas_unicas - piezas_con_defecto
        intentos_sin_defecto = sum(
            row["intentos"] for row in formatted_rows if not row["fue_reparacion"]
        )
        porcentaje_pass_real = (
            round((correcto_real / total_real) * 100, 2) if total_real else 0
        )
        porcentaje_fail_real = (
            round((ng_real / total_real) * 100, 2) if total_real else 0
        )
        porcentaje_deteccion = (
            round((defectos_detectados / piezas_con_defecto) * 100, 2)
            if piezas_con_defecto else 0
        )
        porcentaje_falso_negativo = (
            round((falsos_negativos / piezas_con_defecto) * 100, 2)
            if piezas_con_defecto else 0
        )
        porcentaje_falso_fail = (
            round((falsos_fail / intentos_sin_defecto) * 100, 2)
            if intentos_sin_defecto else 0
        )

        return jsonify({
            "summary": {
                "fecha": fecha,
                "linea": linea,
                "ict": ict,
                "turno": turno,
                "numero_parte": numero_parte,
                "total_intentos": total_intentos,
                "ok_total": ok_total,
                "ng_total": ng_total,
                "total_real": total_real,
                "ok_real": ok_real,
                "defectos_detectados": defectos_detectados,
                "falsos_negativos": falsos_negativos,
                "falsos_fail": falsos_fail,
                "correcto_real": correcto_real,
                "ng_real": ng_real,
                "piezas_unicas": piezas_unicas,
                "piezas_repetidas": piezas_repetidas,
                "piezas_reparacion": piezas_reparacion,
                "piezas_con_defecto": piezas_con_defecto,
                "piezas_sin_defecto": piezas_sin_defecto,
                "intentos_sin_defecto": intentos_sin_defecto,
                "pass_real": correcto_real,
                "porcentaje_pass_real": porcentaje_pass_real,
                "porcentaje_fail_real": porcentaje_fail_real,
                "porcentaje_deteccion": porcentaje_deteccion,
                "porcentaje_falso_negativo": porcentaje_falso_negativo,
                "porcentaje_falso_fail": porcentaje_falso_fail,
            },
            "rows": formatted_rows,
        })
    except ValueError:
        return jsonify({"error": "Fecha invalida para consultar detalle."}), 400
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@bp.route("/api/ict/pass-fail/export")
@login_requerido
def ict_pass_fail_export():
    """Exportar resumen Pass/Fail de ICT a Excel.

    Respeta el modo activo en pantalla (?modo=normal|detallado):
      - normal:    intentos crudos (Intentos, OK, NG, % Pass, % Fail + barra).
      - detallado: metricas reales reclasificadas (Total real, Correctos,
                   OK real, Detectados, F. fail, % Correcto, etc. + barra).
    En detallado se ignora la reparacion cuando el ICT estuvo OK
    (ng_count = 0); por eso ya no hay columna de F. negativos.
    """
    from app.api.shared.excel_helpers import (
        _create_vision_pass_fail_excel_image,
        _send_excel_download,
    )

    try:
        modo = request.args.get("modo", "").strip().lower()
        modo = "detallado" if modo == "detallado" else "normal"

        sql, params = _build_history_ict_pass_fail_summary_query()
        rows = execute_query(sql, params, fetch="all") or []

        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "ICT Pass Fail"

        header_fill = PatternFill(start_color="3f6b6e", end_color="3f6b6e", fill_type="solid")
        cell_fill = PatternFill(start_color="a1a09c", end_color="a1a09c", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        if modo == "detallado":
            headers = [
                "Fecha", "Linea", "ICT", "Turno", "Numero de parte",
                "Total real", "Correctos", "OK real",
                "Detectados", "F. fail",
                "% Correcto", "% Deteccion", "% F. fail",
                "PORCENTAJE",
            ]
            column_widths = [14, 20, 10, 18, 28, 14, 12, 12, 14, 12, 14, 14, 14, 58]
        else:
            headers = [
                "Fecha", "Linea", "ICT", "Turno", "Numero de parte",
                "Intentos", "OK", "NG",
                "% Pass", "% Fail",
                "PORCENTAJE",
            ]
            column_widths = [14, 20, 10, 18, 28, 14, 12, 12, 14, 14, 58]

        bar_col = len(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_idx, row in enumerate(rows, start=2):
            total = int(row.get("total") or 0)
            ok_count = int(row.get("ok_count") or 0)
            total_intentos = int(row.get("total_intentos") or 0)
            ok_count_raw = int(row.get("ok_count_raw") or 0)
            ng_count_raw = int(row.get("ng_count_raw") or 0)
            piezas_con_defecto = int(row.get("piezas_con_defecto") or 0)
            intentos_sin_defecto = int(row.get("intentos_sin_defecto") or 0)
            defectos_detectados = int(row.get("defectos_detectados") or 0)
            falsos_fail = int(row.get("falsos_fail") or 0)

            # % reales (modo detallado)
            porcentaje_correcto = round((ok_count / total) * 100, 2) if total else 0
            porcentaje_deteccion = (
                round((defectos_detectados / piezas_con_defecto) * 100, 2)
                if piezas_con_defecto else 0
            )
            porcentaje_falso_fail = (
                round((falsos_fail / intentos_sin_defecto) * 100, 2)
                if intentos_sin_defecto else 0
            )

            # % crudos de intentos (modo normal). Coinciden con la pantalla.
            porcentaje_pass = (
                round((ok_count_raw / total_intentos) * 100, 2)
                if total_intentos else 0
            )
            porcentaje_fail = (
                round((ng_count_raw / total_intentos) * 100, 2)
                if total_intentos else 0
            )

            fecha_fmt = _ict_format_row({"fecha": row.get("fecha")}).get("fecha", "")
            linea = row.get("linea", "") or ""
            ict = row.get("ict", "") or ""
            turno = row.get("turno", "") or ""
            numero_parte = row.get("numero_parte", "") or ""

            if modo == "detallado":
                values = [
                    fecha_fmt, linea, ict, turno, numero_parte,
                    total,
                    ok_count,
                    int(row.get("ok_real") or 0),
                    defectos_detectados,
                    falsos_fail,
                    porcentaje_correcto,
                    porcentaje_deteccion,
                    porcentaje_falso_fail,
                ]
                bar_ok, bar_ng = porcentaje_correcto, round(100 - porcentaje_correcto, 2)
            else:
                values = [
                    fecha_fmt, linea, ict, turno, numero_parte,
                    total_intentos,
                    ok_count_raw,
                    ng_count_raw,
                    porcentaje_pass,
                    porcentaje_fail,
                ]
                bar_ok, bar_ng = porcentaje_pass, porcentaje_fail

            for col_num, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_num, value=value)
                cell.fill = cell_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            image_cell = ws.cell(row=row_idx, column=bar_col, value="")
            image_cell.fill = cell_fill
            image_cell.alignment = Alignment(horizontal="center", vertical="center")
            image_cell.border = border

            excel_image = _create_vision_pass_fail_excel_image(bar_ok, bar_ng)
            ws.add_image(excel_image, f"{image_cell.column_letter}{row_idx}")
            ws.row_dimensions[row_idx].height = 24

        for col_num, width in enumerate(column_widths, start=1):
            column_letter = ws.cell(row=1, column=col_num).column_letter
            ws.column_dimensions[column_letter].width = width

        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = (
            f"historial_ict_pass_fail_{modo}_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        return _send_excel_download(output, filename)
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500
