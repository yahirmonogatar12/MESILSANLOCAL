"""Modulo web de Inventario Exceso QA.

Consumido por LISTA_DE_CONTROL_DE_RESULTADOS / Control de inventario.
La captura pieza por pieza vive en la app movil; este modulo consulta el
inventario activo y gestiona el cierre mensual con CSV fisico.

WF_001: boton/sidebar en LISTA_DE_CONTROL_DE_RESULTADOS.
WF_002: template/CSS/JS propios con prefijo inventario-exceso-*.
WF_003: blueprint dueno de template, APIs JSON, exportacion y cierre.
WF_004: CSS persistente en MainTemplate.html y garantizado por JS.
"""

import csv
import hashlib
import io
import json
import logging
import re

from flask import Blueprint, Response, jsonify, render_template, request, send_file, session

from app.api.pda.excess_inventory import EXCESS_TABLES, init_excess_inventory_tables
from app.api.pda.shipping import get_db_connection
from app.api.pda.shipping_material import (
    OQC_RELEASE_BOXES_TABLE,
    SHIPPING_TABLES,
    get_dict_cursor,
    normalize_integer,
    normalize_part_number,
    normalize_search,
    serialize_row,
)
from app.api.shared import (
    login_requerido,
    obtener_fecha_hora_mexico,
    requiere_permiso_dropdown,
)

logger = logging.getLogger(__name__)

bp = Blueprint("control_resultados_inventario_exceso", __name__)

PERMISO_PAGINA = "LISTA_DE_CONTROL_DE_RESULTADOS"
PERMISO_SECCION = "Control de inventario"
PERMISO_BOTON = "Inventario Exceso"

_requiere_permiso_inventario_exceso = requiere_permiso_dropdown(
    PERMISO_PAGINA, PERMISO_SECCION, PERMISO_BOTON
)


def _usuario_actual():
    return (
        session.get("nombre_completo")
        or session.get("usuario")
        or session.get("username")
        or "Sistema"
    )


def _json_default(value):
    return str(value)


def _normalizar_header(value):
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _serializar_json(data):
    return json.dumps(data, ensure_ascii=False, sort_keys=True, default=_json_default)


def _hash_rows(rows):
    payload = json.dumps(
        rows,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=_json_default,
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _normalizar_numero(value):
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _resolver_part_number_desde_barcode(raw_value, catalog_part_numbers):
    barcode = normalize_search(raw_value).upper()
    if not barcode:
        return ""

    normalized_catalog = [
        normalize_part_number(part_number)
        for part_number in catalog_part_numbers
        if normalize_part_number(part_number)
    ]
    for part_number in sorted(normalized_catalog, key=len, reverse=True):
        if barcode.startswith(part_number):
            return part_number

    match = re.search(r"[A-Z]{2,5}\d{6,12}", barcode)
    candidate = normalize_part_number(match.group(0)) if match else ""
    return candidate if candidate in set(normalized_catalog) else ""


def _obtener_historial_cierres(cursor, limit=20):
    cursor.execute(
        f"""
        SELECT
          id,
          closure_label,
          closure_month,
          DATE_FORMAT(closed_at, '%%Y-%%m-%%d %%H:%%i:%%s') AS closed_at,
          DATE_FORMAT(created_at, '%%Y-%%m-%%d %%H:%%i:%%s') AS created_at,
          DATE_FORMAT(confirmed_at, '%%Y-%%m-%%d %%H:%%i:%%s') AS confirmed_at,
          created_by,
          confirmed_by,
          status,
          summary_json
        FROM `{EXCESS_TABLES['closure_batches']}`
        ORDER BY COALESCE(confirmed_at, created_at) DESC, id DESC
        LIMIT %s
        """,
        (limit,),
    )
    rows = []
    for row in cursor.fetchall():
        summary = {}
        if row.get("summary_json"):
            try:
                summary = json.loads(row["summary_json"])
            except (TypeError, json.JSONDecodeError):
                summary = {}
        row = serialize_row(row)
        row["summary"] = summary
        row.pop("summary_json", None)
        rows.append(row)
    return rows


def _sincronizar_entradas_exceso(cursor):
    cursor.execute(
        f"""
        INSERT INTO `{EXCESS_TABLES['entries']}` (
          piece_id,
          scan_code,
          raw_code,
          part_number,
          quantity,
          source,
          registered_by_user_id,
          registered_by,
          device_id,
          notes,
          entry_at
        )
        SELECT
          p.id,
          p.scan_code,
          p.raw_code,
          p.part_number,
          p.quantity,
          CASE
            WHEN p.device_id = 'baseline-cierre-mayo'
              OR p.registered_by = 'Carga inicial'
            THEN 'baseline'
            ELSE 'mobile'
          END AS source,
          p.registered_by_user_id,
          p.registered_by,
          p.device_id,
          p.notes,
          p.scanned_at
        FROM `{EXCESS_TABLES['pieces']}` p
        LEFT JOIN `{EXCESS_TABLES['entries']}` e
          ON e.piece_id = p.id
        WHERE p.status = 'active'
          AND e.id IS NULL
        """
    )


def _sincronizar_salidas_exceso(cursor):
    lock_name = "qa_exceso_salidas_sync"
    cursor.execute("SELECT GET_LOCK(%s, 0) AS lock_acquired", (lock_name,))
    lock_row = cursor.fetchone() or {}
    lock_acquired = (
        lock_row.get("lock_acquired")
        if isinstance(lock_row, dict)
        else lock_row[0] if lock_row else 0
    )
    if int(lock_acquired or 0) != 1:
        logger.info("Sincronizacion de salidas de exceso ya esta en ejecucion; se omite esta llamada")
        return 0

    try:
        cursor.execute("SET SESSION innodb_lock_wait_timeout = 5")
        cursor.execute(
            f"""
        INSERT INTO `{EXCESS_TABLES['exits']}` (
          piece_id,
          scan_code,
          raw_code,
          part_number,
          box_scan_id,
          box_code,
          oqc_release_box_id,
          oqc_folio,
          oqc_box_quantity,
          oqc_status,
          qc_passed,
          released_at,
          lqc_last_scan,
          source
        )
        SELECT
          p.id,
          p.scan_code,
          p.raw_code,
          p.part_number,
          NULL AS box_scan_id,
          b.box_code,
          o.id,
          o.oqc_folio,
          o.quantity,
          o.status,
          o.qc_passed,
          o.released_at,
          b.last_scan,
          'oqc_lqc_match'
        FROM `{EXCESS_TABLES['pieces']}` p
        STRAIGHT_JOIN `box_scans` b FORCE INDEX (idx_box_scans_serial)
          ON b.serial = p.scan_code
        STRAIGHT_JOIN `{OQC_RELEASE_BOXES_TABLE}` o FORCE INDEX (idx_oqc_release_box_code)
          ON o.box_code = b.box_code COLLATE utf8mb4_0900_ai_ci
         AND o.part_number = p.part_number COLLATE utf8mb4_0900_ai_ci
        LEFT JOIN (
          SELECT c1.*
          FROM `{EXCESS_TABLES['closures']}` c1
          INNER JOIN (
            SELECT
              part_number AS part_number_key,
              MAX(closed_at) AS max_closed_at
            FROM `{EXCESS_TABLES['closures']}`
            GROUP BY part_number
          ) lc
            ON lc.part_number_key = c1.part_number
           AND lc.max_closed_at = c1.closed_at
        ) latest_x
          ON latest_x.part_number = p.part_number
        LEFT JOIN `{EXCESS_TABLES['exits']}` existing_exit
          ON existing_exit.piece_id = p.id
        WHERE existing_exit.id IS NULL
          AND p.status = 'active'
          AND o.status IN ('released', 'received_shipping')
          AND COALESCE(o.qc_passed, 0) = 1
          AND o.released_at IS NOT NULL
          AND p.scanned_at < o.released_at
          AND o.released_at > COALESCE(latest_x.closed_at, '1000-01-01')
          AND o.released_at >= (
            SELECT COALESCE(MIN(closed_at), '1000-01-01')
            FROM `{EXCESS_TABLES['closures']}`
          )
        ON DUPLICATE KEY UPDATE
          scan_code = VALUES(scan_code),
          raw_code = VALUES(raw_code),
          part_number = VALUES(part_number),
          box_code = VALUES(box_code),
          oqc_release_box_id = VALUES(oqc_release_box_id),
          oqc_folio = VALUES(oqc_folio),
          oqc_box_quantity = VALUES(oqc_box_quantity),
          oqc_status = VALUES(oqc_status),
          qc_passed = VALUES(qc_passed),
          released_at = VALUES(released_at),
          lqc_last_scan = VALUES(lqc_last_scan),
          source = VALUES(source)
        """
        )
        return cursor.rowcount
    finally:
        cursor.execute("SELECT RELEASE_LOCK(%s)", (lock_name,))


def _sincronizar_movimientos_exceso(cursor, sync_exits=True):
    _sincronizar_entradas_exceso(cursor)
    if sync_exits:
        _sincronizar_salidas_exceso(cursor)


def _obtener_inventario_actual(cursor, search=None, limit=2000, sync_exits=False):
    where = ["1 = 1"]
    params = []
    search_value = normalize_search(search)
    part_key = "CONVERT({column} USING utf8mb4) COLLATE utf8mb4_unicode_ci"
    _sincronizar_movimientos_exceso(cursor, sync_exits=sync_exits)
    if search_value:
        like = f"%{search_value}%"
        where.append(
            "("
            "i.part_number LIKE %s OR "
            "i.product_model LIKE %s OR "
            "i.customer LIKE %s"
            ")"
        )
        params.extend([like, like, like])

    cursor.execute(
        f"""
        SELECT
          i.part_number,
          i.product_model,
          i.customer,
          COALESCE(latest.initial_quantity, 0) AS initial_quantity,
          COALESCE(entradas.entries_qty, 0) AS entries_qty,
          COALESCE(salidas.exits_qty, 0) AS exits_qty,
          latest.closed_at AS period_start,
          DATE_FORMAT(latest.closed_at, '%%Y-%%m-%%d %%H:%%i:%%s') AS period_start_text,
          latest.closure_label
        FROM `{SHIPPING_TABLES['inventory']}` i
        LEFT JOIN (
          SELECT c1.*
          FROM `{EXCESS_TABLES['closures']}` c1
          INNER JOIN (
            SELECT
              {part_key.format(column='part_number')} AS part_number_key,
              MAX(closed_at) AS max_closed_at
            FROM `{EXCESS_TABLES['closures']}`
            GROUP BY {part_key.format(column='part_number')}
          ) lc
            ON lc.part_number_key = {part_key.format(column='c1.part_number')}
           AND lc.max_closed_at = c1.closed_at
        ) latest
          ON {part_key.format(column='latest.part_number')} = {part_key.format(column='i.part_number')}
        LEFT JOIN (
          SELECT
            e.part_number,
            COALESCE(SUM(e.quantity), 0) AS entries_qty
          FROM `{EXCESS_TABLES['entries']}` e
          INNER JOIN `{EXCESS_TABLES['pieces']}` p
            ON p.id = e.piece_id
          LEFT JOIN (
            SELECT c1.*
            FROM `{EXCESS_TABLES['closures']}` c1
            INNER JOIN (
              SELECT
                {part_key.format(column='part_number')} AS part_number_key,
                MAX(closed_at) AS max_closed_at
              FROM `{EXCESS_TABLES['closures']}`
              GROUP BY {part_key.format(column='part_number')}
            ) lc
              ON lc.part_number_key = {part_key.format(column='c1.part_number')}
             AND lc.max_closed_at = c1.closed_at
          ) latest_p
            ON {part_key.format(column='latest_p.part_number')} = {part_key.format(column='e.part_number')}
          WHERE p.status = 'active'
            AND e.entry_at > COALESCE(latest_p.closed_at, '1000-01-01')
          GROUP BY {part_key.format(column='e.part_number')}, e.part_number
        ) entradas
          ON {part_key.format(column='entradas.part_number')} = {part_key.format(column='i.part_number')}
        LEFT JOIN (
          SELECT
            x.part_number,
            COUNT(DISTINCT x.piece_id) AS exits_qty
          FROM `{EXCESS_TABLES['exits']}` x
          INNER JOIN `{EXCESS_TABLES['pieces']}` p
            ON p.id = x.piece_id
          LEFT JOIN (
            SELECT c1.*
            FROM `{EXCESS_TABLES['closures']}` c1
            INNER JOIN (
              SELECT
                {part_key.format(column='part_number')} AS part_number_key,
                MAX(closed_at) AS max_closed_at
              FROM `{EXCESS_TABLES['closures']}`
              GROUP BY {part_key.format(column='part_number')}
            ) lc
              ON lc.part_number_key = {part_key.format(column='c1.part_number')}
             AND lc.max_closed_at = c1.closed_at
          ) latest_x
            ON {part_key.format(column='latest_x.part_number')} = {part_key.format(column='x.part_number')}
          WHERE p.status = 'active'
            AND x.released_at > COALESCE(latest_x.closed_at, '1000-01-01')
          GROUP BY {part_key.format(column='x.part_number')}, x.part_number
        ) salidas
          ON {part_key.format(column='salidas.part_number')} = {part_key.format(column='i.part_number')}
        WHERE {' AND '.join(where)}
        ORDER BY i.part_number ASC
        LIMIT %s
        """,
        tuple(params + [limit]),
    )
    result_rows = []
    for row in cursor.fetchall():
        initial_quantity = _normalizar_numero(row.get("initial_quantity"))
        entries_qty = _normalizar_numero(row.get("entries_qty"))
        exits_qty = _normalizar_numero(row.get("exits_qty"))
        if initial_quantity == 0 and entries_qty == 0 and exits_qty == 0:
            continue

        current_quantity = initial_quantity + entries_qty - exits_qty
        result_rows.append(
            {
                "part_number": normalize_part_number(row.get("part_number")),
                "product_model": normalize_search(row.get("product_model")),
                "customer": normalize_search(row.get("customer")),
                "initial_quantity": initial_quantity,
                "entries_qty": entries_qty,
                "exits_qty": exits_qty,
                "current_quantity": current_quantity,
                "period_start": serialize_row(row).get("period_start_text") or "",
                "closure_label": normalize_search(row.get("closure_label")),
            }
        )
    return result_rows


def _construir_preview(current_rows, csv_quantities=None):
    csv_quantities = csv_quantities or {}
    rows = []
    matching_rows = 0
    total_system = 0
    total_physical = 0
    absolute_difference = 0

    for row in current_rows:
        part_number = normalize_part_number(row.get("part_number"))
        system_quantity = int(row.get("current_quantity") or 0)
        physical_quantity = csv_quantities.get(part_number)
        difference_quantity = (
            physical_quantity - system_quantity
            if physical_quantity is not None
            else None
        )

        if physical_quantity is not None:
            total_system += system_quantity
            total_physical += physical_quantity
            absolute_difference += abs(difference_quantity or 0)
            if difference_quantity == 0:
                matching_rows += 1

        rows.append(
            {
                "part_number": part_number,
                "product_model": row.get("product_model") or "",
                "customer": row.get("customer") or "",
                "system_quantity": system_quantity,
                "physical_quantity": physical_quantity,
                "difference_quantity": difference_quantity,
                "initial_quantity": physical_quantity,
                "status": (
                    "pendiente"
                    if physical_quantity is None
                    else "igual"
                    if difference_quantity == 0
                    else "diferencia"
                ),
            }
        )

    total_rows = len(rows)
    difference_rows = total_rows - matching_rows if csv_quantities else 0
    accuracy_pct = 0
    if csv_quantities:
        if total_system <= 0:
            accuracy_pct = 100 if total_physical == 0 else 0
        else:
            accuracy_pct = max(0, round((1 - (absolute_difference / total_system)) * 100, 2))

    serializable_rows = [
        {
            "part_number": row["part_number"],
            "system_quantity": row["system_quantity"],
            "physical_quantity": row["physical_quantity"],
            "difference_quantity": row["difference_quantity"],
            "initial_quantity": row["initial_quantity"],
        }
        for row in rows
    ]
    rows_hash = _hash_rows(serializable_rows)

    return {
        "rows": rows,
        "summary": {
            "totalRows": total_rows,
            "matchingRows": matching_rows,
            "differenceRows": difference_rows,
            "systemQuantity": total_system,
            "physicalQuantity": total_physical,
            "absoluteDifference": absolute_difference,
            "accuracyPct": accuracy_pct,
            "rowsHash": rows_hash,
        },
    }


def _parsear_csv_cierre(file_storage, expected_part_numbers):
    raw_bytes = file_storage.read()
    file_hash = hashlib.sha256(raw_bytes).hexdigest()

    try:
        decoded = raw_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            decoded = raw_bytes.decode("latin-1")
        except UnicodeDecodeError as exc:
            return {
                "valid": False,
                "errors": [f"No fue posible leer el CSV: {exc}"],
                "fileHash": file_hash,
                "rows": {},
            }

    reader = csv.reader(io.StringIO(decoded))
    try:
        headers = next(reader)
    except StopIteration:
        return {
            "valid": False,
            "errors": ["El archivo CSV esta vacio."],
            "fileHash": file_hash,
            "rows": {},
        }

    normalized_headers = [_normalizar_header(header) for header in headers]
    if normalized_headers in (["barcode"], ["scan_code"]):
        barcode_idx = 0
        parsed_rows = {
            normalize_part_number(part_number): 0
            for part_number in expected_part_numbers
            if normalize_part_number(part_number)
        }
        errors = []

        for line_number, row in enumerate(reader, start=2):
            if not row or not any(str(cell or "").strip() for cell in row):
                continue
            barcode = normalize_search(row[barcode_idx]).upper()
            if not barcode:
                errors.append(f"Linea {line_number}: BARCODE es obligatorio.")
                continue

            part_number = _resolver_part_number_desde_barcode(
                barcode,
                expected_part_numbers,
            )
            if not part_number:
                errors.append(
                    f"Linea {line_number}: no se pudo relacionar {barcode} con el catalogo de embarques."
                )
                continue
            parsed_rows[part_number] = parsed_rows.get(part_number, 0) + 1

        return {
            "valid": not errors,
            "errors": errors,
            "fileHash": file_hash,
            "rows": parsed_rows,
        }

    if set(normalized_headers) != {"part_number", "current_qty"} or len(normalized_headers) != 2:
        return {
            "valid": False,
            "errors": [
                "El CSV debe contener BARCODE o exactamente las columnas part_number y current_qty."
            ],
            "fileHash": file_hash,
            "rows": {},
        }

    part_idx = normalized_headers.index("part_number")
    qty_idx = normalized_headers.index("current_qty")
    parsed_rows = {}
    errors = []

    for line_number, row in enumerate(reader, start=2):
        if not row or not any(str(cell or "").strip() for cell in row):
            continue
        if len(row) <= max(part_idx, qty_idx):
            errors.append(f"Linea {line_number}: faltan columnas requeridas.")
            continue

        part_number = normalize_part_number(row[part_idx])
        qty_raw = str(row[qty_idx] or "").strip()
        if not part_number:
            errors.append(f"Linea {line_number}: part_number es obligatorio.")
            continue
        if not re.fullmatch(r"\d+", qty_raw):
            errors.append(
                f"Linea {line_number}: current_qty debe ser un entero sin negativos para {part_number}."
            )
            continue
        if part_number in parsed_rows:
            errors.append(f"Linea {line_number}: part_number duplicado {part_number}.")
            continue
        parsed_rows[part_number] = int(qty_raw)

    expected_set = set(expected_part_numbers)
    uploaded_set = set(parsed_rows.keys())
    missing_parts = sorted(expected_set - uploaded_set)
    extra_parts = sorted(uploaded_set - expected_set)

    if missing_parts:
        sample = ", ".join(missing_parts[:8])
        errors.append(
            f"Faltan {len(missing_parts)} numeros de parte del inventario activo. Ejemplo: {sample}"
        )
    if extra_parts:
        sample = ", ".join(extra_parts[:8])
        errors.append(
            f"El CSV contiene {len(extra_parts)} numeros de parte fuera del inventario activo. Ejemplo: {sample}"
        )

    return {
        "valid": not errors,
        "errors": errors,
        "fileHash": file_hash,
        "rows": parsed_rows,
    }


@bp.route("/control_resultados/inventario_exceso")
@login_requerido
@_requiere_permiso_inventario_exceso
def inventario_exceso_ajax():
    try:
        init_excess_inventory_tables()
        return render_template("Control de resultados/inventario_exceso_ajax.html")
    except Exception as exc:
        logger.exception("Error cargando Inventario Exceso")
        return f"Error al cargar el contenido: {exc}", 500


@bp.route("/api/inventario_exceso/inventory", methods=["GET"])
@login_requerido
def api_inventario_exceso_inventory():
    init_excess_inventory_tables()
    conn = get_db_connection()
    cursor = get_dict_cursor(conn)
    try:
        rows = _obtener_inventario_actual(cursor, request.args.get("q"))
        total_quantity = sum(int(row.get("current_quantity") or 0) for row in rows)
        return jsonify(
            {
                "success": True,
                "items": rows,
                "totalParts": len(rows),
                "totalQuantity": total_quantity,
            }
        )
    finally:
        cursor.close()
        conn.close()


@bp.route("/api/inventario_exceso/exits/sync", methods=["POST"])
@login_requerido
def api_inventario_exceso_exits_sync():
    init_excess_inventory_tables()
    conn = get_db_connection()
    cursor = get_dict_cursor(conn)
    try:
        _sincronizar_movimientos_exceso(cursor, sync_exits=True)
        return jsonify({"success": True, "message": "Salidas sincronizadas."})
    finally:
        cursor.close()
        conn.close()


@bp.route("/api/inventario_exceso/exits/detail", methods=["GET"])
@login_requerido
def api_inventario_exceso_exits_detail():
    init_excess_inventory_tables()
    limit = min(max(normalize_integer(request.args.get("limit")) or 500, 1), 2000)
    search = normalize_search(request.args.get("q") or request.args.get("search"))
    part_number = normalize_part_number(request.args.get("partNumber"))
    box_code = normalize_search(request.args.get("boxCode")).upper()

    where = ["p.status = 'active'"]
    params = []

    if part_number:
        where.append("x.part_number = %s")
        params.append(part_number)
    if box_code:
        where.append("x.box_code = %s")
        params.append(box_code)
    if search:
        like = f"%{search}%"
        where.append(
            "("
            "x.scan_code LIKE %s OR "
            "x.part_number LIKE %s OR "
            "x.box_code LIKE %s OR "
            "x.oqc_folio LIKE %s"
            ")"
        )
        params.extend([like, like, like, like])

    conn = get_db_connection()
    cursor = get_dict_cursor(conn)
    try:
        _sincronizar_movimientos_exceso(cursor)
        cursor.execute(
            f"""
            SELECT
              x.piece_id,
              x.part_number,
              x.scan_code,
              x.raw_code,
              x.box_code,
              x.oqc_folio,
              x.oqc_release_box_id,
              x.oqc_box_quantity,
              x.oqc_status,
              x.qc_passed,
              DATE_FORMAT(x.released_at, '%%Y-%%m-%%d %%H:%%i:%%s') AS released_at,
              DATE_FORMAT(x.lqc_last_scan, '%%Y-%%m-%%d %%H:%%i:%%s') AS lqc_last_scan,
              DATE_FORMAT(p.scanned_at, '%%Y-%%m-%%d %%H:%%i:%%s') AS excess_scanned_at
            FROM `{EXCESS_TABLES['exits']}` x
            INNER JOIN `{EXCESS_TABLES['pieces']}` p
              ON p.id = x.piece_id
            WHERE {' AND '.join(where)}
            ORDER BY x.released_at DESC, x.lqc_last_scan DESC, x.piece_id DESC
            LIMIT %s
            """,
            tuple(params + [limit]),
        )
        return jsonify(
            {
                "success": True,
                "items": [serialize_row(row) for row in cursor.fetchall()],
            }
        )
    finally:
        cursor.close()
        conn.close()


@bp.route("/api/inventario_exceso/inventory/export", methods=["GET"])
@login_requerido
def api_inventario_exceso_inventory_export():
    init_excess_inventory_tables()
    conn = get_db_connection()
    cursor = get_dict_cursor(conn)
    try:
        rows = _obtener_inventario_actual(
            cursor,
            request.args.get("search"),
            limit=20000,
            sync_exits=True,
        )

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Inventario Exceso"

        columns = [
            ("No. Parte", "part_number"),
            ("Modelo", "product_model"),
            ("Cliente", "customer"),
            ("Inventario inicial", "initial_quantity"),
            ("Entradas", "entries_qty"),
            ("Salidas", "exits_qty"),
            ("Cantidad total", "current_quantity"),
            ("Inicio de periodo", "period_start"),
            ("Cierre", "closure_label"),
        ]

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx, (label, _) in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, (_, key) in enumerate(columns, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=row.get(key))

        for col_idx, (label, key) in enumerate(columns, start=1):
            values = [str(row.get(key) or "") for row in rows[:200]]
            width = max([len(label), *(len(value) for value in values)]) + 2
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 12), 28)

        workbook.save(output)
        output.seek(0)
        filename = f"inventario_exceso_{obtener_fecha_hora_mexico().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        cursor.close()
        conn.close()


@bp.route("/api/inventario_exceso/movements", methods=["GET"])
@login_requerido
def api_inventario_exceso_movements():
    init_excess_inventory_tables()
    limit = min(max(normalize_integer(request.args.get("limit")) or 200, 1), 1000)
    search = normalize_search(request.args.get("q"))
    from_date = normalize_search(request.args.get("fromDate"))
    to_date = normalize_search(request.args.get("toDate"))
    where = ["1 = 1"]
    params = []
    if search:
        like = f"%{search}%"
        where.append("(m.part_number LIKE %s OR m.scan_code LIKE %s OR m.registered_by LIKE %s)")
        params.extend([like, like, like])
    if from_date:
        where.append("DATE(m.movement_at) >= %s")
        params.append(from_date)
    if to_date:
        where.append("DATE(m.movement_at) <= %s")
        params.append(to_date)

    conn = get_db_connection()
    cursor = get_dict_cursor(conn)
    try:
        cursor.execute(
            f"""
            SELECT
              m.id,
              m.scan_code,
              m.movement_type,
              m.part_number,
              m.quantity,
              m.registered_by_user_id,
              m.registered_by,
              m.device_id,
              m.notes,
              DATE_FORMAT(m.movement_at, '%%Y-%%m-%%d %%H:%%i:%%s') AS movement_at
            FROM `{EXCESS_TABLES['movements']}` m
            WHERE {' AND '.join(where)}
            ORDER BY m.movement_at DESC, m.id DESC
            LIMIT %s
            """,
            tuple(params + [limit]),
        )
        return jsonify(
            {
                "success": True,
                "items": [serialize_row(row) for row in cursor.fetchall()],
            }
        )
    finally:
        cursor.close()
        conn.close()


@bp.route("/api/inventario_exceso/cierre/bootstrap", methods=["GET"])
@login_requerido
def api_inventario_exceso_cierre_bootstrap():
    init_excess_inventory_tables()
    fecha_actual = obtener_fecha_hora_mexico()
    conn = get_db_connection()
    cursor = get_dict_cursor(conn)
    try:
        current_rows = _obtener_inventario_actual(cursor, sync_exits=True)
        preview = _construir_preview(current_rows)
        metadata = {
            "closureDate": fecha_actual.strftime("%Y-%m-%d"),
            "closureDateTime": fecha_actual.strftime("%Y-%m-%d %H:%M:%S"),
            "closureMonth": fecha_actual.strftime("%Y-%m"),
            "closureMonthLabel": fecha_actual.strftime("%Y-%m"),
            "closureLabel": f"Cierre Exceso QA {fecha_actual.strftime('%Y-%m')}",
            "closureUser": _usuario_actual(),
        }
        return jsonify(
            {
                "success": True,
                "valid": False,
                "metadata": metadata,
                "preview": preview,
                "history": _obtener_historial_cierres(cursor),
            }
        )
    finally:
        cursor.close()
        conn.close()


@bp.route("/api/inventario_exceso/cierre/template", methods=["GET"])
@login_requerido
def api_inventario_exceso_cierre_template():
    init_excess_inventory_tables()
    conn = get_db_connection()
    cursor = get_dict_cursor(conn)
    try:
        rows = _obtener_inventario_actual(cursor, sync_exits=True)
        output = io.StringIO()
        writer = csv.writer(output, lineterminator="\n")
        writer.writerow(["part_number", "current_qty"])
        for row in rows:
            writer.writerow([row.get("part_number") or "", row.get("current_quantity") or 0])
        filename = f"plantilla_cierre_inventario_exceso_{obtener_fecha_hora_mexico().strftime('%Y%m%d')}.csv"
        return Response(
            output.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    finally:
        cursor.close()
        conn.close()


@bp.route("/api/inventario_exceso/cierre/preview", methods=["POST"])
@login_requerido
def api_inventario_exceso_cierre_preview():
    init_excess_inventory_tables()
    file_storage = request.files.get("closure_file")
    if not file_storage:
        return jsonify({"success": False, "error": "Se requiere el CSV de inventario fisico."}), 400

    fecha_actual = obtener_fecha_hora_mexico()
    conn = get_db_connection()
    cursor = get_dict_cursor(conn)
    try:
        current_rows = _obtener_inventario_actual(cursor, sync_exits=True)
        if not current_rows:
            return jsonify({"success": False, "error": "No hay inventario activo para cerrar."}), 400

        expected_parts = [row["part_number"] for row in current_rows]
        parsed = _parsear_csv_cierre(file_storage, expected_parts)
        preview = _construir_preview(current_rows, parsed["rows"])
        metadata = {
            "closureDate": fecha_actual.strftime("%Y-%m-%d"),
            "closureDateTime": fecha_actual.strftime("%Y-%m-%d %H:%M:%S"),
            "closureMonth": fecha_actual.strftime("%Y-%m"),
            "closureMonthLabel": fecha_actual.strftime("%Y-%m"),
            "closureLabel": f"Cierre Exceso QA {fecha_actual.strftime('%Y-%m')}",
            "closureUser": _usuario_actual(),
        }

        payload = {
            "metadata": metadata,
            "preview": preview,
            "errors": parsed["errors"],
            "sourceFilename": normalize_search(file_storage.filename),
        }

        batch_id = None
        if parsed["valid"]:
            cursor.execute(
                f"""
                INSERT INTO `{EXCESS_TABLES['closure_batches']}` (
                  closure_label,
                  closure_month,
                  created_by,
                  status,
                  source_filename,
                  file_hash,
                  rows_hash,
                  payload_json,
                  summary_json
                ) VALUES (%s, %s, %s, 'pending', %s, %s, %s, %s, %s)
                """,
                (
                    metadata["closureLabel"],
                    metadata["closureMonth"],
                    metadata["closureUser"],
                    payload["sourceFilename"],
                    parsed["fileHash"],
                    preview["summary"]["rowsHash"],
                    _serializar_json(payload),
                    _serializar_json(preview["summary"]),
                ),
            )
            batch_id = cursor.lastrowid
            conn.commit()

        return jsonify(
            {
                "success": True,
                "valid": parsed["valid"],
                "errors": parsed["errors"],
                "batchId": batch_id,
                "metadata": metadata,
                "preview": preview,
                "history": _obtener_historial_cierres(cursor),
            }
        )
    finally:
        cursor.close()
        conn.close()


@bp.route("/api/inventario_exceso/cierre/confirm", methods=["POST"])
@login_requerido
def api_inventario_exceso_cierre_confirm():
    init_excess_inventory_tables()
    data = request.get_json(silent=True) or {}
    batch_id = normalize_integer(data.get("batchId") or data.get("batch_id"))
    if not batch_id:
        return jsonify({"success": False, "error": "Se requiere batchId."}), 400

    conn = get_db_connection()
    conn.autocommit(False)
    cursor = get_dict_cursor(conn)
    try:
        cursor.execute(
            f"""
            SELECT *
            FROM `{EXCESS_TABLES['closure_batches']}`
            WHERE id = %s
              AND status = 'pending'
            FOR UPDATE
            """,
            (batch_id,),
        )
        batch = cursor.fetchone()
        if not batch:
            conn.rollback()
            return jsonify({"success": False, "error": "El preview ya no esta disponible."}), 404

        payload = json.loads(batch.get("payload_json") or "{}")
        rows = payload.get("preview", {}).get("rows") or []
        if not rows:
            conn.rollback()
            return jsonify({"success": False, "error": "El preview no contiene filas."}), 400

        closed_at = obtener_fecha_hora_mexico()
        closed_by = _usuario_actual()
        closure_label = batch.get("closure_label")

        insert_rows = []
        for row in rows:
            physical_quantity = int(row.get("physical_quantity") or 0)
            insert_rows.append(
                (
                    batch_id,
                    closure_label,
                    row.get("part_number"),
                    row.get("product_model") or None,
                    row.get("customer") or None,
                    int(row.get("system_quantity") or 0),
                    physical_quantity,
                    int(row.get("difference_quantity") or 0),
                    physical_quantity,
                    closed_at,
                    closed_by,
                )
            )

        cursor.executemany(
            f"""
            INSERT INTO `{EXCESS_TABLES['closures']}` (
              closure_batch_id,
              closure_label,
              part_number,
              product_model,
              customer,
              system_quantity,
              physical_quantity,
              difference_quantity,
              initial_quantity,
              closed_at,
              closed_by
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            insert_rows,
        )
        cursor.execute(
            f"""
            UPDATE `{EXCESS_TABLES['closure_batches']}`
            SET status = 'confirmed',
                closed_at = %s,
                confirmed_at = %s,
                confirmed_by = %s
            WHERE id = %s
            """,
            (closed_at, closed_at, closed_by, batch_id),
        )
        conn.commit()

        history = _obtener_historial_cierres(cursor)
        return jsonify(
            {
                "success": True,
                "message": "Cierre de inventario de exceso confirmado.",
                "batchId": batch_id,
                "history": history,
            }
        )
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()


@bp.route("/api/inventario_exceso/cierre/history", methods=["GET"])
@login_requerido
def api_inventario_exceso_cierre_history():
    init_excess_inventory_tables()
    conn = get_db_connection()
    cursor = get_dict_cursor(conn)
    try:
        return jsonify({"success": True, "history": _obtener_historial_cierres(cursor, 50)})
    finally:
        cursor.close()
        conn.close()


@bp.route("/api/inventario_exceso/cierre/history/<int:batch_id>", methods=["GET"])
@login_requerido
def api_inventario_exceso_cierre_history_detail(batch_id):
    init_excess_inventory_tables()
    conn = get_db_connection()
    cursor = get_dict_cursor(conn)
    try:
        cursor.execute(
            f"""
            SELECT *
            FROM `{EXCESS_TABLES['closure_batches']}`
            WHERE id = %s
            LIMIT 1
            """,
            (batch_id,),
        )
        batch = cursor.fetchone()
        if not batch:
            return jsonify({"success": False, "error": "Cierre no encontrado."}), 404
        payload = {}
        if batch.get("payload_json"):
            payload = json.loads(batch["payload_json"])
        return jsonify(
            {
                "success": True,
                "batch": serialize_row(batch),
                "payload": payload,
            }
        )
    finally:
        cursor.close()
        conn.close()


@bp.route("/api/inventario_exceso/cierre/history/<int:batch_id>/export", methods=["GET"])
@login_requerido
def api_inventario_exceso_cierre_history_export(batch_id):
    conn = get_db_connection()
    cursor = get_dict_cursor(conn)
    try:
        cursor.execute(
            f"""
            SELECT
              part_number,
              product_model,
              customer,
              system_quantity,
              physical_quantity,
              difference_quantity,
              initial_quantity,
              DATE_FORMAT(closed_at, '%%Y-%%m-%%d %%H:%%i:%%s') AS closed_at,
              closed_by
            FROM `{EXCESS_TABLES['closures']}`
            WHERE closure_batch_id = %s
            ORDER BY part_number
            """,
            (batch_id,),
        )
        output = io.StringIO()
        writer = csv.writer(output, lineterminator="\n")
        writer.writerow(
            [
                "part_number",
                "product_model",
                "customer",
                "system_quantity",
                "physical_quantity",
                "difference_quantity",
                "initial_quantity",
                "closed_at",
                "closed_by",
            ]
        )
        for row in cursor.fetchall():
            writer.writerow(
                [
                    row.get("part_number"),
                    row.get("product_model"),
                    row.get("customer"),
                    row.get("system_quantity"),
                    row.get("physical_quantity"),
                    row.get("difference_quantity"),
                    row.get("initial_quantity"),
                    row.get("closed_at"),
                    row.get("closed_by"),
                ]
            )
        return Response(
            output.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f"attachment; filename=cierre_inventario_exceso_{batch_id}.csv"
            },
        )
    finally:
        cursor.close()
        conn.close()
