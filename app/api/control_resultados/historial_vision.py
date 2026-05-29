"""Endpoints HTTP del modulo "Historial de maquina vision".

Consumido por LISTA_DE_CONTROL_DE_RESULTADOS / Historial de maquinas calidad.
JS cliente: app/static/js/history_vision.js
Template:   app/templates/Control de resultados/history_vision.html

Rutas:
  GET /historial_vision/ajax    -> render template (canonica)
  GET /api/vision/data          -> listar registros con filtros
  GET /api/vision/image-info    -> metadata de la imagen asociada a un registro
  GET /api/vision/image-file    -> servir imagen resuelta (Result share UNC)
  GET /api/vision/export        -> exportar listado filtrado a Excel

Alias 2026-05-27:
  GET /historial-vision         -> 301 a /historial_vision/ajax
  GET /historial-vision-ajax    -> 301 a /historial_vision/ajax

Migrado desde app/routes.py el 2026-05-27. Las 4 APIs ya tenian
@login_requerido en el original.

Helpers compartidos con historial_vision_pass_fail.py viven en
app/api/shared/vision_helpers.py.
"""

import re
from datetime import datetime
from pathlib import Path

from flask import (
    Blueprint, jsonify, redirect, render_template, request, send_file, url_for,
)

from app.api.shared import execute_query, login_requerido
from app.api.shared.vision_helpers import (
    _build_history_vision_query,
    _get_history_vision_record,
    _resolve_history_vision_image,
    _vision_format_value,
    _vision_is_safe_path,
)

import logging
logger = logging.getLogger(__name__)


bp = Blueprint("historial_vision", __name__)


# ---------------------------------------------------------------------------
# Render template
# ---------------------------------------------------------------------------


@bp.route("/historial_vision/ajax")
@login_requerido
def historial_vision_ajax():
    """Render canonico del template Historial Vision."""
    try:
        return render_template("Control de resultados/history_vision.html")
    except Exception as e:
        logger.error(f"Error al cargar Historial Vision: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@bp.route("/historial-vision")
@bp.route("/historial-vision-ajax")
def alias_legacy_historial_vision():
    """Aliases 301 -> /historial_vision/ajax."""
    return redirect("/historial_vision/ajax", code=301)


# ---------------------------------------------------------------------------
# APIs
# ---------------------------------------------------------------------------


@bp.route("/api/vision/data")
@login_requerido
def vision_data_api():
    """Obtener registros recientes del historial de Vision con filtros opcionales."""
    import traceback
    try:
        sql, params = _build_history_vision_query()
        rows = execute_query(sql, params, fetch="all") or []

        result = []
        for row in rows:
            result.append({
                "id": row.get("id"),
                "linea": _vision_format_value(row.get("linea", "")) or "",
                "fecha": _vision_format_value(row.get("fecha", "")) or "",
                "hora": _vision_format_value(row.get("hora", "")) or "",
                "numero_parte": _vision_format_value(row.get("numero_parte", "")) or "",
                "qr": _vision_format_value(row.get("qr", "")) or "",
                "barcode": _vision_format_value(row.get("barcode", "")) or "",
                "resultado": _vision_format_value(row.get("resultado", "")) or "",
            })

        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@bp.route("/api/vision/image-info")
@login_requerido
def vision_image_info_api():
    """Resolver metadata de la imagen asociada a un registro de history_vision."""
    import traceback
    record_id = request.args.get("id", "").strip()
    if not re.fullmatch(r"\d+", record_id):
        return jsonify({"error": "ID de registro invalido."}), 400

    try:
        record = _get_history_vision_record(int(record_id))
        if not record:
            return jsonify({"error": "Registro de vision no encontrado."}), 404

        resolution = _resolve_history_vision_image(record)
        response_payload = {
            "record_id": record.get("id"),
            "linea": _vision_format_value(record.get("machine_name", "")) or "",
            "numero_parte": _vision_format_value(record.get("part_code", "")) or "",
            "resultado": _vision_format_value(record.get("result", "")) or "",
            "fecha_hora": resolution.get("reference_datetime") or "",
            "resolved_path": resolution.get("resolved_path", ""),
            "share_name": resolution.get("share_name", ""),
            "side_folder": resolution.get("side_folder", ""),
            "delta_seconds": resolution.get("delta_seconds"),
            "searched_paths": resolution.get("searched_paths", []),
            "source_file": _vision_format_value(record.get("source_file", "")) or "",
            "machine_ip": _vision_format_value(record.get("machine_ip", "")) or "",
            "image_url": "",
        }

        if resolution.get("resolved_path"):
            response_payload["image_url"] = url_for(
                "historial_vision.vision_image_file_api", id=record.get("id")
            )
            return jsonify(response_payload)

        response_payload["error"] = resolution.get(
            "error", "No se encontro imagen para el registro."
        )
        return jsonify(response_payload), 404
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@bp.route("/api/vision/image-file")
@login_requerido
def vision_image_file_api():
    """Servir la imagen resuelta de un registro de history_vision."""
    import traceback
    record_id = request.args.get("id", "").strip()
    if not re.fullmatch(r"\d+", record_id):
        return jsonify({"error": "ID de registro invalido."}), 400

    try:
        record = _get_history_vision_record(int(record_id))
        if not record:
            return jsonify({"error": "Registro de vision no encontrado."}), 404

        resolution = _resolve_history_vision_image(record)
        resolved_path = resolution.get("resolved_path", "")
        share_roots = [Path(path) for path in resolution.get("share_roots", [])]

        if not resolved_path:
            return (
                jsonify({
                    "error": resolution.get(
                        "error", "No se encontro imagen para el registro."
                    ),
                    "searched_paths": resolution.get("searched_paths", []),
                }),
                404,
            )

        resolved_file = Path(resolved_path)
        if not _vision_is_safe_path(resolved_file, share_roots):
            return jsonify({"error": "La ruta resuelta no es segura."}), 403

        if not resolved_file.is_file():
            return (
                jsonify({
                    "error": "La imagen resuelta ya no esta disponible.",
                    "resolved_path": str(resolved_file),
                }),
                404,
            )

        mimetype_map = {
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".bmp": "image/bmp",
        }
        mimetype = mimetype_map.get(
            resolved_file.suffix.lower(), "application/octet-stream"
        )

        return send_file(
            str(resolved_file),
            mimetype=mimetype,
            as_attachment=False,
            download_name=resolved_file.name,
            conditional=True,
        )
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@bp.route("/api/vision/export")
@login_requerido
def export_vision_excel():
    """Exportar el historial Vision filtrado a un archivo de Excel."""
    import traceback
    try:
        sql, params = _build_history_vision_query()
        rows = execute_query(sql, params, fetch="all") or []

        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Historial Vision"

        header_fill = PatternFill(start_color="3f6b6e", end_color="3f6b6e", fill_type="solid")
        cell_fill = PatternFill(start_color="a1a09c", end_color="a1a09c", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        headers = [
            "Linea", "Fecha", "Hora", "Numero de parte", "QR", "Barcode", "Resultado",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_idx, row in enumerate(rows, start=2):
            values = [
                _vision_format_value(row.get("linea", "")) or "",
                _vision_format_value(row.get("fecha", "")) or "",
                _vision_format_value(row.get("hora", "")) or "",
                _vision_format_value(row.get("numero_parte", "")) or "",
                _vision_format_value(row.get("qr", "")) or "",
                _vision_format_value(row.get("barcode", "")) or "",
                _vision_format_value(row.get("resultado", "")) or "",
            ]

            for col_num, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_num, value=value)
                cell.fill = cell_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        column_widths = [22, 14, 14, 28, 36, 24, 14]
        for col_num, width in enumerate(column_widths, start=1):
            column_letter = ws.cell(row=1, column=col_num).column_letter
            ws.column_dimensions[column_letter].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"historial_vision_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500
