"""Endpoints HTTP del modulo "Historial de cambios de parametros ICT".

Consumido por LISTA_DE_CONTROL_DE_RESULTADOS / Historial de maquinas calidad.
JS cliente: app/static/js/historial_cambios_parametros_ict.js
Template:   app/templates/Control de resultados/historial_cambios_parametros_ict_ajax.html

Rutas:
  GET /historial_cambios_parametros_ict/ajax  -> render template (canonica)
  GET /api/ict/param-changes/progress         -> polling de progreso del calculo
  GET /api/ict/param-changes                  -> detectar cambios entre archivos LGD
  GET /api/ict/param-changes/export           -> exportar cambios a Excel
  GET /api/ict/param-changes/detail           -> localizar el archivo .lgd donde
                                                 aparece por primera vez un valor

Alias 2026-05-27:
  GET /historial-cambios-parametros-ict-ajax  -> 301 a la canonica

Helpers DDL (idempotentes): crear_indice_history_ict_audit() y
crear_indice_history_ict_ts_nopart() se invocan desde app/startup_init.py.

Migrado desde app/routes.py el 2026-05-27. Las 4 APIs ya tenian
@login_requerido en el original. _ict_compute_parameter_changes()
preserva 1:1 (incluyendo el progress tracker en memoria y el modo
"pares consecutivos" con ThreadPoolExecutor).
"""

import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, timedelta
from datetime import time as dt_time
from decimal import Decimal

from flask import Blueprint, jsonify, redirect, render_template, request

from app.api.shared import execute_query, login_requerido
from app.services.ict_lgd_parser import (
    IctLgdError,
    get_lgd_parameters_for_barcode,
    resolve_lgd_path,
)


bp = Blueprint("historial_cambios_parametros_ict", __name__)


# ---------------------------------------------------------------------------
# DDL (llamado desde startup_init.py)
# ---------------------------------------------------------------------------


def crear_indice_history_ict_audit():
    """Crea indice cubriente para acelerar el modulo de Cambios de Parametros ICT.

    El SQL de _ict_compute_parameter_changes filtra por (ict, ts) y agrupa por
    (no_parte, linea, ict, fuente_archivo). Este indice cubre el WHERE + GROUP BY
    sin tocar la tabla base.
    """
    try:
        execute_query(
            "CREATE INDEX idx_history_ict_audit "
            "ON history_ict (ict, ts, no_parte, linea, fuente_archivo)"
        )
        print("Indice idx_history_ict_audit creado")
    except Exception as e:
        msg = str(e)
        if "1061" in msg or "Duplicate key name" in msg:
            return
        print(f"(info) idx_history_ict_audit no se pudo crear: {e}")


def crear_indice_history_ict_ts_nopart():
    """Indice cubriente para consultas sin filtro por ict (ict_all=1).

    El indice idx_history_ict_audit empieza con (ict, ts, ...). Cuando la query
    omite `AND ict = ?`, MySQL no puede usar el prefijo y cae a un scan completo.
    Este segundo indice arranca en (ts, ...) para cubrir esos casos sin scan.
    """
    try:
        execute_query(
            "CREATE INDEX idx_history_ict_ts_nopart "
            "ON history_ict (ts, no_parte, linea, fuente_archivo)"
        )
        print("Indice idx_history_ict_ts_nopart creado")
    except Exception as e:
        msg = str(e)
        if "1061" in msg or "Duplicate key name" in msg:
            return
        print(f"(info) idx_history_ict_ts_nopart no se pudo crear: {e}")


# ---------------------------------------------------------------------------
# Constantes + estado de modulo
# ---------------------------------------------------------------------------


_ICT_PARAM_JORNADA_START = dt_time(7, 30)

# Progreso del calculo en memoria, identificado por progress_id del frontend.
_ICT_PARAM_PROGRESS = {}
_ICT_PARAM_PROGRESS_LOCK = threading.Lock()
_ICT_PARAM_PROGRESS_TTL = 120  # segundos


def _ict_param_progress_set(progress_id, **fields):
    if not progress_id:
        return
    with _ICT_PARAM_PROGRESS_LOCK:
        now = time.monotonic()
        # Evict entradas viejas para evitar crecimiento ilimitado
        stale = [pid for pid, st in _ICT_PARAM_PROGRESS.items()
                 if now - st.get("started", now) > _ICT_PARAM_PROGRESS_TTL]
        for pid in stale:
            _ICT_PARAM_PROGRESS.pop(pid, None)
        state = _ICT_PARAM_PROGRESS.setdefault(progress_id, {"started": now})
        state.update(fields)


def _ict_param_progress_increment(progress_id, by=1):
    if not progress_id:
        return
    with _ICT_PARAM_PROGRESS_LOCK:
        state = _ICT_PARAM_PROGRESS.get(progress_id)
        if state:
            state["done"] = state.get("done", 0) + by


_ICT_PARAM_CHANGE_FIELDS = (
    ("std_value", "STD"),
    ("std_unit", "UNIT (STD)"),
    ("hlim_pct", "HLIM %"),
    ("llim_pct", "LLIM %"),
    ("hp_value", "HP"),
    ("lp_value", "LP"),
    ("ws_value", "WS"),
    ("ds_value", "DS"),
    ("rc_value", "RC"),
    ("p_flag", "P"),
    ("j_flag", "J"),
)


# ---------------------------------------------------------------------------
# Helpers de parsing + comparacion
# ---------------------------------------------------------------------------


def _ict_param_parse_date(value, field_name):
    value = (value or "").strip()
    if not value:
        raise ValueError(f"{field_name} es requerido.")
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"{field_name} debe tener formato YYYY-MM-DD.") from exc


def _ict_param_parse_time(value, field_name):
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(value, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"{field_name} debe tener formato HH:MM.")


def _ict_param_as_datetime(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, dt_time.min)
    if isinstance(value, str):
        raw = value.strip()
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M:%S.%f",
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%dT%H:%M:%S.%f",
        ):
            try:
                return datetime.strptime(raw, fmt)
            except ValueError:
                continue
    return None


def _ict_param_parse_ict_filter(raw_value):
    raw = (raw_value or "").strip().upper()
    if not raw:
        raise ValueError("ICT es requerido.")

    ict_match = re.search(r"\bICT\s*([0-9]+)\b", raw)
    if ict_match:
        ict_value = int(ict_match.group(1))
    elif raw.isdigit():
        ict_value = int(raw)
    else:
        raise ValueError("ICT debe ser un numero o texto como ICT1.")

    line_match = re.search(r"\b(M|DP|H)\s*([0-9]+)\b", raw)
    line_value = f"{line_match.group(1)}{line_match.group(2)}" if line_match else ""
    return ict_value, line_value


def _ict_param_format_ict(linea, ict):
    linea = str(linea or "").strip()
    ict = str(ict or "").strip()
    if linea and ict:
        return f"{linea} ICT{ict}"
    if ict:
        return f"ICT{ict}"
    return linea


def _ict_param_jornada_label(ts_value):
    ts = _ict_param_as_datetime(ts_value)
    if not ts:
        return ""
    jornada = ts.date() if ts.time() >= _ICT_PARAM_JORNADA_START else ts.date() - timedelta(days=1)
    return jornada.isoformat()


def _ict_param_time_allowed(ts_value, hora_desde, hora_hasta):
    ts = _ict_param_as_datetime(ts_value)
    if not ts:
        return False
    current = ts.time()
    if hora_desde and hora_hasta:
        if hora_desde <= hora_hasta:
            return hora_desde <= current <= hora_hasta
        return current >= hora_desde or current <= hora_hasta
    if hora_desde:
        return current >= hora_desde
    if hora_hasta:
        return current <= hora_hasta
    return True


def _ict_param_compare_token(value):
    if value is None:
        return ("empty", "")
    raw = str(value).strip()
    if raw == "":
        return ("empty", "")
    try:
        return ("num", Decimal(raw.replace(",", "")).normalize())
    except Exception:
        return ("text", raw.upper())


def _ict_param_display_value(value):
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, float):
        return f"{value:g}"
    return str(value).strip()


def _ict_param_component_label(componente, pinref):
    componente = str(componente or "").strip()
    pinref = str(pinref or "").strip()
    if componente and pinref:
        return f"{componente} / {pinref}"
    return componente or pinref


def _ict_param_build_snapshot(param_rows):
    """Snapshot por (componente, pinref) con valores crudos y tokens precomputados.

    Estructura: {(comp, pin): {"values": {field: raw}, "tokens": {field: token}}}.
    Precomputar el token (que invoca Decimal.normalize) ahorra ~22k normalizaciones
    por grupo en el bucle de comparacion.
    """
    snapshot = {}
    fields = _ICT_PARAM_CHANGE_FIELDS
    for row in param_rows:
        componente = str(row.get("componente") or "").strip()
        pinref = str(row.get("pinref") or "").strip()
        if not componente and not pinref:
            continue
        values = {}
        tokens = {}
        for field_key, _label in fields:
            v = row.get(field_key)
            values[field_key] = v
            tokens[field_key] = _ict_param_compare_token(v)
        snapshot[(componente, pinref)] = {"values": values, "tokens": tokens}
    return snapshot


def _ict_param_load_snapshot(source_file, barcode, warnings):
    source_file = str(source_file or "").strip()
    barcode = str(barcode or "").strip()
    if not source_file:
        warnings.append("Se omitio un registro sin fuente_archivo.")
        return None
    if not barcode:
        warnings.append(f"Se omitio {source_file}: no tiene barcode representativo.")
        return None

    try:
        lgd_path = resolve_lgd_path(source_file)
        param_rows = get_lgd_parameters_for_barcode(str(lgd_path), barcode)
    except (IctLgdError, OSError) as exc:
        warnings.append(f"No se pudo leer {source_file}: {exc}")
        return None
    except Exception as exc:
        warnings.append(f"No se pudo parsear {source_file}: {exc}")
        return None

    if not param_rows:
        warnings.append(f"Sin parametros para {barcode} en {source_file}.")
        return None
    return _ict_param_build_snapshot(param_rows)


# ---------------------------------------------------------------------------
# Core: detectar cambios entre archivos LGD consecutivos
# ---------------------------------------------------------------------------


def _ict_compute_parameter_changes(
    fecha_desde,
    fecha_hasta,
    hora_desde,
    hora_hasta,
    ict_filter,
    no_parte_filter="",
    componente_filter="",
    parametro_filter="",
    limit=1000,
    ict_all=False,
    progress_id=None,
):
    _ict_param_progress_set(progress_id, total=0, done=0, phase="iniciando")
    start_date = _ict_param_parse_date(fecha_desde, "fecha_desde")
    end_date = _ict_param_parse_date(fecha_hasta or fecha_desde, "fecha_hasta")
    if end_date < start_date:
        raise ValueError("fecha_hasta no puede ser menor que fecha_desde.")

    if ict_all:
        ict_value = None
        line_filter = ""
    else:
        ict_value, line_filter = _ict_param_parse_ict_filter(ict_filter)
    hora_desde_value = _ict_param_parse_time(hora_desde, "hora_desde")
    hora_hasta_value = _ict_param_parse_time(hora_hasta, "hora_hasta")
    no_parte_filter = (no_parte_filter or "").strip()
    componente_filter = (componente_filter or "").strip().lower()
    parametro_filter = (parametro_filter or "").strip().lower()

    jornada_start = datetime.combine(start_date, _ICT_PARAM_JORNADA_START)
    jornada_end = datetime.combine(end_date + timedelta(days=1), _ICT_PARAM_JORNADA_START)

    sql = (
        "SELECT MIN(barcode) AS barcode, MIN(ts) AS first_ts, "
        "COALESCE(NULLIF(TRIM(no_parte), ''), 'SIN NUMERO DE PARTE') AS no_parte, "
        "COALESCE(NULLIF(TRIM(linea), ''), 'SIN LINEA') AS linea, "
        "ict, fuente_archivo "
        "FROM history_ict "
        "WHERE ts >= %s AND ts < %s "
        "AND fuente_archivo IS NOT NULL AND fuente_archivo <> ''"
    )
    params = [jornada_start, jornada_end]

    if ict_value is not None:
        sql += " AND ict = %s"
        params.append(ict_value)
    if line_filter:
        sql += " AND linea = %s"
        params.append(line_filter)
    if no_parte_filter:
        sql += " AND no_parte LIKE %s"
        params.append(f"{no_parte_filter}%")

    sql += (
        " GROUP BY "
        "COALESCE(NULLIF(TRIM(no_parte), ''), 'SIN NUMERO DE PARTE'), "
        "COALESCE(NULLIF(TRIM(linea), ''), 'SIN LINEA'), "
        "ict, fuente_archivo "
        "ORDER BY no_parte ASC, linea ASC, ict ASC, first_ts ASC"
    )

    _ict_param_progress_set(progress_id, phase="consultando_db")
    source_rows = execute_query(sql, tuple(params), fetch="all") or []
    warnings = []
    warnings_lock = threading.Lock()

    # Modo "pares consecutivos": para cada grupo (linea, ict, no_parte)
    # ordenamos los archivos .lgd por ts y comparamos cada par consecutivo
    # (i, i+1). Esto detecta cambios intermedios (ej. 0->5->0) que un modo
    # primer-vs-ultimo perderia. Cada flip distinto se emite como una fila.
    groups_raw = {}
    for source_row in source_rows:
        ts = _ict_param_as_datetime(source_row.get("first_ts"))
        if not ts or not _ict_param_time_allowed(ts, hora_desde_value, hora_hasta_value):
            continue
        source_file = source_row.get("fuente_archivo") or ""
        if not source_file:
            continue
        gkey = (
            str(source_row.get("linea") or ""),
            str(source_row.get("ict") or ""),
            str(source_row.get("no_parte") or ""),
        )
        groups_raw.setdefault(gkey, []).append({
            "ts": ts,
            "source_file": source_file,
            "barcode": source_row.get("barcode") or "",
        })

    plan = {}  # gkey -> [items ordenados, archivos distintos consecutivos]
    for gkey, items in groups_raw.items():
        items.sort(key=lambda item: item["ts"])
        deduped = []
        for it in items:
            if deduped and deduped[-1]["source_file"] == it["source_file"]:
                continue
            deduped.append(it)
        if len(deduped) >= 2:
            plan[gkey] = deduped

    unique_files = {}
    for items in plan.values():
        for it in items:
            existing = unique_files.get(it["source_file"])
            if existing is None or it["ts"] < existing[0]:
                unique_files[it["source_file"]] = (it["ts"], it["barcode"])

    _ict_param_progress_set(
        progress_id,
        total=len(unique_files),
        done=0,
        phase="parseando_archivos",
    )

    def _parse_one(item):
        source_file, (_ts, barcode) = item
        local_warnings = []
        snapshot = _ict_param_load_snapshot(source_file, barcode, local_warnings)
        _ict_param_progress_increment(progress_id)
        return source_file, snapshot, local_warnings

    snapshot_by_file = {}
    if unique_files:
        max_workers = min(8, len(unique_files))
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            for source_file, snapshot, local_warnings in executor.map(_parse_one, list(unique_files.items())):
                snapshot_by_file[source_file] = snapshot
                if local_warnings:
                    with warnings_lock:
                        warnings.extend(local_warnings)

    _ict_param_progress_set(progress_id, phase="detectando_cambios")
    files_read = sum(1 for snap in snapshot_by_file.values() if snap is not None)

    # Optimizacion 1: filtrar campos por parametro_filter UNA VEZ por request,
    # no por (par x componente).
    fields_to_check = [
        (k, l) for k, l in _ICT_PARAM_CHANGE_FIELDS
        if not parametro_filter or parametro_filter in l.lower()
    ]

    events = []
    grupos_con_cambios = 0
    if fields_to_check:
        for gkey in sorted(plan):
            items = plan[gkey]
            linea_value, ict_value_g, no_parte_value = gkey
            grupo_tuvo_cambio = False
            # Optimizacion 2: cachear shared_keys del grupo (interseccion de
            # (comp, pin)). Se invalida solo si cambia la identidad del par de
            # snapshots, lo cual rara vez sucede dentro del mismo grupo.
            shared_keys_cache = None

            for previous, current in zip(items, items[1:]):
                prev_snapshot = snapshot_by_file.get(previous["source_file"])
                curr_snapshot = snapshot_by_file.get(current["source_file"])
                if prev_snapshot is None or curr_snapshot is None:
                    continue

                if (
                    shared_keys_cache is None
                    or shared_keys_cache[0] is not prev_snapshot
                    or shared_keys_cache[1] is not curr_snapshot
                ):
                    shared = sorted(
                        set(prev_snapshot.keys()) & set(curr_snapshot.keys()),
                        key=lambda item: (item[0], item[1]),
                    )
                    if componente_filter:
                        shared = [
                            ck for ck in shared
                            if componente_filter in _ict_param_component_label(*ck).lower()
                        ]
                    shared_keys_cache = (prev_snapshot, curr_snapshot, shared)
                shared_keys = shared_keys_cache[2]

                for comp_key in shared_keys:
                    componente_raw, pinref_raw = comp_key
                    prev_cell = prev_snapshot[comp_key]
                    curr_cell = curr_snapshot[comp_key]
                    prev_tokens = prev_cell["tokens"]
                    curr_tokens = curr_cell["tokens"]
                    prev_values = prev_cell["values"]
                    curr_values = curr_cell["values"]
                    component_label = None  # lazy

                    for field_key, field_label in fields_to_check:
                        # Optimizacion 3: comparacion con tokens precomputados,
                        # cero llamadas a Decimal.normalize en este bucle.
                        if prev_tokens[field_key] == curr_tokens[field_key]:
                            continue
                        if component_label is None:
                            component_label = _ict_param_component_label(componente_raw, pinref_raw)

                        grupo_tuvo_cambio = True
                        events.append({
                            "jornada": _ict_param_jornada_label(current["ts"]),
                            "hora_anterior": previous["ts"].strftime("%H:%M:%S"),
                            "hora_cambio": current["ts"].strftime("%H:%M:%S"),
                            "ict": _ict_param_format_ict(linea_value, ict_value_g),
                            "ict_num": ict_value_g,
                            "linea": linea_value,
                            "no_parte": no_parte_value,
                            "std": no_parte_value,
                            "componente": component_label,
                            "componente_raw": componente_raw,
                            "pinref": pinref_raw,
                            "parametro": field_label,
                            "field_key": field_key,
                            "valor_anterior": _ict_param_display_value(prev_values[field_key]),
                            "valor_nuevo": _ict_param_display_value(curr_values[field_key]),
                            "archivo_anterior": previous["source_file"],
                            "archivo_cambio": current["source_file"],
                            "barcode_anterior": previous["barcode"],
                            "barcode_cambio": current["barcode"],
                        })
            if grupo_tuvo_cambio:
                grupos_con_cambios += 1

    events.sort(
        key=lambda row: (row.get("jornada", ""), row.get("hora_cambio", "")),
        reverse=True,
    )
    total_events = len(events)
    limited_events = events[:limit] if limit else events
    file_warning_count = len(warnings)
    if limit and total_events > limit:
        warnings.append(f"Se muestran {limit} de {total_events} cambios. Use filtros para reducir la consulta.")

    _ict_param_progress_set(progress_id, phase="completado")

    return {
        "rows": limited_events,
        "warnings": warnings,
        "meta": {
            "archivos_consultados": len(source_rows),
            "grupos_total": len(groups_raw),
            "grupos_con_cambios": grupos_con_cambios,
            "archivos_unicos": len(unique_files),
            "archivos_leidos": files_read,
            "archivos_faltantes": file_warning_count,
            "eventos": total_events,
            "limite": limit,
            "jornada_inicio": jornada_start.strftime("%Y-%m-%d %H:%M:%S"),
            "jornada_fin": jornada_end.strftime("%Y-%m-%d %H:%M:%S"),
        },
    }


# ---------------------------------------------------------------------------
# Render template
# ---------------------------------------------------------------------------


@bp.route("/historial_cambios_parametros_ict/ajax")
@login_requerido
def historial_cambios_parametros_ict_ajax():
    """Render canonico del template Historial de Cambios de Parametros ICT."""
    try:
        return render_template(
            "Control de resultados/historial_cambios_parametros_ict_ajax.html"
        )
    except Exception as e:
        print(f"Error al cargar template de Historial Cambios Parametros ICT: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@bp.route("/historial-cambios-parametros-ict-ajax")
def alias_legacy_cambios_parametros():
    """Alias 301 -> /historial_cambios_parametros_ict/ajax."""
    return redirect("/historial_cambios_parametros_ict/ajax", code=301)


# ---------------------------------------------------------------------------
# APIs
# ---------------------------------------------------------------------------


@bp.route("/api/ict/param-changes/progress")
@login_requerido
def ict_param_changes_progress():
    """Devuelve el progreso del calculo de cambios de parametros ICT.

    El frontend manda progress_id al iniciar la consulta principal y luego
    pollea este endpoint para alimentar la barra de progreso.
    """
    pid = request.args.get("id", "").strip()
    with _ICT_PARAM_PROGRESS_LOCK:
        state = _ICT_PARAM_PROGRESS.get(pid)
        if not state:
            return jsonify({"total": 0, "done": 0, "phase": "desconocido"})
        return jsonify({
            "total": state.get("total", 0),
            "done": state.get("done", 0),
            "phase": state.get("phase", ""),
        })


@bp.route("/api/ict/param-changes")
@login_requerido
def ict_param_changes_api():
    """API para obtener cambios de parametros ICT desde archivos LGD locales."""
    try:
        fecha = request.args.get("fecha", "").strip()
        fecha_desde = request.args.get("fecha_desde", "").strip()
        fecha_hasta = request.args.get("fecha_hasta", "").strip()
        payload = _ict_compute_parameter_changes(
            fecha_desde=fecha_desde or fecha,
            fecha_hasta=fecha_hasta or fecha_desde or fecha,
            hora_desde=request.args.get("hora_desde", "").strip(),
            hora_hasta=request.args.get("hora_hasta", "").strip(),
            ict_filter=request.args.get("ict", "").strip(),
            no_parte_filter=(
                request.args.get("no_parte", "").strip()
                or request.args.get("numero_parte", "").strip()
                or request.args.get("std", "").strip()
            ),
            componente_filter=request.args.get("componente", "").strip(),
            parametro_filter=request.args.get("parametro", "").strip(),
            limit=1000,
            ict_all=request.args.get("ict_all", "").strip() == "1",
            progress_id=request.args.get("progress_id", "").strip() or None,
        )
        return jsonify(payload)
    except ValueError as e:
        return jsonify({"error": str(e), "rows": [], "warnings": [], "meta": {}}), 400
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@bp.route("/api/ict/param-changes/export")
@login_requerido
def ict_param_changes_export():
    """Exportar cambios de parametros ICT calculados desde LGD locales."""
    from app.api.shared.excel_helpers import _send_excel_download

    try:
        fecha = request.args.get("fecha", "").strip()
        fecha_desde = request.args.get("fecha_desde", "").strip()
        fecha_hasta = request.args.get("fecha_hasta", "").strip()
        payload = _ict_compute_parameter_changes(
            fecha_desde=fecha_desde or fecha,
            fecha_hasta=fecha_hasta or fecha_desde or fecha,
            hora_desde=request.args.get("hora_desde", "").strip(),
            hora_hasta=request.args.get("hora_hasta", "").strip(),
            ict_filter=request.args.get("ict", "").strip(),
            no_parte_filter=(
                request.args.get("no_parte", "").strip()
                or request.args.get("numero_parte", "").strip()
                or request.args.get("std", "").strip()
            ),
            componente_filter=request.args.get("componente", "").strip(),
            parametro_filter=request.args.get("parametro", "").strip(),
            limit=5000,
            ict_all=request.args.get("ict_all", "").strip() == "1",
            progress_id=request.args.get("progress_id", "").strip() or None,
        )
        rows = payload.get("rows", [])

        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Cambios Parametros ICT"

        headers = [
            "Jornada", "Hora Anterior", "Hora Cambio", "ICT", "Linea",
            "No Parte", "Componente", "Parametro", "Valor Anterior",
            "Valor Nuevo", "Archivo Anterior", "Archivo Cambio",
        ]
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row.get("jornada", "") or "")
            ws.cell(row=row_idx, column=2, value=row.get("hora_anterior", "") or "")
            ws.cell(row=row_idx, column=3, value=row.get("hora_cambio", "") or "")
            ws.cell(row=row_idx, column=4, value=row.get("ict", "") or "")
            ws.cell(row=row_idx, column=5, value=row.get("linea", "") or "")
            ws.cell(row=row_idx, column=6, value=row.get("no_parte", "") or "")
            ws.cell(row=row_idx, column=7, value=row.get("componente", "") or "")
            ws.cell(row=row_idx, column=8, value=row.get("parametro", "") or "")
            ws.cell(row=row_idx, column=9, value=row.get("valor_anterior", "") or "")
            ws.cell(row=row_idx, column=10, value=row.get("valor_nuevo", "") or "")
            ws.cell(row=row_idx, column=11, value=row.get("archivo_anterior", "") or "")
            ws.cell(row=row_idx, column=12, value=row.get("archivo_cambio", "") or "")

        col_widths = [12, 12, 12, 14, 8, 22, 25, 22, 20, 20, 48, 48]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = width
        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"cambios_parametros_ict_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return _send_excel_download(output, filename)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        import traceback
        print(f"Error exportando Cambios Parametros ICT: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


@bp.route("/api/ict/param-changes/detail")
@login_requerido
def ict_param_changes_detail():
    """Localiza el archivo .lgd mas temprano donde aparece valor_nuevo para un
    parametro especifico de un (linea, ict, no_parte). Usa busqueda binaria
    sobre los archivos del rango y un barrido lineal hacia atras para
    parametros que oscilan."""
    try:
        fecha = request.args.get("fecha", "").strip()
        fecha_desde_raw = request.args.get("fecha_desde", "").strip() or fecha
        fecha_hasta_raw = (
            request.args.get("fecha_hasta", "").strip()
            or fecha_desde_raw
            or fecha
        )
        linea = request.args.get("linea", "").strip()
        ict_raw = request.args.get("ict", "").strip()
        no_parte = request.args.get("no_parte", "").strip()
        componente = request.args.get("componente", "")
        pinref = request.args.get("pinref", "")
        field_key = request.args.get("parametro", "").strip()
        valor_nuevo = request.args.get("valor_nuevo", "")

        valid_fields = {fk for fk, _ in _ICT_PARAM_CHANGE_FIELDS}
        if field_key not in valid_fields:
            return jsonify({"error": "Parametro desconocido."}), 400
        if not no_parte:
            return jsonify({"error": "no_parte es requerido."}), 400
        if not ict_raw:
            return jsonify({"error": "ict es requerido."}), 400

        try:
            ict_value = int(ict_raw)
        except ValueError:
            return jsonify({"error": "ict debe ser numerico."}), 400

        start_date = _ict_param_parse_date(fecha_desde_raw, "fecha_desde")
        end_date = _ict_param_parse_date(fecha_hasta_raw, "fecha_hasta")
        if end_date < start_date:
            return jsonify({"error": "fecha_hasta no puede ser menor que fecha_desde."}), 400

        jornada_start = datetime.combine(start_date, _ICT_PARAM_JORNADA_START)
        jornada_end = datetime.combine(end_date + timedelta(days=1), _ICT_PARAM_JORNADA_START)

        sql = (
            "SELECT MIN(barcode) AS barcode, MIN(ts) AS first_ts, fuente_archivo "
            "FROM history_ict "
            "WHERE ts >= %s AND ts < %s "
            "AND ict = %s AND no_parte = %s "
            "AND fuente_archivo IS NOT NULL AND fuente_archivo <> ''"
        )
        params = [jornada_start, jornada_end, ict_value, no_parte]
        if linea:
            sql += " AND linea = %s"
            params.append(linea)
        sql += (
            " GROUP BY fuente_archivo "
            "ORDER BY first_ts ASC"
        )

        rows = execute_query(sql, tuple(params), fetch="all") or []
        files = []
        for row in rows:
            ts = _ict_param_as_datetime(row.get("first_ts"))
            source_file = row.get("fuente_archivo") or ""
            barcode = row.get("barcode") or ""
            if not ts or not source_file:
                continue
            files.append({"ts": ts, "file": source_file, "barcode": barcode})

        if not files:
            return jsonify({"found": False, "reason": "Sin archivos en el rango."})

        target_token = _ict_param_compare_token(valor_nuevo)
        comp_key = (str(componente or ""), str(pinref or ""))
        snap_cache = {}
        warnings_local = []

        def _matches(idx):
            entry = files[idx]
            cache_key = (entry["file"], entry["barcode"])
            snap = snap_cache.get(cache_key)
            if snap is None and cache_key not in snap_cache:
                snap = _ict_param_load_snapshot(
                    entry["file"], entry["barcode"], warnings_local
                )
                snap_cache[cache_key] = snap
            if snap is None:
                return None  # archivo no leible
            cell = snap.get(comp_key)
            if cell is None:
                return False
            # Snapshot ahora tiene la forma {"values": {...}, "tokens": {...}}.
            # Usamos los tokens precomputados para evitar Decimal.normalize aqui.
            return cell["tokens"].get(field_key) == target_token

        last_idx = len(files) - 1
        last_match = _matches(last_idx)
        if last_match is None:
            return jsonify({"found": False, "reason": "No se pudo leer el ultimo archivo."})
        if last_match is False:
            return jsonify({"found": False, "reason": "El valor nuevo no aparece en archivos posteriores."})

        lo, hi = 0, last_idx
        while lo < hi:
            mid = (lo + hi) // 2
            mid_match = _matches(mid)
            if mid_match is True:
                hi = mid
            else:
                # archivo no leible o sin valor_nuevo: buscar mas adelante
                lo = mid + 1

        # Barrido lineal hacia atras para tolerar oscilaciones
        while lo > 0:
            prev_match = _matches(lo - 1)
            if prev_match is True:
                lo -= 1
            else:
                break

        target = files[lo]
        return jsonify({
            "found": True,
            "hora": target["ts"].strftime("%H:%M:%S"),
            "fecha": target["ts"].strftime("%Y-%m-%d"),
            "ts": target["ts"].strftime("%Y-%m-%d %H:%M:%S"),
            "archivo": target["file"],
            "barcode": target["barcode"],
            "archivos_consultados": len(files),
            "archivos_parseados": len(snap_cache),
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500
