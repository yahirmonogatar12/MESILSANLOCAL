"""Helper compartido para exportar listados a Excel (estilo ILSAN).

Extraido de inventario_reparacion_smd/_assy (2026-06-10) para que cualquier
modulo `app.api.*` genere descargas .xlsx con el mismo formato de encabezados
sin duplicar el codigo.

Dos variantes:
  - excel_response:     encabezado azul ILSAN (#1F4E79), celdas sin relleno.
  - excel_response_ict: estilo de los historiales ICT/Vision (#3f6b6e,
    celdas grises #a1a09c, bordes finos, todo centrado).

Consumidores actuales:
  - control_resultados.inventario_reparacion_smd / _assy
  - control_resultados.historial_operadores_maquina
  - control_resultados.historial_ict (export y export-defects)
  - control_resultados.historial_vision (export)
  - control_resultados.historial_cambios_parametros_ict (export)
  - control_calidad.stations_qa

Para exports con layout especial (imagenes de barras Pass/Fail, comparacion
de parametros con resaltado de diferencias) ver excel_helpers.py y el codigo
propio de cada modulo.
"""

from io import BytesIO

from flask import Response

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _workbook_response(wb, filename):
    """Serializa el Workbook y lo devuelve como descarga adjunta."""
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype=_XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


def excel_response(rows, headers, keys, widths, sheet, filename, freeze=None):
    """Construye un .xlsx con encabezados ILSAN y lo devuelve como descarga.

    rows:     lista de dicts (cada dict es una fila).
    headers:  titulos de columna visibles.
    keys:     llave del dict a usar por columna (mismo orden que headers).
    widths:   ancho de cada columna (mismo orden que headers).
    sheet:    nombre de la hoja (se recorta a 31 chars, limite de Excel).
    filename: nombre del archivo SIN extension.
    freeze:   celda para freeze_panes (p.ej. "A2"), opcional.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet[:31]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    if freeze:
        ws.freeze_panes = freeze

    return _workbook_response(wb, filename)


def excel_response_ict(rows, headers, keys, widths, sheet, filename, freeze=None):
    """Variante con el estilo de los historiales ICT/Vision.

    Encabezado verde petroleo (#3f6b6e) con fuente blanca bold 10, celdas
    grises (#a1a09c), bordes finos negros y todo el texto centrado. Misma
    firma que excel_response.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = sheet[:31]

    header_fill = PatternFill(start_color="3f6b6e", end_color="3f6b6e", fill_type="solid")
    cell_fill = PatternFill(start_color="a1a09c", end_color="a1a09c", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(keys, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
            cell.fill = cell_fill
            cell.alignment = center
            cell.border = border

    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    if freeze:
        ws.freeze_panes = freeze

    return _workbook_response(wb, filename)
