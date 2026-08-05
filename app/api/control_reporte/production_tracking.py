"""Endpoints HTTP del modulo "Production Tracking By Process".

Reporte de SOLO LECTURA sobre la tabla `Tracking`: el operador escanea un QR o
un barcode y se le devuelve por que etapas del proceso paso esa pieza y cuando.

`Tracking` NO la escribe esta app: la llenan 13 triggers de la base compartida,
uno por etapa (SMT, IMD, Assy, Vision, ICT_QA, FCT, Packing_QA, Rep_SMD,
Rep_Assy, Releases_OQC, Embarques, Scrap). El detalle de cada trigger y sus
tablas fuente esta en:
    Modulo IMPRESION/backend/sql/2026-08-04_tracking_etapas.sql

Piezas sin fila en Tracking: se reconstruye su historial leyendo directamente
las tablas de origen (ver _historial_desde_fuentes). Pasa con las lineas que no
escanean el barcode en ASSY, donde el QR y el barcode nunca se ven juntos y por
lo tanto la fila nunca se creo. Lo que no se puede determinar se marca como
"SIN QR" o "SIN BARCODE", nunca como "-": en trazabilidad, "no se sabe" y "no
paso por ahi" no son lo mismo.

JS cliente: app/static/js/production-tracking-process.js
Template:   app/templates/Control de reporte/production_tracking_process_ajax.html

Rutas:
  GET  /control_reporte/production_tracking       -> render template
  GET  /api/production_tracking/buscar            -> historial de la pieza (JSON)
  GET  /api/production_tracking/buscar/export     -> idem (.xlsx)
  POST /api/production_tracking/lote              -> muchos codigos desde un .xlsx
                                                     (JSON, o .xlsx con formato=excel)
"""

import logging
import re
import traceback

from flask import Blueprint, jsonify, render_template, request

from app.api.shared import excel_response, execute_query, login_requerido

logger = logging.getLogger(__name__)


bp = Blueprint("control_reporte_production_tracking", __name__)


# (columna flag, columna fecha, etiqueta). El orden es el del flujo fisico y
# coincide con el orden de las columnas en la tabla.
_ETAPAS = [
    ("SMT", "SMT_At", "SMT"),
    ("IMD", "IMD_At", "IMD"),
    ("Assy", "Assy_At", "Assy"),
    ("Vision", "Vision_At", "Vision"),
    ("ICT_QA", "ICT_QA_At", "ICT"),
    ("FCT", "FCT_At", "FCT"),
    ("Packing_QA", "Packing_QA_At", "Packing QA"),
    ("Rep_SMD", "Rep_SMD_At", "Reparacion SMD"),
    ("Rep_Assy", "Rep_Assy_At", "Reparacion Assy"),
    ("Releases_OQC", "Releases_OQC_At", "Releases OQC"),
    ("Embarques", "Embarques_At", "Embarques"),
    ("Scrap", "Scrap_At", "Scrap"),
]

# Etapas por las que una pieza sana NO tiene que pasar. Si no las alcanzo se
# muestran como "N/A" en vez de "-": no haberlas tocado es lo normal, mientras
# que una etapa del flujo sin registrar si es algo que falto.
_ETAPAS_EXCEPCION = {"Rep_SMD", "Rep_Assy", "Scrap"}

# Etapas cuyos triggers cruzan por Barcode. Si la pieza no tiene Barcode no se
# pueden determinar NUNCA, y un "-" se leeria como "no paso por ahi" cuando en
# realidad es "no se puede saber".
#
# Pasa con las lineas M1, D1 y D2: no escanean el codigo de barras en ASSY
# (99.2%, 100% y 99.9% de sus filas vienen sin el), asi que el QR y el barcode
# de esas piezas no coinciden en ninguna tabla. Se busco un puente en
# history_vision, que si guarda qr_payload y barcode juntos, pero de 98,147 QR
# sin barcode solo 23 tenian ambos: la camara les lee el QR y no el barcode.
# Mientras no se capture el barcode en ASSY, esto no se arregla con SQL.
_ETAPAS_POR_BARCODE = {"ICT_QA", "FCT", "Packing_QA", "Releases_OQC", "Embarques"}

# El espejo: etapas cuyos triggers cruzan SOLO por QR. Una pieza que tiene
# barcode pero nunca se vinculo a un QR no puede tenerlas. (Vision, las dos de
# reparacion y Scrap NO estan aqui: sus triggers aceptan cualquiera de los dos.)
_ETAPAS_POR_QR = {"SMT", "IMD", "Assy"}

# Minimo de caracteres para permitir la busqueda por prefijo. Sin esto, teclear
# una letra suelta barre la tabla completa (582k filas).
_MIN_PREFIJO = 4

# Busqueda por lote (Excel importado).
_MAX_CODIGOS_LOTE = 5000
_MAX_BYTES_LOTE = 10 * 1024 * 1024
# Codigos por sentencia. Cada IN se resuelve por indice; partirlo evita armar
# una query gigante y mantiene el uso de memoria del driver acotado.
_LOTE_CHUNK = 500


# ---------------------------------------------------------------------------
# Render template
# ---------------------------------------------------------------------------


@bp.route("/control_reporte/production_tracking")
@login_requerido
def production_tracking_ajax():
    """Ruta AJAX canonica para cargar el contenido de Production Tracking."""
    try:
        return render_template("Control de reporte/production_tracking_process_ajax.html")
    except Exception as e:
        logger.error("Error al cargar template Production Tracking AJAX: %s", e)
        logger.info(traceback.format_exc())
        return f"Error al cargar el contenido: {str(e)}", 500


# ---------------------------------------------------------------------------
# Consulta
# ---------------------------------------------------------------------------


def _buscar_filas(codigo, limit):
    """Busca la(s) pieza(s) del codigo dado. Exacto primero, prefijo despues.

    Exacto cubre el caso normal (pistola de codigo de barras) y resuelve con
    index_merge de PRIMARY + idx_tracking_barcode: 2 filas leidas.
    """
    rows = execute_query(
        f"SELECT * FROM Tracking WHERE QR = %s OR Barcode = %s LIMIT {int(limit)}",
        [codigo, codigo],
        fetch="all",
    ) or []

    if rows or len(codigo) < _MIN_PREFIJO:
        return rows

    # Prefijo (el operador pego un fragmento): una query por columna. Con
    # `QR LIKE %s OR Barcode LIKE %s` en una sola sentencia el optimizador
    # descarta ambos indices y barre las 582k filas; separadas, cada una es un
    # range scan sobre su indice. El nombre de columna sale de la tupla fija.
    pref = f"{codigo}%"
    for col in ("QR", "Barcode"):
        rows = execute_query(
            f"SELECT * FROM Tracking WHERE {col} LIKE %s LIMIT {int(limit)}",
            [pref],
            fetch="all",
        ) or []
        if rows:
            return rows

    return []


# Reconstruccion desde las tablas de origen, para barcodes que NO tienen fila en
# Tracking. Pasa con las lineas que no escanean el barcode en ASSY (M1/D1/D2):
# ICT, FCT, etc. si registraron la pieza, pero sus triggers actualizan por
# Barcode y no encontraron a quien. Sin esto el modulo dice "sin historial"
# cuando el historial existe.
#
# Es SOLO LECTURA: no se crea ninguna fila. Crearlas fue el "escalon 2" que se
# descarto, porque partiria la historia de la tarjeta en dos mitades que no se
# pueden unir (ver sql/2026-08-05_tracking_barcodes_sin_qr.sql, PASO 4).
#
# Cada entrada: (columna de etapa, SQL con {ph} para los placeholders). La
# consulta devuelve k = barcode y ts = primer registro. Todas cruzan por una
# columna indexada.
_FUENTES_BARCODE = [
    ("Vision", """
        SELECT barcode AS k, MIN(log_datetime) AS ts FROM history_vision
        WHERE result = 'OK' AND barcode IN ({ph}) GROUP BY barcode"""),
    ("ICT_QA", """
        SELECT barcode AS k, MIN(ts) AS ts FROM history_ict
        WHERE resultado = 'OK' AND barcode IN ({ph}) GROUP BY barcode"""),
    ("FCT", """
        SELECT serial_number AS k, MIN(end_at) AS ts FROM fct_test_results
        WHERE final_result = 'PASS' AND serial_number IN ({ph}) GROUP BY serial_number"""),
    ("Packing_QA", """
        SELECT serial AS k, MIN(last_scan) AS ts FROM box_scans
        WHERE serial IN ({ph}) GROUP BY serial"""),
    ("Scrap", """
        SELECT COALESCE(NULLIF(scanned_original_norm, ''), raw_barcode) AS k,
               MIN(fecha_registro) AS ts
        FROM scrap_records
        WHERE scanned_original_norm IN ({ph}) OR raw_barcode IN ({ph})
        GROUP BY k"""),
]

# Releases_OQC va aparte porque exit_records guarda las cajas como texto libre
# ('...LGB123...LGB124...') y sacarlas en SQL sale caro:
#   JOIN ... ON observations LIKE '%box_code%'   -> 2,035 ms (sin indice posible)
#   CTE recursivo con REGEXP_SUBSTR              -> 1,310 ms (~205k llamadas:
#                                                   30 iteraciones x 6,837 filas)
#   traer las 6,837 filas y parsear en Python    ->   ~40 ms
# Son pocas filas y el trabajo es de cadenas, que es justo lo que MySQL hace mal
# y Python bien.
_SQL_BOX_DE_SERIAL = """
    SELECT serial AS k, box_code FROM box_scans
    WHERE serial IN ({ph}) AND box_code IS NOT NULL AND box_code <> ''"""

_SQL_CAJAS_LIBERADAS = """
    SELECT observations, exit_date FROM exit_records
    WHERE qc_passed = 1 AND observations IS NOT NULL
      AND observations REGEXP 'LGB[0-9]+'"""

_RE_CAJA = re.compile(r"LGB[0-9]+")


def _oqc_por_barcode(barcodes, ph):
    """{barcode: primera fecha de liberacion OQC}, via su caja."""
    cajas_de = execute_query(_SQL_BOX_DE_SERIAL.format(ph=ph), barcodes, fetch="all") or []
    if not cajas_de:
        return {}

    liberada = {}  # box_code -> primera exit_date
    for r in execute_query(_SQL_CAJAS_LIBERADAS, None, fetch="all") or []:
        ts = r.get("exit_date")
        if not ts:
            continue
        for caja in _RE_CAJA.findall(r.get("observations") or ""):
            if caja not in liberada or ts < liberada[caja]:
                liberada[caja] = ts

    salida = {}
    for r in cajas_de:
        ts = liberada.get(r.get("box_code"))
        k = r.get("k")
        if ts and (k not in salida or ts < salida[k]):
            salida[k] = ts
    return salida

# Embarques va aparte: no interesa la primera fecha sino el ULTIMO movimiento,
# porque de el depende si la pieza sigue en embarques o ya salio.
_SQL_EMBARQUES = """
    SELECT bs.serial AS k, m.movement_type AS tipo, m.movement_at AS ts
    FROM box_scans bs
    JOIN embarques_movimiento_cajas m ON m.box_code = bs.box_code
    JOIN (
      SELECT bs2.serial AS s, MAX(m2.movement_at) AS ult
      FROM box_scans bs2
      JOIN embarques_movimiento_cajas m2 ON m2.box_code = bs2.box_code
      WHERE bs2.serial IN ({ph}) GROUP BY bs2.serial
    ) u ON u.s = bs.serial AND u.ult = m.movement_at
    WHERE bs.serial IN ({ph})"""


def _historial_desde_fuentes(barcodes):
    """{barcode: {columna_etapa: valor}} leido de las tablas de origen."""
    if not barcodes:
        return {}

    encontrado = {}
    pedidos = set(barcodes)

    def anota(k, col, valor):
        # Solo lo que se pidio: la consulta de Scrap mira dos columnas y puede
        # devolver una clave que no estaba en la lista.
        if k in pedidos and valor is not None:
            encontrado.setdefault(k, {})[col] = valor

    ph = ", ".join(["%s"] * len(barcodes))

    # Sin try/except a proposito: si una fuente falla, la etapa saldria como "-"
    # y eso se lee como "no paso por ahi". En un modulo de trazabilidad una
    # respuesta equivocada es peor que un error, asi que se deja reventar (misma
    # regla fail-loud de execute_query, WF_003).
    for col, plantilla in _FUENTES_BARCODE:
        # Scrap mira dos columnas, asi que necesita la lista dos veces.
        veces = plantilla.count("{ph}")
        filas = execute_query(plantilla.format(ph=ph), barcodes * veces, fetch="all") or []
        for r in filas:
            anota(r.get("k"), col, r.get("ts"))

    for k, ts in _oqc_por_barcode(barcodes, ph).items():
        anota(k, "Releases_OQC", ts)

    for r in execute_query(_SQL_EMBARQUES.format(ph=ph), barcodes * 2, fetch="all") or []:
        anota(r.get("k"), "Embarques",
              "SALIO" if r.get("tipo") == "exit" else "EN_EMBARQUES")
        anota(r.get("k"), "Embarques_At", r.get("ts"))

    return encontrado


def _fila_sintetica(barcode, etapas):
    """Fila con la forma de Tracking, armada desde las fuentes. NO se guarda."""
    fila = {"QR": None, "Barcode": barcode}
    for col, col_at, _ in _ETAPAS:
        if col == "Embarques":
            fila[col] = etapas.get("Embarques")
            fila[col_at] = etapas.get("Embarques_At")
        else:
            ts = etapas.get(col)
            fila[col] = 1 if ts is not None else 0
            fila[col_at] = ts
    return fila


def _fila_a_item(row):
    """Convierte una fila de Tracking en {cabecera + lista de etapas}."""
    etapas = []
    registradas = 0
    ultima = None  # (etiqueta, timestamp) de la etapa mas reciente con fecha

    # Sin Barcode (o sin QR), las etapas que cruzan por ese lado son
    # indeterminables, no ausentes. Son los dos lados del mismo hueco.
    sin_barcode = not (row.get("Barcode") or "").strip()
    sin_qr = not (row.get("QR") or "").strip()

    for col, col_at, label in _ETAPAS:
        valor = row.get(col)
        ts = row.get(col_at)

        ok = bool(valor)

        # "NO VINCULADO" y no "sin": la pieza casi seguro TIENE el otro codigo,
        # con su propio historial. Lo que falta es la union entre ambos, que
        # nadie registro. Decir "sin QR" sugeriria que la pieza no lo trae.
        if not ok and sin_barcode and col in _ETAPAS_POR_BARCODE:
            estado, clase = "BARCODE NO VINCULADO", "nd"
        elif not ok and sin_qr and col in _ETAPAS_POR_QR:
            estado, clase = "QR NO VINCULADO", "nd"
        elif col == "Embarques":
            # Unica etapa que no es booleana: enum EN_EMBARQUES / SALIO.
            estado = {"SALIO": "SALIO", "EN_EMBARQUES": "EN EMBARQUES"}.get(valor or "", "-")
            clase = "salio" if valor == "SALIO" else ("embarques" if ok else "pendiente")
        elif col == "Scrap":
            estado, clase = ("SCRAP", "scrap") if ok else ("N/A", "na")
        elif col in _ETAPAS_EXCEPCION:
            estado, clase = ("OK", "rep") if ok else ("N/A", "na")
        else:
            estado, clase = ("OK", "hecho") if ok else ("-", "pendiente")

        if ok:
            registradas += 1
            if ts and (ultima is None or ts > ultima[1]):
                ultima = (label, ts)

        # `celda`: el texto que va en la vista de lote (una columna por etapa).
        # Ahi el dato util es la fecha, salvo en Embarques (el estado cambia el
        # significado) y en lo no alcanzado ("-" / "N/A"). Lo calcula el backend
        # para que la tabla del navegador y el Excel digan exactamente lo mismo.
        if not ts:
            celda = estado
        elif col == "Embarques":
            celda = f"{estado} {ts}"
        else:
            celda = str(ts)

        etapas.append({
            "etapa": label,
            "ok": ok,
            "estado": estado,
            "clase": clase,
            "fecha": str(ts) if ts else "",
            "celda": celda,
        })

    # Estado global: scrap y embarques mandan sobre la ultima etapa cronologica.
    if row.get("Scrap"):
        estado_global, clase_global = "SCRAP", "scrap"
    elif row.get("Embarques") == "SALIO":
        estado_global, clase_global = "Salio de embarques", "salio"
    elif row.get("Embarques") == "EN_EMBARQUES":
        estado_global, clase_global = "En embarques", "embarques"
    elif ultima:
        estado_global, clase_global = ultima[0], "hecho"
    else:
        estado_global, clase_global = "Sin registro de proceso", "pendiente"

    # El denominador cuenta solo lo que se puede saber. Decir "4 / 12" en una
    # pieza sin barcode sugiere 8 etapas faltantes cuando 5 son indeterminables.
    total = len(_ETAPAS)
    if sin_barcode:
        total -= len(_ETAPAS_POR_BARCODE)
    if sin_qr:
        total -= len(_ETAPAS_POR_QR)

    return {
        "qr": row.get("QR") or "",
        "barcode": row.get("Barcode") or "",
        "sin_barcode": sin_barcode,
        "sin_qr": sin_qr,
        "estado_global": estado_global,
        "clase_global": clase_global,
        "registradas": registradas,
        "total": total,
        "ultima_etapa": ultima[0] if ultima else "",
        "ultima_fecha": str(ultima[1]) if ultima else "",
        "etapas": etapas,
    }


def _query_items(limit):
    codigo = request.args.get("codigo", "", type=str).strip()
    if not codigo:
        return []

    filas = _buscar_filas(codigo, limit)
    if filas:
        return [dict(_fila_a_item(r), reconstruido=False) for r in filas]

    # No hay fila en Tracking. Puede seguir habiendo historial: si la pieza
    # nunca se vinculo a un QR, ICT/FCT/Packing/OQC igual la registraron por
    # barcode. Se arma la vista desde esas tablas, sin guardar nada.
    fuentes = _historial_desde_fuentes([codigo])
    if codigo in fuentes:
        return [dict(_fila_a_item(_fila_sintetica(codigo, fuentes[codigo])),
                     reconstruido=True)]
    return []


@bp.route("/api/production_tracking/buscar", methods=["GET"])
@login_requerido
def api_production_tracking_buscar():
    """Historial de etapas de la pieza cuyo QR o barcode es `codigo`."""
    try:
        codigo = request.args.get("codigo", "", type=str).strip()
        if not codigo:
            return jsonify({"status": "success", "items": [],
                            "message": "Escanea o escribe un QR / barcode"})

        items = _query_items(limit=20)
        if not items:
            return jsonify({"status": "success", "items": [],
                            "message": "Sin historial para ese codigo"})
        return jsonify({"status": "success", "items": items})
    except Exception as e:
        logger.error("Error en api_production_tracking_buscar: %s", e)
        return jsonify({"status": "error", "message": str(e), "items": []}), 500


@bp.route("/api/production_tracking/buscar/export", methods=["GET"])
@login_requerido
def api_production_tracking_export():
    """Exportar el historial a Excel: una fila por etapa de cada pieza."""
    try:
        filas = []
        for item in _query_items(limit=200):
            for etapa in item["etapas"]:
                filas.append({
                    "qr": item["qr"],
                    "barcode": item["barcode"],
                    "estado_global": item["estado_global"],
                    "etapa": etapa["etapa"],
                    "estado": etapa["estado"],
                    "fecha": etapa["fecha"],
                })

        return excel_response(
            filas,
            ["QR", "Barcode", "Estado de la pieza", "Etapa", "Estado", "Fecha / Hora"],
            ["qr", "barcode", "estado_global", "etapa", "estado", "fecha"],
            [40, 30, 20, 18, 14, 20],
            sheet="Production Tracking",
            filename="production_tracking_by_process",
            freeze="A2",
        )
    except Exception as e:
        logger.exception("Error exportando Production Tracking: %s", e)
        return jsonify({"status": "error", "message": str(e)}), 500


# ---------------------------------------------------------------------------
# Busqueda por lote (importar Excel con muchos codigos)
#
# Una fila por codigo de ENTRADA, en el orden del archivo, con una columna por
# etapa. Los codigos que no existen en Tracking tambien salen (marcados como no
# encontrados): el operador necesita saber cuales de sus 1000 no aparecieron.
# ---------------------------------------------------------------------------


def _codigos_desde_excel(file_bytes):
    """Extrae los codigos de un .xlsx: toda celda no vacia, en orden y sin repetir.

    Se leen todas las columnas a proposito, para no obligar al operador a poner
    los codigos en una columna concreta ni a quitar encabezados: lo que no sea un
    QR o un barcode simplemente no hace match contra Tracking y sale reportado
    como no encontrado.
    """
    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        vistos = {}  # dict preserva el orden de insercion
        for hoja in wb.worksheets:
            for fila in hoja.iter_rows(values_only=True):
                for celda in fila:
                    if celda is None:
                        continue
                    # Excel devuelve los codigos 100% numericos como float; sin
                    # esto '123456' llegaria como '123456.0' y nunca cruzaria.
                    if isinstance(celda, float) and celda.is_integer():
                        celda = int(celda)
                    cod = str(celda).strip()
                    if cod and cod not in vistos:
                        vistos[cod] = True
                        if len(vistos) >= _MAX_CODIGOS_LOTE:
                            return list(vistos)
        return list(vistos)
    finally:
        wb.close()


def _buscar_lote(codigos):
    """Una entrada por codigo pedido, en el mismo orden que traia el archivo."""
    por_qr, por_barcode = {}, {}

    for i in range(0, len(codigos), _LOTE_CHUNK):
        chunk = codigos[i:i + _LOTE_CHUNK]
        marcas = ", ".join(["%s"] * len(chunk))
        rows = execute_query(
            f"SELECT * FROM Tracking WHERE QR IN ({marcas}) OR Barcode IN ({marcas})",
            chunk + chunk,
            fetch="all",
        ) or []
        for r in rows:
            por_qr[r.get("QR")] = r
            if r.get("Barcode"):
                por_barcode.setdefault(r["Barcode"], r)

    # Lo que no esta en Tracking puede seguir teniendo historial por barcode
    # (piezas que nunca se vincularon a un QR). Se reconstruye desde las tablas
    # de origen, tambien por tandas.
    faltantes = [c for c in codigos if c not in por_qr and c not in por_barcode]
    fuentes = {}
    for i in range(0, len(faltantes), _LOTE_CHUNK):
        fuentes.update(_historial_desde_fuentes(faltantes[i:i + _LOTE_CHUNK]))

    items = []
    for cod in codigos:
        row = por_qr.get(cod) or por_barcode.get(cod)
        reconstruido = False

        if row is None and cod in fuentes:
            row = _fila_sintetica(cod, fuentes[cod])
            reconstruido = True

        if row is None:
            items.append({
                "codigo": cod,
                "encontrado": False,
                "reconstruido": False,
                "qr": "",
                "barcode": "",
                "sin_barcode": False,
                "sin_qr": False,
                "estado_global": "No encontrado",
                "clase_global": "pendiente",
                "registradas": 0,
                "total": len(_ETAPAS),
                "ultima_etapa": "",
                "ultima_fecha": "",
                "etapas": [
                    {"etapa": label, "ok": False, "estado": "", "clase": "pendiente",
                     "fecha": "", "celda": ""}
                    for _, _, label in _ETAPAS
                ],
            })
            continue

        item = _fila_a_item(row)
        item["codigo"] = cod
        item["encontrado"] = True
        item["reconstruido"] = reconstruido
        items.append(item)

    return items


def _leer_lote_del_request():
    """(items, error). error es (payload, status) listo para devolver."""
    archivo = request.files.get("file")
    if archivo is None or not archivo.filename:
        return None, ({"status": "error", "message": "No se recibio archivo.", "items": []}, 400)

    file_bytes = archivo.read()
    if not file_bytes:
        return None, ({"status": "error", "message": "El archivo esta vacio.", "items": []}, 400)
    if len(file_bytes) > _MAX_BYTES_LOTE:
        return None, ({"status": "error", "message": "El archivo supera 10 MB.", "items": []}, 400)

    try:
        codigos = _codigos_desde_excel(file_bytes)
    except Exception as e:
        logger.warning("production_tracking: Excel ilegible (%s): %s", archivo.filename, e)
        return None, ({"status": "error", "items": [],
                       "message": "El archivo esta corrupto o no es un Excel (.xlsx) valido."}, 400)

    if not codigos:
        return None, ({"status": "error", "message": "El archivo no tiene celdas con datos.",
                       "items": []}, 400)

    return _buscar_lote(codigos), None


@bp.route("/api/production_tracking/lote", methods=["POST"])
@login_requerido
def api_production_tracking_lote():
    """Busca muchos codigos de un .xlsx. Devuelve JSON, o el .xlsx si formato=excel."""
    try:
        items, error = _leer_lote_del_request()
        if error is not None:
            payload, status = error
            return jsonify(payload), status

        if request.form.get("formato") == "excel":
            etiquetas = [label for _, _, label in _ETAPAS]
            filas = [
                {
                    "codigo": it["codigo"],
                    "qr": it["qr"],
                    "barcode": it["barcode"],
                    "estado_global": it["estado_global"],
                    **{f"et{i}": e["celda"] for i, e in enumerate(it["etapas"])},
                }
                for it in items
            ]
            return excel_response(
                filas,
                ["Codigo buscado", "QR", "Barcode", "Estado de la pieza"] + etiquetas,
                ["codigo", "qr", "barcode", "estado_global"]
                + [f"et{i}" for i in range(len(_ETAPAS))],
                [30, 40, 30, 20] + [20] * len(_ETAPAS),
                sheet="Tracking por lote",
                filename="production_tracking_lote",
                freeze="A2",
            )

        encontrados = sum(1 for it in items if it["encontrado"])
        return jsonify({
            "status": "success",
            "items": items,
            "etapas": [label for _, _, label in _ETAPAS],
            "total": len(items),
            "encontrados": encontrados,
            "no_encontrados": len(items) - encontrados,
        })
    except Exception as e:
        logger.exception("Error en api_production_tracking_lote: %s", e)
        return jsonify({"status": "error", "message": str(e), "items": []}), 500
