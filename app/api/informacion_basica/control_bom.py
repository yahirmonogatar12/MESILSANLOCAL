"""API de Control BOM y flujo ECO para Informacion Basica.

Mantiene las URLs legacy que consume el template de Control BOM, pero las
registra como Blueprint conforme a WF_003.
"""

from datetime import datetime
from decimal import Decimal
from functools import wraps
import io
import re
import tempfile
import traceback

import pandas as pd
from flask import Blueprint, jsonify, render_template, request, send_file, session

from app.api.shared import auth_system, execute_query, login_requerido, obtener_fecha_hora_mexico
from app.auth_system import ECO_APPROVE_PERMISSION, ECO_CREATE_PERMISSION
from app.db_mysql import get_connection
from .control_bom_data import (
    _eco_excel_row_ref,
    _ks_fetch_current_bom_items,
    _ks_fetch_bom_items_multi,
    aprobar_eco,
    cancelar_eco,
    contar_ecos,
    crear_eco,
    crear_eco_desde_excel,
    crear_eco_familia_desde_excel,
    eliminar_eco,
    importar_items_eco_desde_dataframe,
    insertar_bom_desde_dataframe,
    listar_bom_por_modelo,
    listar_ecos,
    obtener_modelos_bom,
    obtener_diff_eco,
    obtener_ecn_ks,
    obtener_eco_detalle,
    obtener_scope_eco,
    resolver_familia,
    siguiente_revision_bom_eco,
)


bp = Blueprint("control_bom_api", __name__)


def _usuario_puede_crear_eco(username=None):
    return _usuario_tiene_permiso_eco(ECO_CREATE_PERMISSION, username)


def _usuario_puede_aprobar_eco(username=None):
    return _usuario_tiene_permiso_eco(ECO_APPROVE_PERMISSION, username)


def _usuario_tiene_permiso_eco(permission, username=None):
    username = username or session.get("usuario")
    if not username:
        return False
    if auth_system.obtener_rol_principal_usuario(username) == "superadmin":
        return True
    return auth_system.verificar_permiso_boton(
        username,
        permission["pagina"],
        permission["seccion"],
        permission["boton"],
    )


def _requiere_permiso_eco(permission, error_message):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not _usuario_tiene_permiso_eco(permission):
                return jsonify({
                    "success": False,
                    "error": error_message,
                    "permiso_requerido": (
                        f"{permission['pagina']} > "
                        f"{permission['seccion']} > "
                        f"{permission['boton']}"
                    ),
                }), 403
            return f(*args, **kwargs)

        return decorated_function

    return decorator


def requiere_permiso_crear_eco(f):
    return _requiere_permiso_eco(
        ECO_CREATE_PERMISSION,
        "No tienes permiso para crear o modificar ECOs.",
    )(f)


def requiere_permiso_aprobar_eco(f):
    return _requiere_permiso_eco(
        ECO_APPROVE_PERMISSION,
        "No tienes permiso para aprobar ECOs.",
    )(f)


def _render_control_bom_template():
    return render_template(
        "INFORMACION BASICA/CONTROL_DE_BOM.html",
        modelos=obtener_modelos_bom(),
        puede_crear_eco=_usuario_puede_crear_eco(),
        puede_aprobar_eco=_usuario_puede_aprobar_eco(),
    )


@bp.route("/informacion_basica/control_de_bom")
@login_requerido
def control_de_bom_ajax():
    """Render del fragmento AJAX de Control BOM."""
    try:
        return _render_control_bom_template()
    except Exception as e:
        print(f"Error al cargar template Control de BOM: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@bp.route("/control-bom-ajax")
@login_requerido
def control_bom_ajax():
    """Alias AJAX historico de Control BOM servido por el Blueprint."""
    try:
        return _render_control_bom_template()
    except Exception as e:
        print(f"Error al cargar template Control BOM AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


# Helpers _ks_current_bom_revision, _eco_for_part_revision y _bom_revision_catalog
# consolidados en app/api/shared/bom_revisions.py (2026-05-25). Antes vivian
# duplicados aqui y en app/routes.py.
from app.api.shared.bom_revisions import (  # noqa: E402, F401
    _ks_current_bom_revision,
    _eco_for_part_revision,
    _bom_revision_catalog,
)


@bp.route("/importar_excel_bom", methods=["POST"])
@login_requerido
def importar_excel_bom():
    if "file" not in request.files:
        return jsonify({"success": False, "error": "No se encontró el archivo"})

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"success": False, "error": "No se seleccionó ningún archivo"})

    try:
        print("--- Iniciando importación de BOM ---")
        df = pd.read_excel(file)

        # Imprime las columnas detectadas para depuración
        print(f"Columnas detectadas en el Excel: {df.columns.tolist()}")

        registrador = session.get("usuario", "desconocido")

        # Llamar a la nueva función de la base de datos
        resultado = insertar_bom_desde_dataframe(df, registrador)

        insertados = resultado.get("insertados", 0)
        omitidos = resultado.get("omitidos", 0)

        mensaje = f"Importación completada: {insertados} registros guardados."
        if omitidos > 0:
            mensaje += f" Se omitieron {omitidos} filas por no tener 'Modelo' o 'Número de parte'."

        print(f"--- Finalizando importación: {mensaje} ---")

        return jsonify({"success": True, "message": mensaje})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": f"Ocurrió un error: {str(e)}"})


@bp.route("/listar_modelos_bom", methods=["GET"])
@login_requerido
def listar_modelos_bom():
    """
    Devuelve la lista de modelos únicos disponibles en la tabla BOM
    """
    try:
        modelos = obtener_modelos_bom()
        return jsonify(modelos)
    except Exception as e:
        print(f"Error al obtener modelos BOM: {e}")
        return jsonify({"error": str(e)}), 500


@bp.route("/listar_bom", methods=["POST"])
@login_requerido
def listar_bom():
    """
    Lista los registros de BOM, opcionalmente filtrados por modelo, revision y classification
    """
    try:
        data = request.get_json()
        modelo = data.get("modelo", "todos") if data else "todos"
        classification = data.get("classification", None) if data else None
        bom_revision = data.get("bom_revision", None) if data else None

        bom_data = listar_bom_por_modelo(modelo, classification, bom_revision)
        return jsonify(bom_data)

    except Exception as e:
        print(f"Error al listar BOM: {e}")
        return jsonify({"error": str(e)}), 500


@bp.route("/api/bom/revisions", methods=["GET"])
@login_requerido
def api_bom_revisions():
    """Catalogo KS de revisiones disponibles para Control BOM."""
    try:
        modelo = (request.args.get("modelo") or request.args.get("part_no") or "").strip()
        if not modelo:
            return jsonify({"error": "modelo requerido"}), 400
        return jsonify({"success": True, "data": _bom_revision_catalog(modelo)})
    except Exception as e:
        print(f"Error al listar revisiones BOM KS: {e}")
        return jsonify({"error": str(e)}), 500


@bp.route("/consultar_bom", methods=["GET"])
@login_requerido
def consultar_bom():
    """
    Consulta datos de BOM con filtros GET para la interfaz de Control de salida
    """
    try:
        # Obtener filtros de los parámetros de consulta
        modelo = request.args.get("modelo", "").strip()
        numero_parte = request.args.get("numero_parte", "").strip()

        # Si no hay filtros específicos, obtener todos los datos
        if not modelo and not numero_parte:
            bom_data = listar_bom_por_modelo("todos")
        else:
            # Aplicar filtros
            bom_data = listar_bom_por_modelo(modelo if modelo else "todos")

            # Filtrar por número de parte si se proporciona
            if numero_parte and bom_data:
                bom_data = [
                    item
                    for item in bom_data
                    if numero_parte.lower() in str(item.get("numeroParte", "")).lower()
                ]

        return jsonify(bom_data)

    except Exception as e:
        print(f"Error al consultar BOM: {e}")
        return jsonify({"error": str(e)}), 500


def _json_safe_datetime(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def _serialize_eco_row(row):
    if not row:
        return row
    data = dict(row)
    for key in data:
        if (
            key in (
                "effective_at", "created_at", "approved_at", "updated_at",
                "source_updated_at", "synced_at", "sb_date",
            )
            or isinstance(data.get(key), Decimal)
        ):
            data[key] = _json_safe_datetime(data[key])
    return data


@bp.route("/api/ecos", methods=["GET"])
@login_requerido
def api_ecos_list():
    """Listar ECOs/cambios de ingenieria (historial unificado MES + KS)."""
    try:
        status = request.args.get("status")
        part_no = request.args.get("part_no")
        origen = request.args.get("origen")
        eco_no = request.args.get("eco_no")
        date_from = request.args.get("date_from")
        date_to = request.args.get("date_to")
        filters_active = any([status, part_no, origen, eco_no, date_from, date_to])
        page = max(1, int(request.args.get("page", 1) or 1))
        page_size = max(1, min(int(request.args.get("page_size", 500) or 500), 500))
        if filters_active:
            limit = page_size
            offset = (page - 1) * page_size
        else:
            page = 1
            page_size = 200
            limit = 200
            offset = 0
        rows = listar_ecos(
            status=status,
            part_no=part_no,
            limit=limit,
            origen=origen,
            eco_no=eco_no,
            date_from=date_from,
            date_to=date_to,
            offset=offset,
        )
        total = contar_ecos(
            status=status,
            part_no=part_no,
            origen=origen,
            eco_no=eco_no,
            date_from=date_from,
            date_to=date_to,
        )
        return jsonify({
            "success": True,
            "data": [_serialize_eco_row(r) for r in rows],
            "meta": {
                "total": total,
                "page": page,
                "page_size": page_size,
                "filters_active": filters_active,
                "default_latest_limit": 200,
                "has_more": (offset + len(rows)) < total,
            },
        })
    except Exception as e:
        print(f"Error listando ECOs: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/ecos/export", methods=["GET"])
@login_requerido
def api_ecos_export():
    """Exportar a Excel los ECOs filtrados."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        status = request.args.get("status")
        part_no = request.args.get("part_no")
        origen = request.args.get("origen")
        eco_no = request.args.get("eco_no")
        date_from = request.args.get("date_from")
        date_to = request.args.get("date_to")
        filters_active = any([status, part_no, origen, eco_no, date_from, date_to])
        rows = listar_ecos(
            status=status,
            part_no=part_no,
            limit=None if filters_active else 200,
            origen=origen,
            eco_no=eco_no,
            date_from=date_from,
            date_to=date_to,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "ECOs"
        headers = [
            "Origen", "ECO", "ID", "Modelo representante", "Familia/Scope",
            "Revision", "Fecha efectiva", "Estatus", "Items",
            "Creado por", "Aprobado por", "Creado", "Aprobado", "Actualizado",
        ]
        ws.append(headers)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            ws.append([
                row.get("origen"),
                row.get("eco_no"),
                row.get("id"),
                row.get("part_no"),
                row.get("scope_parts") or row.get("ks_family_prefix") or "",
                row.get("bom_revision"),
                row.get("effective_at"),
                row.get("status"),
                row.get("item_count"),
                row.get("created_by"),
                row.get("approved_by"),
                row.get("created_at"),
                row.get("approved_at"),
                row.get("updated_at"),
            ])

        ws.freeze_panes = "A2"
        widths = [12, 18, 18, 24, 38, 12, 22, 14, 10, 18, 18, 22, 22, 22]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

        meta = wb.create_sheet("_filtros")
        meta.append(["Filtro", "Valor"])
        meta.append(["tipo", origen or "TODOS"])
        meta.append(["estatus", status or "TODOS"])
        meta.append(["modelo", part_no or ""])
        meta.append(["eco_id", eco_no or ""])
        meta.append(["desde", date_from or ""])
        meta.append(["hasta", date_to or ""])
        meta.append(["total_exportado", len(rows)])
        meta.append(["nota", "Sin filtros exporta los ultimos 200; con filtros exporta todo el resultado filtrado."])

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M")
        return send_file(
            bio,
            as_attachment=True,
            download_name=f"ECOS_FILTRADOS_{ts}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        print(f"Error exportando ECOs: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/ecos/<int:eco_id>", methods=["GET"])
@login_requerido
def api_ecos_detail(eco_id):
    """Obtener ECO con items."""
    try:
        eco = obtener_eco_detalle(eco_id)
        if not eco:
            return jsonify({"success": False, "error": "ECO no encontrado"}), 404
        eco = _serialize_eco_row(eco)
        eco["items"] = [_serialize_eco_row(i) for i in eco.get("items", [])]
        eco["scope"] = [_serialize_eco_row(i) for i in eco.get("scope", [])]
        return jsonify({"success": True, "data": eco})
    except Exception as e:
        print(f"Error obteniendo ECO: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/ecn-ks/<int:hist_seq>", methods=["GET"])
@login_requerido
def api_ecn_ks_detail(hist_seq):
    """Obtener detalle de un ECN sincronizado desde K-system."""
    try:
        ecn = obtener_ecn_ks(hist_seq)
        if not ecn:
            return jsonify({"success": False, "error": "ECN no encontrado"}), 404
        ecn = _serialize_eco_row(ecn)
        return jsonify({"success": True, "data": ecn})
    except Exception as e:
        print(f"Error obteniendo ECN KS {hist_seq}: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/ecos", methods=["POST"])
@login_requerido
@requiere_permiso_crear_eco
def api_ecos_create():
    """Crear ECO DRAFT y copiar el BOM actual del modelo si se solicita."""
    try:
        data = request.get_json() or {}
        usuario = session.get("usuario", "desconocido")
        eco = crear_eco(data, usuario)
        if not eco:
            return jsonify({"success": False, "error": "No se pudo crear el ECO"}), 500
        eco = _serialize_eco_row(eco)
        eco["items"] = [_serialize_eco_row(i) for i in eco.get("items", [])]
        return jsonify({"success": True, "data": eco}), 201
    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        print(f"Error creando ECO: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


BOM_EXCEL_COLUMNS = [
    ("__row_id", "id"),
    ("item_no", "item_no"),
    ("item_name", "item_name"),
    ("spec", "spec"),
    ("qty", "qty"),
    ("unit", "unit"),
    ("location_text", "location_text"),
    ("maker", "maker"),
    ("supplier", "supplier"),
    ("item_class", "item_class"),
    ("item_process", "item_process"),
    ("process_name", "process_name"),
    ("valid_from", "valid_from"),
    ("valid_to", "valid_to"),
    ("is_alternate", "is_alternate"),
    ("alt_item_no", "alt_item_no"),
    ("alt_item_name", "alt_item_name"),
    ("alt_spec", "alt_spec"),
    ("alt_maker", "alt_maker"),
    ("remark", "remark"),
    ("bom_level", "bom_level"),
    ("item_seq", "item_seq"),
]


BOM_KS_TEMPLATE_HEADERS = [
    "BOM level", "Matrl No.", "Matrl Name", "Matrl Spec",
    "Total Request AMT", "Location", "Maker", "세부공정",
    "대체자재번호", "품목공정", "Matrl Abbr", "Matrl Eng Name",
    "Matrl Eng Abbr", "Matrl Remark", "Unit", "Item Asset Category",
    "Remark", "Appl Start Date", "Closing Date of Appl",
    "PurchaseUnit Price", "Purchase AMT", "Customer", "Currency",
    "Current stock", "대체자재명", "대체자재규격", "자재대분류",
    "대체자재Maker", "IC(PIN)", "품목상태", "대표단가",
]


BOM_EXCEL_HEADER_ALIASES = {
    "__row_id": ["row_id", "id", "__row_key", "row key"],
    "item_no": ["matrl no.", "matrl no", "material no.", "material no", "material", "part no", "part_no"],
    "item_name": ["matrl name", "material name", "name"],
    "spec": ["matrl spec", "material spec", "specification"],
    "qty": ["total request amt", "total request amount", "request amt", "qty", "quantity", "cantidad"],
    "unit": ["unit", "uom"],
    "location_text": ["location", "ubicacion", "ubicación"],
    "maker": ["maker", "manufacturer", "fabricante"],
    "supplier": ["supplier", "customer", "vendor", "proveedor"],
    "item_class": ["item asset category", "asset category", "item class", "classification", "자재대분류"],
    "item_process": ["품목공정", "item process", "process"],
    "process_name": ["세부공정", "process name", "detail process"],
    "valid_from": ["appl start date", "valid from", "start date"],
    "valid_to": ["closing date of appl", "valid to", "end date"],
    "is_alternate": ["is alternate", "alternate", "alt"],
    "alt_item_no": ["대체자재번호", "alt item no", "alternate item no"],
    "alt_item_name": ["대체자재명", "alt item name", "alternate item name"],
    "alt_spec": ["대체자재규격", "alt spec", "alternate spec"],
    "alt_maker": ["대체자재maker", "대체자재 maker", "alt maker", "alternate maker"],
    "remark": ["remark", "matrl remark", "material remark"],
    "bom_level": ["bom level", "bom_level", "level"],
    "item_seq": ["item seq", "item_seq", "seq"],
}


def _excel_header_token(value):
    return "".join(
        char
        for char in str(value or "").strip().lower()
        if char.isalnum()
    )


def _build_bom_excel_idx_map(headers):
    normalized = {}
    for idx, header in enumerate(headers):
        token = _excel_header_token(header)
        if token and token not in normalized:
            normalized[token] = idx

    idx_map = {}
    for label, _key in BOM_EXCEL_COLUMNS:
        candidates = [label] + BOM_EXCEL_HEADER_ALIASES.get(label, [])
        for candidate in candidates:
            token = _excel_header_token(candidate)
            if token in normalized:
                idx_map[label] = normalized[token]
                break
    return idx_map


def _normalize_bom_excel_row(obj):
    process_name = str(obj.get("process_name") or "").strip()
    if not obj.get("item_process") and process_name and not process_name.replace("/", "").isdigit():
        obj["item_process"] = process_name
    if not obj.get("process_name") and obj.get("item_process"):
        obj["process_name"] = obj.get("item_process")
    return obj


def _infer_part_no_from_filename(filename):
    match = re.search(r"([A-Z]{2,}\d{5,})", str(filename or "").upper())
    return match.group(1) if match else ""


@bp.route("/api/bom/download-excel", methods=["GET"])
@login_requerido
@requiere_permiso_crear_eco
def api_bom_download_excel():
    """Descargar BOM actual de un modelo en formato Excel editable."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    part_no = (request.args.get("part_no") or "").strip().upper()
    bom_rev = (request.args.get("bom_rev") or "").strip().upper() or None
    mode = (request.args.get("mode") or "").strip().upper()
    new_bom_template = mode in ("NEW", "NEW_BOM", "NUEVO", "NUEVO_BOM")
    if not part_no:
        return jsonify({"success": False, "error": "part_no requerido"}), 400

    try:
        rows = [] if new_bom_template else (_ks_fetch_current_bom_items(part_no, bom_rev) or [])

        wb = Workbook()
        ws = wb.active
        ws.title = "BOM"

        headers = BOM_KS_TEMPLATE_HEADERS if new_bom_template else [label for label, _key in BOM_EXCEL_COLUMNS]
        ws.append(headers)
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        if not new_bom_template:
            for row in rows:
                ws.append([
                    _eco_excel_row_ref(row) if label == "__row_id" else row.get(key)
                    for label, key in BOM_EXCEL_COLUMNS
                ])

        ws.freeze_panes = "A2" if new_bom_template else "B2"
        if not new_bom_template:
            ws.column_dimensions["A"].hidden = True
        start_width_col = 1 if new_bom_template else 2
        for col_idx in range(start_width_col, len(headers) + 1):
            header = str(ws.cell(row=1, column=col_idx).value or "")
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(14, min(len(header) + 4, 24))

        ws_meta = wb.create_sheet("_meta")
        ws_meta.append(["part_no", part_no])
        ws_meta.append(["bom_rev", rows[0].get("bom_rev") if rows else (bom_rev or "")])
        ws_meta.append(["mode", "NEW_BOM" if new_bom_template else "ECO"])
        ws_meta.append(["exported_at", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")])
        ws_meta.sheet_state = "hidden"

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M")
        download_name = (
            f"PLANTILLA_BOM_NUEVO_{part_no}_{ts}.xlsx"
            if new_bom_template
            else f"BOM_{part_no}_{ts}.xlsx"
        )
        return send_file(
            bio,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        print(f"Error descargando BOM Excel: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/bom/next-eco-revision", methods=["GET"])
@login_requerido
@requiere_permiso_crear_eco
def api_bom_next_eco_revision():
    """Calcular la siguiente revision KS que un ECO nuevo reservara."""
    part_no = (request.args.get("part_no") or "").strip().upper()
    scope_parts = [
        part.strip().upper()
        for part in (request.args.get("scope_parts") or "").split(",")
        if part.strip()
    ]
    part_numbers = scope_parts or ([part_no] if part_no else [])
    if not part_numbers:
        return jsonify({"success": False, "error": "part_no o scope_parts requerido"}), 400
    try:
        revision = siguiente_revision_bom_eco(part_numbers)
        return jsonify({
            "success": True,
            "data": {
                "bom_revision": revision,
                "part_numbers": part_numbers,
            },
        })
    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        print(f"Error calculando siguiente revision ECO: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/bom/resolve-family", methods=["GET"])
@login_requerido
@requiere_permiso_crear_eco
def api_bom_resolve_family():
    """Resolver una familia + sufijos a part_no existentes en ks_part_catalog."""
    family = (request.args.get("family") or "").strip()
    suffixes = request.args.get("suffixes") or ""
    if not family:
        return jsonify({"success": False, "error": "family requerido"}), 400
    try:
        result = resolver_familia(family, suffixes)
        return jsonify({"success": True, "data": result})
    except Exception as e:
        print(f"Error resolviendo familia: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


def _build_family_excel(part_numbers, bom_revision=None):
    """Construir Workbook con union de items de los part_no del scope.

    Cada fila lleva __row_key = item_no|bom_level y modelos_afectados = lista de part_no.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    bom_by_part = _ks_fetch_bom_items_multi(part_numbers, bom_revision)

    # Indexar por (item_no, bom_level)
    union = {}  # key -> {row_canonica, modelos_afectados:set, by_part:{pn:row}}
    for pn, rows in bom_by_part.items():
        for r in rows:
            key = f"{(r.get('item_no') or '').upper()}|{(r.get('bom_level') or '').strip()}"
            if not key.strip('|'):
                continue
            if key not in union:
                union[key] = {"row": dict(r), "modelos": set(), "by_part": {}}
            union[key]["modelos"].add(pn)
            union[key]["by_part"][pn] = r

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM_FAMILIA"

    columns = [
        ("__row_key", "row_key"),
        ("modelos_afectados", "modelos"),
        ("item_no", "item_no"),
        ("item_name", "item_name"),
        ("spec", "spec"),
        ("qty", "qty"),
        ("unit", "unit"),
        ("location_text", "location_text"),
        ("maker", "maker"),
        ("supplier", "supplier"),
        ("item_class", "item_class"),
        ("item_process", "item_process"),
        ("process_name", "process_name"),
        ("valid_from", "valid_from"),
        ("valid_to", "valid_to"),
        ("is_alternate", "is_alternate"),
        ("alt_item_no", "alt_item_no"),
        ("alt_item_name", "alt_item_name"),
        ("alt_spec", "alt_spec"),
        ("alt_maker", "alt_maker"),
        ("remark", "remark"),
        ("bom_level", "bom_level"),
        ("item_seq", "item_seq"),
    ]
    headers = [label for label, _ in columns]
    ws.append(headers)
    header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    sorted_keys = sorted(union.keys(), key=lambda k: (union[k]["row"].get("bom_level") or "", k))
    origin_rows = []
    for key in sorted_keys:
        entry = union[key]
        row = entry["row"]
        modelos = ",".join(sorted(entry["modelos"]))
        excel_row = []
        for label, source_key in columns:
            if source_key == "row_key":
                excel_row.append(key)
            elif source_key == "modelos":
                excel_row.append(modelos)
            else:
                excel_row.append(row.get(source_key))
        ws.append(excel_row)
        origin_rows.append([key] + [row.get(field) for field in FAMILY_EXCEL_ORIGIN_FIELDS])

    ws.freeze_panes = "C2"
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["B"].width = 40
    for col_idx in range(3, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18

    ws_meta = wb.create_sheet("_meta")
    ws_meta.append(["scope_kind", "FAMILY"])
    ws_meta.append(["part_numbers", ",".join(part_numbers)])
    ws_meta.append(["bom_rev", bom_revision or ""])
    ws_meta.append(["exported_at", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")])
    ws_meta.sheet_state = "hidden"

    ws_origin = wb.create_sheet("_family_origin")
    ws_origin.append(["__row_key"] + FAMILY_EXCEL_ORIGIN_FIELDS)
    for origin_row in origin_rows:
        ws_origin.append(origin_row)
    ws_origin.sheet_state = "hidden"

    return wb


@bp.route("/api/bom/download-excel-family", methods=["GET"])
@login_requerido
@requiere_permiso_crear_eco
def api_bom_download_excel_family():
    """Descargar Excel multi-modelo: union de items de los modelos del scope."""
    family = (request.args.get("family") or "").strip().upper()
    suffixes = request.args.get("suffixes") or ""
    bom_rev = (request.args.get("bom_rev") or "").strip().upper() or None
    if not family:
        return jsonify({"success": False, "error": "family requerido"}), 400
    try:
        resolved = resolver_familia(family, suffixes)
        parts = [p["part_no"] for p in resolved.get("parts", [])]
        if not parts:
            return jsonify({"success": False, "error": "Ningun part_no resuelto", "data": resolved}), 400

        wb = _build_family_excel(parts, bom_rev)
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M")
        return send_file(
            bio,
            as_attachment=True,
            download_name=f"BOM_FAMILIA_{family}_{ts}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        print(f"Error descargando BOM familia: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/ecos/from-excel", methods=["POST"])
@login_requerido
@requiere_permiso_crear_eco
def api_ecos_from_excel():
    """Crear ECO DRAFT desde Excel modificado: valida, calcula diff y persiste."""
    from openpyxl import load_workbook

    if "file" not in request.files:
        return jsonify({"success": False, "errors": ["No se encontro el archivo"]}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"success": False, "errors": ["No se selecciono ningun archivo"]}), 400

    metadata = {
        "eco_no": request.form.get("eco_no"),
        "part_no": request.form.get("part_no") or _infer_part_no_from_filename(file.filename),
        "effective_at": request.form.get("effective_at"),
        "item_name": request.form.get("item_name"),
        "notes": request.form.get("notes"),
        "bom_mode": request.form.get("bom_mode"),
    }

    try:
        wb = load_workbook(file, data_only=True)
        if "BOM" in wb.sheetnames:
            ws = wb["BOM"]
        else:
            ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return jsonify({"success": False, "errors": ["El Excel no tiene encabezados"]}), 400
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        expected = [label for label, _ in BOM_EXCEL_COLUMNS]
        idx_map = _build_bom_excel_idx_map(headers)
        required = ["item_no", "qty", "bom_level"]
        missing = [col for col in required if col not in idx_map]
        if missing:
            return jsonify({
                "success": False,
                "errors": [
                    "Faltan columnas requeridas en el Excel: "
                    f"{', '.join(missing)}. Tambien se aceptan encabezados KS como "
                    "Matrl No., Total Request AMT y BOM level."
                ],
            }), 400

        excel_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(v in (None, "") for v in row):
                continue
            obj = {}
            for col in expected:
                i = idx_map.get(col)
                obj[col] = row[i] if i is not None and i < len(row) else None
            excel_rows.append(_normalize_bom_excel_row(obj))

        usuario = session.get("usuario", "desconocido")
        result = crear_eco_desde_excel(metadata, excel_rows, usuario)
        status_code = 201 if result.get("success") else 400
        return jsonify(result), status_code
    except Exception as e:
        print(f"Error creando ECO desde Excel: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "errors": [str(e)]}), 500


FAMILY_EXCEL_ORIGIN_FIELDS = [
    "item_no", "item_name", "spec", "qty", "unit", "location_text",
    "maker", "supplier", "item_class", "item_process", "process_name",
    "valid_from", "valid_to", "is_alternate", "alt_item_no",
    "alt_item_name", "alt_spec", "alt_maker", "remark",
]


FAMILY_EXCEL_COLUMNS = [
    "__row_key", "modelos_afectados", "item_no", "item_name", "spec", "qty",
    "unit", "location_text", "maker", "supplier", "item_class", "item_process",
    "process_name", "valid_from", "valid_to", "is_alternate", "alt_item_no",
    "alt_item_name", "alt_spec", "alt_maker", "remark", "bom_level", "item_seq",
]


@bp.route("/api/ecos/from-excel-family", methods=["POST"])
@login_requerido
@requiere_permiso_crear_eco
def api_ecos_from_excel_family():
    """Crear ECO de familia DRAFT desde Excel multi-modelo."""
    from openpyxl import load_workbook

    if "file" not in request.files:
        return jsonify({"success": False, "errors": ["No se encontro el archivo"]}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"success": False, "errors": ["No se selecciono ningun archivo"]}), 400

    metadata = {
        "eco_no": request.form.get("eco_no"),
        "family_prefix": request.form.get("family_prefix"),
        "effective_at": request.form.get("effective_at"),
        "item_name": request.form.get("item_name"),
        "notes": request.form.get("notes"),
    }
    scope_parts_raw = request.form.get("scope_parts") or ""
    scope_parts = [p.strip().upper() for p in scope_parts_raw.split(",") if p.strip()]
    if not scope_parts:
        return jsonify({"success": False, "errors": ["scope_parts requerido"]}), 400

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb["BOM_FAMILIA"] if "BOM_FAMILIA" in wb.sheetnames else wb.active
        if "_family_origin" not in wb.sheetnames:
            return jsonify({
                "success": False,
                "errors": [
                    "Este Excel de familia no tiene hoja de control _family_origin. Descargue de nuevo el BOM de familia y vuelva a aplicar sus cambios para detectar el diff completo."
                ],
            }), 400

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return jsonify({"success": False, "errors": ["El Excel no tiene encabezados"]}), 400
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        missing = [col for col in FAMILY_EXCEL_COLUMNS if col not in headers]
        if missing:
            return jsonify({
                "success": False,
                "errors": [f"Faltan columnas en el Excel: {', '.join(missing)}"],
            }), 400

        idx_map = {col: headers.index(col) for col in FAMILY_EXCEL_COLUMNS}
        origin_by_key = {}
        ws_origin = wb["_family_origin"]
        origin_header = next(ws_origin.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if origin_header:
            origin_headers = [str(h).strip() if h is not None else "" for h in origin_header]
            origin_idx = {col: origin_headers.index(col) for col in origin_headers if col}
            missing_origin = [
                col for col in (["__row_key"] + FAMILY_EXCEL_ORIGIN_FIELDS)
                if col not in origin_idx
            ]
            if missing_origin:
                return jsonify({
                    "success": False,
                    "errors": [
                        "Este Excel de familia tiene formato de origen anterior. Descargue de nuevo el BOM de familia para detectar todos los campos editables."
                    ],
                }), 400
            if "__row_key" in origin_idx:
                for origin_row in ws_origin.iter_rows(min_row=2, values_only=True):
                    row_key_idx = origin_idx["__row_key"]
                    row_key = origin_row[row_key_idx] if row_key_idx < len(origin_row) else None
                    row_key = str(row_key or "").strip()
                    if not row_key:
                        continue
                    origin_by_key[row_key] = {}
                    for field in FAMILY_EXCEL_ORIGIN_FIELDS:
                        if field not in origin_idx:
                            continue
                        field_idx = origin_idx[field]
                        origin_by_key[row_key][field] = origin_row[field_idx] if field_idx < len(origin_row) else None
        if not origin_by_key:
            return jsonify({
                "success": False,
                "errors": [
                    "La hoja _family_origin esta vacia. Descargue de nuevo el BOM de familia antes de crear el ECO."
                ],
            }), 400

        excel_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(v in (None, "") for v in row):
                continue
            obj = {}
            for col in FAMILY_EXCEL_COLUMNS:
                i = idx_map[col]
                obj[col] = row[i] if i < len(row) else None
            row_key = str(obj.get("__row_key") or "").strip()
            if row_key in origin_by_key:
                obj["__origin_values"] = origin_by_key[row_key]
            excel_rows.append(obj)

        usuario = session.get("usuario", "desconocido")
        result = crear_eco_familia_desde_excel(metadata, excel_rows, scope_parts, usuario)
        status_code = 201 if result.get("success") else 400
        return jsonify(result), status_code
    except Exception as e:
        print(f"Error creando ECO familia desde Excel: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "errors": [str(e)]}), 500


@bp.route("/api/ecos/<int:eco_id>/scope", methods=["GET"])
@login_requerido
def api_ecos_scope(eco_id):
    """Listar scope (part_no afectados) de un ECO."""
    try:
        rows = obtener_scope_eco(eco_id)
        return jsonify({"success": True, "data": rows})
    except Exception as e:
        print(f"Error obteniendo scope ECO {eco_id}: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/ecos/<int:eco_id>/diff", methods=["GET"])
@login_requerido
def api_ecos_diff(eco_id):
    """Obtener diff persistido de un ECO."""
    try:
        rows = obtener_diff_eco(eco_id)
        rows = [_serialize_eco_row(r) for r in rows]
        added = [r for r in rows if r.get("action") == "ADD"]
        modified = [r for r in rows if r.get("action") == "MODIFY"]
        removed = [r for r in rows if r.get("action") == "REMOVE"]

        per_part = {}
        for r in rows:
            pn = r.get("part_no") or ""
            entry = per_part.setdefault(pn, {"added": 0, "modified": 0, "removed": 0})
            action = r.get("action")
            if action == "ADD":
                entry["added"] += 1
            elif action == "MODIFY":
                entry["modified"] += 1
            elif action == "REMOVE":
                entry["removed"] += 1

        return jsonify({
            "success": True,
            "data": {
                "added": added,
                "modified": modified,
                "removed": removed,
                "counts": {
                    "added": len(added),
                    "modified": len(modified),
                    "removed": len(removed),
                },
                "per_part": per_part,
            },
        })
    except Exception as e:
        print(f"Error obteniendo diff ECO {eco_id}: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/ecos/<int:eco_id>/items/import", methods=["POST"])
@login_requerido
@requiere_permiso_crear_eco
def api_ecos_import_items(eco_id):
    """Reemplazar items del ECO con un Excel completo de BOM."""
    if "file" not in request.files:
        return jsonify({"success": False, "error": "No se encontro el archivo"}), 400
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"success": False, "error": "No se selecciono ningun archivo"}), 400

    try:
        df = pd.read_excel(file)
        result = importar_items_eco_desde_dataframe(eco_id, df)
        return jsonify({"success": True, "data": result})
    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        print(f"Error importando items ECO: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/ecos/<int:eco_id>/approve", methods=["POST"])
@login_requerido
@requiere_permiso_aprobar_eco
def api_ecos_approve(eco_id):
    """Aprobar ECO DRAFT. Desde este punto queda inmutable."""
    try:
        usuario = session.get("usuario", "desconocido")
        result = aprobar_eco(eco_id, usuario)
        if not result.get("success"):
            return jsonify({"success": False, "errors": result.get("errors", [])}), 400
        eco = obtener_eco_detalle(eco_id)
        eco = _serialize_eco_row(eco)
        eco["items"] = [_serialize_eco_row(i) for i in eco.get("items", [])]
        return jsonify({"success": True, "data": eco})
    except Exception as e:
        print(f"Error aprobando ECO: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/ecos/<int:eco_id>/cancel", methods=["POST"])
@login_requerido
@requiere_permiso_crear_eco
def api_ecos_cancel(eco_id):
    """Cancelar ECO DRAFT; los aprobados no se modifican."""
    try:
        usuario = session.get("usuario", "desconocido")
        result = cancelar_eco(eco_id, usuario)
        if not result.get("success"):
            return jsonify({"success": False, "error": result.get("error", "No se pudo cancelar")}), 400
        return jsonify({"success": True})
    except Exception as e:
        print(f"Error cancelando ECO: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/api/ecos/<int:eco_id>", methods=["DELETE"])
@login_requerido
@requiere_permiso_crear_eco
def api_ecos_delete(eco_id):
    """Eliminar fisicamente un ECO no aprobado."""
    try:
        result = eliminar_eco(eco_id)
        if not result.get("success"):
            return jsonify({"success": False, "error": result.get("error", "No se pudo borrar")}), 400
        return jsonify({"success": True})
    except Exception as e:
        print(f"Error borrando ECO: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

def exportar_bom_a_excel(modelo=None, classification=None, bom_revision=None):
    """
    Función auxiliar para exportar datos de BOM a Excel con filtros opcionales
    """
    try:
        import os
        import tempfile

        result = listar_bom_por_modelo(modelo or "todos", classification, bom_revision)

        if not result:
            print(
                "No se encontraron datos de BOM para exportar "
                f"(modelo={modelo}, classification={classification}, bom_revision={bom_revision})"
            )
            return None

        # Crear DataFrame
        df = pd.DataFrame(result)
        df = df.drop(columns=["posicionAssy"], errors="ignore")

        # Renombrar columnas para mejor legibilidad
        column_mapping = {
            "modelo": "Modelo",
            "numeroParte": "Número de Parte",
            "side": "Side",
            "tipoMaterial": "Tipo de Material",
            "classification": "Classification",
            "especificacionMaterial": "Especificación de Material",
            "vender": "Maker",
            "cantidadTotal": "Cantidad Total",
            "ubicacion": "Ubicación",
            "materialSustituto": "Material Sustituto",
            "materialOriginal": "Material Original",
            "registrador": "Registrador",
            "fechaRegistro": "Fecha de Registro",
            "bomRevision": "BOM Rev",
        }

        df = df.rename(columns=column_mapping)

        # Crear archivo temporal
        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, mode="wb")

        # Escribir a Excel
        with pd.ExcelWriter(temp_file.name, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="BOM_Data", index=False)

            # Obtener el workbook y worksheet para formateo
            workbook = writer.book
            worksheet = writer.sheets["BOM_Data"]

            # Ajustar ancho de columnas automáticamente
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
                worksheet.column_dimensions[column_letter].width = adjusted_width

        temp_file.close()
        return temp_file.name

    except Exception as e:
        print(f"Error en exportar_bom_a_excel: {e}")
        traceback.print_exc()
        return None


@bp.route("/exportar_excel_bom", methods=["GET"])
@login_requerido
def exportar_excel_bom():
    """
    Exporta datos de BOM a un archivo Excel, filtrados por modelo, revision y classification
    """
    try:
        # Obtener parámetros de consulta
        modelo = request.args.get("modelo", None)
        classification = request.args.get("classification", None)
        bom_revision = request.args.get("bom_revision", None)

        if modelo and modelo.strip() and modelo != "todos":
            # Exportar modelo específico con filtro opcional de classification
            archivo_temp = exportar_bom_a_excel(modelo, classification, bom_revision)

            # Construir nombre del archivo
            nombre_base = f"bom_export_{modelo}"
            if bom_revision:
                nombre_base += f"_{bom_revision}"
            if classification and classification != "TODOS":
                nombre_base += f"_{classification}"
            download_name = (
                f"{nombre_base}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
        else:
            # Exportar todos con filtro opcional
            archivo_temp = exportar_bom_a_excel(None, classification, bom_revision)
            nombre_base = "bom_export_todos"
            if classification and classification != "TODOS":
                nombre_base += f"_{classification}"
            download_name = (
                f"{nombre_base}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )

        if archivo_temp:
            return send_file(
                archivo_temp,
                as_attachment=True,
                download_name=download_name,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            return jsonify({"error": "Error al generar el archivo Excel"}), 500

    except Exception as e:
        print(f"Error al exportar BOM: {e}")
        return jsonify({"error": str(e)}), 500


@bp.route("/api/bom/update", methods=["POST"])
@login_requerido
def api_bom_update():
    """Mantener bloqueada la edicion directa; BOM KS se publica via ECO."""
    return jsonify({
        "success": False,
        "error": "La edicion directa de bom esta deshabilitada. Use Crear ECO para modificar/publicar el BOM KS.",
    }), 409


@bp.route("/api/bom/update-posiciones-assy", methods=["POST"])
@login_requerido
def api_bom_update_posiciones_assy():
    """Mantener bloqueada la edicion legacy de posiciones ASSY."""
    return jsonify({
        "success": False,
        "error": "La edicion directa de posiciones en bom esta deshabilitada. Use Crear ECO.",
    }), 409


# ---------------------------------------------------------------------------
# Fase 4 (2026-05-28): /api/bom-smt-data migrado desde routes.py.
# Lee bom_smt_f / bom_smt_r para el panel "Control de Operacion SMT" via
# mapeo de linea SMT (A/B/C/D) a numero de linea fisica (2/2/3/4).
# ---------------------------------------------------------------------------


@bp.route("/api/bom-smt-data", methods=["GET"])
@login_requerido
def api_bom_smt_data():
    """API para obtener datos del BOM SMT basado en linea y modelo"""
    try:
        linea = request.args.get("linea", "")
        model_code = request.args.get("model_code", "")

        if not linea or not model_code:
            return jsonify(
                {"success": False, "error": "Linea y modelo son requeridos"}
            ), 400

        print(f"API BOM SMT - Filtros:")
        print(f"  Linea: {linea}")
        print(f"  Modelo: {model_code}")

        conn = get_connection()
        cursor = conn.cursor()

        # Mapear linea SMT a numero de linea
        mapeo_lineas = {
            "SMT A": "2",
            "SMT B": "2",
            "SMT C": "3",
            "SMT D": "4",
            "1LINE": "2",
            "2LINE": "2",
            "3LINE": "3",
            "4LINE": "4",
        }

        linea_numero = mapeo_lineas.get(linea, "2")

        # Consultar ambas tablas (bom_smt_f y bom_smt_r) - solo elementos con cantidad > 0
        query_f = """
            SELECT
                id, linea, model_code, mounter, slot, material_code,
                description, feeder_info, qty, raw_filename,
                created_at, updated_at, 'FRONT' as tabla_tipo
            FROM bom_smt_f
            WHERE linea = %s AND model_code LIKE %s AND qty > 0
            ORDER BY mounter, slot
        """

        query_r = """
            SELECT
                id, linea, model_code, mounter, slot, material_code,
                description, feeder_info, qty, raw_filename,
                created_at, updated_at, 'REAR' as tabla_tipo
            FROM bom_smt_r
            WHERE linea = %s AND model_code LIKE %s AND qty > 0
            ORDER BY mounter, slot
        """

        # Buscar por modelo (puede contener EBR)
        model_pattern = f"%{model_code}%"

        cursor.execute(query_f, [linea_numero, model_pattern])
        resultados_f = cursor.fetchall()

        cursor.execute(query_r, [linea_numero, model_pattern])
        resultados_r = cursor.fetchall()

        todos_resultados = list(resultados_f) + list(resultados_r)

        print(
            f"Encontrados {len(todos_resultados)} registros BOM ({len(resultados_f)} F + {len(resultados_r)} R)"
        )
        print(
            f"Parametros de busqueda - Linea numero: {linea_numero}, Patron modelo: {model_pattern}"
        )

        # Formatear datos - solo incluir elementos con cantidad > 0
        formatted_data = []
        for row in todos_resultados:
            try:
                qty_value = row[8] if len(row) > 8 else 0

                if qty_value <= 0:
                    continue

                formatted_row = {
                    "id": row[0] if len(row) > 0 else "",
                    "linea": row[1] if len(row) > 1 else "",
                    "model_code": row[2] if len(row) > 2 else "",
                    "mounter": row[3] if len(row) > 3 else "",
                    "slot": row[4] if len(row) > 4 else "",
                    "material_code": row[5] if len(row) > 5 else "",
                    "description": row[6] if len(row) > 6 else "",
                    "feeder_info": row[7] if len(row) > 7 else "",
                    "qty": qty_value,
                    "raw_filename": row[9] if len(row) > 9 else "",
                    "created_at": str(row[10]) if len(row) > 10 and row[10] else "",
                    "updated_at": str(row[11]) if len(row) > 11 and row[11] else "",
                    "tabla_tipo": row[12] if len(row) > 12 else "",
                    "status": "pending",
                }
                formatted_data.append(formatted_row)

            except Exception as row_error:
                print(f"Error procesando fila BOM: {row_error}")
                continue

        cursor.close()
        conn.close()

        print(f"BOM filtrado: {len(formatted_data)} elementos con qty > 0")

        return jsonify(
            {
                "success": True,
                "data": formatted_data,
                "total": len(formatted_data),
                "linea": linea,
                "model_code": model_code,
                "total_raw": len(todos_resultados),
                "total_filtered": len(formatted_data),
            }
        )

    except Exception as e:
        print(f"Error en api_bom_smt_data: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
