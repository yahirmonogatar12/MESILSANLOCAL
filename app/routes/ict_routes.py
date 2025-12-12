# -*- coding: utf-8 -*-
"""
ICT Routes - Rutas para el historial ICT y defectos detallados
"""

from flask import Blueprint, request, jsonify, render_template, session, send_file
from functools import wraps
from datetime import datetime, date, time as dt_time, timedelta
from io import BytesIO
from ..database.db_mysql import execute_query

ict_bp = Blueprint('ict', __name__)

# ============================================================================
# DECORADOR DE AUTENTICACIÓN
# ============================================================================

def login_requerido(f):
    """Decorador para verificar que el usuario está autenticado"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario' not in session:
            return jsonify({'error': 'No autorizado', 'redirect': '/login'}), 401
        return f(*args, **kwargs)
    return decorated_function


# ============================================================================
# FUNCIONES AUXILIARES
# ============================================================================

def convertir_linea_smt(linea_nombre):
    """
    Convierte nombres de línea SMT a formato de BD
    SMT A = 1line
    SMT B = 2line  
    SMT C = 3line
    SMT D = 4line
    """
    conversion = {
        'SMT A': '1line',
        'SMT B': '2line', 
        'SMT C': '3line',
        'SMT D': '4line'
    }
    return conversion.get(linea_nombre, linea_nombre)


def convertir_linea_smt_reverso(linea_bd):
    """
    Convierte formato de BD a nombres de línea SMT
    1line = SMT A
    2line = SMT B
    3line = SMT C  
    4line = SMT D
    """
    conversion = {
        '1line': 'SMT A',
        '2line': 'SMT B',
        '3line': 'SMT C', 
        '4line': 'SMT D'
    }
    return conversion.get(linea_bd, linea_bd)


def _ict_format_row(row):
    """Convertir campos fecha/hora a cadenas serializables."""
    if not row:
        return {}

    formatted = {}
    for key, value in row.items():
        if isinstance(value, datetime):
            formatted[key] = value.isoformat(sep=' ')
        elif isinstance(value, date):
            formatted[key] = value.isoformat()
        elif isinstance(value, dt_time):
            formatted[key] = value.strftime('%H:%M:%S')
        elif isinstance(value, timedelta):
            formatted[key] = str(value)
        else:
            formatted[key] = value
    return formatted


# ============================================================================
# PÁGINA PRINCIPAL ICT
# ============================================================================

@ict_bp.route('/historial-ict')
@ict_bp.route('/ict/front-full-defects2')
@login_requerido
def ict_front_full_defects2():
    """Vista principal del historial ICT con defectos detallados."""
    try:
        return render_template('Control de resultados/history_ict.html')
    except Exception as e:
        print(f"Error al cargar History ICT: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


# ============================================================================
# API: DATOS ICT
# ============================================================================

@ict_bp.route('/api/ict/data')
@login_requerido
def ict_data_api():
    """Obtener registros recientes del historial ICT con filtros opcionales."""
    try:
        fecha = request.args.get('fecha', '').strip()
        linea = request.args.get('linea', '').strip()
        resultado = request.args.get('resultado', '').strip()
        barcode_like = request.args.get('barcode_like', '').strip()

        sql = (
            "SELECT fecha, TIME(ts) AS hora, linea, ict, resultado, no_parte, barcode, "
            "ts, fuente_archivo, defect_code, defect_valor "
            "FROM history_ict WHERE 1=1"
        )
        params = []

        if fecha:
            sql += " AND fecha=%s"
            params.append(fecha)
        if linea:
            sql += " AND linea=%s"
            params.append(linea)
        if resultado:
            sql += " AND resultado=%s"
            params.append(resultado)
        if barcode_like:
            sql += " AND barcode LIKE %s"
            params.append(f"%{barcode_like}%")

        sql += " ORDER BY ts DESC LIMIT 500"
        rows = execute_query(sql, tuple(params), fetch='all') or []

        return jsonify([_ict_format_row(row) for row in rows])
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


# ============================================================================
# API: DEFECTOS ICT
# ============================================================================

@ict_bp.route('/api/ict/defects')
@login_requerido
def ict_defects_api():
    """Obtener defectos asociados a un barcode específico."""
    barcode = request.args.get("barcode", "").strip()
    if not barcode:
        return jsonify([])

    try:
        sql = (
            "SELECT d.barcode, h.linea, h.ict, d.componente, d.pinref, d.act_value, d.act_unit, "
            "d.std_value, d.std_unit, d.meas_value, "
            "d.m_value, d.r_value, d.hlim_pct, d.llim_pct, "
            "d.hp_value, d.lp_value, d.ws_value, d.ds_value, d.rc_value, "
            "d.p_flag, d.j_flag, d.resultado_local, d.defecto_tipo, d.ts "
            "FROM history_ict_defects d "
            "LEFT JOIN history_ict h ON d.barcode COLLATE utf8mb4_unicode_ci = h.barcode COLLATE utf8mb4_unicode_ci "
            "AND d.ts = h.ts "
            "WHERE d.barcode=%s "
            "ORDER BY d.ts DESC, d.componente LIMIT 1000"
        )

        rows = execute_query(sql, (barcode,), fetch='all') or []
        return jsonify([_ict_format_row(row) for row in rows])
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


# ============================================================================
# API: EXPORTAR ICT A EXCEL
# ============================================================================

@ict_bp.route('/api/ict/export')
@login_requerido
def export_ict_excel():
    """Exportar el historial ICT filtrado a un archivo de Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        fecha = request.args.get("fecha", "").strip()
        linea = request.args.get("linea", "").strip()
        resultado = request.args.get("resultado", "").strip()
        barcode_like = request.args.get("barcode_like", "").strip()

        sql = (
            "SELECT fecha, TIME(ts) AS hora, linea, ict, resultado, no_parte, barcode, "
            "fuente_archivo, defect_code, defect_valor "
            "FROM history_ict WHERE 1=1"
        )
        params = []

        if fecha:
            sql += " AND fecha=%s"
            params.append(fecha)
        if linea:
            sql += " AND linea=%s"
            params.append(linea)
        if resultado:
            sql += " AND resultado=%s"
            params.append(resultado)
        if barcode_like:
            sql += " AND barcode LIKE %s"
            params.append(f"%{barcode_like}%")

        sql += " ORDER BY ts DESC LIMIT 500"
        rows = execute_query(sql, tuple(params), fetch='all') or []

        wb = Workbook()
        ws = wb.active
        ws.title = "Historial ICT"

        header_fill = PatternFill(start_color="3f6b6e", end_color="3f6b6e", fill_type="solid")
        cell_fill = PatternFill(start_color="a1a09c", end_color="a1a09c", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        headers = ["Fecha", "Hora", "Línea", "ICT", "Resultado", "No Parte", "Barcode", "Fuente", "Defect Code", "Defect Valor"]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        for row_idx, row in enumerate(rows, start=2):
            formatted = _ict_format_row(row)
            values = [
                formatted.get('fecha', ''),
                formatted.get('hora', ''),
                formatted.get('linea', ''),
                formatted.get('ict', ''),
                formatted.get('resultado', ''),
                formatted.get('no_parte', ''),
                formatted.get('barcode', ''),
                formatted.get('fuente_archivo', ''),
                formatted.get('defect_code', ''),
                formatted.get('defect_valor', ''),
            ]

            for col_num, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_num, value=value)
                cell.fill = cell_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

        for col in range(1, len(headers) + 1):
            column_letter = ws.cell(row=1, column=col).column_letter
            ws.column_dimensions[column_letter].width = 16

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"historial_ict_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


# ============================================================================
# API: EXPORTAR DEFECTOS A EXCEL
# ============================================================================

@ict_bp.route('/api/ict/export-defects')
@login_requerido
def export_ict_defects_excel():
    """Exportar detalles de defectos ICT a un archivo de Excel."""
    barcode = request.args.get("barcode", "").strip()
    resultado_filter = request.args.get("resultado", "").strip()

    if not barcode:
        return jsonify({"error": "Barcode requerido"}), 400

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        sql = (
            "SELECT d.barcode, h.linea, h.ict, d.componente, d.pinref, d.act_value, d.act_unit, "
            "d.std_value, d.std_unit, d.meas_value, "
            "d.m_value, d.r_value, d.hlim_pct, d.llim_pct, "
            "d.hp_value, d.lp_value, d.ws_value, d.ds_value, d.rc_value, "
            "d.p_flag, d.j_flag, d.resultado_local, d.defecto_tipo, d.ts, "
            "DATE(d.ts) AS fecha, TIME(d.ts) AS hora "
            "FROM history_ict_defects d "
            "LEFT JOIN history_ict h ON d.barcode COLLATE utf8mb4_unicode_ci = h.barcode COLLATE utf8mb4_unicode_ci "
            "AND d.ts = h.ts "
            "WHERE d.barcode=%s "
        )
        params = [barcode]

        if resultado_filter:
            sql += " AND d.resultado_local=%s"
            params.append(resultado_filter)

        sql += " ORDER BY d.ts DESC, d.componente LIMIT 1000"

        rows = execute_query(sql, tuple(params), fetch='all') or []

        wb = Workbook()
        ws = wb.active
        ws.title = f"Parametros {barcode[:20]}"

        header_fill = PatternFill(start_color="3f6b6e", end_color="3f6b6e", fill_type="solid")
        cell_fill = PatternFill(start_color="a1a09c", end_color="a1a09c", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        headers = [
            "Fecha", "Hora", "Línea", "ICT", "Barcode", "Componente", "Pinref",
            "ACT", "Unit", "STD", "Unit", "MEAS", "M", "R", "HLIM", "LLIM",
            "H.P", "L.P", "WS", "DS", "RC", "P", "J", "Resultado", "Tipo Defecto"
        ]

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        for row_idx, row in enumerate(rows, start=2):
            formatted = _ict_format_row(row)
            hlim = formatted.get('hlim_pct', '')
            llim = formatted.get('llim_pct', '')

            row_values = [
                formatted.get('fecha', ''),
                formatted.get('hora', ''),
                formatted.get('linea', ''),
                formatted.get('ict', ''),
                formatted.get('barcode', ''),
                formatted.get('componente', ''),
                formatted.get('pinref', ''),
                formatted.get('act_value', ''),
                formatted.get('act_unit', ''),
                formatted.get('std_value', ''),
                formatted.get('std_unit', ''),
                formatted.get('meas_value', ''),
                formatted.get('m_value', ''),
                formatted.get('r_value', ''),
                f"{hlim}%" if hlim else '',
                f"{llim}%" if llim else '',
                formatted.get('hp_value', ''),
                formatted.get('lp_value', ''),
                formatted.get('ws_value', ''),
                formatted.get('ds_value', ''),
                formatted.get('rc_value', ''),
                formatted.get('p_flag', ''),
                formatted.get('j_flag', ''),
                formatted.get('resultado_local', ''),
                formatted.get('defecto_tipo', '')
            ]

            for col_num, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_num, value=value)
                cell.fill = cell_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

        for col in range(1, len(headers) + 1):
            column_letter = ws.cell(row=1, column=col).column_letter
            ws.column_dimensions[column_letter].width = 12

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"parametros_{barcode}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500
