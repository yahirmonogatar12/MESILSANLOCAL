import json
import os
import re
import socket
import subprocess
import csv
import hashlib
import io
import tempfile
import threading
import time
import traceback
import struct
import zlib
from concurrent.futures import ThreadPoolExecutor

# Intentar importar MySQLdb, si no está disponible usar PyMySQL como alternativa
try:
    import MySQLdb
except ImportError:
    import pymysql

    pymysql.install_as_MySQLdb()
    import MySQLdb

from datetime import date, datetime, timedelta
from datetime import time as dt_time
from decimal import Decimal
from functools import wraps
from pathlib import Path


def obtener_fecha_hora_mexico():
    """Obtener fecha y hora actual en zona horaria de México (GMT-6)"""
    try:
        # Calcular hora de México Central (GMT-6)
        utc_now = datetime.utcnow()
        mexico_time = utc_now - timedelta(hours=6)
        return mexico_time
    except Exception as e:
        # Fallback a hora local
        return datetime.now()


import pandas as pd
from flask import (
    Flask,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from werkzeug.utils import secure_filename

# Migrados a app/api/ y registrados via registrar_blueprints_api() en app_factory.py:
#   admin_api          -> app/api/admin/permisos.py
#   api_po_wo          -> app/api/control_produccion/po_wo.py
#                         (2026-05-27: absorbio /api/work-orders, /api/work-orders/import,
#                          /api/wo/exportar que aun vivian en este archivo)
#   api_raw_modelos    -> app/api/shared/raw_modelos.py

# Importar sistema de autenticación mejorado
from .auth_system import AuthSystem
from .db import (
    agregar_control_material_almacen,
    get_db_connection,
    init_db,
    migrar_datos_sqlite,
    obtener_control_material_almacen,
    test_database_connection,
)
from .db_mysql import (
    actualizar_inventario,
    actualizar_material_completo,
    cargar_configuracion,
    execute_query,
    guardar_configuracion,
    guardar_material,
    obtener_inventario,
    obtener_materiales,
)
from .config_mysql import get_pooled_connection

# Importar modelos y funciones PO → WO
from .po_wo_models import (
    crear_tablas_po_wo,
    generar_codigo_po,
    generar_codigo_wo,
    listar_pos_con_filtros,
    listar_pos_por_estado,
    listar_wos_con_filtros,
    listar_wos_por_po,
    obtener_po_por_codigo,
    obtener_wo_por_codigo,
    validar_codigo_po,
    validar_codigo_wo,
    verificar_po_existe,
    verificar_wo_existe,
)
# shipping_api migrado a app/api/pda/shipping.py
# Se registra via registrar_blueprints_api() en app_factory.py
from .api.pda.shipping import init_shipping_tables
# shipping_material_api migrado a app/api/pda/shipping_material.py
# Se registra via registrar_blueprints_api() en app_factory.py
from .api.pda.shipping_material import (
    SHIPPING_TABLES,
    adjust_shipping_movement_record,
    assign_exit_departure_value,
    delete_shipping_movement_record,
    ensure_inventory_record,
    generate_movement_folio,
    get_departure_history_records,
    get_dict_cursor,
    init_shipping_material_tables,
    normalize_integer,
    normalize_part_number,
    normalize_search,
    rebuild_part_inventory_state,
    to_sql_datetime,
)
from .services.ict_lgd_parser import (
    IctLgdError,
    IctLgdNotFoundError,
    IctLgdPathError,
    get_lgd_parameters_for_barcode,
    resolve_lgd_path,
)

# Migracion 2026-05-27: Helpers movidos a app/api/control_proceso/almacen_embarques.py
# Reexportados aqui porque "Control de salida de lineas" los consume:
#   - _normalizar_texto_embarques_historial (usado por _obtener_control_salida_lineas)
#   - _exportar_historial_embarques_excel  (usado por export_control_salida_lineas)
from .api.control_proceso.almacen_embarques import (
    _exportar_historial_embarques_excel,
    _normalizar_texto_embarques_historial,
)
# tickets_portal migrado a app/api/portal/tickets.py
# Se registra via registrar_blueprints_api() en app_factory.py
# user_admin migrado a app/api/admin/usuarios.py

app = Flask(__name__)
app.secret_key = os.getenv(
    "SECRET_KEY", "fallback_key_for_development_only"
)  # Necesario para usar sesiones


def _env_flag(name, default=False):
    val = os.getenv(name)
    if val is None:
        return default
    return str(val).strip().lower() in ("1", "true", "yes", "on", "si")


def should_run_startup_init():
    # Overrides explícitos
    if _env_flag("MES_FORCE_STARTUP_INIT", False):
        return True
    if _env_flag("MES_SKIP_STARTUP_INIT", False):
        return False

    # Si estamos en dev con reloader, solo correr en el proceso real
    if _env_flag("MES_USE_RELOADER", False):
        return os.environ.get("WERKZEUG_RUN_MAIN") == "true"

    # Sin reloader: correr init normalmente
    return True


STARTUP_INIT_ENABLED = should_run_startup_init()
_startup_t0 = time.time()


def _startup_log(msg):
    elapsed = round(time.time() - _startup_t0, 2)
    print(f"[startup {elapsed}s] {msg}")


# smd_inventory ahora se registra via app/api/__init__.py

# shipping y shipping_material (apps moviles PDA) registrados
# via registrar_blueprints_api() en app_factory.py

# api_po_wo y api_raw migrados a app/api/ y se registran centralmente
# desde app_factory.py via registrar_blueprints_api().

# Las inicializaciones (init_db, auth_system.init_database, shipping)
# se movieron a app/startup_init.py. Solo se instancia auth_system aqui
# porque otros decoradores en este archivo lo usan a nivel de modulo.
auth_system = AuthSystem()

# Registrar Blueprints de administración


@app.route("/api/health", methods=["GET"])
def api_health():
    database_status = "unavailable"

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT 1")
        cursor.fetchone()
        cursor.close()
        conn.close()
        database_status = "ok"
    except Exception:
        database_status = "error"

    return jsonify(
        {
            "status": "OK",
            "service": "MESILSANLOCAL",
            "database": database_status,
            "timestamp": obtener_fecha_hora_mexico().strftime("%Y-%m-%d %H:%M:%S"),
        }
    )

# smt_routes_date_fixed migrado a app/api/control_calidad/smt_historial_simple.py
# smt_routes_clean      migrado a app/api/control_calidad/smt_historial.py
# Ambos se registran via registrar_blueprints_api() en app_factory.py

# user_admin (con url_prefix="/admin") y admin_bp ahora se registran
# desde app/api/__init__.py via registrar_blueprints_api()
# tickets_portal: registro movido a app/api/__init__.py


def requiere_permiso_dropdown(pagina, seccion, boton):
    """Decorador para verificar permisos específicos de dropdowns"""

    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if "usuario" not in session:
                return jsonify(
                    {"error": "Usuario no autenticado", "redirect": "/login"}
                ), 401

            try:
                username = session["usuario"]
                print(
                    f" Verificando permisos para usuario: {username}, página: {pagina}, sección: {seccion}, botón: {boton}"
                )

                rol_nombre = auth_system.obtener_rol_principal_usuario(username)
                if not rol_nombre:
                    print(" Usuario sin roles asignados")
                    return jsonify({"error": "Usuario sin roles asignados"}), 403

                print(f" Rol del usuario: {rol_nombre}")
                if rol_nombre == "superadmin":
                    print(" Superadmin: permiso concedido")
                    return f(*args, **kwargs)

                tiene_permiso = auth_system.verificar_permiso_boton(
                    username, pagina, seccion, boton
                )
                print(f" Tiene permiso: {tiene_permiso}")

                if not tiene_permiso:
                    print(f" Sin permisos para: {pagina} > {seccion} > {boton}")
                    # Respuesta diferente para AJAX vs navegación directa
                    if (
                        request.headers.get("Content-Type") == "application/json"
                        or request.is_json
                    ):
                        return jsonify(
                            {
                                "error": f"No tienes permisos para acceder a: {boton}",
                                "permiso_requerido": f"{pagina} > {seccion} > {boton}",
                            }
                        ), 403
                    else:
                        # Para carga AJAX de HTML, devolver mensaje de error
                        return (
                            f"""
                        <div style="
                            display: flex;
                            flex-direction: column;
                            align-items: center;
                            justify-content: center;
                            height: 400px;
                            background: #2c2c2c;
                            color: #e0e0e0;
                            border-radius: 10px;
                            margin: 20px;
                            text-align: center;
                        ">
                            <i class="fas fa-lock" style="font-size: 3rem; color: #dc3545; margin-bottom: 20px;"></i>
                            <h3>Acceso Denegado</h3>
                            <p>No tienes permisos para acceder a: <strong>{boton}</strong></p>
                            <p style="font-size: 0.9rem; opacity: 0.7;">Permiso requerido: {pagina} > {seccion} > {boton}</p>
                        </div>
                        """,
                            403,
                        )

                print(f" Permisos verificados correctamente, ejecutando función...")
                return f(*args, **kwargs)

            except Exception as e:
                print(f" Error verificando permisos: {e}")
                import traceback

                traceback.print_exc()
                return jsonify({"error": "Error interno del servidor"}), 500

        return decorated_function

    return decorator


# Filtros de Jinja2 para permisos de botones
@app.template_filter("tiene_permiso_boton")
def tiene_permiso_boton(nombre_boton):
    """Filtro para verificar si el usuario actual tiene permiso para un botón específico"""
    try:
        # Obtener el usuario de la sesión actual
        if "usuario" not in session:
            return False

        username = session["usuario"]
        if auth_system.obtener_rol_principal_usuario(username) == "superadmin":
            return True

        permisos_botones = auth_system.obtener_permisos_botones_usuario(username)
        for secciones in permisos_botones.values():
            for botones in secciones.values():
                for item in botones:
                    if isinstance(item, dict) and item.get("boton") == nombre_boton:
                        return True
                    if item == nombre_boton:
                        return True

        return False

    except Exception as e:
        print(f"Error verificando permiso de botón '{nombre_boton}': {e}")
        return False


@app.template_filter("permisos_botones_pagina")
def permisos_botones_pagina(usuario, pagina):
    """Filtro para obtener todos los permisos de botones de una página"""
    if not usuario:
        return {}
    return auth_system.obtener_permisos_botones_usuario(usuario, pagina)


# DEPRECADO: Función antigua para compatibilidad temporal
def cargar_usuarios():
    """Función deprecada - se mantiene para compatibilidad"""
    ruta = os.path.join(os.path.dirname(__file__), "database", "usuarios.json")
    ruta = os.path.abspath(ruta)
    try:
        with open(ruta, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(" usuarios.json no encontrado, usando solo sistema de BD")
        return {}


# ACTUALIZADO: Usar el sistema de autenticación avanzado
def login_requerido(f):
    @wraps(f)
    def decorada(*args, **kwargs):
        print(" Verificando sesión avanzada:", session.get("usuario"))

        # Verificar si hay usuario en sesión
        if "usuario" not in session:
            print("No hay usuario en sesión")
            return redirect(url_for("inicio"))

        usuario = session.get("usuario")

        # Evita un round-trip a MySQL remoto en cada request AJAX.
        now_ts = int(time.time())
        last_touch_ts = int(session.get("_last_activity_touch_ts", 0) or 0)
        if now_ts - last_touch_ts >= 300:
            auth_system._actualizar_actividad_sesion(usuario)
            session["_last_activity_touch_ts"] = now_ts
            session.modified = True

        return f(*args, **kwargs)

    return decorada


PUBLIC_ROUTE_ENDPOINTS = {
    "index",
    "inicio",
    "login",
    "api_health",
    "favicon",
    "front_plan_static",
    "static",
}


def _request_expects_json():
    accept = request.headers.get("Accept", "")
    requested_with = request.headers.get("X-Requested-With", "")
    return (
        request.path.startswith("/api/")
        or request.is_json
        or "application/json" in accept
        or requested_with == "XMLHttpRequest"
    )


def _is_public_api_route():
    """APIs que no dependen de la sesión web del portal."""
    path = request.path.rstrip("/")
    exact_paths = {
        "/api/shipping/auth/login",
        "/api/shipping/auth/logout",
        "/api/shipping/permissions/available",
        "/api/shipping/departments",
        "/api/shipping/cargos",
        "/api/shipping/entries",
        "/api/shipping/stats/today",
        "/api/shipping/stats/summary",
        "/api/shipping/material/entries",
        "/api/shipping/material/entries/boxes",
        "/api/shipping/material/exits",
        "/api/shipping/material/exits/boxes",
        "/api/shipping/material/returns",
        "/api/shipping/material/inventory",
        "/api/shipping/material/stats/today",
    }
    public_prefixes = (
        "/api/shipping/auth/verify/",
        "/api/shipping/users/",
        "/api/shipping/quality/",
        "/api/shipping/entries/",
        "/api/shipping/material/boxes/",
    )
    return path in exact_paths or any(path.startswith(prefix) for prefix in public_prefixes)


@app.before_request
def require_login_by_default():
    """Protect routes by default and keep a short explicit public allowlist."""
    endpoint = request.endpoint

    # Let Flask resolve 404/405 normally when there is no matched endpoint.
    if endpoint is None:
        return None

    # Static assets and a small set of public endpoints remain accessible.
    if endpoint in PUBLIC_ROUTE_ENDPOINTS or endpoint.endswith(".static"):
        return None

    # Las apps PDA autentican contra /api/shipping/auth/login y luego consumen
    # endpoints del mismo blueprint sin sesión web del portal corporativo.
    if _is_public_api_route():
        return None

    if "usuario" in session:
        return None

    if _request_expects_json():
        return jsonify({"error": "Usuario no autenticado", "redirect": "/login"}), 401

    return redirect(url_for("inicio"))


def render_landing_page(login_error=None, login_username=None):
    """Renderiza la landing page con o sin sesión activa."""
    authenticated = "usuario" in session
    nombre_completo = None
    permisos = {}
    roles = []

    if authenticated:
        usuario = session.get("usuario")
        nombre_completo = session.get("nombre_completo", usuario)
        permisos = session.get("permisos", {})
        roles = session.get("roles") or auth_system.obtener_roles_usuario(usuario)

    upcoming_apps = [
        {
            "name": "Más Herramientas",
            "description": "Expansión futura",
            "long_description": "Nuevas aplicaciones serán agregadas pronto.",
            "icon": "rocket",
        }
    ]

    return render_template(
        "landing.html",
        nombre_usuario=nombre_completo,
        permisos=permisos,
        roles=roles,
        upcoming_apps=upcoming_apps,
        usuario_autenticado=authenticated,
        login_error=login_error,
        login_username=login_username,
    )


@app.route("/")
def index():
    return redirect(url_for("inicio"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return redirect(url_for("inicio"))

    is_ajax = request.headers.get(
        "X-Requested-With"
    ) == "XMLHttpRequest" or "application/json" in request.headers.get("Accept", "")

    user = request.form.get("username", "").strip()
    pw = request.form.get("password", "")
    print(f" Intento de login: {user}")

    # PRIORIDAD 1: Intentar con el nuevo sistema de BD
    resultado_auth = auth_system.verificar_usuario(user, pw)

    # verificar_usuario devuelve (success, message) en lugar de diccionario
    if isinstance(resultado_auth, tuple):
        auth_success, auth_message = resultado_auth
    else:
        auth_success = (
            resultado_auth.get("success", False)
            if isinstance(resultado_auth, dict)
            else False
        )
        auth_message = (
            resultado_auth.get("message", "Error desconocido")
            if isinstance(resultado_auth, dict)
            else str(resultado_auth)
        )

    if auth_success:
        print(f" Login exitoso con sistema BD: {user}")
        session["usuario"] = user

        # Obtener información completa del usuario
        info_usuario = auth_system.obtener_informacion_usuario(user)
        if info_usuario:
            session["nombre_completo"] = info_usuario["nombre_completo"]
            session["email"] = info_usuario["email"]
            session["departamento"] = info_usuario["departamento"]
            print(f" Información completa cargada para {user}:")
            print(f"  - Nombre completo: {info_usuario['nombre_completo']}")
            print(f"  - Email: {info_usuario['email']}")
            print(f"  - Departamento: {info_usuario['departamento']}")
        else:
            # Fallback si no se puede obtener la información
            session["nombre_completo"] = user  # Usar username como fallback
            print(
                f"⚠️ No se pudo cargar información completa para {user}, usando username como fallback"
            )

        # Registrar auditoría
        auth_system.registrar_auditoria(
            usuario=user,
            modulo="sistema",
            accion="login",
            descripcion="Inicio de sesión exitoso",
            resultado="EXITOSO",
        )

        # Obtener permisos del usuario
        permisos_resultado = auth_system.obtener_permisos_usuario(user)

        # Verificar si devuelve tupla (permisos, rol_id) o solo permisos
        if isinstance(permisos_resultado, tuple):
            permisos, rol_id = permisos_resultado
        else:
            permisos = permisos_resultado
            rol_id = None

        session["permisos"] = permisos
        session["roles"] = auth_system.obtener_roles_usuario(user)
        session["rol_principal"] = session["roles"][0] if session["roles"] else None
        session.modified = True
        print(f" Permisos establecidos en sesión para {user}: {permisos}")

        # Redirigir siempre al hub de aplicaciones (landing page)
        print(f" Login exitoso para {user}, redirigiendo al hub de aplicaciones")
        redirect_url = url_for("inicio")
        if is_ajax:
            return jsonify({"success": True, "redirect": redirect_url})
        return redirect(redirect_url)

    # FALLBACK: Intentar con el sistema antiguo (usuarios.json)
    try:
        usuarios_json = cargar_usuarios()
        if user in usuarios_json and usuarios_json[user] == pw:
            print(f" Login exitoso con sistema JSON (fallback): {user}")
            session["usuario"] = user

            # Para usuarios del sistema JSON, usar el username como nombre completo
            session["nombre_completo"] = (
                user  # Fallback para usuarios del sistema antiguo
            )
            session["email"] = ""  # Sin email para usuarios del sistema antiguo
            session["departamento"] = (
                ""  # Sin departamento para usuarios del sistema antiguo
            )
            print(f"⚠️ Usuario del sistema JSON (fallback): {user}")

            # Registrar auditoría del fallback
            auth_system.registrar_auditoria(
                usuario=user,
                modulo="sistema",
                accion="login_json",
                descripcion="Inicio de sesión con sistema JSON (fallback)",
                resultado="EXITOSO",
            )

            # Redirigir según el usuario (lógica original)
            redirect_url = url_for("inicio")
            if user.startswith("Materiales") or user == "1111":
                redirect_url = url_for("material")
            elif user.startswith("Produccion") or user == "2222":
                redirect_url = url_for("produccion")
            elif user.startswith("DDESARROLLO") or user == "3333":
                redirect_url = url_for("desarrollo")

            if is_ajax:
                return jsonify({"success": True, "redirect": redirect_url})
            return redirect(redirect_url)
    except Exception as e:
        print(f" Error en fallback JSON: {e}")

    # Si llega aquí, login falló
    print(f" Login falló: {user} ({auth_message})")
    auth_system.registrar_auditoria(
        usuario=user,
        modulo="sistema",
        accion="login_failed",
        descripcion="Intento de login fallido - credenciales incorrectas",
        resultado="ERROR",
    )

    error_message = "Usuario o contraseña incorrectos. Por favor, intente de nuevo"

    if is_ajax:
        return jsonify({"success": False, "message": error_message}), 401

    return render_landing_page(login_error=error_message, login_username=user)


@app.route("/inicio")
def inicio():
    """Landing page / Hub de aplicaciones"""
    return render_landing_page()


@app.route("/api/mi-perfil", methods=["GET", "POST"])
def api_mi_perfil():
    """Consultar o actualizar el perfil del usuario autenticado."""
    usuario = session.get("usuario")
    if not usuario:
        return jsonify({"success": False, "message": "Sesion expirada"}), 401

    roles = session.get("roles") or auth_system.obtener_roles_usuario(usuario) or []
    rol_principal = session.get("rol_principal") or (roles[0] if roles else "")
    if roles and session.get("roles") != roles:
        session["roles"] = roles
        session["rol_principal"] = rol_principal
        session.modified = True

    info_usuario = auth_system.obtener_informacion_usuario(usuario)

    if request.method == "GET":
        nombre_completo = session.get("nombre_completo") or usuario
        email = session.get("email") or ""
        editable = bool(info_usuario)

        if info_usuario:
            nombre_completo = info_usuario.get("nombre_completo") or nombre_completo
            email = info_usuario.get("email") or email
            session["nombre_completo"] = nombre_completo
            session["email"] = email
            session.modified = True

        return jsonify(
            {
                "success": True,
                "perfil": {
                    "username": usuario,
                    "nombre_completo": nombre_completo,
                    "email": email,
                    "rol": rol_principal,
                    "roles": roles,
                    "editable": editable,
                },
            }
        )

    payload = request.get_json(silent=True) or request.form
    nombre_completo = (payload.get("nombre_completo") or "").strip()
    email = (payload.get("email") or "").strip()
    current_password = payload.get("current_password") or ""
    new_password = payload.get("new_password") or ""
    confirm_password = payload.get("confirm_password") or ""
    change_password = bool(current_password or new_password or confirm_password)

    if not nombre_completo:
        return (
            jsonify(
                {
                    "success": False,
                    "message": "El nombre completo es obligatorio",
                }
            ),
            400,
        )

    if len(nombre_completo) > 120:
        return (
            jsonify(
                {
                    "success": False,
                    "message": "El nombre completo es demasiado largo",
                }
            ),
            400,
        )

    if email and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
        return (
            jsonify(
                {
                    "success": False,
                    "message": "El correo electronico no es valido",
                }
            ),
            400,
        )

    if change_password:
        if not current_password:
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "Debes escribir tu contrasena actual",
                    }
                ),
                400,
            )

        if not new_password:
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "Debes escribir una nueva contrasena",
                    }
                ),
                400,
            )

        if len(new_password) < 6:
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "La nueva contrasena debe tener al menos 6 caracteres",
                    }
                ),
                400,
            )

        if new_password != confirm_password:
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "La confirmacion de la nueva contrasena no coincide",
                    }
                ),
                400,
            )

        if new_password == current_password:
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "La nueva contrasena debe ser diferente a la actual",
                    }
                ),
                400,
            )

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        if conn is None:
            raise RuntimeError("No se pudo obtener conexion a la base de datos")

        cursor = conn.cursor()
        if change_password:
            cursor.execute(
                """
                SELECT password_hash
                FROM usuarios_sistema
                WHERE username = %s
                """,
                (usuario,),
            )
            usuario_row = cursor.fetchone()
            password_hash_actual = ""

            if isinstance(usuario_row, dict):
                password_hash_actual = usuario_row.get("password_hash") or ""
            elif usuario_row:
                password_hash_actual = usuario_row[0]

            if not password_hash_actual:
                conn.rollback()
                return (
                    jsonify(
                        {
                            "success": False,
                            "message": "No fue posible validar tu contrasena actual",
                        }
                    ),
                    404,
                )

            if password_hash_actual != auth_system.hash_password(current_password):
                conn.rollback()
                return (
                    jsonify(
                        {
                            "success": False,
                            "message": "La contrasena actual es incorrecta",
                        }
                    ),
                    400,
                )

            cursor.execute(
                """
                UPDATE usuarios_sistema
                SET nombre_completo = %s,
                    email = %s,
                    password_hash = %s,
                    modificado_por = %s,
                    fecha_modificacion = %s,
                    intentos_fallidos = 0,
                    bloqueado_hasta = NULL
                WHERE username = %s
                """,
                (
                    nombre_completo,
                    email,
                    auth_system.hash_password(new_password),
                    usuario,
                    auth_system.get_mexico_time_mysql(),
                    usuario,
                ),
            )
        else:
            cursor.execute(
                """
                UPDATE usuarios_sistema
                SET nombre_completo = %s,
                    email = %s,
                    modificado_por = %s,
                    fecha_modificacion = %s
                WHERE username = %s
                """,
                (
                    nombre_completo,
                    email,
                    usuario,
                    auth_system.get_mexico_time_mysql(),
                    usuario,
                ),
            )

        if cursor.rowcount == 0:
            conn.rollback()
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "No fue posible actualizar este usuario",
                    }
                ),
                404,
            )

        conn.commit()
    except Exception as exc:
        if conn:
            conn.rollback()
        print(f"Error actualizando perfil de {usuario}: {exc}")
        return (
            jsonify(
                {
                    "success": False,
                    "message": "No fue posible guardar los cambios",
                }
            ),
            500,
        )
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    session["nombre_completo"] = nombre_completo
    session["email"] = email
    session.modified = True

    auth_system.registrar_auditoria(
        usuario=usuario,
        modulo="sistema",
        accion="actualizar_perfil",
        descripcion=(
            "Actualizacion de perfil y contrasena desde la landing page"
            if change_password
            else "Actualizacion de perfil desde la landing page"
        ),
        resultado="EXITOSO",
    )

    return jsonify(
        {
            "success": True,
            "message": (
                "Perfil y contrasena actualizados correctamente"
                if change_password
                else "Perfil actualizado correctamente"
            ),
            "perfil": {
                "username": usuario,
                "nombre_completo": nombre_completo,
                "email": email,
                "rol": rol_principal,
                "roles": roles,
                "editable": True,
            },
        }
    )


@app.route("/calendario")
@login_requerido
def calendario():
    """Página del calendario de producción"""
    return render_template("calendario.html")


@app.route("/defect-management")
@login_requerido
def defect_management():
    """Módulo de Gestión de Defectos (En Desarrollo)"""
    # TODO: Implementar módulo completo de gestión de defectos
    return render_template(
        "info.html",
        titulo="Gestión de Defectos",
        mensaje="Módulo en desarrollo. Próximamente disponible.",
        tipo="warning",
    )


@app.route("/favicon.eco")
def favicon():
    """Servir favicon usando un icono existente"""
    return send_from_directory(
        os.path.join(app.root_path, "static", "icons"),
        "produccion.png",
        mimetype="image/png",
    )


@app.route("/sistemas")
@login_requerido
def sistemas():
    """Redirige al hub de inicio"""
    return redirect(url_for("inicio"))


@app.route("/soporte")
@login_requerido
def soporte():
    """Página de soporte técnico"""
    return (
        render_template("soporte.html")
        if os.path.exists("app/templates/soporte.html")
        else f"<h1>Soporte Técnico</h1><p>En construcción. <a href='/inicio'>Volver al inicio</a></p>"
    )


@app.route("/documentacion")
@login_requerido
def documentacion():
    """Página de documentación"""
    return (
        render_template("documentacion.html")
        if os.path.exists("app/templates/documentacion.html")
        else f"<h1>Documentación</h1><p>En construcción. <a href='/inicio'>Volver al inicio</a></p>"
    )


@app.route("/ILSAN-ELECTRONICS")
@login_requerido
def material():
    usuario = session.get("usuario", "Invitado")
    nombre_completo = session.get("nombre_completo", None)

    # Si no tenemos el nombre completo en la sesión, obtenerlo de la BD
    if not nombre_completo and usuario != "Invitado":
        print(
            f"⚠️ Nombre completo no encontrado en sesión para {usuario}, obteniendo de BD..."
        )
        from .auth_system import auth_system

        info_usuario = auth_system.obtener_informacion_usuario(usuario)
        if info_usuario and info_usuario["nombre_completo"]:
            nombre_completo = info_usuario["nombre_completo"]
            session["nombre_completo"] = (
                nombre_completo  # Guardar en sesión para futuras consultas
            )
            print(f" Nombre completo obtenido de BD: {nombre_completo}")
        else:
            nombre_completo = usuario  # Fallback al username
            session["nombre_completo"] = usuario
            print(
                f"⚠️ No se pudo obtener nombre completo de BD, usando username: {usuario}"
            )

    # Si todavía no tenemos nombre completo, usar el username
    if not nombre_completo:
        nombre_completo = usuario

    permisos = session.get("permisos", {})

    # Verificar si tiene permisos de administración de usuarios
    tiene_permisos_usuarios = False
    if isinstance(permisos, dict) and "sistema" in permisos:
        tiene_permisos_usuarios = "usuarios" in permisos["sistema"]

    return render_template(
        "MainTemplate.html",
        usuario=nombre_completo,  # Pasar nombre completo en lugar de username
        tiene_permisos_usuarios=tiene_permisos_usuarios,
    )


@app.route("/dashboard")
@login_requerido
def dashboard():
    """Alias para la página principal (MainTemplate)"""
    usuario = session.get("usuario")
    nombre_completo = session.get("nombre_completo")

    # Si no tenemos nombre completo, intentar obtenerlo
    if not nombre_completo and usuario:
        try:
            query = "SELECT nombre_completo FROM users WHERE usuario = %s"
            result = execute_query(query, (usuario,), fetch="one")
            if result and result.get("nombre_completo"):
                nombre_completo = result["nombre_completo"]
                session["nombre_completo"] = nombre_completo
        except Exception as e:
            print(f"⚠️ Error obteniendo nombre completo para dashboard: {e}")
            nombre_completo = usuario

    if not nombre_completo:
        nombre_completo = usuario

    permisos = session.get("permisos", {})
    tiene_permisos_usuarios = False
    if isinstance(permisos, dict) and "sistema" in permisos:
        tiene_permisos_usuarios = "usuarios" in permisos["sistema"]

    return render_template(
        "MainTemplate.html",
        usuario=nombre_completo,
        tiene_permisos_usuarios=tiene_permisos_usuarios,
    )


@app.route("/logout")
def logout():
    usuario = session.get("usuario", "unknown")

    # Registrar auditoría del logout
    if usuario != "unknown":
        auth_system.registrar_auditoria(
            usuario=usuario,
            modulo="sistema",
            accion="logout",
            descripcion="Cierre de sesión",
            resultado="EXITOSO",
        )
        print(f"🚪 Logout exitoso: {usuario}")

    # Limpiar sesión completa
    session.clear()

    return redirect(url_for("inicio"))


# =============================
# FRONT PLAN: Vistas y estáticos
# =============================


# Alias para servir los assets originales de FRONT PLAN ubicados en
# app/FRONT PLAN/static sin depender de moverlos físicamente.
@app.route("/front-plan/static/<path:filename>")
def front_plan_static(filename):
    try:
        base_dir = os.path.join(os.path.dirname(__file__), "FRONT PLAN", "static")
        return send_from_directory(base_dir, filename)
    except Exception as e:
        return jsonify({"error": f"Recurso no encontrado: {str(e)}"}), 404


# Migracion 2026-05-26: view_plan_main movido a app/api/control_produccion/plan_assy.py


# Limpieza 2026-05-27: view_control_main eliminado (modulo Control de operacion de linea Main borrado)


# Rutas AJAX para cargar módulos en el área de Control de Proceso (prompts)
# Migracion 2026-05-26: plan_main_assy_ajax movido a app/api/control_produccion/plan_assy.py


# Migracion 2026-05-26: plan_main_imd_ajax movido a app/api/control_produccion/plan_imd.py


@app.route("/plan-main-smt-ajax")
@login_requerido
def plan_main_smt_ajax():
    try:
        return render_template("Control de proceso/Control_produccion_smt_plan.html")
    except Exception as e:
        return f"Error al cargar el contenido: {str(e)}", 500


# Limpieza 2026-05-27: ctrl_operacion_linea_main_ajax eliminado (modulo Control de operacion de linea Main borrado)



# =============================
# CONTROL DE CUCHILLAS DE CORTE (ASSY)
# =============================
# Migrado a app/api/control_produccion/cuchillas_corte.py (2026-05-25).
# Las 17 rutas se registran via blueprint en app/api/__init__.py.
# Reexportamos los helpers y funciones de bootstrap aqui para preservar
# `_routes.crear_tablas_cuchillas_corte()` (startup_init) y
# `from app.routes import _cuchillas_rows_to_json` (app.api.shared).
from app.api.control_produccion.cuchillas_corte import (  # noqa: E402, F401
    CUCHILLAS_PERMISO_PAGINA,
    CUCHILLAS_PERMISO_SECCION,
    CUCHILLAS_PERMISO_BOTON,
    CUCHILLAS_SOURCE_DEFAULT,
    CUCHILLAS_SOURCE_ALLOWED,
    CUCHILLAS_HOURLY_SYNC_SECONDS,
    _cuchillas_row_to_json,
    _cuchillas_rows_to_json,
    _cuchillas_execute_raw,
    crear_tablas_cuchillas_corte,
    crear_trigger_cuchillas_corte_plan_main,
    iniciar_cuchillas_hourly_sync_worker,
)
# Bootstrap de cuchillas movido a app/startup_init.py


# Migracion 2026-05-27: snapshot inventario (constantes, DDL, captura,
# worker daemon, endpoints) movido a app/api/shared/snapshot_inventario.py.
# startup_init.py importa crear_tablas_snapshot_inventario() e
# iniciar_snapshot_inv_worker() desde el blueprint.

# Bootstrap de snapshot inventario movido a app/startup_init.py



# =============================
# FRONT PLAN: API mínima plan_main
# =============================


# Migracion 2026-05-26: _fp_safe_date y _fp_generate_lot_no consolidados en
# app/api/shared/plan_lot_no.py. Antes vivian duplicados aqui y en plan_smt.py.
from app.api.shared.plan_lot_no import (  # noqa: E402, F401
    _fp_generate_lot_no,
    _fp_safe_date,
)

# Migracion 2026-05-26: _ensure_plan_bom_assignment_columns, _plan_has_ks_snapshot
# y _validate_plan_bom_assignment consolidados en app/api/shared/bom_revisions.py
# (junto a los helpers existentes _ks_current_bom_revision, etc.).
# Antes vivian duplicados aqui y en plan_assy.py.
from app.api.shared.bom_revisions import (  # noqa: E402, F401
    _ks_current_bom_revision,
    _eco_for_part_revision,
    _plan_bom_revision_catalog,
    _ensure_plan_bom_assignment_columns,
    _plan_has_ks_snapshot,
    _validate_plan_bom_assignment,
)


# Migracion 2026-05-26: api_plan_list movido a app/api/control_produccion/plan_assy.py


# Migracion 2026-05-26: helpers _PLAN_BOM_ASSIGNMENT_COLUMNS_READY, etc. movidos a shared/bom_revisions.py


# Migracion 2026-05-26: api_plan_bom_revisions movido a app/api/control_produccion/plan_assy.py


# Migracion 2026-05-26: api_plan_input_main_scan_lots movido a app/api/control_produccion/plan_assy.py


# Migracion 2026-05-26: api_plan_input_main_assign_lot movido a app/api/control_produccion/plan_assy.py


# Migracion 2026-05-26: api_plan_input_main_create_plan movido a app/api/control_produccion/plan_assy.py


# Migracion 2026-05-26: api_plan_create movido a app/api/control_produccion/plan_assy.py


# Migracion 2026-05-26: api_plan_update movido a app/api/control_produccion/plan_assy.py


@app.route("/api/raw/search", methods=["GET"])
@login_requerido
def api_raw_search():
    """Buscar datos en la tabla RAW por part_no o model"""
    try:
        part_no = request.args.get("part_no", "").strip()
        if not part_no:
            return jsonify({"error": "part_no requerido"}), 400

        # Buscar con múltiples campos para mayor flexibilidad
        # Usar TRIM para ignorar espacios y comparación case-insensitive
        sql = """
            SELECT part_no, model, project, c_t as ct, uph
            FROM raw
            WHERE TRIM(model) = %s
               OR TRIM(part_no) = %s
               OR TRIM(part_no) LIKE %s
               OR UPPER(TRIM(part_no)) = UPPER(%s)
            LIMIT 1
        """
        params = (part_no, part_no, f"%{part_no}%", part_no)

        # CRÍTECO: Usar fetch='all' para obtener los datos, no el rowcount
        result = execute_query(sql, params, fetch="all")

        # Verificar que result sea una lista/tupla antes de usar len()
        if result and isinstance(result, (list, tuple)) and len(result) > 0:
            row = result[0]

            # execute_query con fetch='all' retorna lista de diccionarios
            # Acceder como diccionario, no como tupla
            data = {
                "part_no": row.get("part_no", "")
                if row.get("part_no") is not None
                else "",
                "model": row.get("model", "") if row.get("model") is not None else "",
                "model_code": row.get("model", "")
                if row.get("model") is not None
                else "",  # Alias
                "project": row.get("project", "")
                if row.get("project") is not None
                else "",
                "ct": str(row.get("ct", "0")) if row.get("ct") is not None else "0",
                "uph": str(row.get("uph", "0")) if row.get("uph") is not None else "0",
            }
            return jsonify([data])
        else:
            return jsonify([])

    except Exception as e:
        print(f"Error en api_raw_search: {e}")
        return jsonify({"error": str(e)}), 500


# Migracion 2026-05-26: api_plan_status movido a app/api/control_produccion/plan_assy.py


# Migracion 2026-05-26: api_plan_save_sequences movido a app/api/control_produccion/plan_assy.py


# Migracion 2026-05-26: api_plan_pending movido a app/api/control_produccion/plan_assy.py


# Migracion 2026-05-26: api_plan_reschedule movido a app/api/control_produccion/plan_assy.py


# Migracion 2026-05-26: api_plan_export_excel movido a app/api/control_produccion/plan_assy.py


# ====== RUTAS API PARA PLAN IMD ======


# Migracion 2026-05-26: api_plan_imd_list movido a app/api/control_produccion/plan_imd.py


# Migracion 2026-05-26: api_plan_imd_create movido a app/api/control_produccion/plan_imd.py


# Migracion 2026-05-26: api_plan_imd_batch_update movido a app/api/control_produccion/plan_imd.py


# Migracion 2026-05-26: api_plan_imd_update movido a app/api/control_produccion/plan_imd.py


# Migracion 2026-05-26: api_plan_imd_save_sequences movido a app/api/control_produccion/plan_imd.py


# Migracion 2026-05-26: api_plan_imd_pending movido a app/api/control_produccion/plan_imd.py


# Migracion 2026-05-26: api_plan_imd_pending_reschedule movido a app/api/control_produccion/plan_imd.py


# Migracion 2026-05-26: api_plan_imd_reschedule movido a app/api/control_produccion/plan_imd.py


# Migracion 2026-05-26: api_plan_imd_export_excel movido a app/api/control_produccion/plan_imd.py


# Migracion 2026-05-26: api_plan_imd_import_excel movido a app/api/control_produccion/plan_imd.py


# ====== RUTAS API PARA PLAN SMT (tabla plan_smt) ======
# Migrado a app/api/control_produccion/plan_smt.py (2026-05-25).
# Las 9 rutas se registran via blueprint en app/api/__init__.py.
# Reexportamos `crear_tabla_plan_smt_v2` para preservar `startup_init.py`.
from app.api.control_produccion.plan_smt import (  # noqa: E402, F401
    crear_tabla_plan_smt_v2,
)


# Migracion 2026-05-26: api_plan_main_list movido a app/api/control_produccion/plan_assy.py


# Migracion 2026-05-27: /api/work-orders/import movido a
# app/api/control_produccion/po_wo.py como POST /api/work_orders/import.
# La URL legacy /api/work-orders/import responde con 308 redirect (alias).


@app.route("/cargar_template", methods=["POST"])
@login_requerido
def cargar_template():
    template_path = None  # Initialize template_path
    try:
        data = request.get_json()
        template_path = data.get("template_path")

        if not template_path:
            return jsonify({"error": "No se especificó la ruta del template"}), 400

        # Validar que la ruta del template sea segura
        if ".." in template_path or template_path.startswith("/"):
            return jsonify({"error": "Ruta de template no válida"}), 400

        # Renderizar el template y devolver el HTML
        html_content = render_template(template_path)
        return html_content

    except Exception as e:
        template_name = template_path if template_path else "unknown"
        print(f"Error al cargar template {template_name}: {str(e)}")
        return jsonify({"error": f"Error al cargar el template: {str(e)}"}), 500


# Limpieza 2026-05-26: rutas /guardar_entrada_aereo y /listar_entradas_aereo eliminadas.
# Sin consumidores en app/static/js/ ni en app/templates/. Tabla entrada_aereo (SQLite legacy)
# tampoco se lee desde ningun otro lado. Documentadas como huerfanas en
# Documentacion/VERIFICACION_DETALLADA_HUERFANAS.md lineas 39-40.


# Función de conexión movida a db.py - usar get_db_connection() importada


@app.route("/guardar_cliente_seleccionado", methods=["POST"])
@login_requerido
def guardar_cliente_seleccionado():
    """Guardar la selección de cliente del usuario"""
    try:
        data = request.get_json()
        if not data or "cliente" not in data:
            return jsonify({"success": False, "error": "Cliente no proporcionado"}), 400

        cliente = data["cliente"]
        usuario = session.get("usuario", "default")

        # Guardar la configuración
        if guardar_configuracion_usuario(usuario, "cliente_seleccionado", cliente):
            return jsonify(
                {"success": True, "message": "Cliente guardado exitosamente"}
            )
        else:
            return jsonify({"success": False, "error": "Error al guardar cliente"}), 500

    except Exception as e:
        print(f"Error en guardar_cliente_seleccionado: {str(e)}")
        return jsonify({"success": False, "error": f"Error interno: {str(e)}"}), 500


@app.route("/cargar_cliente_seleccionado", methods=["GET"])
@login_requerido
def cargar_cliente_seleccionado():
    """Cargar la última selección de cliente del usuario"""
    try:
        usuario = session.get("usuario", "default")
        config = cargar_configuracion_usuario(usuario)
        cliente = config.get("cliente_seleccionado", "") if config else ""

        return jsonify({"success": True, "cliente": cliente})

    except Exception as e:
        print(f"Error en cargar_cliente_seleccionado: {str(e)}")
        return jsonify({"success": False, "error": f"Error interno: {str(e)}"}), 500


@app.route("/informacion_basica/control_de_material")
@login_requerido
def control_de_material_ajax():
    """Ruta para cargar dinámicamente el contenido de Control de Material"""
    try:
        return render_template("INFORMACION BASICA/CONTROL_DE_MATERIAL.html")
    except Exception as e:
        print(f"Error al cargar template Control de Material: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


# Rutas para cargar contenido dinámicamente (AJAX)
@app.route("/listas/informacion_basica")
@login_requerido
def lista_informacion_basica():
    """Cargar dinámicamente la lista de Información Básica"""
    try:
        return render_template("LISTAS/LISTA_INFORMACIONBASICA.html")
    except Exception as e:
        print(f"Error al cargar LISTA_INFORMACIONBASICA: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/listas/control_material")
@login_requerido
def lista_control_material():
    """Cargar dinámicamente la lista de Control de Material"""
    try:
        return render_template("LISTAS/LISTA_DE_MATERIALES.html")
    except Exception as e:
        print(f"Error al cargar LISTA_DE_MATERIALES: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/listas/control_produccion")
@login_requerido
def lista_control_produccion():
    """Cargar dinámicamente la lista de Control de Producción"""
    try:
        return render_template("LISTAS/LISTA_CONTROLDEPRODUCCION.html")
    except Exception as e:
        print(f"Error al cargar LISTA_CONTROLDEPRODUCCION: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


# Renders HTML de Control de produccion migrados a
# app/api/control_produccion/views.py (2026-05-25):
#   /control_produccion/control_embarque
#   /Control de embarque
#   /control_produccion/crear_plan
#   /control_produccion/plan_smt
# Se registran via blueprint en app/api/__init__.py.


# ==========================================
# AGENTE GENERADOR DE PLAN SMD
# ==========================================
# Migrado a app/api/control_produccion/plan_smd.py (2026-05-25).
# Las 4 rutas se registran via blueprint en app/api/__init__.py.
# Reexportamos las funciones de bootstrap para preservar `startup_init.py`.
from app.api.control_produccion.plan_smd import (  # noqa: E402, F401
    crear_tabla_plan_smd,
    crear_tabla_plan_smd_runs,
)


# Migracion 2026-05-27: GET /api/work-orders movido a
# app/api/control_produccion/po_wo.py como GET /api/work_orders
# (con flag ?include_import_status=1 para el shape legacy).
# La URL legacy /api/work-orders responde con 301 redirect (alias).


@app.route("/api/inventario/modelo/<codigo_modelo>", methods=["GET"])
@login_requerido
def api_inventario_modelo(codigo_modelo):
    """API para obtener inventario por código de modelo"""
    try:
        query = """
        SELECT modelo, nparte, stock_total, ubicaciones,
               ultima_entrada, ultima_salida, updated_at
        FROM inv_resumen_modelo
        WHERE nparte = %s
        """

        inventario = execute_query(query, (codigo_modelo,), fetch="all")

        # Formatear respuesta
        resultado = []
        for item in inventario:
            resultado.append(
                {
                    "modelo": item["modelo"],
                    "nparte": item["nparte"],
                    "stock_total": item["stock_total"] or 0,
                    "ubicaciones": item["ubicaciones"] or "",
                    "ultima_entrada": item["ultima_entrada"].strftime(
                        "%Y-%m-%d %H:%M:%S"
                    )
                    if item["ultima_entrada"]
                    else "",
                    "ultima_salida": item["ultima_salida"].strftime("%Y-%m-%d %H:%M:%S")
                    if item["ultima_salida"]
                    else "",
                    "updated_at": item["updated_at"].strftime("%Y-%m-%d %H:%M:%S")
                    if item["updated_at"]
                    else "",
                }
            )

        return jsonify(resultado)

    except Exception as e:
        print(f" Error en API inventario modelo {codigo_modelo}: {e}")
        return jsonify({"error": str(e)}), 500



# Limpieza 2026-05-27: control_produccion_smt_ajax eliminado (modulo descartado)


# Ruta eliminada - Control de operacion de linea SMT será reemplazado por Control BOM


# Limpieza 2026-05-27: control_operacion_linea_smt_ajax eliminado (modulo descartado)


# Rutas AJAX para todos los módulos de Control de Proceso
# Limpieza 2026-05-27: control_impresion_identificacion_smt_ajax eliminado (modulo descartado)


# Limpieza 2026-05-27: control_registro_identificacion_smt_ajax eliminado (modulo descartado)


@app.route("/historial-operacion-proceso-ajax")
@login_requerido
def historial_operacion_proceso_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Historial de operación de proceso"""
    try:
        return render_template(
            "Control de proceso/historial_operacion_proceso_ajax.html"
        )
    except Exception as e:
        print(f"Error al cargar template Historial de operación de proceso AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/bom-management-process-ajax")
@login_requerido
def bom_management_process_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de BOM Management Process"""
    try:
        return render_template("Control de proceso/bom_management_process_ajax.html")
    except Exception as e:
        print(f"Error al cargar template BOM Management Process AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/reporte-diario-inspeccion-smt-ajax")
@login_requerido
def reporte_diario_inspeccion_smt_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Reporte diario de inspección SMT"""
    try:
        return render_template(
            "Control de proceso/reporte_diario_inspeccion_smt_ajax.html"
        )
    except Exception as e:
        print(f"Error al cargar template Reporte diario de inspección SMT AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/control-diario-inspeccion-smt-ajax")
@login_requerido
def control_diario_inspeccion_smt_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Control diario de inspección SMT"""
    try:
        return render_template(
            "Control de proceso/control_diario_inspeccion_smt_ajax.html"
        )
    except Exception as e:
        print(f"Error al cargar template Control diario de inspección SMT AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/reporte-diario-inspeccion-proceso-ajax")
@login_requerido
def reporte_diario_inspeccion_proceso_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Reporte diario de inspección de proceso"""
    try:
        return render_template(
            "Control de proceso/reporte_diario_inspeccion_proceso_ajax.html"
        )
    except Exception as e:
        print(
            f"Error al cargar template Reporte diario de inspección de proceso AJAX: {e}"
        )
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/control-unidad-empaque-modelo-ajax")
@login_requerido
def control_unidad_empaque_modelo_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Control de unidad de empaque modelo"""
    try:
        return render_template(
            "Control de proceso/control_unidad_empaque_modelo_ajax.html"
        )
    except Exception as e:
        print(f"Error al cargar template Control de unidad de empaque modelo AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/packaging-register-management-ajax")
@login_requerido
def packaging_register_management_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Packaging Register Management"""
    try:
        return render_template(
            "Control de proceso/packaging_register_management_ajax.html"
        )
    except Exception as e:
        print(f"Error al cargar template Packaging Register Management AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/search-packaging-history-ajax")
@login_requerido
def search_packaging_history_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Search Packaging History"""
    try:
        return render_template("Control de proceso/search_packaging_history_ajax.html")
    except Exception as e:
        print(f"Error al cargar template Search Packaging History AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/shipping-register-management-ajax")
@login_requerido
def shipping_register_management_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Shipping Register Management"""
    try:
        return render_template(
            "Control de proceso/shipping_register_management_ajax.html"
        )
    except Exception as e:
        print(f"Error al cargar template Shipping Register Management AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/search-shipping-history-ajax")
@login_requerido
def search_shipping_history_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Search Shipping History"""
    try:
        return render_template("Control de proceso/search_shipping_history_ajax.html")
    except Exception as e:
        print(f"Error al cargar template Search Shipping History AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500



def _parse_fecha_control_salida_lineas(value, fallback):
    """Parsear fechas de filtros del modulo Control de salida de lineas."""
    try:
        return datetime.strptime(str(value or "").strip(), "%Y-%m-%d").date()
    except Exception:
        return fallback


def _calcular_estado_control_salida_lineas(produccion, oqc, almacen):
    if produccion == 0 and oqc == 0 and almacen > 0:
        return "Solo almacen"
    if oqc > produccion or almacen > oqc:
        return "Revisar"
    if produccion > 0 and oqc >= produccion and almacen >= oqc:
        return "Completo"
    if oqc < produccion:
        return "Pendiente OQC"
    if almacen < oqc:
        return "Pendiente almacen"
    return "Sin datos"


def _obtener_control_salida_lineas(limit=500):
    """Consultar produccion, liberacion OQC y entradas, acumuladas por parte en el rango."""
    today = date.today()
    default_from = today - timedelta(days=7)
    fecha_desde = _parse_fecha_control_salida_lineas(
        request.args.get("fecha_desde"),
        default_from,
    )
    fecha_hasta = _parse_fecha_control_salida_lineas(
        request.args.get("fecha_hasta"),
        today,
    )
    if fecha_hasta < fecha_desde:
        fecha_desde, fecha_hasta = fecha_hasta, fecha_desde

    fecha_inicio_sql = fecha_desde.strftime("%Y-%m-%d")
    fecha_fin_sql = (fecha_hasta + timedelta(days=1)).strftime("%Y-%m-%d")
    part_number = (request.args.get("part_number", "") or "").strip()
    part_like = f"%{part_number}%"
    output_collation = "utf8mb4_0900_ai_ci"

    production_part_expr = (
        "COALESCE(NULLIF(p.part_no, ''), "
        "NULLIF(LEFT(b.serial, GREATEST(CHAR_LENGTH(b.serial) - 12, 1)), ''), "
        "'SIN PARTE')"
    )
    production_model_expr = "COALESCE(NULLIF(p.model_code, ''), '')"
    oqc_part_expr = "COALESCE(NULLIF(pn.part_number, ''), CONCAT('ID-', er.part_number_id))"
    oqc_model_expr = "COALESCE(NULLIF(pn.model, ''), '')"
    entry_part_expr = "COALESCE(NULLIF(e.part_number, ''), 'SIN PARTE')"
    entry_model_expr = "COALESCE(NULLIF(e.product_model, ''), '')"
    production_part_select = f"({production_part_expr}) COLLATE {output_collation}"
    production_model_select = f"({production_model_expr}) COLLATE {output_collation}"
    oqc_part_select = f"({oqc_part_expr}) COLLATE {output_collation}"
    oqc_model_select = f"({oqc_model_expr}) COLLATE {output_collation}"
    entry_part_select = f"({entry_part_expr}) COLLATE {output_collation}"
    entry_model_select = f"({entry_model_expr}) COLLATE {output_collation}"
    empty_text_select = f"'' COLLATE {output_collation}"

    production_where = ["b.last_scan >= %s", "b.last_scan < %s"]
    production_params = [fecha_inicio_sql, fecha_fin_sql]
    oqc_where = [
        "COALESCE(er.exit_date, er.created_at, CAST(er.inspection_date AS DATETIME)) >= %s",
        "COALESCE(er.exit_date, er.created_at, CAST(er.inspection_date AS DATETIME)) < %s",
        "COALESCE(er.status, '') <> 'cancelled'",
        "COALESCE(er.qc_passed, 1) = 1",
    ]
    oqc_params = [fecha_inicio_sql, fecha_fin_sql]
    entry_where = [
        "COALESCE(e.movement_at, e.created_at) >= %s",
        "COALESCE(e.movement_at, e.created_at) < %s",
        "COALESCE(e.is_fifo_layer_only, 0) = 0",
    ]
    entry_params = [fecha_inicio_sql, fecha_fin_sql]

    if part_number:
        production_where.append(f"{production_part_select} LIKE %s")
        production_params.append(part_like)
        oqc_where.append(f"COALESCE(pn.part_number, '') COLLATE {output_collation} LIKE %s")
        oqc_params.append(part_like)
        entry_where.append(f"e.part_number COLLATE {output_collation} LIKE %s")
        entry_params.append(part_like)

    sql = f"""
        SELECT
            MIN(fuente.fecha) AS fecha_inicio,
            MAX(fuente.fecha) AS fecha_fin,
            fuente.part_number,
            MAX(NULLIF(fuente.product_model, '')) AS product_model,
            SUM(fuente.produced_quantity) AS produced_quantity,
            SUM(fuente.production_boxes) AS production_boxes,
            SUM(fuente.oqc_quantity) AS oqc_quantity,
            SUM(fuente.oqc_records) AS oqc_records,
            SUM(fuente.warehouse_quantity) AS warehouse_quantity,
            SUM(fuente.warehouse_records) AS warehouse_records,
            GROUP_CONCAT(DISTINCT NULLIF(fuente.lineas, '') ORDER BY fuente.lineas SEPARATOR ', ') AS lineas,
            GROUP_CONCAT(DISTINCT NULLIF(fuente.lotes, '') ORDER BY fuente.lotes SEPARATOR ', ') AS lotes,
            GROUP_CONCAT(DISTINCT NULLIF(fuente.oqc_statuses, '') ORDER BY fuente.oqc_statuses SEPARATOR ', ') AS oqc_statuses
        FROM (
            SELECT
                DATE(b.last_scan) AS fecha,
                {production_part_select} AS part_number,
                MAX({production_model_select}) AS product_model,
                COUNT(*) AS produced_quantity,
                COUNT(DISTINCT NULLIF(b.box_code, '')) AS production_boxes,
                0 AS oqc_quantity,
                0 AS oqc_records,
                0 AS warehouse_quantity,
                0 AS warehouse_records,
                (GROUP_CONCAT(DISTINCT NULLIF(p.line, '') ORDER BY p.line SEPARATOR ', ')) COLLATE {output_collation} AS lineas,
                (GROUP_CONCAT(DISTINCT NULLIF(b.lot_no, '') ORDER BY b.lot_no SEPARATOR ', ')) COLLATE {output_collation} AS lotes,
                {empty_text_select} AS oqc_statuses
            FROM box_scans b
            LEFT JOIN plan_main p ON p.lot_no = b.lot_no
            WHERE {" AND ".join(production_where)}
            GROUP BY DATE(b.last_scan), {production_part_select}

            UNION ALL

            SELECT
                DATE(COALESCE(er.exit_date, er.created_at, CAST(er.inspection_date AS DATETIME))) AS fecha,
                {oqc_part_select} AS part_number,
                MAX({oqc_model_select}) AS product_model,
                0 AS produced_quantity,
                0 AS production_boxes,
                SUM(er.quantity) AS oqc_quantity,
                COUNT(*) AS oqc_records,
                0 AS warehouse_quantity,
                0 AS warehouse_records,
                {empty_text_select} AS lineas,
                {empty_text_select} AS lotes,
                (GROUP_CONCAT(DISTINCT COALESCE(er.status, 'pending') ORDER BY er.status SEPARATOR ', ')) COLLATE {output_collation} AS oqc_statuses
            FROM exit_records er
            LEFT JOIN part_numbers pn ON pn.id = er.part_number_id
            WHERE {" AND ".join(oqc_where)}
            GROUP BY DATE(COALESCE(er.exit_date, er.created_at, CAST(er.inspection_date AS DATETIME))), {oqc_part_select}

            UNION ALL

            SELECT
                DATE(COALESCE(e.movement_at, e.created_at)) AS fecha,
                {entry_part_select} AS part_number,
                MAX({entry_model_select}) AS product_model,
                0 AS produced_quantity,
                0 AS production_boxes,
                0 AS oqc_quantity,
                0 AS oqc_records,
                SUM(e.quantity) AS warehouse_quantity,
                COUNT(*) AS warehouse_records,
                {empty_text_select} AS lineas,
                {empty_text_select} AS lotes,
                {empty_text_select} AS oqc_statuses
            FROM embarques_entrada_material e
            WHERE {" AND ".join(entry_where)}
            GROUP BY DATE(COALESCE(e.movement_at, e.created_at)), {entry_part_select}
        ) fuente
        WHERE COALESCE(fuente.part_number, '') <> ''
        GROUP BY fuente.part_number
        ORDER BY fecha_fin DESC, fuente.part_number
        LIMIT %s
    """
    params = production_params + oqc_params + entry_params + [int(limit)]
    rows = execute_query(sql, tuple(params), fetch="all") or []

    result_rows = []
    summary = {
        "produced_quantity": 0,
        "production_boxes": 0,
        "oqc_quantity": 0,
        "oqc_records": 0,
        "warehouse_quantity": 0,
        "warehouse_records": 0,
        "pending_oqc": 0,
        "pending_warehouse": 0,
    }
    for row in rows:
        part = _normalizar_texto_embarques_historial(row.get("part_number"))
        produced = int(row.get("produced_quantity") or 0)
        production_boxes = int(row.get("production_boxes") or 0)
        oqc = int(row.get("oqc_quantity") or 0)
        oqc_records = int(row.get("oqc_records") or 0)
        warehouse = int(row.get("warehouse_quantity") or 0)
        warehouse_records = int(row.get("warehouse_records") or 0)
        pending_oqc = max(produced - oqc, 0)
        pending_warehouse = max(oqc - warehouse, 0)
        if pending_oqc > 0:
            estado = "Pendiente OQC"
        elif pending_warehouse > 0:
            estado = "Pendiente almacen"
        else:
            estado = _calcular_estado_control_salida_lineas(produced, oqc, warehouse)
        fecha_inicio = _normalizar_texto_embarques_historial(row.get("fecha_inicio"))
        fecha_fin = _normalizar_texto_embarques_historial(row.get("fecha_fin"))
        fecha_periodo = (
            fecha_inicio
            if fecha_inicio == fecha_fin or not fecha_fin
            else f"{fecha_inicio} a {fecha_fin}"
        )

        summary["produced_quantity"] += produced
        summary["production_boxes"] += production_boxes
        summary["oqc_quantity"] += oqc
        summary["oqc_records"] += oqc_records
        summary["warehouse_quantity"] += warehouse
        summary["warehouse_records"] += warehouse_records
        summary["pending_oqc"] += pending_oqc
        summary["pending_warehouse"] += pending_warehouse

        result_rows.append(
            {
                "fecha": fecha_periodo,
                "fecha_inicio": fecha_inicio,
                "fecha_fin": fecha_fin,
                "part_number": part,
                "product_model": _normalizar_texto_embarques_historial(row.get("product_model")),
                "lineas": _normalizar_texto_embarques_historial(row.get("lineas")),
                "lotes": _normalizar_texto_embarques_historial(row.get("lotes")),
                "produced_quantity": produced,
                "production_boxes": production_boxes,
                "oqc_quantity": oqc,
                "oqc_records": oqc_records,
                "warehouse_quantity": warehouse,
                "warehouse_records": warehouse_records,
                "pending_oqc": pending_oqc,
                "pending_warehouse": pending_warehouse,
                "produced_cutoff": produced,
                "oqc_cutoff": oqc,
                "warehouse_cutoff": warehouse,
                "oqc_statuses": _normalizar_texto_embarques_historial(row.get("oqc_statuses")),
                "estado": estado,
            }
        )

    return {
        "rows": result_rows,
        "summary": summary,
        "filters": {
            "fecha_desde": fecha_desde.isoformat(),
            "fecha_hasta": fecha_hasta.isoformat(),
            "part_number": part_number,
        },
    }



@app.route("/control-salida-lineas-ajax")
@login_requerido
def control_salida_lineas_ajax():
    """Ruta AJAX para consultar salida de lineas contra OQC y almacen de embarques."""
    try:
        return render_template("Control de proceso/control_salida_lineas_ajax.html")
    except Exception as e:
        print(f"Error al cargar template Control de salida de lineas AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500



@app.route("/api/control-salida-lineas")
@login_requerido
def api_control_salida_lineas():
    """Obtener produccion, liberacion OQC y entradas de almacen por parte/fecha."""
    try:
        payload = _obtener_control_salida_lineas()
        payload["success"] = True
        return jsonify(payload)
    except Exception as e:
        print(f"Error API Control de salida de lineas: {e}\n{traceback.format_exc()}")
        return jsonify({"success": False, "error": str(e), "rows": []}), 500


@app.route("/api/control-salida-lineas/export")
@login_requerido
def export_control_salida_lineas():
    """Exportar Control de salida de lineas a Excel."""
    try:
        payload = _obtener_control_salida_lineas(limit=5000)
        return _exportar_historial_embarques_excel(
            "Salida Lineas",
            "control_salida_lineas.xlsx",
            {
                "Periodo": "fecha",
                "No. Parte": "part_number",
                "Produccion": "produced_quantity",
                "Liberacion OQC": "oqc_quantity",
                "Pendiente OQC": "pending_oqc",
                "Entradas Almacen": "warehouse_quantity",
                "Pendientes Almacen": "pending_warehouse",
            },
            payload["rows"],
        )
    except Exception as e:
        print(f"Error exportando Control de salida de lineas: {e}\n{traceback.format_exc()}")
        return jsonify({"success": False, "error": str(e)}), 500



@app.route("/registro-movimiento-identificacion-ajax")
@login_requerido
def registro_movimiento_identificacion_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Registro Movimiento Identificación"""
    try:
        return render_template(
            "Control de proceso/registro_movimiento_identificacion_ajax.html"
        )
    except Exception as e:
        print(f"Error al cargar template Registro Movimiento Identificación AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/control-otras-identificaciones-ajax")
@login_requerido
def control_otras_identificaciones_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Control Otras Identificaciones"""
    try:
        return render_template(
            "Control de proceso/control_otras_identificaciones_ajax.html"
        )
    except Exception as e:
        print(f"Error al cargar template Control Otras Identificaciones AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/control-movimiento-ns-producto-ajax")
@login_requerido
def control_movimiento_ns_producto_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Control Movimiento NS Producto"""
    try:
        return render_template(
            "Control de proceso/control_movimiento_ns_producto_ajax.html"
        )
    except Exception as e:
        print(f"Error al cargar template Control Movimiento NS Producto AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/model-sn-management-ajax")
@login_requerido
def model_sn_management_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Model SN Management"""
    try:
        return render_template("Control de proceso/model_sn_management_ajax.html")
    except Exception as e:
        print(f"Error al cargar template Model SN Management AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/control-scrap-ajax")
@login_requerido
def control_scrap_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Control Scrap"""
    try:
        return render_template("Control de proceso/control_scrap_ajax.html")
    except Exception as e:
        print(f"Error al cargar template Control Scrap AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


# Rutas AJAX para módulos de Control de Producción
@app.route("/line-material-status-ajax")
@login_requerido
def line_material_status_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Line Material Status_es"""
    try:
        return render_template(
            "Control de produccion/line_material_status_es_ajax.html"
        )
    except Exception as e:
        print(f"Error al cargar template Line Material Status_es AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


# Migracion 2026-05-26: 3 rutas template de Control de SMT (Metal Mask,
# Squeegee, Caja Metal Mask) movidas a sus blueprints en
# app/api/control_produccion/{metal_mask,squeegee,caja_metal_mask}.py


@app.route("/estandares-soldadura-ajax")
@login_requerido
def estandares_soldadura_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Estandares sobre control de soldadura"""
    try:
        return render_template("Control de produccion/estandares_soldadura_ajax.html")
    except Exception as e:
        print(
            f"Error al cargar template Estandares sobre control de soldadura AJAX: {e}"
        )
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/registro-recibo-soldadura-ajax")
@login_requerido
def registro_recibo_soldadura_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Registro de recibo de soldadura"""
    try:
        return render_template(
            "Control de produccion/registro_recibo_soldadura_ajax.html"
        )
    except Exception as e:
        print(f"Error al cargar template Registro de recibo de soldadura AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/control-salida-soldadura-ajax")
@login_requerido
def control_salida_soldadura_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Control de salida de soldadura"""
    try:
        return render_template(
            "Control de produccion/control_salida_soldadura_ajax.html"
        )
    except Exception as e:
        print(f"Error al cargar template Control de salida de soldadura AJAX: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/historial-tension-mask-metal-ajax")
@login_requerido
def historial_tension_mask_metal_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Historial de tension de mask de metal"""
    try:
        return render_template(
            "Control de produccion/historial_tension_mask_metal_ajax.html"
        )
    except Exception as e:
        print(
            f"Error al cargar template Historial de tension de mask de metal AJAX: {e}"
        )
        return f"Error al cargar el contenido: {str(e)}", 500


# Migracion 2026-05-27: ruta canonica movida a
# app/api/control_resultados/inventario_imd.py
# (/control_resultados/inventario_imd_terminado). Esta ruta legacy queda
# como redirect 301 para no romper sidebars cacheados.
@app.route("/control_proceso/inventario_imd_terminado")
def inventario_imd_terminado_legacy_redirect():
    return redirect("/control_resultados/inventario_imd_terminado", code=301)


@app.route("/listas/control_proceso")
@login_requerido
def lista_control_proceso():
    """Cargar dinámicamente la lista de Control de Proceso"""
    try:
        return render_template("LISTAS/LISTA_CONTROL_DE_PROCESO.html")
    except Exception as e:
        print(f"Error al cargar LISTA_CONTROL_DE_PROCESO: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/listas/control_calidad")
@login_requerido
def lista_control_calidad():
    """Cargar dinámicamente la lista de Control de Calidad"""
    try:
        return render_template("LISTAS/LISTA_CONTROL_DE_CALIDAD.html")
    except Exception as e:
        print(f"Error al cargar LISTA_CONTROL_DE_CALIDAD: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/listas/control_resultados")
@login_requerido
def lista_control_resultados():
    """Cargar dinámicamente la lista de Control de Resultados"""
    try:
        return render_template("LISTAS/LISTA_DE_CONTROL_DE_RESULTADOS.html")
    except Exception as e:
        print(f"Error al cargar LISTA_DE_CONTROL_DE_RESULTADOS: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/historial-aoi")
@login_requerido
def historial_aoi():
    """Servir la página de Historial AOI"""
    try:
        return render_template("Control de resultados/Historial AOI.html")
    except Exception as e:
        print(f"Error al cargar Historial AOI: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/historial-ict-ajax")
@login_requerido
def historial_ict_ajax():
    """Ruta AJAX para cargar el Historial ICT"""
    try:
        return render_template("Control de resultados/history_ict.html")
    except Exception as e:
        print(f"Error al cargar template de Historial ICT: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500

@app.route("/historial-maquina-ict-pass-fail")
@app.route("/historial-maquina-ict-pass-fail-ajax")
@login_requerido
def historial_maquina_ict_pass_fail():
    """Servir la página de Historial maquina ICT % Pass/Fail"""
    try:
        return render_template("Control de resultados/history_ict_Pass_Fail.html")
    except Exception as e:
        print(f"Error al cargar Historial maquina ICT % Pass/Fail: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


def _ict_pass_fail_fecha_jornada_expr(ts_column="ts"):
    return (
        f"CASE WHEN TIME({ts_column}) >= '07:30:00' "
        f"THEN DATE({ts_column}) ELSE DATE(DATE_SUB({ts_column}, INTERVAL 1 DAY)) END"
    )


def _ict_pass_fail_turno_expr(ts_column="ts"):
    return (
        "CASE "
        f"WHEN TIME({ts_column}) >= '07:30:00' AND TIME({ts_column}) < '17:30:00' THEN 'DIA' "
        f"WHEN TIME({ts_column}) >= '17:30:00' AND TIME({ts_column}) < '22:00:00' THEN 'TIEMPO EXTRA' "
        "ELSE 'NOCHE' "
        "END"
    )


def _build_history_ict_pass_fail_summary_query():
    """Construir query para resumen Pass/Fail de ICT por jornada, turno, linea y numero de parte."""
    fecha_desde = (
        request.args.get("fecha_desde", "").strip()
        or request.args.get("fecha", "").strip()
    )
    fecha_hasta = request.args.get("fecha_hasta", "").strip() or fecha_desde
    numero_parte = (
        request.args.get("numero_parte", "").strip()
        or request.args.get("no_parte", "").strip()
    )
    turno = request.args.get("turno", "").strip().upper()
    barcode = (
        request.args.get("barcode", "").strip()
        or request.args.get("barcode_like", "").strip()
    )

    fecha_jornada_expr = _ict_pass_fail_fecha_jornada_expr("ts")
    turno_expr = _ict_pass_fail_turno_expr("ts")

    sql = (
        "SELECT "
        f"{fecha_jornada_expr} AS fecha, "
        "COALESCE(NULLIF(TRIM(linea), ''), 'SIN LINEA') AS linea, "
        "COALESCE(ict, 0) AS ict, "
        f"{turno_expr} AS turno, "
        "COALESCE(NULLIF(TRIM(no_parte), ''), 'SIN NUMERO DE PARTE') AS numero_parte, "
        "COUNT(*) AS total, "
        "SUM(CASE WHEN UPPER(COALESCE(resultado, '')) = 'OK' THEN 1 ELSE 0 END) AS ok_count, "
        "SUM(CASE WHEN UPPER(COALESCE(resultado, '')) = 'NG' THEN 1 ELSE 0 END) AS ng_count "
        "FROM history_ict WHERE 1=1"
    )
    params = []

    if fecha_desde:
        start_date = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
        params.append(datetime.combine(start_date, dt_time(7, 30)))
        sql += " AND ts >= %s"
    if fecha_hasta:
        end_date = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
        params.append(datetime.combine(end_date + timedelta(days=1), dt_time(7, 30)))
        sql += " AND ts < %s"
    if numero_parte:
        sql += " AND no_parte LIKE %s"
        params.append(f"{numero_parte}%")
    if barcode:
        sql += " AND barcode LIKE %s"
        params.append(f"{barcode}%")
    if turno in {"DIA", "TIEMPO EXTRA", "NOCHE"}:
        sql += f" AND {turno_expr}=%s"
        params.append(turno)

    sql += (
        f" GROUP BY {fecha_jornada_expr},"
        " COALESCE(NULLIF(TRIM(linea), ''), 'SIN LINEA'),"
        " COALESCE(ict, 0),"
        f" {turno_expr},"
        " COALESCE(NULLIF(TRIM(no_parte), ''), 'SIN NUMERO DE PARTE')"
        " ORDER BY fecha ASC, linea ASC, ict ASC, FIELD(turno, 'DIA', 'TIEMPO EXTRA', 'NOCHE'), total DESC, numero_parte ASC"
    )

    return sql, tuple(params) if params else None


@app.route("/api/ict/pass-fail")
@login_requerido
def ict_pass_fail_api():
    """Obtener resumen Pass/Fail de ICT con el mismo formato que Vision."""
    try:
        sql, params = _build_history_ict_pass_fail_summary_query()
        rows = execute_query(sql, params, fetch="all") or []

        result = []
        for row in rows:
            total = int(row.get("total") or 0)
            ok_count = int(row.get("ok_count") or 0)
            ng_count = int(row.get("ng_count") or 0)
            porcentaje_ok = round((ok_count / total) * 100, 2) if total else 0
            porcentaje_ng = round((ng_count / total) * 100, 2) if total else 0

            result.append(
                {
                    "fecha": _ict_format_row({"fecha": row.get("fecha")}).get("fecha", ""),
                    "linea": row.get("linea", "") or "",
                    "ict": row.get("ict", "") or "",
                    "turno": row.get("turno", "") or "",
                    "numero_parte": row.get("numero_parte", "") or "",
                    "total": total,
                    "ok_count": ok_count,
                    "ng_count": ng_count,
                    "porcentaje_ok": porcentaje_ok,
                    "porcentaje_ng": porcentaje_ng,
                }
            )

        return jsonify(result)
    except Exception as e:
        import traceback

        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500
@app.route("/api/ict/pass-fail/detail")
@login_requerido
def ict_pass_fail_detail_api():
    """Obtener detalle de barcodes para una fila del resumen ICT Pass/Fail."""
    try:
        fecha = request.args.get("fecha", "").strip()
        linea = request.args.get("linea", "").strip()
        ict = request.args.get("ict", "").strip()
        turno = request.args.get("turno", "").strip().upper()
        numero_parte = request.args.get("numero_parte", "").strip()
        barcode = request.args.get("barcode", "").strip()

        try:
            min_intentos = max(
                1, int(request.args.get("min_intentos", "").strip() or 1)
            )
        except ValueError:
            min_intentos = 1

        if not all([fecha, linea, ict, turno, numero_parte]):
            return jsonify({"error": "Faltan filtros para consultar el detalle."}), 400

        if turno not in {"DIA", "TIEMPO EXTRA", "NOCHE"}:
            return jsonify({"error": "Turno invalido para consultar detalle."}), 400

        fecha_jornada = datetime.strptime(fecha, "%Y-%m-%d").date()
        fecha_inicio = datetime.combine(fecha_jornada, dt_time(7, 30))
        fecha_fin = datetime.combine(fecha_jornada + timedelta(days=1), dt_time(7, 30))

        linea_expr = "COALESCE(NULLIF(TRIM(h.linea), ''), 'SIN LINEA')"
        ict_expr = "COALESCE(h.ict, 0)"
        numero_parte_expr = (
            "COALESCE(NULLIF(TRIM(h.no_parte), ''), 'SIN NUMERO DE PARTE')"
        )
        turno_expr = _ict_pass_fail_turno_expr("h.ts")

        sql = (
            "SELECT detalle.barcode, detalle.numero_parte, detalle.linea, "
            "detalle.ict, detalle.turno, detalle.primer_test, detalle.ultimo_test, "
            "detalle.intentos, detalle.ok_count, detalle.ng_count, "
            "detalle.resultado_primer, detalle.resultado_final, "
            "COALESCE(reparacion.defect_count, 0) AS defect_count, "
            "CASE WHEN reparacion.defect_count IS NULL THEN 0 ELSE 1 END AS fue_reparacion, "
            "reparacion.primera_reparacion, "
            "COALESCE(reparacion.defectos, '') AS defectos "
            "FROM ("
            " SELECT h.barcode AS barcode, "
            f" MIN({numero_parte_expr}) AS numero_parte, "
            f" MIN({linea_expr}) AS linea, "
            f" MIN({ict_expr}) AS ict, "
            f" MIN({turno_expr}) AS turno, "
            " MIN(h.ts) AS primer_test, "
            " MAX(h.ts) AS ultimo_test, "
            " COUNT(*) AS intentos, "
            " SUM(CASE WHEN UPPER(COALESCE(h.resultado, '')) = 'OK' THEN 1 ELSE 0 END) AS ok_count, "
            " SUM(CASE WHEN UPPER(COALESCE(h.resultado, '')) = 'NG' THEN 1 ELSE 0 END) AS ng_count, "
            " SUBSTRING_INDEX(GROUP_CONCAT(UPPER(COALESCE(h.resultado, '')) ORDER BY h.ts ASC, h.id ASC SEPARATOR ','), ',', 1) AS resultado_primer, "
            " SUBSTRING_INDEX(GROUP_CONCAT(UPPER(COALESCE(h.resultado, '')) ORDER BY h.ts DESC, h.id DESC SEPARATOR ','), ',', 1) AS resultado_final "
            " FROM history_ict h "
            " WHERE h.ts >= %s AND h.ts < %s "
            f" AND {linea_expr}=%s "
            f" AND {ict_expr}=%s "
            f" AND {turno_expr}=%s "
            f" AND {numero_parte_expr}=%s "
        )
        params = [fecha_inicio, fecha_fin, linea, ict, turno, numero_parte]

        if barcode:
            sql += " AND h.barcode LIKE %s "
            params.append(f"{barcode}%")

        sql += (
            " GROUP BY h.barcode "
            " HAVING COUNT(*) >= %s "
            ") detalle "
            "LEFT JOIN ("
            " SELECT codigo, COUNT(*) AS defect_count, MIN(fecha) AS primera_reparacion, "
            " GROUP_CONCAT(DISTINCT NULLIF(TRIM(defecto), '') ORDER BY defecto SEPARATOR ', ') AS defectos "
            " FROM defect_data GROUP BY codigo"
            ") reparacion ON reparacion.codigo = detalle.barcode "
            "ORDER BY detalle.intentos DESC, detalle.ultimo_test DESC, detalle.barcode ASC"
        )
        params.append(min_intentos)

        rows = execute_query(sql, tuple(params), fetch="all") or []
        formatted_rows = []
        for row in rows:
            formatted = _ict_format_row(row)
            intentos = int(formatted.get("intentos") or 0)
            ok_count = int(formatted.get("ok_count") or 0)
            ng_count = int(formatted.get("ng_count") or 0)
            defect_count = int(formatted.get("defect_count") or 0)
            fue_reparacion = bool(int(formatted.get("fue_reparacion") or 0))
            resultado_primer = (formatted.get("resultado_primer") or "").upper()
            resultado_final = (formatted.get("resultado_final") or "").upper()

            formatted["intentos"] = intentos
            formatted["ok_count"] = ok_count
            formatted["ng_count"] = ng_count
            formatted["defect_count"] = defect_count
            formatted["fue_reparacion"] = fue_reparacion
            formatted["resultado_primer"] = resultado_primer
            formatted["resultado_final"] = resultado_final
            formatted["pass_real"] = resultado_primer == "OK" or (
                fue_reparacion and ok_count > 0
            )
            formatted_rows.append(formatted)

        total_intentos = sum(row["intentos"] for row in formatted_rows)
        ok_total = sum(row["ok_count"] for row in formatted_rows)
        ng_total = sum(row["ng_count"] for row in formatted_rows)
        piezas_unicas = len(formatted_rows)
        piezas_repetidas = sum(1 for row in formatted_rows if row["intentos"] > 1)
        piezas_reparacion = sum(1 for row in formatted_rows if row["fue_reparacion"])
        pass_real = sum(1 for row in formatted_rows if row["pass_real"])
        porcentaje_pass_real = (
            round((pass_real / piezas_unicas) * 100, 2) if piezas_unicas else 0
        )

        return jsonify(
            {
                "summary": {
                    "fecha": fecha,
                    "linea": linea,
                    "ict": ict,
                    "turno": turno,
                    "numero_parte": numero_parte,
                    "total_intentos": total_intentos,
                    "ok_total": ok_total,
                    "ng_total": ng_total,
                    "piezas_unicas": piezas_unicas,
                    "piezas_repetidas": piezas_repetidas,
                    "piezas_reparacion": piezas_reparacion,
                    "pass_real": pass_real,
                    "porcentaje_pass_real": porcentaje_pass_real,
                },
                "rows": formatted_rows,
            }
        )
    except ValueError:
        return jsonify({"error": "Fecha invalida para consultar detalle."}), 400
    except Exception as e:
        import traceback

        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/api/ict/pass-fail/export")
@login_requerido
def ict_pass_fail_export():
    """Exportar resumen Pass/Fail de ICT a un archivo de Excel."""
    try:
        sql, params = _build_history_ict_pass_fail_summary_query()
        rows = execute_query(sql, params, fetch="all") or []

        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "ICT Pass Fail"

        header_fill = PatternFill(
            start_color="3f6b6e", end_color="3f6b6e", fill_type="solid"
        )
        cell_fill = PatternFill(
            start_color="a1a09c", end_color="a1a09c", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        headers = [
            "Fecha",
            "Linea",
            "ICT",
            "Turno",
            "Numero de parte",
            "Total",
            "OK",
            "NG",
            "% Pass",
            "% Fail",
            "PORCENTAJE",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_idx, row in enumerate(rows, start=2):
            total = int(row.get("total") or 0)
            ok_count = int(row.get("ok_count") or 0)
            ng_count = int(row.get("ng_count") or 0)
            porcentaje_ok = round((ok_count / total) * 100, 2) if total else 0
            porcentaje_ng = round((ng_count / total) * 100, 2) if total else 0

            values = [
                _ict_format_row({"fecha": row.get("fecha")}).get("fecha", ""),
                row.get("linea", "") or "",
                row.get("ict", "") or "",
                row.get("turno", "") or "",
                row.get("numero_parte", "") or "",
                total,
                ok_count,
                ng_count,
                porcentaje_ok,
                porcentaje_ng,
            ]

            for col_num, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_num, value=value)
                cell.fill = cell_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            image_cell = ws.cell(row=row_idx, column=11, value="")
            image_cell.fill = cell_fill
            image_cell.alignment = Alignment(horizontal="center", vertical="center")
            image_cell.border = border

            excel_image = _create_vision_pass_fail_excel_image(
                porcentaje_ok, porcentaje_ng
            )
            ws.add_image(excel_image, f"K{row_idx}")
            ws.row_dimensions[row_idx].height = 24

        column_widths = [14, 20, 10, 18, 28, 14, 12, 12, 14, 14, 58]
        for col_num, width in enumerate(column_widths, start=1):
            column_letter = ws.cell(row=1, column=col_num).column_letter
            ws.column_dimensions[column_letter].width = width

        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = (
            f"historial_ict_pass_fail_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        return _send_excel_download(output, filename)
    except Exception as e:
        import traceback

        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500

@app.route("/historial-cambios-parametros-ict-ajax")
@login_requerido
def historial_cambios_parametros_ict_ajax():
    """Ruta AJAX para cargar el Historial de Cambios de Parámetros ICT"""
    try:
        return render_template(
            "Control de resultados/historial_cambios_parametros_ict_ajax.html"
        )
    except Exception as e:
        print(f"Error al cargar template de Historial Cambios Parámetros ICT: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


def crear_indice_history_ict_audit():
    """Crea indice cubriente para acelerar el modulo de Cambios de Parametros ICT.

    El SQL de _ict_compute_parameter_changes filtra por (ict, ts) y agrupa por
    (no_parte, linea, ict, fuente_archivo). Este indice cubre el WHERE + GROUP BY
    sin tocar la tabla base.
    """
    try:
        execute_query(
            "CREATE INDEX idx_history_ict_audit "
            "ON history_ict (ict, ts, no_parte, linea, fuente_archivo)"
        )
        print("Indice idx_history_ict_audit creado")
    except Exception as e:
        msg = str(e)
        if "1061" in msg or "Duplicate key name" in msg:
            return
        print(f"(info) idx_history_ict_audit no se pudo crear: {e}")


def crear_indice_history_ict_ts_nopart():
    """Indice cubriente para consultas sin filtro por ict (ict_all=1).

    El indice idx_history_ict_audit empieza con (ict, ts, ...). Cuando la query
    omite `AND ict = ?`, MySQL no puede usar el prefijo y cae a un scan completo.
    Este segundo indice arranca en (ts, ...) para cubrir esos casos sin scan.
    """
    try:
        execute_query(
            "CREATE INDEX idx_history_ict_ts_nopart "
            "ON history_ict (ts, no_parte, linea, fuente_archivo)"
        )
        print("Indice idx_history_ict_ts_nopart creado")
    except Exception as e:
        msg = str(e)
        if "1061" in msg or "Duplicate key name" in msg:
            return
        print(f"(info) idx_history_ict_ts_nopart no se pudo crear: {e}")


# Indices history_ict movidos a app/startup_init.py


_ICT_PARAM_JORNADA_START = dt_time(7, 30)

# ----- Progreso de calculo de cambios de parametros ICT -----
# Diccionario en memoria con el progreso de cada request activo, identificado
# por progress_id que envia el frontend. Permite barra de progreso real.
_ICT_PARAM_PROGRESS = {}
_ICT_PARAM_PROGRESS_LOCK = threading.Lock()
_ICT_PARAM_PROGRESS_TTL = 120  # segundos


def _ict_param_progress_set(progress_id, **fields):
    if not progress_id:
        return
    with _ICT_PARAM_PROGRESS_LOCK:
        now = time.monotonic()
        # Evict entradas viejas para evitar crecimiento ilimitado
        stale = [pid for pid, st in _ICT_PARAM_PROGRESS.items()
                 if now - st.get("started", now) > _ICT_PARAM_PROGRESS_TTL]
        for pid in stale:
            _ICT_PARAM_PROGRESS.pop(pid, None)
        state = _ICT_PARAM_PROGRESS.setdefault(progress_id, {"started": now})
        state.update(fields)


def _ict_param_progress_increment(progress_id, by=1):
    if not progress_id:
        return
    with _ICT_PARAM_PROGRESS_LOCK:
        state = _ICT_PARAM_PROGRESS.get(progress_id)
        if state:
            state["done"] = state.get("done", 0) + by


_ICT_PARAM_CHANGE_FIELDS = (
    ("std_value", "STD"),
    ("std_unit", "UNIT (STD)"),
    ("hlim_pct", "HLIM %"),
    ("llim_pct", "LLIM %"),
    ("hp_value", "HP"),
    ("lp_value", "LP"),
    ("ws_value", "WS"),
    ("ds_value", "DS"),
    ("rc_value", "RC"),
    ("p_flag", "P"),
    ("j_flag", "J"),
)


def _ict_param_parse_date(value, field_name):
    value = (value or "").strip()
    if not value:
        raise ValueError(f"{field_name} es requerido.")
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"{field_name} debe tener formato YYYY-MM-DD.") from exc


def _ict_param_parse_time(value, field_name):
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(value, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"{field_name} debe tener formato HH:MM.")


def _ict_param_as_datetime(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, dt_time.min)
    if isinstance(value, str):
        raw = value.strip()
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M:%S.%f",
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%dT%H:%M:%S.%f",
        ):
            try:
                return datetime.strptime(raw, fmt)
            except ValueError:
                continue
    return None


def _ict_param_parse_ict_filter(raw_value):
    raw = (raw_value or "").strip().upper()
    if not raw:
        raise ValueError("ICT es requerido.")

    ict_match = re.search(r"\bICT\s*([0-9]+)\b", raw)
    if ict_match:
        ict_value = int(ict_match.group(1))
    elif raw.isdigit():
        ict_value = int(raw)
    else:
        raise ValueError("ICT debe ser un numero o texto como ICT1.")

    line_match = re.search(r"\b(M|DP|H)\s*([0-9]+)\b", raw)
    line_value = f"{line_match.group(1)}{line_match.group(2)}" if line_match else ""
    return ict_value, line_value


def _ict_param_format_ict(linea, ict):
    linea = str(linea or "").strip()
    ict = str(ict or "").strip()
    if linea and ict:
        return f"{linea} ICT{ict}"
    if ict:
        return f"ICT{ict}"
    return linea


def _ict_param_jornada_label(ts_value):
    ts = _ict_param_as_datetime(ts_value)
    if not ts:
        return ""
    jornada = ts.date() if ts.time() >= _ICT_PARAM_JORNADA_START else ts.date() - timedelta(days=1)
    return jornada.isoformat()


def _ict_param_time_allowed(ts_value, hora_desde, hora_hasta):
    ts = _ict_param_as_datetime(ts_value)
    if not ts:
        return False
    current = ts.time()
    if hora_desde and hora_hasta:
        if hora_desde <= hora_hasta:
            return hora_desde <= current <= hora_hasta
        return current >= hora_desde or current <= hora_hasta
    if hora_desde:
        return current >= hora_desde
    if hora_hasta:
        return current <= hora_hasta
    return True


def _ict_param_compare_token(value):
    if value is None:
        return ("empty", "")
    raw = str(value).strip()
    if raw == "":
        return ("empty", "")
    try:
        return ("num", Decimal(raw.replace(",", "")).normalize())
    except Exception:
        return ("text", raw.upper())


def _ict_param_display_value(value):
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, float):
        return f"{value:g}"
    return str(value).strip()


def _ict_param_component_label(componente, pinref):
    componente = str(componente or "").strip()
    pinref = str(pinref or "").strip()
    if componente and pinref:
        return f"{componente} / {pinref}"
    return componente or pinref


def _ict_param_build_snapshot(param_rows):
    """Snapshot por (componente, pinref) con valores crudos y tokens
    precomputados.

    Estructura: {(comp, pin): {"values": {field: raw}, "tokens": {field: token}}}.
    Precomputar el token (que invoca Decimal.normalize) ahorra ~22k normalizaciones
    por grupo en el bucle de comparacion.
    """
    snapshot = {}
    fields = _ICT_PARAM_CHANGE_FIELDS
    for row in param_rows:
        componente = str(row.get("componente") or "").strip()
        pinref = str(row.get("pinref") or "").strip()
        if not componente and not pinref:
            continue
        values = {}
        tokens = {}
        for field_key, _label in fields:
            v = row.get(field_key)
            values[field_key] = v
            tokens[field_key] = _ict_param_compare_token(v)
        snapshot[(componente, pinref)] = {"values": values, "tokens": tokens}
    return snapshot


def _ict_param_load_snapshot(source_file, barcode, warnings):
    source_file = str(source_file or "").strip()
    barcode = str(barcode or "").strip()
    if not source_file:
        warnings.append("Se omitio un registro sin fuente_archivo.")
        return None
    if not barcode:
        warnings.append(f"Se omitio {source_file}: no tiene barcode representativo.")
        return None

    try:
        lgd_path = resolve_lgd_path(source_file)
        param_rows = get_lgd_parameters_for_barcode(str(lgd_path), barcode)
    except (IctLgdError, OSError) as exc:
        warnings.append(f"No se pudo leer {source_file}: {exc}")
        return None
    except Exception as exc:
        warnings.append(f"No se pudo parsear {source_file}: {exc}")
        return None

    if not param_rows:
        warnings.append(f"Sin parametros para {barcode} en {source_file}.")
        return None
    return _ict_param_build_snapshot(param_rows)


def _ict_compute_parameter_changes(
    fecha_desde,
    fecha_hasta,
    hora_desde,
    hora_hasta,
    ict_filter,
    no_parte_filter="",
    componente_filter="",
    parametro_filter="",
    limit=1000,
    ict_all=False,
    progress_id=None,
):
    _ict_param_progress_set(progress_id, total=0, done=0, phase="iniciando")
    start_date = _ict_param_parse_date(fecha_desde, "fecha_desde")
    end_date = _ict_param_parse_date(fecha_hasta or fecha_desde, "fecha_hasta")
    if end_date < start_date:
        raise ValueError("fecha_hasta no puede ser menor que fecha_desde.")

    if ict_all:
        ict_value = None
        line_filter = ""
    else:
        ict_value, line_filter = _ict_param_parse_ict_filter(ict_filter)
    hora_desde_value = _ict_param_parse_time(hora_desde, "hora_desde")
    hora_hasta_value = _ict_param_parse_time(hora_hasta, "hora_hasta")
    no_parte_filter = (no_parte_filter or "").strip()
    componente_filter = (componente_filter or "").strip().lower()
    parametro_filter = (parametro_filter or "").strip().lower()

    jornada_start = datetime.combine(start_date, _ICT_PARAM_JORNADA_START)
    jornada_end = datetime.combine(end_date + timedelta(days=1), _ICT_PARAM_JORNADA_START)

    sql = (
        "SELECT MIN(barcode) AS barcode, MIN(ts) AS first_ts, "
        "COALESCE(NULLIF(TRIM(no_parte), ''), 'SIN NUMERO DE PARTE') AS no_parte, "
        "COALESCE(NULLIF(TRIM(linea), ''), 'SIN LINEA') AS linea, "
        "ict, fuente_archivo "
        "FROM history_ict "
        "WHERE ts >= %s AND ts < %s "
        "AND fuente_archivo IS NOT NULL AND fuente_archivo <> ''"
    )
    params = [jornada_start, jornada_end]

    if ict_value is not None:
        sql += " AND ict = %s"
        params.append(ict_value)
    if line_filter:
        sql += " AND linea = %s"
        params.append(line_filter)
    if no_parte_filter:
        sql += " AND no_parte LIKE %s"
        params.append(f"{no_parte_filter}%")

    sql += (
        " GROUP BY "
        "COALESCE(NULLIF(TRIM(no_parte), ''), 'SIN NUMERO DE PARTE'), "
        "COALESCE(NULLIF(TRIM(linea), ''), 'SIN LINEA'), "
        "ict, fuente_archivo "
        "ORDER BY no_parte ASC, linea ASC, ict ASC, first_ts ASC"
    )

    _ict_param_progress_set(progress_id, phase="consultando_db")
    source_rows = execute_query(sql, tuple(params), fetch="all") or []
    warnings = []
    warnings_lock = threading.Lock()

    # Modo "pares consecutivos": para cada grupo (linea, ict, no_parte)
    # ordenamos los archivos .lgd por ts y comparamos cada par consecutivo
    # (i, i+1). Esto detecta cambios intermedios (ej. 0->5->0) que un modo
    # primer-vs-ultimo perderia. Cada flip distinto se emite como una fila.
    groups_raw = {}
    for source_row in source_rows:
        ts = _ict_param_as_datetime(source_row.get("first_ts"))
        if not ts or not _ict_param_time_allowed(ts, hora_desde_value, hora_hasta_value):
            continue
        source_file = source_row.get("fuente_archivo") or ""
        if not source_file:
            continue
        gkey = (
            str(source_row.get("linea") or ""),
            str(source_row.get("ict") or ""),
            str(source_row.get("no_parte") or ""),
        )
        groups_raw.setdefault(gkey, []).append({
            "ts": ts,
            "source_file": source_file,
            "barcode": source_row.get("barcode") or "",
        })

    plan = {}  # gkey -> [items ordenados, archivos distintos consecutivos]
    for gkey, items in groups_raw.items():
        items.sort(key=lambda item: item["ts"])
        deduped = []
        for it in items:
            if deduped and deduped[-1]["source_file"] == it["source_file"]:
                continue
            deduped.append(it)
        if len(deduped) >= 2:
            plan[gkey] = deduped

    unique_files = {}
    for items in plan.values():
        for it in items:
            existing = unique_files.get(it["source_file"])
            if existing is None or it["ts"] < existing[0]:
                unique_files[it["source_file"]] = (it["ts"], it["barcode"])

    _ict_param_progress_set(
        progress_id,
        total=len(unique_files),
        done=0,
        phase="parseando_archivos",
    )

    def _parse_one(item):
        source_file, (_ts, barcode) = item
        local_warnings = []
        snapshot = _ict_param_load_snapshot(source_file, barcode, local_warnings)
        _ict_param_progress_increment(progress_id)
        return source_file, snapshot, local_warnings

    snapshot_by_file = {}
    if unique_files:
        max_workers = min(8, len(unique_files))
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            for source_file, snapshot, local_warnings in executor.map(_parse_one, list(unique_files.items())):
                snapshot_by_file[source_file] = snapshot
                if local_warnings:
                    with warnings_lock:
                        warnings.extend(local_warnings)

    _ict_param_progress_set(progress_id, phase="detectando_cambios")
    files_read = sum(1 for snap in snapshot_by_file.values() if snap is not None)

    # Optimizacion 1: filtrar campos por parametro_filter UNA VEZ por request,
    # no por (par x componente).
    fields_to_check = [
        (k, l) for k, l in _ICT_PARAM_CHANGE_FIELDS
        if not parametro_filter or parametro_filter in l.lower()
    ]

    events = []
    grupos_con_cambios = 0
    if fields_to_check:
        for gkey in sorted(plan):
            items = plan[gkey]
            linea_value, ict_value_g, no_parte_value = gkey
            grupo_tuvo_cambio = False
            # Optimizacion 2: cachear shared_keys del grupo (interseccion de
            # (comp, pin)). Se invalida solo si cambia la identidad del par de
            # snapshots, lo cual rara vez sucede dentro del mismo grupo.
            shared_keys_cache = None

            for previous, current in zip(items, items[1:]):
                prev_snapshot = snapshot_by_file.get(previous["source_file"])
                curr_snapshot = snapshot_by_file.get(current["source_file"])
                if prev_snapshot is None or curr_snapshot is None:
                    continue

                if (
                    shared_keys_cache is None
                    or shared_keys_cache[0] is not prev_snapshot
                    or shared_keys_cache[1] is not curr_snapshot
                ):
                    shared = sorted(
                        set(prev_snapshot.keys()) & set(curr_snapshot.keys()),
                        key=lambda item: (item[0], item[1]),
                    )
                    if componente_filter:
                        shared = [
                            ck for ck in shared
                            if componente_filter in _ict_param_component_label(*ck).lower()
                        ]
                    shared_keys_cache = (prev_snapshot, curr_snapshot, shared)
                shared_keys = shared_keys_cache[2]

                for comp_key in shared_keys:
                    componente_raw, pinref_raw = comp_key
                    prev_cell = prev_snapshot[comp_key]
                    curr_cell = curr_snapshot[comp_key]
                    prev_tokens = prev_cell["tokens"]
                    curr_tokens = curr_cell["tokens"]
                    prev_values = prev_cell["values"]
                    curr_values = curr_cell["values"]
                    component_label = None  # lazy

                    for field_key, field_label in fields_to_check:
                        # Optimizacion 3: comparacion con tokens precomputados,
                        # cero llamadas a Decimal.normalize en este bucle.
                        if prev_tokens[field_key] == curr_tokens[field_key]:
                            continue
                        if component_label is None:
                            component_label = _ict_param_component_label(componente_raw, pinref_raw)

                        grupo_tuvo_cambio = True
                        events.append({
                            "jornada": _ict_param_jornada_label(current["ts"]),
                            "hora_anterior": previous["ts"].strftime("%H:%M:%S"),
                            "hora_cambio": current["ts"].strftime("%H:%M:%S"),
                            "ict": _ict_param_format_ict(linea_value, ict_value_g),
                            "ict_num": ict_value_g,
                            "linea": linea_value,
                            "no_parte": no_parte_value,
                            "std": no_parte_value,
                            "componente": component_label,
                            "componente_raw": componente_raw,
                            "pinref": pinref_raw,
                            "parametro": field_label,
                            "field_key": field_key,
                            "valor_anterior": _ict_param_display_value(prev_values[field_key]),
                            "valor_nuevo": _ict_param_display_value(curr_values[field_key]),
                            "archivo_anterior": previous["source_file"],
                            "archivo_cambio": current["source_file"],
                            "barcode_anterior": previous["barcode"],
                            "barcode_cambio": current["barcode"],
                        })
            if grupo_tuvo_cambio:
                grupos_con_cambios += 1

    events.sort(
        key=lambda row: (row.get("jornada", ""), row.get("hora_cambio", "")),
        reverse=True,
    )
    total_events = len(events)
    limited_events = events[:limit] if limit else events
    file_warning_count = len(warnings)
    if limit and total_events > limit:
        warnings.append(f"Se muestran {limit} de {total_events} cambios. Use filtros para reducir la consulta.")

    _ict_param_progress_set(progress_id, phase="completado")

    return {
        "rows": limited_events,
        "warnings": warnings,
        "meta": {
            "archivos_consultados": len(source_rows),
            "grupos_total": len(groups_raw),
            "grupos_con_cambios": grupos_con_cambios,
            "archivos_unicos": len(unique_files),
            "archivos_leidos": files_read,
            "archivos_faltantes": file_warning_count,
            "eventos": total_events,
            "limite": limit,
            "jornada_inicio": jornada_start.strftime("%Y-%m-%d %H:%M:%S"),
            "jornada_fin": jornada_end.strftime("%Y-%m-%d %H:%M:%S"),
        },
    }


@app.route("/api/ict/param-changes/progress")
@login_requerido
def ict_param_changes_progress():
    """Devuelve el progreso del calculo de cambios de parametros ICT.

    El frontend manda progress_id al iniciar la consulta principal y luego
    pollea este endpoint para alimentar la barra de progreso.
    """
    pid = request.args.get("id", "").strip()
    with _ICT_PARAM_PROGRESS_LOCK:
        state = _ICT_PARAM_PROGRESS.get(pid)
        if not state:
            return jsonify({"total": 0, "done": 0, "phase": "desconocido"})
        return jsonify({
            "total": state.get("total", 0),
            "done": state.get("done", 0),
            "phase": state.get("phase", ""),
        })


@app.route("/api/ict/param-changes")
@login_requerido
def ict_param_changes_api():
    """API para obtener cambios de parametros ICT desde archivos LGD locales."""
    try:
        fecha = request.args.get("fecha", "").strip()
        fecha_desde = request.args.get("fecha_desde", "").strip()
        fecha_hasta = request.args.get("fecha_hasta", "").strip()
        payload = _ict_compute_parameter_changes(
            fecha_desde=fecha_desde or fecha,
            fecha_hasta=fecha_hasta or fecha_desde or fecha,
            hora_desde=request.args.get("hora_desde", "").strip(),
            hora_hasta=request.args.get("hora_hasta", "").strip(),
            ict_filter=request.args.get("ict", "").strip(),
            no_parte_filter=(
                request.args.get("no_parte", "").strip()
                or request.args.get("numero_parte", "").strip()
                or request.args.get("std", "").strip()
            ),
            componente_filter=request.args.get("componente", "").strip(),
            parametro_filter=request.args.get("parametro", "").strip(),
            limit=1000,
            ict_all=request.args.get("ict_all", "").strip() == "1",
            progress_id=request.args.get("progress_id", "").strip() or None,
        )
        return jsonify(payload)
    except ValueError as e:
        return jsonify({"error": str(e), "rows": [], "warnings": [], "meta": {}}), 400
    except Exception as e:
        import traceback

        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/api/ict/param-changes/export")
@login_requerido
def ict_param_changes_export():
    """Exportar cambios de parametros ICT calculados desde LGD locales."""
    try:
        fecha = request.args.get("fecha", "").strip()
        fecha_desde = request.args.get("fecha_desde", "").strip()
        fecha_hasta = request.args.get("fecha_hasta", "").strip()
        payload = _ict_compute_parameter_changes(
            fecha_desde=fecha_desde or fecha,
            fecha_hasta=fecha_hasta or fecha_desde or fecha,
            hora_desde=request.args.get("hora_desde", "").strip(),
            hora_hasta=request.args.get("hora_hasta", "").strip(),
            ict_filter=request.args.get("ict", "").strip(),
            no_parte_filter=(
                request.args.get("no_parte", "").strip()
                or request.args.get("numero_parte", "").strip()
                or request.args.get("std", "").strip()
            ),
            componente_filter=request.args.get("componente", "").strip(),
            parametro_filter=request.args.get("parametro", "").strip(),
            limit=5000,
            ict_all=request.args.get("ict_all", "").strip() == "1",
            progress_id=request.args.get("progress_id", "").strip() or None,
        )
        rows = payload.get("rows", [])

        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Cambios Parametros ICT"

        headers = [
            "Jornada",
            "Hora Anterior",
            "Hora Cambio",
            "ICT",
            "Linea",
            "No Parte",
            "Componente",
            "Parametro",
            "Valor Anterior",
            "Valor Nuevo",
            "Archivo Anterior",
            "Archivo Cambio",
        ]
        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row.get("jornada", "") or "")
            ws.cell(row=row_idx, column=2, value=row.get("hora_anterior", "") or "")
            ws.cell(row=row_idx, column=3, value=row.get("hora_cambio", "") or "")
            ws.cell(row=row_idx, column=4, value=row.get("ict", "") or "")
            ws.cell(row=row_idx, column=5, value=row.get("linea", "") or "")
            ws.cell(row=row_idx, column=6, value=row.get("no_parte", "") or "")
            ws.cell(row=row_idx, column=7, value=row.get("componente", "") or "")
            ws.cell(row=row_idx, column=8, value=row.get("parametro", "") or "")
            ws.cell(row=row_idx, column=9, value=row.get("valor_anterior", "") or "")
            ws.cell(row=row_idx, column=10, value=row.get("valor_nuevo", "") or "")
            ws.cell(row=row_idx, column=11, value=row.get("archivo_anterior", "") or "")
            ws.cell(row=row_idx, column=12, value=row.get("archivo_cambio", "") or "")

        col_widths = [12, 12, 12, 14, 8, 22, 25, 22, 20, 20, 48, 48]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = width
        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = (
            f"cambios_parametros_ict_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        return _send_excel_download(output, filename)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        import traceback

        print(f"Error exportando Cambios Parámetros ICT: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/ict/param-changes/detail")
@login_requerido
def ict_param_changes_detail():
    """Localiza el archivo .lgd mas temprano donde aparece valor_nuevo para un
    parametro especifico de un (linea, ict, no_parte). Usa busqueda binaria
    sobre los archivos del rango y un barrido lineal hacia atras para
    parametros que oscilan."""
    try:
        fecha = request.args.get("fecha", "").strip()
        fecha_desde_raw = request.args.get("fecha_desde", "").strip() or fecha
        fecha_hasta_raw = (
            request.args.get("fecha_hasta", "").strip()
            or fecha_desde_raw
            or fecha
        )
        linea = request.args.get("linea", "").strip()
        ict_raw = request.args.get("ict", "").strip()
        no_parte = request.args.get("no_parte", "").strip()
        componente = request.args.get("componente", "")
        pinref = request.args.get("pinref", "")
        field_key = request.args.get("parametro", "").strip()
        valor_nuevo = request.args.get("valor_nuevo", "")

        valid_fields = {fk for fk, _ in _ICT_PARAM_CHANGE_FIELDS}
        if field_key not in valid_fields:
            return jsonify({"error": "Parametro desconocido."}), 400
        if not no_parte:
            return jsonify({"error": "no_parte es requerido."}), 400
        if not ict_raw:
            return jsonify({"error": "ict es requerido."}), 400

        try:
            ict_value = int(ict_raw)
        except ValueError:
            return jsonify({"error": "ict debe ser numerico."}), 400

        start_date = _ict_param_parse_date(fecha_desde_raw, "fecha_desde")
        end_date = _ict_param_parse_date(fecha_hasta_raw, "fecha_hasta")
        if end_date < start_date:
            return jsonify({"error": "fecha_hasta no puede ser menor que fecha_desde."}), 400

        jornada_start = datetime.combine(start_date, _ICT_PARAM_JORNADA_START)
        jornada_end = datetime.combine(end_date + timedelta(days=1), _ICT_PARAM_JORNADA_START)

        sql = (
            "SELECT MIN(barcode) AS barcode, MIN(ts) AS first_ts, fuente_archivo "
            "FROM history_ict "
            "WHERE ts >= %s AND ts < %s "
            "AND ict = %s AND no_parte = %s "
            "AND fuente_archivo IS NOT NULL AND fuente_archivo <> ''"
        )
        params = [jornada_start, jornada_end, ict_value, no_parte]
        if linea:
            sql += " AND linea = %s"
            params.append(linea)
        sql += (
            " GROUP BY fuente_archivo "
            "ORDER BY first_ts ASC"
        )

        rows = execute_query(sql, tuple(params), fetch="all") or []
        files = []
        for row in rows:
            ts = _ict_param_as_datetime(row.get("first_ts"))
            source_file = row.get("fuente_archivo") or ""
            barcode = row.get("barcode") or ""
            if not ts or not source_file:
                continue
            files.append({"ts": ts, "file": source_file, "barcode": barcode})

        if not files:
            return jsonify({"found": False, "reason": "Sin archivos en el rango."})

        target_token = _ict_param_compare_token(valor_nuevo)
        comp_key = (str(componente or ""), str(pinref or ""))
        snap_cache = {}
        warnings_local = []

        def _matches(idx):
            entry = files[idx]
            cache_key = (entry["file"], entry["barcode"])
            snap = snap_cache.get(cache_key)
            if snap is None and cache_key not in snap_cache:
                snap = _ict_param_load_snapshot(
                    entry["file"], entry["barcode"], warnings_local
                )
                snap_cache[cache_key] = snap
            if snap is None:
                return None  # archivo no leible
            cell = snap.get(comp_key)
            if cell is None:
                return False
            # Snapshot ahora tiene la forma {"values": {...}, "tokens": {...}}.
            # Usamos los tokens precomputados para evitar Decimal.normalize aqui.
            return cell["tokens"].get(field_key) == target_token

        last_idx = len(files) - 1
        last_match = _matches(last_idx)
        if last_match is None:
            return jsonify({"found": False, "reason": "No se pudo leer el ultimo archivo."})
        if last_match is False:
            return jsonify({"found": False, "reason": "El valor nuevo no aparece en archivos posteriores."})

        lo, hi = 0, last_idx
        while lo < hi:
            mid = (lo + hi) // 2
            mid_match = _matches(mid)
            if mid_match is True:
                hi = mid
            else:
                # archivo no leible o sin valor_nuevo: buscar mas adelante
                lo = mid + 1

        # Barrido lineal hacia atras para tolerar oscilaciones
        while lo > 0:
            prev_match = _matches(lo - 1)
            if prev_match is True:
                lo -= 1
            else:
                break

        target = files[lo]
        return jsonify({
            "found": True,
            "hora": target["ts"].strftime("%H:%M:%S"),
            "fecha": target["ts"].strftime("%Y-%m-%d"),
            "ts": target["ts"].strftime("%Y-%m-%d %H:%M:%S"),
            "archivo": target["file"],
            "barcode": target["barcode"],
            "archivos_consultados": len(files),
            "archivos_parseados": len(snap_cache),
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        import traceback

        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/historial-aoi-ajax")
@login_requerido
def historial_aoi_ajax():
    """Ruta AJAX para cargar dinámicamente el contenido de Historial AOI"""
    try:
        return render_template("Control de resultados/Historial AOI.html")
    except Exception as e:
        print(f"Error al cargar template de Historial AOI: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/listas/control_reporte")
@login_requerido
def lista_control_reporte():
    """Cargar dinámicamente la lista de Control de Reporte"""
    try:
        return render_template("LISTAS/LISTA_DE_CONTROL_DE_REPORTE.html")
    except Exception as e:
        print(f"Error al cargar LISTA_DE_CONTROL_DE_REPORTE: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/listas/configuracion_programa")
@login_requerido
def lista_configuracion_programa():
    """Cargar dinámicamente la lista de Configuración de Programa"""
    try:
        return render_template("LISTAS/LISTA_DE_CONFIGPG.html")
    except Exception as e:
        print(f"Error al cargar LISTA_DE_CONFIGPG: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/material/info")
@login_requerido
def material_info():
    """Cargar dinámicamente la información general de material"""
    try:
        return render_template("info.html")
    except Exception as e:
        print(f"Error al cargar info.html: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500

@app.route("/templates/LISTAS/<filename>")
def serve_list_template(filename):
    """Servir plantillas de listas para el menú móvil"""
    try:
        # Verificar que el archivo existe y es uno de los permitidos
        allowed_files = [
            "LISTA_INFORMACIONBASICA.html",
            "LISTA_DE_MATERIALES.html",
            "LISTA_CONTROLDEPRODUCCION.html",
            "LISTA_CONTROL_DE_PROCESO.html",
            "LISTA_CONTROL_DE_CALIDAD.html",
            "LISTA_DE_CONFIGPG.html",
            "LISTA_DE_CONTROL_DE_REPORTE.html",
            "LISTA_DE_CONTROL_DE_RESULTADOS.html",
        ]

        if filename not in allowed_files:
            return "Archivo no encontrado", 404

        # Leer el archivo directamente
        template_folder = app.template_folder or "templates"
        template_path = os.path.join(template_folder, "LISTAS", filename)

        if not os.path.exists(template_path):
            return f"Archivo no encontrado: {template_path}", 404

        with open(template_path, "r", encoding="utf-8") as f:
            content = f.read()

        return content, 200, {"Content-Type": "text/html; charset=utf-8"}

    except Exception as e:
        print(f"Error sirviendo plantilla {filename}: {str(e)}")
        return f"Error cargando la plantilla: {str(e)}", 500


# ===== RUTAS PARA EL SISTEMA DE PERMISOS DROPDOWNS =====


@app.route("/verificar_permiso_dropdown", methods=["POST"])
def verificar_permiso_dropdown():
    """
    Verificar si el usuario actual tiene permiso para un dropdown específico
    """
    try:
        if "usuario" not in session:
            return jsonify(
                {"tiene_permiso": False, "error": "Usuario no autenticado"}
            ), 401

        # Obtener datos desde JSON
        data = request.get_json()
        if not data:
            return jsonify(
                {"tiene_permiso": False, "error": "Datos JSON requeridos"}
            ), 400

        pagina = data.get("pagina", "").strip()
        seccion = data.get("seccion", "").strip()
        boton = data.get("boton", "").strip()

        if not all([pagina, seccion, boton]):
            return jsonify(
                {"tiene_permiso": False, "error": "Parámetros incompletos"}
            ), 400

        username = session.get("usuario")
        rol_nombre = auth_system.obtener_rol_principal_usuario(username)

        if not rol_nombre:
            return jsonify(
                {"tiene_permiso": False, "error": "Usuario no encontrado o sin roles"}
            ), 404

        tiene_permiso = auth_system.verificar_permiso_boton(
            username, pagina, seccion, boton
        )

        return jsonify(
            {
                "tiene_permiso": tiene_permiso,
                "usuario": username,
                "rol": rol_nombre,
                "permiso": f"{pagina} > {seccion} > {boton}",
            }
        )

    except Exception as e:
        print(f"Error verificando permiso: {e}")
        return jsonify({"tiene_permiso": False, "error": str(e)}), 500


@app.route("/obtener_permisos_usuario_actual", methods=["GET"])
@login_requerido
def obtener_permisos_usuario_actual():
    """
    Obtener todos los permisos del usuario actual para caché en frontend
    """
    try:
        if "usuario" not in session:
            return jsonify({"permisos": [], "error": "Usuario no autenticado"}), 401

        username = session["usuario"]
        rol_nombre = auth_system.obtener_rol_principal_usuario(username)
        if not rol_nombre:
            return jsonify(
                {"permisos": {}, "error": "Usuario no encontrado o sin roles"}
            ), 404

        permisos = auth_system.obtener_permisos_botones_usuario(username)

        # Formatear permisos para JavaScript en estructura jerárquica
        permisos_jerarquicos = {}
        total_permisos = 0

        for pagina, secciones in permisos.items():
            permisos_jerarquicos[pagina] = {}
            for seccion, botones in secciones.items():
                permisos_jerarquicos[pagina][seccion] = []
                for item in botones:
                    boton = item.get("boton") if isinstance(item, dict) else item
                    if not boton:
                        continue
                    permisos_jerarquicos[pagina][seccion].append(boton)
                    total_permisos += 1

        return jsonify(
            {
                "permisos": permisos_jerarquicos,
                "usuario": username,
                "rol": rol_nombre,
                "total_permisos": total_permisos,
            }
        )

    except Exception as e:
        print(f"Error obteniendo permisos: {e}")
        return jsonify({"permisos": [], "error": str(e)}), 500


# ============== CSV VIEWER ROUTES ==============
@app.route("/csv-viewer")
@login_requerido
def csv_viewer():
    """Página principal del visor de CSV"""
    try:
        return render_template("csv-viewer.html")
    except Exception as e:
        print(f"Error al cargar CSV viewer: {e}")
        return f"Error al cargar la página: {str(e)}", 500


# Nueva ruta para historial de cambio de material de SMT
@app.route("/historial-cambio-material-smt")
@login_requerido
def historial_cambio_material_smt():
    """Página del historial de cambio de material de SMT"""
    try:
        return render_template("Control de calidad/historial_cambio_material_smt.html")
    except Exception as e:
        print(f"Error al cargar historial de cambio de material SMT: {e}")
        return f"Error al cargar la página: {str(e)}", 500


@app.route("/historial-cambio-material-smt-ajax")
def historial_cambio_material_smt_ajax():
    if "usuario" not in session:
        return redirect(url_for("login"))
    try:
        return render_template(
            "Control de calidad/historial_cambio_material_smt_ajax.html"
        )
    except Exception as e:
        print(f"Error en historial_cambio_material_smt_ajax: {e}")
        return f"Error interno del servidor: {e}", 500


@app.route("/api/csv_data")
@login_requerido
def get_csv_data():
    """API para obtener datos SMT desde MySQL (no archivos CSV)"""
    try:
        folder = request.args.get("folder", "")
        print(f" Solicitud recibida para carpeta: '{folder}'")

        if not folder:
            print(" No se proporcionó parámetro de carpeta")
            return jsonify(
                {"success": False, "error": "Folder parameter required"}
            ), 400

        # Conectar a MySQL directamente
        import mysql.connector

        mysql_config = {
            "host": os.getenv("MYSQL_HOST"),
            "port": int(os.getenv("MYSQL_PORT", 3306)),
            "user": os.getenv("MYSQL_USER"),
            "password": os.getenv("MYSQL_PASSWORD"),
            "database": os.getenv("MYSQL_DATABASE"),
            "charset": "utf8mb4",
        }

        conn = mysql.connector.connect(**mysql_config)
        cursor = conn.cursor(dictionary=True)

        print(f" Consultando datos SMT desde MySQL para carpeta: {folder}")

        # Query para obtener datos de la tabla MySQL
        query = """
            SELECT
            ScanDate,
            ScanTime,
            SlotNo,
            Result,
            PreviousBarcode,
            Productdate,
            PartName,
            Quantity,
            SEQ,
            Vendor,
            LOTNO,
            Barcode,
            FeederBase,
            archivo,
            linea,
            maquina,
            fecha_subida
        FROM logs_maquina
        WHERE archivo LIKE %s OR linea LIKE %s OR maquina LIKE %s
        ORDER BY ScanDate DESC, ScanTime DESC
        LIMIT 1000
        """

        cursor.execute(query, (f"%{folder}%", f"%{folder}%", f"%{folder}%"))
        resultados = cursor.fetchall()

        print(f"✓ Encontrados {len(resultados)} registros en MySQL")

        # Convertir datos para JSON
        all_data = []
        for resultado in resultados:
            cleaned_record = {}
            for key, value in resultado.items():
                if hasattr(value, "isoformat"):  # Es una fecha/datetime
                    cleaned_record[key] = value.isoformat()
                elif value is None:
                    cleaned_record[key] = None
                else:
                    cleaned_record[key] = str(value)

            # Mapear nombres para compatibilidad con frontend (usar nombres de la nueva tabla)
            cleaned_record["ScanDate"] = cleaned_record.get("ScanDate", "")
            cleaned_record["ScanTime"] = cleaned_record.get("ScanTime", "")
            cleaned_record["SlotNo"] = cleaned_record.get("SlotNo", "")
            cleaned_record["Result"] = cleaned_record.get("Result", "")
            cleaned_record["PartName"] = cleaned_record.get("PartName", "")
            cleaned_record["SourceFile"] = cleaned_record.get(
                "archivo", ""
            )  # Mapear archivo a SourceFile
            cleaned_record["LOTNO"] = cleaned_record.get("LOTNO", "")
            cleaned_record["Barcode"] = cleaned_record.get("Barcode", "")
            cleaned_record["Quantity"] = cleaned_record.get("Quantity", "")
            cleaned_record["Vendor"] = cleaned_record.get("Vendor", "")
            cleaned_record["FeederBase"] = cleaned_record.get("FeederBase", "")
            cleaned_record["PreviousBarcode"] = cleaned_record.get(
                "PreviousBarcode", ""
            )

            all_data.append(cleaned_record)

        cursor.close()
        conn.close()

        return jsonify(
            {
                "success": True,
                "data": all_data,
                "message": f"Datos MySQL cargados para {folder}: {len(all_data)} registros",
                "files_processed": len(
                    set([d.get("SourceFile", "") for d in all_data])
                ),
                "source": "mysql_logs_maquina",
            }
        )

    except Exception as e:
        print(f" Error obteniendo datos desde MySQL: {e}")
        print(f" Traceback: {traceback.format_exc()}")
        return jsonify(
            {
                "success": False,
                "error": f"Error al consultar base de datos MySQL: {str(e)}",
            }
        ), 500


@app.route("/api/csv_stats")
@login_requerido
def get_csv_stats():
    """API para obtener estadísticas SMT desde MySQL (no archivos CSV)"""
    try:
        folder = request.args.get("folder", "")
        print(f" Solicitud recibida para estadísticas de carpeta: '{folder}'")

        if not folder:
            print(" No se proporcionó parámetro de carpeta")
            return jsonify(
                {"success": False, "error": "Folder parameter required"}
            ), 400

        # Conectar a MySQL directamente
        import mysql.connector

        mysql_config = {
            "host": os.getenv("MYSQL_HOST"),
            "port": int(os.getenv("MYSQL_PORT", 3306)),
            "user": os.getenv("MYSQL_USER"),
            "password": os.getenv("MYSQL_PASSWORD"),
            "database": os.getenv("MYSQL_DATABASE"),
            "charset": "utf8mb4",
        }

        conn = mysql.connector.connect(**mysql_config)
        cursor = conn.cursor(dictionary=True)

        print(f" Consultando estadísticas SMT desde MySQL para carpeta: {folder}")

        # Query para obtener estadísticas de la tabla MySQL
        query = """
            SELECT
            COUNT(*) as total_records,
            COUNT(DISTINCT archivo) as total_files,
            COUNT(DISTINCT ScanDate) as total_days,
            COUNT(CASE WHEN Result = 'OK' THEN 1 END) as ok_count,
            COUNT(CASE WHEN Result = 'NG' THEN 1 END) as ng_count,
            MIN(ScanDate) as first_date,
            MAX(ScanDate) as last_date
        FROM logs_maquina
        WHERE archivo LIKE %s OR linea LIKE %s OR maquina LIKE %s
        """

        cursor.execute(query, (f"%{folder}%", f"%{folder}%", f"%{folder}%"))
        stats = cursor.fetchone()

        # Query para obtener archivos únicos
        files_query = """
        SELECT DISTINCT archivo, COUNT(*) as records
        FROM logs_maquina
        WHERE archivo LIKE %s OR linea LIKE %s OR maquina LIKE %s
        GROUP BY archivo
        ORDER BY archivo
        """

        cursor.execute(files_query, (f"%{folder}%", f"%{folder}%", f"%{folder}%"))
        files_info = cursor.fetchall()

        cursor.close()
        conn.close()

        print(
            f" Estadísticas obtenidas: {stats['total_records']} registros de {stats['total_files']} archivos"
        )

        return jsonify(
            {
                "success": True,
                "stats": {
                    "total_records": stats["total_records"] or 0,
                    "total_files": stats["total_files"] or 0,
                    "total_days": stats["total_days"] or 0,
                    "ok_count": stats["ok_count"] or 0,
                    "ng_count": stats["ng_count"] or 0,
                    "first_date": stats["first_date"].isoformat()
                    if stats["first_date"]
                    else None,
                    "last_date": stats["last_date"].isoformat()
                    if stats["last_date"]
                    else None,
                },
                "files": [
                    {"name": f["source_file"], "records": f["records"]}
                    for f in files_info
                ],
                "message": f"Estadísticas MySQL para {folder}",
                "source": "mysql",
            }
        )

    except Exception as e:
        print(f" Error obteniendo estadísticas desde MySQL: {e}")
        print(f" Traceback: {traceback.format_exc()}")
        return jsonify(
            {
                "success": False,
                "error": f"Error al consultar estadísticas MySQL: {str(e)}",
            }
        ), 500

        print(f" Encontrados {len(csv_files)} archivos CSV")

        # Leer y combinar todos los archivos CSV
        all_data = []

        for csv_file in csv_files:
            try:
                print(
                    f" Leyendo archivo: {os.path.basename(csv_file)} (tamaño: {os.path.getsize(csv_file)} bytes)"
                )

                # Intentar lectura simple primero
                try:
                    df = pd.read_csv(csv_file, encoding="utf-8", on_bad_lines="skip")
                    print(f" Lectura exitosa con pandas básico: {len(df)} filas")
                except Exception as simple_error:
                    print(f" Error con lectura simple: {str(simple_error)}")

                    # Leer el archivo como texto primero para limpiar formato
                    with open(csv_file, "r", encoding="utf-8", errors="ignore") as f:
                        content = f.read()

                    print(f" Contenido leído: {len(content)} caracteres")

                    # Limpiar saltos de línea incorrectos en el contenido
                    lines = content.strip().split("\n")
                    cleaned_lines = []

                    for line in lines:
                        # Si la línea no termina con una coma y la siguiente no empieza con una fecha,
                        # probablemente es una línea cortada
                        if line and not line.endswith(","):
                            cleaned_lines.append(line)
                        elif line.endswith(","):
                            # Línea que termina en coma, probablemente incompleta
                            if cleaned_lines:
                                cleaned_lines[-1] += line
                            else:
                                cleaned_lines.append(line)

                    print(
                        f"🧹 Líneas limpiadas: {len(cleaned_lines)} de {len(lines)} originales"
                    )

                    # Crear DataFrame desde el contenido limpio
                    from io import StringIO

                    cleaned_content = "\n".join(cleaned_lines)

                    # Leer el archivo CSV con pandas usando el contenido limpio
                    df = pd.read_csv(
                        StringIO(cleaned_content), encoding="utf-8", on_bad_lines="skip"
                    )

                print(f" DataFrame creado: {len(df)} filas, {len(df.columns)} columnas")
                print(f" Columnas: {list(df.columns)}")

                # Verificar que el DataFrame tenga las columnas esperadas
                expected_columns = [
                    "ScanDate",
                    "ScanTime",
                    "SlotNo",
                    "Result",
                    "PartName",
                ]
                missing_columns = [
                    col for col in expected_columns if col not in df.columns
                ]

                if missing_columns:
                    print(f" Columnas faltantes en {csv_file}: {missing_columns}")
                    # Intentar leer de forma más básica
                    df = pd.read_csv(
                        csv_file, encoding="utf-8", on_bad_lines="skip", sep=","
                    )

                # Convertir a diccionarios y agregar nombre del archivo fuente
                file_data = df.to_dict("records")

                # Limpiar valores NaN y convertir a tipos JSON válidos
                cleaned_data = []
                for record in file_data:
                    cleaned_record = {}
                    for key, value in record.items():
                        # Convertir NaN y valores problemáticos a None (null en JSON)
                        if pd.isna(value) or str(value).lower() == "nan":
                            cleaned_record[key] = None
                        elif isinstance(value, (int, float)) and (
                            value != value
                        ):  # Check for NaN
                            cleaned_record[key] = None
                        else:
                            # Convertir a string para asegurar compatibilidad JSON
                            cleaned_record[key] = (
                                str(value) if value is not None else None
                            )

                    cleaned_record["SourceFile"] = os.path.basename(csv_file)
                    cleaned_data.append(cleaned_record)

                print(
                    f" Datos procesados y limpiados: {len(cleaned_data)} registros del archivo {os.path.basename(csv_file)}"
                )
                all_data.extend(cleaned_data)

            except Exception as file_error:
                print(f" Error definitivo leyendo {csv_file}: {str(file_error)}")
                print(f" Tipo de error: {type(file_error).__name__}")
                import traceback

                print(f" Traceback: {traceback.format_exc()}")
                continue

        if not all_data:
            return jsonify(
                {
                    "success": False,
                    "error": "No se pudieron leer datos de los archivos CSV",
                    "files_found": len(csv_files),
                }
            ), 500

        print(
            f" Datos cargados: {len(all_data)} registros de {len(csv_files)} archivos"
        )

        return jsonify(
            {
                "success": True,
                "data": all_data,
                "message": f"Datos cargados para {folder}: {len(all_data)} registros",
                "files_processed": len(csv_files),
                "path": folder_path,
            }
        )

    except Exception as e:
        print(f" Error obteniendo datos CSV: {e}")
        print(f" Traceback: {traceback.format_exc()}")
        return jsonify(
            {
                "success": False,
                "error": f"Error al acceder a los archivos CSV: {str(e)}",
            }
        ), 500


@app.route("/api/filter_data", methods=["POST"])
@login_requerido
def filter_csv_data():
    """API para filtrar datos SMT desde MySQL (no archivos CSV)"""
    try:
        filters = request.get_json()
        folder = filters.get("folder", "")
        part_name = filters.get("partName", "")
        result = filters.get("result", "")
        date_from = filters.get("dateFrom", "")
        date_to = filters.get("dateTo", "")

        if not folder:
            return jsonify(
                {"success": False, "error": "Folder parameter required"}
            ), 400

        print(f" Filtrando datos MySQL para carpeta: {folder}")
        print(
            f" Filtros: partName={part_name}, result={result}, dateFrom={date_from}, dateTo={date_to}"
        )

        # Conectar a MySQL directamente
        import mysql.connector

        mysql_config = {
            "host": os.getenv("MYSQL_HOST"),
            "port": int(os.getenv("MYSQL_PORT", 3306)),
            "user": os.getenv("MYSQL_USER"),
            "password": os.getenv("MYSQL_PASSWORD"),
            "database": os.getenv("MYSQL_DATABASE"),
            "charset": "utf8mb4",
        }

        conn = mysql.connector.connect(**mysql_config)
        cursor = conn.cursor(dictionary=True)

        # Construir query dinámicamente con filtros
        where_conditions = ["source_file LIKE %s"]
        params = [f"%{folder}%"]

        if part_name:
            where_conditions.append("part_name LIKE %s")
            params.append(f"%{part_name}%")

        if result:
            where_conditions.append("result = %s")
            params.append(result.upper())

        if date_from:
            where_conditions.append("ScanDate >= %s")
            params.append(date_from.replace("-", ""))  # Convertir YYYY-MM-DD a YYYYMMDD

        if date_to:
            where_conditions.append("ScanDate <= %s")
            params.append(date_to.replace("-", ""))  # Convertir YYYY-MM-DD a YYYYMMDD

        where_clause = " AND ".join(where_conditions)

        query = f"""
            SELECT
            scan_date,
            scan_time,
            slot_no,
            result,
            previous_barcode,
            product_date,
            part_name,
            quantity,
            seq,
            vendor,
            lot_no,
            barcode,
            feeder_base,
            source_file,
            created_at
        FROM historial_cambio_material_smt
        WHERE {where_clause}
        ORDER BY ScanDate DESC, ScanTime DESC
            LIMIT 5000
        """

        cursor.execute(query, params)
        resultados = cursor.fetchall()

        print(f" Encontrados {len(resultados)} registros con filtros aplicados")

        # Convertir datos para JSON y mapear nombres para compatibilidad
        filtered_data = []
        for resultado in resultados:
            cleaned_record = {}
            for key, value in resultado.items():
                if hasattr(value, "isoformat"):  # Es una fecha/datetime
                    cleaned_record[key] = value.isoformat()
                elif value is None:
                    cleaned_record[key] = None
                else:
                    cleaned_record[key] = str(value)

            # Mapear nombres para compatibilidad con frontend
            cleaned_record["ScanDate"] = cleaned_record.get("ScanDate", "")
            cleaned_record["ScanTime"] = cleaned_record.get("scan_time", "")
            cleaned_record["SlotNo"] = cleaned_record.get("slot_no", "")
            cleaned_record["Result"] = cleaned_record.get("result", "")
            cleaned_record["PartName"] = cleaned_record.get("part_name", "")
            cleaned_record["SourceFile"] = cleaned_record.get("source_file", "")
            cleaned_record["LOTNO"] = cleaned_record.get("lot_no", "")
            cleaned_record["Barcode"] = cleaned_record.get("barcode", "")
            cleaned_record["Quantity"] = cleaned_record.get("quantity", "")
            cleaned_record["Vendor"] = cleaned_record.get("vendor", "")
            cleaned_record["FeederBase"] = cleaned_record.get("feeder_base", "")
            cleaned_record["PreviousBarcode"] = cleaned_record.get(
                "previous_barcode", ""
            )

            filtered_data.append(cleaned_record)

        # Calcular estadísticas de los datos filtrados
        stats = {
            "total_records": len(filtered_data),
            "ok_count": len(
                [d for d in filtered_data if str(d.get("Result", "")).upper() == "OK"]
            ),
            "ng_count": len(
                [d for d in filtered_data if str(d.get("Result", "")).upper() == "NG"]
            ),
        }

        cursor.close()
        conn.close()

        print(f" Datos filtrados desde MySQL: {len(filtered_data)} registros")

        return jsonify(
            {
                "success": True,
                "data": filtered_data,
                "stats": stats,
                "source": "mysql",
                "message": f"Datos filtrados desde MySQL para {folder}",
            }
        )

    except Exception as e:
        print(f" Error filtrando datos desde MySQL: {e}")
        print(f" Traceback: {traceback.format_exc()}")
        return jsonify(
            {"success": False, "error": f"Error al filtrar datos MySQL: {str(e)}"}
        ), 500


def crear_patron_caracteres(
    texto_original, part_start, part_length, lot_start, lot_length
):
    """
    Crea un patrón de caracteres donde:
    - Caracteres específicos se mantienen como están
    - Números se marcan como 'N'
    - Letras se marcan como 'A'
    - Las zonas de número de parte y lote se marcan como 'X' (cualquier carácter)
    """
    patron = list(texto_original)

    # Marcar la zona del número de parte como 'X' (cualquier carácter)
    for i in range(part_start, part_start + part_length):
        if i < len(patron):
            patron[i] = "X"

    # Marcar la zona del lote como 'X' solo si existe lote
    if lot_start != -1 and lot_length > 0:
        for i in range(lot_start, lot_start + lot_length):
            if i < len(patron):
                patron[i] = "X"

    # Para el resto de caracteres, determinar el tipo
    for i, char in enumerate(patron):
        if char != "X":  # Si no es una zona variable
            if char.isdigit():
                patron[i] = "N"  # Número específico
            elif char.isalpha():
                patron[i] = "A"  # Letra específica
            # Los caracteres especiales y espacios se mantienen como están

    return "".join(patron)


def cargar_configuracion_usuario(usuario):
    """Cargar configuración específica del usuario"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT configuracion FROM usuarios_sistema
            WHERE username = %s
        """,
            (usuario,),
        )

        result = cursor.fetchone()
        cursor.close()
        conn.close()

        if result and result[0]:  # Usar índice en lugar de clave de diccionario
            import json

            return json.loads(result[0])
        else:
            return {}

    except Exception as e:
        print(f"Error cargando configuración del usuario {usuario}: {e}")
        return {}


def guardar_configuracion_usuario(usuario, config, valor=None):
    """Guardar configuración específica del usuario.

    Acepta:
    - guardar_configuracion_usuario(usuario, config_dict)
    - guardar_configuracion_usuario(usuario, clave, valor)
    """
    try:
        import json

        if valor is not None:
            config_actual = cargar_configuracion_usuario(usuario) or {}
            config_actual[str(config)] = valor
            config = config_actual
        elif not isinstance(config, dict):
            raise ValueError("config debe ser un diccionario o una clave con valor")

        conn = get_db_connection()
        cursor = conn.cursor()

        config_json = json.dumps(config)

        cursor.execute(
            """
            UPDATE usuarios_sistema
            SET configuracion = %s
            WHERE username = %s
        """,
            (config_json, usuario),
        )

        success = cursor.rowcount > 0
        conn.commit()
        cursor.close()
        conn.close()

        return success

    except Exception as e:
        print(f"Error guardando configuración del usuario {usuario}: {e}")
        return False

@app.route("/importar_excel_plan_produccion", methods=["POST"])
@login_requerido
def importar_excel_plan_produccion():
    """Importar plan de producción desde Excel"""
    conn = None
    cursor = None
    temp_path = None

    try:
        if "file" not in request.files:
            return jsonify(
                {"success": False, "error": "No se proporcionó archivo"}
            ), 400

        file = request.files["file"]
        if file.filename == "":
            return jsonify({"success": False, "error": "No se seleccionó archivo"}), 400

        if (
            not file
            or not file.filename
            or not file.filename.lower().endswith((".xlsx", ".xls"))
        ):
            return jsonify(
                {
                    "success": False,
                    "error": "Formato de archivo no válido. Use .xlsx o .xls",
                }
            ), 400

        # Guardar el archivo temporalmente
        filename = secure_filename(file.filename)
        temp_path = os.path.join(os.path.dirname(__file__), "temp_" + filename)
        file.save(temp_path)

        # Leer el archivo Excel
        try:
            # Intentar leer con encabezados primero
            df = pd.read_excel(
                temp_path, engine="openpyxl" if filename.endswith(".xlsx") else "xlrd"
            )

            # Si las primeras filas contienen datos directamente (sin encabezados claros)
            # y las columnas tienen nombres genéricos como 0, 1, 2, etc.
            if all(isinstance(col, int) for col in df.columns):
                # Leer sin encabezados y asignar nombres de columnas
                df = pd.read_excel(
                    temp_path,
                    header=None,
                    engine="openpyxl" if filename.endswith(".xlsx") else "xlrd",
                )
                # Asignar nombres basados en la posición
                if len(df.columns) >= 3:
                    df.columns = ["Modelo", "Numero_Parte", "Cantidad"] + [
                        f"Col_{i}" for i in range(3, len(df.columns))
                    ]
                elif len(df.columns) == 2:
                    df.columns = ["Modelo", "Cantidad"]
                else:
                    df.columns = ["Modelo"]
        except Exception as e:
            try:
                # Intentar leer sin encabezados como respaldo
                df = pd.read_excel(temp_path, header=None)
                if len(df.columns) >= 3:
                    df.columns = ["Modelo", "Numero_Parte", "Cantidad"] + [
                        f"Col_{i}" for i in range(3, len(df.columns))
                    ]
                elif len(df.columns) == 2:
                    df.columns = ["Modelo", "Cantidad"]
                else:
                    df.columns = ["Modelo"]
            except Exception as e2:
                return jsonify(
                    {
                        "success": False,
                        "error": f"Error al leer el archivo Excel: {str(e2)}",
                    }
                ), 500

        # Verificar que el DataFrame no esté vacío
        if df.empty:
            return jsonify(
                {"success": False, "error": "El archivo Excel está vacío"}
            ), 400

        # Obtener usuario de la sesión
        usuario_actual = session.get("usuario", "USUARIO_EXCEL")

        # Obtener fecha de operación seleccionada por el usuario
        fecha_operacion_usuario = request.form.get("fecha_operacion", "").strip()
        if fecha_operacion_usuario:
            print(
                f" Fecha de operación personalizada seleccionada: {fecha_operacion_usuario}"
            )
        else:
            print(" Usando fechas del Excel o fecha actual como respaldo")

        # Función auxiliar para obtener nombre del modelo desde raw
        def obtener_nombre_modelo(codigo_modelo):
            """Obtener nombre (project) desde raw por part_no"""
            try:
                if not codigo_modelo:
                    return ""
                cursor.execute(
                    "SELECT project FROM raw WHERE TRIM(part_no)=TRIM(%s) ORDER BY id DESC LIMIT 1",
                    (codigo_modelo,),
                )
                row = cursor.fetchone()
                return (row.get("project") if row else "") or ""
            except Exception as e:
                print(f"Error obteniendo nombre modelo para {codigo_modelo}: {e}")
                return ""

        # Conectar a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor()

        # Asegurar que la tabla work_orders tenga la columna linea
        try:
            cursor.execute("SHOW COLUMNS FROM work_orders LIKE 'linea'")
            if not cursor.fetchone():
                cursor.execute("ALTER TABLE work_orders ADD COLUMN linea VARCHAR(32)")
                print(" Columna 'linea' agregada a work_orders")
        except Exception as e:
            print(f"Error agregando columna linea: {e}")

        registros_insertados = 0
        registros_actualizados = 0
        errores = []

        # Mapeo de columnas (flexible para diferentes nombres)
        mapeo_columnas = {
            "linea": ["Linea", "linea", "Line", "LINEA", "Línea"],
            "modelo": ["Modelo", "modelo", "Model", "MODELO"],
            "numero_parte": [
                "Numero de parte",
                "Número de parte",
                "numero_parte",
                "Part Number",
                "NUMERO_PARTE",
                "Numero_Parte",
            ],
            "cantidad": ["Cantidad", "cantidad", "Quantity", "CANTIDAD"],
            "fecha_operacion": [
                "Fecha",
                "fecha_operacion",
                "Fecha de operación",
                "Date",
                "FECHA",
            ],
            "codigo_po": [
                "PO",
                "codigo_po",
                "Código PO",
                "Purchase Order",
                "CODIGO_PO",
            ],
        }

        # Debug: Mostrar información del DataFrame
        print(f"Columnas en el DataFrame: {list(df.columns)}")
        print(f"Primeras 3 filas del DataFrame:")
        print(df.head(3))

        # Detectar columnas disponibles
        columnas_detectadas = {}
        for campo, posibles_nombres in mapeo_columnas.items():
            for nombre in posibles_nombres:
                if nombre in df.columns:
                    columnas_detectadas[campo] = nombre
                    break

        print(f"Columnas detectadas: {columnas_detectadas}")

        # Verificar que al menos tengamos modelo y cantidad
        if "modelo" not in columnas_detectadas or "cantidad" not in columnas_detectadas:
            # Información detallada para debugging
            error_msg = (
                f"El archivo debe contener al menos las columnas: Modelo y Cantidad. "
            )
            error_msg += f"Columnas encontradas: {list(df.columns)}. "
            error_msg += f"Mapeo detectado: {columnas_detectadas}"

            return jsonify({"success": False, "error": error_msg}), 400

        # Procesar cada fila del DataFrame
        for index, row in df.iterrows():
            try:
                # Extraer datos de la fila
                modelo = str(row.get(columnas_detectadas["modelo"], "")).strip()
                cantidad = row.get(columnas_detectadas["cantidad"], 0)

                # Validar datos básicos
                if not modelo or modelo == "nan":
                    errores.append(f"Fila {index + 2}: Modelo vacío")
                    continue

                try:
                    cantidad = (
                        int(float(cantidad))
                        if cantidad and str(cantidad) != "nan"
                        else 0
                    )
                except (ValueError, TypeError):
                    cantidad = 0

                if cantidad <= 0:
                    errores.append(f"Fila {index + 2}: Cantidad inválida ({cantidad})")
                    continue

                # Datos opcionales
                linea = str(row.get(columnas_detectadas.get("linea", ""), "")).strip()
                if linea == "nan":
                    linea = ""

                numero_parte = str(
                    row.get(columnas_detectadas.get("numero_parte", ""), "")
                ).strip()
                if numero_parte == "nan":
                    numero_parte = ""

                codigo_po = str(
                    row.get(columnas_detectadas.get("codigo_po", ""), "SIN-PO")
                ).strip()
                if codigo_po == "nan" or not codigo_po:
                    codigo_po = "SIN-PO"

                # Fecha de operación - priorizar la fecha seleccionada por el usuario
                fecha_operacion_usuario = request.form.get(
                    "fecha_operacion", ""
                ).strip()

                if fecha_operacion_usuario:
                    # Usar la fecha seleccionada por el usuario
                    fecha_operacion = fecha_operacion_usuario
                else:
                    # Usar la fecha del Excel o fecha actual como respaldo
                    fecha_operacion = row.get(
                        columnas_detectadas.get("fecha_operacion", ""), ""
                    )
                    if pd.isna(fecha_operacion) or fecha_operacion == "nan":
                        fecha_operacion = datetime.now().strftime("%Y-%m-%d")
                    else:
                        try:
                            if isinstance(fecha_operacion, str):
                                # Intentar convertir string a fecha
                                from dateutil import parser

                                fecha_operacion = parser.parse(
                                    fecha_operacion
                                ).strftime("%Y-%m-%d")
                            else:
                                # Es datetime o similar
                                fecha_operacion = fecha_operacion.strftime("%Y-%m-%d")
                        except:
                            fecha_operacion = datetime.now().strftime("%Y-%m-%d")

                # Generar código WO único
                fecha_codigo = datetime.now().strftime("%y%m%d")

                # Buscar el último número de secuencia para hoy
                cursor.execute(
                    """
                    SELECT codigo_wo FROM work_orders
                    WHERE codigo_wo LIKE %s
                    ORDER BY codigo_wo DESC LIMIT 1
                """,
                    (f"WO-{fecha_codigo}-%",),
                )

                ultimo_wo = cursor.fetchone()
                if ultimo_wo:
                    try:
                        ultimo_numero = int(ultimo_wo["codigo_wo"].split("-")[-1])
                        nuevo_numero = ultimo_numero + 1
                    except:
                        nuevo_numero = 1
                else:
                    nuevo_numero = 1

                codigo_wo = f"WO-{fecha_codigo}-{nuevo_numero:04d}"

                # Obtener nombre del modelo desde la tabla raw
                codigo_modelo = modelo
                nombre_modelo = obtener_nombre_modelo(codigo_modelo)

                # Insertar nueva WO
                cursor.execute(
                    """
                    INSERT INTO work_orders
                    (codigo_wo, codigo_po, modelo, codigo_modelo, nombre_modelo, linea,
                     cantidad_planeada, fecha_operacion, usuario_creacion, modificador, estado)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'CREADA')
                """,
                    (
                        codigo_wo,
                        codigo_po,
                        modelo
                        if modelo
                        else numero_parte,  # Usar modelo o numero_parte como respaldo
                        codigo_modelo,
                        nombre_modelo,
                        linea,
                        cantidad,
                        fecha_operacion,
                        usuario_actual,
                        usuario_actual,
                    ),
                )

                registros_insertados += 1

            except Exception as e:
                errores.append(f"Fila {index + 2}: {str(e)}")
                continue

        # Confirmar transacción
        conn.commit()

        # Preparar respuesta
        mensaje = f"Importación completada. {registros_insertados} WOs creadas."
        if fecha_operacion_usuario:
            mensaje += f" Fecha de operación aplicada: {fecha_operacion_usuario}."
        if errores:
            mensaje += f" {len(errores)} errores encontrados."

        return jsonify(
            {
                "success": True,
                "message": mensaje,
                "registros_procesados": registros_insertados,
                "errores": len(errores),
                "fecha_aplicada": fecha_operacion_usuario or "Fechas del Excel/Actual",
                "detalles": {
                    "insertados": registros_insertados,
                    "errores": errores[:10]
                    if errores
                    else [],  # Solo primeros 10 errores
                },
            }
        )

    except Exception as e:
        print(f"Error general en importar_excel_plan_produccion: {str(e)}")
        return jsonify({"success": False, "error": f"Error interno: {str(e)}"}), 500

    finally:
        try:
            if cursor:
                cursor.close()
            if conn:
                conn.close()
            if temp_path and os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception as e:
            print(f"Error en cleanup: {str(e)}")


@app.route("/produccion/info")
@login_requerido
def produccion_info():
    try:
        return render_template("CONTROL DE PRODUCCION/info_produccion.html")
    except Exception as e:
        return f"Error al cargar información de producción: {str(e)}", 500


# Migracion 2026-05-27: GET /api/wo/exportar movido a
# app/api/control_produccion/po_wo.py (misma URL, mismo comportamiento).



@app.route("/api/inventario", methods=["GET"])
@login_requerido
def api_inventario():
    """API para consultar inventario por modelo y/o nparte"""
    try:
        modelo = request.args.get("modelo", "").strip()
        nparte = request.args.get("nparte", "").strip()

        if not modelo:
            return jsonify({"error": "Parámetro modelo es requerido"}), 400

        if nparte:
            # Consultar inventario específico por modelo y nparte
            query = """
            SELECT modelo, nparte, stock_total, ubicaciones, ultima_entrada, ultima_salida, updated_at
            FROM inv_resumen_modelo
            WHERE modelo = %s AND nparte = %s
            """
            result = execute_query(query, (modelo, nparte), fetch="one")

            if result:
                return jsonify(
                    {
                        "modelo": result["modelo"],
                        "nparte": result["nparte"],
                        "stock_total": result["stock_total"] or 0,
                    }
                )
            else:
                return jsonify({"modelo": modelo, "nparte": nparte, "stock_total": 0})
        else:
            # Consultar inventario total del modelo
            query = """
            SELECT modelo, SUM(stock_total) as stock_total
            FROM inv_resumen_modelo
            WHERE modelo = %s
            GROUP BY modelo
            """
            result = execute_query(query, (modelo,), fetch="one")

            if result:
                return jsonify(
                    {
                        "modelo": result["modelo"],
                        "stock_total": result["stock_total"] or 0,
                    }
                )
            else:
                return jsonify({"modelo": modelo, "stock_total": 0})

    except Exception as e:
        print(f" Error consultando inventario: {e}")
        return jsonify({"error": str(e)}), 500


# ============================================================================
# RUTAS PARA CONTROL DE CALIDAD
# ============================================================================


@app.route("/control-resultado-reparacion-ajax")
@login_requerido
def control_resultado_reparacion_ajax():
    """Template para Control de resultado de reparación"""
    return render_template("Control de calidad/control_resultado_reparacion_ajax.html")


@app.route("/control-item-reparado-ajax")
@login_requerido
def control_item_reparado_ajax():
    """Template para Control de item reparado"""
    return render_template("Control de calidad/control_item_reparado_ajax.html")


# Eliminado 2026-05-27: "Historial de cambio de material por maquina" dado de baja
# (modulo no usado). Rutas borradas:
#   /historial-cambio-material-maquina-ajax  (render)
#   /api/historial-cambio-material-maquina   (GET)


# ==========================
# Historial SMT (ultimo por linea/maquina/slot)
# ==========================
@app.route("/api/historial_smt_latest", methods=["GET"])
@login_requerido
def api_historial_smt_latest():
    """Devuelve el ultimo escaneo por (linea, maquina, SlotNo) desde la tabla
    historial_cambio_material_smt. Pensado para el panel de Control de Operacion SMT
    que requiere el ultimo material escaneado para hacer match con el BOM.

    Parametros:
      - linea: opcional. Ej: 'SMT B'. Si se omite, devuelve para todas las lineas.
    """
    try:
        linea = request.args.get("linea", "").strip()

        from .db_mysql import get_connection

        conn = get_connection()
        cursor = conn.cursor()

        where_sub = ""
        params = []
        if linea:
            where_sub = "WHERE linea = %s"
            params.append(linea)

        # Seleccionar el ultimo registro por grupo usando fecha_subida
        query = f"""
            SELECT h.id, h.linea, h.maquina, h.archivo, h.ScanDate, h.ScanTime,
                   h.SlotNo, h.Result, h.PreviousBarcode, h.Productdate,
                   h.PartName, h.Quantity, h.SEQ, h.Vendor, h.LOTNO,
                   h.Barcode, h.FeederBase, h.fecha_subida,
                   CASE WHEN UPPER(h.FeederBase) LIKE '%%F%%' THEN 'FRONT'
                        WHEN UPPER(h.FeederBase) LIKE '%%R%%' THEN 'REAR'
                        ELSE 'UNKNOWN' END AS side_norm
            FROM historial_cambio_material_smt h
            INNER JOIN (
                SELECT linea, maquina, SlotNo,
                       CASE WHEN UPPER(FeederBase) LIKE '%%F%%' THEN 'FRONT'
                            WHEN UPPER(FeederBase) LIKE '%%R%%' THEN 'REAR'
                            ELSE 'UNKNOWN' END AS side_norm,
                       MAX(fecha_subida) AS max_fecha
                FROM historial_cambio_material_smt
                {where_sub}
                GROUP BY linea, maquina, SlotNo, side_norm
            ) m
            ON h.linea = m.linea AND h.maquina = m.maquina
               AND h.SlotNo = m.SlotNo AND h.fecha_subida = m.max_fecha
               AND (
                    (CASE WHEN UPPER(h.FeederBase) LIKE '%%F%%' THEN 'FRONT'
                          WHEN UPPER(h.FeederBase) LIKE '%%R%%' THEN 'REAR'
                          ELSE 'UNKNOWN' END) = m.side_norm
               )
            {("WHERE h.linea = %s" if linea else "")}
            ORDER BY h.linea, h.maquina, h.SlotNo, side_norm
        """

        if linea:
            # Parametros para subconsulta y para el filtro externo
            cursor.execute(query, params + params)
        else:
            cursor.execute(query)

        rows = cursor.fetchall()

        data = []
        for r in rows:
            # Indices alineados al SELECT de arriba
            linea_v = r[1] if len(r) > 1 else ""
            maquina_v = r[2] if len(r) > 2 else ""
            scan_date = r[4] if len(r) > 4 else ""
            scan_time = r[5] if len(r) > 5 else ""
            slot_no = r[6] if len(r) > 6 else ""
            part_name = r[10] if len(r) > 10 else ""
            quantity = r[11] if len(r) > 11 else 0
            vendor = r[13] if len(r) > 13 else ""
            feeder_base = r[16] if len(r) > 16 else ""

            # Normalizaciones amigables para el frontend existente
            formatted = {
                "linea": linea_v,
                "maquina": maquina_v,  # usado para extraer mounter (mN)
                "Equipment": maquina_v,  # alias por compatibilidad
                "SlotNo": slot_no,
                "FeederBase": feeder_base,
                "RegistDate": scan_date,  # el frontend ya acepta varios nombres
                "fecha_formateada": scan_date,
                "PartName": part_name,  # se usa para matching contra BOM Material Code
                "Quantity": quantity,
                "Vendor": vendor,
                "ScanDate": scan_date,
                "ScanTime": scan_time,
            }
            data.append(formatted)

        cursor.close()
        conn.close()

        return jsonify({"success": True, "data": data, "total": len(data)})
    except Exception as e:
        print(f"Error en api_historial_smt_latest: {e}")
        print(traceback.format_exc())
        return jsonify({"success": False, "error": str(e)}), 500


# Variante robusta con lado FRONT/REAR agrupado explicitamente
@app.route("/api/historial_smt_latest_v2", methods=["GET"])
@login_requerido
def api_historial_smt_latest_v2():
    try:
        linea_input = request.args.get("linea", "").strip()
        # Convertir nombre de línea a formato de BD
        linea = convertir_linea_smt(linea_input)

        from .db_mysql import get_connection

        conn = get_connection()
        cursor = conn.cursor()

        where_sub = ""
        params = []
        if linea:
            where_sub = "WHERE linea = %s"
            params.append(linea)

        query = f"""
            SELECT h.id, h.linea, h.maquina, h.archivo, h.ScanDate, h.ScanTime,
                   h.SlotNo, h.Result, h.PreviousBarcode, h.Productdate,
                   h.PartName, h.Quantity, h.SEQ, h.Vendor, h.LOTNO,
                   h.Barcode, h.FeederBase, h.fecha_subida,
                   CASE WHEN UPPER(h.FeederBase) LIKE '%%F%%' THEN 'FRONT'
                        WHEN UPPER(h.FeederBase) LIKE '%%R%%' THEN 'REAR'
                        ELSE 'UNKNOWN' END AS side_norm
            FROM historial_cambio_material_smt h
            INNER JOIN (
                SELECT linea, maquina, SlotNo,
                       (CASE WHEN UPPER(FeederBase) LIKE '%%F%%' THEN 'FRONT'
                             WHEN UPPER(FeederBase) LIKE '%%R%%' THEN 'REAR'
                             ELSE 'UNKNOWN' END) AS side_norm,
                       MAX(fecha_subida) AS max_fecha
                FROM historial_cambio_material_smt
                {where_sub}
                GROUP BY linea, maquina, SlotNo,
                         (CASE WHEN UPPER(FeederBase) LIKE '%%F%%' THEN 'FRONT'
                               WHEN UPPER(FeederBase) LIKE '%%R%%' THEN 'REAR'
                               ELSE 'UNKNOWN' END)
            ) m
              ON h.linea = m.linea AND h.maquina = m.maquina
             AND h.SlotNo = m.SlotNo AND h.fecha_subida = m.max_fecha
             AND (
                 (CASE WHEN UPPER(h.FeederBase) LIKE '%%F%%' THEN 'FRONT'
                       WHEN UPPER(h.FeederBase) LIKE '%%R%%' THEN 'REAR'
                       ELSE 'UNKNOWN' END) = m.side_norm
             )
            {("WHERE h.linea = %s" if linea else "")}
            ORDER BY h.linea, h.maquina, h.SlotNo, m.side_norm
        """

        if linea:
            cursor.execute(query, params + params)
        else:
            cursor.execute(query)

        rows = cursor.fetchall()
        data = []
        for r in rows:
            linea_v = r[1] if len(r) > 1 else ""
            maquina_v = r[2] if len(r) > 2 else ""
            scan_date = r[4] if len(r) > 4 else ""
            scan_time = r[5] if len(r) > 5 else ""
            slot_no = r[6] if len(r) > 6 else ""
            part_name = r[10] if len(r) > 10 else ""
            quantity = r[11] if len(r) > 11 else 0
            vendor = r[13] if len(r) > 13 else ""
            feeder_base = r[16] if len(r) > 16 else ""

            formatted = {
                "linea": linea_v,
                "maquina": maquina_v,
                "Equipment": maquina_v,
                "SlotNo": slot_no,
                "FeederBase": feeder_base,
                "RegistDate": scan_date,
                "fecha_formateada": scan_date,
                "PartName": part_name,
                "Quantity": quantity,
                "Vendor": vendor,
                "ScanDate": scan_date,
                "ScanTime": scan_time,
            }
            data.append(formatted)

        cursor.close()
        conn.close()

        return jsonify({"success": True, "data": data, "total": len(data)})
    except Exception as e:
        print("Error en api_historial_smt_latest_v2:", e)
        print(traceback.format_exc())
        return jsonify({"success": False, "error": str(e)}), 500


# ==========================
# Metal Mask info lookup
# ==========================
@app.route("/api/masks/info", methods=["GET"])
@login_requerido
def api_masks_info():
    try:
        code = request.args.get("code", "").strip()
        if not code:
            return jsonify({"success": False, "error": "code requerido"}), 400

        from .db_mysql import get_connection

        conn = get_connection()
        cursor = conn.cursor()

        q = """
            SELECT management_no, storage_box, pcb_code, side, production_date,
                   used_count, max_count, allowance, model_name, tension_min,
                   tension_max, thickness, supplier, registration_date, disuse
            FROM masks
            WHERE management_no = %s
            LIMIT 1
        """
        cursor.execute(q, [code])
        row = cursor.fetchone()
        cursor.close()
        conn.close()

        if not row:
            return jsonify(
                {"success": False, "found": False, "message": "No encontrado"}
            )

        fields = [
            "management_no",
            "storage_box",
            "pcb_code",
            "side",
            "production_date",
            "used_count",
            "max_count",
            "allowance",
            "model_name",
            "tension_min",
            "tension_max",
            "thickness",
            "supplier",
            "registration_date",
            "disuse",
        ]
        data = {
            fields[i]: (row[i] if i < len(row) else None) for i in range(len(fields))
        }

        def to_int(v):
            try:
                return int(v)
            except Exception:
                try:
                    return int(float(v))
                except Exception:
                    return 0

        data["used_count"] = to_int(data.get("used_count"))
        data["max_count"] = to_int(data.get("max_count"))
        data["allowance"] = to_int(data.get("allowance"))

        return jsonify({"success": True, "found": True, "data": data})
    except Exception as e:
        print("Error en api_masks_info:", e)
        print(traceback.format_exc())
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/historial-uso-pegamento-soldadura-ajax")
@login_requerido
def historial_uso_pegamento_soldadura_ajax():
    """Template para Historial de uso de pegamento de soldadura"""
    return render_template(
        "Control de calidad/historial_uso_pegamento_soldadura_ajax.html"
    )


# ==========================
# API para historial de Metal Mask
# ==========================
@app.route("/api/metal-mask/history", methods=["POST"])
@login_requerido
def api_save_metal_mask_history():
    """Guardar historial de uso de Metal Mask"""
    try:
        data = request.get_json()

        # Validar datos requeridos
        required_fields = ["mask_code", "model_code", "linea", "quantity_used"]
        for field in required_fields:
            if not data.get(field):
                return jsonify(
                    {"success": False, "error": f"{field} es requerido"}
                ), 400

        from .db_mysql import get_connection

        conn = get_connection()
        cursor = conn.cursor()

        # Crear tabla si no existe
        create_table_query = """
            CREATE TABLE IF NOT EXISTS metal_mask_history (
                id INT AUTO_INCREMENT PRIMARY KEY,
                mask_code VARCHAR(50) NOT NULL,
                model_code VARCHAR(50) NOT NULL,
                linea VARCHAR(20) NOT NULL,
                quantity_used INT NOT NULL DEFAULT 0,
                scan_date DATETIME DEFAULT CURRENT_TIMESTAMP,
                usuario VARCHAR(50),
                plan_id INT,
                run_id INT,
                available_uses INT DEFAULT 0,
                total_uses INT DEFAULT 0,
                status ENUM('OK', 'NG', 'WARNING') DEFAULT 'OK',
                notes TEXT,
                INDEX idx_mask_code (mask_code),
                INDEX idx_model_code (model_code),
                INDEX idx_linea (linea),
                INDEX idx_scan_date (scan_date)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
        cursor.execute(create_table_query)
        conn.commit()

        # Insertar registro de historial
        insert_query = """
            INSERT INTO metal_mask_history
            (mask_code, model_code, linea, quantity_used, usuario, plan_id, run_id,
             available_uses, total_uses, status, notes)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        usuario = session.get("usuario_logueado", "Sistema")
        plan_id = data.get("plan_id")
        run_id = data.get("run_id")
        available_uses = data.get("available_uses", 0)
        total_uses = data.get("total_uses", 0)
        status = data.get("status", "OK")
        notes = data.get("notes", "")

        cursor.execute(
            insert_query,
            [
                data["mask_code"],
                data["model_code"],
                data["linea"],
                data["quantity_used"],
                usuario,
                plan_id,
                run_id,
                available_uses,
                total_uses,
                status,
                notes,
            ],
        )

        history_id = cursor.lastrowid
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(
            {
                "success": True,
                "history_id": history_id,
                "message": "Historial de Metal Mask guardado correctamente",
            }
        )

    except Exception as e:
        print("Error en api_save_metal_mask_history:", e)
        print(traceback.format_exc())
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/metal-mask/history", methods=["GET"])
@login_requerido
def api_get_metal_mask_history():
    """Obtener historial de uso de Metal Mask"""
    try:
        # Parámetros de filtro
        mask_code = request.args.get("mask_code", "").strip()
        model_code = request.args.get("model_code", "").strip()
        linea = request.args.get("linea", "").strip()
        date_from = request.args.get("date_from", "").strip()
        date_to = request.args.get("date_to", "").strip()
        limit = int(request.args.get("limit", 100))

        from .db_mysql import get_connection

        conn = get_connection()
        cursor = conn.cursor()

        # Construir consulta con filtros
        where_conditions = []
        params = []

        if mask_code:
            where_conditions.append("mask_code = %s")
            params.append(mask_code)

        if model_code:
            where_conditions.append("model_code = %s")
            params.append(model_code)

        if linea:
            where_conditions.append("linea = %s")
            params.append(linea)

        if date_from:
            where_conditions.append("scan_date >= %s")
            params.append(date_from)

        if date_to:
            where_conditions.append("scan_date <= %s")
            params.append(date_to + " 23:59:59")

        where_clause = (
            "WHERE " + " AND ".join(where_conditions) if where_conditions else ""
        )

        query = f"""
            SELECT id, mask_code, model_code, linea, quantity_used,
                   DATE_FORMAT(scan_date, '%%Y-%%m-%%d %%H:%%i:%%s') as scan_date,
                   usuario, plan_id, run_id, available_uses, total_uses,
                   status, notes
            FROM metal_mask_history
            {where_clause}
            ORDER BY scan_date DESC
            LIMIT %s
        """

        params.append(limit)
        cursor.execute(query, params)
        rows = cursor.fetchall()

        # Convertir a lista de diccionarios
        columns = [
            "id",
            "mask_code",
            "model_code",
            "linea",
            "quantity_used",
            "scan_date",
            "usuario",
            "plan_id",
            "run_id",
            "available_uses",
            "total_uses",
            "status",
            "notes",
        ]

        data = [dict(zip(columns, row)) for row in rows]

        cursor.close()
        conn.close()

        return jsonify({"success": True, "data": data, "count": len(data)})

    except Exception as e:
        print(" Error en api_get_metal_mask_history:", e)
        print(" Traceback completo:")
        import traceback

        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/metal-mask/update-used-count", methods=["POST"])
@login_requerido
def api_update_metal_mask_used_count():
    """Actualizar used_count de Metal Mask al finalizar plan"""
    try:
        data = request.get_json()
        plan_id = data.get("plan_id")
        cantidad_producida = int(data.get("cantidad_producida", 0))

        if not plan_id or cantidad_producida <= 0:
            return jsonify(
                {
                    "success": False,
                    "error": "plan_id y cantidad_producida son requeridos",
                }
            )

        from .db_mysql import get_connection

        conn = get_connection()
        cursor = conn.cursor()

        # 1. Obtener información del plan para saber el modelo y línea
        cursor.execute(
            """
            SELECT modelo, linea, nparte
            FROM plan_smd
            WHERE id = %s
        """,
            (plan_id,),
        )
        plan_info = cursor.fetchone()

        if not plan_info:
            return jsonify({"success": False, "error": "Plan no encontrado"})

        modelo, linea, nparte = plan_info

        # 2. Buscar Metal Masks que se usaron para este modelo/línea
        # Prioridad 1: Buscar por plan_id específico
        cursor.execute(
            """
            SELECT DISTINCT mask_code, COUNT(*) as usage_count
            FROM metal_mask_history
            WHERE plan_id = %s
            GROUP BY mask_code
            ORDER BY usage_count DESC, scan_date DESC
        """,
            (plan_id,),
        )

        mask_codes = [row[0] for row in cursor.fetchall()]

        # Prioridad 2: Si no hay historial del plan, buscar por modelo/línea reciente
        if not mask_codes:
            cursor.execute(
                """
                SELECT DISTINCT mask_code, COUNT(*) as usage_count
                FROM metal_mask_history
                WHERE model_code = %s AND linea = %s
                AND scan_date >= DATE_SUB(NOW(), INTERVAL 7 DAY)
                GROUP BY mask_code
                ORDER BY usage_count DESC, scan_date DESC
                LIMIT 3
            """,
                (modelo, linea),
            )
            mask_codes = [row[0] for row in cursor.fetchall()]

        # Prioridad 3: Buscar cualquier mask para el modelo (último recurso)
        if not mask_codes:
            cursor.execute(
                """
                SELECT DISTINCT mask_code
                FROM metal_mask_history
                WHERE model_code = %s
                ORDER BY scan_date DESC
                LIMIT 1
            """,
                (modelo,),
            )
            mask_codes = [row[0] for row in cursor.fetchall()]

        updated_count = 0

        # 3. Actualizar used_count en la tabla masks para cada mask_code encontrada
        for mask_code in mask_codes:
            cursor.execute(
                """
                UPDATE masks
                SET used_count = used_count + %s
                WHERE management_no = %s
            """,
                (cantidad_producida, mask_code),
            )

            if cursor.rowcount > 0:
                updated_count += cursor.rowcount
                print(
                    f" Metal Mask {mask_code} - used_count incrementado en {cantidad_producida}"
                )

        # 4. Registrar el update en el historial para cada mask actualizada
        for mask_code in mask_codes:
            # Obtener información actualizada de la mask
            cursor.execute(
                """
                SELECT used_count, max_count, allowance
                FROM masks
                WHERE management_no = %s
            """,
                (mask_code,),
            )

            mask_info = cursor.fetchone()
            if mask_info:
                used_count, max_count, allowance = mask_info
                available_uses = max(0, (max_count + allowance) - used_count)

                cursor.execute(
                    """
                    INSERT INTO metal_mask_history
                    (mask_code, model_code, linea, quantity_used, plan_id,
                     available_uses, total_uses, status, notes, usuario, scan_date)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                """,
                    (
                        mask_code,
                        modelo,
                        linea,
                        cantidad_producida,
                        plan_id,
                        available_uses,
                        used_count,
                        "END_PLAN",
                        f"Finalización de plan {plan_id} - Producido: {cantidad_producida}",
                        session.get("usuario", "sistema"),
                    ),
                )

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(
            {
                "success": True,
                "updated_masks": updated_count,
                "cantidad_producida": cantidad_producida,
                "plan_id": plan_id,
                "masks_actualizadas": mask_codes,
            }
        )

    except Exception as e:
        print(" Error actualizando used_count:", e)
        import traceback

        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/historial-uso-mask-metal-ajax")
@login_requerido
def historial_uso_mask_metal_ajax():
    """Template para Historial de uso de mask de metal"""
    return render_template("Control de calidad/historial_uso_mask_metal_ajax.html")


@app.route("/historial-uso-squeegee-ajax")
@login_requerido
def historial_uso_squeegee_ajax():
    """Template para Historial de uso de squeegee"""
    return render_template("Control de calidad/historial_uso_squeegee_ajax.html")


@app.route("/process-interlock-history-ajax")
@login_requerido
def process_interlock_history_ajax():
    """Template para Process interlock History"""
    return render_template("Control de calidad/process_interlock_history_ajax.html")


@app.route("/control-master-sample-smt-ajax")
@login_requerido
def control_master_sample_smt_ajax():
    """Template para Control de Master Sample de SMT"""
    return render_template("Control de calidad/control_master_sample_smt_ajax.html")


@app.route("/historial-inspeccion-master-sample-smt-ajax")
@login_requerido
def historial_inspeccion_master_sample_smt_ajax():
    """Template para Historial de inspección de Master Sample de SMT"""
    return render_template(
        "Control de calidad/historial_inspeccion_master_sample_smt_ajax.html"
    )


@app.route("/control-inspeccion-oqc-ajax")
@login_requerido
def control_inspeccion_oqc_ajax():
    """Template para Control de inspección de OQC"""
    return render_template("Control de calidad/control_inspeccion_oqc_ajax.html")


@app.route("/historial-liberacion-lqc-ajax")
@login_requerido
def historial_liberacion_lqc_ajax():
    """Template para Historial de liberacion LQC"""
    return render_template("Control de calidad/historial_liberacion_lqc_ajax.html")


def _lqc_fecha_operativa(now=None):
    from datetime import datetime, timedelta
    now = now or datetime.now()
    corte = now.replace(hour=7, minute=30, second=0, microsecond=0)
    if now >= corte:
        return now.date()
    return (now - timedelta(days=1)).date()


def _lqc_parse_fecha(value, fallback):
    from datetime import datetime
    try:
        return datetime.strptime(str(value or ""), "%Y-%m-%d").date()
    except Exception:
        return fallback


def _lqc_ventana_operativa(fecha_inicio, fecha_fin):
    from datetime import datetime, time, timedelta
    inicio = datetime.combine(fecha_inicio, time(7, 30))
    fin = datetime.combine(fecha_fin + timedelta(days=1), time(7, 30))
    return inicio, fin


@app.route("/api/smt-scanner/lineas")
@login_requerido
def api_smt_scanner_lineas():
    """Resumen de escaneos LQC por linea real usando box_scans + plan_main."""
    try:
        from datetime import datetime, timedelta
        ahora_dt = datetime.now()
        fecha_operativa = _lqc_fecha_operativa(ahora_dt)
        ventana_inicio, ventana_fin = _lqc_ventana_operativa(fecha_operativa, fecha_operativa)
        ventana_fin_consulta = min(ahora_dt, ventana_fin)
        hoy = fecha_operativa.isoformat()
        ahora = ahora_dt.strftime('%Y-%m-%d %H:%M:%S')
        inicio_sql = ventana_inicio.strftime('%Y-%m-%d %H:%M:%S')
        fin_sql = ventana_fin_consulta.strftime('%Y-%m-%d %H:%M:%S')
        lineas_config = ['M1','M2','M3','M4','DP1','DP2','DP3','H1']
        line_expr = "COALESCE(NULLIF(p.line, ''), 'SIN PLAN')"
        slot_expr = """
                    DATE_ADD(
                        STR_TO_DATE(
                            DATE_FORMAT(DATE_SUB(b.last_scan, INTERVAL 30 MINUTE), '%%Y-%%m-%%d %%H:00:00'),
                            '%%Y-%%m-%%d %%H:%%i:%%s'
                        ),
                        INTERVAL 30 MINUTE
                    )
        """

        conn = get_pooled_connection()
        if conn is None:
            return jsonify({"success": True, "lineas": [], "uph_hoy": {}})

        cursor = get_dict_cursor(conn)
        try:
            # Total por linea de la fecha operativa actual; corte 07:30.
            cursor.execute(f"""
                SELECT
                    {line_expr} AS linea,
                    COUNT(*) AS total_hoy
                FROM box_scans b
                LEFT JOIN plan_main p ON p.lot_no = b.lot_no
                WHERE b.last_scan >= %s
                  AND b.last_scan < %s
                GROUP BY linea
            """, (inicio_sql, fin_sql))
            rows = cursor.fetchall()
            counts = {r['linea']: int(r['total_hoy'] or 0) for r in rows}

            # UPH por bloque real de hora; evita mezclar horas iguales de dias distintos.
            cursor.execute(f"""
                SELECT
                    {line_expr} AS linea,
                    DATE_FORMAT({slot_expr}, '%%Y-%%m-%%d %%H:%%i:%%s') AS slot_key,
                    HOUR({slot_expr}) AS hora,
                    COUNT(*) AS total
                FROM box_scans b
                LEFT JOIN plan_main p ON p.lot_no = b.lot_no
                WHERE b.last_scan >= %s
                  AND b.last_scan < %s
                GROUP BY linea, slot_key, hora
                ORDER BY slot_key
            """, (inicio_sql, fin_sql))
            uph_rows = cursor.fetchall()
            uph_hoy = {}
            uph_slots_map = {}
            for r in uph_rows:
                l = r['linea']
                slot_key = r['slot_key']
                if l not in uph_hoy:
                    uph_hoy[l] = {}
                uph_hoy[l][slot_key] = int(r['total'] or 0)
                if slot_key not in uph_slots_map:
                    uph_slots_map[slot_key] = {
                        "key": slot_key,
                        "label": "",
                        "hour": int(r['hora'] or 0),
                    }
            uph_slots = []
            if ventana_fin_consulta >= ventana_inicio:
                start_slot = ventana_inicio
                end_slot = ventana_fin
                current_slot = start_slot
                while current_slot < end_slot:
                    key = current_slot.strftime('%Y-%m-%d %H:%M:%S')
                    slot = uph_slots_map.get(key, {
                        "key": key,
                        "hour": current_slot.hour,
                    })
                    slot["label"] = current_slot.strftime('%H:%M')
                    uph_slots.append(slot)
                    current_slot += timedelta(hours=1)

            ordered_lines = lineas_config[:]
            ordered_lines.extend(
                sorted([line for line in counts.keys() if line not in ordered_lines])
            )
            lineas = [{"linea": l, "total_hoy": counts.get(l, 0)} for l in ordered_lines]
            return jsonify({
                "success": True,
                "fecha_consulta": hoy,
                "actualizado": ahora,
                "ventana_inicio": inicio_sql,
                "ventana_fin": fin_sql,
                "lineas": lineas,
                "uph_hoy": uph_hoy,
                "uph_slots": uph_slots,
            })
        finally:
            cursor.close()
            conn.close()
    except Exception as e:
        return jsonify({"success": False, "error": str(e), "lineas": [], "uph_hoy": {}}), 500


@app.route("/api/smt-scanner/datos")
@login_requerido
def api_smt_scanner_datos():
    """Registros LQC desde box_scans filtrados por linea real y rango de fechas."""
    try:
        from datetime import datetime
        fecha_default = _lqc_fecha_operativa()
        lineas = [l.strip() for l in request.args.getlist("lineas") if l.strip()]
        fecha_inicio = _lqc_parse_fecha(request.args.get("fecha_inicio"), fecha_default)
        fecha_fin = _lqc_parse_fecha(request.args.get("fecha_fin"), fecha_default)
        if fecha_fin < fecha_inicio:
            fecha_inicio, fecha_fin = fecha_fin, fecha_inicio
        ventana_inicio, ventana_fin = _lqc_ventana_operativa(fecha_inicio, fecha_fin)
        inicio_sql = ventana_inicio.strftime('%Y-%m-%d %H:%M:%S')
        fin_sql = ventana_fin.strftime('%Y-%m-%d %H:%M:%S')
        turno_filtro = request.args.get("turno", "Todos")
        line_expr = "COALESCE(NULLIF(p.line, ''), 'SIN PLAN')"
        fecha_operativa_expr = """
                    CASE
                        WHEN TIME(b.last_scan) >= '07:30:00' THEN DATE(b.last_scan)
                        ELSE DATE_SUB(DATE(b.last_scan), INTERVAL 1 DAY)
                    END
        """
        turno_expr = """
                    CASE
                        WHEN TIME(b.last_scan) >= '07:30:00' AND TIME(b.last_scan) < '17:30:00' THEN 'DIA'
                        WHEN TIME(b.last_scan) >= '17:30:00' AND TIME(b.last_scan) < '22:00:00' THEN 'TIEMPO EXTRA'
                        ELSE 'NOCHE'
                    END
        """

        conn = get_pooled_connection()
        if conn is None:
            return jsonify({"success": True, "records": []})

        cursor = get_dict_cursor(conn)
        try:
            where_conditions = ["b.last_scan >= %s", "b.last_scan < %s"]
            params = [inicio_sql, fin_sql]

            if lineas:
                placeholders = ','.join(['%s'] * len(lineas))
                where_conditions.append(f"{line_expr} IN ({placeholders})")
                params.extend(lineas)

            if turno_filtro != 'Todos':
                where_conditions.append(f"{turno_expr} = %s")
                params.append(turno_filtro)

            where_clause = " AND ".join(where_conditions)
            query = f"""
                SELECT
                    {line_expr} AS linea,
                    {fecha_operativa_expr} AS fecha,
                    COALESCE(NULLIF(p.part_no, ''), LEFT(b.serial, GREATEST(CHAR_LENGTH(b.serial) - 12, 1))) AS part,
                    COALESCE(NULLIF(p.model_code, ''), '') AS model_code,
                    COALESCE(b.lot_no, '') AS lot_no,
                    b.box_code,
                    b.id AS scan_id,
                    b.serial,
                    b.status AS box_status,
                    b.first_scan,
                    b.last_scan,
                    {turno_expr} AS turno
                FROM box_scans b
                LEFT JOIN plan_main p ON p.lot_no = b.lot_no
                WHERE {where_clause}
                ORDER BY b.last_scan DESC, b.id DESC
            """
            cursor.execute(query, params)
            rows = cursor.fetchall()

            historico_por_serial = {}
            seriales = sorted({
                str(r.get('serial') or '').strip()
                for r in rows
                if str(r.get('serial') or '').strip()
            })
            for i in range(0, len(seriales), 1000):
                serial_chunk = seriales[i:i + 1000]
                placeholders = ','.join(['%s'] * len(serial_chunk))
                cursor.execute(f"""
                    SELECT serial
                    FROM box_scans
                    WHERE serial IN ({placeholders})
                    GROUP BY serial
                    HAVING COUNT(*) > 1
                """, serial_chunk)
                for h in cursor.fetchall():
                    serial = str(h.get('serial') or '').strip()
                    if serial:
                        historico_por_serial[serial] = []

            seriales_repetidos = sorted(historico_por_serial.keys())
            for i in range(0, len(seriales_repetidos), 1000):
                serial_chunk = seriales_repetidos[i:i + 1000]
                placeholders = ','.join(['%s'] * len(serial_chunk))
                cursor.execute(f"""
                    SELECT
                        id,
                        serial,
                        CASE
                            WHEN TIME(last_scan) >= '07:30:00' THEN DATE(last_scan)
                            ELSE DATE_SUB(DATE(last_scan), INTERVAL 1 DAY)
                        END AS fecha_operativa,
                        last_scan
                    FROM box_scans
                    WHERE serial IN ({placeholders})
                    ORDER BY serial, last_scan, id
                """, serial_chunk)
                for h in cursor.fetchall():
                    serial = str(h.get('serial') or '').strip()
                    if not serial:
                        continue
                    historico_por_serial[serial].append({
                        "id": h.get('id'),
                        "fecha": str(h.get('fecha_operativa')) if h.get('fecha_operativa') else '',
                        "last_scan": str(h.get('last_scan')).replace('T', ' ') if h.get('last_scan') else '',
                    })

            records = []
            for r in rows:
                serial = r['serial'] or ''
                historico = historico_por_serial.get(serial, [])
                scan_id = r.get('scan_id')
                otros_escaneos = [
                    h for h in historico
                    if scan_id is None or h.get('id') != scan_id
                ]
                fechas_repetidas = []
                scans_repetidos = []
                for h in otros_escaneos:
                    fecha_hist = h.get('fecha') or ''
                    scan_hist = h.get('last_scan') or ''
                    if fecha_hist and fecha_hist not in fechas_repetidas:
                        fechas_repetidas.append(fecha_hist)
                    if scan_hist and scan_hist not in scans_repetidos:
                        scans_repetidos.append(scan_hist)
                es_duplicado_historico = len(otros_escaneos) > 0
                # El primer escaneo cronologico (id mas chico) es el ORIGINAL.
                # Los siguientes son los duplicados reales.
                es_original = False
                if historico:
                    primer_id = min((h.get('id') for h in historico if h.get('id') is not None), default=None)
                    es_original = (primer_id is not None and scan_id == primer_id)
                status = r.get('box_status') or ''
                if es_duplicado_historico and not es_original:
                    status = 'Duplicado'
                records.append({
                    "linea": r['linea'],
                    "fecha": str(r['fecha']) if r['fecha'] else '',
                    "part": r['part'] or '',
                    "model_code": r.get('model_code') or '',
                    "lot_no": r.get('lot_no') or '',
                    "box_code": r.get('box_code') or '',
                    "serial": serial,
                    "status": status,
                    "duplicado_historico": es_duplicado_historico and not es_original,
                    "es_original": es_original,
                    "total_repeticiones": len(historico),
                    "fechas_repetidas": ', '.join(fechas_repetidas),
                    "escaneos_repetidos": ' | '.join(scans_repetidos[:5]),
                    "first_scan": str(r.get('first_scan')).replace('T', ' ') if r.get('first_scan') else '',
                    "last_scan": str(r['last_scan']).replace('T', ' ') if r['last_scan'] else '',
                    "turno": r['turno'],
                })
            return jsonify({"success": True, "records": records, "total": len(records)})
        finally:
            cursor.close()
            conn.close()
    except Exception as e:
        return jsonify({"success": False, "error": str(e), "records": []}), 500


# Migracion 2026-05-27:
#   /api/inventario_general, /api/ubicacion, /api/movimientos
#     -> app/api/control_resultados/inventario_imd.py
#   /api/snapshot_inventario/{fechas,general,ubicacion,trigger}
#     + DDL + worker daemon + funcion de captura
#     -> app/api/shared/snapshot_inventario.py
# Las 6 GETs ahora exigen @login_requerido (antes eran publicas).


# === RUTAS MIGRADAS ===
# /api/mysql, /api/mysql/columns, /api/mysql/data, /api/mysql/update,
# /api/mysql/create, /api/mysql/usuario-actual, /api/mysql/delete,
# /visor-mysql, /control-modelos-visor-ajax
# Migradas a app/api/informacion_basica/control_modelos_visor.py
# Se registran automaticamente via registrar_blueprints_api() en app_factory.py



def generar_lot_no_secuencial(q, like, prefix, fecha):
    """Genera un numero de lote secuencial basado en la consulta"""
    last = execute_query(q, (like,), fetch="one")
    if last and last.get("lot_no"):
        try:
            seq = int(last["lot_no"].split("-")[-1]) + 1
        except Exception:
            seq = 1
    else:
        seq = 1
    return f"{prefix}{fecha}-{seq:04d}"


# Limpieza 2026-05-27: api_plan_run_start eliminado (modulo Control de operacion de linea Main borrado)


# Limpieza 2026-05-27: api_plan_run_end eliminado (modulo Control de operacion de linea Main borrado)


# Limpieza 2026-05-27: api_plan_run_pause eliminado (modulo Control de operacion de linea Main borrado)


# Limpieza 2026-05-27: api_plan_run_resume eliminado (modulo Control de operacion de linea Main borrado)


# Limpieza 2026-05-27: api_plan_run_status eliminado (modulo Control de operacion de linea Main borrado)


def crear_tabla_trazabilidad():
    """Crear tabla de trazabilidad (LOTE por WO/LINEA con estados)."""
    try:
        query = """
        CREATE TABLE IF NOT EXISTS trazabilidad (
            id INT AUTO_INCREMENT PRIMARY KEY,
            linea VARCHAR(32) NOT NULL,
            lot_no VARCHAR(32) NOT NULL,
            plan_id INT NULL,
            codigo_wo VARCHAR(32) NULL,
            estado ENUM('PLANEADO','INICIADO','PAUSA','FINALIZADO') DEFAULT 'PLANEADO',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            usuario VARCHAR(64) DEFAULT 'sistema',
            INDEX idx_linea (linea),
            INDEX idx_lot (lot_no),
            INDEX idx_estado (estado)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """
        execute_query(query)
        print(" Tabla trazabilidad creada/verificada")
    except Exception as e:
        print(f" Error creando tabla trazabilidad: {e}")


# crear_tabla_trazabilidad movido a app/startup_init.py


###############################################
# Metal Mask: poginas y API (integracion)
###############################################


def init_metal_mask_tables():
    """Crea/ajusta tablas usadas por Metal Mask si no existen."""
    try:
        # Tabla principal de masks con nombres de columnas en inglos (usadas por el frontend)
        execute_query(
            """
            CREATE TABLE IF NOT EXISTS masks (
                id INT AUTO_INCREMENT PRIMARY KEY,
                management_no VARCHAR(64) UNIQUE,
                storage_box VARCHAR(64),
                pcb_code VARCHAR(64),
                side VARCHAR(16),
                production_date DATE,
                used_count INT DEFAULT 0,
                max_count INT DEFAULT 0,
                allowance INT DEFAULT 0,
                model_name VARCHAR(255),
                tension_min DECIMAL(6,2),
                tension_max DECIMAL(6,2),
                thickness DECIMAL(6,2),
                supplier VARCHAR(128),
                registration_date VARCHAR(64),
                disuse ENUM('Uso','Desuso','Scrap') DEFAULT 'Uso',
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )

        # Asegurar valores del ENUM en caso de historiales previos (migracion suave)
        try:
            execute_query(
                "ALTER TABLE masks MODIFY COLUMN disuse ENUM('Use','Disuse','Uso','Desuso','Scrap') DEFAULT 'Uso'"
            )
            execute_query("UPDATE masks SET disuse='Uso' WHERE disuse='Use'")
            execute_query("UPDATE masks SET disuse='Desuso' WHERE disuse='Disuse'")
            execute_query(
                "ALTER TABLE masks MODIFY COLUMN disuse ENUM('Uso','Desuso','Scrap') DEFAULT 'Uso'"
            )
        except Exception as _:
            # Si falla (p.ej. por no existir la tabla/columna aon), continuar
            pass

        # Tabla de cajas de almacenamiento
        execute_query(
            """
            CREATE TABLE IF NOT EXISTS storage_boxes (
                id INT AUTO_INCREMENT PRIMARY KEY,
                management_no VARCHAR(64) UNIQUE,
                code VARCHAR(64),
                name VARCHAR(64),
                location VARCHAR(64),
                storage_status ENUM('Disponible','Ocupado','Mantenimiento') DEFAULT 'Disponible',
                used_status ENUM('Usado','No Usado') DEFAULT 'Usado',
                note TEXT,
                registration_date VARCHAR(64),
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )
        print(" Tablas Metal Mask creadas/verificadas")
    except Exception as e:
        print(f"Error creando/verificando tablas Metal Mask: {e}")


# init_metal_mask_tables movido a app/startup_init.py


# Migracion 2026-05-26: rutas alternas /control/metal-mask y
# /control/metal-mask/caja eliminadas (no consumidas por el frontend; el
# sidebar y scriptMain.js usan /control-mask-metal-ajax y
# /control-caja-mask-metal-ajax que ahora viven en sus blueprints).

# Migracion 2026-05-26: api_list_masks + api_create_mask + api_update_mask movidos a
# app/api/control_produccion/metal_mask.py


# Migracion 2026-05-26: api_get_storage + api_add_storage + api_update_storage
# movidos a app/api/control_produccion/caja_metal_mask.py (sigue siendo
# consumida tambien por MetalMask.js -- los blueprints comparten URL space).


@app.route("/api/bom-smt-data", methods=["GET"])
@login_requerido
def api_bom_smt_data():
    """API para obtener datos del BOM SMT basado en lonea y modelo"""
    try:
        # Obtener parometros
        linea = request.args.get("linea", "")
        model_code = request.args.get("model_code", "")

        if not linea or not model_code:
            return jsonify(
                {"success": False, "error": "Lonea y modelo son requeridos"}
            ), 400

        print(f"API BOM SMT - Filtros:")
        print(f"  Linea: {linea}")
        print(f"  Modelo: {model_code}")

        from .db_mysql import get_connection

        conn = get_connection()
        cursor = conn.cursor()

        # Mapear lonea SMT a nomero de lonea
        mapeo_lineas = {
            "SMT A": "2",
            "SMT B": "2",
            "SMT C": "3",
            "SMT D": "4",
            "1LINE": "2",
            "2LINE": "2",
            "3LINE": "3",
            "4LINE": "4",
        }

        linea_numero = mapeo_lineas.get(linea, "2")

        # Consultar ambas tablas (bom_smt_f y bom_smt_r) - solo elementos con cantidad > 0
        query_f = """
            SELECT
                id, linea, model_code, mounter, slot, material_code,
                description, feeder_info, qty, raw_filename,
                created_at, updated_at, 'FRONT' as tabla_tipo
            FROM bom_smt_f
            WHERE linea = %s AND model_code LIKE %s AND qty > 0
            ORDER BY mounter, slot
        """

        query_r = """
            SELECT
                id, linea, model_code, mounter, slot, material_code,
                description, feeder_info, qty, raw_filename,
                created_at, updated_at, 'REAR' as tabla_tipo
            FROM bom_smt_r
            WHERE linea = %s AND model_code LIKE %s AND qty > 0
            ORDER BY mounter, slot
        """

        # Buscar por modelo (puede contener EBR)
        model_pattern = f"%{model_code}%"

        # Ejecutar consultas
        cursor.execute(query_f, [linea_numero, model_pattern])
        resultados_f = cursor.fetchall()

        cursor.execute(query_r, [linea_numero, model_pattern])
        resultados_r = cursor.fetchall()

        # Combinar resultados
        todos_resultados = list(resultados_f) + list(resultados_r)

        print(
            f"Encontrados {len(todos_resultados)} registros BOM ({len(resultados_f)} F + {len(resultados_r)} R)"
        )
        print(
            f"Parametros de busqueda - Linea numero: {linea_numero}, Patron modelo: {model_pattern}"
        )

        # Formatear datos - solo incluir elementos con cantidad > 0
        formatted_data = []
        for row in todos_resultados:
            try:
                qty_value = row[8] if len(row) > 8 else 0

                # Solo incluir si qty > 0
                if qty_value <= 0:
                    continue

                formatted_row = {
                    "id": row[0] if len(row) > 0 else "",
                    "linea": row[1] if len(row) > 1 else "",
                    "model_code": row[2] if len(row) > 2 else "",
                    "mounter": row[3] if len(row) > 3 else "",
                    "slot": row[4] if len(row) > 4 else "",
                    "material_code": row[5] if len(row) > 5 else "",
                    "description": row[6] if len(row) > 6 else "",
                    "feeder_info": row[7] if len(row) > 7 else "",
                    "qty": qty_value,
                    "raw_filename": row[9] if len(row) > 9 else "",
                    "created_at": str(row[10]) if len(row) > 10 and row[10] else "",
                    "updated_at": str(row[11]) if len(row) > 11 and row[11] else "",
                    "tabla_tipo": row[12] if len(row) > 12 else "",
                    "status": "pending",  # Por defecto pendiente, se actualizaro con el mapeo
                }
                formatted_data.append(formatted_row)

            except Exception as row_error:
                print(f"Error procesando fila BOM: {row_error}")
                continue

        cursor.close()
        conn.close()

        print(f"BOM filtrado: {len(formatted_data)} elementos con qty > 0")

        return jsonify(
            {
                "success": True,
                "data": formatted_data,
                "total": len(formatted_data),
                "linea": linea,
                "model_code": model_code,
                "total_raw": len(todos_resultados),
                "total_filtered": len(formatted_data),
            }
        )

    except Exception as e:
        print(f"Error en api_bom_smt_data: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


# ====== FUNCIÓN PARA CONVERTIR NOMBRES DE LÍNEA ======
def convertir_linea_smt(linea_nombre):
    """
    Convierte nombres de línea SMT a formato de BD
    SMT A = 1line
    SMT B = 2line
    SMT C = 3line
    SMT D = 4line
    """
    conversion = {
        "SMT A": "1line",
        "SMT B": "2line",
        "SMT C": "3line",
        "SMT D": "4line",
    }
    return conversion.get(linea_nombre, linea_nombre)


def convertir_linea_smt_reverso(linea_bd):
    """
    Convierte formato de BD a nombres de línea SMT
    1line = SMT A
    2line = SMT B
    3line = SMT C
    4line = SMT D
    """
    conversion = {
        "1line": "SMT A",
        "2line": "SMT B",
        "3line": "SMT C",
        "4line": "SMT D",
    }
    return conversion.get(linea_bd, linea_bd)


# ====== HISTORIAL ICT (FRONT FULL DEFECTS 2) ======
def _append_indexable_text_filter(sql, params, column_name, raw_value, exact_min_length=12):
    """Agregar filtro exacto o por prefijo para evitar LIKE con comodin inicial."""
    value = (raw_value or "").strip()
    if not value:
        return sql

    if len(value) >= exact_min_length:
        sql += f" AND {column_name}=%s"
        params.append(value)
    else:
        sql += f" AND {column_name} LIKE %s"
        params.append(f"{value}%")
    return sql


def _ict_format_row(row):
    """Convertir campos fecha/hora a cadenas serializables."""
    if not row:
        return {}

    formatted = {}
    for key, value in row.items():
        if isinstance(value, datetime):
            formatted[key] = value.isoformat(sep=" ")
        elif isinstance(value, date):
            formatted[key] = value.isoformat()
        elif isinstance(value, dt_time):
            formatted[key] = value.strftime("%H:%M:%S")
        elif isinstance(value, timedelta):
            formatted[key] = str(value)
        else:
            formatted[key] = value
    return formatted


def _ict_find_history_record(barcode, ts=None):
    """Buscar el registro resumen que apunta al archivo LGD local."""
    sql = (
        "SELECT barcode, ts, fuente_archivo, linea, ict "
        "FROM history_ict WHERE barcode=%s"
    )
    params = [barcode]
    if ts:
        sql += " AND ts=%s"
        params.append(ts)
    sql += " ORDER BY ts DESC LIMIT 1"
    return execute_query(sql, tuple(params), fetch="one")


def _ict_load_local_parameters(barcode, ts=None):
    history_row = _ict_find_history_record(barcode, ts)
    if not history_row:
        raise IctLgdNotFoundError("No se encontro el registro ICT solicitado.")

    source_file = (history_row.get("fuente_archivo") or "").strip()
    if not source_file:
        raise IctLgdPathError("El registro no tiene fuente_archivo.")

    lgd_path = resolve_lgd_path(source_file)
    cached_rows = get_lgd_parameters_for_barcode(str(lgd_path), barcode)
    # Copia local porque vamos a setear linea/ict/fuente_archivo. El resultado
    # de get_lgd_parameters_for_barcode comparte memoria con el cache global.
    rows = [dict(row) for row in cached_rows]
    for row in rows:
        row.setdefault("linea", history_row.get("linea"))
        row.setdefault("ict", history_row.get("ict"))
        row.setdefault("fuente_archivo", source_file)
    return rows, source_file


def _vision_format_value(value):
    """Convertir valores de history_vision a cadenas serializables."""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, dt_time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, timedelta):
        total_seconds = int(value.total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    if isinstance(value, str) and re.fullmatch(r"\d{2}:\d{2}:\d{2}\.\d+", value):
        return value.split(".", 1)[0]
    return value


def _vision_parse_datetime(value):
    """Normalizar distintos formatos de fecha/hora a datetime."""
    if isinstance(value, datetime):
        return value

    if isinstance(value, str):
        raw_value = value.strip()
        if not raw_value:
            return None
        try:
            return datetime.fromisoformat(raw_value)
        except ValueError:
            for fmt in (
                "%Y-%m-%d %H:%M:%S.%f",
                "%Y-%m-%d %H:%M:%S",
                "%Y/%m/%d %H:%M:%S.%f",
                "%Y/%m/%d %H:%M:%S",
            ):
                try:
                    return datetime.strptime(raw_value, fmt)
                except ValueError:
                    continue
    return None


def _vision_parse_date(value):
    """Normalizar valores de fecha a date."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        raw_value = value.strip()
        if not raw_value:
            return None
        try:
            return date.fromisoformat(raw_value)
        except ValueError:
            return None
    return None


def _vision_parse_time(value):
    """Normalizar valores de hora a time."""
    if isinstance(value, dt_time):
        return value
    if isinstance(value, timedelta):
        total_seconds = int(value.total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return dt_time(hour=hours % 24, minute=minutes, second=seconds)
    if isinstance(value, str):
        raw_value = value.strip()
        if not raw_value:
            return None
        try:
            return dt_time.fromisoformat(raw_value)
        except ValueError:
            for fmt in ("%H:%M:%S.%f", "%H:%M:%S"):
                try:
                    return datetime.strptime(raw_value, fmt).time()
                except ValueError:
                    continue
    return None


def _vision_reference_datetime(record):
    """Obtener la fecha/hora mas confiable para resolver la imagen."""
    parsed_value = _vision_parse_datetime(record.get("log_datetime"))
    if parsed_value:
        return parsed_value

    log_date = _vision_parse_date(record.get("log_date"))
    log_time = _vision_parse_time(record.get("log_time"))
    if log_date and log_time:
        return datetime.combine(log_date, log_time)

    for key in ("captured_at_utc", "created_at"):
        parsed_value = _vision_parse_datetime(record.get(key))
        if parsed_value:
            return parsed_value

    return None


def _vision_extract_host_from_source_file(source_file):
    """Extraer el host UNC desde source_file."""
    raw_source = str(source_file or "").strip()
    match = re.match(r"^\\\\([^\\]+)\\", raw_source)
    return match.group(1).strip() if match else ""


def _vision_unique_values(values):
    """Eliminar duplicados preservando el orden original."""
    unique_values = []
    seen_values = set()
    for value in values:
        normalized = str(value or "").strip()
        if not normalized:
            continue
        value_key = normalized.lower()
        if value_key in seen_values:
            continue
        seen_values.add(value_key)
        unique_values.append(normalized)
    return unique_values


def _vision_candidate_share_roots(record):
    """Construir raices UNC candidatas para buscar imagenes."""
    machine_name = str(record.get("machine_name") or "").strip()
    candidate_hosts = _vision_unique_values(
        [
            _vision_extract_host_from_source_file(record.get("source_file")),
            record.get("machine_ip"),
        ]
    )

    roots = []
    seen_roots = set()
    for host in candidate_hosts:
        root_candidates = []
        if machine_name:
            root_candidates.append(Path(rf"\\{host}\Result {machine_name}"))
        root_candidates.append(Path(rf"\\{host}\Result"))

        for candidate_root in root_candidates:
            root_key = os.path.normcase(os.path.normpath(str(candidate_root)))
            if root_key in seen_roots:
                continue
            seen_roots.add(root_key)
            roots.append(candidate_root)

    return roots


def _vision_candidate_hour_directories(root_path, part_code, reference_dt):
    """Construir directorios candidatos por hora, incluyendo borde de hora."""
    if not part_code or not reference_dt:
        return []

    base_hour = reference_dt.replace(minute=0, second=0, microsecond=0)
    candidate_hours = [base_hour]

    previous_delta = (reference_dt - base_hour).total_seconds()
    if previous_delta <= 1.0:
        candidate_hours.append(base_hour - timedelta(hours=1))

    next_hour = base_hour + timedelta(hours=1)
    next_delta = (next_hour - reference_dt).total_seconds()
    if next_delta <= 1.0:
        candidate_hours.append(next_hour)

    ebr_prefix = str(part_code or "").strip().upper()[:11]
    candidate_part_directories = []
    seen_part_directories = set()

    def append_part_directory(path_obj):
        path_key = os.path.normcase(os.path.normpath(str(path_obj)))
        if path_key in seen_part_directories:
            return
        seen_part_directories.add(path_key)
        candidate_part_directories.append(path_obj)

    append_part_directory(root_path / part_code)

    if ebr_prefix and ebr_prefix != str(part_code or "").strip().upper():
        append_part_directory(root_path / ebr_prefix)

    try:
        if root_path.is_dir() and ebr_prefix:
            matching_directories = sorted(
                [
                    child
                    for child in root_path.iterdir()
                    if child.is_dir()
                    and child.name.strip().upper()[:11] == ebr_prefix
                ],
                key=lambda path: path.name.lower(),
            )
            for matching_directory in matching_directories:
                append_part_directory(matching_directory)
    except OSError:
        pass

    directories = []
    seen_directories = set()
    for part_directory in candidate_part_directories:
        for candidate_hour in candidate_hours:
            base_dir = (
                part_directory
                / "Image(Process)"
                / candidate_hour.strftime("%Y")
                / candidate_hour.strftime("%m")
                / candidate_hour.strftime("%d")
                / candidate_hour.strftime("%H")
            )
            dir_key = os.path.normcase(os.path.normpath(str(base_dir)))
            if dir_key in seen_directories:
                continue
            seen_directories.add(dir_key)
            directories.append(base_dir)

    return directories


def _vision_share_name(root_path):
    """Obtener el nombre visible del share UNC."""
    normalized = os.path.normpath(str(root_path)).rstrip("\\/")
    share_name = os.path.basename(normalized)
    if share_name:
        return share_name
    return normalized.split("\\")[-1] if normalized else ""


def _vision_is_safe_path(target_path, allowed_roots):
    """Validar que la ruta final quede contenida dentro de una raiz candidata."""
    normalized_target = os.path.normcase(os.path.normpath(str(target_path)))
    for allowed_root in allowed_roots:
        normalized_root = os.path.normcase(os.path.normpath(str(allowed_root)))
        try:
            if os.path.commonpath([normalized_target, normalized_root]) == normalized_root:
                return True
        except ValueError:
            continue
    return False


def _get_history_vision_record(record_id):
    """Obtener un registro completo de history_vision por id."""
    sql = (
        "SELECT id, source_uid, machine_name, machine_ip, result, log_timestamp, "
        "log_datetime, log_date, log_time, serial_qr, barcode, qr_payload, work_area, "
        "part_code, source_file, captured_at_utc, created_at "
        "FROM history_vision WHERE id=%s LIMIT 1"
    )
    return execute_query(sql, (record_id,), fetch="one")


def _resolve_history_vision_image(record):
    """Resolver la imagen mas cercana a un registro de history_vision."""
    image_extensions = {".jpg", ".jpeg", ".png", ".bmp"}
    expected_result = str(record.get("result") or "").strip().upper()
    part_code = str(record.get("part_code") or "").strip()
    part_code_prefix = part_code.upper()[:11]
    reference_dt = _vision_reference_datetime(record)
    share_roots = _vision_candidate_share_roots(record)

    result_payload = {
        "record_id": record.get("id"),
        "resolved_path": "",
        "share_name": "",
        "side_folder": "",
        "delta_seconds": None,
        "searched_paths": [],
        "reference_datetime": _vision_format_value(reference_dt) if reference_dt else "",
        "share_roots": [str(root) for root in share_roots],
    }

    if not part_code:
        result_payload["error"] = "El registro no tiene numero de parte."
        return result_payload

    if expected_result not in {"OK", "NG"}:
        result_payload["error"] = "El resultado del registro no es valido."
        return result_payload

    if not reference_dt:
        result_payload["error"] = "No se pudo determinar la fecha/hora del registro."
        return result_payload

    if not share_roots:
        result_payload["error"] = "No se encontraron rutas compartidas candidatas."
        return result_payload

    filename_part_prefix = part_code_prefix or part_code.upper()
    filename_pattern = re.compile(
        rf"^{re.escape(filename_part_prefix)}(?:[^_]*)_(?P<side>[^_]+)_(?P<stamp>\d{{8}}_\d{{9}})_(?P<result>OK|NG)\.(?P<ext>jpg|jpeg|png|bmp)$",
        re.IGNORECASE,
    )

    best_candidate = None
    searched_paths = []
    seen_search_paths = set()

    for share_root in share_roots:
        for hour_dir in _vision_candidate_hour_directories(
            share_root, part_code, reference_dt
        ):
            hour_dir_str = str(hour_dir)
            hour_dir_key = os.path.normcase(os.path.normpath(hour_dir_str))
            if hour_dir_key not in seen_search_paths:
                seen_search_paths.add(hour_dir_key)
                searched_paths.append(hour_dir_str)

            try:
                if not hour_dir.is_dir():
                    continue
                side_directories = sorted(
                    [child for child in hour_dir.iterdir() if child.is_dir()],
                    key=lambda path: path.name.lower(),
                )
            except OSError:
                continue

            for side_dir in side_directories:
                result_dir = side_dir / expected_result
                result_dir_str = str(result_dir)
                result_dir_key = os.path.normcase(os.path.normpath(result_dir_str))
                if result_dir_key not in seen_search_paths:
                    seen_search_paths.add(result_dir_key)
                    searched_paths.append(result_dir_str)

                try:
                    if not result_dir.is_dir():
                        continue
                    image_candidates = sorted(
                        [child for child in result_dir.iterdir() if child.is_file()],
                        key=lambda path: path.name.lower(),
                    )
                except OSError:
                    continue

                for image_path in image_candidates:
                    if image_path.suffix.lower() not in image_extensions:
                        continue

                    match = filename_pattern.match(image_path.name)
                    if not match or match.group("result").upper() != expected_result:
                        continue

                    try:
                        file_dt = datetime.strptime(
                            match.group("stamp"), "%Y%m%d_%H%M%S%f"
                        )
                    except ValueError:
                        continue

                    delta_seconds = abs((file_dt - reference_dt).total_seconds())
                    candidate = {
                        "path": image_path,
                        "share_name": _vision_share_name(share_root),
                        "side_folder": side_dir.name,
                        "delta_seconds": delta_seconds,
                    }

                    if best_candidate is None or delta_seconds < best_candidate["delta_seconds"]:
                        best_candidate = candidate

    result_payload["searched_paths"] = searched_paths

    if best_candidate and best_candidate["delta_seconds"] <= 1.0:
        result_payload.update(
            {
                "resolved_path": str(best_candidate["path"]),
                "share_name": best_candidate["share_name"],
                "side_folder": best_candidate["side_folder"],
                "delta_seconds": round(best_candidate["delta_seconds"], 3),
            }
        )
        return result_payload

    result_payload["error"] = "No se encontro una imagen dentro del umbral de 1 segundo."
    return result_payload


def _build_history_vision_query():
    """Construir query y parámetros para consultar history_vision."""
    fecha_desde = request.args.get("fecha_desde", "").strip()
    fecha_hasta = request.args.get("fecha_hasta", "").strip()
    linea = request.args.get("linea", "").strip()
    resultado = request.args.get("resultado", "").strip()
    numero_parte = request.args.get("numero_parte", "").strip()
    qr = request.args.get("qr", "").strip()
    barcode = request.args.get("barcode", "").strip()

    sql = (
        "SELECT "
        "id, "
        "machine_name AS linea, "
        "log_date AS fecha, "
        "log_time AS hora, "
        "part_code AS numero_parte, "
        "qr_payload AS qr, "
        "barcode, "
        "result AS resultado "
        "FROM history_vision WHERE 1=1"
    )
    params = []

    if fecha_desde:
        sql += " AND log_date >= %s"
        params.append(fecha_desde)
    if fecha_hasta:
        sql += " AND log_date <= %s"
        params.append(fecha_hasta)
    if linea:
        sql += " AND machine_name LIKE %s"
        params.append(f"%{linea}%")
    if resultado:
        sql += " AND result=%s"
        params.append(resultado)
    if numero_parte:
        sql += " AND part_code LIKE %s"
        params.append(f"%{numero_parte}%")
    if qr:
        sql += " AND qr_payload LIKE %s"
        params.append(f"%{qr}%")
    if barcode:
        sql += " AND barcode LIKE %s"
        params.append(f"%{barcode}%")

    sql += " ORDER BY COALESCE(log_datetime, created_at) DESC, id DESC"

    return sql, tuple(params) if params else None


def _build_history_vision_pass_fail_summary_query():
    """Construir query y parámetros para el resumen Pass/Fail de history_vision por línea y número de parte."""
    fecha_desde = request.args.get("fecha_desde", "").strip()
    fecha_hasta = request.args.get("fecha_hasta", "").strip()
    numero_parte = request.args.get("numero_parte", "").strip()

    sql = (
        "SELECT "
        "COALESCE(NULLIF(TRIM(machine_name), ''), 'SIN LINEA') AS linea, "
        "COALESCE(NULLIF(TRIM(part_code), ''), 'SIN NUMERO DE PARTE') AS numero_parte, "
        "COUNT(*) AS total, "
        "SUM(CASE WHEN UPPER(COALESCE(result, '')) = 'OK' THEN 1 ELSE 0 END) AS ok_count, "
        "SUM(CASE WHEN UPPER(COALESCE(result, '')) = 'NG' THEN 1 ELSE 0 END) AS ng_count "
        "FROM history_vision WHERE 1=1"
    )
    params = []

    if fecha_desde:
        sql += " AND log_date >= %s"
        params.append(fecha_desde)
    if fecha_hasta:
        sql += " AND log_date <= %s"
        params.append(fecha_hasta)
    if numero_parte:
        sql += " AND part_code LIKE %s"
        params.append(f"%{numero_parte}%")

    sql += (
        " GROUP BY COALESCE(NULLIF(TRIM(machine_name), ''), 'SIN LINEA'),"
        " COALESCE(NULLIF(TRIM(part_code), ''), 'SIN NUMERO DE PARTE')"
        " ORDER BY linea ASC, total DESC, numero_parte ASC"
    )

    return sql, tuple(params) if params else None


VISION_PASS_FAIL_EXCEL_IMAGE_WIDTH = 430
VISION_PASS_FAIL_EXCEL_IMAGE_HEIGHT = 28


def _measure_vision_pass_fail_excel_text(draw, text, font):
    if hasattr(draw, "textbbox"):
        left, top, right, bottom = draw.textbbox((0, 0), text, font=font)
        return right - left, bottom - top

    if hasattr(draw, "textsize"):
        return draw.textsize(text, font=font)

    if hasattr(font, "getbbox"):
        left, top, right, bottom = font.getbbox(text)
        return right - left, bottom - top

    if hasattr(font, "getsize"):
        return font.getsize(text)

    return (max(len(text), 1) * 8, 12)


def _load_vision_pass_fail_excel_font(size=12, bold=False):
    from PIL import ImageFont

    windows_font_dir = Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts"
    preferred_font_names = [
        "DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf",
        "Arial Bold.ttf" if bold else "Arial.ttf",
        "Calibri Bold.ttf" if bold else "Calibri.ttf",
        "arialbd.ttf" if bold else "arial.ttf",
        "calibrib.ttf" if bold else "calibri.ttf",
    ]

    for font_name in preferred_font_names:
        try:
            return ImageFont.truetype(font_name, size=size)
        except OSError:
            continue

    font_candidates = [
        windows_font_dir / ("arialbd.ttf" if bold else "arial.ttf"),
        windows_font_dir / ("calibrib.ttf" if bold else "calibri.ttf"),
        windows_font_dir / ("DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf"),
    ]

    for font_path in font_candidates:
        try:
            if font_path.is_file():
                return ImageFont.truetype(str(font_path), size=size)
        except OSError:
            continue

    return ImageFont.load_default()


def _build_vision_pass_fail_excel_bar_image(porcentaje_ok, porcentaje_ng):
    from io import BytesIO

    from PIL import Image, ImageDraw

    pass_rate = max(0.0, min(100.0, float(porcentaje_ok or 0)))
    fail_rate = max(0.0, min(100.0, float(porcentaje_ng or 0)))

    canvas_width = 430
    canvas_height = 28
    bar_width = 350
    bar_height = 20
    label_gap = 8
    label_width = canvas_width - bar_width - label_gap
    radius = bar_height // 2

    ok_width = int(round((pass_rate / 100.0) * bar_width))
    ok_width = max(0, min(bar_width, ok_width))
    ng_width = max(0, bar_width - ok_width)

    image = Image.new("RGBA", (canvas_width, canvas_height), (0, 0, 0, 0))
    draw = ImageDraw.Draw(image)

    bar_x = 0
    bar_y = (canvas_height - bar_height) // 2

    # Base dark bar
    draw.rounded_rectangle(
        (bar_x, bar_y, bar_x + bar_width, bar_y + bar_height),
        radius=radius,
        fill="#1A2740",
    )

    # Draw segments inside a rounded mask for clean edges
    segments = Image.new("RGBA", (bar_width, bar_height), (0, 0, 0, 0))
    segments_draw = ImageDraw.Draw(segments)
    if ok_width > 0:
        segments_draw.rectangle((0, 0, ok_width, bar_height), fill="#4CAF63")
    if ng_width > 0:
        segments_draw.rectangle(
            (bar_width - ng_width, 0, bar_width, bar_height), fill="#E45454"
        )

    mask = Image.new("L", (bar_width, bar_height), 0)
    mask_draw = ImageDraw.Draw(mask)
    mask_draw.rounded_rectangle((0, 0, bar_width, bar_height), radius=radius, fill=255)
    image.paste(segments, (bar_x, bar_y), mask)

    pass_label = f"{pass_rate:.2f}%"
    fail_label = f"{fail_rate:.2f}%"

    pass_font = _load_vision_pass_fail_excel_font(size=12, bold=True)
    fail_font = _load_vision_pass_fail_excel_font(size=12, bold=True)

    pass_text_width, pass_text_height = _measure_vision_pass_fail_excel_text(
        draw, pass_label, pass_font
    )
    pass_center_x = bar_x + max(ok_width / 2, min(bar_width / 2, 70))
    pass_text_x = int(
        max(
            bar_x + 10,
            min(pass_center_x - pass_text_width / 2, bar_width - pass_text_width - 10),
        )
    )
    pass_text_y = int(bar_y + (bar_height - pass_text_height) / 2 - 1)
    draw.text(
        (pass_text_x, pass_text_y),
        pass_label,
        font=pass_font,
        fill="#FFFFFF",
    )

    _, fail_text_height = _measure_vision_pass_fail_excel_text(
        draw, fail_label, fail_font
    )
    fail_text_x = bar_width + label_gap
    fail_text_y = int(bar_y + (bar_height - fail_text_height) / 2 - 1)
    draw.text(
        (fail_text_x, fail_text_y),
        fail_label,
        font=fail_font,
        fill="#1F1F1F",
    )

    output = BytesIO()
    image.save(output, format="PNG")
    output.seek(0)
    return output


def _vision_pass_fail_text_width(text, char_width=10, spacing=3):
    if not text:
        return 0

    width = 0
    for idx, char in enumerate(text):
        width += _get_vision_pass_fail_char_width(char, char_width)
        if idx < len(text) - 1:
            width += spacing
    return width


def _set_vision_pass_fail_pixel(pixels, width, x, y, color):
    if x < 0 or y < 0:
        return

    idx = ((y * width) + x) * 4
    if idx < 0 or idx + 3 >= len(pixels):
        return

    pixels[idx] = color[0]
    pixels[idx + 1] = color[1]
    pixels[idx + 2] = color[2]
    pixels[idx + 3] = color[3]


def _fill_vision_pass_fail_rounded_rect(
    pixels, canvas_width, x, y, width, height, radius, color
):
    radius = max(0, min(radius, width // 2, height // 2))
    radius_sq = radius * radius

    for local_y in range(height):
        for local_x in range(width):
            if radius == 0:
                inside = True
            elif radius <= local_x < width - radius or radius <= local_y < height - radius:
                inside = True
            else:
                corner_x = radius if local_x < radius else width - radius - 1
                corner_y = radius if local_y < radius else height - radius - 1
                delta_x = local_x - corner_x
                delta_y = local_y - corner_y
                inside = (delta_x * delta_x) + (delta_y * delta_y) <= radius_sq

            if inside:
                _set_vision_pass_fail_pixel(
                    pixels, canvas_width, x + local_x, y + local_y, color
                )


def _draw_vision_pass_fail_disc(pixels, canvas_width, center_x, center_y, radius, color):
    radius_sq = radius * radius
    for local_y in range(-radius, radius + 1):
        for local_x in range(-radius, radius + 1):
            if (local_x * local_x) + (local_y * local_y) <= radius_sq:
                _set_vision_pass_fail_pixel(
                    pixels,
                    canvas_width,
                    center_x + local_x,
                    center_y + local_y,
                    color,
                )


def _draw_vision_pass_fail_line(
    pixels, canvas_width, x1, y1, x2, y2, thickness, color
):
    dx = x2 - x1
    dy = y2 - y1
    steps = max(abs(dx), abs(dy), 1)
    radius = max(1, thickness // 2)

    for step in range(steps + 1):
        point_x = int(round(x1 + (dx * step / steps)))
        point_y = int(round(y1 + (dy * step / steps)))
        _draw_vision_pass_fail_disc(
            pixels, canvas_width, point_x, point_y, radius, color
        )


def _draw_vision_pass_fail_segment(
    pixels,
    canvas_width,
    origin_x,
    origin_y,
    segment_name,
    color,
    char_width=10,
    char_height=14,
    thickness=2,
):
    mid_y = origin_y + (char_height // 2)
    bottom_y = origin_y + char_height - thickness
    right_x = origin_x + char_width - thickness
    horizontal_width = char_width - (thickness * 2)
    vertical_height = (char_height // 2) - thickness - 1
    radius = max(1, thickness // 2)

    if segment_name == "top":
        _fill_vision_pass_fail_rounded_rect(
            pixels,
            canvas_width,
            origin_x + thickness,
            origin_y,
            horizontal_width,
            thickness,
            radius,
            color,
        )
    elif segment_name == "middle":
        _fill_vision_pass_fail_rounded_rect(
            pixels,
            canvas_width,
            origin_x + thickness,
            mid_y - (thickness // 2),
            horizontal_width,
            thickness,
            radius,
            color,
        )
    elif segment_name == "bottom":
        _fill_vision_pass_fail_rounded_rect(
            pixels,
            canvas_width,
            origin_x + thickness,
            bottom_y,
            horizontal_width,
            thickness,
            radius,
            color,
        )
    elif segment_name == "upper_left":
        _fill_vision_pass_fail_rounded_rect(
            pixels,
            canvas_width,
            origin_x,
            origin_y + thickness,
            thickness,
            vertical_height,
            radius,
            color,
        )
    elif segment_name == "upper_right":
        _fill_vision_pass_fail_rounded_rect(
            pixels,
            canvas_width,
            right_x,
            origin_y + thickness,
            thickness,
            vertical_height,
            radius,
            color,
        )
    elif segment_name == "lower_left":
        _fill_vision_pass_fail_rounded_rect(
            pixels,
            canvas_width,
            origin_x,
            mid_y + 1,
            thickness,
            vertical_height,
            radius,
            color,
        )
    elif segment_name == "lower_right":
        _fill_vision_pass_fail_rounded_rect(
            pixels,
            canvas_width,
            right_x,
            mid_y + 1,
            thickness,
            vertical_height,
            radius,
            color,
        )


def _get_vision_pass_fail_segments(char):
    segment_map = {
        "0": (
            "top",
            "upper_left",
            "upper_right",
            "lower_left",
            "lower_right",
            "bottom",
        ),
        "1": ("upper_right", "lower_right"),
        "2": ("top", "upper_right", "middle", "lower_left", "bottom"),
        "3": ("top", "upper_right", "middle", "lower_right", "bottom"),
        "4": ("upper_left", "upper_right", "middle", "lower_right"),
        "5": ("top", "upper_left", "middle", "lower_right", "bottom"),
        "6": (
            "top",
            "upper_left",
            "middle",
            "lower_left",
            "lower_right",
            "bottom",
        ),
        "7": ("top", "upper_right", "lower_right"),
        "8": (
            "top",
            "upper_left",
            "upper_right",
            "middle",
            "lower_left",
            "lower_right",
            "bottom",
        ),
        "9": (
            "top",
            "upper_left",
            "upper_right",
            "middle",
            "lower_right",
            "bottom",
        ),
    }
    return segment_map.get(char, ())


def _get_vision_pass_fail_char_width(char, default_width=10):
    if char == "1":
        return default_width - 2
    if char == ".":
        return 4
    if char == "%":
        return default_width + 2
    if char == " ":
        return max(3, default_width // 2)
    return default_width


def _draw_vision_pass_fail_vector_char(
    pixels,
    canvas_width,
    x,
    y,
    char,
    color,
    char_width=10,
    char_height=14,
    thickness=2,
):
    actual_width = _get_vision_pass_fail_char_width(char, char_width)

    if char.isdigit():
        for segment_name in _get_vision_pass_fail_segments(char):
            _draw_vision_pass_fail_segment(
                pixels,
                canvas_width,
                x,
                y,
                segment_name,
                color,
                char_width=actual_width,
                char_height=char_height,
                thickness=thickness,
            )
        return actual_width

    if char == ".":
        radius = max(1, thickness)
        _draw_vision_pass_fail_disc(
            pixels,
            canvas_width,
            x + radius,
            y + char_height - radius - 1,
            radius,
            color,
        )
        return actual_width

    if char == "%":
        disc_radius = max(1, thickness)
        _draw_vision_pass_fail_disc(
            pixels, canvas_width, x + 2, y + 3, disc_radius, color
        )
        _draw_vision_pass_fail_disc(
            pixels,
            canvas_width,
            x + actual_width - 3,
            y + char_height - 3,
            disc_radius,
            color,
        )
        _draw_vision_pass_fail_line(
            pixels,
            canvas_width,
            x + actual_width - 3,
            y + 1,
            x + 2,
            y + char_height - 2,
            thickness,
            color,
        )
        return actual_width

    return actual_width


def _draw_vision_pass_fail_vector_text(
    pixels,
    canvas_width,
    x,
    y,
    text,
    color,
    char_width=10,
    char_height=14,
    thickness=2,
    spacing=3,
):
    cursor_x = x
    for idx, char in enumerate(text):
        cursor_x += _draw_vision_pass_fail_vector_char(
            pixels,
            canvas_width,
            cursor_x,
            y,
            char,
            color,
            char_width=char_width,
            char_height=char_height,
            thickness=thickness,
        )
        if idx < len(text) - 1:
            cursor_x += spacing


def _build_vision_pass_fail_excel_bar_png_bytes(porcentaje_ok, porcentaje_ng):
    pass_rate = max(0.0, min(100.0, float(porcentaje_ok or 0)))
    fail_rate = max(0.0, min(100.0, float(porcentaje_ng or 0)))

    canvas_width = VISION_PASS_FAIL_EXCEL_IMAGE_WIDTH
    canvas_height = VISION_PASS_FAIL_EXCEL_IMAGE_HEIGHT
    bar_width = 350
    bar_height = 20
    label_gap = 8
    radius = bar_height // 2
    bar_x = 0
    bar_y = (canvas_height - bar_height) // 2

    ok_width = int(round((pass_rate / 100.0) * bar_width))
    ok_width = max(0, min(bar_width, ok_width))
    ng_width = max(0, bar_width - ok_width)

    pixels = bytearray(canvas_width * canvas_height * 4)

    _fill_vision_pass_fail_rounded_rect(
        pixels, canvas_width, bar_x, bar_y, bar_width, bar_height, radius, (26, 39, 64, 255)
    )

    for local_y in range(bar_height):
        for local_x in range(bar_width):
            if radius <= local_x < bar_width - radius or radius <= local_y < bar_height - radius:
                inside = True
            else:
                corner_x = radius if local_x < radius else bar_width - radius - 1
                corner_y = radius if local_y < radius else bar_height - radius - 1
                delta_x = local_x - corner_x
                delta_y = local_y - corner_y
                inside = (delta_x * delta_x) + (delta_y * delta_y) <= radius * radius

            if not inside:
                continue

            if ok_width > 0 and local_x < ok_width:
                color = (76, 175, 99, 255)
            elif ng_width > 0 and local_x >= bar_width - ng_width:
                color = (228, 84, 84, 255)
            else:
                color = (26, 39, 64, 255)

            _set_vision_pass_fail_pixel(
                pixels, canvas_width, bar_x + local_x, bar_y + local_y, color
            )

    pass_label = f"{pass_rate:.2f}%"
    fail_label = f"{fail_rate:.2f}%"

    char_width = 10
    char_height = 14
    thickness = 2
    spacing = 2
    pass_text_width = _vision_pass_fail_text_width(
        pass_label, char_width=char_width, spacing=spacing
    )
    pass_center_x = bar_x + max(ok_width / 2, min(bar_width / 2, 70))
    pass_text_x = int(
        max(
            bar_x + 10,
            min(pass_center_x - pass_text_width / 2, bar_width - pass_text_width - 10),
        )
    )
    pass_text_y = bar_y + max(0, (bar_height - char_height) // 2)
    _draw_vision_pass_fail_vector_text(
        pixels,
        canvas_width,
        pass_text_x,
        pass_text_y,
        pass_label,
        (255, 255, 255, 255),
        char_width=char_width,
        char_height=char_height,
        thickness=thickness,
        spacing=spacing,
    )

    fail_text_x = bar_width + label_gap
    fail_text_y = bar_y + max(0, (bar_height - char_height) // 2)
    _draw_vision_pass_fail_vector_text(
        pixels,
        canvas_width,
        fail_text_x,
        fail_text_y,
        fail_label,
        (31, 31, 31, 255),
        char_width=char_width,
        char_height=char_height,
        thickness=thickness,
        spacing=spacing,
    )

    raw_rows = bytearray()
    stride = canvas_width * 4
    for row_idx in range(canvas_height):
        raw_rows.append(0)
        start = row_idx * stride
        raw_rows.extend(pixels[start : start + stride])

    compressed = zlib.compress(bytes(raw_rows), 9)

    def png_chunk(chunk_type, data):
        return (
            struct.pack(">I", len(data))
            + chunk_type
            + data
            + struct.pack(">I", zlib.crc32(chunk_type + data) & 0xFFFFFFFF)
        )

    png_bytes = bytearray()
    png_bytes.extend(b"\x89PNG\r\n\x1a\n")
    png_bytes.extend(
        png_chunk(
            b"IHDR",
            struct.pack(
                ">IIBBBBB",
                canvas_width,
                canvas_height,
                8,
                6,
                0,
                0,
                0,
            ),
        )
    )
    png_bytes.extend(png_chunk(b"IDAT", compressed))
    png_bytes.extend(png_chunk(b"IEND", b""))
    return bytes(png_bytes)


def _create_vision_pass_fail_excel_image(porcentaje_ok, porcentaje_ng):
    from openpyxl.drawing.image import Image as XLImage

    try:
        image_buffer = _build_vision_pass_fail_excel_bar_image(
            porcentaje_ok, porcentaje_ng
        )
        excel_image = XLImage(image_buffer)
    except Exception:
        png_bytes = _build_vision_pass_fail_excel_bar_png_bytes(
            porcentaje_ok, porcentaje_ng
        )

        class RawPngExcelImage(XLImage):
            def __init__(self, raw_bytes, width, height):
                self.ref = io.BytesIO(raw_bytes)
                self._raw_bytes = raw_bytes
                self.width = width
                self.height = height
                self.format = "png"

            def _data(self):
                return self._raw_bytes

        excel_image = RawPngExcelImage(
            png_bytes,
            VISION_PASS_FAIL_EXCEL_IMAGE_WIDTH,
            VISION_PASS_FAIL_EXCEL_IMAGE_HEIGHT,
        )

    excel_image.width = VISION_PASS_FAIL_EXCEL_IMAGE_WIDTH
    excel_image.height = VISION_PASS_FAIL_EXCEL_IMAGE_HEIGHT
    return excel_image


def _send_excel_download(output, filename):
    send_kwargs = {
        "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "as_attachment": True,
    }

    try:
        return send_file(output, download_name=filename, **send_kwargs)
    except TypeError:
        output.seek(0)
        return send_file(output, attachment_filename=filename, **send_kwargs)


@app.route("/historial-vision")
@app.route("/historial-vision-ajax")
@login_requerido
def historial_vision():
    """Servir la página de Historial Vision."""
    try:
        return render_template("Control de resultados/history_vision.html")
    except Exception as e:
        print(f"Error al cargar Historial Vision: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/historial-vision-pass-fail")
@app.route("/historial-vision-pass-fail-ajax")
@login_requerido
def historial_vision_pass_fail():
    """Servir la página de Historial Vision % Pass/Fail."""
    try:
        return render_template("Control de resultados/history_vision_pass_fail.html")
    except Exception as e:
        print(f"Error al cargar Historial Vision % Pass/Fail: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/api/vision/data")
@login_requerido
def vision_data_api():
    """Obtener registros recientes del historial de Vision con filtros opcionales."""
    try:
        sql, params = _build_history_vision_query()
        rows = execute_query(sql, params, fetch="all") or []

        result = []
        for row in rows:
            result.append(
                {
                    "id": row.get("id"),
                    "linea": _vision_format_value(row.get("linea", "")) or "",
                    "fecha": _vision_format_value(row.get("fecha", "")) or "",
                    "hora": _vision_format_value(row.get("hora", "")) or "",
                    "numero_parte": _vision_format_value(
                        row.get("numero_parte", "")
                    )
                    or "",
                    "qr": _vision_format_value(row.get("qr", "")) or "",
                    "barcode": _vision_format_value(row.get("barcode", "")) or "",
                    "resultado": _vision_format_value(row.get("resultado", ""))
                    or "",
                }
            )

        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/api/vision/pass-fail-summary")
@login_requerido
def vision_pass_fail_summary_api():
    """Obtener resumen agrupado por línea y número de parte para history_vision."""
    try:
        sql, params = _build_history_vision_pass_fail_summary_query()
        rows = execute_query(sql, params, fetch="all") or []

        result = []
        for row in rows:
            total = int(row.get("total") or 0)
            ok_count = int(row.get("ok_count") or 0)
            ng_count = int(row.get("ng_count") or 0)
            porcentaje_ok = round((ok_count / total) * 100, 2) if total else 0
            porcentaje_ng = round((ng_count / total) * 100, 2) if total else 0

            result.append(
                {
                    "linea": _vision_format_value(row.get("linea", "")) or "",
                    "numero_parte": _vision_format_value(
                        row.get("numero_parte", "")
                    )
                    or "",
                    "total": total,
                    "ok_count": ok_count,
                    "ng_count": ng_count,
                    "porcentaje_ok": porcentaje_ok,
                    "porcentaje_ng": porcentaje_ng,
                }
            )

        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/api/vision/pass-fail-summary/export")
@login_requerido
def export_vision_pass_fail_summary_excel():
    """Exportar resumen Pass/Fail de Vision a un archivo de Excel."""
    try:
        sql, params = _build_history_vision_pass_fail_summary_query()
        rows = execute_query(sql, params, fetch="all") or []

        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Vision Pass Fail"

        header_fill = PatternFill(
            start_color="3f6b6e", end_color="3f6b6e", fill_type="solid"
        )
        cell_fill = PatternFill(
            start_color="a1a09c", end_color="a1a09c", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        headers = [
            "Linea",
            "Numero de parte",
            "Total",
            "OK",
            "NG",
            "% Pass",
            "% Fail",
            "PORCENTAJE",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_idx, row in enumerate(rows, start=2):
            total = int(row.get("total") or 0)
            ok_count = int(row.get("ok_count") or 0)
            ng_count = int(row.get("ng_count") or 0)
            porcentaje_ok = round((ok_count / total) * 100, 2) if total else 0
            porcentaje_ng = round((ng_count / total) * 100, 2) if total else 0

            values = [
                _vision_format_value(row.get("linea", "")) or "",
                _vision_format_value(row.get("numero_parte", "")) or "",
                total,
                ok_count,
                ng_count,
                porcentaje_ok,
                porcentaje_ng,
            ]

            for col_num, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_num, value=value)
                cell.fill = cell_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            image_cell = ws.cell(row=row_idx, column=8, value="")
            image_cell.fill = cell_fill
            image_cell.alignment = Alignment(horizontal="center", vertical="center")
            image_cell.border = border

            excel_image = _create_vision_pass_fail_excel_image(
                porcentaje_ok, porcentaje_ng
            )
            ws.add_image(excel_image, f"H{row_idx}")
            ws.row_dimensions[row_idx].height = 24

        column_widths = [20, 28, 14, 12, 12, 14, 14, 58]
        for col_num, width in enumerate(column_widths, start=1):
            column_letter = ws.cell(row=1, column=col_num).column_letter
            ws.column_dimensions[column_letter].width = width

        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = (
            f"historial_vision_pass_fail_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        return _send_excel_download(output, filename)
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/api/vision/image-info")
@login_requerido
def vision_image_info_api():
    """Resolver metadata de la imagen asociada a un registro de history_vision."""
    record_id = request.args.get("id", "").strip()
    if not re.fullmatch(r"\d+", record_id):
        return jsonify({"error": "ID de registro invalido."}), 400

    try:
        record = _get_history_vision_record(int(record_id))
        if not record:
            return jsonify({"error": "Registro de vision no encontrado."}), 404

        resolution = _resolve_history_vision_image(record)
        response_payload = {
            "record_id": record.get("id"),
            "linea": _vision_format_value(record.get("machine_name", "")) or "",
            "numero_parte": _vision_format_value(record.get("part_code", "")) or "",
            "resultado": _vision_format_value(record.get("result", "")) or "",
            "fecha_hora": resolution.get("reference_datetime") or "",
            "resolved_path": resolution.get("resolved_path", ""),
            "share_name": resolution.get("share_name", ""),
            "side_folder": resolution.get("side_folder", ""),
            "delta_seconds": resolution.get("delta_seconds"),
            "searched_paths": resolution.get("searched_paths", []),
            "source_file": _vision_format_value(record.get("source_file", "")) or "",
            "machine_ip": _vision_format_value(record.get("machine_ip", "")) or "",
            "image_url": "",
        }

        if resolution.get("resolved_path"):
            response_payload["image_url"] = url_for(
                "vision_image_file_api", id=record.get("id")
            )
            return jsonify(response_payload)

        response_payload["error"] = resolution.get(
            "error", "No se encontro imagen para el registro."
        )
        return jsonify(response_payload), 404
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/api/vision/image-file")
@login_requerido
def vision_image_file_api():
    """Servir la imagen resuelta de un registro de history_vision."""
    record_id = request.args.get("id", "").strip()
    if not re.fullmatch(r"\d+", record_id):
        return jsonify({"error": "ID de registro invalido."}), 400

    try:
        record = _get_history_vision_record(int(record_id))
        if not record:
            return jsonify({"error": "Registro de vision no encontrado."}), 404

        resolution = _resolve_history_vision_image(record)
        resolved_path = resolution.get("resolved_path", "")
        share_roots = [Path(path) for path in resolution.get("share_roots", [])]

        if not resolved_path:
            return (
                jsonify(
                    {
                        "error": resolution.get(
                            "error", "No se encontro imagen para el registro."
                        ),
                        "searched_paths": resolution.get("searched_paths", []),
                    }
                ),
                404,
            )

        resolved_file = Path(resolved_path)
        if not _vision_is_safe_path(resolved_file, share_roots):
            return jsonify({"error": "La ruta resuelta no es segura."}), 403

        if not resolved_file.is_file():
            return (
                jsonify(
                    {
                        "error": "La imagen resuelta ya no esta disponible.",
                        "resolved_path": str(resolved_file),
                    }
                ),
                404,
            )

        mimetype_map = {
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".bmp": "image/bmp",
        }
        mimetype = mimetype_map.get(
            resolved_file.suffix.lower(), "application/octet-stream"
        )

        return send_file(
            str(resolved_file),
            mimetype=mimetype,
            as_attachment=False,
            download_name=resolved_file.name,
            conditional=True,
        )
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/api/vision/export")
@login_requerido
def export_vision_excel():
    """Exportar el historial Vision filtrado a un archivo de Excel."""
    try:
        sql, params = _build_history_vision_query()
        rows = execute_query(sql, params, fetch="all") or []

        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Historial Vision"

        header_fill = PatternFill(
            start_color="3f6b6e", end_color="3f6b6e", fill_type="solid"
        )
        cell_fill = PatternFill(
            start_color="a1a09c", end_color="a1a09c", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        headers = [
            "Linea",
            "Fecha",
            "Hora",
            "Numero de parte",
            "QR",
            "Barcode",
            "Resultado",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_idx, row in enumerate(rows, start=2):
            values = [
                _vision_format_value(row.get("linea", "")) or "",
                _vision_format_value(row.get("fecha", "")) or "",
                _vision_format_value(row.get("hora", "")) or "",
                _vision_format_value(row.get("numero_parte", "")) or "",
                _vision_format_value(row.get("qr", "")) or "",
                _vision_format_value(row.get("barcode", "")) or "",
                _vision_format_value(row.get("resultado", "")) or "",
            ]

            for col_num, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_num, value=value)
                cell.fill = cell_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        column_widths = [22, 14, 14, 28, 36, 24, 14]
        for col_num, width in enumerate(column_widths, start=1):
            column_letter = ws.cell(row=1, column=col_num).column_letter
            ws.column_dimensions[column_letter].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = (
            f"historial_vision_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/historial-ict")
@app.route("/ict/front-full-defects2")
@login_requerido
def ict_front_full_defects2():
    """Vista principal del historial ICT con defectos detallados."""
    try:
        return render_template("Control de resultados/history_ict.html")
    except Exception as e:
        print(f"Error al cargar History ICT: {e}")
        return f"Error al cargar el contenido: {str(e)}", 500


@app.route("/api/ict/data")
@login_requerido
def ict_data_api():
    """Obtener registros del historial ICT con filtros y paginacion.

    Soporta `fecha` (igualdad, retro-compatible) o `fecha_desde`/`fecha_hasta`
    (rango). Si se envian ambos, prevalece el rango.

    Paginacion: `page` (1-based) y `per_page` (default 200, max 1000).
    Cuando se envia `page`, la respuesta es un objeto con metadata; si no se
    envia, se devuelve el array plano (retro-compatible) con LIMIT 500.
    """
    try:
        fecha = request.args.get("fecha", "").strip()
        fecha_desde = request.args.get("fecha_desde", "").strip()
        fecha_hasta = request.args.get("fecha_hasta", "").strip()
        no_parte = request.args.get("no_parte", "").strip()
        linea = request.args.get("linea", "").strip()
        ict_filter = request.args.get("ict", "").strip()
        resultado = request.args.get("resultado", "").strip()
        barcode_like = request.args.get("barcode_like", "").strip()
        page_raw = request.args.get("page", "").strip()
        per_page_raw = request.args.get("per_page", "").strip()
        paginated = bool(page_raw)
        if barcode_like and len(barcode_like) < 6:
            if paginated:
                return jsonify({"rows": [], "total": 0, "page": 1, "per_page": 0, "total_pages": 0})
            return jsonify([])

        where_sql = "WHERE 1=1"
        params = []

        def _add(clause, *vals):
            nonlocal where_sql
            where_sql += " " + clause
            params.extend(vals)

        if fecha_desde or fecha_hasta:
            if fecha_desde:
                _add("AND fecha>=%s", fecha_desde)
            if fecha_hasta:
                _add("AND fecha<=%s", fecha_hasta)
        elif fecha:
            _add("AND fecha=%s", fecha)
        if no_parte:
            _add("AND no_parte LIKE %s", f"{no_parte}%")
        if linea:
            _add("AND linea=%s", linea)
        if ict_filter:
            try:
                _add("AND ict=%s", int(ict_filter))
            except ValueError:
                return jsonify({"error": "ict debe ser numerico"}), 400
        if resultado:
            _add("AND resultado=%s", resultado)
        if barcode_like:
            # Reaplicar la misma logica de _append_indexable_text_filter pero
            # sobre el WHERE acumulado (no sobre el SELECT completo).
            value = barcode_like.strip()
            if len(value) >= 12:
                where_sql += " AND barcode=%s"
                params.append(value)
            else:
                where_sql += " AND barcode LIKE %s"
                params.append(f"{value}%")

        select_cols = (
            "SELECT fecha, TIME(ts) AS hora, linea, ict, resultado, no_parte, barcode, "
            "ts, fuente_archivo, defect_code, defect_valor "
            "FROM history_ict "
        )

        if paginated:
            try:
                page = max(1, int(page_raw))
            except ValueError:
                page = 1
            try:
                per_page = int(per_page_raw) if per_page_raw else 200
            except ValueError:
                per_page = 200
            per_page = max(1, min(per_page, 1000))

            count_sql = "SELECT COUNT(*) AS n FROM history_ict " + where_sql
            count_row = execute_query(count_sql, tuple(params), fetch="one") or {}
            total = int(count_row.get("n", 0))

            offset = (page - 1) * per_page
            data_sql = select_cols + where_sql + " ORDER BY ts DESC LIMIT %s OFFSET %s"
            rows = execute_query(data_sql, tuple(params) + (per_page, offset), fetch="all") or []

            total_pages = (total + per_page - 1) // per_page if per_page else 0
            return jsonify({
                "rows": [_ict_format_row(row) for row in rows],
                "total": total,
                "page": page,
                "per_page": per_page,
                "total_pages": total_pages,
            })

        # Modo legacy sin paginacion (retro-compatible)
        data_sql = select_cols + where_sql + " ORDER BY ts DESC LIMIT 500"
        rows = execute_query(data_sql, tuple(params), fetch="all") or []
        return jsonify([_ict_format_row(row) for row in rows])
    except Exception as e:
        import traceback

        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/api/ict/defects")
@login_requerido
def ict_defects_api():
    """Obtener defectos asociados a un barcode espec��fico."""
    barcode = request.args.get("barcode", "").strip()
    ts = request.args.get("ts", "").strip()
    if not barcode:
        return jsonify([])

    try:
        rows, _ = _ict_load_local_parameters(barcode, ts)
        return jsonify([_ict_format_row(row) for row in rows])
    except IctLgdNotFoundError as e:
        return jsonify({"error": str(e)}), 404
    except IctLgdPathError as e:
        return jsonify({"error": str(e)}), 400
    except IctLgdError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        import traceback

        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/api/ict/export")
@login_requerido
def export_ict_excel():
    """Exportar el historial ICT filtrado a un archivo de Excel."""
    try:
        fecha = request.args.get("fecha", "").strip()
        fecha_desde = request.args.get("fecha_desde", "").strip()
        fecha_hasta = request.args.get("fecha_hasta", "").strip()
        no_parte = request.args.get("no_parte", "").strip()
        linea = request.args.get("linea", "").strip()
        ict_filter = request.args.get("ict", "").strip()
        resultado = request.args.get("resultado", "").strip()
        barcode_like = request.args.get("barcode_like", "").strip()
        if barcode_like and len(barcode_like) < 6:
            return jsonify({"error": "Barcode demasiado corto para exportar"}), 400

        sql = (
            "SELECT fecha, TIME(ts) AS hora, linea, ict, resultado, no_parte, barcode, "
            "fuente_archivo, defect_code, defect_valor "
            "FROM history_ict WHERE 1=1"
        )
        params = []

        if fecha_desde or fecha_hasta:
            if fecha_desde:
                sql += " AND fecha>=%s"
                params.append(fecha_desde)
            if fecha_hasta:
                sql += " AND fecha<=%s"
                params.append(fecha_hasta)
        elif fecha:
            sql += " AND fecha=%s"
            params.append(fecha)
        if no_parte:
            sql += " AND no_parte LIKE %s"
            params.append(f"{no_parte}%")
        if linea:
            sql += " AND linea=%s"
            params.append(linea)
        if ict_filter:
            try:
                sql += " AND ict=%s"
                params.append(int(ict_filter))
            except ValueError:
                return jsonify({"error": "ict debe ser numerico"}), 400
        if resultado:
            sql += " AND resultado=%s"
            params.append(resultado)
        if barcode_like:
            sql = _append_indexable_text_filter(sql, params, "barcode", barcode_like)

        sql += " ORDER BY ts DESC LIMIT 500"
        rows = execute_query(sql, tuple(params), fetch="all") or []

        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Historial ICT"

        header_fill = PatternFill(
            start_color="3f6b6e", end_color="3f6b6e", fill_type="solid"
        )
        cell_fill = PatternFill(
            start_color="a1a09c", end_color="a1a09c", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        headers = [
            "Fecha",
            "Hora",
            "L��nea",
            "ICT",
            "Resultado",
            "No Parte",
            "Barcode",
            "Fuente",
            "Defect Code",
            "Defect Valor",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_idx, row in enumerate(rows, start=2):
            formatted = _ict_format_row(row)
            values = [
                formatted.get("fecha", ""),
                formatted.get("hora", ""),
                formatted.get("linea", ""),
                formatted.get("ict", ""),
                formatted.get("resultado", ""),
                formatted.get("no_parte", ""),
                formatted.get("barcode", ""),
                formatted.get("fuente_archivo", ""),
                formatted.get("defect_code", ""),
                formatted.get("defect_valor", ""),
            ]

            for col_num, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_num, value=value)
                cell.fill = cell_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        for col in range(1, len(headers) + 1):
            column_letter = ws.cell(row=1, column=col).column_letter
            ws.column_dimensions[column_letter].width = 16

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"historial_ict_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except IctLgdNotFoundError as e:
        return jsonify({"error": str(e)}), 404
    except IctLgdPathError as e:
        return jsonify({"error": str(e)}), 400
    except IctLgdError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        import traceback

        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/api/ict/export-defects")
@login_requerido
def export_ict_defects_excel():
    """Exportar detalles de defectos ICT a un archivo de Excel."""
    barcode = request.args.get("barcode", "").strip()
    ts = request.args.get("ts", "").strip()
    resultado_filter = request.args.get("resultado", "").strip()

    if not barcode:
        return jsonify({"error": "Barcode requerido"}), 400

    try:
        rows, _ = _ict_load_local_parameters(barcode, ts)
        if resultado_filter:
            rows = [row for row in rows if row.get("resultado_local") == resultado_filter]

        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = f"Parametros {barcode[:20]}"

        header_fill = PatternFill(
            start_color="3f6b6e", end_color="3f6b6e", fill_type="solid"
        )
        cell_fill = PatternFill(
            start_color="a1a09c", end_color="a1a09c", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        headers = [
            "Fecha",
            "Hora",
            "L��nea",
            "ICT",
            "Barcode",
            "Componente",
            "Pinref",
            "ACT",
            "Unit",
            "STD",
            "Unit",
            "MEAS",
            "M",
            "R",
            "HLIM",
            "LLIM",
            "H.P",
            "L.P",
            "WS",
            "DS",
            "RC",
            "P",
            "J",
            "Resultado",
            "Tipo Defecto",
        ]

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_idx, row in enumerate(rows, start=2):
            formatted = _ict_format_row(row)
            hlim = formatted.get("hlim_pct", "")
            llim = formatted.get("llim_pct", "")

            row_values = [
                formatted.get("fecha", ""),
                formatted.get("hora", ""),
                formatted.get("linea", ""),
                formatted.get("ict", ""),
                formatted.get("barcode", ""),
                formatted.get("componente", ""),
                formatted.get("pinref", ""),
                formatted.get("act_value", ""),
                formatted.get("act_unit", ""),
                formatted.get("std_value", ""),
                formatted.get("std_unit", ""),
                formatted.get("meas_value", ""),
                formatted.get("m_value", ""),
                formatted.get("r_value", ""),
                f"{hlim}%" if hlim else "",
                f"{llim}%" if llim else "",
                formatted.get("hp_value", ""),
                formatted.get("lp_value", ""),
                formatted.get("ws_value", ""),
                formatted.get("ds_value", ""),
                formatted.get("rc_value", ""),
                formatted.get("p_flag", ""),
                formatted.get("j_flag", ""),
                formatted.get("resultado_local", ""),
                formatted.get("defecto_tipo", ""),
            ]

            for col_num, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_num, value=value)
                cell.fill = cell_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        for col in range(1, len(headers) + 1):
            column_letter = ws.cell(row=1, column=col).column_letter
            ws.column_dimensions[column_letter].width = 12

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = (
            f"parametros_{barcode}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except IctLgdNotFoundError as e:
        return jsonify({"error": str(e)}), 404
    except IctLgdPathError as e:
        return jsonify({"error": str(e)}), 400
    except IctLgdError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        import traceback

        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/api/ict/export-compare", methods=["POST"])
@login_requerido
def export_ict_compare_excel():
    """Exportar comparacion de parametros ICT entre varias ejecuciones (auditoria).

    Body JSON: { runs: [{barcode, ts, fecha, hora, linea, resultado}, ...], only_diffs: bool }
    """
    try:
        payload = request.get_json(silent=True) or {}
        runs_input = payload.get("runs") or []
        only_diffs = bool(payload.get("only_diffs", True))

        if len(runs_input) < 2:
            return jsonify({"error": "Se requieren al menos 2 ejecuciones"}), 400

        runs = []
        for idx, rec in enumerate(runs_input, start=1):
            barcode = (rec.get("barcode") or "").strip()
            ts = (rec.get("ts") or "").strip()
            if not barcode:
                continue
            rows, _ = _ict_load_local_parameters(barcode, ts)
            runs.append({
                "run_index": idx,
                "barcode": barcode,
                "ts": ts,
                "fecha": rec.get("fecha") or "",
                "hora": rec.get("hora") or "",
                "linea": rec.get("linea") or "",
                "resultado": rec.get("resultado") or "",
                "defects": [_ict_format_row(row) for row in rows],
            })

        if len(runs) < 2:
            return jsonify({"error": "No se pudieron cargar parametros de las ejecuciones"}), 400

        compare_fields = [
            ("act_value", "ACT", ""),
            ("act_unit", "UNIT", ""),
            ("std_value", "STD", ""),
            ("std_unit", "UNIT", ""),
            ("m_value", "M", ""),
            ("r_value", "R", ""),
            ("hlim_pct", "HLIM %", "%"),
            ("llim_pct", "LLIM %", "%"),
            ("hp_value", "HP", ""),
            ("lp_value", "LP", ""),
            ("ws_value", "WS", ""),
            ("ds_value", "DS", ""),
            ("rc_value", "RC", ""),
            ("p_flag", "P", ""),
            ("j_flag", "J", ""),
        ]

        groups = {}
        for run in runs:
            for d in run["defects"]:
                key = (d.get("componente") or "", d.get("pinref") or "")
                if key not in groups:
                    groups[key] = {"componente": key[0], "pinref": key[1], "by_run": {}}
                groups[key]["by_run"][run["run_index"]] = d

        def _norm(v):
            if v is None:
                return ""
            return str(v).strip()

        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Comparacion ICT"

        header_fill = PatternFill(start_color="3f6b6e", end_color="3f6b6e", fill_type="solid")
        cell_fill = PatternFill(start_color="a1a09c", end_color="a1a09c", fill_type="solid")
        diff_fill = PatternFill(start_color="f4b3ad", end_color="f4b3ad", fill_type="solid")
        missing_fill = PatternFill(start_color="d9d9d9", end_color="d9d9d9", fill_type="solid")
        info_fill = PatternFill(start_color="dde6e8", end_color="dde6e8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        title_font = Font(bold=True, color="000000", size=12)
        info_font = Font(bold=True, color="000000", size=9)
        diff_font = Font(bold=True, color="8b0000", size=10)
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.cell(row=1, column=1, value="COMPARACION DE PARAMETROS ICT - AUDITORIA").font = title_font
        ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=2, column=2, value=f"Solo diferencias: {'SI' if only_diffs else 'NO'}")

        info_headers = ["#", "Barcode", "Fecha", "Hora", "Linea", "Resultado"]
        for col_num, h in enumerate(info_headers, start=1):
            c = ws.cell(row=4, column=col_num, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

        for i, run in enumerate(runs, start=5):
            for col_num, value in enumerate([
                f"#{run['run_index']}",
                run["barcode"],
                run["fecha"],
                run["hora"],
                run["linea"],
                run["resultado"],
            ], start=1):
                c = ws.cell(row=i, column=col_num, value=value)
                c.fill = info_fill
                c.font = info_font
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border

        table_start_row = 5 + len(runs) + 2
        table_headers = ["Componente", "Pinref", "Ejecucion"] + [f[1] for f in compare_fields]
        for col_num, h in enumerate(table_headers, start=1):
            c = ws.cell(row=table_start_row, column=col_num, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

        sorted_keys = sorted(groups.keys(), key=lambda k: (k[0] or "", k[1] or ""))

        current_row = table_start_row + 1
        for key in sorted_keys:
            group = groups[key]

            diff_keys = set()
            for f in compare_fields:
                values = set()
                for run in runs:
                    d = group["by_run"].get(run["run_index"])
                    if d:
                        values.add(_norm(d.get(f[0])))
                if len(values) > 1:
                    diff_keys.add(f[0])

            missing_in_some = any(run["run_index"] not in group["by_run"] for run in runs)
            has_any_diff = bool(diff_keys) or missing_in_some

            if only_diffs and not has_any_diff:
                continue

            for run in runs:
                d = group["by_run"].get(run["run_index"])
                run_label = f"#{run['run_index']} - {run['barcode']} {run['fecha']} {run['hora']}"
                row_values = [group["componente"], group["pinref"], run_label]

                for f in compare_fields:
                    if not d:
                        row_values.append("--")
                        continue
                    raw = d.get(f[0])
                    if raw is None or raw == "":
                        row_values.append("")
                    else:
                        suffix = f[2]
                        row_values.append(f"{raw}{suffix}" if suffix else raw)

                for col_num, value in enumerate(row_values, start=1):
                    cell = ws.cell(row=current_row, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border

                    if col_num <= 3:
                        cell.fill = cell_fill
                    else:
                        field = compare_fields[col_num - 4]
                        if not d:
                            cell.fill = missing_fill
                        elif field[0] in diff_keys:
                            cell.fill = diff_fill
                            cell.font = diff_font
                        else:
                            cell.fill = cell_fill

                current_row += 1

        widths = [18, 12, 38] + [12] * len(compare_fields)
        for col_idx, width in enumerate(widths, start=1):
            column_letter = ws.cell(row=table_start_row, column=col_idx).column_letter
            ws.column_dimensions[column_letter].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"comparacion_ict_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except IctLgdNotFoundError as e:
        return jsonify({"error": str(e)}), 404
    except IctLgdPathError as e:
        return jsonify({"error": str(e)}), 400
    except IctLgdError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        import traceback

        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500
